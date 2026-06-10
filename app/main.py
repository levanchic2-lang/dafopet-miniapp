from __future__ import annotations

# ── 强制进程时区为北京（Asia/Shanghai），必须在任何 datetime/time 调用前执行 ──
# 系统统一按北京时间运作：无论运行主机在哪个时区、是否挂全局 VPN，
# datetime.now() / date.today() / time.localtime() / strftime 都返回北京时间。
# datetime.utcnow() 不受影响（始终 UTC，显示侧已有的 +8 转换照常正确）。
import os as _os
import time as _time
_os.environ["TZ"] = "Asia/Shanghai"
if hasattr(_time, "tzset"):          # Unix（Linux 服务器 / macOS）生效；Windows 无 tzset 跳过
    _time.tzset()
# ────────────────────────────────────────────────────────────────────────────

import asyncio
import json
import logging
import mimetypes
import re
import secrets
import shutil
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from typing import Annotated, Optional
from urllib.parse import quote

from fastapi import Body, Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from passlib.context import CryptContext
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.templating import Jinja2Templates
import httpx
from sqlalchemy import func, or_
from sqlalchemy.orm import Session, selectinload
from starlette.middleware.sessions import SessionMiddleware
from starlette.staticfiles import StaticFiles

# 启动时若缺 Pillow 则自动安装（openpyxl 图片嵌入依赖）
try:
    import PIL  # noqa: F401
except ImportError:
    try:
        subprocess.run(
            [__import__("sys").executable, "-m", "pip", "install", "Pillow", "-q"],
            capture_output=True, timeout=120,
        )
    except Exception:
        pass

from app.config import settings
from app.database import get_db, init_db
from app.models import (
    AdminUser,
    Application,
    ApplicationStatus,
    Appointment,
    AppointmentCategory,
    AppointmentStatus,
    AuditLog,
    Contract,
    ContractType,
    Customer,
    MediaFile,
    MediaKind,
    Pet,
    Staff,
    StaffStatus,
    Prescription,
    PrescriptionItem,
    SalesOrder,
    SalesOrderItem,
    Visit,
    InventoryItem,
    InventoryTransaction,
    InventoryBatch,
    StocktakeSession,
    StocktakeItem,
    RabiesVaccineRecord,
    AdoptionPet,
    Invoice,
    InvoiceItem,
    Vaccination,
    TnrStoreConfig,
    ExamOrder,
    ExamReport,
    MicroscopyReport,
    CalendarBlock,
    DewormingRecord,
    WeightRecord,
    MedicalDocument,
    PrescriptionTemplate,
    ExamTemplate,
    FollowUp,
    Wallet,
    WalletTransaction,
    PackageProduct,
    CustomerPackage,
    PackageRedemption,
    Deposit,
    Payment,
    Coupon,
    ConsentTemplate,
    ConsentTask,
    ConsentDocument,
    AnesthesiaOrder,
    AnesthesiaOrderItem,
    NarcoticsLedger,
    AnesthesiaMonitorSheet,
    AnesthesiaMonitorEntry,
    FollowUpTemplate,
    Disease,
    GroomingOrder,
    Cage,
    Hospitalization,
    MedicationAdminLog,
    VitalSignsLog,
    IOLog,
    FeedingLog,
    HandoverNote,
)
from app.services.ai_review import apply_auto_status_from_ai, review_application_media
from app.services.breeds import all_breeds as _all_breeds
_BREEDS_ALL = _all_breeds()
from app.services.notify import notify_application_result
from app.services.backup_local import create_backup_zip, is_safe_backup_filename, list_backup_zips
from app.services.wechat_miniapp import push_application_result, push_appointment_status, push_pending_manual_notice, push_rejection_notice, push_surgery_done, push_surgery_reminder, push_vaccine_reminder, wechat_code2session

logger = logging.getLogger(__name__)

app = FastAPI(title=settings.app_name)
app.add_middleware(SessionMiddleware, secret_key=settings.session_secret, session_cookie="tnr_session")

_pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
templates = Jinja2Templates(directory=str(Path(__file__).resolve().parent.parent / "templates"))

# 后台「AI 辅助结论」：把模型 JSON 转成中文可读结构（模板用 | ai_review_view）
_AI_FLAG_ZH = {
    "collar": "可见项圈等装饰",
    "carrier": "猫包 / 航空箱等携带方式",
    "indoor_luxury": "偏家养 / 室内饲养环境线索",
}


def _filter_ai_review_view(raw: str | None) -> dict:
    out: dict = {
        "parse_error": False,
        "stray_zh": "—",
        "confidence_zh": "—",
        "reasons": [],
        "photo_text": "",
        "fraud_lines": [],
        "caveats": [],
        "suggestion_zh": "",
    }
    if not raw or not str(raw).strip():
        return out
    try:
        d = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        out["parse_error"] = True
        return out
    if not isinstance(d, dict):
        out["parse_error"] = True
        return out

    v = d.get("is_likely_stray")
    if v is True:
        out["stray_zh"] = "是（模型判断更接近流浪 / 无主场景）"
    elif v is False:
        out["stray_zh"] = "否（模型判断不够像流浪，或信息不足）"
    else:
        out["stray_zh"] = "未给出明确结论"

    conf = d.get("confidence")
    try:
        c = float(conf)
        if 0 <= c <= 1:
            out["confidence_zh"] = f"{c * 100:.0f}%"
        else:
            out["confidence_zh"] = f"{c:.2f}"
    except (TypeError, ValueError):
        out["confidence_zh"] = "—"

    reasons = d.get("reasons")
    if isinstance(reasons, list):
        out["reasons"] = [str(x).strip() for x in reasons if str(x).strip()]

    kidx = d.get("key_evidence_photo_indexes")
    if isinstance(kidx, list) and kidx:
        nums: list[str] = []
        for x in kidx:
            try:
                nums.append(str(int(x)))
            except (TypeError, ValueError):
                if x is not None and str(x).strip():
                    nums.append(str(x).strip())
        if nums:
            out["photo_text"] = "第 " + "、".join(nums) + " 张（按申请时照片顺序）"

    flags = d.get("anti_fraud_flags")
    if isinstance(flags, list) and flags:
        known: list[str] = []
        unknown_n = 0
        for f in flags:
            key = str(f).strip().lower()
            zh = _AI_FLAG_ZH.get(key)
            if zh:
                if zh not in known:
                    known.append(zh)
            elif str(f).strip():
                unknown_n += 1
        out["fraud_lines"] = known.copy()
        if unknown_n:
            out["fraud_lines"].append(f"另有 {unknown_n} 条内部标记未展开（已由系统参与规则判断）")

    caveats = d.get("caveats")
    if isinstance(caveats, list):
        out["caveats"] = [str(x).strip() for x in caveats if str(x).strip()]

    step = (d.get("suggested_next_step") or "").strip().lower().replace("-", "_")
    if step == "auto_approve_candidate":
        out["suggestion_zh"] = "模型流程建议：可作自动通过候选（实际状态已由系统规则与阈值综合决定，医院仍可拒绝或取消）。"
    elif step == "manual_review":
        out["suggestion_zh"] = "模型流程建议：优先走人工复核。"
    elif step:
        out["suggestion_zh"] = "模型已给出内部流程建议，系统已按规则处理。"
    else:
        out["suggestion_zh"] = "—"

    return out


templates.env.filters["ai_review_view"] = _filter_ai_review_view


def _filter_parse_json(s):
    """Jinja 过滤器：把 JSON 字符串解析成 dict/list，失败返回 None。"""
    if not s:
        return None
    try:
        import json as _json
        return _json.loads(s)
    except Exception:
        return None


templates.env.filters["parse_json"] = _filter_parse_json


def _filter_pet_age(birthday: str) -> str:
    """把 'YYYY-MM-DD' 出生日期渲染成「X 岁」或「Y 个月」可读年龄。
    非日期格式（老数据如 '2岁'）原样返回。"""
    if not birthday:
        return "—"
    s = str(birthday).strip()
    if not s or s == "—":
        return "—"
    # 老自由文本兼容
    if not (len(s) >= 7 and s[4] in ("-", "/") and s[:4].isdigit()):
        return s
    try:
        from datetime import date as _date
        parts = s.replace("/", "-").split("-")
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2]) if len(parts) > 2 else 1
        today = _date.today()
        years = today.year - y - (1 if (today.month, today.day) < (m, d) else 0)
        if years >= 1:
            return f"{years} 岁"
        months = (today.year - y) * 12 + (today.month - m) - (1 if today.day < d else 0)
        months = max(0, months)
        return f"{months} 个月"
    except Exception:
        return s


templates.env.filters["pet_age"] = _filter_pet_age

# 姓名规范性检测：捕获「先生 / 女士 / 小姐 / X小姐 / X先生 / X女士 …」等占位 / 单姓后缀格式
# （实现见下文 _is_invalid_name；这里只声明全局，使 templates 内可 {{ name | is_invalid_name }}）
def _filter_is_invalid_name(name: str) -> bool:
    return _is_invalid_name(name or "")

templates.env.filters["is_invalid_name"] = _filter_is_invalid_name


# 门店级价格覆盖（方案 H）— Jinja 全局过滤器
# 用法：{{ item | eff_price(current_store) }}
from app.services.pricing import (
    effective_sell_price as _eff_sell_price,
    effective_cost_price as _eff_cost_price,
    has_override as _has_override,
    overrides_summary as _overrides_summary,
)

def _filter_eff_price(item, store: str = "") -> float:
    return _eff_sell_price(item, store or "")

def _filter_eff_cost(item, store: str = "") -> float:
    return _eff_cost_price(item, store or "")

templates.env.filters["eff_price"] = _filter_eff_price
templates.env.filters["eff_cost"] = _filter_eff_cost
templates.env.filters["has_override"] = lambda item, store: _has_override(item, store or "")
templates.env.filters["overrides_summary"] = _overrides_summary

_static_dir = Path(__file__).resolve().parent.parent / "static"
if _static_dir.is_dir():
    app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")

Path(settings.upload_dir).mkdir(parents=True, exist_ok=True)
# /uploads 静态文件兜底（生产 nginx 通常会优先匹配；本地 / 无 nginx 时由 FastAPI 直接提供）
_uploads_dir = Path(settings.upload_dir).resolve()
if _uploads_dir.is_dir():
    app.mount("/uploads", StaticFiles(directory=str(_uploads_dir)), name="uploads")

# 中国省 / 市 / 区 / 街道四级数据（static/china_pcas.json，来源见 static/china_pcas.source.txt）
_china_pcas: dict | None = None


def _load_china_pcas() -> dict:
    global _china_pcas
    if _china_pcas is None:
        path = Path(__file__).resolve().parent.parent / "static" / "china_pcas.json"
        if not path.is_file():
            _china_pcas = {}
        else:
            _china_pcas = json.loads(path.read_text(encoding="utf-8"))
    return _china_pcas


@app.on_event("startup")
def _startup():
    # 部署后看这行即可确认进程时区已钉死北京（不受主机/VPN 影响）
    logger.info(
        "时区 TZ=%s | 本地 now=%s | UTC=%s",
        _os.environ.get("TZ", "(未设置)"),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
    )
    init_db()
    asyncio.get_event_loop().create_task(_surgery_reminder_loop())
    asyncio.get_event_loop().create_task(_vaccine_reminder_loop())
    # 回访任务调度器（每小时跑一次）
    try:
        from app.services.followup_dispatch import start_scheduler as _start_fu
        _start_fu()
    except Exception as _e:
        logger.warning("回访调度器启动失败：%s", _e)
    # 住院 漏药 + 接班提醒（每 5 分钟扫漏药 / 6:50 14:50 21:50 接班推送）
    try:
        from app.services.inpatient_dispatch import start_scheduler as _start_ip
        _start_ip()
    except Exception as _e:
        logger.warning("住院调度器启动失败：%s", _e)


@app.on_event("shutdown")
def _shutdown():
    try:
        from app.services.followup_dispatch import stop_scheduler as _stop_fu
        _stop_fu()
    except Exception:
        pass
    try:
        from app.services.inpatient_dispatch import stop_scheduler as _stop_ip
        _stop_ip()
    except Exception:
        pass


async def _surgery_reminder_loop():
    """每天 08:00 检查手术预约并推送前一天+当天提醒。
    先等到下一个 08:00 再执行，避免服务重启时立即触发误推。
    """
    while True:
        now = datetime.now()
        next_run = now.replace(hour=8, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        await asyncio.sleep((next_run - now).total_seconds())
        try:
            _run_surgery_reminders()
        except Exception:
            pass


def _run_surgery_reminders():
    """同步执行：查询手术预约并推送提醒。"""
    from app.database import SessionLocal
    db = SessionLocal()
    try:
        today = datetime.now().date()
        tomorrow = today + timedelta(days=1)
        today_str = today.strftime("%Y-%m-%d")
        tomorrow_str = tomorrow.strftime("%Y-%m-%d")

        # 查询今天和明天的手术/TNR 预约（已确认）
        rows = (
            db.query(Appointment)
            .filter(
                Appointment.category.in_([AppointmentCategory.surgery.value, AppointmentCategory.tnr.value]),
                Appointment.status == AppointmentStatus.confirmed.value,
                Appointment.appointment_date.in_([today_str, tomorrow_str]),
                Appointment.wechat_openid.isnot(None),
                Appointment.wechat_openid != "",
            )
            .all()
        )
        for row in rows:
            openid = (row.wechat_openid or "").strip()
            if not openid:
                continue
            reminder_type = "day_of" if row.appointment_date == today_str else "day_before"
            log_key = f"surgery_reminder_{reminder_type}_{row.id}_{row.appointment_date}"
            # 检查是否已推送过（同一天同一类型不重复发）
            from app.models import NotificationLog as _NL
            already = (
                db.query(_NL)
                .filter(_NL.payload.contains(log_key))
                .first()
            )
            if already:
                continue
            cat_name = (row.pet_name or "猫咪").strip()
            push_surgery_reminder(
                db,
                appointment_id=row.id,
                openid=openid,
                cat_name=cat_name,
                appointment_date=row.appointment_date or "",
                appointment_time=row.appointment_time or "",
                reminder_type=reminder_type,
            )
            # 记录已推送标记
            db.add(_NL(
                application_id=row.related_application_id,
                channel="log",
                payload=log_key,
                success=True,
            ))
            db.commit()
    finally:
        db.close()


_VACC_REMINDER_DAYS = 7  # 提前 N 天推送


def _run_vaccine_reminders(db: Session | None = None) -> dict:
    """查询 N 天内到期的疫苗并推送提醒，返回 {"sent": int, "skipped": int, "errors": int}。"""
    from app.database import SessionLocal
    close_db = db is None
    if db is None:
        db = SessionLocal()
    try:
        today = datetime.now().date()
        deadline = today + timedelta(days=_VACC_REMINDER_DAYS)
        today_str = today.strftime("%Y-%m-%d")
        deadline_str = deadline.strftime("%Y-%m-%d")

        rows = (
            db.query(Vaccination)
            .filter(
                Vaccination.next_due_date >= today_str,
                Vaccination.next_due_date <= deadline_str,
                Vaccination.reminder_sent_at.is_(None),
            )
            .all()
        )

        vacc_type_zh_map = {
            "rabies": "狂犬疫苗", "combo_3": "猫三联", "combo_6": "猫六联",
            "canine_8": "犬八联", "deworming": "驱虫", "other": "其他疫苗",
        }
        sent = skipped = errors = 0
        for row in rows:
            # 找 openid：通过 customer → wechat_openid
            cust = row.customer
            if not cust:
                skipped += 1
                continue
            openid = (cust.wechat_openid or "").strip()
            if not openid:
                skipped += 1
                continue
            pet_name = row.pet.name if row.pet else "宠物"
            vtype_zh = vacc_type_zh_map.get(row.vaccine_type or "", "疫苗")
            try:
                push_vaccine_reminder(
                    db,
                    vaccination_id=row.id,
                    openid=openid,
                    pet_name=pet_name,
                    vaccine_type_zh=vtype_zh,
                    next_due_date=row.next_due_date or "",
                )
                row.reminder_sent_at = datetime.utcnow()
                db.commit()
                sent += 1
            except Exception:
                errors += 1
        return {"sent": sent, "skipped": skipped, "errors": errors}
    finally:
        if close_db:
            db.close()


async def _vaccine_reminder_loop():
    """每天 09:00 执行疫苗到期提醒推送。"""
    while True:
        now = datetime.now()
        next_run = now.replace(hour=9, minute=0, second=0, microsecond=0)
        if now >= next_run:
            next_run += timedelta(days=1)
        await asyncio.sleep((next_run - now).total_seconds())
        try:
            _run_vaccine_reminders()
        except Exception:
            pass


def _upsert_customer(db: Session, name: str, phone: str, openid: str = "", id_number: str = "", address: str = "", source: str = "") -> "Customer":
    """查找或创建客户档案，始终合并最新信息。

    匹配优先级：phone → openid。
    姓名覆盖规则：当老档案名为空 / 占位（"高女士" 等），且新名是合法全名时，覆盖。
    """
    phone = (phone or "").strip()
    cust = db.query(Customer).filter(Customer.phone == phone).first() if phone else None
    if not cust:
        # 尝试通过 openid 查找（openid 非空时）
        if openid and openid.strip():
            cust = db.query(Customer).filter(Customer.wechat_openid == openid.strip()).first()
    if cust:
        # 合并更新
        new_name = (name or "").strip()
        old_name = (cust.name or "").strip()
        # 覆盖姓名条件：
        #   1. 老档案名为空 → 直接填
        #   2. 老档案名是占位（X女士/X先生/X小姐 等）+ 姓氏与新名首字相同
        #      + 新名是真实全名 → 用新名修复（避免同手机号不同人误覆盖）
        if new_name:
            if not old_name:
                cust.name = new_name[:120]
            elif _is_invalid_name(old_name) and not _is_invalid_name(new_name):
                # 占位名格式 "X+后缀"：取首字作为姓氏
                old_surname = old_name[0] if old_name else ""
                new_surname = new_name[0] if new_name else ""
                if not old_surname or old_surname == new_surname:
                    cust.name = new_name[:120]
        if openid and not cust.wechat_openid:
            cust.wechat_openid = openid.strip()
        if id_number and not cust.id_number:
            cust.id_number = id_number[:40]
        if address and not cust.address:
            cust.address = address[:500]
    else:
        cust = Customer(
            name=(name or "")[:120],
            phone=phone[:40],
            wechat_openid=(openid or "").strip()[:64],
            id_number=(id_number or "")[:40],
            address=(address or "")[:500],
            source=(source or "")[:40],
        )
        db.add(cust)
        db.flush()  # get id without commit
    return cust


def _admin_ok(request: Request) -> bool:
    return bool(request.session.get("admin"))


def _admin_role(request: Request) -> str:
    """返回当前登录角色：'superadmin' | 'staff' | ''（未登录）。旧 session 默认当 superadmin。"""
    if not request.session.get("admin"):
        return ""
    return request.session.get("admin_role", "superadmin")


def _is_superadmin(request: Request) -> bool:
    return _admin_role(request) == "superadmin"


def require_admin(request: Request):
    if not _admin_ok(request):
        # 浏览器导航类请求 → 重定向到登录页（友好）
        # API/AJAX 请求 → 401 JSON（让前端拿到错误码）
        accept = (request.headers.get("accept") or "").lower()
        path = request.url.path or ""
        is_html_nav = (
            "text/html" in accept
            and not path.startswith("/api/")
        )
        if is_html_nav:
            raise HTTPException(
                status_code=303,
                headers={"Location": "/admin/login"},
            )
        raise HTTPException(status_code=401, detail="需要医院后台登录")


def require_superadmin(request: Request):
    if not _is_superadmin(request):
        raise HTTPException(status_code=403, detail="需要超级管理员权限")


def _get_csrf_token(request: Request) -> str:
    tok = request.session.get("csrf_token") or ""
    if not isinstance(tok, str) or not tok:
        tok = secrets.token_urlsafe(32)
        request.session["csrf_token"] = tok
    return tok


def _admin_back(request: Request, app_id: int, msg: str = "") -> RedirectResponse:
    """操作完成后跳回后台，保留当前搜索/翻页参数，并定位到对应申请卡片。"""
    from urllib.parse import urlparse, urlencode, parse_qs
    referer = request.headers.get("referer", "")
    qs_keep: dict[str, str] = {}
    if referer:
        try:
            parsed = urlparse(referer)
            params = parse_qs(parsed.query, keep_blank_values=False)
            qs_keep = {k: v[0] for k, v in params.items()
                       if k in ("q", "page", "store", "status", "page_size")}
        except Exception:
            pass
    if msg:
        qs_keep["msg"] = msg
    qs = ("?" + urlencode(qs_keep)) if qs_keep else ""
    return RedirectResponse(f"/admin{qs}#app-{app_id}", status_code=303)


def _require_csrf(request: Request, csrf_token: str) -> None:
    expected = request.session.get("csrf_token") or ""
    if not isinstance(expected, str) or not expected:
        raise HTTPException(status_code=403, detail="CSRF token missing")
    if not isinstance(csrf_token, str) or not secrets.compare_digest(expected, csrf_token):
        raise HTTPException(status_code=403, detail="CSRF token invalid")


def _audit(
    db: Session,
    request: Request,
    action: str,
    *,
    application_id: int | None = None,
    detail: dict | str | None = None,
):
    ip = (request.client.host if request.client else "") or ""
    ua = request.headers.get("user-agent", "") or ""
    if isinstance(detail, dict):
        detail_s = json.dumps(detail, ensure_ascii=False)
    elif isinstance(detail, str):
        detail_s = detail
    else:
        detail_s = ""
    db.add(
        AuditLog(
            action=action,
            actor=request.session.get("admin_username", "admin"),
            application_id=application_id,
            ip=ip,
            user_agent=ua[:300],
            detail=detail_s,
        )
    )


def _require_status_in(row: Application, allowed: set[str], action_label: str) -> None:
    if row.status not in allowed:
        zh = {
            ApplicationStatus.draft.value: "草稿",
            ApplicationStatus.pending_ai.value: "系统处理中",
            ApplicationStatus.pending_manual.value: "待人工审核",
            ApplicationStatus.pre_approved.value: "预通过（待复核）",
            ApplicationStatus.approved.value: "已通过",
            ApplicationStatus.scheduled.value: "已预约",
            ApplicationStatus.no_show.value: "爽约",
            ApplicationStatus.cancelled.value: "已取消",
            ApplicationStatus.rejected.value: "已拒绝",
            ApplicationStatus.arrived_verified.value: "已到店",
            ApplicationStatus.surgery_completed.value: "手术完成",
        }
        allowed_zh = " / ".join(zh.get(s, s) for s in sorted(allowed))
        current_zh = zh.get(row.status, row.status)
        raise HTTPException(409, f"{action_label}仅允许在「{allowed_zh}」状态执行，当前为「{current_zh}」。")


def _application_has_surgery_before_and_after(db: Session, application_id: int) -> bool:
    """术前、术后各至少一条媒体（照片或视频均可）。"""
    rows = (
        db.query(MediaFile.kind)
        .filter(MediaFile.application_id == application_id)
        .filter(MediaFile.kind.in_((MediaKind.surgery_before.value, MediaKind.surgery_after.value)))
        .all()
    )
    kinds = {r[0] for r in rows}
    return MediaKind.surgery_before.value in kinds and MediaKind.surgery_after.value in kinds


@app.get("/api/regions/china")
async def api_regions_china():
    """省 / 市 / 区 / 街道四级行政区划（全量）。数据：modood/Administrative-divisions-of-China dist/pcas.json"""
    return _load_china_pcas()


def _shenzhen_district_streets() -> dict:
    """深圳市：区 → 街道列表（当前业务仅限深圳门店）。"""
    pcas = _load_china_pcas()
    prov = pcas.get("广东省") or {}
    sz = prov.get("深圳市")
    return sz if isinstance(sz, dict) else {}


def _shenzhen_regions_embed() -> dict:
    """优先读 static/shenzhen_regions.json；缺失或损坏时从全量 pcas 推导。"""
    p = Path(__file__).resolve().parent.parent / "static" / "shenzhen_regions.json"
    if p.is_file():
        try:
            d = json.loads(p.read_text(encoding="utf-8"))
            if isinstance(d, dict) and d:
                return d
        except (json.JSONDecodeError, OSError):
            pass
    return _shenzhen_district_streets()


@app.get("/api/regions/shenzhen")
async def api_regions_shenzhen():
    """深圳市 区 / 街道二级数据（体积小，供申请页默认深圳使用）。"""
    return _shenzhen_regions_embed()


@app.get("/api/diag")
async def api_diag():
    k = settings.openai_api_key or ""
    ws = settings.wechat_appsecret or ""
    return {
        "openai_base_url": settings.openai_base_url,
        "openai_model": settings.openai_model,
        "openai_key_set": bool(k.strip()),
        "openai_key_is_ascii": k.isascii() if k else True,
        "openai_key_len": len(k),

        # wechat miniapp (do not expose secret)
        "wechat_appid_set": bool((settings.wechat_appid or "").strip()),
        "wechat_appsecret_set": bool(ws.strip()),
        "wechat_tmpl_application_result_set": bool((settings.wechat_tmpl_application_result or "").strip()),
        "wechat_tmpl_surgery_done_set": bool((settings.wechat_tmpl_surgery_done or "").strip()),
        "wechat_message_page": settings.wechat_message_page,
    }


@app.get("/api/wechat/config")
async def api_wechat_config():
    """给小程序前端下发订阅消息模板配置（不包含任何 secret）。"""
    return {
        "wechat_appid": settings.wechat_appid,
        "wechat_tmpl_application_result": settings.wechat_tmpl_application_result,
        "wechat_tmpl_surgery_done": settings.wechat_tmpl_surgery_done,
        "wechat_tmpl_appointment": settings.wechat_tmpl_appointment,
        "wechat_tmpl_rejection": settings.wechat_tmpl_rejection,
        "wechat_tmpl_pending_manual": settings.wechat_tmpl_pending_manual,
        "wechat_tmpl_surgery_reminder": settings.wechat_tmpl_surgery_reminder,
        "wechat_message_page": settings.wechat_message_page,
        "wechat_fields_application_result": settings.wechat_fields_application_result,
        "wechat_fields_surgery_done": settings.wechat_fields_surgery_done,
        "wechat_fields_appointment": settings.wechat_fields_appointment,
        "wechat_fields_rejection": settings.wechat_fields_rejection,
    }


@app.get("/api/geocode/regeo")
async def api_geocode_regeo(lat: str = "", lng: str = ""):
    """经纬度 -> 地址（高德逆地理编码）。未配置 Key 时返回空。"""
    key = (settings.amap_web_key or "").strip()
    if not key:
        return {"ok": False, "address": "", "detail": "AMAP_WEB_KEY not set", "key_tail": ""}
    lat = (lat or "").strip()
    lng = (lng or "").strip()
    if not lat or not lng:
        raise HTTPException(400, "missing lat/lng")
    url = "https://restapi.amap.com/v3/geocode/regeo"
    params = {
        "key": key,
        "location": f"{lng},{lat}",
        "radius": "1000",
        "extensions": "base",
        "roadlevel": "0",
    }
    try:
        async with httpx.AsyncClient(timeout=6.0) as client:
            r = await client.get(url, params=params)
            r.raise_for_status()
            data = r.json()
    except Exception as e:
        return {"ok": False, "address": "", "detail": str(e)}
    addr = ""
    try:
        if str(data.get("status")) == "1":
            rg = data.get("regeocode") or {}
            addr = (rg.get("formatted_address") or "").strip()
    except Exception:
        addr = ""
    return {
        "ok": bool(addr),
        "address": addr,
        "amap_info": data.get("info"),
        "amap_infocode": data.get("infocode"),
        "key_tail": key[-6:],
        "raw": data if not addr else None,
    }


# ─── 企业微信域名归属校验 ───
# 校验文件由企业微信后台「申请校验域名」生成，必须可通过 https://域名/WW_verify_xxx.txt 访问
@app.get("/WW_verify_f5g3FhGYiTN0VHR8.txt", response_class=Response)
async def wecom_domain_verify():
    return Response(content="f5g3FhGYiTN0VHR8", media_type="text/plain")


# ─── 小程序「普通链接二维码规则」校验文件 + 兜底网页 ───
# 微信扫码 https://dafopet.com/miniapp/* → 打开小程序 pages/index/index（TNR 申请）
# 非微信扫码（支付宝/相机）→ 落到下方兜底 HTML
@app.get("/miniapp/{filename}.txt", response_class=Response)
async def miniapp_verify_file(filename: str):
    """微信小程序后台「下载校验文件」→ 上传到 static/wechat/ → 此路由代理出来。
    防遍历：filename 限制为 [\\w\\-]+，且文件必须真实存在于 static/wechat/。
    """
    import re
    from pathlib import Path as _Path
    if not re.fullmatch(r"[\w\-]+", filename):
        return Response(status_code=404)
    p = _Path(__file__).parent.parent / "static" / "wechat" / f"{filename}.txt"
    if not p.exists():
        return Response(status_code=404)
    return Response(content=p.read_text(encoding="utf-8"), media_type="text/plain")


_MINIAPP_FALLBACK_HTML = """<!doctype html>
<html lang="zh-CN"><head><meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>大风动物医院 · 流浪猫 TNR 申请</title>
<style>
  body { font-family: Georgia, "Times New Roman", "Source Han Serif SC", serif;
         color:#1a1a1a; background:#f7f5f0; margin:0; padding:0;
         display:flex; min-height:100vh; align-items:center; justify-content:center; }
  .box { max-width:520px; padding:48px 36px; text-align:center; }
  h1 { font-size:22pt; letter-spacing:4px; margin:0 0 18px; }
  .sub { font-style:italic; color:#8a8a8a; letter-spacing:2px; font-size:11pt; margin-bottom:32px; }
  hr { border:0; height:4px; border-top:0.5px solid #1a1a1a; border-bottom:0.5px solid #1a1a1a; margin:24px auto 28px; max-width:80px; }
  p { font-size:11pt; line-height:1.85; color:#4a4a4a; letter-spacing:0.5px; }
  .tip { background:#fff; border:0.5px solid #d4d4d4; padding:18px 22px; margin-top:24px; font-size:10.5pt; line-height:1.75; }
  .tip b { letter-spacing:1.5px; }
  a { color:#1a1a1a; border-bottom:0.5px solid #8a8a8a; text-decoration:none; }
</style></head><body>
<div class="box">
  <h1>大风动物医院</h1>
  <div class="sub">始 于 2018</div>
  <hr/>
  <p>本页是小程序入口。<br/>请用 <b>微信「扫一扫」</b> 重新扫描二维码<br/>即可打开「流浪猫 TNR 申请」小程序。</p>
  <div class="tip">
    没装微信？可访问 <a href="/">在线 H5 申请页</a>，或拨打 <b>东环店 / 横岗店</b> 前台电话咨询。
  </div>
</div>
</body></html>
"""

@app.get("/miniapp", response_class=HTMLResponse)
@app.get("/miniapp/", response_class=HTMLResponse)
@app.get("/miniapp/tnr", response_class=HTMLResponse)
@app.get("/miniapp/index", response_class=HTMLResponse)
async def miniapp_fallback():
    return HTMLResponse(content=_MINIAPP_FALLBACK_HTML)


# PWA：Service Worker 必须从能 claim 根作用域的路径加载。/static/sw.js 只能
# claim /static/，所以同一个文件再从根路径 /sw.js 服务一份，scope:'/' 才合法。
@app.get("/sw.js")
async def pwa_service_worker():
    from pathlib import Path as _Path
    p = _Path(__file__).parent.parent / "static" / "sw.js"
    if not p.exists():
        return Response(status_code=404)
    return FileResponse(
        str(p),
        media_type="application/javascript",
        headers={
            "Cache-Control": "no-cache",
            "Service-Worker-Allowed": "/",
        },
    )


# ─── 企微「接收消息」回调（语音/文字 agent 入口） ───
# 配置位置：企业微信管理后台 → 自建应用 → 接收消息 → API 接收
#   URL: https://dafopet.com/wecom/callback
#   Token: settings.wecom_callback_token（你自定义）
#   EncodingAESKey: settings.wecom_callback_aes_key（43 字符）
@app.get("/wecom/callback")
async def wecom_callback_verify(
    request: Request,
    msg_signature: str = Query(""),
    timestamp: str = Query(""),
    nonce: str = Query(""),
    echostr: str = Query(""),
):
    """URL 验证：企微后台保存配置时会 GET 这个端点。"""
    from app.services.wecom_callback_crypto import verify_url, decrypt_msg, WXBizMsgCryptError
    token = settings.wecom_callback_token
    aes_key = settings.wecom_callback_aes_key
    corp_id = settings.wecom_corp_id
    if not (token and aes_key and corp_id):
        raise HTTPException(503, "wecom callback not configured")
    if not verify_url(token, msg_signature, timestamp, nonce, echostr):
        raise HTTPException(401, "signature mismatch")
    try:
        plain = decrypt_msg(echostr, aes_key, corp_id)
    except WXBizMsgCryptError as e:
        raise HTTPException(400, f"decrypt failed: {e}")
    return Response(content=plain, media_type="text/plain")


@app.post("/wecom/callback")
async def wecom_callback_receive(
    request: Request,
    msg_signature: str = Query(""),
    timestamp: str = Query(""),
    nonce: str = Query(""),
):
    """接收消息：员工通过企微发文字/语音给应用 → LLM agent 处理 → 主动 push 回复"""
    from app.services.wecom_callback_crypto import (
        verify_msg, decrypt_msg, parse_encrypt_envelope, parse_inbound_xml,
        WXBizMsgCryptError,
    )
    from app.services.wecom_agent import handle_inbound_message, push_reply
    token = settings.wecom_callback_token
    aes_key = settings.wecom_callback_aes_key
    corp_id = settings.wecom_corp_id
    if not (token and aes_key and corp_id):
        return Response(content="", media_type="text/plain")
    body = (await request.body()).decode("utf-8")
    try:
        encrypt = parse_encrypt_envelope(body)
    except Exception as e:
        logger.warning(f"[wecom callback] bad envelope: {e}")
        return Response(content="", media_type="text/plain")
    if not verify_msg(token, msg_signature, timestamp, nonce, encrypt):
        logger.warning("[wecom callback] sig mismatch")
        return Response(content="", media_type="text/plain")
    try:
        plain = decrypt_msg(encrypt, aes_key, corp_id)
        msg = parse_inbound_xml(plain)
    except WXBizMsgCryptError as e:
        logger.warning(f"[wecom callback] decrypt failed: {e}")
        return Response(content="", media_type="text/plain")
    msg_type = msg.get("MsgType", "")
    userid = msg.get("FromUserName", "")
    # 文字 → Content；语音 → Recognition（开了语音识别开关才有）
    text = ""
    if msg_type == "text":
        text = msg.get("Content", "")
    elif msg_type == "voice":
        text = msg.get("Recognition", "") or "（语音转文字失败 · 请在企微应用后台开启「语音识别」）"
    elif msg_type == "event":
        # 关注/取关 等事件，先忽略
        return Response(content="", media_type="text/plain")
    else:
        text = f"（暂不支持 {msg_type} 消息，请发文字或语音）"
    # 路由到 agent
    try:
        reply = handle_inbound_message(userid, text)
        if reply:
            push_reply(userid, reply)
    except Exception as e:
        logger.exception("[wecom agent] handle failed")
        try:
            push_reply(userid, f"❌ 内部错误：{e}")
        except Exception:
            pass
    # 被动响应空字符串即可（我们用主动 push 回消息）
    return Response(content="", media_type="text/plain")


@app.get("/", response_class=HTMLResponse)
async def page_apply(request: Request):
    return templates.TemplateResponse(
        request,
        "uk/apply.html",
        {
            "title": settings.app_name,
            "shenzhen_regions": _shenzhen_regions_embed(),
        },
    )


_CLINIC_STORES = ("大风动物医院（东环店）", "大风动物医院（横岗店）")
_ALLOWED_CLINIC_STORES = frozenset(_CLINIC_STORES)
# 短名 ↔ 全名映射（用于员工/管理员 store 字段 vs 预约/申请 store 字段）
_STORE_SHORT_TO_FULL = {"东环店": "大风动物医院（东环店）", "横岗店": "大风动物医院（横岗店）"}
_STORE_FULL_TO_SHORT = {"大风动物医院（东环店）": "东环店", "大风动物医院（横岗店）": "横岗店"}


def _get_admin_store(request: "Request") -> str:
    """返回当前登录用户被限制的门店短名（如 '东环店'）。超级管理员返回空字符串（不限）。"""
    if request.session.get("admin_role") == "superadmin":
        return ""
    return request.session.get("admin_store", "")


def _apply_store_filter(query, store_field, store_short: str):
    """给「目录类」表加门店可见性过滤。

    约定：store 字段为空字符串 = 通用，两店共享。
      - staff（store_short 非空）→ 看到 本店 + 通用
      - superadmin（store_short 空）→ 看到全部，不过滤

    使用例：
      q = _apply_store_filter(q, InventoryItem.store, _get_admin_store(request))
    """
    if not store_short:
        return query
    return query.filter(or_(store_field == store_short, store_field == ""))


def _get_op_store(request: "Request") -> str:
    """日常开单场景的"当前操作门店"。所有角色（含超管）都按 session.admin_store 过滤。
    与 _get_admin_store 的区别：超管不再返回空。
    用法：处方 / 检查 / 销售 / 美容 / 疫苗 等单据的品目选择，必须按此过滤，
    避免横岗店登录的超管开单时选到东环店的库存。
    """
    return request.session.get("admin_store", "") or ""


def _resolve_store_for_create(request: "Request", explicit: str = "") -> str:
    """决定新建「目录类」记录归属哪个门店。

    优先级：
      1. 表单显式传了 explicit（superadmin 在切换器里选了具体门店）
      2. staff 自动归本店
    校验：必须是「东环店」或「横岗店」之一，不再允许空字符串（通用）。
    """
    explicit = (explicit or "").strip()
    if request.session.get("admin_role") == "superadmin":
        # superadmin 必须显式选一家店
        if explicit in ("东环店", "横岗店"):
            return explicit
        # 没选 → 退到 session.admin_store；仍为空 → 抛错
        fallback = request.session.get("admin_store", "") or ""
        if fallback in ("东环店", "横岗店"):
            return fallback
        raise HTTPException(400, "请选择归属门店：东环店 或 横岗店（每个品目必须归属一家店）")
    # staff：永远是本店
    fallback = request.session.get("admin_store", "") or ""
    if fallback not in ("东环店", "横岗店"):
        raise HTTPException(400, "员工账号缺少有效门店归属，请联系管理员")
    return fallback


def _age_estimate_to_birthday(s: str) -> str:
    """把 "3岁" / "6个月" / "1岁半" 等自由文本年龄估算转成 YYYY-MM-DD（按今天倒推）。

    无法识别 → 返回空字符串（保持空比存脏数据好，编辑表单能正常显示）。
    """
    import re
    from datetime import date as _date, timedelta
    if not s:
        return ""
    raw = str(s).strip()
    if not raw:
        return ""
    # 中文数字映射（简单覆盖 1-10）
    cn_num = {"一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "七": 7, "八": 8, "九": 9, "十": 10}
    for cn, n in cn_num.items():
        raw = raw.replace(cn, str(n))
    today = _date.today()
    # 「3岁半」「1岁半」→ years + 0.5
    half = ".5" if "半" in raw else ""
    # 「6个月」「8 个月」
    m_month = re.search(r"(\d+(?:\.\d+)?)\s*个?月", raw)
    # 「3岁」「3 岁」「3 年」
    m_year = re.search(r"(\d+(?:\.\d+)?)\s*[岁年]", raw)
    if m_year:
        try:
            yrs = float(m_year.group(1) + half)
            days = int(yrs * 365.25)
            d = today - timedelta(days=days)
            return d.strftime("%Y-%m-%d")
        except Exception:
            return ""
    if m_month:
        try:
            mos = float(m_month.group(1))
            days = int(mos * 30.44)
            d = today - timedelta(days=days)
            return d.strftime("%Y-%m-%d")
        except Exception:
            return ""
    # 已经是 YYYY-MM-DD / YYYY-MM 格式 → 直接返回
    if re.match(r"^\d{4}-\d{1,2}(?:-\d{1,2})?$", raw):
        parts = raw.split("-")
        y = parts[0]
        mo = parts[1].zfill(2) if len(parts) > 1 else "01"
        da = parts[2].zfill(2) if len(parts) > 2 else "01"
        return f"{y}-{mo}-{da}"
    return ""


# 门店首字母（病历号前缀）
_STORE_INITIAL = {"东环店": "D", "横岗店": "H"}


def _gen_medical_record_no(db: "Session", store: str) -> str:
    """生成病历号：{门店首字母}C{YY}{MM}{5位序号}。
    例：DC2605 00001 / HC2605 00012。当月内同店递增。
    """
    letter = _STORE_INITIAL.get(store, "X")
    now = datetime.utcnow()
    prefix = f"{letter}C{now.strftime('%y%m')}"
    # 找当月最大序号
    from sqlalchemy import desc as _desc
    last = (
        db.query(Pet.medical_record_no)
        .filter(Pet.medical_record_no.like(prefix + "%"))
        .order_by(_desc(Pet.medical_record_no))
        .first()
    )
    seq = 1
    if last and last[0]:
        try:
            seq = int(last[0][len(prefix):]) + 1
        except (ValueError, TypeError):
            seq = 1
    return f"{prefix}{seq:05d}"
_ALLOWED_APPOINTMENT_CATEGORIES = frozenset({x.value for x in AppointmentCategory})
_ALLOWED_APPOINTMENT_STATUSES = frozenset({x.value for x in AppointmentStatus})
_APPOINTMENT_CATEGORY_LABELS = {
    AppointmentCategory.tnr.value: "TNR 预约",
    AppointmentCategory.outpatient.value: "门诊预约",
    AppointmentCategory.surgery.value: "手术预约",
    AppointmentCategory.beauty.value: "美容预约",
    AppointmentCategory.grooming.value: "造型预约",   # 历史兼容
    AppointmentCategory.washcare.value: "洗护预约",   # 历史兼容
}
_APPOINTMENT_STATUS_LABELS = {
    AppointmentStatus.pending.value: "待确认",
    AppointmentStatus.confirmed.value: "已确认",
    AppointmentStatus.arrived.value: "已到店",
    AppointmentStatus.completed.value: "已完成",
    AppointmentStatus.cancelled.value: "已取消",
    AppointmentStatus.no_show.value: "未到店",
}
_PET_GENDER_LABELS = {"male": "公", "female": "母", "unknown": "未知"}
_APPOINTMENT_BOOKING_MAX_DAYS_AHEAD = 30


def _assert_appointment_fields(
    *,
    category: str,
    service_name: str,
    customer_name: str,
    phone: str,
    pet_name: str,
    pet_gender: str,
    store: str,
    appointment_date: str,
    appointment_time: str,
    notes: str,
    duration_minutes: str,
) -> dict[str, str | int]:
    def need(label: str, raw: str, max_len: int) -> str:
        s = (raw or "").strip()
        if not s:
            raise HTTPException(400, f"请填写{label}。")
        if len(s) > max_len:
            raise HTTPException(400, f"{label}过长。")
        return s

    out: dict[str, str | int] = {}
    cat = (category or "").strip()
    if cat not in _ALLOWED_APPOINTMENT_CATEGORIES:
        raise HTTPException(400, "请选择有效的预约类型。")
    out["category"] = cat
    out["service_name"] = need("预约项目", service_name, 120)
    out["customer_name"] = need("联系人姓名", customer_name, 120)
    phone_v = need("手机号", phone, 40)
    if not re.fullmatch(r"1\d{10}", phone_v):
        raise HTTPException(400, "请填写 11 位中国大陆手机号。")
    out["phone"] = phone_v
    out["pet_name"] = need("宠物/流浪猫名称", pet_name, 120)
    g = (pet_gender or "").strip().lower()
    if g not in ("male", "female", "unknown"):
        raise HTTPException(400, "请选择性别。")
    out["pet_gender"] = g
    store_v = need("门店", store, 120)
    if store_v not in _ALLOWED_CLINIC_STORES:
        raise HTTPException(400, "请选择有效的预约门店。")
    out["store"] = store_v
    date_v = need("预约日期", appointment_date, 20)
    try:
        date_obj = datetime.strptime(date_v, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "预约日期格式应为 YYYY-MM-DD。")
    if date_obj < datetime.now().date():
        raise HTTPException(400, "预约日期不能早于今天。")
    out["appointment_date"] = date_v
    time_v = need("预约时间", appointment_time, 20)
    if not re.fullmatch(r"([01]\d|2[0-3]):[0-5]\d", time_v):
        raise HTTPException(400, "预约时间格式应为 HH:MM。")
    out["appointment_time"] = time_v
    out["notes"] = (notes or "").strip()[:2000]
    try:
        dur = int((duration_minutes or "30").strip() or "30")
    except ValueError:
        raise HTTPException(400, "时长应为分钟数。")
    if dur < 10 or dur > 480:
        raise HTTPException(400, "时长范围应在 10 到 480 分钟之间。")
    out["duration_minutes"] = dur
    # 与小程序 TNR 预约一致：后台/API 侧也固定项目与时长，避免手改表单绕过
    if out["category"] == AppointmentCategory.tnr.value:
        out["service_name"] = "TNR 手术安排"
        out["duration_minutes"] = 60
    return out


def _mask_phone(phone: str) -> str:
    t = (phone or "").strip()
    if len(t) < 7:
        return t
    return t[:3] + "****" + t[-4:]


_TNR_TIME_START = "11:00"
_TNR_TIME_END = "18:00"
_TNR_DAILY_MAX = 2
_TNR_MONTHLY_QUOTA = 30          # 每月已确认 TNR 预约上限
_TNR_NOSHOW_BAN_COUNT = 3        # 单月爽约次数达到此值触发封禁
_TNR_NOSHOW_BAN_DAYS  = 90       # 封禁天数


def _time_to_minutes(t: str) -> int:
    h, m = t.split(":")
    return int(h) * 60 + int(m)


_BEAUTY_TRACK = {"beauty", "grooming", "washcare"}
_MEDICAL_TRACK = {"tnr", "surgery", "outpatient"}


def _appt_track(category: str) -> str:
    """预约线归类：美容线 vs 医疗线（不同员工不同场地，两线并行不互相挡）。"""
    if category in _BEAUTY_TRACK:
        return "beauty"
    if category in _MEDICAL_TRACK:
        return "medical"
    return "other"


def _check_appointment_conflict(
    db: "Session",
    store: str,
    appointment_date: str,
    appointment_time: str,
    duration_minutes: int,
    exclude_id: int | None = None,
    category: str = "",
) -> "Appointment | None":
    """检查同门店同日同业务线是否存在时间重叠的预约（排除已取消/爽约）。
    业务线划分：
      - 美容线：beauty / grooming / washcare（美容师独立场地）
      - 医疗线：tnr / surgery / outpatient（医生独立场地）
    跨线并行不冲突。
    重叠返回冲突记录，否则返回 None。
    """
    new_track = _appt_track(category)
    q = (
        db.query(Appointment)
        .filter(
            Appointment.store == store,
            Appointment.appointment_date == appointment_date,
            Appointment.status.notin_([AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value]),
        )
    )
    if exclude_id is not None:
        q = q.filter(Appointment.id != exclude_id)
    existing = q.all()
    new_start = _time_to_minutes(appointment_time)
    new_end = new_start + duration_minutes
    for appt in existing:
        # 跨线并行 → 不挡
        if new_track and _appt_track(appt.category) != new_track:
            continue
        a_start = _time_to_minutes(appt.appointment_time)
        a_end = a_start + appt.duration_minutes
        if new_start < a_end and new_end > a_start:
            return appt
    return None


def _get_tnr_store_config(db: "Session", store_name: str) -> "TnrStoreConfig":
    """获取（或自动创建）门店 TNR 配额配置。"""
    cfg = db.query(TnrStoreConfig).filter(TnrStoreConfig.store_name == store_name).first()
    if cfg is None:
        cfg = TnrStoreConfig(
            store_name=store_name,
            tnr_monthly_quota=_TNR_MONTHLY_QUOTA,
            tnr_accepting=True,
        )
        db.add(cfg)
        db.commit()
        db.refresh(cfg)
    return cfg


def _get_tnr_monthly_confirmed_count(db: "Session", store: str, year_month: str) -> int:
    """统计指定门店当月已确认（未取消/爽约）的 TNR 预约数。year_month 格式 YYYY-MM。"""
    _counted_statuses = [
        AppointmentStatus.confirmed.value,
        AppointmentStatus.arrived.value,
        AppointmentStatus.completed.value,
    ]
    return (
        db.query(Appointment)
        .filter(
            Appointment.category == AppointmentCategory.tnr.value,
            Appointment.store == store,
            Appointment.appointment_date.like(f"{year_month}%"),
            Appointment.status.in_(_counted_statuses),
        )
        .count()
    )


def _get_phone_noshow_ban_until(db: "Session", phone: str) -> "datetime | None":
    """
    检查该手机号是否因本月爽约 ≥3 次而被封禁。
    返回封禁截止日期（date 对象），如未被封禁返回 None。
    """
    from collections import defaultdict
    no_shows = (
        db.query(Appointment)
        .filter(
            Appointment.phone == phone,
            Appointment.category == AppointmentCategory.tnr.value,
            Appointment.status == AppointmentStatus.no_show.value,
        )
        .order_by(Appointment.appointment_date)
        .all()
    )
    by_month: dict[str, list] = defaultdict(list)
    for appt in no_shows:
        ym = (appt.appointment_date or "")[:7]
        if ym:
            by_month[ym].append(appt)
    ban_until = None
    for ym, appts in by_month.items():
        if len(appts) >= _TNR_NOSHOW_BAN_COUNT:
            appts_sorted = sorted(appts, key=lambda a: a.appointment_date)
            trigger_date_str = appts_sorted[_TNR_NOSHOW_BAN_COUNT - 1].appointment_date
            try:
                trigger_date = datetime.strptime(trigger_date_str, "%Y-%m-%d").date()
                candidate = trigger_date + timedelta(days=_TNR_NOSHOW_BAN_DAYS)
                if ban_until is None or candidate > ban_until:
                    ban_until = candidate
            except Exception:
                pass
    return ban_until


def _check_tnr_constraints(
    db: "Session",
    category: str,
    store: str,
    appointment_date: str,
    appointment_time: str,
    phone: str = "",
    exclude_id: int | None = None,
) -> str | None:
    """TNR 预约专项校验：时间段限制、每日上限、月度配额、门店开关、爽约封禁。违规返回错误字符串，通过返回 None。"""
    if category != AppointmentCategory.tnr.value:
        return None
    # 时间段校验
    t_minutes = _time_to_minutes(appointment_time)
    start_minutes = _time_to_minutes(_TNR_TIME_START)
    end_minutes = _time_to_minutes(_TNR_TIME_END)
    if t_minutes < start_minutes or t_minutes >= end_minutes:
        return f"TNR 手术预约时间须在 {_TNR_TIME_START}–{_TNR_TIME_END} 之间，请重新选择。"
    # 每店每日上限
    q = (
        db.query(Appointment)
        .filter(
            Appointment.category == AppointmentCategory.tnr.value,
            Appointment.store == store,
            Appointment.appointment_date == appointment_date,
            Appointment.status.notin_([AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value]),
        )
    )
    if exclude_id is not None:
        q = q.filter(Appointment.id != exclude_id)
    count = q.count()
    if count >= _TNR_DAILY_MAX:
        return f"该门店 {appointment_date} TNR 手术预约已约满（每日上限 {_TNR_DAILY_MAX} 个），请改约其他日期。"
    # 门店手动开关校验
    cfg = _get_tnr_store_config(db, store)
    if not cfg.tnr_accepting:
        return f"该门店 TNR 预约暂停接受，请稍后再试或联系门店咨询。"
    # 月度配额校验
    year_month = appointment_date[:7] if appointment_date and len(appointment_date) >= 7 else datetime.now().strftime("%Y-%m")
    monthly_count = _get_tnr_monthly_confirmed_count(db, store, year_month)
    quota = cfg.tnr_monthly_quota
    if monthly_count >= quota:
        return f"该门店 {year_month} 月 TNR 预约已达上限（{quota} 个），如有疑问请联系门店。"
    # 爽约封禁校验
    if phone:
        ban_until = _get_phone_noshow_ban_until(db, phone)
        if ban_until is not None and ban_until >= datetime.now().date():
            return (
                f"您本月 TNR 爽约次数已达 {_TNR_NOSHOW_BAN_COUNT} 次，账户已被限制至 "
                f"{ban_until.strftime('%Y-%m-%d')} 前无法提交新的 TNR 预约，如有疑问请联系门店。"
            )
    return None


def _check_duplicate_application_appointment(
    db: "Session",
    related_application_id: int | None,
    exclude_id: int | None = None,
) -> str | None:
    """检查同一 TNR 申请编号是否已有有效预约（非取消/爽约）。重复则返回错误字符串，否则返回 None。"""
    if not related_application_id:
        return None
    q = (
        db.query(Appointment)
        .filter(
            Appointment.related_application_id == related_application_id,
            Appointment.status.notin_([AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value]),
        )
    )
    if exclude_id is not None:
        q = q.filter(Appointment.id != exclude_id)
    existing = q.first()
    if existing:
        return (
            f"申请 #{related_application_id} 已存在有效预约（预约 #{existing.id}，"
            f"状态：{existing.status}），请先取消原预约再重新预约。"
        )
    return None


async def _resolve_wechat_openid(payload: dict) -> str:
    openid = ((payload or {}).get("openid", "") or "").strip()
    code = ((payload or {}).get("code", "") or "").strip()
    if not openid and code:
        data = wechat_code2session(code)
        openid = (data.get("openid", "") or "").strip()
    if not openid:
        raise HTTPException(400, "missing openid")
    return openid


def _appointment_catalog() -> dict:
    today = datetime.now().date()
    return {
        "stores": list(_CLINIC_STORES),
        "booking_window": {
            "start_date": today.strftime("%Y-%m-%d"),
            "max_days_ahead": _APPOINTMENT_BOOKING_MAX_DAYS_AHEAD,
            "suggestion": "请使用「预约日期 / 预约时间」自主选择到院时段；后续可按门店、医生与服务能力细化排班规则。",
        },
        "categories": [
            {
                "value": AppointmentCategory.tnr.value,
                "label": "流浪动物 TNR",
                "description": "适合流浪动物初诊评估、TNR 手术安排和术后复诊。",
                "booking_tip": "请使用下方「预约日期 / 预约时间」自主选择到院时段；不再展示固定建议时段列表。",
                "supports_related_application": True,
                "time_slots": [],
                "services": [
                    {
                        "name": "TNR 手术安排",
                        "duration_minutes": 60,
                        "description": "用于确认手术时间、门店与到院前准备事项。",
                    },
                ],
            },
            {
                "value": AppointmentCategory.outpatient.value,
                "label": "常规门诊",
                "description": "适合普通门诊、复诊、健康检查及疫苗驱虫等咨询。",
                "booking_tip": "门诊可预约时间：10:00 – 21:00（上午需护理住院动物，晚上 21:00 后不接受新预约）。疫苗/驱虫 30 分钟，其余科目 60 分钟。",
                "supports_related_application": False,
                "time_range": {"start": "10:00", "end": "21:00"},
                "time_slots": [],
                "services": [
                    {"name": "疫苗/驱虫",  "duration_minutes": 30},
                    {"name": "体检",      "duration_minutes": 60},
                    {"name": "呼吸道",    "duration_minutes": 60},
                    {"name": "胃肠道",    "duration_minutes": 60},
                    {"name": "泌尿道",    "duration_minutes": 60},
                    {"name": "皮肤",      "duration_minutes": 60},
                    {"name": "口腔",      "duration_minutes": 60},
                    {"name": "行动异常",   "duration_minutes": 60},
                    {"name": "心内科",    "duration_minutes": 60},
                    {"name": "肾内科",    "duration_minutes": 60},
                ],
            },
            {
                "value": AppointmentCategory.surgery.value,
                "label": "手术预约",
                "description": "适合绝育、骨科、软组织、眼科、神经外科等各类手术预约。",
                "booking_tip": "请使用下方「预约日期 / 预约时间」自主选择到院时段；后续可按手术台与麻醉安排细化。",
                "supports_related_application": False,
                "time_slots": [],
                "services": [
                    {"name": "绝育手术",    "duration_minutes": 60,  "description": "用于常规绝育手术预约。"},
                    {"name": "骨科手术",    "duration_minutes": 120, "description": "用于骨折、关节、脊椎等骨科手术预约。"},
                    {"name": "软组织手术",  "duration_minutes": 90,  "description": "用于皮肤、肿瘤切除、消化道等软组织手术预约。"},
                    {"name": "眼科手术",    "duration_minutes": 60,  "description": "用于眼睑、角膜、晶体等眼科手术预约。"},
                    {"name": "神经外科手术","duration_minutes": 120, "description": "用于脑部、脊髓等神经外科手术预约。"},
                ],
            },
            {
                "value": AppointmentCategory.beauty.value,
                "label": "美容预约",
                "description": "适合猫/犬洗护与造型服务，支持附加项目选择。",
                "booking_tip": "请选择美容项目及附加服务，并提供宠物体型与毛发信息，以便安排合适的美容师与时段。",
                "supports_related_application": False,
                "time_slots": [],
                "services": [
                    {"name": "猫洗护", "duration_minutes": 60},
                    {"name": "猫造型", "duration_minutes": 60},
                    {"name": "犬洗护", "duration_minutes": 90},
                    {"name": "犬造型", "duration_minutes": 90},
                ],
                "addon_options": ["去浮毛", "SPA", "护发素", "纯手剪", "药浴", "去油"],
                "size_options": ["微小型犬（4kg 以下）", "小型犬（4–10kg）", "中型犬（10–15kg）", "中大型犬（15–25kg）", "大型犬（25kg 以上）"],
                "coat_options": ["长毛", "短毛"],
            },
        ],
        "pet_genders": [
            {"value": "female", "label": "母"},
            {"value": "male", "label": "公"},
            {"value": "unknown", "label": "未知"},
        ],
    }


def _serialize_appointment(row: Appointment) -> dict:
    return {
        "id": row.id,
        "category": row.category,
        "category_zh": _APPOINTMENT_CATEGORY_LABELS.get(row.category, row.category),
        "status": row.status,
        "status_zh": _APPOINTMENT_STATUS_LABELS.get(row.status, row.status),
        "service_name": row.service_name,
        "customer_name": row.customer_name,
        "phone_masked": _mask_phone(row.phone),
        "pet_name": row.pet_name,
        "pet_gender": row.pet_gender,
        "pet_gender_zh": _PET_GENDER_LABELS.get(row.pet_gender, row.pet_gender),
        "store": row.store,
        "appointment_date": row.appointment_date,
        "appointment_time": row.appointment_time,
        "duration_minutes": row.duration_minutes,
        "notes": row.notes or "",
        "related_application_id": row.related_application_id,
        "pet_size": row.pet_size or "",
        "coat_length": row.coat_length or "",
        "addon_services": row.addon_services or "",
        "created_at": row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        "updated_at": row.updated_at.strftime("%Y-%m-%d %H:%M") if row.updated_at else "",
    }


def _assert_application_form_fields(
    *,
    applicant_name: str,
    phone: str,
    address: str,
    clinic_store: str,
    appointment_at: str,
    post_surgery_plan: str,
    id_number: str,
    cat_nickname: str,
    cat_gender: str,
    age_estimate: str,
    health_note: str,
) -> dict[str, str]:
    def need(label: str, raw: str, max_len: int) -> str:
        s = (raw or "").strip()
        if not s:
            raise HTTPException(400, f"请填写{label}。")
        if len(s) > max_len:
            raise HTTPException(400, f"{label}过长。")
        return s

    out: dict[str, str] = {}
    out["applicant_name"] = need("申请人姓名", applicant_name, 120)
    out["phone"] = need("手机号", phone, 40)
    if not re.fullmatch(r"1\d{10}", out["phone"]):
        raise HTTPException(400, "请填写 11 位中国大陆手机号。")
    out["address"] = need("完整地址", address, 500)
    cs = need("预约门店", clinic_store, 80)
    if cs not in _ALLOWED_CLINIC_STORES:
        raise HTTPException(400, "请选择有效的预约门店。")
    out["clinic_store"] = cs
    out["appointment_at"] = (appointment_at or "").strip()[:40]  # 可选字段
    out["post_surgery_plan"] = need("术后打算", post_surgery_plan, 120)
    idn_raw = need("身份证号", id_number, 40)
    idn = idn_raw.upper()
    if len(idn) == 18:
        if not re.fullmatch(r"\d{17}[\dX]", idn):
            raise HTTPException(400, "请填写 18 位身份证号（末位可为 X）。")
    elif len(idn) == 15:
        if not idn.isdigit():
            raise HTTPException(400, "请填写 15 位身份证号。")
    else:
        raise HTTPException(400, "请填写 15 或 18 位身份证号。")
    out["id_number"] = idn
    out["cat_nickname"] = need("流浪猫名字", cat_nickname, 120)
    g = (cat_gender or "").strip().lower()
    if g not in ("male", "female", "unknown"):
        raise HTTPException(400, "请选择猫咪性别。")
    out["cat_gender"] = g
    out["age_estimate"] = need("年龄估计", age_estimate, 80)
    out["health_note"] = need("流浪状况说明", health_note, 8000)
    return out


def _count_valid_apply_images(images: list[UploadFile]) -> int:
    n = 0
    for uf in images:
        if not uf.filename:
            continue
        ext = Path(uf.filename).suffix.lower() or ".jpg"
        if ext in (".jpg", ".jpeg", ".png", ".webp"):
            n += 1
    return n


def _assert_application_row_complete(row: Application) -> None:
    _assert_application_form_fields(
        applicant_name=row.applicant_name,
        phone=row.phone,
        address=row.address,
        clinic_store=row.clinic_store,
        appointment_at=row.appointment_at,
        post_surgery_plan=row.post_surgery_plan,
        id_number=row.id_number,
        cat_nickname=row.cat_nickname,
        cat_gender=row.cat_gender,
        age_estimate=row.age_estimate,
        health_note=row.health_note,
    )


@app.post("/api/apply")
async def api_apply(
    request: Request,
    db: Session = Depends(get_db),
    applicant_name: str = Form(...),
    phone: str = Form(...),
    address: str = Form(...),
    clinic_store: str = Form(""),
    appointment_at: str = Form(""),
    location_lat: str = Form(""),
    location_lng: str = Form(""),
    location_address: str = Form(""),
    id_number: str = Form(""),
    post_surgery_plan: str = Form(""),
    cat_nickname: str = Form(""),
    cat_gender: str = Form(...),
    cat_breed: str = Form(""),
    cat_color: str = Form(""),
    age_estimate: str = Form(""),
    health_note: str = Form(""),
    wechat_openid: str = Form(""),
    agree_ear_tip: str = Form("false"),
    agree_no_pet_fraud: str = Form("false"),
    images: Annotated[Optional[list[UploadFile]], File()] = None,
    videos: Annotated[Optional[list[UploadFile]], File()] = None,
):
    images = images or []
    videos = videos or []
    ok_ear = agree_ear_tip.lower() in ("true", "1", "on", "yes")
    ok_fraud = agree_no_pet_fraud.lower() in ("true", "1", "on", "yes")
    if not ok_ear or not ok_fraud:
        raise HTTPException(400, "请勾选同意剪耳标记与承诺非家养猫冒充。")
    if _count_valid_apply_images(images) < 1:
        raise HTTPException(400, "请至少上传 1 张申请照片。")

    f = _assert_application_form_fields(
        applicant_name=applicant_name,
        phone=phone,
        address=address,
        clinic_store=clinic_store,
        appointment_at=appointment_at,
        post_surgery_plan=post_surgery_plan,
        id_number=id_number,
        cat_nickname=cat_nickname,
        cat_gender=cat_gender,
        age_estimate=age_estimate,
        health_note=health_note,
    )
    # 品种和颜色（可选）
    f["cat_breed"] = (cat_breed or "").strip()[:80]
    f["cat_color"] = (cat_color or "").strip()[:80]

    app_row = Application(
        applicant_name=f["applicant_name"],
        phone=f["phone"],
        wechat_openid=wechat_openid.strip(),
        clinic_store=f["clinic_store"],
        appointment_at=f["appointment_at"],
        location_lat=location_lat.strip(),
        location_lng=location_lng.strip(),
        location_address=location_address.strip()[:400],
        id_number=f["id_number"],
        post_surgery_plan=f["post_surgery_plan"],
        address=f["address"],
        cat_nickname=f["cat_nickname"],
        cat_gender=f["cat_gender"],
        cat_breed=f.get("cat_breed", ""),
        cat_color=f.get("cat_color", ""),
        age_estimate=f["age_estimate"],
        weight_estimate="",
        health_note=f["health_note"],
        agree_ear_tip=ok_ear,
        agree_no_pet_fraud=ok_fraud,
        status=ApplicationStatus.pending_ai.value,
    )
    db.add(app_row)
    db.flush()

    aid = app_row.id
    base = Path(settings.upload_dir) / str(aid)
    base.mkdir(parents=True, exist_ok=True)

    image_paths: list[Path] = []
    video_paths: list[Path] = []

    for uf in images:
        if not uf.filename:
            continue
        ext = Path(uf.filename).suffix.lower() or ".jpg"
        if ext not in (".jpg", ".jpeg", ".png", ".webp"):
            continue
        dest = base / f"app_img_{secrets.token_hex(6)}{ext}"
        dest.write_bytes(await uf.read())
        db.add(
            MediaFile(
                application_id=aid,
                kind=MediaKind.application_image.value,
                stored_path=str(dest),
                original_name=uf.filename,
            )
        )
        image_paths.append(dest)

    for uf in videos:
        if not uf.filename:
            continue
        ext = Path(uf.filename).suffix.lower() or ".mp4"
        if ext not in (".mp4", ".webm", ".mov", ".mkv"):
            continue
        dest = base / f"app_vid_{secrets.token_hex(6)}{ext}"
        dest.write_bytes(await uf.read())
        db.add(
            MediaFile(
                application_id=aid,
                kind=MediaKind.application_video.value,
                stored_path=str(dest),
                original_name=uf.filename,
            )
        )
        video_paths.append(dest)

    db.commit()
    db.refresh(app_row)

    ai_result = await review_application_media(image_paths, video_paths)

    app_row.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
    app_row.ai_is_likely_stray = ai_result.get("is_likely_stray")
    conf = ai_result.get("confidence")
    app_row.ai_confidence = float(conf) if conf is not None else None

    new_status, auto_ok = apply_auto_status_from_ai(ai_result)
    app_row.status = new_status
    db.commit()

    if auto_ok:
        notify_application_result(db, aid, app_row.phone, app_row.applicant_name, approved=True, extra="系统根据图像辅助判断已完成预审通过，到院后仍需工作人员核对猫只身份。")
        push_application_result(
            db,
            application_id=aid,
            openid=app_row.wechat_openid,
            applicant_name=app_row.applicant_name,
            status_text="审核已通过",
            phone_masked=app_row.phone,
            note="请按约定时间携带猫咪到院",
            submitted_at=app_row.created_at.strftime("%Y-%m-%d %H:%M") if app_row.created_at else "",
            action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
    elif new_status == ApplicationStatus.pre_approved.value:
        push_application_result(
            db,
            application_id=aid,
            openid=app_row.wechat_openid,
            applicant_name=app_row.applicant_name,
            status_text="预通过（待复核）",
            phone_masked=app_row.phone,
            note="医院将尽快人工复核，请保持手机畅通",
            submitted_at=app_row.created_at.strftime("%Y-%m-%d %H:%M") if app_row.created_at else "",
            action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
    elif new_status == ApplicationStatus.pending_manual.value:
        push_pending_manual_notice(
            db,
            application_id=aid,
            openid=app_row.wechat_openid,
            applicant_name=app_row.applicant_name,
            submitted_at=app_row.created_at.strftime("%Y-%m-%d %H:%M") if app_row.created_at else "",
        )
        # 同步推送到员工企业微信
        try:
            from app.services import wecom_notify as _wn
            _wn.notify_tnr_pending_manual(db, app_row)
        except Exception as _e:
            logger.warning("[wecom] notify_tnr_pending_manual failed: %s", _e)

    return {
        "id": aid,
        "status": app_row.status,
        "ai_summary": {
            "is_likely_stray": app_row.ai_is_likely_stray,
            "confidence": app_row.ai_confidence,
            "auto_approved": auto_ok,
        },
        "message": "提交成功。若未自动通过，请耐心等待医院人工审核，通过后将以登记方式通知您。"
    }


@app.post("/api/apply/create")
async def api_apply_create(
    request: Request,
    db: Session = Depends(get_db),
    applicant_name: str = Form(...),
    phone: str = Form(...),
    address: str = Form(...),
    clinic_store: str = Form(""),
    appointment_at: str = Form(""),
    location_lat: str = Form(""),
    location_lng: str = Form(""),
    location_address: str = Form(""),
    id_number: str = Form(""),
    post_surgery_plan: str = Form(""),
    cat_nickname: str = Form(""),
    cat_gender: str = Form(...),
    cat_breed: str = Form(""),
    cat_color: str = Form(""),
    age_estimate: str = Form(""),
    health_note: str = Form(""),
    wechat_openid: str = Form(""),
    agree_ear_tip: str = Form("false"),
    agree_no_pet_fraud: str = Form("false"),
    is_proxy: str = Form(""),
    proxy_name: str = Form(""),
    proxy_phone: str = Form(""),
    proxy_relation: str = Form(""),
):
    ok_ear = agree_ear_tip.lower() in ("true", "1", "on", "yes")
    ok_fraud = agree_no_pet_fraud.lower() in ("true", "1", "on", "yes")
    if not ok_ear or not ok_fraud:
        raise HTTPException(400, "请勾选同意剪耳标记与承诺非家养猫冒充。")

    f = _assert_application_form_fields(
        applicant_name=applicant_name,
        phone=phone,
        address=address,
        clinic_store=clinic_store,
        appointment_at=appointment_at,
        post_surgery_plan=post_surgery_plan,
        id_number=id_number,
        cat_nickname=cat_nickname,
        cat_gender=cat_gender,
        age_estimate=age_estimate,
        health_note=health_note,
    )
    # 品种和颜色（可选，不强制校验）
    f["cat_breed"] = (cat_breed or "").strip()[:80]
    f["cat_color"] = (cat_color or "").strip()[:80]

    # ── 重复提交检测 ──
    _DUP_STATUS_ZH = {
        "draft": "草稿", "pending_ai": "审核中", "pending_manual": "待人工审核",
        "pre_approved": "预通过", "approved": "已通过", "scheduled": "已预约",
        "no_show": "爽约", "arrived_verified": "已到店",
    }

    # 1a. 清理同手机号的遗留草稿（网络中断/关闭小程序导致未完成的提交）
    old_drafts = (
        db.query(Application)
        .filter(Application.phone == f["phone"])
        .filter(Application.status == ApplicationStatus.draft.value)
        .all()
    )
    for _d in old_drafts:
        db.delete(_d)
    if old_drafts:
        db.commit()

    # 1b. 审核进行中时（pending_ai 之后）不允许再提交新申请
    _PENDING_STATUSES = [
        ApplicationStatus.pending_ai.value,
        ApplicationStatus.pending_manual.value,
        ApplicationStatus.pre_approved.value,
    ]
    pending_dup = (
        db.query(Application)
        .filter(Application.phone == f["phone"])
        .filter(Application.status.in_(_PENDING_STATUSES))
        .order_by(Application.id.desc())
        .first()
    )
    if pending_dup:
        status_label = _DUP_STATUS_ZH.get(pending_dup.status, pending_dup.status)
        raise HTTPException(
            409,
            f"您已有一份申请正在审核中（编号 #{pending_dup.id}，当前状态：{status_label}），"
            f"请等待审核通过后再提交新的申请。如需取消，请联系医院前台。",
        )

    # 2. 同一手机号 + 同一猫咪名称，不能重复提交（终结状态除外）
    _TERMINAL_STATUSES = [
        ApplicationStatus.rejected.value,
        ApplicationStatus.cancelled.value,
        ApplicationStatus.surgery_completed.value,
    ]
    same_cat_dup = (
        db.query(Application)
        .filter(Application.phone == f["phone"])
        .filter(Application.cat_nickname == f["cat_nickname"])
        .filter(Application.status.notin_(_TERMINAL_STATUSES))
        .order_by(Application.id.desc())
        .first()
    )
    if same_cat_dup:
        status_label = _DUP_STATUS_ZH.get(same_cat_dup.status, same_cat_dup.status)
        raise HTTPException(
            409,
            f"「{f['cat_nickname']}」已有进行中的申请（编号 #{same_cat_dup.id}，当前状态：{status_label}），"
            f"请勿为同一只猫重复提交申请。",
        )

    # ── 自动创建/合并客户档案 ──
    try:
        _cust = _upsert_customer(
            db,
            name=f["applicant_name"],
            phone=f["phone"],
            openid=wechat_openid.strip(),
            id_number=f["id_number"],
            address=f["address"],
            source="tnr",
        )
        _cust_id = _cust.id
    except Exception:
        _cust_id = None

    app_row = Application(
        applicant_name=f["applicant_name"],
        phone=f["phone"],
        wechat_openid=wechat_openid.strip(),
        clinic_store=f["clinic_store"],
        appointment_at=f["appointment_at"],
        location_lat=location_lat.strip(),
        location_lng=location_lng.strip(),
        location_address=location_address.strip()[:400],
        id_number=f["id_number"],
        post_surgery_plan=f["post_surgery_plan"],
        address=f["address"],
        cat_nickname=f["cat_nickname"],
        cat_gender=f["cat_gender"],
        cat_breed=f.get("cat_breed", ""),
        cat_color=f.get("cat_color", ""),
        age_estimate=f["age_estimate"],
        weight_estimate="",
        health_note=f["health_note"],
        agree_ear_tip=ok_ear,
        agree_no_pet_fraud=ok_fraud,
        is_proxy=is_proxy.lower() in ("true", "1", "on", "yes"),
        proxy_name=proxy_name.strip()[:120],
        proxy_phone=proxy_phone.strip()[:40],
        proxy_relation=proxy_relation.strip()[:40],
        status=ApplicationStatus.draft.value,
        customer_id=_cust_id,
    )
    db.add(app_row)
    db.commit()
    db.refresh(app_row)

    # ── 自动创建/复用宠物档案 ──
    # 同一客户同名流浪猫（特别是申请后取消又重新申请的）应复用同一条 Pet，
    # 避免客户宠物档案出现多个重复的「彪 / 小灰灰 …」
    if _cust_id and f.get("cat_nickname"):
        try:
            _cat_name = f["cat_nickname"].strip()[:120]
            # 把 "3岁" / "6个月" 等自由文本转成「出生日期」估算（YYYY-MM-DD），
            # 这样宠物编辑的 <input type="date"> 能渲染回来。
            _est_birthday = _age_estimate_to_birthday(f.get("age_estimate", ""))
            # 门店：从 clinic_store 全名取短名，给病历号生成用
            _short_store = _STORE_FULL_TO_SHORT.get(f.get("clinic_store", ""), "")

            _existing_pet = (
                db.query(Pet)
                .filter(
                    Pet.customer_id == _cust_id,
                    Pet.species == "cat",
                    Pet.name == _cat_name,
                )
                .first()
            )
            if _existing_pet:
                # 复用已有宠物：补全可能新增的信息（不覆盖已有字段）
                if not _existing_pet.gender or _existing_pet.gender == "unknown":
                    _existing_pet.gender = f.get("cat_gender", "unknown")
                if not _existing_pet.birthday_estimate and _est_birthday:
                    _existing_pet.birthday_estimate = _est_birthday
                if not _existing_pet.notes and f.get("health_note"):
                    _existing_pet.notes = f.get("health_note", "")[:500]
                if not _existing_pet.breed and f.get("cat_breed"):
                    _existing_pet.breed = f.get("cat_breed", "")[:80]
                if not _existing_pet.color_pattern and f.get("cat_color"):
                    _existing_pet.color_pattern = f.get("cat_color", "")[:80]
                if not _existing_pet.store and _short_store:
                    _existing_pet.store = _short_store
                if not _existing_pet.medical_record_no and (_existing_pet.store or _short_store):
                    _existing_pet.medical_record_no = _gen_medical_record_no(db, _existing_pet.store or _short_store)
                app_row.pet_id = _existing_pet.id
            else:
                _pet = Pet(
                    customer_id=_cust_id,
                    name=_cat_name,
                    species="cat",
                    gender=f.get("cat_gender", "unknown"),
                    breed=f.get("cat_breed", "")[:80],
                    color_pattern=f.get("cat_color", "")[:80],
                    birthday_estimate=_est_birthday,
                    is_stray=True,
                    notes=f.get("health_note", "")[:500],
                    store=_short_store,
                    medical_record_no=_gen_medical_record_no(db, _short_store) if _short_store else "",
                )
                db.add(_pet)
                db.flush()
                app_row.pet_id = _pet.id
            db.commit()
        except Exception as _pet_err:
            logger.warning("[tnr pet create] %s", _pet_err)

    base = Path(settings.upload_dir) / str(app_row.id)
    base.mkdir(parents=True, exist_ok=True)
    return {"id": app_row.id, "status": app_row.status, "message": "申请已创建，请继续上传照片/视频后提交。"}


@app.post("/api/apply/{app_id}/upload-media")
async def api_apply_upload_media(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    kind: str = Form("image"),  # image / video
    file: UploadFile = File(...),
):
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404, "not found")

    base = Path(settings.upload_dir) / str(app_id)
    base.mkdir(parents=True, exist_ok=True)

    if kind == "video":
        ext = _video_ext(file.filename or "")
        dest = base / f"app_vid_{secrets.token_hex(6)}{ext}"
        dest.write_bytes(await file.read())
        m = MediaFile(
            application_id=app_id,
            kind=MediaKind.application_video.value,
            stored_path=str(dest),
            original_name=file.filename or "",
        )
    else:
        ext = _image_ext(file.filename or "")
        dest = base / f"app_img_{secrets.token_hex(6)}{ext}"
        dest.write_bytes(await file.read())
        m = MediaFile(
            application_id=app_id,
            kind=MediaKind.application_image.value,
            stored_path=str(dest),
            original_name=file.filename or "",
        )
    db.add(m)
    db.commit()
    db.refresh(m)
    return {"ok": True, "media_id": m.id}


@app.post("/api/apply/{app_id}/finalize")
async def api_apply_finalize(app_id: int, request: Request, db: Session = Depends(get_db)):
    row = (
        db.query(Application)
        .options(selectinload(Application.media))
        .filter(Application.id == app_id)
        .first()
    )
    if not row:
        raise HTTPException(404, "not found")

    _assert_application_row_complete(row)

    image_paths = [Path(m.stored_path) for m in (row.media or []) if m.kind == MediaKind.application_image.value]
    video_paths = [Path(m.stored_path) for m in (row.media or []) if m.kind == MediaKind.application_video.value]
    if not image_paths:
        raise HTTPException(400, "请至少上传 1 张申请照片。")

    row.status = ApplicationStatus.pending_ai.value
    db.commit()

    ai_result = await review_application_media(image_paths, video_paths)
    row.ai_raw_json = json.dumps(ai_result, ensure_ascii=False)
    row.ai_is_likely_stray = ai_result.get("is_likely_stray")
    conf = ai_result.get("confidence")
    row.ai_confidence = float(conf) if conf is not None else None

    new_status, auto_ok = apply_auto_status_from_ai(ai_result)
    row.status = new_status
    db.commit()

    if auto_ok:
        notify_application_result(db, app_id, row.phone, row.applicant_name, approved=True, extra="系统根据图像辅助判断已完成预审通过，到院后仍需工作人员核对猫只身份。")
        push_application_result(
            db,
            application_id=app_id,
            openid=row.wechat_openid,
            applicant_name=row.applicant_name,
            status_text="审核已通过",
            phone_masked=row.phone,
            note="请按约定时间携带猫咪到院",
            submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
            action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
    elif new_status == ApplicationStatus.pre_approved.value:
        push_application_result(
            db,
            application_id=app_id,
            openid=row.wechat_openid,
            applicant_name=row.applicant_name,
            status_text="预通过（待复核）",
            phone_masked=row.phone,
            note="医院将尽快人工复核，请保持手机畅通",
            submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
            action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
        )
    elif new_status == ApplicationStatus.pending_manual.value:
        push_pending_manual_notice(
            db,
            application_id=app_id,
            openid=row.wechat_openid,
            applicant_name=row.applicant_name,
            submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        )
        # 同步推送到员工企业微信
        try:
            from app.services import wecom_notify as _wn
            _wn.notify_tnr_pending_manual(db, row)
        except Exception as _e:
            logger.warning("[wecom] notify_tnr_pending_manual failed: %s", _e)

    _STATUS_ZH = {
        "draft": "草稿", "pending_ai": "系统处理中", "pending_manual": "待人工审核",
        "pre_approved": "预通过（待复核）", "approved": "已通过", "scheduled": "已预约",
        "arrived_verified": "已到店", "surgery_completed": "手术已完成",
        "rejected": "未通过", "cancelled": "已取消", "no_show": "爽约",
    }
    return {
        "id": row.id,
        "status": row.status,
        "status_zh": _STATUS_ZH.get(row.status, row.status),
        "ai_summary": {
            "is_likely_stray": row.ai_is_likely_stray,
            "confidence": row.ai_confidence,
            "auto_approved": auto_ok,
        },
        "message": "提交成功。若未自动通过，请耐心等待医院人工审核，通过后将以登记方式通知您。",
    }


@app.get("/api/app/{app_id}/status")
async def api_app_status(app_id: int, db: Session = Depends(get_db)):
    row = (
        db.query(Application)
        .options(selectinload(Application.notifications))
        .filter(Application.id == app_id)
        .first()
    )
    if not row:
        raise HTTPException(404, "not found")

    def mask_phone(s: str) -> str:
        t = (s or "").strip()
        if len(t) < 7:
            return t
        return t[:3] + "****" + t[-4:]

    notes = (row.reject_reason or "").strip()
    if len(notes) > 80:
        notes = notes[:80] + "…"

    return {
        "id": row.id,
        "status": row.status,
        "clinic_store": row.clinic_store,
        "appointment_at": row.appointment_at,
        "applicant_name": row.applicant_name or "",
        "phone": row.phone or "",
        "phone_masked": mask_phone(row.phone),
        "cat_nickname": row.cat_nickname or "",
        "cat_gender": row.cat_gender or "",
        "age_estimate": row.age_estimate or "",
        "health_note": row.health_note or "",
        "address": row.address or "",
        "note": notes,
        "reject_reason": notes,
        "created_at": row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        "updated_at": row.updated_at.strftime("%Y-%m-%d %H:%M") if row.updated_at else "",
        "notifications": [
            {
                "channel": n.channel,
                "success": bool(n.success),
                "created_at": n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else "",
            }
            for n in (row.notifications or [])
        ],
    }


@app.get("/admin", response_class=HTMLResponse)
async def page_admin(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return templates.TemplateResponse(request, "uk/login.html",
            {"request": request, "title": "医院后台登录", "csrf_token": _get_csrf_token(request)},
        )
    qp = request.query_params
    status = (qp.get("status") or "").strip()
    store = (qp.get("store") or "").strip()
    qtext = (qp.get("q") or "").strip()
    consent = (qp.get("consent") or "").strip()  # any/true/false
    verified = (qp.get("verified") or "").strip()  # any/true/false
    has_media = (qp.get("has_media") or "").strip()  # any/true
    date_from = (qp.get("from") or "").strip()  # YYYY-MM-DD
    date_to = (qp.get("to") or "").strip()  # YYYY-MM-DD
    page = int((qp.get("page") or "1").strip() or 1)
    page_size = int((qp.get("page_size") or "50").strip() or 50)
    page = max(1, page)
    page_size = min(max(10, page_size), 100)

    base_q = db.query(Application)
    # 门店权限过滤（非超级管理员只看自己门店的数据）
    admin_store = _get_admin_store(request)
    if admin_store:
        full_store = _STORE_SHORT_TO_FULL.get(admin_store, "")
        if full_store:
            base_q = base_q.filter(Application.clinic_store == full_store)
    if status == "showcase_hidden":
        # 虚拟筛选：手术完成 + 管理员手动关闭公开
        base_q = base_q.filter(
            Application.status == ApplicationStatus.surgery_completed.value,
            Application.showcase_consent.is_(False),
        )
    elif status == "cat_verified":
        # 虚拟筛选：已现场确认（员工核对过"到的就是这只猫"）
        base_q = base_q.filter(Application.staff_cat_verified.is_(True))
    elif status:
        base_q = base_q.filter(Application.status == status)
    if store:
        base_q = base_q.filter(Application.clinic_store == store)
    if consent == "true":
        base_q = base_q.filter(Application.showcase_consent.is_(True))
    elif consent == "false":
        base_q = base_q.filter(Application.showcase_consent.is_(False))
    if verified == "true":
        base_q = base_q.filter(Application.staff_cat_verified.is_(True))
    elif verified == "false":
        base_q = base_q.filter(Application.staff_cat_verified.is_(False))
    if date_from:
        try:
            dt = datetime.strptime(date_from, "%Y-%m-%d")
            base_q = base_q.filter(Application.created_at >= dt)
        except Exception:
            pass
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d")
            base_q = base_q.filter(Application.created_at < (dt + timedelta(days=1)))
        except Exception:
            pass
    if qtext:
        if qtext.isdigit():
            base_q = base_q.filter(or_(Application.id == int(qtext), Application.phone.contains(qtext)))
        else:
            base_q = base_q.filter(
                or_(
                    Application.applicant_name.contains(qtext),
                    Application.phone.contains(qtext),
                    Application.address.contains(qtext),
                    Application.cat_nickname.contains(qtext),
                )
            )

    # 统计（按门店权限过滤，非超级管理员只统计自己门店）
    _stat_q = db.query(Application)
    if admin_store:
        _stat_full = _STORE_SHORT_TO_FULL.get(admin_store, "")
        if _stat_full:
            _stat_q = _stat_q.filter(Application.clinic_store == _stat_full)
    overall_by_status = dict(_stat_q.with_entities(Application.status, func.count(Application.id)).group_by(Application.status).all())
    # 虚拟状态：手术完成但被管理员手动关闭公开 — 用 "showcase_hidden" key 加进统计
    _hidden_count = _stat_q.filter(
        Application.status == ApplicationStatus.surgery_completed.value,
        Application.showcase_consent.is_(False),
    ).with_entities(func.count(Application.id)).scalar() or 0
    if _hidden_count > 0:
        overall_by_status["showcase_hidden"] = _hidden_count
    # 虚拟统计：已现场确认（一次性事件，跨状态）
    _verified_count = _stat_q.filter(
        Application.staff_cat_verified.is_(True),
    ).with_entities(func.count(Application.id)).scalar() or 0
    if _verified_count > 0:
        overall_by_status["cat_verified"] = _verified_count
    today0 = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_new = _stat_q.filter(Application.created_at >= today0).with_entities(func.count(Application.id)).scalar() or 0
    pending_todo = (
        _stat_q.filter(Application.status.in_([ApplicationStatus.pending_manual.value, ApplicationStatus.pre_approved.value]))
        .with_entities(func.count(Application.id))
        .scalar()
        or 0
    )

    total = base_q.count()

    rows = (
        base_q.options(selectinload(Application.media))
        .order_by(Application.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    if has_media == "true":
        rows = [a for a in rows if any(m.kind in (MediaKind.application_image.value, MediaKind.application_video.value) for m in (a.media or []))]
        total = len(rows) if page == 1 else total
    try:
        backup_files = list_backup_zips()
    except Exception:
        backup_files = []
    appointments = (
        db.query(Appointment)
        .options(selectinload(Appointment.application))
        .order_by(Appointment.created_at.desc())
        .limit(30)
        .all()
    )
    return templates.TemplateResponse(request, "uk/admin_tnr.html",  # B 补 - UK 重写 TNR 审核
        {
            "request": request,
            "title": "TNR 审核与手术登记",
            "apps": rows,
            "appointments": appointments,
            "csrf_token": _get_csrf_token(request),
            "backup_files": backup_files,
            "filters": {
                "status": status,
                "store": store,
                "q": qtext,
                "consent": consent,
                "verified": verified,
                "has_media": has_media,
                "from": date_from,
                "to": date_to,
                "page": page,
                "page_size": page_size,
            },
            "stats": {"overall_by_status": overall_by_status, "today_new": today_new, "pending_todo": pending_todo, "total": total},
        },
    )


def _admin_appointment_redirect_base(redirect_after: str) -> str:
    """预约表单提交后回到哪一页；仅允许固定路径，避免开放重定向。"""
    val = (redirect_after or "").strip().lower()
    if val == "appointments":
        return "/admin/appointments"
    if val == "calendar":
        return "/admin/calendar"
    if val == "mobile":
        return "/m/appointments"
    if val == "mobile_today":
        return "/m"
    # 从客户档案进来：return_to=customer:123 → 跳回 /admin/customers/123
    if val.startswith("customer:"):
        cid = val.split(":", 1)[1]
        if cid.isdigit():
            return f"/admin/customers/{cid}"
    if val.startswith("mobile_customer:"):
        cid = val.split(":", 1)[1]
        if cid.isdigit():
            return f"/m/customer/{cid}"
    if val.startswith("mobile_appt:"):
        aid = val.split(":", 1)[1]
        if aid.isdigit():
            return f"/m/appointment/{aid}"
    return "/admin"


def _admin_appointment_redirect(
    next_val: str | None,
    *,
    ok: str | None = None,
    err: str | None = None,
    anchor: str | None = None,
) -> RedirectResponse:
    base = "/admin/appointments"
    parts: list[str] = []
    if ok:
        parts.append(f"appointment_ok={ok}")
    if err:
        parts.append("appointment_err=" + quote(str(err)[:200], safe=""))
    url = base + ("?" + "&".join(parts) if parts else "")
    if anchor:
        url += f"#{anchor}"
    return RedirectResponse(url, status_code=303)


# ── 美容时长估算 ──────────────────────────────────────────────────────────────
# 美容师工作时间（分钟，从午夜起算）
_BEAUTY_WORK_START = 13 * 60   # 13:00
_BEAUTY_WORK_END   = 22 * 60   # 22:00
_BEAUTY_SLOT_STEP  = 30        # 每 30 分钟一个可选时段


def _calc_beauty_duration(service_name: str, pet_size: str, coat_length: str, addon_services: str = "") -> int:
    """
    根据美容项目、体型、毛发长度、附加服务估算占用分钟数。
    附加服务用逗号分隔，每项 +30 分钟。
    """
    sn    = service_name or ""
    sz    = pet_size     or ""
    is_long  = (coat_length or "") == "长毛"
    is_wash  = "洗护" in sn
    # is_groom = "造型" in sn  # 不需要显式判断，else 分支即为造型

    base = 60  # fallback

    if "犬" in sn:
        if "微小型" in sz:
            base = 30 if is_wash else 90
        elif "小型犬" in sz:           # 注意：elif 保证 "微小型" 已被排除
            base = 60 if is_wash else 120
        elif "中大型" in sz:
            base = (120 if is_long else 90) if is_wash else 180
        elif "中型犬" in sz:
            base = (90 if is_long else 60) if is_wash else 150
        elif "大型犬" in sz:           # "中大型犬" 已被前面 elif 排除
            base = (150 if is_long else 120) if is_wash else 210
    elif "猫" in sn:
        if "大型猫" in sz:
            base = (150 if is_long else 120) if is_wash else 150
        elif "中型猫" in sz:
            base = (120 if is_long else 90) if is_wash else 120
        elif "小型猫" in sz:
            base = (90 if is_long else 60) if is_wash else 120

    # 附加服务：每项 +30 分钟
    if addon_services:
        addon_count = len([a for a in addon_services.split(",") if a.strip()])
        base += addon_count * 30

    return base


def _beauty_slots_for_date(
    db: "Session",
    store: str,
    date_str: str,
    new_service: str,
    new_duration: int,
) -> list[str]:
    """
    返回指定日期门店美容师可接受的开始时间列表（HH:MM 字符串）。
    实现猫进烘干机时可并发做 ≤60min 犬洗护的规则。
    """
    _inactive = {AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value}
    bookings = (
        db.query(Appointment)
        .filter(
            Appointment.appointment_date == date_str,
            Appointment.store == store,
            Appointment.category.in_(["beauty", "grooming", "washcare"]),
            Appointment.status.notin_(list(_inactive)),
        )
        .all()
    )

    # 解析已有预约 → 时间块（分钟）
    dog_blocks: list[tuple[int, int]] = []       # 犬：完全占用美容师
    cat_active: list[tuple[int, int]] = []       # 猫主动护理阶段
    cat_dryer:  list[tuple[int, int]] = []       # 猫烘干机阶段（可并发 ≤60min 犬服务）

    for b in bookings:
        try:
            bh, bm = b.appointment_time.split(":")
            b_s = int(bh) * 60 + int(bm)
            b_e = b_s + (b.duration_minutes or 60)
        except Exception:
            continue
        sn = b.service_name or ""
        if "猫" in sn:
            # 主动护理阶段固定为最初60分钟（洗猫+护理+冲洗）
            # 之后才是烘干机阶段（此阶段可并发小型犬服务）
            cat_active_end = b_s + 60
            if cat_active_end < b_e:
                cat_active.append((b_s, cat_active_end))
                cat_dryer.append((cat_active_end, b_e))
            else:
                # 总时长 ≤60 分钟：全程主动护理，无烘干机窗口
                cat_active.append((b_s, b_e))
        else:
            dog_blocks.append((b_s, b_e))

    is_new_dog = "犬" in new_service
    is_new_cat = "猫" in new_service

    def _overlaps(s1: int, e1: int, s2: int, e2: int) -> bool:
        return s1 < e2 and e1 > s2

    def _slot_ok(start: int) -> bool:
        end = start + new_duration
        if start < _BEAUTY_WORK_START or end > _BEAUTY_WORK_END:
            return False
        # 犬类预约：不能与 dog_blocks 重叠
        for bs, be in dog_blocks:
            if _overlaps(start, end, bs, be):
                return False
        if is_new_dog:
            # 不能与猫的主动护理阶段重叠
            for bs, be in cat_active:
                if _overlaps(start, end, bs, be):
                    return False
            # 烘干机窗口：仅 ≤60min 且完全落在窗口内才允许
            for dws, dwe in cat_dryer:
                if _overlaps(start, end, dws, dwe):
                    if not (new_duration <= 60 and start >= dws and end <= dwe):
                        return False
        elif is_new_cat:
            # 猫：需要完全空闲（不能与任何阶段重叠）
            for bs, be in cat_active:
                if _overlaps(start, end, bs, be):
                    return False
            for dws, dwe in cat_dryer:
                if _overlaps(start, end, dws, dwe):
                    return False
        return True

    slots = []
    t = _BEAUTY_WORK_START
    while t + new_duration <= _BEAUTY_WORK_END:
        if _slot_ok(t):
            slots.append(f"{t // 60:02d}:{t % 60:02d}")
        t += _BEAUTY_SLOT_STEP
    return slots


# ── 门诊/手术容量规则 ─────────────────────────────────────────────────────────
# 每门店、每时段（上午/下午/晚上）的总容量单位数上限
# 各时段容量上限（按时长等比，每小时 3 单位）
# 上午 3h=9，下午 6h=18，晚上 4h=12
_SLOT_CAPACITY = {
    "morning":   9,
    "afternoon": 18,
    "evening":   12,
    "other":     9,
}
# 疫苗/驱虫 = 1 单位；普通门诊 = 3 单位；TNR/手术 = 4 单位；美容 = 0（不参与）

_OUTPATIENT_SERVICES = [
    "疫苗/驱虫", "体检", "呼吸道", "胃肠道", "泌尿道",
    "皮肤", "口腔", "行动异常", "心内科", "肾内科",
]
_VACCINE_KEYWORDS = ("疫苗", "驱虫")

_SLOT_BOUNDS = {
    "morning":   ("09:00", "12:00"),
    "afternoon": ("12:00", "18:00"),
    "evening":   ("18:00", "22:00"),
    "other":     ("00:00", "23:59"),
}
_SLOT_NAME_ZH = {"morning": "上午", "afternoon": "下午", "evening": "晚上", "other": "该时段"}


def _capacity_units(category: str, service_name: str) -> int:
    """返回该预约消耗的容量单位（0 = 不纳入容量管控）。
    单位换算：
      疫苗/驱虫 = 1 单位
      普通门诊   = 3 单位
      TNR/手术   = 4 单位（术前检查少，相对快，2台=8单位 ≤ 上限9）
      美容/洗护  = 0 单位（不参与）
    """
    if category in ("tnr", "surgery"):
        return 4
    if category == "outpatient":
        sn = service_name or ""
        if any(kw in sn for kw in _VACCINE_KEYWORDS):
            return 1
        return 3
    return 0  # beauty / grooming / washcare 不参与容量管控


_OUTPATIENT_TIME_START = "10:00"   # 门诊最早开始时间（上午护理住院动物）
_OUTPATIENT_TIME_END   = "21:00"   # 门诊最晚开始时间（避免加班）


def _check_outpatient_time(category: str, appointment_time: str) -> str | None:
    """门诊/疫苗时间限制：10:00 之后、21:00 之前。返回错误描述或 None。"""
    if category != AppointmentCategory.outpatient.value:
        return None
    t = (appointment_time or "")[:5]
    if t < _OUTPATIENT_TIME_START:
        return f"门诊预约最早从 {_OUTPATIENT_TIME_START} 开始（上午需护理住院动物），请选择 {_OUTPATIENT_TIME_START} 或之后的时间。"
    if t >= _OUTPATIENT_TIME_END:
        return f"门诊预约最晚在 {_OUTPATIENT_TIME_END} 之前，请选择更早的时间。"
    return None


def _check_slot_capacity(
    db: "Session",
    store: str,
    appointment_date: str,
    appointment_time: str,
    category: str,
    service_name: str,
    exclude_id: int | None = None,
) -> str | None:
    """返回错误提示（str）或 None（通过）。"""
    new_units = _capacity_units(category, service_name)
    if new_units == 0:
        return None  # 不受容量限制

    slot_key = _appt_time_slot(appointment_time)
    t_from, t_to = _SLOT_BOUNDS.get(slot_key, ("00:00", "23:59"))
    slot_zh = _SLOT_NAME_ZH.get(slot_key, "该时段")

    _inactive = {AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value}
    q = (
        db.query(Appointment)
        .filter(
            Appointment.store == store,
            Appointment.appointment_date == appointment_date,
            Appointment.appointment_time >= t_from,
            Appointment.appointment_time < t_to,
            Appointment.status.notin_(list(_inactive)),
        )
    )
    if exclude_id:
        q = q.filter(Appointment.id != exclude_id)

    slot_max = _SLOT_CAPACITY.get(slot_key, 9)
    used = sum(_capacity_units(a.category, a.service_name or "") for a in q.all())
    avail = slot_max - used

    if avail < new_units:
        type_zh = (
            "手术" if category in ("tnr", "surgery")
            else ("疫苗/驱虫" if new_units == 1 else "门诊")
        )
        if avail <= 0:
            return f"该门店 {slot_zh}时段预约容量已满（上限 {slot_max} 单位），请选择其他时段或日期。"
        return (
            f"该门店 {slot_zh}时段剩余容量不足：剩余 {avail} 单位，"
            f"此{type_zh}需要 {new_units} 单位，请选择其他时段或日期。"
        )
    return None


def _appt_time_slot(time_str: str) -> str:
    try:
        h = int((time_str or "00:00").split(":")[0])
        if 9 <= h < 12:
            return "morning"
        if 12 <= h < 18:
            return "afternoon"
        if 18 <= h < 22:
            return "evening"
        return "other"
    except Exception:
        return "other"


@app.get("/api/admin/pending-count")
async def api_admin_pending_count(request: Request, db: Session = Depends(get_db)):
    """返回待确认预约数量（仅限已登录后台）"""
    if not request.session.get("admin"):
        return {"count": 0}
    from datetime import date as _date
    today_str = _date.today().isoformat()
    count = (
        db.query(func.count(Appointment.id))
        .filter(
            Appointment.status == AppointmentStatus.pending.value,
            Appointment.appointment_date >= today_str,
        )
        .scalar()
    ) or 0
    return {"count": count}


@app.get("/api/admin/feedback-count")
async def api_admin_feedback_count(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return {"count": 0}
    from app.models import Feedback
    count = db.query(func.count(Feedback.id)).filter(Feedback.status == "pending").scalar() or 0
    return {"count": count}


_HEALTH_WARNING_KEYWORDS = [
    ("过敏", "🚫", "过敏"),
    ("青霉素", "🚫", "青霉素过敏"),
    ("CKD", "💊", "慢性肾病 CKD"),
    ("慢性肾", "💊", "慢性肾病"),
    ("肾衰", "💊", "肾衰"),
    ("糖尿病", "💊", "糖尿病"),
    ("HCM", "❤️", "肥厚性心肌病"),
    ("心脏病", "❤️", "心脏病"),
    ("FIP", "⚠", "FIP（治疗中）"),
    ("癫痫", "⚡", "癫痫"),
    ("特应性", "💊", "特应性皮炎"),
    ("FIV", "🩸", "FIV"),
    ("FeLV", "🩸", "FeLV"),
    ("管控药", "⚠", "管控药敏感"),
]


def _detect_health_warnings(text: str) -> list:
    """从 pet.notes / visit.diagnosis 文本里扫描健康警示关键词。返回 [{emoji, label}] 去重。"""
    if not text:
        return []
    out = []
    seen = set()
    for kw, emoji, label in _HEALTH_WARNING_KEYWORDS:
        if kw in text and label not in seen:
            out.append({"emoji": emoji, "label": label})
            seen.add(label)
    return out


@app.get("/api/admin/customer-context")
async def api_admin_customer_context(
    request: Request, db: Session = Depends(get_db),
    customer_id: int = Query(0), pet_id: int = Query(0),
):
    """开单页右侧 sidebar 异步拉取的客户上下文：
    钱包余额 / 健康警示 / 未付单据 / 最近就诊 / 防疫近况 / 体重变化。
    """
    if not request.session.get("admin"):
        return {}
    data: dict = {}

    # 档案摘要（始终返回 — 客户/宠物基本统计，避免右栏空白）
    if customer_id:
        cust = db.get(Customer, customer_id)
        if cust:
            pet_count = db.query(func.count(Pet.id)).filter(Pet.customer_id == customer_id).scalar() or 0
            paid_total = db.query(func.coalesce(func.sum(Invoice.total_amount), 0)).filter(
                Invoice.customer_id == customer_id,
                Invoice.payment_status == "paid",
            ).scalar() or 0
            last_visit = db.query(Visit.visit_date).filter(Visit.customer_id == customer_id)\
                .order_by(Visit.visit_date.desc(), Visit.id.desc()).first()
            data["summary"] = {
                "register_date": cust.created_at.strftime("%Y-%m-%d") if cust.created_at else "",
                "pet_count": int(pet_count),
                "lifetime_paid": round(float(paid_total), 2),
                "last_visit": (last_visit[0] if last_visit else "") or "",
            }

    # 钱包
    if customer_id:
        w = db.query(Wallet).filter(Wallet.customer_id == customer_id).first()
        if w and (w.balance or 0) > 0:
            data["wallet"] = {
                "balance": round(float(w.balance or 0), 2),
                "lifetime_recharge": round(float(w.lifetime_recharge or 0), 2),
            }

    # 健康警示 — 从 pet.notes + 最近 5 次 visit.diagnosis 扫
    warnings: list = []
    if pet_id:
        pet = db.get(Pet, pet_id)
        if pet:
            warnings.extend(_detect_health_warnings(pet.notes or ""))
            recent_diags = db.query(Visit.diagnosis)\
                .filter(Visit.pet_id == pet_id, Visit.diagnosis != "")\
                .order_by(Visit.visit_date.desc(), Visit.id.desc()).limit(5).all()
            for (diag,) in recent_diags:
                for w in _detect_health_warnings(diag or ""):
                    if w not in warnings:
                        warnings.append(w)
    if warnings:
        data["warnings"] = warnings[:5]

    # 未付单据
    if customer_id:
        unpaid_rows = db.query(Invoice).filter(
            Invoice.customer_id == customer_id,
            Invoice.payment_status == "unpaid",
        ).all()
        if unpaid_rows:
            data["unpaid"] = {
                "count": len(unpaid_rows),
                "total": round(sum((r.total_amount or 0) for r in unpaid_rows), 2),
            }

    # 最近 3 次就诊
    if pet_id:
        recent_visits = db.query(Visit).filter(Visit.pet_id == pet_id)\
            .order_by(Visit.visit_date.desc(), Visit.id.desc()).limit(3).all()
        if recent_visits:
            data["recent_visits"] = [{
                "id": v.id,
                "date": v.visit_date or "—",
                "diagnosis": ((v.diagnosis or v.chief_complaint or "")[:30]) or "—",
                "vet": v.vet_name or "",
            } for v in recent_visits]

    # 防疫近况：最近一针狂犬 + 联苗 + 体内驱 + 体外驱
    if pet_id:
        from datetime import date as _date
        today = _date.today().isoformat()
        immun = []
        # 疫苗最近一次（按 vaccine_type 取最新）
        for vt, label in [("rabies", "狂犬"), ("combo", "联苗"), ("combo_3", "联苗"), ("combo_6", "联苗")]:
            v = db.query(Vaccination).filter(
                Vaccination.pet_id == pet_id,
                Vaccination.vaccine_type == vt,
                Vaccination.status != "voided",
            ).order_by(Vaccination.vaccinated_date.desc(), Vaccination.id.desc()).first()
            if v:
                state = "ok"
                if v.next_due_date:
                    if v.next_due_date < today:
                        state = "expired"
                    else:
                        try:
                            from datetime import datetime as _dt
                            days_left = (_dt.strptime(v.next_due_date, "%Y-%m-%d") - _dt.strptime(today, "%Y-%m-%d")).days
                            if 0 <= days_left <= 14:
                                state = "due_soon"
                        except Exception:
                            pass
                immun.append({"type": label, "date": v.vaccinated_date, "next": v.next_due_date, "state": state})
                if vt == "combo_3" or vt == "combo_6":
                    break  # 联苗只取一次
        # 驱虫
        for dt_kind, label in [("internal", "体内驱虫"), ("external", "体外驱虫"), ("combo", "体内外驱虫")]:
            d = db.query(DewormingRecord).filter(
                DewormingRecord.pet_id == pet_id,
                DewormingRecord.deworm_type == dt_kind,
                DewormingRecord.status != "voided",
            ).order_by(DewormingRecord.deworm_date.desc(), DewormingRecord.id.desc()).first()
            if d:
                state = "ok"
                if d.next_due_date and d.next_due_date < today:
                    state = "expired"
                immun.append({"type": label, "date": d.deworm_date, "next": d.next_due_date, "state": state})
        if immun:
            data["immunity"] = immun[:6]

    # 体重
    if pet_id:
        weights = db.query(WeightRecord).filter(WeightRecord.pet_id == pet_id)\
            .order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).limit(2).all()
        if weights:
            cur = weights[0]
            delta = None
            if len(weights) >= 2:
                delta = round(float(cur.weight_kg or 0) - float(weights[1].weight_kg or 0), 2)
            data["weight"] = {
                "current": round(float(cur.weight_kg or 0), 2),
                "date": cur.record_date or "",
                "delta": delta,
            }

    return data


@app.get("/admin/appointments", response_class=HTMLResponse)
async def page_admin_appointments(
    request: Request,
    db: Session = Depends(get_db),
    df: str = Query(""),        # date_from  YYYY-MM-DD
    dt: str = Query(""),        # date_to    YYYY-MM-DD
    preset: str = Query("today"),  # today / 3days / week / month / custom
    appt_status: str = Query(""),
    appt_store: str = Query(""),
    appt_category: str = Query(""),
):
    if not _admin_ok(request):
        return templates.TemplateResponse(request, "uk/login.html",
            {"request": request, "title": "医院后台登录", "csrf_token": _get_csrf_token(request)},
        )

    today = datetime.now().date()
    today_str = today.strftime("%Y-%m-%d")
    tomorrow_str = (today + timedelta(days=1)).strftime("%Y-%m-%d")

    # ── 日期范围推导 ──────────────────────────────────────────
    _preset = (preset or "today").strip()
    if _preset == "3days":
        df_d, dt_d = today, today + timedelta(days=2)
    elif _preset == "week":
        # 本周：当周一到当周日
        df_d = today - timedelta(days=today.weekday())
        dt_d = df_d + timedelta(days=6)
    elif _preset == "month":
        # 本月：本月1日到本月最后一天
        import calendar as _calendar
        df_d = today.replace(day=1)
        last_day = _calendar.monthrange(today.year, today.month)[1]
        dt_d = today.replace(day=last_day)
    elif _preset == "custom":
        try:
            df_d = datetime.strptime(df, "%Y-%m-%d").date() if df else today
        except ValueError:
            df_d = today
        try:
            dt_d = datetime.strptime(dt, "%Y-%m-%d").date() if dt else df_d + timedelta(days=6)
        except ValueError:
            dt_d = df_d + timedelta(days=6)
        if dt_d < df_d:
            dt_d = df_d
    else:  # today (default)
        _preset = "today"
        df_d, dt_d = today, today

    df_str = df_d.strftime("%Y-%m-%d")
    dt_str = dt_d.strftime("%Y-%m-%d")

    # ── 查询 ──────────────────────────────────────────────────
    q = (
        db.query(Appointment)
        .options(selectinload(Appointment.application))
        .filter(
            Appointment.appointment_date >= df_str,
            Appointment.appointment_date <= dt_str,
        )
    )
    if appt_status:
        q = q.filter(Appointment.status == appt_status)
    if appt_store:
        q = q.filter(Appointment.store == appt_store)
    if appt_category:
        q = q.filter(Appointment.category == appt_category)
    # 门店权限过滤
    _appt_admin_store = _get_admin_store(request)
    if _appt_admin_store:
        _appt_full_store = _STORE_SHORT_TO_FULL.get(_appt_admin_store, "")
        if _appt_full_store:
            q = q.filter(Appointment.store == _appt_full_store)

    appointments_raw = q.order_by(
        Appointment.appointment_date, Appointment.appointment_time
    ).all()

    # ── 按日期→部门→时段分组 ──────────────────────────────────
    _SLOT_ORDER = ["morning", "afternoon", "evening", "other"]
    _SLOT_LABELS = {
        "morning":   "上午  09:00 – 12:00",
        "afternoon": "下午  12:00 – 18:00",
        "evening":   "晚上  18:00 – 22:00",
        "other":     "其他时段",
    }
    _WEEKDAYS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    _BEAUTY_CATS = {"beauty", "grooming", "washcare"}
    _DEPT_ORDER  = ["medical", "beauty"]
    _DEPT_LABELS = {"medical": "医疗", "beauty": "美容"}

    # date → dept → slot → [appts]
    date_buckets: dict[str, dict[str, dict[str, list]]] = {}
    for appt in appointments_raw:
        d = appt.appointment_date
        if d not in date_buckets:
            date_buckets[d] = {
                "medical": {s: [] for s in _SLOT_ORDER},
                "beauty":  {s: [] for s in _SLOT_ORDER},
            }
        dept = "beauty" if (appt.category or "") in _BEAUTY_CATS else "medical"
        date_buckets[d][dept][_appt_time_slot(appt.appointment_time)].append(appt)

    def _date_display(d_str: str) -> str:
        try:
            d_obj = datetime.strptime(d_str, "%Y-%m-%d").date()
            wd = _WEEKDAYS[d_obj.weekday()]
            label = f"{d_obj.month}月{d_obj.day}日（{wd}）"
            if d_str == today_str:
                label += " · 今天"
            elif d_str == tomorrow_str:
                label += " · 明天"
            return label
        except Exception:
            return d_str

    grouped_appointments = []
    _inactive = {AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value}
    for d_str in sorted(date_buckets):
        all_day: list = []
        dept_groups = []
        for dept_key in _DEPT_ORDER:
            dept_appts = [a for sl in _SLOT_ORDER for a in date_buckets[d_str][dept_key][sl]]
            all_day.extend(dept_appts)
            dept_slots = [
                {"key": sl, "label": _SLOT_LABELS[sl], "appts": date_buckets[d_str][dept_key][sl]}
                for sl in _SLOT_ORDER
                if date_buckets[d_str][dept_key][sl]
            ]
            if dept_slots:
                dept_groups.append({
                    "key":   dept_key,
                    "label": _DEPT_LABELS[dept_key],
                    "slots": dept_slots,
                    "total": len(dept_appts),
                })

        active   = [a for a in all_day if a.status not in _inactive]
        pending  = [a for a in all_day if a.status == AppointmentStatus.pending.value]
        tnr_ct   = sum(1 for a in active if a.category == AppointmentCategory.tnr.value)
        grouped_appointments.append({
            "date":          d_str,
            "date_display":  _date_display(d_str),
            "is_today":      d_str == today_str,
            "total":         len(all_day),
            "active_count":  len(active),
            "pending_count": len(pending),
            "tnr_count":     tnr_ct,
            "dept_groups":   dept_groups,
        })

    # ── 顶部统计（始终基于今日全量，不受筛选影响） ────────────
    _today_appts = (
        db.query(Appointment)
        .filter(
            Appointment.appointment_date == today_str,
            Appointment.status.notin_(list(_inactive)),
        )
        .all()
    )
    # TNR 月度配额状态
    _this_month = today_str[:7]
    _tnr_quota_status = []
    for _store in _CLINIC_STORES:
        _cfg = _get_tnr_store_config(db, _store)
        _monthly = _get_tnr_monthly_confirmed_count(db, _store, _this_month)
        _tnr_quota_status.append({
            "store": _store,
            "store_index": list(_CLINIC_STORES).index(_store),
            "accepting": _cfg.tnr_accepting,
            "monthly_count": _monthly,
            "monthly_quota": _cfg.tnr_monthly_quota,
            "is_open": _cfg.tnr_accepting and (_monthly < _cfg.tnr_monthly_quota),
        })
    stats = {
        "today_active":  len(_today_appts),
        "pending_total": db.query(Appointment)
            .filter(
                Appointment.status == AppointmentStatus.pending.value,
                Appointment.appointment_date >= today_str,
            ).count(),
        "tnr_today":     sum(1 for a in _today_appts if a.category == AppointmentCategory.tnr.value),
        "tnr_daily_max": _TNR_DAILY_MAX,
        "tnr_quota_status": _tnr_quota_status,
    }

    return templates.TemplateResponse(request, "uk/appointments.html",  # B9 UK 重写
        {
            "request":              request,
            "title":                "预约管理",
            "appointments":         appointments_raw,          # 创建表单 JS 仍使用
            "grouped_appointments": grouped_appointments,
            "stats":                stats,
            "filters": {
                "date_from": df_str,
                "date_to":   dt_str,
                "preset":    _preset,
                "status":    appt_status,
                "store":     appt_store,
                "category":  appt_category,
            },
            "stores":     list(_CLINIC_STORES),
            "csrf_token": _get_csrf_token(request),
        },
    )


@app.get("/admin/api/tnr-application/{application_id}/for-appointment")
async def admin_api_tnr_application_for_appointment(
    application_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """后台新建预约时：按 TNR 申请编号拉取联系人、宠物、门店等用于自动填充（需已登录）。"""
    require_admin(request)
    row = db.get(Application, application_id)
    if not row:
        raise HTTPException(404, detail="申请不存在")
    store = (row.clinic_store or "").strip()
    if store not in _ALLOWED_CLINIC_STORES:
        store = ""
    g = (row.cat_gender or "unknown").strip().lower()
    if g not in ("male", "female", "unknown"):
        g = "unknown"
    appt_date = ""
    raw_at = (row.appointment_at or "").strip()
    if len(raw_at) >= 10 and raw_at[4] == "-" and raw_at[7] == "-":
        try:
            datetime.strptime(raw_at[:10], "%Y-%m-%d")
            appt_date = raw_at[:10]
        except ValueError:
            pass
    return {
        "ok": True,
        "application_id": row.id,
        "customer_name": (row.applicant_name or "").strip(),
        "phone": (row.phone or "").strip(),
        "pet_name": (row.cat_nickname or "").strip(),
        "pet_gender": g,
        "store": store,
        "appointment_date_suggestion": appt_date,
    }


@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    """GET: 展示登录页（点击「登录」链接 / 未登录访问任意页 都会到这里）。"""
    # 已登录直接回工作台
    if _admin_ok(request):
        return RedirectResponse(_post_login_redirect(request), status_code=303)
    return templates.TemplateResponse(request, "uk/login.html", {
        "title": "医院后台登录",
        "csrf_token": _get_csrf_token(request),
    })


# ═════════════════════════════════════════════════════════════
# M1 · 手机 PWA 入口判断
# - UA 检测 + cookie 「force_desktop」让用户强制桌面版
# - mobile_role auto → 按 role 给默认
# ═════════════════════════════════════════════════════════════
_MOBILE_UA_PAT = re.compile(r"(iPhone|Android|iPod|Mobile|BlackBerry|IEMobile)", re.I)


def _is_mobile_ua(request: Request) -> bool:
    """是否手机 UA。平板（iPad）当桌面，因为屏幕够大。

    优先级（高 → 低）：
      ?mobile=1 强制手机（覆盖一切，最高优先级）
      ?desktop=1 强制桌面
      force_desktop cookie（来自 /m/desktop，30 天有效）
      UA 自动检测
    """
    # 显式 ?mobile=1 优先级最高，可以覆盖 force_desktop cookie（用于清理）
    if request.query_params.get("mobile") == "1":
        return True
    if request.cookies.get("force_desktop") == "1":
        return False
    if request.query_params.get("desktop") == "1":
        return False
    ua = request.headers.get("user-agent", "") or ""
    if "iPad" in ua:
        return False
    return bool(_MOBILE_UA_PAT.search(ua))


def _resolve_mobile_role(session_role: str, mobile_role: str) -> str:
    """auto → 按角色推断；其他直接返回。"""
    mr = (mobile_role or "auto").strip()
    if mr in ("doctor", "nurse", "groomer"):
        return mr
    # auto：superadmin 默认 doctor，staff 默认 nurse
    return "doctor" if session_role == "superadmin" else "nurse"


def _post_login_redirect(request: Request) -> str:
    """登录成功后跳哪。
    优先级：?next=（白名单：以 / 开头不含 //）→ 手机 /m → 桌面 /admin。
    """
    nxt = (request.query_params.get("next") or "").strip()
    if nxt and nxt.startswith("/") and not nxt.startswith("//"):
        return nxt
    # 也支持表单提交时 referer 里带的 next（POST 登录会丢 query_params）
    referer = request.headers.get("referer", "") or ""
    if "?next=" in referer:
        try:
            from urllib.parse import urlparse, parse_qs
            qs = parse_qs(urlparse(referer).query)
            r_next = (qs.get("next", [""])[0] or "").strip()
            if r_next.startswith("/") and not r_next.startswith("//"):
                return r_next
        except Exception:
            pass
    if _is_mobile_ua(request):
        return "/m"
    return "/admin/customers"


@app.post("/admin/login")
async def admin_login(
    request: Request,
    db: Session = Depends(get_db),
    username: str = Form(...),
    password: str = Form(...),
    csrf_token: str = Form(""),
):
    _require_csrf(request, csrf_token)
    username = username.strip()

    # 优先查 DB 账号（精确匹配）
    user = db.query(AdminUser).filter(AdminUser.username == username, AdminUser.is_active == True).first()
    # 兜底：DB 里历史数据可能有首尾空格 → 用 TRIM 再匹配一次
    if not user:
        from sqlalchemy import func as _f
        user = db.query(AdminUser).filter(
            _f.trim(AdminUser.username) == username,
            AdminUser.is_active == True,
        ).first()
    if user and _pwd_ctx.verify(password, user.password_hash):
        request.session["admin"] = True
        request.session["admin_role"] = user.role
        request.session["admin_username"] = (user.username or "").strip()
        request.session["admin_store"] = user.store or ""
        request.session["mobile_role"] = (user.mobile_role or "auto")
        return RedirectResponse(_post_login_redirect(request), status_code=303)

    # 兜底：环境变量密码（用于迁移期 / 紧急登录，用户名须为 admin）
    if username == "admin" and password == settings.admin_password:
        request.session["admin"] = True
        request.session["admin_role"] = "superadmin"
        request.session["admin_username"] = "admin"
        request.session["mobile_role"] = "auto"
        return RedirectResponse(_post_login_redirect(request), status_code=303)

    # 失败时记日志，方便诊断（控制台 / journalctl 看得到）
    _diag_user = db.query(AdminUser).filter(AdminUser.username.like(f"%{username}%")).first()
    if _diag_user:
        logger.warning(
            "[login fail] 输入='%s' 找到相似账号 id=%s username=%r is_active=%s",
            username, _diag_user.id, _diag_user.username, _diag_user.is_active,
        )
    else:
        logger.warning("[login fail] 输入='%s' 数据库里找不到该账号", username)

    return templates.TemplateResponse(request, "uk/login.html",
        {"request": request, "title": "医院后台登录", "error": "账号或密码不正确", "csrf_token": _get_csrf_token(request)},
        status_code=401,
    )


# ═════════════════════════════════════════════════════════════
# 企业微信单点登录（Phase 1）
# 流程：员工在企微内点应用 → 跳 /admin/wecom-login → 重定向到企微 OAuth →
#       企微带 code 回 /admin/wecom-callback → 用 code 换 userid →
#       按 wecom_userid 找到 AdminUser → 写 session → 跳 /admin/customers
# ═════════════════════════════════════════════════════════════
@app.get("/admin/wecom-login")
async def admin_wecom_login(request: Request, next: str = "/admin/customers"):
    from app.services import wecom_client as _wc
    if not _wc.enabled():
        return HTMLResponse(
            "<p style='font-family:system-ui;padding:2rem;color:#b91c1c;'>"
            "企业微信集成未配置。请管理员在服务器 <code>.env</code> 中配置 "
            "WECOM_CORP_ID / WECOM_AGENT_ID / WECOM_SECRET。</p>",
            status_code=503,
        )
    base = (settings.public_base_url or "").rstrip("/") or str(request.base_url).rstrip("/")
    # state 里塞 next，回调时取出
    state = quote(next or "/admin/customers", safe="")
    redirect_uri = f"{base}/admin/wecom-callback"
    url = _wc.build_oauth_url(redirect_uri, state=state)
    return RedirectResponse(url, status_code=302)


@app.get("/admin/wecom-callback")
async def admin_wecom_callback(
    request: Request,
    code: str = "",
    state: str = "",
    db: Session = Depends(get_db),
):
    from app.services import wecom_client as _wc
    if not code:
        return HTMLResponse("缺少 code 参数", status_code=400)
    try:
        info = _wc.code_to_userid(code)
    except Exception as e:
        logger.warning("[wecom oauth] code_to_userid failed: %s", e)
        return HTMLResponse(
            f"<p style='font-family:system-ui;padding:2rem;color:#b91c1c;'>企业微信登录失败：{e}</p>"
            "<p><a href='/admin/login'>← 用账号密码登录</a></p>",
            status_code=502,
        )
    userid = (info.get("userid") or "").strip()
    if not userid:
        # 外部成员（客户）登录，不允许进后台
        return HTMLResponse(
            "<p style='font-family:system-ui;padding:2rem;color:#b91c1c;'>"
            "检测到您是外部联系人，无后台访问权限。</p>",
            status_code=403,
        )
    user = db.query(AdminUser).filter(
        AdminUser.wecom_userid == userid,
        AdminUser.is_active == True,
    ).first()
    if not user:
        return HTMLResponse(
            f"<div style='font-family:system-ui;padding:2rem;max-width:480px;margin:auto;'>"
            f"<h3 style='color:#b91c1c;'>企业微信账号未绑定</h3>"
            f"<p>你的企微 userid：<code>{userid}</code></p>"
            f"<p>请联系超级管理员，在「员工管理」中把这个 userid 填到你的账号上，再重新打开应用。</p>"
            f"<p><a href='/admin/login'>← 先用账号密码登录</a></p>"
            f"</div>",
            status_code=403,
        )
    request.session["admin"] = True
    request.session["admin_role"] = user.role
    request.session["admin_username"] = (user.username or "").strip()
    request.session["admin_store"] = user.store or ""
    # state 中带回的跳转目标
    from urllib.parse import unquote as _unquote
    try:
        next_url = _unquote(state) if state else "/admin/customers"
    except Exception:
        next_url = "/admin/customers"
    if not next_url.startswith("/"):
        next_url = "/admin/customers"
    return RedirectResponse(next_url, status_code=303)


_DEPLOY_TOKEN_FILE = Path("/srv/tnr-app/deploy_token.txt")

@app.post("/api/webhook/deploy")
async def webhook_deploy(request: Request):
    token = request.headers.get("X-Deploy-Token", "")
    try:
        expected = _DEPLOY_TOKEN_FILE.read_text().strip()
    except Exception:
        raise HTTPException(status_code=503, detail="deploy token not configured")
    if not token or not secrets.compare_digest(token, expected):
        raise HTTPException(status_code=403, detail="forbidden")
    subprocess.Popen(
        "sleep 3 && git -C /srv/tnr-app/releases/current pull origin main && systemctl restart tnr-app",
        shell=True, start_new_session=True,
    )
    return {"status": "deploying"}


@app.get("/admin/logout")
async def admin_logout(request: Request):
    request.session.clear()
    return RedirectResponse("/admin", status_code=303)


# ── 账号管理（仅 superadmin）────────────────────────────────────────────

@app.get("/admin/run-seed-2604", response_class=HTMLResponse)
async def admin_run_seed_2604(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    import traceback
    lines = []
    records = [
        dict(applicant_name="郑香玉", phone="15323455977", clinic_store="龙华店",
             appointment_at="2026-04-25", location_address="中国广东省深圳市",
             id_number="632123199706180526", address="秋港花园；秋港花园D 5楼下灌木丛",
             cat_nickname="黑猫带一点白", cat_gender="male",
             age_estimate="6个月-1岁（最佳）", weight_estimate="6",
             health_note="花色特征：黑猫带一点白；亲人程度：亲人，随便摸",
             post_surgery_plan="医院住院", status="surgery_completed",
             created_at=datetime(2026, 4, 24, 12, 31),
             updated_at=datetime(2026, 4, 24, 12, 31)),
        dict(applicant_name="张春晓", phone="19856109910", clinic_store="龙华店",
             appointment_at="2026-04-26", location_address="中国广东省深圳市",
             id_number="340603199502220224", address="1980科技文化产业园；停车场",
             cat_nickname="黑白", cat_gender="female",
             age_estimate="6个月-1岁（最佳）", weight_estimate="3.5",
             health_note="花色特征：黑白；怀孕/哺乳：是，肚子很大/乳头红肿有奶；亲人程度：可摸但警惕",
             post_surgery_plan="医院住院", status="cancelled",
             created_at=datetime(2026, 4, 25, 20, 33),
             updated_at=datetime(2026, 4, 25, 20, 33)),
    ]
    try:
        for r in records:
            exists = db.query(Application).filter(Application.phone == r["phone"]).first()
            if exists:
                lines.append(f"跳过（已存在）：{r['applicant_name']} id={exists.id}")
                continue
            app_row = Application(wechat_openid="", agree_ear_tip=True,
                                  agree_no_pet_fraud=True, is_proxy=False, **r)
            db.add(app_row)
            db.flush()
            lines.append(f"✅ 已插入：{r['applicant_name']} id={app_row.id}")
        db.commit()
    except Exception:
        db.rollback()
        tb = traceback.format_exc()
        return HTMLResponse(f"<pre style='color:red;padding:2rem'>{tb}</pre>", status_code=500)
    msg = "\n".join(lines)
    return HTMLResponse(f"<pre style='padding:2rem;font-size:1.1rem'>{msg}\n\n<a href='/admin'>← 返回后台</a></pre>")


@app.get("/admin/changelog", response_class=HTMLResponse)
async def admin_changelog_page(request: Request):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login", status_code=303)
    import subprocess, re

    # ── 提交中文描述映射表（短哈希 → 中文说明）──────────────────
    _COMMIT_ZH: dict[str, str] = {
        "853e7be": "问题反馈页新增两家门店联系电话、微信及地址",
        "5447ee7": "修复：预约看板突破页面宽度限制，真正撑满全屏",
        "5468e59": "预约列表改版为三栏看板（上午 / 下午 / 晚上）",
        "cb8460c": "补充美容预约规则和各体型时长对照表至预约规则说明",
        "25c5731": "新增重复提交检测；TNR 提交成功后自动订阅手术完成通知",
        "1e8b16e": "修复：代申请人信息输入框高度过小、文字不显示",
        "0b25cf5": "修复：提交 TNR/手术预约时同步订阅手术完成通知模板",
        "4216c8c": "新增客户问题反馈功能（文字 + 截图上传，后台管理与处理）",
        "15a3f57": "修复：Application 模型缺失代预约字段导致提交失败",
        "0e4c84c": "修复：TNR 提交失败时前端显示详细错误信息",
        "e936282": "后台新增开发日志页面（自动从 Git 记录生成）",
        "67f6784": "新增专属手术提醒订阅消息模板（预约提醒）",
        "db64cc6": "修复：服务重启后不再误发手术提醒推送",
        "8433391": "新增客户自助改约功能（仅限待确认状态）",
        "338c13f": "TNR 申请页新增代预约功能",
        "b52b2d0": "修复：时段容量池说明在后台页面不显示的问题",
        "4b0d67a": "时段容量池按时长等比调整（上午9 / 下午18 / 晚上12 单位）",
        "fc28caa": "TNR/手术容量单位调整为4，预约管理页新增容量池规则说明",
        "caa7a7e": "修复：TNR/手术不再占用时段容量池，避免连续手术被误拒",
        "5315231": "新增代预约功能，记录代预约人姓名、电话及与实际申请人关系",
        "cdebebc": "后台新预约待确认通知：导航橙色角标 + 页面提示条",
        "d60f47d": "修复：爱心展示入口移至「我的预约」按钮下方",
        "7b3af92": "小程序新增 TNR 爱心展示页面",
        "439c2b0": "新增员工档案与合同管理模块，优化账号管理",
        "bf68614": "修复：固定 bcrypt 版本以兼容 passlib 1.7.4",
        "28a860e": "新增多账号权限管理系统",
        "94e3369": "补全预约规则两条缺失说明",
        "0f6a818": "更新预约规则说明，同步当前通知开放状态",
        "b548fe8": "预约页体验优化四项改进",
        "025cb96": "安全加固后台登录页，修复 textarea 跳顶问题",
        "1ed1c53": "优化订阅通知错误处理，添加诊断日志并修复编码问题",
        "5ecc4cf": "修复：遵守微信每次最多3个模板限制，拆分订阅时机",
        "357caa2": "新增「待人工审核」小程序推送通知",
        "7708658": "新增「审核不通过」小程序推送通知",
        "59f0201": "修复：爱心展示默认关闭，手术完成后需管理员手动授权才公开",
        "3023e71": "预约页 TNR 板块仅对申请已通过的用户开放",
        "7b2f570": "保存并展示定位文字地址，删除调试提示文字",
        "fb578bd": "修复：模板订阅改为 API + Storage 合并取值，互补不互斥",
        "027dacf": "修复：订阅授权每次先从 API 拉取全部模板，Storage 仅降级备用",
        "f982c6b": "修复：订阅授权始终从 API 拉取模板 ID，避免 Storage 缓存导致漏订阅",
        "24bf5bf": "新增预约通知模板订阅；/api/wechat/config 返回预约模板字段",
        "34c91aa": "预约通知改用独立模板，正确映射各字段",
        "a1b84a9": "修复预约通知日志：加 rollback 防止 NOT NULL 报错崩溃",
        "3256ed1": "数据清理功能完善：支持删除预约、一键清空全部数据",
        "5626011": "修复提交成功页：状态显示中文，AI 结论为空时不显示",
        "dbc0739": "期望手术日期改为可选字段",
        "b943bb1": "网页端降级为备用入口，申请页顶部加小程序引导",
        "3381914": "功能完善七项改进 + HTTPS 上线配置",
        "bbb7f1d": "项目初始化导入",
    }

    commits = []
    try:
        result = subprocess.run(
            ["git", "log", "--format=%H|%h|%s|%an|%ad", "--date=format:%Y-%m-%d %H:%M", "-500"],
            capture_output=True, text=True, timeout=8,
            cwd=Path(__file__).parent.parent,
        )
        for line in result.stdout.strip().splitlines():
            parts = line.split("|", 4)
            if len(parts) == 5:
                full_hash, short_hash, subject, author, date = parts
                # 提取类型前缀（feat/fix/chore/…）
                m = re.match(r"^(feat|fix|chore|refactor|docs|style|test|perf|build|ci|revert)(\(.+?\))?:\s*(.+)$", subject)
                if m:
                    kind = m.group(1)
                    scope = (m.group(2) or "").strip("()")
                    msg_raw = m.group(3)
                else:
                    kind = "update"
                    scope = ""
                    msg_raw = subject
                # 优先使用中文说明，无则保留原文
                msg = _COMMIT_ZH.get(short_hash, msg_raw)
                commits.append({
                    "full_hash": full_hash,
                    "short_hash": short_hash,
                    "subject": subject,
                    "msg": msg,
                    "kind": kind,
                    "scope": scope,
                    "author": author,
                    "date": date,
                })
    except Exception as e:
        commits = [{"short_hash": "—", "msg": f"无法读取 git log：{e}", "kind": "error", "scope": "", "author": "", "date": "", "subject": "", "full_hash": ""}]
    return templates.TemplateResponse(request, "uk/changelog.html", {
        "request": request,
        "title": "开发日志",
        "commits": commits,
    })


# ────────────────────────────────────────────────────────────────
# 审计日志：单据作废 / 复制 / 锁定相关操作的全部痕迹
# ────────────────────────────────────────────────────────────────
_AUDIT_ACTION_GROUPS = {
    "void":      ("作废",      "#dc2626"),
    "copy_from": ("复制为新单", "#3b82f6"),
    "cert_locked": ("证号录入锁定", "#7c3aed"),
    "delete":    ("删除",      "#dc2626"),
    "manual_approve": ("人工通过", "#10b981"),
    "reject":    ("驳回",      "#dc2626"),
    "surgery_done": ("手术完成", "#10b981"),
}

_AUDIT_DOC_TYPE_ZH = {
    "prescription":   "处方单",
    "sales_order":    "销售单",
    "anesthesia":     "麻醉单",
    "exam_order":     "检查单",
    "vaccination":    "疫苗单",
    "deworming":      "驱虫单",
    "rabies":         "狂犬登记",
    "application":    "TNR 申请",
}


def _parse_audit_detail(detail: str) -> dict:
    """解析 audit detail 字符串，形如 'id=42 reason=用错药 src=41'"""
    out = {}
    if not detail:
        return out
    # 简单 key=value 分隔（允许 value 含空格直到下一个 key=）
    import re
    pairs = re.findall(r"(\w+)=([^=]+?)(?=\s+\w+=|$)", detail)
    for k, v in pairs:
        out[k] = v.strip()
    if not pairs:
        out["note"] = detail
    return out


@app.get("/admin/audit-logs", response_class=HTMLResponse)
async def page_admin_audit_logs(
    request: Request,
    db: Session = Depends(get_db),
    q: str = Query(""),
    action: str = Query(""),
    actor: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    page: int = Query(1),
):
    """审计日志查询页：单据作废 / 复制 / TNR 审核动作 等"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    # 仅超管可见
    if request.session.get("admin_role") != "superadmin":
        raise HTTPException(403, "仅超级管理员可访问审计日志")

    query = db.query(AuditLog)
    if action:
        # 支持模糊匹配前缀（prescription.void 等）
        if "." in action:
            query = query.filter(AuditLog.action == action)
        else:
            query = query.filter(AuditLog.action.ilike(f"%{action}%"))
    if actor:
        query = query.filter(AuditLog.actor.ilike(f"%{actor}%"))
    if date_from:
        query = query.filter(AuditLog.created_at >= date_from)
    if date_to:
        query = query.filter(AuditLog.created_at <= date_to + " 23:59:59")
    if q:
        like = f"%{q}%"
        query = query.filter((AuditLog.detail.ilike(like)) | (AuditLog.action.ilike(like)))

    total = query.count()
    page_size = 50
    rows = query.order_by(AuditLog.id.desc())\
        .offset((page - 1) * page_size).limit(page_size).all()

    # 解析每行的 doc_type/action_kind + detail
    for r in rows:
        if "." in (r.action or ""):
            doc_type, action_kind = r.action.split(".", 1)
        else:
            doc_type, action_kind = "", (r.action or "")
        r._doc_type = doc_type
        r._doc_type_zh = _AUDIT_DOC_TYPE_ZH.get(doc_type, doc_type)
        r._action_kind = action_kind
        label, color = _AUDIT_ACTION_GROUPS.get(action_kind, (action_kind, "#64748b"))
        r._action_zh = label
        r._action_color = color
        r._detail_kv = _parse_audit_detail(r.detail)

    # 统计概览（最近 30 天按 action_kind 分组）
    from datetime import date as _date, timedelta as _td
    thirty_days_ago = (_date.today() - _td(days=30)).isoformat()
    stats_rows = db.query(AuditLog.action, AuditLog.id).filter(
        AuditLog.created_at >= thirty_days_ago
    ).all()
    stats: dict = {}
    for a, _ in stats_rows:
        if not a:
            continue
        kind = a.split(".", 1)[1] if "." in a else a
        stats[kind] = stats.get(kind, 0) + 1

    # 操作类型下拉候选
    distinct_actions = [r[0] for r in db.query(AuditLog.action).distinct().limit(50).all() if r[0]]

    total_pages = max(1, (total + page_size - 1) // page_size)
    return templates.TemplateResponse(request, "uk/audit_logs.html", {
        "rows": rows,
        "total": total, "page": page, "total_pages": total_pages, "page_size": page_size,
        "q": q, "action": action, "actor": actor, "date_from": date_from, "date_to": date_to,
        "action_groups": _AUDIT_ACTION_GROUPS,
        "doc_type_zh": _AUDIT_DOC_TYPE_ZH,
        "stats": stats,
        "distinct_actions": sorted(distinct_actions),
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/hr", response_class=HTMLResponse)
async def admin_hr_page(
    request: Request,
    db: Session = Depends(get_db),
    msg: str = Query(""),
    err: str = Query(""),
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login")
    _hr_admin_store = _get_admin_store(request)
    _staff_q = db.query(Staff)
    if _hr_admin_store:
        _staff_q = _staff_q.filter(Staff.store == _hr_admin_store)
    active_staff = _staff_q.filter(Staff.status != StaffStatus.resigned.value).order_by(Staff.hire_date).all()
    resigned_staff = _staff_q.filter(Staff.status == StaffStatus.resigned.value).order_by(Staff.resign_date.desc()).all()
    expiring = _expiring_contracts(db)
    all_users = db.query(AdminUser).order_by(AdminUser.created_at).all()
    return templates.TemplateResponse(request, "uk/admin_hr.html", {  # B10 UK 重写
        "request": request, "title": "人事管理",
        "active_staff": active_staff, "resigned_staff": resigned_staff,
        "expiring": expiring,
        "active_users": [u for u in all_users if u.is_active],
        "inactive_users": [u for u in all_users if not u.is_active],
        "current_username": request.session.get("admin_username", ""),
        "csrf_token": _get_csrf_token(request),
        "msg": msg, "err": err,
    })


@app.get("/admin/users", response_class=HTMLResponse)
async def admin_users_page(request: Request):
    """旧的独立账号管理页已合并到 /admin/hr 底部，永久跳转。"""
    return RedirectResponse("/admin/hr", status_code=302)


@app.post("/admin/users/create", name="admin_users_create")
async def admin_users_create(
    request: Request,
    db: Session = Depends(get_db),
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form("staff"),
    store: str = Form(""),
    display_name: str = Form(""),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    username = username.strip()
    display_name = display_name.strip()
    # 没填显示名 → 用用户名当显示名（用户没在 UI 里手动填，但同一个字段保存两份）
    if not display_name:
        display_name = username
    if not username or not password:
        return RedirectResponse("/admin/hr?err=用户名和密码不能为空", status_code=303)
    if len(password) < 6:
        return RedirectResponse("/admin/hr?err=密码不能少于6位", status_code=303)
    if role not in ("superadmin", "staff"):
        role = "staff"
    store = store.strip() if store.strip() in _STORE_OPTIONS else ""
    existing = db.query(AdminUser).filter(AdminUser.username == username).first()
    if existing:
        return RedirectResponse(f"/admin/hr?err=用户名已存在：{username}", status_code=303)
    new_user = AdminUser(username=username, password_hash=_pwd_ctx.hash(password),
                          role=role, is_active=True, store=store, display_name=display_name)
    db.add(new_user)
    db.flush()  # 获取 new_user.id
    _audit(db, request, "admin_user_create", application_id=None, detail={"username": username, "role": role})
    db.commit()
    return RedirectResponse(f"/admin/hr?msg=已创建账号：{username}", status_code=303)


@app.post("/admin/users/{user_id}/toggle", name="admin_users_toggle")
async def admin_users_toggle(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    if user.username == request.session.get("admin_username", ""):
        return RedirectResponse("/admin/hr?err=不能停用当前登录账号", status_code=303)
    user.is_active = not user.is_active
    _audit(db, request, "admin_user_toggle", application_id=None, detail={"username": user.username, "active": user.is_active})
    db.commit()
    status_zh = "启用" if user.is_active else "停用"
    return RedirectResponse(f"/admin/hr?msg=已{status_zh}账号：{user.username}", status_code=303)


@app.post("/admin/users/{user_id}/reset-password", name="admin_users_reset_password")
async def admin_users_reset_password(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    new_password: str = Form(...),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    new_password = (new_password or "").strip()
    if not new_password or len(new_password) < 6:
        return RedirectResponse("/admin/hr?err=新密码不能少于6位", status_code=303)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    # 顺手 trim 用户名 — 历史数据可能 import 时带了首尾空格，导致登录时 username 无法 exact-match
    _orig = user.username or ""
    if _orig != _orig.strip():
        user.username = _orig.strip()
    user.password_hash = _pwd_ctx.hash(new_password)
    # 顺手把账号激活（万一被停用了）
    user.is_active = True
    _audit(db, request, "admin_user_reset_password", application_id=None, detail={"username": user.username})
    db.commit()
    # 显示带引号的用户名，方便发现首尾空格、全角字符等问题
    return RedirectResponse(
        f"/admin/hr?msg=密码已重置 · 登录用户名：「{user.username}」 长度{len(user.username)} · 已确保账号启用",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/set-role", name="admin_users_set_role")
async def admin_users_set_role(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    role: str = Form(...),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    if role not in ("superadmin", "staff"):
        return RedirectResponse("/admin/hr?err=角色参数无效", status_code=303)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    if user.username == request.session.get("admin_username", ""):
        return RedirectResponse("/admin/hr?err=不能修改自己的角色", status_code=303)
    user.role = role
    _audit(db, request, "admin_user_set_role", application_id=None, detail={"username": user.username, "role": role})
    db.commit()
    role_zh = "超级管理员" if role == "superadmin" else "员工"
    return RedirectResponse(f"/admin/hr?msg=已将「{user.username}」的角色改为{role_zh}", status_code=303)


@app.post("/admin/users/{user_id}/set-display-name", name="admin_users_set_display_name")
async def admin_users_set_display_name(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    display_name: str = Form(""),
    csrf_token: str = Form(""),
):
    """设置/修改账号的显示名（医生真名）。回访任务按此匹配 Visit.vet_name。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    user.display_name = (display_name or "").strip()[:80]
    _audit(db, request, "admin_user_set_display_name", application_id=None,
           detail={"username": user.username, "display_name": user.display_name})
    db.commit()
    return RedirectResponse(
        f"/admin/hr?msg=已将「{user.username}」的显示名改为{user.display_name or '（已清空）'}",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/set-wecom-userid", name="admin_users_set_wecom_userid")
async def admin_users_set_wecom_userid(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    wecom_userid: str = Form(""),
    csrf_token: str = Form(""),
):
    """绑定企业微信 userid（Phase 1 单点登录用）。
    在企业微信管理后台 → 通讯录 → 点员工 → 看「账号」字段就是 userid。
    """
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    new_uid = (wecom_userid or "").strip()[:80]
    # 同一企微 userid 只允许绑一个账号
    if new_uid:
        clash = db.query(AdminUser).filter(
            AdminUser.wecom_userid == new_uid,
            AdminUser.id != user_id,
        ).first()
        if clash:
            return RedirectResponse(
                f"/admin/hr?err=该企微 userid 已绑定到账号「{clash.username}」",
                status_code=303,
            )
    user.wecom_userid = new_uid
    _audit(db, request, "admin_user_set_wecom_userid", application_id=None,
           detail={"username": user.username, "wecom_userid": new_uid})
    db.commit()
    return RedirectResponse(
        f"/admin/hr?msg=已为「{user.username}」绑定企微 userid={new_uid or '（已清空）'}",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/set-mobile-role", name="admin_users_set_mobile_role")
async def admin_users_set_mobile_role(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    mobile_role: str = Form("auto"),
    csrf_token: str = Form(""),
):
    """设置员工手机端默认身份（doctor / nurse / groomer / auto）。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    mr = (mobile_role or "auto").strip().lower()
    if mr not in ("auto", "doctor", "nurse", "groomer"):
        mr = "auto"
    user.mobile_role = mr
    _audit(db, request, "admin_user_set_mobile_role", application_id=None,
           detail={"username": user.username, "mobile_role": mr})
    db.commit()
    return RedirectResponse(f"/admin/hr?msg=已为「{user.username}」设手机端身份={mr}", status_code=303)


@app.post("/admin/wecom-notify/dispatch-now", name="admin_wecom_dispatch_now")
async def admin_wecom_dispatch_now(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """手动触发：把当前每个员工的「今日待办」推送到他们的企业微信。

    遍历所有 active + 已绑 wecom_userid 的员工：按各自门店计算 workbench，
    没有待办则跳过，有待办推一张摘要卡。
    """
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    from app.services import wecom_notify as _notify
    try:
        stat = _notify.dispatch_workbench_to_all(db)
    except Exception as e:
        logger.warning("[wecom dispatch] failed: %s", e)
        return RedirectResponse(
            f"/admin/customers?err=推送失败：{str(e)[:100]}",
            status_code=303,
        )
    msg_parts = [f"已推送 {stat['sent']} 人"]
    if stat["skipped"]:
        msg_parts.append(f"{stat['skipped']} 人无待办跳过")
    if stat["failed"]:
        msg_parts.append(f"失败 {len(stat['failed'])}：" + " | ".join(stat["failed"][:2]))
    return RedirectResponse(
        f"/admin/customers?msg=" + "，".join(msg_parts),
        status_code=303,
    )


# ═════════════════════════════════════════════════════════════
# Phase 3 Step 1 — 企微外部联系人 ↔ Customer 映射
# ═════════════════════════════════════════════════════════════

@app.get("/api/wecom/jssdk-config", name="api_wecom_jssdk_config")
async def api_wecom_jssdk_config(request: Request, url: str = Query(..., description="完整页面 URL（去 hash）")):
    """返回企业微信 JS-SDK 鉴权配置。

    不要求登录：签名本身不敏感（只是 ticket + url + nonce 的 SHA1），
    且企微 iframe 内 cookies 不一定能传过来。

    用法（前端）：
      fetch('/api/wecom/jssdk-config?url=' + encodeURIComponent(location.href))
        .then(r => r.json())
        .then(cfg => {
          wx.config({...cfg.config, jsApiList: ['agentConfig']});
          wx.ready(() => {
            wx.agentConfig({...cfg.agent_config, jsApiList: ['getCurExternalContact'], success() { ... }});
          });
        });
    """
    from app.services import wecom_client as _wc
    if not _wc.enabled():
        return {"error": "wecom not configured"}
    try:
        corp_cfg = _wc.build_jsapi_signature(url, agent=False)
        agent_cfg = _wc.build_jsapi_signature(url, agent=True)
    except Exception as e:
        logger.warning("[jssdk-config] %s", e)
        raise HTTPException(500, f"签名生成失败：{e}")
    # wx.config 用 corp 级别
    config = {
        "beta": True,
        "debug": False,
        "appId": corp_cfg["appId"],
        "timestamp": corp_cfg["timestamp"],
        "nonceStr": corp_cfg["nonceStr"],
        "signature": corp_cfg["signature"],
    }
    # wx.agentConfig 用 agent 级别（同时带 corpid + agentid）
    agent_config = {
        "corpid": agent_cfg["appId"],
        "agentid": agent_cfg["agentid"],
        "timestamp": agent_cfg["timestamp"],
        "nonceStr": agent_cfg["nonceStr"],
        "signature": agent_cfg["signature"],
    }
    return {"config": config, "agent_config": agent_config}


@app.get("/admin/wecom-sidebar", response_class=HTMLResponse, name="admin_wecom_sidebar")
async def admin_wecom_sidebar(
    request: Request,
    db: Session = Depends(get_db),
    external_userid: str = Query("", description="企微外部联系人 ID（聊天客户）"),
):
    """企微聊天侧边栏 H5：员工跟客户聊天时显示客户档案。

    Phase 3 Step 2。
    企微侧边栏会自动在 URL 拼 ?external_userid=xxx（部分版本需要 JS-SDK 主动调用获取）。
    """
    # 未登录 → 走企微 OAuth 静默登录 → 回到这个 URL（带原 external_userid 参数）
    if not _admin_ok(request):
        from urllib.parse import quote as _q
        next_url = f"/admin/wecom-sidebar?external_userid={_q(external_userid)}"
        return RedirectResponse(f"/admin/wecom-login?next={_q(next_url, safe='')}", status_code=302)

    from app.models import WecomCustomerLink

    link = None
    cust = None
    pets = []
    recent_visits = []
    wallet = None
    if external_userid:
        link = db.query(WecomCustomerLink).filter(
            WecomCustomerLink.external_userid == external_userid
        ).first()
        if link and link.customer_id:
            cust = db.get(Customer, link.customer_id)
            if cust:
                pets = db.query(Pet).filter(Pet.customer_id == cust.id).all()
                recent_visits = (
                    db.query(Visit)
                    .filter(Visit.customer_id == cust.id)
                    .order_by(Visit.id.desc())
                    .limit(5).all()
                )
                wallet = db.query(Wallet).filter(Wallet.customer_id == cust.id).first()

    # 计算待办：未付费 / 未签协议 / 押金未结算
    pending_invoices = 0
    pending_consents = 0
    held_deposits = 0
    if cust:
        from app.models import Invoice, ConsentTask, Deposit
        pending_invoices = db.query(Invoice).filter(
            Invoice.customer_id == cust.id,
            Invoice.payment_status.in_(("unpaid", "partial")),
        ).count()
        pending_consents = db.query(ConsentTask).filter(
            ConsentTask.customer_id == cust.id,
            ConsentTask.status == "pending",
        ).count()
        held_deposits = db.query(Deposit).filter(
            Deposit.customer_id == cust.id,
            Deposit.status == "held",
        ).count()

    return templates.TemplateResponse(request, "uk/wecom_sidebar.html", {
        "external_userid": external_userid,
        "link": link,
        "cust": cust,
        "pets": pets,
        "recent_visits": recent_visits,
        "wallet": wallet,
        "pending_invoices": pending_invoices,
        "pending_consents": pending_consents,
        "held_deposits": held_deposits,
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/wecom-customers", response_class=HTMLResponse, name="admin_wecom_customers_list")
async def admin_wecom_customers_list(
    request: Request,
    db: Session = Depends(get_db),
    status: str = Query("", description="筛选：matched/unmatched/created/ignored，空=全部"),
    q: str = Query("", description="按 remark_name / remark_mobile / name 搜"),
    page: int = Query(1, ge=1),
):
    require_admin(request)
    require_superadmin(request)
    from app.models import WecomCustomerLink
    query = db.query(WecomCustomerLink)
    if status:
        query = query.filter(WecomCustomerLink.sync_status == status)
    if q.strip():
        kw = f"%{q.strip()}%"
        query = query.filter(or_(
            WecomCustomerLink.remark_name.like(kw),
            WecomCustomerLink.remark_mobile.like(kw),
            WecomCustomerLink.name.like(kw),
        ))
    total = query.count()
    page_size = 50
    links = (
        query.order_by(WecomCustomerLink.sync_status.asc(), WecomCustomerLink.id.desc())
        .offset((page - 1) * page_size).limit(page_size).all()
    )
    # 统计各状态数量
    from sqlalchemy import func as _f
    counts_raw = (
        db.query(WecomCustomerLink.sync_status, _f.count(WecomCustomerLink.id))
        .group_by(WecomCustomerLink.sync_status).all()
    )
    counts = {s: n for s, n in counts_raw}
    counts["total"] = sum(counts.values())
    return templates.TemplateResponse(request, "uk/wecom_customers.html", {
        "links": links, "counts": counts, "status": status, "q": q,
        "page": page, "total": total,
        "total_pages": max(1, (total + page_size - 1) // page_size),
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
        "err": request.query_params.get("err"),
    })


@app.post("/admin/wecom-customers/sync", name="admin_wecom_customers_sync")
async def admin_wecom_customers_sync(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    dry_run: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    from app.services.wecom_customers import sync_all
    is_dry = bool(dry_run)
    try:
        stats = sync_all(db, dry_run=is_dry)
    except Exception as e:
        logger.warning("[wecom sync] failed: %s", e)
        return RedirectResponse(f"/admin/wecom-customers?err=同步失败：{str(e)[:120]}", status_code=303)
    prefix = "试运行：" if is_dry else "同步完成："
    msg = (
        f"{prefix}拉取 {stats['pulled']} 个客户，自动匹配 {stats['matched']}，"
        f"待匹配 {stats['unmatched']}"
    )
    if not is_dry:
        msg += f"（新建 {stats['created_links']} 条 / 更新 {stats['updated_links']} 条）"
    if stats["errors"]:
        msg += f" · 错误 {len(stats['errors'])}：" + (stats["errors"][0][:80])
    return RedirectResponse(f"/admin/wecom-customers?msg={msg}", status_code=303)


@app.post("/admin/wecom-customers/batch-create", name="admin_wecom_customers_batch_create")
async def admin_wecom_customers_batch_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """批量为多条 unmatched 链接新建 Customer 档案。

    每条 link 用 remark_name（企微备注名）作为 Customer.name；
    若 remark_name 为空，回退到 wechat name；都没有则 "（企微未命名）"。
    备注手机号若有则填入 Customer.phone。
    """
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    form = await request.form()
    link_ids = form.getlist("link_ids") if hasattr(form, "getlist") else []
    if not link_ids:
        # Starlette UploadFile-style; try multi-value getter
        raw = form.get("link_ids", "")
        if isinstance(raw, str) and raw:
            link_ids = [x for x in raw.split(",") if x.strip()]
    if not link_ids:
        return RedirectResponse("/admin/wecom-customers?err=未选择任何记录", status_code=303)

    from app.models import WecomCustomerLink
    created = 0
    skipped = 0
    for raw_id in link_ids:
        try:
            lid = int(raw_id)
        except (TypeError, ValueError):
            continue
        link = db.get(WecomCustomerLink, lid)
        if not link or link.customer_id:
            skipped += 1
            continue
        cust_name = (link.remark_name or link.name or "（企微未命名）").strip()[:120]
        cust = Customer(
            name=cust_name,
            phone=link.remark_mobile or "",
            address="",
            source="wecom",
        )
        db.add(cust)
        db.flush()
        link.customer_id = cust.id
        link.sync_status = "created"
        created += 1
    db.commit()
    return RedirectResponse(
        f"/admin/wecom-customers?msg=批量新建完成：新建 {created} 个档案，跳过 {skipped} 条",
        status_code=303,
    )


@app.post("/admin/wecom-customers/{link_id}/create-customer", name="admin_wecom_customers_create")
async def admin_wecom_customers_create(
    link_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """为 unmatched 链接新建一个 Customer 记录 + 自动 link。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    from app.models import WecomCustomerLink
    link = db.get(WecomCustomerLink, link_id)
    if not link:
        raise HTTPException(404)
    if link.customer_id:
        return RedirectResponse(f"/admin/wecom-customers?err=该客户已有匹配档案（#{link.customer_id}）", status_code=303)
    # 用备注名 > 微信昵称作为客户姓名
    cust_name = (link.remark_name or link.name or "").strip() or "（企微未命名）"
    cust = Customer(
        name=cust_name,
        phone=link.remark_mobile or "",
        address="",
        source="wecom",
    )
    db.add(cust)
    db.flush()
    link.customer_id = cust.id
    link.sync_status = "created"
    db.commit()
    return RedirectResponse(
        f"/admin/wecom-customers?msg=已新建客户档案「{cust_name}」并绑定（Customer #{cust.id}）",
        status_code=303,
    )


@app.get("/api/admin/customer/find-by-phone", name="admin_api_customer_find_by_phone")
async def admin_api_customer_find_by_phone(
    request: Request,
    db: Session = Depends(get_db),
    phone: str = Query(""),
):
    """按手机号搜系统内已有客户，给企微客户绑定界面用。"""
    require_admin(request)
    p = "".join(c for c in (phone or "") if c.isdigit())
    if len(p) < 6:
        return {"found": False}
    cust = db.query(Customer).filter(Customer.phone == p).first()
    if not cust:
        return {"found": False}
    pet_count = db.query(Pet).filter(Pet.customer_id == cust.id).count()
    return {
        "found": True,
        "id": cust.id,
        "name": cust.name or "（未命名）",
        "phone": cust.phone,
        "pets": pet_count,
    }


@app.post("/admin/wecom-customers/{link_id}/match", name="admin_wecom_customers_match")
async def admin_wecom_customers_match(
    link_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(...),
):
    """手动把企微客户绑到已有 Customer 上。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    from app.models import WecomCustomerLink
    link = db.get(WecomCustomerLink, link_id)
    if not link:
        raise HTTPException(404)
    cust = db.get(Customer, customer_id)
    if not cust:
        return RedirectResponse(f"/admin/wecom-customers?err=客户 #{customer_id} 不存在", status_code=303)
    link.customer_id = cust.id
    link.sync_status = "matched"
    db.commit()
    return RedirectResponse(
        f"/admin/wecom-customers?msg=已绑定 {link.remark_name or link.name or '客户'} → 档案 {cust.name}（#{cust.id}）",
        status_code=303,
    )


@app.post("/admin/wecom-customers/{link_id}/ignore", name="admin_wecom_customers_ignore")
async def admin_wecom_customers_ignore(
    link_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    from app.models import WecomCustomerLink
    link = db.get(WecomCustomerLink, link_id)
    if not link:
        raise HTTPException(404)
    link.sync_status = "ignored"
    db.commit()
    return RedirectResponse(
        f"/admin/wecom-customers?msg=已忽略 {link.remark_name or link.name or '该客户'}",
        status_code=303,
    )


@app.get("/admin/wecom-customers/probe", response_class=HTMLResponse, name="admin_wecom_customers_probe")
async def admin_wecom_customers_probe(request: Request):
    """Phase 3 探测：调用客户联系基础 API，看 errcode 决定下一步。

    顺序：
      1. get_follow_user_list — 列出能用客户联系的员工
      2. 拿第一个员工的客户联系 list — 看能不能拉客户 external_userid
      3. 拿第一个客户的详情 — 看返回字段有什么
    """
    require_admin(request)
    require_superadmin(request)
    from app.services import wecom_client as _wc
    if not _wc.enabled():
        return HTMLResponse("<p>企业微信未配置</p>", status_code=503)

    import json as _json
    blocks = []
    follow_users: list[str] = []

    # Step 1
    try:
        r1 = _wc.external_get_follow_user_list()
        blocks.append(("① 列出客户联系成员", r1))
        if r1.get("errcode") in (0, "0", None):
            follow_users = r1.get("follow_user", []) or []
    except Exception as e:
        blocks.append(("① 列出客户联系成员", {"exception": str(e)}))

    # Step 2 - 拿第一个员工的 external_userid 列表
    sample_external_userid = ""
    if follow_users:
        try:
            r2 = _wc.external_list_by_userid(follow_users[0])
            blocks.append((f"② 列出 {follow_users[0]} 名下的客户 external_userid", r2))
            eu_list = r2.get("external_userid", []) or []
            if eu_list:
                sample_external_userid = eu_list[0]
        except Exception as e:
            blocks.append(("②", {"exception": str(e)}))

    # Step 3 - 拿一个客户详情
    if sample_external_userid:
        try:
            r3 = _wc.external_get_detail(sample_external_userid)
            blocks.append((f"③ 客户详情（{sample_external_userid[:20]}…）", r3))
        except Exception as e:
            blocks.append(("③", {"exception": str(e)}))

    html_parts = ["""
    <html><head><meta charset="utf-8"><title>客户联系 API 探测</title>
    <style>
      body{font-family:system-ui;padding:1.5rem;max-width:920px;margin:auto;color:#111;}
      h1{font-size:1.2rem;margin:0 0 1rem;}
      h2{font-size:1rem;margin:1.2rem 0 .4rem;color:#1d4ed8;}
      pre{background:#f5f5f5;padding:.85rem;border-radius:8px;overflow:auto;font-size:.82rem;line-height:1.5;}
      .ok{color:#15803d;font-weight:600;}
      .err{color:#b91c1c;font-weight:600;}
      .hint{background:#fef3c7;border-left:3px solid #f59e0b;padding:.7rem 1rem;margin:1rem 0;font-size:.88rem;border-radius:0 8px 8px 0;}
      a{color:#1d4ed8;text-decoration:none;}
    </style></head><body>
    <p><a href="/admin/customers">← 返回工作台</a></p>
    <h1>🔍 客户联系 API 探测</h1>
    """]
    for title, data in blocks:
        ec = data.get("errcode") if isinstance(data, dict) else None
        status = '<span class="ok">errcode=0 ✓</span>' if ec in (0, "0", None) else f'<span class="err">errcode={ec} ✗</span>'
        html_parts.append(f"<h2>{title} {status}</h2>")
        html_parts.append(f"<pre>{_json.dumps(data, ensure_ascii=False, indent=2)}</pre>")

    # 智能提示
    first_errcode = blocks[0][1].get("errcode") if blocks and isinstance(blocks[0][1], dict) else None
    if first_errcode == 60011:
        html_parts.append('<div class="hint"><b>诊断：</b>应用没有客户联系 API 权限。<br>去 <b>客户联系 → 权限配置 → 「可调用接口的应用」</b> 把「大风动物医院 TNR」加进去。</div>')
    elif first_errcode == 48002:
        html_parts.append('<div class="hint"><b>诊断：</b>接口未在白名单。同上，去权限配置授权。</div>')
    elif first_errcode == 60020:
        html_parts.append('<div class="hint"><b>诊断：</b>IP 不在企业可信 IP。但我们之前加过了，可能企微 token 缓存还没刷新，等 2 小时或重启服务。</div>')
    elif first_errcode in (0, "0", None) and blocks:
        n = len(follow_users)
        html_parts.append(f'<div class="hint" style="background:#d1fae5;border-color:#10b981;"><b>✅ 客户联系 API 完全可用！</b><br>当前 {n} 个员工配置了客户联系。可以开始 Phase 3 Step 1：拉取 611 个客户、建映射表。</div>')

    html_parts.append("</body></html>")
    return HTMLResponse("".join(html_parts))


@app.get("/admin/users/{user_id}/notify-prefs", name="admin_users_notify_prefs", response_class=HTMLResponse)
async def admin_users_notify_prefs(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """员工的企微通知偏好（哪些事件要收 / 不收）。"""
    require_admin(request)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    # 超管可以改任何人；普通员工只能改自己
    if request.session.get("admin_role") != "superadmin":
        if request.session.get("admin_username") != user.username:
            raise HTTPException(403, "只能修改自己的通知偏好")
    from app.services.wecom_notify import EVENT_KEYS
    disabled_set = {k.strip() for k in (user.wecom_notify_disabled or "").split(",") if k.strip()}
    return templates.TemplateResponse(request, "uk/user_notify_prefs.html", {
        "user": user,
        "event_keys": EVENT_KEYS,
        "disabled_set": disabled_set,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/users/{user_id}/notify-prefs", name="admin_users_notify_prefs_save")
async def admin_users_notify_prefs_save(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    if request.session.get("admin_role") != "superadmin":
        if request.session.get("admin_username") != user.username:
            raise HTTPException(403)
    _require_csrf(request, csrf_token)
    form = await request.form()
    from app.services.wecom_notify import EVENT_KEYS
    # 表单里 enabled_<key>=on 表示开（收），未提交则视为关
    disabled = [k for k in EVENT_KEYS.keys() if form.get(f"enabled_{k}") != "on"]
    user.wecom_notify_disabled = ",".join(disabled)
    db.commit()
    _audit(db, request, "admin_user_notify_prefs",
           detail={"username": user.username, "disabled": disabled})
    return RedirectResponse(
        f"/admin/users/{user.id}/notify-prefs?msg=偏好已保存（开 {len(EVENT_KEYS) - len(disabled)} 类 / 关 {len(disabled)} 类）",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/wecom-test", name="admin_users_wecom_test")
async def admin_users_wecom_test(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """推送一条测试卡片到该员工的企业微信，验证推送链路是否通畅。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    if not user.wecom_userid:
        return RedirectResponse(
            f"/admin/hr?err=「{user.username}」还没绑定企微 userid，请先绑定",
            status_code=303,
        )
    from app.services import wecom_notify as _notify
    try:
        result = _notify.push_test(user.wecom_userid)
    except Exception as e:
        logger.warning("[wecom test push] failed: %s", e)
        return RedirectResponse(
            f"/admin/hr?err=推送失败：{str(e)[:100]}",
            status_code=303,
        )
    errcode = result.get("errcode")
    if errcode in (0, "0", None):
        return RedirectResponse(
            f"/admin/hr?msg=已向「{user.username}」（{user.wecom_userid}）推送测试卡片，请打开企业微信查看",
            status_code=303,
        )
    return RedirectResponse(
        f"/admin/hr?err=推送失败 errcode={errcode}: {result.get('errmsg', '')[:200]}",
        status_code=303,
    )


@app.post("/admin/users/{user_id}/set-store", name="admin_users_set_store")
async def admin_users_set_store(
    user_id: int,
    request: Request,
    db: Session = Depends(get_db),
    store: str = Form(...),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    store = store.strip()
    if store not in ("", *_STORE_OPTIONS):
        return RedirectResponse("/admin/hr?err=门店参数无效", status_code=303)
    user = db.query(AdminUser).filter(AdminUser.id == user_id).first()
    if not user:
        raise HTTPException(404)
    user.store = store
    _audit(db, request, "admin_user_set_store", application_id=None, detail={"username": user.username, "store": store})
    db.commit()
    store_label = store or "（不限门店）"
    return RedirectResponse(f"/admin/hr?msg=已将「{user.username}」的门店改为{store_label}", status_code=303)


# ── 员工档案 & 合同管理 ─────────────────────────────────────────────────

_STAFF_STATUS_ZH = {"probation": "试用中", "active": "在职", "resigned": "离职"}
_CONTRACT_TYPE_ZH = {"formal": "正式合同", "probation": "试用期合同", "parttime": "兼职合同", "labor": "劳务合同"}
_POSITION_OPTIONS = ["前台", "医生", "美容师", "助理", "收银", "其他"]
_STORE_OPTIONS = ["东环店", "横岗店"]


def _expiring_contracts(db: Session, days: int = 30) -> list:
    """返回 days 天内到期的合同（end_date 非空）。"""
    from datetime import date, timedelta
    today = date.today().isoformat()
    deadline = (date.today() + timedelta(days=days)).isoformat()
    rows = (
        db.query(Contract)
        .filter(Contract.end_date != "", Contract.end_date >= today, Contract.end_date <= deadline)
        .all()
    )
    return rows


@app.get("/admin/staff", response_class=HTMLResponse)
async def admin_staff_list(request: Request):
    """旧的员工列表页已合并到 /admin/hr，永久跳转。"""
    return RedirectResponse("/admin/hr", status_code=302)


@app.get("/admin/staff/create", response_class=HTMLResponse)
async def admin_staff_create_get(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return templates.TemplateResponse(request, "uk/login.html", {"request": request, "title": "医院后台登录", "csrf_token": _get_csrf_token(request)})
    require_superadmin(request)
    return templates.TemplateResponse(request, "uk/staff_form.html", {
        "request": request, "title": "新增员工", "staff": None,
        "position_options": _POSITION_OPTIONS,
        "store_options": _STORE_OPTIONS, "csrf_token": _get_csrf_token(request), "err": "",
    })


@app.post("/admin/staff/create", name="admin_staff_create")
async def admin_staff_create_post(
    request: Request, db: Session = Depends(get_db),
    name: str = Form(...), gender: str = Form(""), birthday: str = Form(""),
    phone: str = Form(""), id_number: str = Form(""), store: str = Form(""),
    position: str = Form(""), hire_date: str = Form(""), probation_end_date: str = Form(""),
    status: str = Form("active"), emergency_contact_name: str = Form(""),
    emergency_contact_phone: str = Form(""), emergency_contact_relation: str = Form(""),
    notes: str = Form(""), csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    name = name.strip()
    if not name:
        return RedirectResponse("/admin/staff/create?err=姓名不能为空", status_code=303)
    s = Staff(
        name=name, gender=gender, birthday=birthday, phone=phone.strip(),
        id_number=id_number.strip(), store=store, position=position,
        hire_date=hire_date, probation_end_date=probation_end_date, status=status,
        emergency_contact_name=emergency_contact_name, emergency_contact_phone=emergency_contact_phone,
        emergency_contact_relation=emergency_contact_relation, notes=notes,
    )
    db.add(s)
    _audit(db, request, "staff_create", application_id=None, detail={"name": name})
    db.commit()
    return RedirectResponse(f"/admin/staff/{s.id}?msg=员工档案已创建", status_code=303)


@app.get("/admin/staff/{staff_id}", response_class=HTMLResponse)
async def admin_staff_detail(
    staff_id: int, request: Request, db: Session = Depends(get_db),
    msg: str = Query(""), err: str = Query(""),
):
    if not _admin_ok(request):
        return templates.TemplateResponse(request, "uk/login.html", {"request": request, "title": "医院后台登录", "csrf_token": _get_csrf_token(request)})
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(404)
    contracts = db.query(Contract).filter(Contract.staff_id == staff_id).order_by(Contract.start_date.desc()).all()
    from datetime import date, timedelta
    today = date.today().isoformat()
    expiry_30 = (date.today() + timedelta(days=30)).isoformat()
    return templates.TemplateResponse(request, "uk/staff_detail.html", {
        "request": request, "title": f"员工档案 · {staff.name}",
        "staff": staff, "contracts": contracts,
        "status_zh": _STAFF_STATUS_ZH, "contract_type_zh": _CONTRACT_TYPE_ZH,
        "csrf_token": _get_csrf_token(request), "msg": msg, "err": err,
        "is_superadmin": _is_superadmin(request),
        "now_date": today, "expiry_30": expiry_30,
    })


@app.get("/admin/staff/{staff_id}/edit", response_class=HTMLResponse)
async def admin_staff_edit_get(staff_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return templates.TemplateResponse(request, "uk/login.html", {"request": request, "title": "医院后台登录", "csrf_token": _get_csrf_token(request)})
    require_superadmin(request)
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(404)
    return templates.TemplateResponse(request, "uk/staff_form.html", {
        "request": request, "title": f"编辑员工 · {staff.name}", "staff": staff,
        "position_options": _POSITION_OPTIONS,
        "store_options": _STORE_OPTIONS, "csrf_token": _get_csrf_token(request), "err": "",
    })


@app.post("/admin/staff/{staff_id}/edit", name="admin_staff_edit")
async def admin_staff_edit_post(
    staff_id: int, request: Request, db: Session = Depends(get_db),
    name: str = Form(...), gender: str = Form(""), birthday: str = Form(""),
    phone: str = Form(""), id_number: str = Form(""), store: str = Form(""),
    position: str = Form(""), hire_date: str = Form(""), probation_end_date: str = Form(""),
    status: str = Form("active"), resign_date: str = Form(""), resign_reason: str = Form(""),
    emergency_contact_name: str = Form(""), emergency_contact_phone: str = Form(""),
    emergency_contact_relation: str = Form(""),
    notes: str = Form(""), csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(404)
    staff.name = name.strip()
    staff.gender = gender; staff.birthday = birthday; staff.phone = phone.strip()
    staff.id_number = id_number.strip(); staff.store = store; staff.position = position
    staff.hire_date = hire_date; staff.probation_end_date = probation_end_date
    staff.status = status; staff.resign_date = resign_date; staff.resign_reason = resign_reason
    staff.emergency_contact_name = emergency_contact_name
    staff.emergency_contact_phone = emergency_contact_phone
    staff.emergency_contact_relation = emergency_contact_relation
    staff.notes = notes
    _audit(db, request, "staff_edit", application_id=None, detail={"staff_id": staff_id, "name": staff.name})
    db.commit()
    return RedirectResponse(f"/admin/staff/{staff_id}?msg=已保存", status_code=303)


@app.post("/admin/staff/{staff_id}/contracts/create", name="admin_contract_create")
async def admin_contract_create(
    staff_id: int, request: Request, db: Session = Depends(get_db),
    contract_type: str = Form("formal"), start_date: str = Form(""),
    end_date: str = Form(""), notes: str = Form(""),
    file: Optional[UploadFile] = File(None), csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    staff = db.query(Staff).filter(Staff.id == staff_id).first()
    if not staff:
        raise HTTPException(404)
    file_path = ""
    original_filename = ""
    if file and file.filename:
        import aiofiles
        ext = Path(file.filename).suffix.lower()
        if ext not in (".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx"):
            return RedirectResponse(f"/admin/staff/{staff_id}?err=合同文件仅支持 PDF/图片/Word", status_code=303)
        save_dir = Path(settings.upload_dir) / "contracts" / str(staff_id)
        save_dir.mkdir(parents=True, exist_ok=True)
        fname = f"{secrets.token_hex(8)}{ext}"
        save_path = save_dir / fname
        async with aiofiles.open(save_path, "wb") as f:
            content = await file.read()
            await f.write(content)
        file_path = str(save_path)
        original_filename = file.filename
    c = Contract(
        staff_id=staff_id, contract_type=contract_type,
        start_date=start_date, end_date=end_date,
        file_path=file_path, original_filename=original_filename, notes=notes,
    )
    db.add(c)
    _audit(db, request, "contract_create", application_id=None, detail={"staff_id": staff_id, "type": contract_type})
    db.commit()
    return RedirectResponse(f"/admin/staff/{staff_id}?msg=合同已添加", status_code=303)


@app.post("/admin/staff/{staff_id}/contracts/{contract_id}/delete", name="admin_contract_delete")
async def admin_contract_delete(
    staff_id: int, contract_id: int, request: Request,
    db: Session = Depends(get_db), csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    c = db.query(Contract).filter(Contract.id == contract_id, Contract.staff_id == staff_id).first()
    if not c:
        raise HTTPException(404)
    if c.file_path and Path(c.file_path).exists():
        Path(c.file_path).unlink(missing_ok=True)
    db.delete(c)
    _audit(db, request, "contract_delete", application_id=None, detail={"staff_id": staff_id, "contract_id": contract_id})
    db.commit()
    return RedirectResponse(f"/admin/staff/{staff_id}?msg=合同已删除", status_code=303)


@app.get("/admin/staff/{staff_id}/contracts/{contract_id}/file")
async def admin_contract_file(
    staff_id: int, contract_id: int, request: Request, db: Session = Depends(get_db),
):
    require_admin(request)
    c = db.query(Contract).filter(Contract.id == contract_id, Contract.staff_id == staff_id).first()
    if not c or not c.file_path or not Path(c.file_path).exists():
        raise HTTPException(404)
    return FileResponse(c.file_path, filename=c.original_filename or Path(c.file_path).name)


@app.post("/admin/backup/create", name="admin_backup_create")
async def admin_backup_create(request: Request, db: Session = Depends(get_db), csrf_token: str = Form("")):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    try:
        out = create_backup_zip()
    except Exception as e:
        return RedirectResponse("/admin?backup_err=1&reason=" + quote(str(e)[:120], safe=""), status_code=303)
    _audit(db, request, "backup_create", application_id=None, detail={"file": out.name, "size": out.stat().st_size})
    db.commit()
    return RedirectResponse("/admin?backup_ok=1&file=" + quote(out.name, safe=""), status_code=303)


@app.get("/admin/backup/download/{filename}", name="admin_backup_download")
async def admin_backup_download(request: Request, filename: str, db: Session = Depends(get_db)):
    require_admin(request)
    require_superadmin(request)
    if not is_safe_backup_filename(filename):
        raise HTTPException(404)
    root = Path(settings.backup_dir).resolve()
    path = (root / filename).resolve()
    if not str(path).startswith(str(root)) or not path.is_file():
        raise HTTPException(404)
    _audit(db, request, "backup_download", application_id=None, detail={"file": filename})
    db.commit()
    return FileResponse(path, filename=filename, media_type="application/zip")


def _rmtree_app_uploads(app_id: int) -> None:
    base = Path(settings.upload_dir) / str(app_id)
    if base.exists() and base.is_dir():
        shutil.rmtree(base, ignore_errors=True)


@app.get("/admin/purge", response_class=HTMLResponse)
async def admin_purge_get_hint():
    """避免误用 GET 打开 /admin/purge 时出现 JSON 404；清理必须用后台表单 POST。"""
    return HTMLResponse(
        '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8"/><title>数据清理</title></head>'
        '<body style="font-family:sans-serif;padding:1.5rem;">'
        "<p>数据清理仅接受 <strong>POST</strong>（请在医院后台页面内提交表单）。</p>"
        '<p><a href="/admin">返回医院后台</a></p></body></html>',
        status_code=200,
    )


async def _admin_purge_run(
    request: Request,
    db: Session,
    csrf_token: str,
    scope: str,
    confirm: str,
):
    """一键清理：scope=all/drafts/appointments/everything"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    scope = (scope or "all").strip().lower()
    if scope not in ("all", "drafts", "appointments", "everything"):
        return RedirectResponse("/admin?purge_err=1", status_code=303)
    confirm = (confirm or "").strip()

    CONFIRM_MAP = {
        "drafts":       "确认删除全部草稿",
        "all":          "确认删除全部申请数据",
        "appointments": "确认删除全部预约数据",
        "everything":   "确认删除全部数据",
    }
    if confirm != CONFIRM_MAP[scope]:
        return RedirectResponse("/admin?purge_err=1", status_code=303)

    n = 0
    if scope in ("drafts", "all", "everything"):
        if scope == "drafts":
            q = db.query(Application).filter(Application.status == ApplicationStatus.draft.value)
        else:
            q = db.query(Application)
        rows = q.all()
        n += len(rows)
        for row in rows:
            _rmtree_app_uploads(row.id)
            db.delete(row)
        if scope in ("all", "everything"):
            db.query(AuditLog).delete(synchronize_session=False)

    if scope in ("appointments", "everything"):
        appt_count = db.query(Appointment).delete(synchronize_session=False)
        n += appt_count

    db.commit()
    _audit(db, request, "purge_" + scope, application_id=None, detail={"deleted": n})
    db.commit()
    return RedirectResponse(f"/admin?purge_ok=1&deleted={n}&what={scope}", status_code=303)


@app.post("/admin/media/{media_id}/delete", name="admin_media_delete")
async def admin_media_delete(
    media_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login", status_code=303)
    form = await request.form()
    if form.get("csrf_token") != request.session.get("csrf_token"):
        return RedirectResponse("/admin?err=csrf", status_code=303)
    m = db.get(MediaFile, media_id)
    if not m:
        return RedirectResponse("/admin?err=文件不存在", status_code=303)
    app_id = m.application_id
    try:
        p = Path(m.stored_path)
        if p.exists():
            p.unlink()
    except Exception:
        pass
    db.delete(m)
    db.commit()
    return _admin_back(request, app_id, "文件已删除")


@app.post("/admin/application/{app_id}/edit-cat", name="admin_edit_cat")
async def admin_edit_cat(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login", status_code=303)
    form = await request.form()
    if form.get("csrf_token") != request.session.get("csrf_token"):
        return RedirectResponse("/admin?err=csrf", status_code=303)
    row = db.get(Application, app_id)
    if not row:
        return RedirectResponse("/admin?err=申请不存在", status_code=303)
    cat_nickname = (form.get("cat_nickname") or "").strip()
    cat_gender = (form.get("cat_gender") or "").strip()
    age_estimate = (form.get("age_estimate") or "").strip()
    health_note = (form.get("health_note") or "").strip()
    clinic_store = (form.get("clinic_store") or "").strip()
    if cat_nickname:
        row.cat_nickname = cat_nickname
    if cat_gender in ("male", "female", "unknown"):
        row.cat_gender = cat_gender
    row.age_estimate = age_estimate
    row.health_note = health_note
    if clinic_store:
        row.clinic_store = clinic_store
    db.commit()
    return _admin_back(request, app_id, f"猫咪信息已更新")


@app.post("/admin/purge", name="admin_purge")
async def admin_purge(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    scope: str = Form("all"),
    confirm: str = Form(""),
):
    return await _admin_purge_run(request, db, csrf_token, scope, confirm)


@app.post("/admin/system/purge", name="admin_purge_system")
async def admin_purge_system(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    scope: str = Form("all"),
    confirm: str = Form(""),
):
    """备用地址：若 /admin/purge 被拦截，可改表单指向此路径。"""
    return await _admin_purge_run(request, db, csrf_token, scope, confirm)


@app.get("/admin/appointments/create", response_class=HTMLResponse)
async def admin_appointment_create_form(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = Query(0),
    pet_id: int = Query(0),
    date: str = Query(""),
    time: str = Query(""),
    category: str = Query(""),
    return_to: str = Query(""),
):
    """新建预约表单页（GET）。从客户档案 / 日历点击都走这里。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    # 该客户的所有宠物（用于下拉切换）
    pets = db.query(Pet).filter(Pet.customer_id == customer_id).all() if customer_id else []
    admin_store = _get_admin_store(request)
    # 默认门店：限店员工锁本店；否则用宠物归属门店；否则空
    default_store_short = admin_store or (pet.store if pet else "") or ""
    default_store_full = _STORE_SHORT_TO_FULL.get(default_store_short, "")
    return templates.TemplateResponse(request, "uk/appointment_create.html", {  # B 补 - UK 重写
        "cust": cust, "pet": pet, "pets": pets,
        "default_date": (date or "")[:10],
        "default_time": (time or "")[:5],
        "default_category": category,
        "default_store_full": default_store_full,
        "store_options": _CLINIC_STORES,   # 全名
        "admin_store_short": admin_store,
        "category_labels": _APPOINTMENT_CATEGORY_LABELS,
        "gender_labels": _PET_GENDER_LABELS,
        "return_to": return_to or "",
        "csrf_token": _get_csrf_token(request),
        "title": "新建预约",
    })


@app.post("/admin/appointments/create", name="admin_appointment_create")
async def admin_appointment_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    category: str = Form(...),
    service_name: str = Form(...),
    customer_name: str = Form(...),
    phone: str = Form(...),
    pet_name: str = Form(...),
    pet_gender: str = Form(...),
    store: str = Form(...),
    appointment_date: str = Form(...),
    appointment_time: str = Form(...),
    duration_minutes: str = Form("30"),
    notes: str = Form(""),
    related_application_id: str = Form(""),
    redirect_after: str = Form("admin"),
    pet_size: str = Form(""),
    coat_length: str = Form(""),
    addon_services: list[str] = Form([]),
    is_proxy: str = Form(""),
    proxy_name: str = Form(""),
    proxy_phone: str = Form(""),
    proxy_relation: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    redirect_base = _admin_appointment_redirect_base(redirect_after)
    try:
        fields = _assert_appointment_fields(
            category=category,
            service_name=service_name,
            customer_name=customer_name,
            phone=phone,
            pet_name=pet_name,
            pet_gender=pet_gender,
            store=store,
            appointment_date=appointment_date,
            appointment_time=appointment_time,
            notes=notes,
            duration_minutes=duration_minutes,
        )
        related_id: int | None = None
        raw_related = (related_application_id or "").strip()
        if str(fields["category"]) == AppointmentCategory.tnr.value and raw_related:
            if not raw_related.isdigit():
                raise HTTPException(400, "关联 TNR 申请编号应为数字。")
            related_id = int(raw_related)
            exists = db.query(Application.id).filter(Application.id == related_id).scalar()
            if not exists:
                raise HTTPException(404, "关联的 TNR 申请不存在。")
        # TNR 规则校验（管理员创建不检查爽约封禁）
        tnr_err = _check_tnr_constraints(
            db,
            category=str(fields["category"]),
            store=str(fields["store"]),
            appointment_date=str(fields["appointment_date"]),
            appointment_time=str(fields["appointment_time"]),
            phone="",
        )
        if tnr_err:
            raise HTTPException(400, tnr_err)
        # 重复申请编号校验
        dup_err = _check_duplicate_application_appointment(db, related_id)
        if dup_err:
            raise HTTPException(400, dup_err)
        # 门诊时间限制校验
        time_err = _check_outpatient_time(str(fields["category"]), str(fields["appointment_time"]))
        if time_err:
            raise HTTPException(400, time_err)
        # 时段容量校验
        cap_err = _check_slot_capacity(
            db,
            store=str(fields["store"]),
            appointment_date=str(fields["appointment_date"]),
            appointment_time=str(fields["appointment_time"]),
            category=str(fields["category"]),
            service_name=str(fields["service_name"]),
        )
        if cap_err:
            raise HTTPException(400, cap_err)
        # 时间冲突检测
        conflict = _check_appointment_conflict(
            db,
            store=str(fields["store"]),
            appointment_date=str(fields["appointment_date"]),
            appointment_time=str(fields["appointment_time"]),
            duration_minutes=int(fields["duration_minutes"]),
            category=str(fields["category"]),
        )
        if conflict:
            raise HTTPException(
                400,
                f"时间冲突：该门店 {conflict.appointment_date} {conflict.appointment_time} 已有预约"
                f"（#{conflict.id} {conflict.customer_name}），请换一个时间段。",
            )
        _is_beauty = str(fields["category"]) == AppointmentCategory.beauty.value
        _is_proxy_bool = bool(is_proxy and is_proxy.strip())
        # ── 自动创建/合并客户档案 ──
        # 优先用表单传入的 customer_id（从客户档案页发起的新建预约会带）
        _admin_appt_cust_id = customer_id if customer_id else None
        if not _admin_appt_cust_id:
            try:
                _admin_appt_cust = _upsert_customer(
                    db,
                    name=str(fields["customer_name"]),
                    phone=str(fields["phone"]),
                    source=str(fields["category"]),
                )
                _admin_appt_cust_id = _admin_appt_cust.id
            except Exception:
                _admin_appt_cust_id = None
        # 校验 pet_id 是否真属于这个客户（防伪造）
        _admin_appt_pet_id = None
        if pet_id and _admin_appt_cust_id:
            _pet_ok = db.query(Pet.id).filter(
                Pet.id == pet_id, Pet.customer_id == _admin_appt_cust_id
            ).first()
            if _pet_ok:
                _admin_appt_pet_id = pet_id
        row = Appointment(
            category=str(fields["category"]),
            # 后台/日历建的：员工录入即视为已确认，不再走"待确认"中间态
            status=AppointmentStatus.confirmed.value,
            service_name=str(fields["service_name"]),
            customer_name=str(fields["customer_name"]),
            phone=str(fields["phone"]),
            pet_name=str(fields["pet_name"]),
            pet_gender=str(fields["pet_gender"]),
            store=str(fields["store"]),
            appointment_date=str(fields["appointment_date"]),
            appointment_time=str(fields["appointment_time"]),
            duration_minutes=int(fields["duration_minutes"]),
            notes=str(fields["notes"]),
            source="admin",
            related_application_id=related_id,
            pet_size=(pet_size.strip() or None) if _is_beauty else None,
            coat_length=(coat_length.strip() or None) if _is_beauty else None,
            addon_services=(",".join(s.strip() for s in addon_services if s.strip()) or None) if _is_beauty else None,
            is_proxy=_is_proxy_bool,
            proxy_name=proxy_name.strip() if _is_proxy_bool else "",
            proxy_phone=proxy_phone.strip() if _is_proxy_bool else "",
            proxy_relation=proxy_relation.strip() if _is_proxy_bool else "",
            customer_id=_admin_appt_cust_id,
            pet_id=_admin_appt_pet_id,
        )
        db.add(row)
        db.flush()
        _audit(
            db,
            request,
            "appointment_create",
            application_id=related_id,
            detail={
                "appointment_id": row.id,
                "category": row.category,
                "service_name": row.service_name,
                "store": row.store,
                "appointment_date": row.appointment_date,
                "appointment_time": row.appointment_time,
            },
        )
        db.commit()
        # /m 路由用 msg/err；admin 路由保持 appointment_ok/err 兼容现有锚点
        if redirect_base.startswith("/m"):
            return RedirectResponse(redirect_base + "?msg=" + quote(f"预约 #{row.id} 已创建", safe=""), status_code=303)
        return RedirectResponse(redirect_base + f"?appointment_ok=create#appt-{row.id}", status_code=303)
    except HTTPException as e:
        db.rollback()
        if redirect_base.startswith("/m"):
            return RedirectResponse(redirect_base + "?err=" + quote(str(e.detail)[:160], safe=""), status_code=303)
        return RedirectResponse(
            redirect_base + "?appointment_err=" + quote(str(e.detail)[:160], safe=""),
            status_code=303,
        )


@app.post("/admin/appointments/{appointment_id}/status", name="admin_appointment_status")
async def admin_appointment_status(
    appointment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    status: str = Form(...),
    redirect_after: str = Form("admin"),
    cancel_reason: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    redirect_base = _admin_appointment_redirect_base(redirect_after)
    status = (status or "").strip()
    anchor = f"appt-{appointment_id}"
    if status not in _ALLOWED_APPOINTMENT_STATUSES:
        if redirect_base.startswith("/m"):
            return RedirectResponse(redirect_base + "?err=" + quote("无效的预约状态", safe=""), status_code=303)
        return RedirectResponse(redirect_base + "?appointment_err=" + quote("无效的预约状态", safe="") + f"#{anchor}", status_code=303)
    row = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not row:
        return RedirectResponse(redirect_base + "?appointment_err=" + quote("预约记录不存在", safe="") + f"#{anchor}", status_code=303)
    # 限店员工只能改本店预约
    admin_store = _get_admin_store(request)
    if admin_store:
        full_store = _STORE_SHORT_TO_FULL.get(admin_store, admin_store)
        if row.store and row.store != full_store:
            return RedirectResponse(redirect_base + "?appointment_err=" + quote("无权操作其他门店的预约", safe="") + f"#{anchor}", status_code=303)
    old_appt_status = row.status
    row.status = status
    row.updated_at = datetime.utcnow()
    reason_clean = (cancel_reason or "").strip()[:300]
    audit_detail: dict = {"appointment_id": row.id, "old_status": old_appt_status, "status": row.status}
    if reason_clean:
        audit_detail["cancel_reason"] = reason_clean
        # 将取消原因追加到备注，方便列表页直接查看
        existing_notes = (row.notes or "").strip()
        row.notes = (existing_notes + f"\n[取消原因] {reason_clean}").strip()
    _audit(
        db,
        request,
        "appointment_status_update",
        application_id=row.related_application_id,
        detail=audit_detail,
    )
    # 同步关联 TNR 申请状态
    if row.related_application_id:
        app_row = db.get(Application, row.related_application_id)
        if app_row:
            if status == AppointmentStatus.confirmed.value and app_row.status in (
                ApplicationStatus.approved.value,
                ApplicationStatus.pre_approved.value,
            ):
                app_row.status = ApplicationStatus.scheduled.value
                app_row.appointment_at = row.appointment_date
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.arrived.value and app_row.status in (
                ApplicationStatus.scheduled.value, ApplicationStatus.approved.value,
            ):
                app_row.status = ApplicationStatus.arrived_verified.value
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.cancelled.value and app_row.status == ApplicationStatus.scheduled.value:
                app_row.status = ApplicationStatus.approved.value
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.no_show.value:
                app_row.status = ApplicationStatus.no_show.value
                app_row.updated_at = datetime.utcnow()
    db.commit()
    # 预约确认/取消后推送通知给用户（#5）
    openid_for_push = (row.wechat_openid or "").strip()
    if openid_for_push and status in (AppointmentStatus.confirmed.value, AppointmentStatus.cancelled.value):
        status_label = "已确认，请按约定时间到院" if status == AppointmentStatus.confirmed.value else "已取消"
        push_appointment_status(
            db,
            appointment_id=row.id,
            openid=openid_for_push,
            status_text=status_label,
            service_name=row.service_name or "",
            store=row.store or "",
            appointment_date=row.appointment_date or "",
            appointment_time=row.appointment_time or "",
            phone=row.phone or "",
            customer_name=row.customer_name or "",
            note=reason_clean or status_label,
        )
    # 尽量回到操作前的筛选页面（保留 preset/df/dt/appt_status 等参数）
    referer = request.headers.get("referer", "")
    # /m 路由用 msg/err；admin 路由保持 appointment_ok 兼容现有锚点
    if redirect_base.startswith("/m"):
        return RedirectResponse(redirect_base + "?msg=" + quote("状态已更新", safe=""), status_code=303)
    if referer and "/admin/appointments" in referer:
        from urllib.parse import urlparse as _urlparse, urlencode as _urlencode, parse_qs as _parse_qs
        _parsed = _urlparse(referer)
        _params = {k: v[0] for k, v in _parse_qs(_parsed.query, keep_blank_values=True).items()
                   if k not in ("appointment_ok", "appointment_err")}
        _qs = _urlencode(_params)
        _base_qs = f"/admin/appointments?{_qs}&appointment_ok=status" if _qs else "/admin/appointments?appointment_ok=status"
        return RedirectResponse(f"{_base_qs}#{anchor}", status_code=303)
    return RedirectResponse(redirect_base + f"?appointment_ok=status#{anchor}", status_code=303)


@app.post("/admin/appointments/{appointment_id}/reschedule", name="admin_appointment_reschedule")
async def admin_appointment_reschedule(
    appointment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    next: str = Form("/admin/appointments"),
    new_store: str = Form(""),
    new_date: str = Form(...),
    new_time: str = Form(...),
    new_duration: str = Form(""),
    reschedule_reason: str = Form(""),
):
    """后台改约：修改预约的门店、日期、时间，自动做冲突检测并记录通知日志。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    _anchor = f"appt-{appointment_id}"
    row = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not row:
        return _admin_appointment_redirect(next, err="预约记录不存在", anchor=_anchor)
    # 限店员工只能改本店预约
    admin_store = _get_admin_store(request)
    if admin_store:
        full_store = _STORE_SHORT_TO_FULL.get(admin_store, admin_store)
        if row.store and row.store != full_store:
            return _admin_appointment_redirect(next, err="无权操作其他门店的预约", anchor=_anchor)
    if row.status in (AppointmentStatus.cancelled.value, AppointmentStatus.completed.value, AppointmentStatus.no_show.value):
        return _admin_appointment_redirect(next, err="当前状态不允许改约", anchor=_anchor)

    new_date = (new_date or "").strip()
    new_time = (new_time or "").strip()
    if not new_date or not new_time:
        return _admin_appointment_redirect(next, err="请填写新的预约日期和时间", anchor=_anchor)
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", new_date):
        return _admin_appointment_redirect(next, err="日期格式应为 YYYY-MM-DD", anchor=_anchor)
    if not re.fullmatch(r"([01]\d|2[0-3]):[0-5]\d", new_time):
        return _admin_appointment_redirect(next, err="时间格式应为 HH:MM", anchor=_anchor)
    try:
        date_obj = datetime.strptime(new_date, "%Y-%m-%d").date()
    except ValueError:
        return _admin_appointment_redirect(next, err="无效的日期", anchor=_anchor)
    if date_obj < datetime.now().date():
        return _admin_appointment_redirect(next, err="改约日期不能早于今天", anchor=_anchor)

    target_store = (new_store or "").strip() or row.store
    if target_store not in _ALLOWED_CLINIC_STORES:
        return _admin_appointment_redirect(next, err="无效的门店", anchor=_anchor)
    dur_raw = (new_duration or "").strip()
    target_duration = int(dur_raw) if dur_raw and dur_raw.isdigit() and 10 <= int(dur_raw) <= 480 else row.duration_minutes

    # TNR 规则校验（管理员改约不检查爽约封禁）
    tnr_err = _check_tnr_constraints(
        db,
        category=row.category,
        store=target_store,
        appointment_date=new_date,
        appointment_time=new_time,
        phone="",
        exclude_id=appointment_id,
    )
    if tnr_err:
        return _admin_appointment_redirect(next, err=tnr_err, anchor=_anchor)
    # 门诊时间限制校验（改约）
    time_err = _check_outpatient_time(row.category, new_time)
    if time_err:
        return _admin_appointment_redirect(next, err=time_err, anchor=_anchor)
    # 时段容量校验（改约）
    cap_err = _check_slot_capacity(
        db,
        store=target_store,
        appointment_date=new_date,
        appointment_time=new_time,
        category=row.category,
        service_name=row.service_name or "",
        exclude_id=appointment_id,
    )
    if cap_err:
        return _admin_appointment_redirect(next, err=cap_err, anchor=_anchor)
    # 冲突检测（排除自身）
    conflict = _check_appointment_conflict(
        db,
        store=target_store,
        appointment_date=new_date,
        appointment_time=new_time,
        duration_minutes=target_duration,
        exclude_id=appointment_id,
        category=row.category,
    )
    if conflict:
        return _admin_appointment_redirect(
            next,
            err=f"时间冲突：该门店 {conflict.appointment_date} {conflict.appointment_time} 已有预约"
            f"（#{conflict.id} {conflict.customer_name}），请换一个时间段。",
            anchor=_anchor,
        )

    old_store = row.store
    old_date = row.appointment_date
    old_time = row.appointment_time
    old_duration = row.duration_minutes

    row.store = target_store
    row.appointment_date = new_date
    row.appointment_time = new_time
    row.duration_minutes = target_duration
    row.updated_at = datetime.utcnow()

    reason_text = (reschedule_reason or "").strip()[:500]
    detail = {
        "appointment_id": row.id,
        "old": {"store": old_store, "date": old_date, "time": old_time, "duration": old_duration},
        "new": {"store": row.store, "date": row.appointment_date, "time": row.appointment_time, "duration": row.duration_minutes},
        "reason": reason_text,
    }
    _audit(db, request, "appointment_reschedule", application_id=row.related_application_id, detail=detail)

    # 同步关联申请的预约日期
    if row.related_application_id:
        app_row = db.get(Application, row.related_application_id)
        if app_row and app_row.status == ApplicationStatus.scheduled.value:
            app_row.appointment_at = new_date
            app_row.updated_at = datetime.utcnow()
        notify_payload = (
            f"预约改约通知：#{row.id} {row.customer_name} 由 {old_store} {old_date} {old_time} "
            f"改为 {row.store} {row.appointment_date} {row.appointment_time}"
        )
        if reason_text:
            notify_payload += f"（原因：{reason_text}）"
        from app.models import NotificationLog
        db.add(NotificationLog(
            application_id=row.related_application_id,
            channel="log",
            payload=notify_payload,
            success=True,
        ))

    db.commit()
    return _admin_appointment_redirect(next, ok="reschedule", anchor=_anchor)


@app.post("/admin/app/{app_id}/manual-approve")
async def manual_approve(app_id: int, request: Request, db: Session = Depends(get_db), csrf_token: str = Form(""), next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.pending_ai.value,
            ApplicationStatus.pending_manual.value,
            ApplicationStatus.pre_approved.value,
        },
        "人工通过",
    )
    row.status = ApplicationStatus.approved.value
    _audit(db, request, "manual_approve", application_id=app_id)
    db.commit()
    notify_application_result(db, app_id, row.phone, row.applicant_name, approved=True)
    push_application_result(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        applicant_name=row.applicant_name,
        status_text="审核已通过",
        phone_masked=row.phone,
        note="请按约定时间携带猫咪到院",
        submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    if next_url:
        return RedirectResponse(_safe_next(next_url, "/admin"), status_code=303)
    return _admin_back(request, app_id)


@app.post("/admin/appointments/{appointment_id}/delete")
async def admin_appointment_delete(
    appointment_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """删除预约：
       - 仅 source=admin (员工后台建的) 且 status=cancelled 可删
       - source=miniapp (小程序自助) 必须保留作记录，绝不能删
    """
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Appointment, appointment_id)
    if not row:
        raise HTTPException(404)
    if row.status != AppointmentStatus.cancelled.value:
        raise HTTPException(400, "只能删除已取消的预约。请先取消。")
    src = (row.source or "").strip().lower()
    # admin / empty / 任何非 miniapp 的来源都允许删；miniapp / wechat 必须留档
    if src in ("miniapp", "wechat"):
        raise HTTPException(403, "小程序自助预约必须留档，不可删除。")
    _audit(db, request, "delete_appointment", application_id=None,
           detail={"appointment_id": appointment_id, "phone": row.phone, "source": src})
    db.delete(row)
    db.commit()
    referer = request.headers.get("referer", "")
    if referer and "/admin/" in referer:
        return RedirectResponse(referer, status_code=303)
    return RedirectResponse("/admin/appointments?msg=预约已删除", status_code=303)


@app.post("/admin/app/{app_id}/delete-draft")
async def admin_delete_draft(
    app_id: int,
    request: Request,
    csrf_token: str = Form(""),
    db: Session = Depends(get_db),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    if row.status != ApplicationStatus.draft.value:
        raise HTTPException(400, "只能删除草稿状态的申请。")
    _audit(db, request, "delete_draft", application_id=app_id, detail={"phone": row.phone})
    db.delete(row)
    db.commit()
    return RedirectResponse("/admin?msg=草稿已删除，客户现可重新提交申请。", status_code=303)


@app.post("/admin/app/{app_id}/reject")
async def manual_reject(
    app_id: int,
    request: Request,
    reason: str = Form(""),
    csrf_token: str = Form(""),
    db: Session = Depends(get_db),
    next_url: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.pending_ai.value,
            ApplicationStatus.pending_manual.value,
            ApplicationStatus.pre_approved.value,
            ApplicationStatus.approved.value,
            ApplicationStatus.scheduled.value,
        },
        "拒绝",
    )
    row.status = ApplicationStatus.rejected.value
    row.reject_reason = reason.strip()
    _audit(db, request, "manual_reject", application_id=app_id, detail={"reason": row.reject_reason})
    db.commit()
    notify_application_result(db, app_id, row.phone, row.applicant_name, approved=False, extra=reason)
    push_rejection_notice(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        cat_nickname=row.cat_nickname or "",
        reason=(reason or "不符合申请条件")[:20],
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    if next_url:
        return RedirectResponse(_safe_next(next_url, "/admin"), status_code=303)
    return _admin_back(request, app_id)


@app.post("/admin/app/{app_id}/verify-cat")
async def verify_cat(app_id: int, request: Request, db: Session = Depends(get_db), csrf_token: str = Form(""), _NOTE_="""
确认「到的这只猫 = 申请单上那只猫」。仅设 staff_cat_verified=True，
不再连带改 status。原因：客户到店（status=arrived_verified）和核实猫身份是
两个独立动作 —— 可能到了店但发现不是同一只猫，此时不应自动设确认。
"""):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.approved.value,
            ApplicationStatus.scheduled.value,
            ApplicationStatus.arrived_verified.value,  # 已到店但还没确认是同一只猫
        },
        "现场确认",
    )
    # 只设核实标记 — 「到店」和「核实是同一只猫」是两个独立动作
    row.staff_cat_verified = True
    _audit(db, request, "verify_cat", application_id=app_id)
    db.commit()
    return _admin_back(request, app_id)


@app.post("/admin/app/{app_id}/surgery-done")
async def surgery_done(app_id: int, request: Request, db: Session = Depends(get_db), csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.approved.value,
            ApplicationStatus.arrived_verified.value,
            ApplicationStatus.scheduled.value,
        },
        "标记手术完成",
    )
    if not _application_has_surgery_before_and_after(db, app_id):
        return RedirectResponse(
            "/admin?surgery_media_err="
            + quote(
                "标记手术完成前，须在本申请下各上传至少 1 条术前资料与 1 条术后资料（照片或视频均可）。",
                safe="",
            )
            + f"#app-{app_id}",
            status_code=303,
        )
    row.status = ApplicationStatus.surgery_completed.value
    # 业务约束：手术完成天然蕴含「已现场确认」（不可能没核对就做手术）
    # 员工跳步骤直接点手术完成的，这里自动补上
    row.staff_cat_verified = True
    # 默认公开公布：标记手术完成时自动开启公益展示
    # 如果客户/医院不想公开，员工可在卡片上点「拒绝公开」手动关掉
    row.showcase_consent = True
    _audit(db, request, "surgery_done", application_id=app_id)
    db.commit()
    notify_application_result(
        db,
        app_id,
        row.phone,
        row.applicant_name,
        approved=True,
        extra="手术已完成。请遵医嘱护理；公猫放归时间请听从医嘱。若同意公开展示，您可在本院 TNR 展示页查看术前术后资料（脱敏处理）。",
    )
    push_surgery_done(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        cat_name=row.cat_nickname or "猫咪",
        note="手术已完成，请按医嘱护理",
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return _admin_back(request, app_id)


@app.post("/api/wechat/login")
async def api_wechat_login(payload: dict = Body(...)):
    """小程序端：传 {code: js_code}，换取 openid。"""
    code = (payload or {}).get("code", "")
    if not code:
        raise HTTPException(400, "missing code")
    try:
        data = wechat_code2session(code)
    except Exception as e:
        raise HTTPException(400, str(e))
    # 生产环境不建议把 session_key 返回给前端；这里只返回 openid 供演示
    return {"openid": data.get("openid", "")}


@app.get("/api/wechat/my-tnr-status")
async def api_my_tnr_status(openid: str = Query(""), db: Session = Depends(get_db)):
    """小程序端：传 openid，返回该用户是否有可预约的已通过 TNR 申请。"""
    if not openid.strip():
        return {"has_approved": False, "approved_apps": []}
    _APPROVED_STATUSES = {
        ApplicationStatus.approved.value,
        ApplicationStatus.pre_approved.value,
        ApplicationStatus.scheduled.value,
    }
    rows = (
        db.query(Application)
        .filter(
            Application.wechat_openid == openid.strip(),
            Application.status.in_(_APPROVED_STATUSES),
        )
        .order_by(Application.created_at.desc())
        .limit(10)
        .all()
    )
    apps = [{"id": r.id, "status": r.status, "cat_nickname": r.cat_nickname} for r in rows]
    return {"has_approved": len(apps) > 0, "approved_apps": apps}


@app.get("/api/appointments/config")
async def api_appointments_config():
    return _appointment_catalog()


@app.get("/api/tnr-store-status")
async def api_tnr_store_status(phone: str = Query(""), db: Session = Depends(get_db)):
    """
    小程序调用：获取各门店 TNR 预约开放状态及当月已用配额。
    可选传 phone 参数以检查爽约封禁。
    """
    today = datetime.now()
    year_month = today.strftime("%Y-%m")
    stores_info = []
    for store in _CLINIC_STORES:
        cfg = _get_tnr_store_config(db, store)
        monthly_count = _get_tnr_monthly_confirmed_count(db, store, year_month)
        quota = cfg.tnr_monthly_quota
        is_open = cfg.tnr_accepting and (monthly_count < quota)
        stores_info.append({
            "store": store,
            "accepting": cfg.tnr_accepting,
            "monthly_count": monthly_count,
            "monthly_quota": quota,
            "is_open": is_open,
        })
    # 爽约封禁检查
    ban_until_str = None
    is_banned = False
    if phone:
        ban_until = _get_phone_noshow_ban_until(db, phone)
        if ban_until is not None and ban_until >= today.date():
            is_banned = True
            ban_until_str = ban_until.strftime("%Y-%m-%d")
    return {
        "stores": stores_info,
        "is_banned": is_banned,
        "ban_until": ban_until_str,
    }


@app.post("/admin/tnr-quota/{store_index}/toggle")
async def admin_tnr_quota_toggle(
    request: Request,
    store_index: int,
    db: Session = Depends(get_db),
):
    """管理员手动开关门店 TNR 预约接受状态。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login", status_code=302)
    stores = list(_CLINIC_STORES)
    if store_index < 0 or store_index >= len(stores):
        raise HTTPException(400, "无效的门店编号")
    store_name = stores[store_index]
    cfg = _get_tnr_store_config(db, store_name)
    cfg.tnr_accepting = not cfg.tnr_accepting
    cfg.updated_by = request.session.get("admin_user", "admin")
    cfg.updated_at = datetime.utcnow()
    db.commit()
    next_url = request.headers.get("referer") or "/admin/appointments"
    return RedirectResponse(next_url, status_code=302)


@app.get("/api/appointments/beauty-slots")
async def api_beauty_slots(
    service:     str = Query(""),
    pet_size:    str = Query(""),
    coat_length: str = Query(""),
    addons:      str = Query(""),   # 逗号分隔的附加项目
    date:        str = Query(""),   # YYYY-MM-DD
    store:       str = Query(""),
    db: Session = Depends(get_db),
):
    """
    美容预约：计算估算时长 + 可预约时段。
    时段生成考虑猫进烘干机可并发小型犬洗护的规则。
    """
    duration = _calc_beauty_duration(service, pet_size, coat_length, addons)

    # 时长显示字符串
    h, m = divmod(duration, 60)
    if h == 0:
        dur_display = f"{m}分钟"
    elif m:
        dur_display = f"{h}小时{m}分钟"
    else:
        dur_display = f"{h}小时"

    slots: list[str] = []
    if date and store:
        try:
            datetime.strptime(date, "%Y-%m-%d")
            slots = _beauty_slots_for_date(db, store, date, service, duration)
        except Exception:
            slots = []

    return {
        "duration_minutes": duration,
        "duration_display": dur_display,
        "available_slots":  slots,
        "disclaimer": (
            "以上占用时间为估算时间。"
            "如动物不配合或毛量超出预估，实际服务时长可能有所浮动，"
            "具体以门店实际执行时间为准。"
        ),
    }


@app.post("/api/appointments/create")
async def api_appointments_create(payload: dict = Body(...), db: Session = Depends(get_db)):
    openid = await _resolve_wechat_openid(payload)
    fields = _assert_appointment_fields(
        category=(payload or {}).get("category", ""),
        service_name=(payload or {}).get("service_name", ""),
        customer_name=(payload or {}).get("customer_name", ""),
        phone=(payload or {}).get("phone", ""),
        pet_name=(payload or {}).get("pet_name", ""),
        pet_gender=(payload or {}).get("pet_gender", ""),
        store=(payload or {}).get("store", ""),
        appointment_date=(payload or {}).get("appointment_date", ""),
        appointment_time=(payload or {}).get("appointment_time", ""),
        notes=(payload or {}).get("notes", ""),
        duration_minutes=str((payload or {}).get("duration_minutes", "30")),
    )
    related_id: int | None = None
    raw_related = (payload or {}).get("related_application_id")
    if raw_related not in (None, ""):
        try:
            related_id = int(raw_related)
        except (TypeError, ValueError):
            raise HTTPException(400, "关联申请编号格式不正确。")
        exists = db.query(Application.id).filter(Application.id == related_id).scalar()
        if not exists:
            raise HTTPException(404, "关联的 TNR 申请不存在。")
    # TNR 规则校验（含爽约封禁检查）
    tnr_err = _check_tnr_constraints(
        db,
        category=str(fields["category"]),
        store=str(fields["store"]),
        appointment_date=str(fields["appointment_date"]),
        appointment_time=str(fields["appointment_time"]),
        phone=str(fields.get("phone", "")),
    )
    if tnr_err:
        raise HTTPException(400, tnr_err)
    # 重复申请编号校验
    dup_err = _check_duplicate_application_appointment(db, related_id)
    if dup_err:
        raise HTTPException(400, dup_err)
    # 时段容量校验
    # 门诊时间限制校验
    time_err = _check_outpatient_time(str(fields["category"]), str(fields["appointment_time"]))
    if time_err:
        raise HTTPException(400, time_err)
    cap_err = _check_slot_capacity(
        db,
        store=str(fields["store"]),
        appointment_date=str(fields["appointment_date"]),
        appointment_time=str(fields["appointment_time"]),
        category=str(fields["category"]),
        service_name=str(fields["service_name"]),
    )
    if cap_err:
        raise HTTPException(400, cap_err)
    # 时间冲突检测
    conflict = _check_appointment_conflict(
        db,
        store=str(fields["store"]),
        appointment_date=str(fields["appointment_date"]),
        appointment_time=str(fields["appointment_time"]),
        duration_minutes=int(fields["duration_minutes"]),
        category=str(fields["category"]),
    )
    if conflict:
        raise HTTPException(
            400,
            f"时间冲突：该门店 {conflict.appointment_date} {conflict.appointment_time} 已有预约"
            f"（#{conflict.id} {conflict.customer_name}），请换一个时间段。",
        )
    _is_beauty_api = str(fields["category"]) == AppointmentCategory.beauty.value
    _pet_size_raw    = ((payload or {}).get("pet_size", "") or "").strip()
    _coat_length_raw = ((payload or {}).get("coat_length", "") or "").strip()
    _addon_raw       = ((payload or {}).get("addon_services", "") or "").strip()
    _is_proxy_api    = bool((payload or {}).get("is_proxy", False))
    _proxy_name_raw  = ((payload or {}).get("proxy_name", "") or "").strip()
    _proxy_phone_raw = ((payload or {}).get("proxy_phone", "") or "").strip()
    _proxy_rel_raw   = ((payload or {}).get("proxy_relation", "") or "").strip()
    # ── 自动创建/合并客户档案 ──
    _appt_cust_id = None
    try:
        _appt_cust = _upsert_customer(
            db,
            name=str(fields["customer_name"]),
            phone=str(fields["phone"]),
            openid=openid,
            source=str(fields["category"]),
        )
        _appt_cust_id = _appt_cust.id
    except Exception:
        _appt_cust_id = None
    # ── 解析前端传的 pet_id（老顾客在预约页选了已有宠物 chip）──
    _appt_pet_id: int | None = None
    _raw_pet_id = (payload or {}).get("pet_id")
    if _raw_pet_id not in (None, "", 0, "0"):
        try:
            _maybe_pet_id = int(_raw_pet_id)
            _maybe_pet = db.get(Pet, _maybe_pet_id) if _maybe_pet_id > 0 else None
            # 安全：宠物必须属于该客户，不允许跨客户绑
            if _maybe_pet and _appt_cust_id and _maybe_pet.customer_id == _appt_cust_id:
                _appt_pet_id = _maybe_pet_id
        except (TypeError, ValueError):
            pass
    row = Appointment(
        wechat_openid=openid,
        category=str(fields["category"]),
        status=AppointmentStatus.pending.value,
        service_name=str(fields["service_name"]),
        customer_name=str(fields["customer_name"]),
        phone=str(fields["phone"]),
        pet_name=str(fields["pet_name"]),
        pet_gender=str(fields["pet_gender"]),
        store=str(fields["store"]),
        appointment_date=str(fields["appointment_date"]),
        appointment_time=str(fields["appointment_time"]),
        duration_minutes=int(fields["duration_minutes"]),
        source="miniapp",
        notes=str(fields["notes"]),
        related_application_id=related_id,
        pet_size    =(_pet_size_raw    or None) if _is_beauty_api else None,
        coat_length =(_coat_length_raw or None) if _is_beauty_api else None,
        addon_services=(_addon_raw     or None) if _is_beauty_api else None,
        is_proxy=_is_proxy_api,
        proxy_name=_proxy_name_raw if _is_proxy_api else "",
        proxy_phone=_proxy_phone_raw if _is_proxy_api else "",
        proxy_relation=_proxy_rel_raw if _is_proxy_api else "",
        customer_id=_appt_cust_id,
        pet_id=_appt_pet_id,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    # 推送给对应门店员工
    try:
        from app.services import wecom_notify as _wn
        _wn.notify_appointment_created(db, row)
    except Exception as _e:
        logger.warning("[wecom] notify_appointment_created failed: %s", _e)
    return {"ok": True, "appointment": _serialize_appointment(row)}


@app.post("/api/wechat/my-apps")
async def api_wechat_my_apps(payload: dict = Body(...), db: Session = Depends(get_db)):
    """小程序端：传 {code} 或 {openid}，返回该 openid 的申请列表（用于“我的订单”）。"""
    code = (payload or {}).get("code", "") or ""
    openid = (payload or {}).get("openid", "") or ""
    if code:
        try:
            data = wechat_code2session(code)
        except Exception as e:
            raise HTTPException(400, str(e))
        openid = data.get("openid", "") or ""
    openid = (openid or "").strip()
    if not openid:
        raise HTTPException(400, "missing openid")

    def mask_phone(s: str) -> str:
        t = (s or "").strip()
        if len(t) < 7:
            return t
        return t[:3] + "****" + t[-4:]

    rows = (
        db.query(Application)
        .filter(Application.wechat_openid == openid)
        .order_by(Application.created_at.desc())
        .limit(50)
        .all()
    )
    items = []
    for row in rows:
        notes = (row.reject_reason or "").strip()
        if len(notes) > 80:
            notes = notes[:80] + "…"
        hn = (row.health_note or "").strip()
        if len(hn) > 60:
            hn = hn[:60] + "…"
        items.append(
            {
                "id": row.id,
                "status": row.status,
                "clinic_store": row.clinic_store,
                "appointment_at": row.appointment_at,
                "phone_masked": mask_phone(row.phone),
                "note": notes,
                "cat_nickname": row.cat_nickname or "",
                "cat_gender": row.cat_gender or "",
                "age_estimate": row.age_estimate or "",
                "health_note_brief": hn,
                "created_at": row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
                "updated_at": row.updated_at.strftime("%Y-%m-%d %H:%M") if row.updated_at else "",
            }
        )
    return {"openid": openid, "items": items}


@app.post("/api/wechat/my-appointments")
async def api_wechat_my_appointments(payload: dict = Body(...), db: Session = Depends(get_db)):
    openid = await _resolve_wechat_openid(payload)
    rows = (
        db.query(Appointment)
        .filter(Appointment.wechat_openid == openid)
        .order_by(Appointment.created_at.desc())
        .limit(50)
        .all()
    )
    return {"openid": openid, "items": [_serialize_appointment(x) for x in rows]}


@app.post("/api/wechat/appointments/{appointment_id}/cancel")
async def api_wechat_appointment_cancel(
    appointment_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    """小程序端用户取消自己的预约（仅限 pending/confirmed 状态，且 openid 匹配）。"""
    openid = await _resolve_wechat_openid(payload)
    row = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not row:
        raise HTTPException(404, "预约记录不存在")
    if (row.wechat_openid or "") != openid:
        raise HTTPException(403, "无权操作此预约")
    if row.status not in (AppointmentStatus.pending.value, AppointmentStatus.confirmed.value):
        raise HTTPException(400, f"当前状态（{row.status}）不可取消，请联系医院处理")
    row.status = AppointmentStatus.cancelled.value
    row.updated_at = datetime.utcnow()
    # 若关联 TNR 申请且已处于 scheduled 状态，回退为 approved
    if row.related_application_id:
        app_row = db.get(Application, row.related_application_id)
        if app_row and app_row.status == ApplicationStatus.scheduled.value:
            app_row.status = ApplicationStatus.approved.value
            app_row.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "id": appointment_id}


@app.post("/api/wechat/appointments/{appointment_id}/get")
async def api_wechat_appointment_get(
    appointment_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    """小程序端获取单条预约详情（openid 校验）。"""
    openid = await _resolve_wechat_openid(payload)
    row = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not row:
        raise HTTPException(404, "预约记录不存在")
    if (row.wechat_openid or "") != openid:
        raise HTTPException(403, "无权查看此预约")
    return _serialize_appointment(row)


@app.post("/api/wechat/appointments/{appointment_id}/update")
async def api_wechat_appointment_update(
    appointment_id: int,
    payload: dict = Body(...),
    db: Session = Depends(get_db),
):
    """小程序端修改自己的预约（仅限 pending 状态，openid 校验，时段容量重新检验）。"""
    openid = await _resolve_wechat_openid(payload)
    row = db.query(Appointment).filter(Appointment.id == appointment_id).first()
    if not row:
        raise HTTPException(404, "预约记录不存在")
    if (row.wechat_openid or "") != openid:
        raise HTTPException(403, "无权操作此预约")
    if row.status != AppointmentStatus.pending.value:
        raise HTTPException(400, "只有「待确认」状态的预约才能修改，如需调整请联系医院")

    new_date  = str(payload.get("appointment_date", "") or row.appointment_date).strip()
    new_time  = str(payload.get("appointment_time", "") or row.appointment_time).strip()
    new_store = str(payload.get("store", "") or row.store or "").strip()

    # 若日期、时间或门店有变，重新做容量检查（exclude 自身，避免自我碰撞）
    if new_date != row.appointment_date or new_time != row.appointment_time or new_store != (row.store or ""):
        err = _check_slot_capacity(
            db, new_store, new_date, new_time,
            row.category, row.service_name or "",
            exclude_id=appointment_id,
        )
        if err:
            raise HTTPException(400, err)

    row.appointment_date = new_date
    row.appointment_time = new_time
    if new_store:
        row.store = new_store
    if "customer_name" in payload and payload["customer_name"]:
        row.customer_name = str(payload["customer_name"])[:80]
    if "phone" in payload and payload["phone"]:
        row.phone = str(payload["phone"])[:20]
    if "pet_name" in payload and payload["pet_name"]:
        row.pet_name = str(payload["pet_name"])[:80]
    if "pet_gender" in payload and payload["pet_gender"]:
        row.pet_gender = str(payload["pet_gender"])[:20]
    if "notes" in payload:
        row.notes = str(payload.get("notes") or "")[:500]

    row.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(row)
    return {"ok": True, "appointment": _serialize_appointment(row)}


@app.post("/api/wechat/feedback/create")
async def api_wechat_feedback_create(payload: dict = Body(...), db: Session = Depends(get_db)):
    from app.models import Feedback
    openid = str((payload or {}).get("openid", "") or "").strip()
    content = str((payload or {}).get("content", "") or "").strip()
    if not content:
        raise HTTPException(400, "请填写反馈内容")
    fb = Feedback(openid=openid, content=content[:2000])
    db.add(fb)
    db.commit()
    db.refresh(fb)
    # create upload dir
    fb_dir = Path(settings.upload_dir) / "feedback" / str(fb.id)
    fb_dir.mkdir(parents=True, exist_ok=True)
    return {"id": fb.id, "ok": True}


@app.post("/api/wechat/feedback/{feedback_id}/upload")
async def api_wechat_feedback_upload(
    feedback_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    from app.models import Feedback
    import json as _json
    fb = db.get(Feedback, feedback_id)
    if not fb:
        raise HTTPException(404, "反馈记录不存在")
    ext = Path(file.filename or "img.jpg").suffix.lower() or ".jpg"
    safe_ext = ext if ext in (".jpg", ".jpeg", ".png", ".gif", ".webp") else ".jpg"
    fb_dir = Path(settings.upload_dir) / "feedback" / str(feedback_id)
    fb_dir.mkdir(parents=True, exist_ok=True)
    existing = _json.loads(fb.image_paths or "[]")
    if len(existing) >= 6:
        raise HTTPException(400, "最多上传6张截图")
    fname = f"{len(existing)+1}{safe_ext}"
    dest = fb_dir / fname
    content_bytes = await file.read()
    dest.write_bytes(content_bytes)
    stored = str(dest)
    existing.append(stored)
    fb.image_paths = _json.dumps(existing, ensure_ascii=False)
    db.commit()
    return {"ok": True, "path": stored}


@app.get("/admin/feedback", response_class=HTMLResponse)
async def admin_feedback_page(request: Request, status: str = "", db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login", status_code=303)
    from app.models import Feedback
    q = db.query(Feedback).order_by(Feedback.created_at.desc())
    if status == "pending":
        q = q.filter(Feedback.status == "pending")
    elif status == "resolved":
        q = q.filter(Feedback.status == "resolved")
    items = q.limit(100).all()
    # Build image URLs for each feedback
    import json as _json
    feed_list = []
    for fb in items:
        paths = _json.loads(fb.image_paths or "[]")
        img_urls = []
        for p in paths:
            # Convert stored path to URL
            try:
                rel = Path(p).relative_to(Path(settings.upload_dir))
                img_urls.append("/uploads/" + str(rel).replace("\\", "/"))
            except Exception:
                pass
        feed_list.append({"fb": fb, "img_urls": img_urls})
    pending_count = db.query(Feedback).filter(Feedback.status == "pending").count()
    return templates.TemplateResponse(request, "uk/feedback.html", {
        "request": request,
        "title": "客户反馈",
        "feed_list": feed_list,
        "pending_count": pending_count,
        "status_filter": status,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/feedback/{feedback_id}/resolve")
async def admin_feedback_resolve(
    feedback_id: int,
    request: Request,
    admin_note: str = Form(""),
    db: Session = Depends(get_db),
):
    if not request.session.get("admin"):
        raise HTTPException(403)
    from app.models import Feedback
    fb = db.get(Feedback, feedback_id)
    if not fb:
        raise HTTPException(404)
    fb.status = "resolved"
    fb.admin_note = admin_note.strip()[:500]
    fb.resolved_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/admin/feedback?msg=已处理#{feedback_id}", status_code=303)


@app.post("/api/wechat/claim-apps")
async def api_wechat_claim_apps(payload: dict = Body(...), db: Session = Depends(get_db)):
    """
    小程序端：用 {openid}+{phone}+{id_number} 找回历史订单。
    仅对 phone+id_number 匹配且 wechat_openid 为空的记录进行补写，不覆盖已有 openid 的记录。
    """
    code = (payload or {}).get("code", "") or ""
    openid = (payload or {}).get("openid", "") or ""
    phone = (payload or {}).get("phone", "") or ""
    id_number = (payload or {}).get("id_number", "") or ""

    if code and not openid:
        try:
            data = wechat_code2session(code)
        except Exception as e:
            raise HTTPException(400, str(e))
        openid = data.get("openid", "") or ""

    openid = (openid or "").strip()
    phone = (phone or "").strip()
    idn = (id_number or "").strip().upper()

    if not openid:
        raise HTTPException(400, "missing openid")
    if not re.fullmatch(r"1\d{10}", phone):
        raise HTTPException(400, "手机号格式不正确")
    if len(idn) == 18:
        if not re.fullmatch(r"\d{17}[\dX]", idn):
            raise HTTPException(400, "身份证号格式不正确")
    elif len(idn) == 15:
        if not re.fullmatch(r"\d{15}", idn):
            raise HTTPException(400, "身份证号格式不正确")
    else:
        raise HTTPException(400, "身份证号格式不正确")

    rows = (
        db.query(Application)
        .filter(Application.phone == phone)
        .filter(Application.id_number == idn)
        .filter(or_(Application.wechat_openid == None, Application.wechat_openid == ""))
        .order_by(Application.created_at.desc())
        .limit(200)
        .all()
    )
    n = 0
    for row in rows:
        row.wechat_openid = openid
        n += 1
    if n:
        db.commit()
    return {"ok": True, "updated": n}


def _video_ext(name: str) -> str:
    ext = Path(name).suffix.lower()
    return ext if ext in (".mp4", ".webm", ".mov", ".mkv") else ".mp4"


def _image_ext(name: str) -> str:
    ext = Path(name).suffix.lower()
    return ext if ext in (".jpg", ".jpeg", ".png", ".webp") else ".jpg"


def _compress_image(src: Path, max_px: int = 1920, quality: int = 85) -> Path:
    try:
        from PIL import Image, ExifTags
        img = Image.open(src)
        # 按 EXIF 自动旋转
        try:
            exif = img._getexif()
            if exif:
                for tag, val in exif.items():
                    if ExifTags.TAGS.get(tag) == "Orientation":
                        if val == 3:
                            img = img.rotate(180, expand=True)
                        elif val == 6:
                            img = img.rotate(270, expand=True)
                        elif val == 8:
                            img = img.rotate(90, expand=True)
                        break
        except Exception:
            pass
        if max(img.width, img.height) > max_px:
            img.thumbnail((max_px, max_px), Image.LANCZOS)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        dest = src.with_suffix(".jpg")
        img.save(dest, "JPEG", quality=quality, optimize=True)
        if src.suffix.lower() not in (".jpg", ".jpeg"):
            src.unlink(missing_ok=True)
        return dest
    except Exception as e:
        logging.warning(f"图片压缩失败，保留原文件：{e}")
        return src


def _transcode_to_h264(src: Path) -> Path:
    """
    用 ffmpeg 将视频转码为 H.264 MP4（最广兼容格式）。
    转码成功返回新路径（.mp4），失败则返回原路径（不影响上传流程）。
    """
    dest = src.with_suffix(".mp4")
    try:
        result = subprocess.run(
            [
                "ffmpeg", "-y",
                "-i", str(src),
                "-c:v", "libx264",
                "-preset", "fast",
                "-crf", "23",
                "-c:a", "aac",
                "-movflags", "+faststart",  # 元数据前置，支持边下边播
                "-vf", "scale=trunc(iw/2)*2:trunc(ih/2)*2",  # 确保宽高为偶数
                str(dest),
            ],
            timeout=300,
            capture_output=True,
        )
        if result.returncode == 0 and dest.exists():
            if src.suffix.lower() != ".mp4":
                src.unlink(missing_ok=True)  # 删除原始文件节省空间
            return dest
        else:
            logging.warning(f"ffmpeg 转码失败：{result.stderr.decode(errors='replace')[:300]}")
            return src
    except FileNotFoundError:
        logging.warning("ffmpeg 未安装，跳过转码。建议服务器执行：apt install ffmpeg")
        return src
    except Exception as e:
        logging.warning(f"ffmpeg 转码异常：{e}")
        return src


@app.post("/admin/app/{app_id}/upload-surgery")
async def upload_surgery(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    before_images: list[UploadFile] | None = File(None),
    after_images: list[UploadFile] | None = File(None),
    before_videos: list[UploadFile] | None = File(None),
    after_videos: list[UploadFile] | None = File(None),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    base = Path(settings.upload_dir) / str(app_id)
    base.mkdir(parents=True, exist_ok=True)

    async def save_batch(files: list[UploadFile] | None, file_prefix: str, kind: str, *, is_video: bool = False):
        if not files:
            return
        for uf in files:
            if not uf.filename:
                continue
            ext = _video_ext(uf.filename) if is_video else _image_ext(uf.filename)
            dest = base / f"{file_prefix}_{secrets.token_hex(6)}{ext}"
            dest.write_bytes(await uf.read())
            if is_video:
                dest = _transcode_to_h264(dest)
            else:
                dest = _compress_image(dest)  # 压缩图片至 1920px/JPEG85，加快加载
            db.add(
                MediaFile(
                    application_id=app_id,
                    kind=kind,
                    stored_path=str(dest),
                    original_name=uf.filename,
                )
            )

    before_images_n = len(before_images or [])
    after_images_n = len(after_images or [])
    before_videos_n = len(before_videos or [])
    after_videos_n = len(after_videos or [])

    await save_batch(before_images, "surg_bi", MediaKind.surgery_before.value)
    await save_batch(after_images, "surg_ai", MediaKind.surgery_after.value)
    await save_batch(before_videos, "surg_bv", MediaKind.surgery_before.value, is_video=True)
    await save_batch(after_videos, "surg_av", MediaKind.surgery_after.value, is_video=True)
    _audit(
        db,
        request,
        "upload_surgery",
        application_id=app_id,
        detail={
            "before_images": before_images_n,
            "after_images": after_images_n,
            "before_videos": before_videos_n,
            "after_videos": after_videos_n,
        },
    )
    db.commit()
    return _admin_back(request, app_id)


def _media_public_ok(m: MediaFile, app_row: Application) -> bool:
    if app_row.status != ApplicationStatus.surgery_completed.value:
        return False
    # 默认公开：仅当 showcase_consent 明确为 False（管理员手动关闭）才屏蔽
    if app_row.showcase_consent is False:
        return False
    return m.kind in (MediaKind.surgery_before.value, MediaKind.surgery_after.value)


@app.get("/file/{media_id}")
async def serve_file(media_id: int, request: Request, db: Session = Depends(get_db)):
    m = db.get(MediaFile, media_id)
    if not m:
        raise HTTPException(404)
    app_row = db.get(Application, m.application_id)
    if not app_row:
        raise HTTPException(404)
    path = Path(m.stored_path).resolve()
    root = Path(settings.upload_dir).resolve()
    if not str(path).startswith(str(root)) or not path.is_file():
        raise HTTPException(404)

    admin = _admin_ok(request)
    if m.kind in (MediaKind.application_image.value, MediaKind.application_video.value):
        if not admin:
            raise HTTPException(403)
    elif not _media_public_ok(m, app_row) and not admin:
        raise HTTPException(403)

    ctype, _ = mimetypes.guess_type(str(path))
    # 统一将常见视频格式映射为 video/mp4，确保浏览器和小程序正常播放
    ext = path.suffix.lower()
    if ext in (".mp4", ".m4v", ".mov"):
        ctype = "video/mp4"
    elif ext in (".webm",):
        ctype = "video/webm"
    elif ext in (".avi",):
        ctype = "video/x-msvideo"
    return FileResponse(path, media_type=ctype or "application/octet-stream", headers={"Accept-Ranges": "bytes"})


@app.get("/api/showcase")
async def api_showcase(request: Request, db: Session = Depends(get_db)):
    """小程序爱心展示 JSON 接口：返回已完成手术且同意公开展示的条目。"""
    base = str(request.base_url).rstrip("/")
    q = (
        db.query(Application)
        .options(selectinload(Application.media))
        .filter(Application.status == ApplicationStatus.surgery_completed.value)
        .filter(or_(Application.showcase_consent.is_(True), Application.showcase_consent.is_(None)))
        .order_by(Application.updated_at.desc())
    )
    items = []
    for a in q.all():
        before_imgs = [f"{base}/file/{m.id}" for m in a.media
                       if m.kind == MediaKind.surgery_before.value and not m.stored_path.lower().endswith(('.mp4', '.mov', '.avi'))]
        after_imgs  = [f"{base}/file/{m.id}" for m in a.media
                       if m.kind == MediaKind.surgery_after.value and not m.stored_path.lower().endswith(('.mp4', '.mov', '.avi'))]
        before_vids = [f"{base}/file/{m.id}" for m in a.media
                       if m.kind == MediaKind.surgery_before.value and m.stored_path.lower().endswith(('.mp4', '.mov', '.avi'))]
        after_vids  = [f"{base}/file/{m.id}" for m in a.media
                       if m.kind == MediaKind.surgery_after.value and m.stored_path.lower().endswith(('.mp4', '.mov', '.avi'))]
        if not (before_imgs or after_imgs or before_vids or after_vids):
            continue
        # 姓氏保留，名字用 * 替代
        name = a.applicant_name.strip() if a.applicant_name else ""
        if len(name) >= 2:
            masked_name = name[0] + "*" * (len(name) - 1)
        elif len(name) == 1:
            masked_name = name
        else:
            masked_name = "—"
        items.append({
            "id": a.id,
            "cat_nickname": a.cat_nickname or "无名猫咪",
            "cat_gender": a.cat_gender,
            "address": a.address or "",
            "store": a.clinic_store or "",
            "surgery_date": a.updated_at.strftime("%Y-%m-%d") if a.updated_at else "",
            "applicant_masked": masked_name,
            "before_images": before_imgs,
            "after_images": after_imgs,
            "before_videos": before_vids,
            "after_videos": after_vids,
        })
    return {"items": items, "total": len(items)}


@app.get("/showcase", response_class=HTMLResponse)
async def page_showcase(request: Request, db: Session = Depends(get_db),
                        page: int = Query(1)):
    page_size = 4
    page = max(1, page)
    base_q = (
        db.query(Application)
        .options(selectinload(Application.media))
        .filter(Application.status == ApplicationStatus.surgery_completed.value)
        .filter(or_(Application.showcase_consent.is_(True), Application.showcase_consent.is_(None)))
        .order_by(Application.updated_at.desc())
    )
    # 先收集有图的记录（过滤掉无图案例）再分页
    all_apps = base_q.all()
    all_items = []
    for a in all_apps:
        before = [x for x in a.media if x.kind == MediaKind.surgery_before.value]
        after  = [x for x in a.media if x.kind == MediaKind.surgery_after.value]
        # 若无术前专用照片，回退到申请人上传的原始素材
        before_is_fallback = False
        if not before:
            before = [x for x in a.media if x.kind in (
                MediaKind.application_image.value, MediaKind.application_video.value)]
            before_is_fallback = bool(before)
        if before or after:
            all_items.append({"app": a, "before": before, "after": after,
                               "before_is_fallback": before_is_fallback})
    total      = len(all_items)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page        = min(page, total_pages)
    items       = all_items[(page - 1) * page_size : page * page_size]
    return templates.TemplateResponse(request, "uk/showcase.html",
        {
            "request": request,
            "title": "公布展示 · TNR 术前术后",
            "items": items,
            "page": page,
            "total_pages": total_pages,
            "total": total,
        },
    )


@app.post("/admin/app/{app_id}/toggle-showcase")
async def toggle_showcase(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    consent: str = Form(""),  # 兼容老调用；新流程不再使用
    csrf_token: str = Form(""),
):
    """切换该申请的"公开展示"开关。
    新流程：默认全部公开，管理员点一下就 toggle 当前状态。
    老流程兼容：如果传了 consent，按其取值生效。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    if consent != "":
        row.showcase_consent = consent.lower() in ("true", "1", "on", "yes")
    else:
        # 视当前 NULL 为公开（默认公开规则）
        currently_public = (row.showcase_consent is None) or bool(row.showcase_consent)
        row.showcase_consent = not currently_public
    _audit(db, request, "toggle_showcase", application_id=app_id, detail={"consent": row.showcase_consent})
    db.commit()
    return _admin_back(request, app_id)


@app.post("/admin/app/{app_id}/mark-scheduled")
async def mark_scheduled(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    appointment_at: str = Form(""),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.approved.value,
            ApplicationStatus.pre_approved.value,
        },
        "标记已预约",
    )
    row.status = ApplicationStatus.scheduled.value
    if appointment_at.strip():
        row.appointment_at = appointment_at.strip()
    _audit(db, request, "mark_scheduled", application_id=app_id, detail={"appointment_at": row.appointment_at})
    db.commit()
    push_application_result(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        applicant_name=row.applicant_name,
        status_text="已预约",
        phone_masked=row.phone,
        note="请按约定时间携带猫咪到院",
        submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return _admin_back(request, app_id)


@app.post("/admin/app/{app_id}/mark-cancelled")
async def mark_cancelled(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    reason: str = Form(""),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.approved.value,
            ApplicationStatus.pre_approved.value,
            ApplicationStatus.scheduled.value,
        },
        "取消",
    )
    row.status = ApplicationStatus.cancelled.value
    if reason.strip():
        row.reject_reason = reason.strip()
    _audit(db, request, "mark_cancelled", application_id=app_id, detail={"reason": row.reject_reason})
    db.commit()
    push_application_result(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        applicant_name=row.applicant_name,
        status_text="已取消",
        phone_masked=row.phone,
        note=(row.reject_reason or "如需帮助请联系医院")[:20],
        submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return _admin_back(request, app_id)


@app.post("/admin/app/{app_id}/mark-no-show")
async def mark_no_show(
    app_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    _require_status_in(
        row,
        {
            ApplicationStatus.scheduled.value,
        },
        "标记爽约",
    )
    row.status = ApplicationStatus.no_show.value
    _audit(db, request, "mark_no_show", application_id=app_id)
    db.commit()
    push_application_result(
        db,
        application_id=app_id,
        openid=row.wechat_openid,
        applicant_name=row.applicant_name,
        status_text="爽约",
        phone_masked=row.phone,
        note="如需改期请联系医院",
        submitted_at=row.created_at.strftime("%Y-%m-%d %H:%M") if row.created_at else "",
        action_at=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return _admin_back(request, app_id)


@app.post("/admin/wechat/test-send")
async def admin_wechat_test_send(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    application_id: int = Form(...),
    template: str = Form("application_result"),  # application_result / surgery_done
    openid: str = Form(""),
):
    raise HTTPException(status_code=410, detail="该功能已移除")


# ══════════════════════ 客户档案 CRM ══════════════════════

@app.get("/admin/customers", response_class=HTMLResponse)
async def page_admin_customers(
    request: Request,
    db: Session = Depends(get_db),
    q: str = Query(""),
    page: int = Query(1),
    store: str = Query(""),
):
    """今日工作台 + 客户档案速查。

    顶部：13+ 张卡片汇总今天/本周要做的事。
    底部：客户搜索 + 列表（搜索框自动聚焦，按 / 也可聚焦）。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    # 手机 UA 进桌面工作台 → 直接转手机 PWA。force_desktop cookie / ?desktop=1 可绕过
    if _is_mobile_ua(request):
        return RedirectResponse("/m", status_code=302)
    PAGE_SIZE = 30
    query = db.query(Customer)
    q = q.strip()
    # 速查：q 是 11 位手机号且唯一命中 → 直接跳客户档案
    if q and len(q) == 11 and q.isdigit():
        _hit = db.query(Customer).filter(
            or_(Customer.phone == q, Customer.phones_extra.like(f"%{q}%"))
        ).all()
        # 备用号是 CSV 子串 — Python 二次精确过滤
        _hit = [c for c in _hit
                if c.phone == q or q in [x.strip() for x in (c.phones_extra or "").split(",") if x.strip()]]
        if len(_hit) == 1:
            return RedirectResponse(f"/admin/customers/{_hit[0].id}", status_code=303)
    if q:
        # 子查询：宠物名命中的客户 id
        _pet_owner_ids = db.query(Pet.customer_id).filter(Pet.name.ilike(f"%{q}%"))
        query = query.filter(
            or_(
                Customer.name.ilike(f"%{q}%"),
                Customer.phone.ilike(f"%{q}%"),
                Customer.phones_extra.ilike(f"%{q}%"),
                Customer.id.in_(_pet_owner_ids),
            )
        )
    total = query.count()
    customers = query.order_by(Customer.id.desc()).offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()

    # 工作台数据：staff 锁本店；superadmin 可通过 ?store=东环店/横岗店 切
    admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        wb_store = (store or "").strip()
    else:
        wb_store = admin_store

    from app.services.dashboard import build_workbench
    try:
        wb = build_workbench(db, wb_store)
    except Exception as _e:
        logger.warning("[workbench] build failed: %s", _e)
        wb = {"urgent": [], "weekly": [], "stock": []}

    # B1 UK 重写：渲染新工作台模板；旧模板 admin_customers.html 暂留以备回滚
    return templates.TemplateResponse(
        request,
        "uk/workbench.html",
        {
            "customers": customers,
            "q": q,
            "page": page,
            "total": total,
            "page_size": PAGE_SIZE,
            "total_pages": max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE),
            "csrf_token": _get_csrf_token(request),
            "workbench": wb,
            "wb_store": wb_store,
            "is_superadmin": request.session.get("admin_role") == "superadmin",
            "msg": request.query_params.get("msg"),
            "err": request.query_params.get("err"),
        },
    )


@app.post("/admin/customers/create")
async def admin_customer_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""),
    phone: str = Form(""),
    address: str = Form(""),
    notes: str = Form(""),
    source: str = Form("manual"),
    next_url: str = Form(""),
    is_internal: str = Form(""),           # "1" = 员工内购档案
    internal_staff_id: str = Form(""),     # 关联员工 id（可选）
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    phone_clean = phone.strip()[:40]
    # 同手机号已有档案 → 不建新档，直接跳已有档案（提示一下）
    if phone_clean:
        existing = db.query(Customer).filter(Customer.phone == phone_clean).first()
        if not existing:
            # 备用号也算
            cands = db.query(Customer).filter(Customer.phones_extra.like(f"%{phone_clean}%")).all()
            for c in cands:
                extras = [x.strip() for x in (c.phones_extra or "").split(",") if x.strip()]
                if phone_clean in extras:
                    existing = c
                    break
        if existing:
            from urllib.parse import quote as _q
            msg = _q(f"该手机号已有客户档案 #{existing.id}「{existing.name or '未命名'}」，已跳转。如需新增请使用其他手机号。", safe="")
            # next_url 模板支持 {id} 占位
            if next_url:
                target = _safe_next(next_url.replace("{id}", str(existing.id)), f"/admin/customers/{existing.id}")
                sep = "&" if "?" in target else "?"
                return RedirectResponse(f"{target}{sep}msg={msg}", status_code=303)
            return RedirectResponse(f"/admin/customers/{existing.id}?msg={msg}", status_code=303)
    # 员工内购档案：仅超管可建
    _is_internal = (is_internal == "1") and (request.session.get("admin_role") == "superadmin")
    _staff_id: int | None = None
    if _is_internal and internal_staff_id.strip().isdigit():
        _staff_id = int(internal_staff_id.strip())
    cust = Customer(
        name=name.strip()[:120],
        phone=phone_clean,
        address=address.strip()[:500],
        notes=notes.strip(),
        source=("employee_internal" if _is_internal else (source.strip()[:40] or "manual")),
        is_internal=_is_internal,
        internal_staff_id=_staff_id,
    )
    db.add(cust)
    db.commit()
    if next_url:
        target = _safe_next(next_url.replace("{id}", str(cust.id)), f"/admin/customers/{cust.id}")
        sep = "&" if "?" in target else "?"
        return RedirectResponse(f"{target}{sep}msg=客户已创建", status_code=303)
    return RedirectResponse(f"/admin/customers/{cust.id}?msg=客户已创建", status_code=303)


def _customer_blockers(db: Session, customer_id: int) -> list[str]:
    """统计客户的业务关联记录。返回非空 list 表示有记录，不能删。"""
    blockers: list[str] = []
    n_visits = db.query(Visit).filter(Visit.customer_id == customer_id).count()
    if n_visits: blockers.append(f"{n_visits} 条病历")
    n_appts = db.query(Appointment).filter(Appointment.customer_id == customer_id).count()
    if n_appts: blockers.append(f"{n_appts} 条预约")
    n_invoices = db.query(Invoice).filter(Invoice.customer_id == customer_id).count()
    if n_invoices: blockers.append(f"{n_invoices} 张收费单")
    n_presc = db.query(Prescription).filter(Prescription.customer_id == customer_id).count()
    if n_presc: blockers.append(f"{n_presc} 张处方")
    n_so = db.query(SalesOrder).filter(SalesOrder.customer_id == customer_id).count()
    if n_so: blockers.append(f"{n_so} 张销售单")
    n_groom = db.query(GroomingOrder).filter(GroomingOrder.customer_id == customer_id).count()
    if n_groom: blockers.append(f"{n_groom} 张美容单")
    n_vacc = db.query(Vaccination).filter(Vaccination.customer_id == customer_id).count()
    if n_vacc: blockers.append(f"{n_vacc} 条疫苗记录")
    n_dew = db.query(DewormingRecord).filter(DewormingRecord.customer_id == customer_id).count()
    if n_dew: blockers.append(f"{n_dew} 条驱虫记录")
    # 钱包流水 / 套餐 / 押金 / 优惠券
    try:
        wallet = db.query(Wallet).filter(Wallet.customer_id == customer_id).first()
        if wallet:
            n_wtx = db.query(WalletTransaction).filter(WalletTransaction.wallet_id == wallet.id).count()
            if n_wtx: blockers.append(f"{n_wtx} 条钱包流水")
    except Exception:
        pass
    try:
        n_pkg = db.query(CustomerPackage).filter(CustomerPackage.customer_id == customer_id).count()
        if n_pkg: blockers.append(f"{n_pkg} 个套餐")
    except Exception:
        pass
    try:
        n_dep = db.query(Deposit).filter(Deposit.customer_id == customer_id).count()
        if n_dep: blockers.append(f"{n_dep} 笔押金")
    except Exception:
        pass
    try:
        n_coup = db.query(Coupon).filter(Coupon.customer_id == customer_id).count()
        if n_coup: blockers.append(f"{n_coup} 张优惠券")
    except Exception:
        pass
    # 协议任务也算
    try:
        n_consent = db.query(ConsentTask).filter(ConsentTask.customer_id == customer_id).count()
        if n_consent: blockers.append(f"{n_consent} 张协议")
    except Exception:
        pass
    return blockers


def _merge_customers(db: Session, primary_id: int, secondary_ids: list[int]) -> dict:
    """把 secondary 客户们的所有附属数据合并到 primary，然后删 secondary。
    返回每张表迁移的行数 + wallet 余额合并明细。"""
    moved: dict = {}
    # 1) 简单 customer_id 重定向 — 一律 UPDATE SET customer_id = primary
    redirect_tables = [
        ("pet", Pet),
        ("visit", Visit),
        ("appointment", Appointment),
        ("invoice", Invoice),
        ("prescription", Prescription),
        ("sales_order", SalesOrder),
        ("grooming_order", GroomingOrder),
        ("vaccination", Vaccination),
        ("deworming", DewormingRecord),
        ("customer_package", CustomerPackage),
        ("deposit", Deposit),
        ("coupon", Coupon),
        ("consent_task", ConsentTask),
        ("follow_up", FollowUp),
        ("exam_order_via_visit", None),  # ExamOrder 没 customer_id，靠 visit 链
    ]
    for label, cls in redirect_tables:
        if cls is None:
            continue
        if not hasattr(cls, "customer_id"):
            continue
        try:
            n = db.query(cls).filter(cls.customer_id.in_(secondary_ids)).update(
                {cls.customer_id: primary_id}, synchronize_session=False,
            )
            if n: moved[label] = n
        except Exception as _e:
            logger.warning("[merge cust] %s skip: %s", label, _e)
    # 病例号尝试 — Application（流浪猫申请，可能挂 customer_id 也可能不挂，安全跳过）
    # 2) 钱包合并 — 副 wallet 的流水迁到主 wallet，余额累加，删副 wallet
    try:
        primary_wallet = db.query(Wallet).filter(Wallet.customer_id == primary_id).first()
        sec_wallets = db.query(Wallet).filter(Wallet.customer_id.in_(secondary_ids)).all()
        for sw in sec_wallets:
            if primary_wallet is None:
                # 主没钱包，直接改副 wallet 的 customer_id 成主，跳过
                sw.customer_id = primary_id
                primary_wallet = sw
                continue
            # 迁流水
            db.query(WalletTransaction).filter(WalletTransaction.wallet_id == sw.id).update(
                {WalletTransaction.wallet_id: primary_wallet.id}, synchronize_session=False,
            )
            # 余额累加
            primary_wallet.balance = round(float(primary_wallet.balance or 0) + float(sw.balance or 0), 2)
            primary_wallet.balance_principal = round(float(primary_wallet.balance_principal or 0) + float(sw.balance_principal or 0), 2)
            primary_wallet.balance_bonus = round(float(primary_wallet.balance_bonus or 0) + float(sw.balance_bonus or 0), 2)
            primary_wallet.lifetime_recharge = round(float(primary_wallet.lifetime_recharge or 0) + float(sw.lifetime_recharge or 0), 2)
            primary_wallet.lifetime_consume = round(float(primary_wallet.lifetime_consume or 0) + float(sw.lifetime_consume or 0), 2)
            primary_wallet.updated_at = datetime.utcnow()
            db.delete(sw)
            moved["wallet_merged"] = (moved.get("wallet_merged") or 0) + 1
    except Exception as _e:
        logger.warning("[merge cust] wallet skip: %s", _e)
    # 3) 合并 secondary 客户的元信息（备注 / 备用号 / openid / 地址）到 primary 后再删
    primary = db.get(Customer, primary_id)
    if primary is not None:
        sec_custs = db.query(Customer).filter(Customer.id.in_(secondary_ids)).all()
        # 备用号汇总
        all_phones: list[str] = [x.strip() for x in (primary.phones_extra or "").split(",") if x.strip()]
        for sc in sec_custs:
            if sc.phone and sc.phone != primary.phone and sc.phone not in all_phones:
                all_phones.append(sc.phone)
            for x in (sc.phones_extra or "").split(","):
                xs = x.strip()
                if xs and xs != primary.phone and xs not in all_phones:
                    all_phones.append(xs)
            # openid（主没有但副有 → 迁）
            if not primary.wechat_openid and sc.wechat_openid:
                primary.wechat_openid = sc.wechat_openid
            # address（主没有但副有 → 迁）
            if not primary.address and sc.address:
                primary.address = sc.address
            # 备注追加
            if sc.notes and sc.notes.strip():
                tag = f"[合并自 #{sc.id}] {sc.notes.strip()}"
                primary.notes = ((primary.notes or "") + "\n" + tag).strip()
            # 姓名升级：主是占位名 + 副是正常名 → 用副的
            if _is_invalid_name(primary.name) and not _is_invalid_name(sc.name):
                primary.name = sc.name
        primary.phones_extra = ",".join(all_phones)[:500]
        primary.updated_at = datetime.utcnow()
        # 4) 删 secondary
        for sc in sec_custs:
            db.delete(sc)
        moved["customers_deleted"] = len(sec_custs)
    db.flush()
    return moved


@app.get("/admin/customers/duplicates", response_class=HTMLResponse)
async def admin_customers_duplicates_page(request: Request, db: Session = Depends(get_db)):
    """重号客户清理页：列出所有同手机号 2+ 的客户档案组。"""
    require_admin(request)
    require_superadmin(request)
    from sqlalchemy import func as _f
    dup_phones = db.query(Customer.phone).filter(Customer.phone != "").group_by(Customer.phone)\
        .having(_f.count(Customer.id) > 1).all()
    groups = []
    for (ph,) in dup_phones:
        rows = db.query(Customer).filter(Customer.phone == ph).order_by(Customer.id.asc()).all()
        enriched = []
        for c in rows:
            n_pets = db.query(Pet).filter(Pet.customer_id == c.id).count()
            n_visits = db.query(Visit).filter(Visit.customer_id == c.id).count()
            n_invoices = db.query(Invoice).filter(Invoice.customer_id == c.id).count()
            n_appts = db.query(Appointment).filter(Appointment.customer_id == c.id).count()
            # 用 (n_pets, n_visits, n_invoices, -id) 综合分判主候选
            score = n_pets * 10 + n_visits + n_invoices + n_appts
            enriched.append({
                "c": c, "n_pets": n_pets, "n_visits": n_visits,
                "n_invoices": n_invoices, "n_appts": n_appts,
                "score": score,
            })
        # 主候选：评分最高那条
        enriched.sort(key=lambda x: (-x["score"], x["c"].id))
        suggested_primary_id = enriched[0]["c"].id
        groups.append({"phone": ph, "rows": enriched, "suggested_primary_id": suggested_primary_id})
    return templates.TemplateResponse(request, "uk/customers_duplicates.html", {
        "groups": groups,
        "total_groups": len(groups),
        "total_customers": sum(len(g["rows"]) for g in groups),
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/customers/duplicates/merge")
async def admin_customers_merge_post(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    require_superadmin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    try:
        primary_id = int(form.get("primary_id") or 0)
    except ValueError:
        raise HTTPException(400, "primary_id 格式错")
    if not primary_id:
        raise HTTPException(400, "缺 primary_id")
    primary = db.get(Customer, primary_id)
    if not primary:
        raise HTTPException(404, "primary 客户不存在")
    # secondaries：所有同手机号其他客户
    secondaries = db.query(Customer).filter(
        Customer.phone == primary.phone, Customer.id != primary_id,
    ).all()
    if not secondaries:
        return RedirectResponse("/admin/customers/duplicates?msg=该号码已无重号客户", status_code=303)
    sec_ids = [s.id for s in secondaries]
    moved = _merge_customers(db, primary_id, sec_ids)
    _audit(db, request, "merge_customers", application_id=None,
           detail={"primary_id": primary_id, "secondary_ids": sec_ids, "moved": moved})
    db.commit()
    parts = [f"{k}: {v}" for k, v in moved.items()]
    from urllib.parse import quote as _q
    msg = _q(f"已合并 {len(secondaries)} 条到 #{primary_id}（{' / '.join(parts)}）", safe="")
    return RedirectResponse(f"/admin/customers/duplicates?msg={msg}", status_code=303)


@app.post("/admin/customers/{customer_id}/delete")
async def admin_customer_delete(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """删除客户档案 — 仅在没有任何业务关联记录时允许。
    会级联删除 customer 名下所有 Pet（同样未有业务记录的）。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    blockers = _customer_blockers(db, customer_id)
    if blockers:
        msg = "该客户有业务记录（" + " / ".join(blockers) + "），不允许删除。"
        from urllib.parse import quote as _q
        return RedirectResponse(f"/admin/customers/{customer_id}?msg={_q(msg, safe='')}", status_code=303)
    # 客户名下的宠物也要确保没记录（理论上若客户无记录，宠物大概率也没。但稳妥起见再扫一遍）
    pets = db.query(Pet).filter(Pet.customer_id == customer_id).all()
    for pet in pets:
        if (
            db.query(Visit).filter(Visit.pet_id == pet.id).count()
            or db.query(Appointment).filter(Appointment.pet_id == pet.id).count()
            or db.query(Invoice).filter(Invoice.pet_id == pet.id).count()
            or db.query(Vaccination).filter(Vaccination.pet_id == pet.id).count()
            or db.query(Prescription).filter(Prescription.pet_id == pet.id).count()
        ):
            from urllib.parse import quote as _q
            msg = _q(f"宠物「{pet.name}」仍有业务记录，不能删除客户。", safe="")
            return RedirectResponse(f"/admin/customers/{customer_id}?msg={msg}", status_code=303)
    _audit(db, request, "delete_customer", application_id=None,
           detail={"customer_id": customer_id, "name": cust.name, "phone": cust.phone, "pets_deleted": len(pets)})
    # ORM cascade 会自动删 pets（Customer.pets 配了 cascade="all, delete-orphan"）
    db.delete(cust)
    db.commit()
    return RedirectResponse("/admin/customers?msg=客户档案已删除", status_code=303)


# ─────────── 老系统客户 xls 批量导入（superadmin only）───────────
@app.get("/admin/customers/import", response_class=HTMLResponse)
async def page_admin_customers_import(request: Request):
    require_admin(request)
    require_superadmin(request)
    return templates.TemplateResponse(
        request,
        "admin_customers_import.html",
        {
            "csrf_token": _get_csrf_token(request),
            "result": None,
        },
    )


@app.post("/admin/customers/import", response_class=HTMLResponse)
async def admin_customers_import_post(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    file: Optional[UploadFile] = File(None),
    confirm: str = Form(""),
    fallback_store: str = Form("东环店"),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    do_commit = (confirm == "yes")

    # 读取所有 overwrite_<phone>=yes 字段（撞号但用户选择"用 xls 覆盖"的）
    # 以及 import_nophone_<np_idx>=yes 字段（无手机号但用户选择导入的）
    form_data = await request.form()
    overwrite_phones = set()
    import_np_indices = set()
    for key, val in form_data.items():
        if key.startswith("overwrite_") and val == "yes":
            phone_part = key[len("overwrite_"):]
            if phone_part:
                overwrite_phones.add(phone_part)
        elif key.startswith("import_nophone_") and val == "yes":
            try:
                import_np_indices.add(int(key[len("import_nophone_"):]))
            except ValueError:
                pass

    result = {"ok": False, "msg": "", "stats": None, "samples": [], "committed": False}

    if not file or not file.filename:
        result["msg"] = "请选择 xls / xlsx 文件"
        return templates.TemplateResponse(
            request,
            "admin_customers_import.html",
            {"csrf_token": _get_csrf_token(request), "result": result},
        )

    import pandas as pd
    import io as _io

    try:
        raw = await file.read()
        fname_lower = (file.filename or "").lower()
        is_xls = fname_lower.endswith(".xls")
        bio = _io.BytesIO(raw)
        try:
            df = pd.read_excel(bio)
        except ImportError as e_imp:
            # 服务器缺 xlrd（.xls）或 openpyxl（.xlsx）依赖
            if is_xls:
                result["msg"] = (
                    "服务器暂时无法读取 .xls 老格式（缺 xlrd 包）。"
                    "解决：用 Excel 打开此文件 → 文件 → 另存为 → 选「Excel 工作簿 (*.xlsx)」"
                    " → 上传转好的 .xlsx 文件即可。"
                )
            else:
                result["msg"] = f"服务器缺解析依赖：{e_imp}"
            return templates.TemplateResponse(
                request, "admin_customers_import.html",
                {"csrf_token": _get_csrf_token(request), "result": result},
            )
        except Exception as e1:
            bio.seek(0)
            try:
                dfs = pd.read_html(bio, encoding="utf-8")
                df = dfs[0]
            except Exception as e2:
                if is_xls:
                    result["msg"] = (
                        "无法解析 .xls 文件。最简方案：用 Excel 打开 → 另存为 .xlsx → 重新上传。"
                        f"（原始错误：{e1}）"
                    )
                else:
                    result["msg"] = f"无法解析文件：{e1} / {e2}"
                return templates.TemplateResponse(
                    request, "admin_customers_import.html",
                    {"csrf_token": _get_csrf_token(request), "result": result},
                )
    except Exception as e:
        result["msg"] = f"读取失败：{e}"
        return templates.TemplateResponse(
            request, "admin_customers_import.html",
            {"csrf_token": _get_csrf_token(request), "result": result},
        )

    # 字段名兼容（中文列）
    col_map = {
        "name": ["客户姓名", "姓名", "name"],
        "phone": ["联系电话", "电话", "手机", "phone"],
        "member_id": ["会员编号", "编号", "id"],
        "gender": ["性别"],
        "level": ["会员级别"],
        "card_bal": ["会员卡余额"],
        "acc_bal": ["账户余额"],
        "spent": ["累计消费"],
        "org": ["所属机构"],
        "source": ["客户来源", "来源"],
        "notes": ["备注"],
        "created_at": ["登记日期", "创建时间"],
    }
    def pick(row, key):
        for c in col_map.get(key, []):
            if c in row.index:
                v = row[c]
                if pd.isna(v):
                    return None
                return v
        return None

    def _phone(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        if isinstance(v, (int, float)):
            return str(int(v))
        return str(v).strip()

    def _f(v) -> float:
        if v is None:
            return 0.0
        try:
            return float(v)
        except Exception:
            return 0.0

    def map_store(org_value) -> str:
        """老机构名 → 当前系统门店。

        - 龙华 / 东环 → 东环店
        - 横岗 → 横岗店
        - 其他 → fallback_store
        """
        if not org_value:
            return fallback_store
        s = str(org_value)
        if "横岗" in s:
            return "横岗店"
        if "龙华" in s or "东环" in s:
            return "东环店"
        return fallback_store

    def build_notes(row, target_store):
        parts = []
        mid = pick(row, "member_id")
        if mid: parts.append(f"老系统会员号:{mid}")
        g = pick(row, "gender")
        if g: parts.append(f"性别:{g}")
        spent = _f(pick(row, "spent"))
        if spent > 0: parts.append(f"老系统累计消费:¥{spent:,.0f}")
        org = pick(row, "org")
        if org: parts.append(f"原属:{org}→{target_store}")
        lvl = pick(row, "level")
        if lvl and str(lvl).strip() and str(lvl) != "warmsoft客户":
            parts.append(f"等级:{lvl}")
        old_note = pick(row, "notes")
        if old_note and str(old_note).strip():
            parts.append(f"原备注:{old_note}")
        return "[导入] " + " | ".join(parts) if parts else ""

    # 现有手机号
    existing_phones = set(
        p for (p,) in db.query(Customer.phone).filter(Customer.phone.isnot(None)).all() if p
    )

    n_new = 0
    n_skip_dup = 0
    n_skip_no_phone = 0
    n_overwritten = 0
    overwrite_jobs = []  # [(phone, name, notes, created_at, target_store)] 覆盖现有档案
    wallet_jobs = []  # (idx_in_new_customers, total_bal, target_store)
    new_customers = []
    samples = []
    org_breakdown = {}  # {老机构: count}
    store_breakdown = {}  # {目标门店: count}
    dup_details = []     # 撞号客户明细
    no_phone_details = []  # 无手机号明细
    wallet_details = []  # 有余额客户明细（核对）

    # 现有手机号 → 姓名 速查（撞号时方便对比是不是同一人）
    existing_phone_to_name = {}
    for cust_row in db.query(Customer.phone, Customer.name).filter(Customer.phone.isnot(None)).all():
        if cust_row[0]:
            existing_phone_to_name[cust_row[0]] = cust_row[1]

    np_idx_counter = -1  # 无手机号行计数，用于稳定 form key
    for _, row in df.iterrows():
        phone = _phone(pick(row, "phone"))
        name_v = pick(row, "name")
        name_str = str(name_v).strip() if name_v else ""
        if not phone:
            np_idx_counter += 1
            # 用户在 no-phone 表里勾了导入 → fallthrough 到下面新建分支
            if np_idx_counter not in import_np_indices:
                n_skip_no_phone += 1
                if len(no_phone_details) < 100:
                    org_v = pick(row, "org")
                    spent_v = _f(pick(row, "spent"))
                    card_v = _f(pick(row, "card_bal"))
                    acc_v = _f(pick(row, "acc_bal"))
                    no_phone_details.append({
                        "np_idx": np_idx_counter,
                        "name": name_str or "—",
                        "org": str(org_v) if org_v else "—",
                        "spent": spent_v,
                        "balance": card_v + acc_v,
                        "will_import": False,
                    })
                continue
            # else: 用户选了要导入 — 继续走新增流程，但 phone 为空
            if len(no_phone_details) < 100:
                org_v = pick(row, "org")
                spent_v = _f(pick(row, "spent"))
                card_v = _f(pick(row, "card_bal"))
                acc_v = _f(pick(row, "acc_bal"))
                no_phone_details.append({
                    "np_idx": np_idx_counter,
                    "name": name_str or "—",
                    "org": str(org_v) if org_v else "—",
                    "spent": spent_v,
                    "balance": card_v + acc_v,
                    "will_import": True,
                })
        if phone and phone in existing_phones:
            # 撞号 → 默认跳过；除非用户在 dup 表里勾了 overwrite_<phone>=yes
            if phone in overwrite_phones:
                # 排队等真写时一起 update
                org_value_dup = pick(row, "org")
                target_store_dup = map_store(org_value_dup)
                org_breakdown[str(org_value_dup) if org_value_dup else "(空)"] = (
                    org_breakdown.get(str(org_value_dup) if org_value_dup else "(空)", 0) + 1
                )
                store_breakdown[target_store_dup] = store_breakdown.get(target_store_dup, 0) + 1
                notes_dup = build_notes(row, target_store_dup)
                created_at_v_dup = pick(row, "created_at")
                try:
                    created_at_dup = pd.to_datetime(created_at_v_dup).to_pydatetime() if created_at_v_dup else None
                except Exception:
                    created_at_dup = None
                overwrite_jobs.append({
                    "phone": phone,
                    "name": name_str,
                    "notes": notes_dup,
                    "created_at": created_at_dup,
                    "target_store": target_store_dup,
                })
                n_overwritten += 1
                # 撞号仍然显示在 dup_details，但带状态
                if len(dup_details) < 100:
                    dup_details.append({
                        "phone": phone,
                        "old_name": name_str or "—",
                        "existing_name": existing_phone_to_name.get(phone, "?") or "—",
                        "balance": _f(pick(row, "card_bal")) + _f(pick(row, "acc_bal")),
                        "will_overwrite": True,
                    })
                continue
            n_skip_dup += 1
            if len(dup_details) < 100:
                dup_details.append({
                    "phone": phone,
                    "old_name": name_str or "—",
                    "existing_name": existing_phone_to_name.get(phone, "?") or "—",
                    "balance": _f(pick(row, "card_bal")) + _f(pick(row, "acc_bal")),
                    "will_overwrite": False,
                })
            continue

        org_value = pick(row, "org")
        target_store = map_store(org_value)
        org_key = str(org_value) if org_value else "(空)"
        org_breakdown[org_key] = org_breakdown.get(org_key, 0) + 1
        store_breakdown[target_store] = store_breakdown.get(target_store, 0) + 1

        name = pick(row, "name")
        name = str(name).strip() if name else ""
        source = pick(row, "source")
        source = str(source).strip() if source else None
        notes = build_notes(row, target_store)
        created_at_v = pick(row, "created_at")
        try:
            created_at = pd.to_datetime(created_at_v).to_pydatetime() if created_at_v else datetime.now()
        except Exception:
            created_at = datetime.now()

        card_bal = _f(pick(row, "card_bal"))
        acc_bal = _f(pick(row, "acc_bal"))
        total_bal = card_bal + acc_bal

        c = Customer(
            name=name,
            phone=phone,
            source=source,
            notes=notes,
            created_at=created_at,
            updated_at=created_at,
        )
        new_customers.append(c)
        if total_bal > 0:
            wallet_jobs.append((len(new_customers) - 1, total_bal, target_store))
            if len(wallet_details) < 200:
                wallet_details.append({
                    "name": name,
                    "phone": phone,
                    "target_store": target_store,
                    "card_bal": _f(pick(row, "card_bal")),
                    "acc_bal": _f(pick(row, "acc_bal")),
                    "total": total_bal,
                })

        n_new += 1
        existing_phones.add(phone)
        if len(samples) < 5:
            samples.append({
                "name": name, "phone": phone, "source": source or "—",
                "notes": notes[:120],
                "balance": total_bal,
                "target_store": target_store,
            })

    wallet_total = sum(b for _, b, _ in wallet_jobs)

    # 按金额倒序排（有余额的客户）/ 按金额倒序（无手机号但有消费/余额）
    wallet_details.sort(key=lambda x: -x["total"])
    # 无手机号：先按 will_import True/False 分组（True 在前），再按 消费+余额 倒序
    no_phone_details.sort(key=lambda x: (not x.get("will_import", False), -(x["balance"] + x["spent"])))
    dup_details.sort(key=lambda x: -x["balance"])

    result["stats"] = {
        "total_rows": int(len(df)),
        "new": n_new,
        "skip_dup": n_skip_dup,
        "skip_no_phone": n_skip_no_phone,
        "overwritten": n_overwritten,
        "wallets": len(wallet_jobs),
        "wallet_total": wallet_total,
        "fallback_store": fallback_store,
        "org_breakdown": sorted(org_breakdown.items(), key=lambda x: -x[1]),
        "store_breakdown": sorted(store_breakdown.items(), key=lambda x: -x[1]),
    }
    result["samples"] = samples
    result["dup_details"] = dup_details
    result["no_phone_details"] = no_phone_details
    result["wallet_details"] = wallet_details
    result["ok"] = True

    if do_commit and (new_customers or overwrite_jobs):
        # 真写：新增客户
        if new_customers:
            db.add_all(new_customers)
            db.flush()
        for idx, bal, target_store in wallet_jobs:
            cid = new_customers[idx].id
            w = Wallet(
                customer_id=cid,
                balance=bal,
                lifetime_recharge=bal,
                lifetime_consume=0,
            )
            db.add(w)
            db.flush()
            tx = WalletTransaction(
                wallet_id=w.id,
                customer_id=cid,
                type="adjust",
                amount=bal,
                balance_after=bal,
                note=f"老系统历史余额导入（{target_store}）",
                operator=request.session.get("admin", "导入"),
                store=target_store,
            )
            db.add(tx)

        # 真写：覆盖现有档案（用户在 dup 表勾选的）
        for job in overwrite_jobs:
            existing = db.query(Customer).filter(Customer.phone == job["phone"]).first()
            if not existing:
                continue
            if job["name"]:
                existing.name = job["name"][:120]
            # notes：保留旧 notes，新 notes 追加在后面（不丢历史）
            old_notes = (existing.notes or "").strip()
            new_notes_part = job["notes"]
            if new_notes_part:
                if old_notes:
                    existing.notes = old_notes + "\n" + new_notes_part
                else:
                    existing.notes = new_notes_part
            if job["created_at"]:
                # 旧档案的 created_at 通常更早，留旧的；只更新 updated_at
                pass
            existing.updated_at = datetime.now()

        db.commit()
        result["committed"] = True
        parts = []
        if n_new: parts.append(f"导入 {n_new} 个新客户")
        if len(wallet_jobs): parts.append(f"{len(wallet_jobs)} 个钱包")
        if n_overwritten: parts.append(f"覆盖 {n_overwritten} 个已有档案")
        result["msg"] = "已" + " + ".join(parts) if parts else "无变更"

    return templates.TemplateResponse(
        request,
        "admin_customers_import.html",
        {"csrf_token": _get_csrf_token(request), "result": result},
    )


@app.get("/admin/customers/{customer_id}", response_class=HTMLResponse)
async def page_admin_customer_detail(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    pet_id: int = Query(0),    # 选中显示哪只宠物（默认第一只）
    tab: str = Query("visits"),  # 默认激活的标签
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    admin_store = _get_admin_store(request)  # 限店员工的门店短名
    # 宠物列表（限店员工只看本店宠物）
    pets_q = db.query(Pet).filter(Pet.customer_id == customer_id)
    if admin_store:
        pets_q = pets_q.filter((Pet.store == admin_store) | (Pet.store == ""))
    pets = pets_q.order_by(Pet.id.desc()).all()
    pet_map = {p.id: p for p in pets}

    # 默认选中宠物
    if pet_id and pet_id in pet_map:
        active_pet = pet_map[pet_id]
    elif pets:
        active_pet = pets[0]
    else:
        active_pet = None
    active_pet_id = active_pet.id if active_pet else 0

    # ── 客户级数据 ──
    applications = db.query(Application).filter(Application.customer_id == customer_id).order_by(Application.id.desc()).limit(50).all()
    cust_sales_orders = db.query(SalesOrder).filter(SalesOrder.customer_id == customer_id).order_by(SalesOrder.id.desc()).limit(100).all()
    cust_invoices = db.query(Invoice).filter(Invoice.customer_id == customer_id).order_by(Invoice.id.desc()).limit(100).all()

    # ── 所有宠物的"最近一次"疫苗/驱虫（用于宠物列表行上展示）──
    all_pet_ids = [p.id for p in pets]
    latest_vacc_by_pet: dict[int, "Vaccination"] = {}
    latest_deworm_by_pet: dict[int, "DewormingRecord"] = {}
    if all_pet_ids:
        for v in db.query(Vaccination).filter(Vaccination.pet_id.in_(all_pet_ids)).order_by(Vaccination.vaccinated_date.desc(), Vaccination.id.desc()).all():
            if v.pet_id not in latest_vacc_by_pet:
                latest_vacc_by_pet[v.pet_id] = v
        for d in db.query(DewormingRecord).filter(DewormingRecord.pet_id.in_(all_pet_ids)).order_by(DewormingRecord.deworm_date.desc(), DewormingRecord.id.desc()).all():
            if d.pet_id not in latest_deworm_by_pet:
                latest_deworm_by_pet[d.pet_id] = d

    # ── 宠物级数据（仅取选中宠物的，节省查询）──
    if active_pet:
        appointments = db.query(Appointment).filter(
            Appointment.pet_id == active_pet_id
        ).order_by(Appointment.appointment_date.desc(), Appointment.appointment_time.desc()).limit(50).all()
        visits = db.query(Visit).filter(Visit.pet_id == active_pet_id).order_by(Visit.visit_date.desc(), Visit.id.desc()).limit(100).all()
        prescriptions = db.query(Prescription).filter(Prescription.pet_id == active_pet_id).order_by(Prescription.id.desc()).limit(50).all()
        exam_orders = db.query(ExamOrder).join(Visit, ExamOrder.visit_id == Visit.id).filter(Visit.pet_id == active_pet_id).order_by(ExamOrder.id.desc()).limit(50).all()
        vaccinations = db.query(Vaccination).filter(Vaccination.pet_id == active_pet_id).order_by(Vaccination.vaccinated_date.desc()).all()
        dewormings = db.query(DewormingRecord).filter(DewormingRecord.pet_id == active_pet_id).order_by(DewormingRecord.deworm_date.desc()).all()
        for _d in dewormings:
            _l, _r = _is_deworming_locked(db, _d)
            _d._locked = _l
            _d._lock_reason = _r
        for _v in vaccinations:
            _l, _r = _is_vaccination_locked(db, _v)
            _v._locked = _l
            _v._lock_reason = _r
        groomings = db.query(GroomingOrder).filter(GroomingOrder.pet_id == active_pet_id).order_by(GroomingOrder.groom_date.desc(), GroomingOrder.id.desc()).all()
        for _g in groomings:
            _l, _r = _is_grooming_locked(db, _g)
            _g._locked = _l
            _g._lock_reason = _r
        weight_records = db.query(WeightRecord).filter(WeightRecord.pet_id == active_pet_id).order_by(WeightRecord.record_date.asc()).all()
        medical_docs = db.query(MedicalDocument).filter(MedicalDocument.pet_id == active_pet_id).order_by(MedicalDocument.id.desc()).all()
        # 该宠物名下发票 = 直接关联该宠物 OR 通过 visit_id 关联
        visit_ids = {v.id for v in visits}
        pet_invoices = [
            inv for inv in cust_invoices
            if inv.pet_id == active_pet_id or (inv.visit_id and inv.visit_id in visit_ids)
        ]
        # 该宠物的销售单（按 pet_id；无 pet_id 的旧单子归入活跃宠物，避免数据消失）
        pet_sales_orders = [
            so for so in cust_sales_orders
            if so.pet_id == active_pet_id or (not so.pet_id)
        ]
    else:
        appointments, visits, prescriptions, exam_orders = [], [], [], []
        vaccinations, dewormings, weight_records, medical_docs = [], [], [], []
        groomings = []
        pet_invoices = []
        pet_sales_orders = []

    _SO_STATUS_ZH_LOCAL = {"pending": "待付款", "paid": "已收款", "cancelled": "已取消"}
    _INV_STATUS_ZH_LOCAL = {"unpaid": "未支付", "paid": "已支付", "cancelled": "已取消", "refunded": "已退款"}
    from datetime import date, timedelta
    today_str = date.today().isoformat()
    soon_str  = (date.today() + timedelta(days=7)).isoformat()

    # 录入驱虫弹窗用：本店可用的驱虫品目（含最近一批的批号）
    deworm_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category == "antiparasitic",
        InventoryItem.stock_qty > 0,
    ).order_by(InventoryItem.name).all()
    _attach_latest_batch(db, deworm_items)

    # ── 钱包 + 流水 ──
    wallet = db.query(Wallet).filter(Wallet.customer_id == customer_id).first()
    wallet_balance = float(wallet.balance) if wallet else 0.0
    wallet_lifetime_recharge = float(wallet.lifetime_recharge) if wallet else 0.0
    wallet_lifetime_consume = float(wallet.lifetime_consume) if wallet else 0.0
    wallet_txs = []
    if wallet:
        wallet_txs = (
            db.query(WalletTransaction)
            .filter(WalletTransaction.wallet_id == wallet.id)
            .order_by(WalletTransaction.id.desc())
            .limit(50)
            .all()
        )

    # ── 套餐 ──
    customer_packages = (
        db.query(CustomerPackage)
        .filter(CustomerPackage.customer_id == customer_id)
        .order_by(CustomerPackage.status.asc(), CustomerPackage.id.desc())
        .all()
    )
    active_packages_count = sum(1 for p in customer_packages if p.status == "active")
    # 客户档案里售卖套餐：staff 只看本店+通用，避免售错门店的套餐
    package_products = _apply_store_filter(
        db.query(PackageProduct).filter(PackageProduct.is_active == True),
        PackageProduct.store, _get_admin_store(request),
    ).order_by(PackageProduct.id.desc()).all()

    # ── 押金 ──
    deposits = (
        db.query(Deposit)
        .filter(Deposit.customer_id == customer_id)
        .order_by(Deposit.status.asc(), Deposit.id.desc())
        .all()
    )
    held_deposits_count = sum(1 for d in deposits if d.status in ("held", "partial_refund"))

    # ── 优惠券 ──
    coupons = (
        db.query(Coupon)
        .filter(Coupon.customer_id == customer_id)
        .order_by(Coupon.status.asc(), Coupon.id.desc())
        .all()
    )
    active_coupons_count = sum(1 for c in coupons if c.status == "issued" and not _coupon_is_expired(c))

    # ── 头部信号 chip 数据 ──
    unpaid_total = round(sum((i.total_amount or 0) for i in cust_invoices if i.payment_status == "unpaid"), 2)
    held_deposits_total = round(sum((d.amount or 0) for d in deposits if d.status in ("held", "partial_refund")), 2)
    # 押金剩余可用（=收-已抵扣-已退）+ 押金累计已抵扣（用于客户档案顶部押金卡显示）
    deposits_remaining_total = round(sum(
        max(0.0, float(d.amount or 0) - float(d.applied_amount or 0) - float(d.refunded_amount or 0))
        for d in deposits if d.status in ("held", "partial_refund")
    ), 2)
    deposits_applied_total = round(sum(float(d.applied_amount or 0) for d in deposits), 2)

    # ── 协议签署任务 + 已归档 PDF ──
    consent_tasks = (
        db.query(ConsentTask)
        .filter(ConsentTask.customer_id == customer_id)
        .order_by(ConsentTask.id.desc())
        .limit(30)
        .all()
    )
    # 已签的关联 ConsentDocument（用于"医疗文书"区显示 PDF）
    signed_task_ids = [t.id for t in consent_tasks if t.status == "signed"]
    consent_docs_map = {}
    if signed_task_ids:
        for d in db.query(ConsentDocument).filter(ConsentDocument.task_id.in_(signed_task_ids)).all():
            consent_docs_map[d.task_id] = d
    consent_templates_active = (
        db.query(ConsentTemplate)
        .filter(ConsentTemplate.is_active == True)
        .order_by(ConsentTemplate.id.desc())
        .all()
    )

    return templates.TemplateResponse(
        request,
        "uk/customer.html",  # B3.1 UK 重写；旧模板暂留
        {
            "cust": cust,
            "pets": pets,
            "pet_map": pet_map,
            "active_pet": active_pet,
            "active_pet_id": active_pet_id,
            "active_tab": tab,
            # 是否可删客户档案（无任何业务记录才能删）
            "cust_blockers": _customer_blockers(db, customer_id),
            # 客户级
            "applications": applications,
            "sales_orders": cust_sales_orders,
            "cust_invoices": cust_invoices,
            # 宠物级
            "appointments": appointments,
            "visits": visits,
            "prescriptions": prescriptions,
            "exam_orders": exam_orders,
            "vaccinations": vaccinations,
            "dewormings": dewormings,
            "groomings": groomings,
            "latest_vacc_by_pet": latest_vacc_by_pet,
            "latest_deworm_by_pet": latest_deworm_by_pet,
            "weight_records": weight_records,
            "medical_docs": medical_docs,
            "pet_invoices": pet_invoices,
            "pet_sales_orders": pet_sales_orders,
            # 钱包
            "wallet_balance": wallet_balance,
            "wallet_lifetime_recharge": wallet_lifetime_recharge,
            "wallet_lifetime_consume": wallet_lifetime_consume,
            "wallet_txs": wallet_txs,
            # 套餐
            "customer_packages": customer_packages,
            "active_packages_count": active_packages_count,
            "package_products": package_products,
            "package_category_zh": _PACKAGE_CATEGORY_ZH,
            # 押金
            "deposits": deposits,
            "held_deposits_count": held_deposits_count,
            "deposit_category_zh": _DEPOSIT_CATEGORY_ZH,
            "deposit_status_zh": _DEPOSIT_STATUS_ZH,
            # 优惠券
            "coupons": coupons,
            "active_coupons_count": active_coupons_count,
            "unpaid_total": unpaid_total,
            "held_deposits_total": held_deposits_total,
            "deposits_remaining_total": deposits_remaining_total,
            "deposits_applied_total": deposits_applied_total,
            "coupon_kind_zh": _COUPON_KIND_ZH,
            "coupon_status_zh": _COUPON_STATUS_ZH,
            # 协议签署
            "consent_tasks": consent_tasks,
            "consent_templates": consent_templates_active,
            "consent_docs_map": consent_docs_map,
            # 翻译字典
            "visit_type_zh": _VISIT_TYPE_ZH,
            "so_status_zh": _SO_STATUS_ZH_LOCAL,
            "inv_status_zh": _INV_STATUS_ZH_LOCAL,
            "vacc_type_zh": _VACC_TYPE_ZH,
            "today_str": today_str,
            "soon_str": soon_str,
            "csrf_token": _get_csrf_token(request),
            "admin_store": _get_admin_store(request),
            # 品种联想（datalist）
            "breed_dogs": _BREEDS_ALL["dog"],
            "breed_cats": _BREEDS_ALL["cat"],
            "breed_exotic": _BREEDS_ALL["exotic"],
            # 驱虫弹窗用
            "deworm_items": deworm_items,
        },
    )


@app.post("/admin/customers/{customer_id}/edit")
async def admin_customer_edit(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""),
    phone: str = Form(""),
    phones_extra: str = Form(""),
    address: str = Form(""),
    notes: str = Form(""),
    source: str = Form(""),
    is_internal: str = Form(""),   # "1" = 员工内购档案，"" = 取消勾选
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    # 限店员工：客户名下若全部宠物都在其他门店，则无权编辑
    admin_store = _get_admin_store(request)
    if admin_store:
        pets_in_store = db.query(Pet).filter(
            Pet.customer_id == customer_id,
            (Pet.store == admin_store) | (Pet.store == "")
        ).count()
        total_pets = db.query(Pet).filter(Pet.customer_id == customer_id).count()
        # 完全没本店宠物 且 客户有别店宠物 → 拒绝
        if total_pets > 0 and pets_in_store == 0:
            raise HTTPException(403, "无权编辑其他门店的客户")
    cust.name = name.strip()[:120]
    cust.phone = phone.strip()[:40]
    # 备用手机号：用户可用换行 / 逗号 / 中文逗号 / 空格 / 分号分隔，统一存为 CSV
    import re as _re
    raw_extras = _re.split(r"[\s,，;；、]+", (phones_extra or "").strip())
    cleaned = []
    seen = {cust.phone.strip()}
    for p in raw_extras:
        p = p.strip()
        if p and p not in seen:
            seen.add(p)
            cleaned.append(p[:40])
    cust.phones_extra = ",".join(cleaned)[:500]
    cust.address = address.strip()[:500]
    cust.notes = notes.strip()
    # source 只在传了非空值时更新，避免历史"老系统导入"被清空
    new_source = source.strip()[:40]
    if new_source:
        cust.source = new_source
    # 员工内购档案标记：仅超管可改，普通员工传也忽略
    if request.session.get("admin_role") == "superadmin":
        new_internal = (is_internal == "1")
        if new_internal != bool(cust.is_internal):
            cust.is_internal = new_internal
            # 首次标记为内购时自动改 source（除非用户也手填了 source）
            if new_internal and not new_source:
                cust.source = "employee_internal"
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?msg=已保存", status_code=303)


@app.post("/admin/customers/{customer_id}/pets/add")
async def admin_customer_add_pet(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""),
    species: str = Form("cat"),
    breed: str = Form(""),
    gender: str = Form("unknown"),
    birthday_estimate: str = Form(""),
    is_neutered: str = Form(""),
    color_pattern: str = Form(""),
    is_stray: str = Form(""),
    microchip_id: str = Form(""),
    notes: str = Form(""),
    store: str = Form(""),                # 短名：东环店/横岗店
    life_status: str = Form("alive"),     # alive/deceased
    next_url: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    # 限店员工：强制使用自己门店
    admin_store = _get_admin_store(request)
    if admin_store:
        store = admin_store
    store = (store or "").strip()
    mrn = _gen_medical_record_no(db, store) if store else ""
    pet = Pet(
        customer_id=customer_id,
        name=name.strip()[:120],
        species=species.strip()[:40] or "cat",
        breed=breed.strip()[:80],
        gender=gender.strip()[:10] or "unknown",
        birthday_estimate=birthday_estimate.strip()[:40],
        is_neutered=is_neutered.lower() in ("1", "true", "on", "yes"),
        color_pattern=color_pattern.strip()[:80],
        is_stray=is_stray.lower() in ("1", "true", "on", "yes"),
        microchip_id=microchip_id.strip()[:40],
        notes=notes.strip(),
        store=store,
        medical_record_no=mrn,
        life_status=(life_status or "alive").strip()[:20],
    )
    db.add(pet)
    db.commit()
    if next_url:
        target = _safe_next(
            next_url.replace("{pet_id}", str(pet.id)).replace("{id}", str(customer_id)),
            f"/admin/customers/{customer_id}",
        )
        sep = "&" if "?" in target else "?"
        return RedirectResponse(f"{target}{sep}msg=宠物已添加", status_code=303)
    return RedirectResponse(f"/admin/customers/{customer_id}?msg=宠物已添加", status_code=303)


@app.post("/admin/customers/{customer_id}/pets/{pet_id}/edit")
async def admin_customer_edit_pet(
    customer_id: int,
    pet_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""),
    species: str = Form("cat"),
    breed: str = Form(""),
    gender: str = Form("unknown"),
    birthday_estimate: str = Form(""),
    is_neutered: str = Form(""),
    color_pattern: str = Form(""),
    is_stray: str = Form(""),
    microchip_id: str = Form(""),
    notes: str = Form(""),
    store: str = Form(""),
    life_status: str = Form("alive"),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)  # 强制校验：空 token 也会拒绝
    pet = db.get(Pet, pet_id)
    if not pet or pet.customer_id != customer_id:
        raise HTTPException(404, "宠物不存在")
    pet.name = name.strip()[:120]
    pet.species = species.strip()[:40] or "cat"
    pet.breed = breed.strip()[:80]
    pet.gender = gender.strip()[:10] or "unknown"
    pet.birthday_estimate = birthday_estimate.strip()[:40]
    pet.is_neutered = is_neutered.lower() in ("1", "true", "on", "yes")
    pet.color_pattern = color_pattern.strip()[:80]
    pet.is_stray = is_stray.lower() in ("1", "true", "on", "yes")
    pet.microchip_id = microchip_id.strip()[:40]
    pet.notes = notes.strip()
    # 门店变更时，限店员工不允许跨店；新分配病历号
    admin_store = _get_admin_store(request)
    new_store = store.strip()
    if admin_store and new_store and new_store != admin_store:
        # 限店员工不允许把宠物移到其他门店
        new_store = admin_store
    if new_store and new_store != (pet.store or ""):
        pet.store = new_store
        # 若原先无病历号或门店首字母变了，重新生成
        old_letter = (pet.medical_record_no or "")[:1]
        new_letter = _STORE_INITIAL.get(new_store, "X")
        if not pet.medical_record_no or old_letter != new_letter:
            pet.medical_record_no = _gen_medical_record_no(db, new_store)
    elif not pet.store and new_store:
        pet.store = new_store

    # 保底兜底：编辑保存时若仍没有病历号 + 已设置门店 → 自动补号
    # 这样老 pet（导入或早期建档遗漏 MRN 的）只要点一次「保存修改」就能补上
    if not pet.medical_record_no and pet.store:
        pet.medical_record_no = _gen_medical_record_no(db, pet.store)

    pet.life_status = (life_status or "alive").strip()[:20]
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&msg=宠物已更新", status_code=303)


@app.post("/admin/customers/{customer_id}/pets/{pet_id}/merge-into")
async def admin_customer_merge_pet(
    customer_id: int,
    pet_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    target_id: int = Form(...),
):
    """合并宠物：把 pet_id 的所有业务关联指向 target_id，然后删除 pet_id。

    适用场景：客户取消 TNR 又重新申请，导致出现多条同名宠物档案。
    """
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    if pet_id == target_id:
        return RedirectResponse(
            f"/admin/customers/{customer_id}?err=源宠物和目标宠物不能相同",
            status_code=303,
        )
    src = db.get(Pet, pet_id)
    tgt = db.get(Pet, target_id)
    if not src or src.customer_id != customer_id:
        raise HTTPException(404, "源宠物不存在")
    if not tgt or tgt.customer_id != customer_id:
        return RedirectResponse(
            f"/admin/customers/{customer_id}?err=目标宠物 #{target_id} 不属于该客户",
            status_code=303,
        )

    # 已知所有挂 pet_id 的表，全部更新到目标
    from sqlalchemy import update as _upd
    moved = {}
    table_classes = [
        ("申请", Application),
        ("预约", Appointment),
        ("就诊", Visit),
        ("收费单", Invoice),
        ("疫苗", Vaccination),
        ("狂犬", RabiesVaccineRecord),
        ("协议", ConsentTask),
    ]
    # 试图扩展更多模型（按需懒导入）
    try:
        from app.models import Deworming
        table_classes.append(("驱虫", Deworming))
    except ImportError:
        pass
    try:
        from app.models import Prescription
        table_classes.append(("处方", Prescription))
    except ImportError:
        pass
    try:
        from app.models import ExamOrder
        table_classes.append(("检查单", ExamOrder))
    except ImportError:
        pass
    try:
        from app.models import Deposit
        table_classes.append(("押金", Deposit))
    except ImportError:
        pass
    try:
        from app.models import MediaFile
        table_classes.append(("素材", MediaFile))
    except ImportError:
        pass
    try:
        from app.models import WeightRecord
        table_classes.append(("体重", WeightRecord))
    except ImportError:
        pass

    for label, cls in table_classes:
        if not hasattr(cls, "pet_id"):
            continue
        try:
            cnt = db.query(cls).filter(cls.pet_id == pet_id).update(
                {cls.pet_id: target_id}, synchronize_session=False
            )
            if cnt:
                moved[label] = cnt
        except Exception as e:
            logger.warning("[merge pet] move %s failed: %s", label, e)

    # 补全目标宠物字段（源有目标没有）
    for fld in ("breed", "color_pattern", "birthday_estimate", "microchip_id", "notes"):
        if not getattr(tgt, fld, None) and getattr(src, fld, None):
            setattr(tgt, fld, getattr(src, fld))
    if not tgt.gender or tgt.gender == "unknown":
        if src.gender and src.gender != "unknown":
            tgt.gender = src.gender

    src_label = f"{src.name}(#{src.id})"
    tgt_label = f"{tgt.name}(#{tgt.id})"
    db.delete(src)
    _audit(db, request, "pet_merge",
           detail={"src": src_label, "target": tgt_label, "moved": moved})
    db.commit()
    summary = "，".join([f"{k} {v}" for k, v in moved.items()]) or "无关联"
    return RedirectResponse(
        f"/admin/customers/{customer_id}?msg=已合并 {src_label} → {tgt_label}（迁移：{summary}）",
        status_code=303,
    )


@app.post("/admin/customers/{customer_id}/pets/{pet_id}/delete")
async def admin_customer_delete_pet(
    customer_id: int,
    pet_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """删除宠物 — 仅在没有任何业务关联记录时允许。

    检查：Visit / Appointment / Invoice / Vaccination / RabiesVaccineRecord / Prescription
    如果有任意一条，拒绝并返回错误提示（保护业务数据）。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    pet = db.get(Pet, pet_id)
    if not pet or pet.customer_id != customer_id:
        raise HTTPException(404, "宠物不存在")

    # 统计关联记录
    blockers = []
    n_visits = db.query(Visit).filter(Visit.pet_id == pet_id).count()
    if n_visits: blockers.append(f"{n_visits} 条病历")
    n_appts = db.query(Appointment).filter(Appointment.pet_id == pet_id).count()
    if n_appts: blockers.append(f"{n_appts} 条预约")
    n_invoices = db.query(Invoice).filter(Invoice.pet_id == pet_id).count()
    if n_invoices: blockers.append(f"{n_invoices} 张收费单")
    n_vacc = db.query(Vaccination).filter(Vaccination.pet_id == pet_id).count()
    if n_vacc: blockers.append(f"{n_vacc} 条疫苗记录")
    try:
        n_rabies = db.query(RabiesVaccineRecord).filter(RabiesVaccineRecord.pet_id == pet_id).count()
        if n_rabies: blockers.append(f"{n_rabies} 条狂犬登记")
    except Exception:
        pass
    n_presc = db.query(Prescription).filter(Prescription.pet_id == pet_id).count()
    if n_presc: blockers.append(f"{n_presc} 张处方")

    if blockers:
        from urllib.parse import quote as _q
        msg = f"该宠物有关联记录（{' / '.join(blockers)}），不允许删除。如确需清理请先处理对应记录。"
        return RedirectResponse(
            f"/admin/customers/{customer_id}?pet_id={pet_id}&msg={_q(msg, safe='')}",
            status_code=303,
        )

    pet_name = pet.name or "未命名"
    db.delete(pet)
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{customer_id}?msg=已删除宠物「{pet_name}」",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# 客户钱包 (Wallet) — 充值 / 消费 / 退款 / 调账
# ---------------------------------------------------------------------------

_WALLET_TX_TYPE_ZH = {
    "recharge": "充值",
    "consume":  "消费",
    "refund":   "退款",
    "adjust":   "调账",
}


def _get_or_create_wallet(db: Session, customer_id: int) -> Wallet:
    """取或建客户钱包，确保单例。"""
    w = db.query(Wallet).filter(Wallet.customer_id == customer_id).first()
    if w:
        return w
    w = Wallet(customer_id=customer_id, balance=0.0)
    db.add(w)
    db.flush()
    return w


def _wallet_apply_tx(
    db: Session,
    wallet: Wallet,
    *,
    tx_type: str,
    amount: float,
    bonus: float = 0.0,
    pay_method: str = "",
    invoice_id: int | None = None,
    operator: str = "",
    store: str = "",
    note: str = "",
) -> WalletTransaction:
    """对钱包施加一笔流水。amount 正/负由 tx_type 决定：
      recharge → balance + (amount + bonus)，lifetime_recharge += amount
      consume  → balance - amount，lifetime_consume += amount
      refund   → balance - amount（把余额退还客户）
      adjust   → balance += amount（amount 可正可负）
    返回 WalletTransaction 行，未 commit。
    """
    amt = float(amount or 0)
    bns = float(bonus or 0)
    # 兜底：旧数据 balance_principal/bonus 还没拆 → 当前余额全部归本金
    if (wallet.balance_principal or 0) == 0 and (wallet.balance_bonus or 0) == 0 and (wallet.balance or 0) > 0:
        wallet.balance_principal = float(wallet.balance)
        wallet.balance_bonus = 0.0
    consumed_p = 0.0
    consumed_b = 0.0
    if tx_type == "recharge":
        delta = amt + bns
        wallet.balance += delta
        wallet.balance_principal = (wallet.balance_principal or 0) + amt
        wallet.balance_bonus = (wallet.balance_bonus or 0) + bns
        wallet.lifetime_recharge += amt
        signed = delta  # 正
    elif tx_type == "consume":
        if amt > wallet.balance + 1e-6:
            raise HTTPException(400, f"余额不足：当前 ¥{wallet.balance:.2f}，需扣 ¥{amt:.2f}")
        # 按本金:赠送 比例同步扣（用户确认的方案）
        bp = float(wallet.balance_principal or 0)
        bb = float(wallet.balance_bonus or 0)
        total = bp + bb
        if total > 0:
            ratio_p = bp / total
            consumed_p = round(amt * ratio_p, 2)
            consumed_b = round(amt - consumed_p, 2)
            # 兜底防越界（浮点误差导致负数）
            if consumed_p > bp:
                consumed_p = bp
                consumed_b = amt - consumed_p
            if consumed_b > bb:
                consumed_b = bb
                consumed_p = amt - consumed_b
        else:
            consumed_p = amt
            consumed_b = 0.0
        wallet.balance -= amt
        wallet.balance_principal = max(0, bp - consumed_p)
        wallet.balance_bonus = max(0, bb - consumed_b)
        wallet.lifetime_consume += amt
        signed = -amt
    elif tx_type == "refund":
        if amt > wallet.balance + 1e-6:
            raise HTTPException(400, f"退款金额超过当前余额（¥{wallet.balance:.2f}）")
        # 退款优先扣本金（实付的钱），不退赠送
        bp = float(wallet.balance_principal or 0)
        if amt <= bp:
            consumed_p = amt
        else:
            consumed_p = bp
            # 超出本金部分按比例从赠送扣（极少见）
            consumed_b = amt - bp
        wallet.balance -= amt
        wallet.balance_principal = max(0, bp - consumed_p)
        wallet.balance_bonus = max(0, (wallet.balance_bonus or 0) - consumed_b)
        signed = -amt
    elif tx_type == "adjust":
        # amount 可正可负 — 调账走本金
        new_bal = wallet.balance + amt
        if new_bal < -1e-6:
            raise HTTPException(400, "调账后余额不能为负")
        wallet.balance = new_bal
        if amt >= 0:
            wallet.balance_principal = (wallet.balance_principal or 0) + amt
        else:
            # 调减：先扣本金，不够再扣赠送
            need = -amt
            bp = float(wallet.balance_principal or 0)
            take_p = min(need, bp)
            consumed_p = take_p
            wallet.balance_principal = bp - take_p
            need -= take_p
            if need > 0:
                bb = float(wallet.balance_bonus or 0)
                wallet.balance_bonus = max(0, bb - need)
                consumed_b = need
        signed = amt
    else:
        raise HTTPException(400, f"未知流水类型：{tx_type}")

    wallet.updated_at = datetime.utcnow()
    tx = WalletTransaction(
        wallet_id=wallet.id,
        customer_id=wallet.customer_id,
        type=tx_type,
        amount=signed,
        balance_after=wallet.balance,
        pay_method=pay_method or "",
        invoice_id=invoice_id,
        bonus_amount=bns if tx_type == "recharge" else 0.0,
        consumed_principal=consumed_p,
        consumed_bonus=consumed_b,
        store=store or "",
        operator=operator or "",
        note=(note or "")[:500],
    )
    db.add(tx)
    db.flush()
    return tx


@app.post("/admin/wallets/{customer_id}/recharge")
async def admin_wallet_recharge(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    amount: str = Form(...),
    bonus: str = Form("0"),
    note: str = Form(""),
):
    """客户钱包充值 → 生成一张未付收费单进收银台。
    收银台收款完成后，对应的钱包余额才会真正到账（add-payment 钩子里触发）。
    这样收款方式、对账、报表都走统一的 Invoice/Payment 链路。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    try:
        amt = float(amount)
        bns = float(bonus or "0")
    except (TypeError, ValueError):
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=金额无效", status_code=303)
    if amt <= 0:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=充值金额需大于 0", status_code=303)
    if bns < 0:
        bns = 0.0

    admin_name = request.session.get("admin_username", "admin")
    # 生成"钱包充值"收费单（未付）
    desc = f"钱包充值 ¥{amt:.2f}"
    if bns > 0:
        desc += f"（赠送 ¥{bns:.2f}）"
    notes_payload = (note.strip() + " " if note.strip() else "") + f"[wallet_recharge_bonus={bns:.2f}]"
    inv = Invoice(
        invoice_no      = _gen_invoice_no(db),
        customer_id     = customer_id,
        pet_id          = None,
        invoice_date    = datetime.now().strftime("%Y-%m-%d"),
        subtotal        = amt,
        discount_amount = 0.0,
        total_amount    = amt,
        payment_status  = "unpaid",
        notes           = notes_payload,
        store           = _get_op_store(request) or "",
        created_by      = admin_name,
    )
    db.add(inv)
    db.flush()
    db.add(InvoiceItem(
        invoice_id  = inv.id,
        ref_type    = "wallet_recharge",
        ref_id      = customer_id,
        description = desc,
        quantity    = 1.0,
        unit_price  = amt,
        subtotal    = amt,
    ))
    _audit(db, request, "wallet_recharge_invoice_create", application_id=None,
           detail={"customer_id": customer_id, "amount": amt, "bonus": bns, "invoice_id": inv.id})
    db.commit()
    # 跳到收费单收款页，立即可以收款
    return RedirectResponse(
        f"/admin/invoices/{inv.id}?msg=已生成充值单 ¥{amt:.2f}" + (f"（送 ¥{bns:.2f}）" if bns > 0 else "") + "，请选择收款方式",
        status_code=303,
    )


def _maybe_credit_wallet_from_invoice(db: Session, inv: "Invoice", request: Request) -> None:
    """收费单付清后：如果含 ref_type='wallet_recharge' 项目，把钱真正打进钱包。
    用 WalletTransaction.invoice_id 做幂等：已经有同 invoice_id 的 recharge tx 就不再加。"""
    if inv.payment_status != "paid":
        return
    if not inv.customer_id:
        return
    has_recharge_item = db.query(InvoiceItem.id).filter(
        InvoiceItem.invoice_id == inv.id,
        InvoiceItem.ref_type == "wallet_recharge",
    ).first()
    if not has_recharge_item:
        return
    existed = db.query(WalletTransaction.id).filter(
        WalletTransaction.invoice_id == inv.id,
        WalletTransaction.type == "recharge",
    ).first()
    if existed:
        return
    # 从 notes 解析 bonus
    bonus = 0.0
    import re as _re
    m = _re.search(r"\[wallet_recharge_bonus=([\d.]+)\]", inv.notes or "")
    if m:
        try: bonus = float(m.group(1))
        except: bonus = 0.0
    wallet = _get_or_create_wallet(db, inv.customer_id)
    _wallet_apply_tx(
        db, wallet, tx_type="recharge",
        amount=float(inv.total_amount or 0),
        bonus=bonus,
        invoice_id=inv.id,
        operator=request.session.get("admin_username", "admin"),
        store=_get_admin_store(request),
        note=f"充值单 {inv.invoice_no or inv.id}",
    )


@app.post("/admin/wallets/{customer_id}/refund")
async def admin_wallet_refund(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    amount: str = Form(...),
    note: str = Form(""),
):
    """钱包退款（把余额退给客户）。amount > 0。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)  # 退款仅超管
    try:
        amt = float(amount)
    except (TypeError, ValueError):
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=金额无效", status_code=303)
    if amt <= 0:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=退款金额需大于 0", status_code=303)
    wallet = _get_or_create_wallet(db, customer_id)
    try:
        _wallet_apply_tx(
            db, wallet, tx_type="refund", amount=amt,
            operator=request.session.get("admin_username", "admin"),
            store=_get_admin_store(request),
            note=note,
        )
    except HTTPException as he:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg={he.detail}", status_code=303)
    db.commit()
    _audit(db, request, "wallet_refund", application_id=None,
           detail={"customer_id": customer_id, "amount": amt})
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=已退款 ¥{amt:.2f}", status_code=303)


@app.post("/admin/wallets/{customer_id}/adjust")
async def admin_wallet_adjust(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    amount: str = Form(...),
    note: str = Form(""),
):
    """钱包调账（正可加、负可扣，需备注）。仅超管。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    if not (note or "").strip():
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=调账必须填备注", status_code=303)
    try:
        amt = float(amount)
    except (TypeError, ValueError):
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=金额无效", status_code=303)
    if amt == 0:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=调账金额不能为 0", status_code=303)
    wallet = _get_or_create_wallet(db, customer_id)
    try:
        _wallet_apply_tx(
            db, wallet, tx_type="adjust", amount=amt,
            operator=request.session.get("admin_username", "admin"),
            store=_get_admin_store(request),
            note=note,
        )
    except HTTPException as he:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg={he.detail}", status_code=303)
    db.commit()
    _audit(db, request, "wallet_adjust", application_id=None,
           detail={"customer_id": customer_id, "amount": amt, "note": note})
    db.commit()
    sign = "+" if amt > 0 else ""
    return RedirectResponse(f"/admin/customers/{customer_id}?tab=wallet&msg=调账 {sign}¥{amt:.2f}", status_code=303)


# ---------------------------------------------------------------------------
# 协议签署 (Consent) — 模板 + 任务 + PDF 归档
# ---------------------------------------------------------------------------

_CONSENT_CATEGORY_ZH = {
    "anesthesia":    "麻醉知情同意书",
    "surgery":       "手术知情同意书",
    "vaccination":   "疫苗接种同意书",
    "euthanasia":    "安乐死同意书",
    "boarding":      "寄养协议",
    "transfusion":   "输血同意书",
    "general":       "通用协议",
}

_CONSENT_STATUS_ZH = {
    "pending":   "待签署",
    "signed":    "已签署",
    "cancelled": "已取消",
    "expired":   "已过期",
}

# 模板里支持的变量
_CONSENT_VARIABLES = {
    "{{cust_name}}":  "客户姓名",
    "{{cust_phone}}": "客户手机",
    "{{pet_name}}":   "宠物名",
    "{{pet_species}}":"宠物种类",
    "{{pet_breed}}":  "品种",
    "{{pet_gender}}": "宠物性别",
    "{{pet_age}}":    "宠物年龄",
    "{{pet_weight}}": "宠物体重",
    "{{visit_date}}": "就诊日期",
    "{{vet_name}}":   "主治医师",
    "{{date}}":       "今日日期",
    "{{clinic_name}}":"门店名称",
}


def _gen_consent_token() -> str:
    import secrets
    return secrets.token_urlsafe(18)[:32]


@app.get("/admin/consent-templates", response_class=HTMLResponse)
async def admin_consent_templates_list(request: Request, db: Session = Depends(get_db)):
    """协议模板管理（列表 / 启停）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    items = db.query(ConsentTemplate).order_by(
        ConsentTemplate.is_active.desc(), ConsentTemplate.id.desc()
    ).all()
    # 各模板被使用次数（签署任务计数）
    from sqlalchemy import func as _f
    used_rows = (
        db.query(ConsentTask.template_id, _f.count(ConsentTask.id))
        .group_by(ConsentTask.template_id)
        .all()
    )
    used_map = {tid: cnt for tid, cnt in used_rows if tid}
    return templates.TemplateResponse(request, "uk/consent_templates.html", {
        "items": items, "used_map": used_map,
        "category_zh": _CONSENT_CATEGORY_ZH,
        "variables": _CONSENT_VARIABLES,
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/consent-templates/create", response_class=HTMLResponse)
@app.get("/admin/consent-templates/{tid}/edit", response_class=HTMLResponse)
async def admin_consent_template_form(
    request: Request, db: Session = Depends(get_db), tid: int = 0,
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    require_superadmin(request)
    item = db.get(ConsentTemplate, tid) if tid else None
    return templates.TemplateResponse(request, "uk/consent_template_form.html", {
        "item": item,
        "category_zh": _CONSENT_CATEGORY_ZH,
        "variables": _CONSENT_VARIABLES,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/consent-templates/save")
async def admin_consent_template_save(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    template_id: str = Form(""),
    name: str = Form(...),
    category: str = Form("general"),
    body_html: str = Form(""),
    notes: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    if category not in _CONSENT_CATEGORY_ZH:
        category = "general"
    name_v = (name or "").strip()[:120]
    if not name_v:
        return RedirectResponse("/admin/consent-templates?msg=模板名必填", status_code=303)
    tid = int(template_id) if template_id.isdigit() else 0
    if tid:
        item = db.get(ConsentTemplate, tid)
        if not item:
            raise HTTPException(404)
        item.name = name_v
        item.category = category
        item.body_html = body_html or ""
        item.notes = (notes or "").strip()
    else:
        item = ConsentTemplate(
            name=name_v, category=category,
            body_html=body_html or "", notes=(notes or "").strip(),
            created_by=request.session.get("admin_username", "admin"),
            is_active=True,
        )
        db.add(item)
    db.commit()
    return RedirectResponse(
        f"/admin/consent-templates?msg={'已保存' if tid else '已创建'}：{item.name}",
        status_code=303,
    )


def _consent_render_snapshot(template_body: str, *, cust=None, pet=None, visit=None,
                              vet_name="", clinic_name="", pet_weight=0.0, pet_age="") -> str:
    """把模板里的 {{变量}} 替换成实际值。HTML 安全，不再做 escape（Quill 已经是 HTML）。"""
    from datetime import date as _date
    vals = {
        "{{cust_name}}":  (cust.name if cust else ""),
        "{{cust_phone}}": (cust.phone if cust else ""),
        "{{pet_name}}":   (pet.name if pet else ""),
        "{{pet_species}}": ({"cat":"猫","dog": "犬"}.get(pet.species, pet.species) if pet else ""),
        "{{pet_breed}}":  (pet.breed if pet else ""),
        "{{pet_gender}}": ({"male":"公","female":"母","unknown":"未知"}.get(pet.gender, "") if pet else ""),
        "{{pet_age}}":    (pet_age or ""),
        "{{pet_weight}}": (f"{pet_weight:.2f}" if pet_weight else ""),
        "{{visit_date}}": (visit.visit_date if visit else ""),
        "{{vet_name}}":   (vet_name or ""),
        "{{date}}":       _date.today().isoformat(),
        "{{clinic_name}}": (clinic_name or "大风动物医院"),
    }
    out = template_body or ""
    for k, v in vals.items():
        out = out.replace(k, str(v or "—"))
    return out


@app.post("/admin/consent-tasks/create")
async def admin_consent_task_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    template_id: int = Form(...),
    customer_id: int = Form(...),
    pet_id: int = Form(0),
    visit_id: int = Form(0),
    title_override: str = Form(""),
    expires_at: str = Form(""),
    notes: str = Form(""),
):
    """发起一次协议签署：把模板正文 + 变量快照保存，生成唯一 token。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    tpl = db.get(ConsentTemplate, template_id)
    if not tpl or not tpl.is_active:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=docs&msg=模板不存在或已下架", status_code=303)
    cust = db.get(Customer, customer_id)
    if not cust:
        return RedirectResponse("/admin?msg=客户不存在", status_code=303)

    # 重复发起校验：1 小时内同一客户 + 同一模板 + 同一宠物 + 仍 pending 的任务
    # 防止用户连点两下生成两条僵尸任务（客户签的是后一条、前一条永远停在待签署）
    from datetime import timedelta as _td
    one_hour_ago = datetime.utcnow() - _td(hours=1)
    _dup_q = db.query(ConsentTask).filter(
        ConsentTask.customer_id == customer_id,
        ConsentTask.template_id == tpl.id,
        ConsentTask.status == "pending",
        ConsentTask.initiated_at >= one_hour_ago,
    )
    if pet_id:
        _dup_q = _dup_q.filter(ConsentTask.pet_id == pet_id)
    else:
        _dup_q = _dup_q.filter(ConsentTask.pet_id.is_(None))
    _dup = _dup_q.order_by(ConsentTask.id.desc()).first()
    if _dup:
        from urllib.parse import quote as _q
        msg = f"已存在未签的相同协议（#CT{_dup.id:06d}），请等客户先签完或在该任务详情页「取消任务」后再重发"
        return RedirectResponse(
            f"/admin/customers/{customer_id}?tab=docs&msg={_q(msg)}",
            status_code=303,
        )
    pet = db.get(Pet, pet_id) if pet_id else None
    visit = db.get(Visit, visit_id) if visit_id else None
    # 取宠物体重 / 年龄
    pet_weight = 0.0
    if pet:
        last_w = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id).order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if last_w:
            pet_weight = float(last_w.weight_kg or 0)
    pet_age = ""
    if pet and pet.birthday_estimate:
        try:
            from datetime import date as _date
            y, m, _ = (pet.birthday_estimate + "-01").split("-")[:3]
            today = _date.today()
            years = today.year - int(y) - (1 if today.month < int(m) else 0)
            pet_age = f"{years} 岁" if years > 0 else f"{max(0, (today.year-int(y))*12 + (today.month-int(m)))} 个月"
        except Exception:
            pet_age = pet.birthday_estimate or ""
    # 渲染快照
    vet_name = (visit.vet_name if visit else "") or (request.session.get("admin_username", ""))
    clinic_name = "大风动物医院"
    if pet and pet.store:
        clinic_name = f"大风动物医院（{pet.store.replace('店', '分院')}）"
    snapshot = _consent_render_snapshot(
        tpl.body_html, cust=cust, pet=pet, visit=visit,
        vet_name=vet_name, clinic_name=clinic_name,
        pet_weight=pet_weight, pet_age=pet_age,
    )
    task = ConsentTask(
        template_id=tpl.id,
        customer_id=customer_id,
        pet_id=pet_id or None,
        visit_id=visit_id or None,
        title=(title_override.strip() or tpl.name)[:120],
        snapshot_html=snapshot,
        token=_gen_consent_token(),
        status="pending",
        expires_at=(expires_at or "").strip()[:20],
        store=_get_admin_store(request),
        initiated_by=request.session.get("admin_username", "admin"),
        notes=(notes or "").strip(),
    )
    db.add(task); db.commit(); db.refresh(task)
    _audit(db, request, "consent_task_create", application_id=None,
           detail={"task_id": task.id, "template": tpl.name, "customer_id": customer_id})
    db.commit()
    # 自动发短信给客户（有手机号才发；SMS 网关未配 → 静默跳过）
    sms_ok = _try_send_consent_sms(db, task, cust, pet)
    suffix = "并已短信发送签字链接" if sms_ok else "（请手动复制链接发给客户）"
    return RedirectResponse(f"/admin/consent-tasks/{task.id}?msg=已发起签署{suffix}", status_code=303)


def _build_consent_sign_url(token: str) -> str:
    base = (settings.public_base_url or "").strip().rstrip("/")
    return f"{base}/consent/{token}" if base else f"/consent/{token}"


def _try_send_consent_sms(db: Session, task: "ConsentTask", cust: "Customer | None", pet: "Pet | None") -> bool:
    """给客户发短信（含签字链接）。
    优先腾讯云直连，回退到通用网关；均未配 / 无手机号 → 返回 False（静默）。
    """
    if not cust or not (cust.phone or "").strip():
        return False
    pet_name = (pet.name if pet else "") or "您的宝贝"
    phone = cust.phone.strip()

    # 路径 1：腾讯云直连
    if (settings.tencent_sms_tmpl_consent or "").strip():
        from app.services.sms_tencent import send_sms_template, _enabled as _tc_enabled
        if _tc_enabled():
            # 模板参数顺序见 settings.tencent_sms_tmpl_consent 注释：
            #   1=宠物名, 2=协议标题, 3=token
            ok, err = send_sms_template(
                phone,
                settings.tencent_sms_tmpl_consent.strip(),
                [pet_name[:10], (task.title or "诊疗协议")[:14], task.token],
            )
            if ok:
                return True
            logger.warning("[consent] 腾讯云短信失败 task=%s: %s", task.id, err)

    # 路径 2：通用 HTTP 网关（自建/其他供应商）
    if (settings.sms_gateway_url or "").strip():
        sign_url = _build_consent_sign_url(task.token)
        text = (
            f"【大风动物医院】关于{pet_name}的{task.title}请尽快签署：{sign_url}"
            f" 如有疑问请联系您的主治医师。"
        )
        try:
            from app.services.sms_gateway import send_sms
            return send_sms(phone, text, scene="consent")
        except Exception as e:
            logger.warning("[consent] 通用网关 SMS 失败 task=%s: %s", task.id, e)
    return False


# 短链：/c/{token} → 跳协议签字页（短信里链接更短）
@app.get("/c/{token}")
async def consent_short_redirect(token: str):
    return RedirectResponse(f"/consent/{token}", status_code=302)


def _try_push_consent_notice(db: Session, task: "ConsentTask", cust: "Customer | None", pet: "Pet | None") -> bool:
    """尝试给客户推送小程序订阅消息。无 openid / 模板未配 → 返回 False（静默）。"""
    if not cust:
        return False
    openid = (cust.wechat_openid or "").strip()
    if not openid and cust.phone:
        # 兜底：按手机号在 Application 表里找历史 openid
        app_row = (
            db.query(Application)
            .filter(Application.phone == cust.phone, Application.wechat_openid != "")
            .order_by(Application.id.desc())
            .first()
        )
        if app_row and app_row.wechat_openid:
            openid = app_row.wechat_openid.strip()
    if not openid:
        return False
    clinic_name = "大风动物医院"
    if pet and pet.store:
        clinic_name = f"大风动物医院（{pet.store.replace('店', '分院')}）"
    base_url = (settings.public_base_url or "").strip().rstrip("/")
    sign_url = f"{base_url}/consent/{task.token}" if base_url else f"/consent/{task.token}"
    initiated_at = task.initiated_at.strftime("%Y-%m-%d %H:%M") if task.initiated_at else ""
    try:
        from app.services.wechat_miniapp import push_consent_signature
        return push_consent_signature(
            db, openid=openid,
            cust_name=cust.name or "客户",
            clinic_name=clinic_name,
            title=task.title or "诊疗协议",
            initiated_at=initiated_at,
            sign_url=sign_url,
            customer_id=cust.id,
        )
    except Exception as e:
        logger.warning("[consent] 推送小程序通知失败 task=%s: %s", task.id, e)
        return False


@app.post("/admin/consent-tasks/{task_id}/resend")
async def admin_consent_task_resend(
    task_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    next_url: str = Form(""),
):
    """手动重发签字链接短信。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    _base = _safe_next(next_url, f"/admin/consent-tasks/{task_id}")
    def _r(m):
        return RedirectResponse(f"{_base}{'&' if '?' in _base else '?'}msg={m}", status_code=303)
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    if task.status != "pending":
        return _r("仅待签状态可重发")
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    pet = db.get(Pet, task.pet_id) if task.pet_id else None
    if not cust or not (cust.phone or "").strip():
        return _r("客户无手机号，请直接复制链接微信发送")
    has_tencent = bool((settings.tencent_sms_tmpl_consent or "").strip())
    has_gateway = bool((settings.sms_gateway_url or "").strip())
    if not (has_tencent or has_gateway):
        return _r("未配置短信通道，请直接复制链接微信发送")
    ok = _try_send_consent_sms(db, task, cust, pet)
    return _r("已发送短信" if ok else "短信发送失败，请稍后重试或复制链接微信发")


@app.get("/admin/consent-tasks/{task_id}", response_class=HTMLResponse)
async def admin_consent_task_detail(
    task_id: int, request: Request, db: Session = Depends(get_db),
):
    """任务详情：展示签署链接 + 状态 + 快照预览 + 复制链接。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    pet = db.get(Pet, task.pet_id) if task.pet_id else None
    doc = db.query(ConsentDocument).filter(ConsentDocument.task_id == task_id).first()
    from app.models import ConsentAuditLog
    audit_logs = (
        db.query(ConsentAuditLog)
        .filter(ConsentAuditLog.task_id == task_id)
        .order_by(ConsentAuditLog.created_at.asc())
        .all()
    )
    return templates.TemplateResponse(request, "uk/consent_task_detail.html", {
        "task": task, "cust": cust, "pet": pet, "doc": doc,
        "status_zh": _CONSENT_STATUS_ZH,
        "category_zh": _CONSENT_CATEGORY_ZH,
        "csrf_token": _get_csrf_token(request),
        "audit_logs": audit_logs,
        "sms_mode_enabled": _consent_sms_enabled(),
    })


# ─── 客户档案绑定（小程序输入手机号 + 验证码 → 写 openid） ──
# 内存验证码池：{phone: (code, expires_ts, customer_id)}
_BIND_CODES: dict[str, tuple[str, float, int]] = {}
_BIND_CODE_TTL_SECONDS = 300  # 5 分钟
_BIND_LAST_SENT: dict[str, float] = {}  # 防短时间内重复发送
_BIND_THROTTLE_SECONDS = 60


@app.post("/api/customer-binding/send-code")
async def api_binding_send_code(request: Request, db: Session = Depends(get_db)):
    """根据手机号在 Customer 表查档案 + 发验证码。
    返回 {ok, customer_summary, dev_code?}。dev_code 仅 sms_gateway 未配时返回。"""
    body = await request.json()
    phone = (body.get("phone") or "").strip()
    if not phone or not phone.isdigit() or len(phone) != 11:
        return {"ok": False, "error": "请输入 11 位手机号"}
    # 防刷
    import time as _t
    now = _t.time()
    last = _BIND_LAST_SENT.get(phone, 0)
    if now - last < _BIND_THROTTLE_SECONDS:
        wait = int(_BIND_THROTTLE_SECONDS - (now - last))
        return {"ok": False, "error": f"请 {wait} 秒后再获取"}
    # 查档案
    rows = db.query(Customer).filter(Customer.phone == phone).all()
    if not rows:
        return {"ok": False, "error": "未找到此手机号对应的客户档案，请先到院前台建档"}
    # 取最早的一条（防止多账号合并历史）
    cust = sorted(rows, key=lambda c: c.id)[0]
    pet_count = db.query(Pet).filter(Pet.customer_id == cust.id).count()
    # 生成 6 位验证码
    import secrets as _s
    code = "".join(_s.choice("0123456789") for _ in range(6))
    _BIND_CODES[phone] = (code, now + _BIND_CODE_TTL_SECONDS, cust.id)
    _BIND_LAST_SENT[phone] = now
    # 发短信：优先腾讯云直连（已配好的验证码模板），失败兜底 sms_gateway 通用网关
    sent = False
    err_detail = ""
    # 1) 腾讯云
    try:
        from app.services.sms_tencent import send_sms_template, _enabled as _tc_enabled
        if _tc_enabled() and (settings.tencent_sms_tmpl_consent or "").strip():
            ttl_min = max(1, _BIND_CODE_TTL_SECONDS // 60)
            sent, err_detail = send_sms_template(
                phone,
                settings.tencent_sms_tmpl_consent.strip(),
                [code, str(ttl_min)],
            )
            if not sent:
                logger.warning("[binding] 腾讯云发送失败：%s", err_detail)
    except Exception as _e:
        logger.warning("[binding] 腾讯云调用异常：%s", _e)
    # 2) 兜底：通用 sms_gateway
    if not sent:
        try:
            from app.services.sms_gateway import send_sms
            sent = send_sms(
                phone,
                f"【大风动物医院】您的档案绑定验证码：{code}，5 分钟内有效。如非本人操作请忽略。",
                scene="binding",
            )
        except Exception as _e:
            logger.warning("[binding] sms_gateway 发送失败：%s", _e)
    resp = {
        "ok": True,
        "customer": {
            "id": cust.id, "name": cust.name or "—",
            "pet_count": pet_count,
            "address": (cust.address or "")[:40],
        },
        "sms_sent": sent,
    }
    if not sent:
        # SMS 未发出（未配网关或网关错误）→ 把 code 返回给前端，并提示走人工核对
        resp["dev_code"] = code
        resp["dev_warning"] = "短信未发送，仅自助测试用；上线前必须配 SMS_GATEWAY_URL"
    return resp


@app.post("/api/customer-binding/verify")
async def api_binding_verify(request: Request, db: Session = Depends(get_db)):
    """校验验证码 + 写 Customer.wechat_openid。需要 js_code（用 wx.login 换 openid）。"""
    body = await request.json()
    phone = (body.get("phone") or "").strip()
    code  = (body.get("code") or "").strip()
    openid = (body.get("openid") or "").strip()
    js_code = (body.get("js_code") or "").strip()
    if not (phone and code):
        return {"ok": False, "error": "手机号 + 验证码必填"}
    # 验证码核对
    import time as _t
    entry = _BIND_CODES.get(phone)
    if not entry:
        return {"ok": False, "error": "请先获取验证码"}
    saved_code, exp_ts, cust_id = entry
    if _t.time() > exp_ts:
        _BIND_CODES.pop(phone, None)
        return {"ok": False, "error": "验证码已过期，请重新获取"}
    if code != saved_code:
        return {"ok": False, "error": "验证码不正确"}
    # 没传 openid → 用 js_code 换
    if not openid and js_code:
        try:
            sess_data = wechat_code2session(js_code)
            openid = (sess_data.get("openid") or "").strip()
        except Exception as e:
            logger.warning("[binding] code2session 失败：%s", e)
            return {"ok": False, "error": f"微信登录失败：{e}"}
    if not openid:
        return {"ok": False, "error": "缺少微信登录凭证（openid 或 js_code）"}
    cust = db.get(Customer, cust_id)
    if not cust:
        return {"ok": False, "error": "客户档案不存在"}
    cust.wechat_openid = openid[:64]
    db.commit()
    _BIND_CODES.pop(phone, None)
    _audit(db, request, "customer_binding", application_id=None,
           detail={"customer_id": cust_id, "phone": phone, "openid": openid[:10] + "..."})
    db.commit()
    return {
        "ok": True,
        "customer_id": cust_id,
        "customer_name": cust.name or "",
    }


# ─── 客户端签署（无登录，token 即凭证） ──────────────────
def _consent_phone_match(phone_input: str, customer_phone: str) -> bool:
    """对比手机号（去除空格/破折号/+86 前缀）。"""
    import re
    a = re.sub(r'\D', '', phone_input or '')
    b = re.sub(r'\D', '', customer_phone or '')
    if not a or not b:
        return False
    if a.startswith('86') and len(a) > 11: a = a[2:]
    if b.startswith('86') and len(b) > 11: b = b[2:]
    return a == b


def _is_consent_verified(request: Request, token: str) -> bool:
    return bool(request.session.get(f"consent_verified_{token}"))


def _phone_mask(phone: str) -> str:
    """脱敏手机号：138****1234"""
    if not phone: return ""
    p = ''.join(ch for ch in phone if ch.isdigit())
    if len(p) >= 7: return p[:3] + "****" + p[-4:]
    if len(p) >= 4: return "****" + p[-4:]
    return ""


def _sha256(data) -> str:
    """计算 SHA256 hex（接受 str 或 bytes）。"""
    import hashlib
    if isinstance(data, str): data = data.encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def _log_consent_audit(
    db: "Session", request: "Request", task_id: int, event: str,
    *, phone: str = "", note: str = "",
    doc_sha256: str = "", sig_sha256: str = "",
) -> None:
    """追加一条签署审计日志。永远不抛异常（日志失败不能阻断业务）。"""
    try:
        from app.models import ConsentAuditLog
        ip = ""
        try:
            ip = (request.client.host if request.client else "")[:60]
            # 兼容反代：优先取 X-Forwarded-For
            xff = request.headers.get("x-forwarded-for") or ""
            if xff:
                ip = xff.split(",")[0].strip()[:60]
        except Exception:
            pass
        ua = (request.headers.get("user-agent") or "")[:500]
        # session_id 不直接存（隐私），存它的 SHA256 — 后续可以核对"同一个 session"但不能反查
        session_id = (request.cookies.get("session") or "")[:200]
        session_hash = _sha256(session_id) if session_id else ""
        db.add(ConsentAuditLog(
            task_id=task_id, event=event,
            ip=ip, user_agent=ua,
            phone_masked=_phone_mask(phone),
            doc_sha256=doc_sha256, sig_sha256=sig_sha256,
            session_hash=session_hash,
            note=note[:2000],
        ))
        db.commit()
    except Exception as _e:
        logger.warning("[consent-audit] 日志写入失败 task=%s event=%s: %s", task_id, event, _e)


# ─── 协议签字 SMS 验证码（中期方案，等腾讯云模板审核通过自动启用） ──
_CONSENT_CODES: dict[str, tuple[str, float]] = {}  # {token: (code, expires_ts)}
_CONSENT_CODE_TTL = 300       # 5 分钟
_CONSENT_SEND_THROTTLE = 60   # 同 token 60 秒内不可重发
_CONSENT_LAST_SENT: dict[str, float] = {}


def _consent_sms_enabled() -> bool:
    """SMS 验证码模式是否启用 — 需腾讯云模板 ID 已配置。"""
    return bool(
        (settings.tencent_sms_tmpl_consent or "").strip()
        and (settings.tencent_sms_secret_id or "").strip()
    )


@app.get("/consent/{token}", response_class=HTMLResponse)
async def consent_sign_page(token: str, request: Request, db: Session = Depends(get_db)):
    task = db.query(ConsentTask).filter(ConsentTask.token == token).first()
    if not task:
        raise HTTPException(404, "协议链接不存在或已失效")
    # 过期检测
    if task.expires_at:
        from datetime import date as _date
        try:
            y, m, d = task.expires_at.split("-")
            if _date(int(y), int(m), int(d)) < _date.today() and task.status == "pending":
                task.status = "expired"
                db.commit()
        except Exception:
            pass
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    pet = db.get(Pet, task.pet_id) if task.pet_id else None
    verified = _is_consent_verified(request, token) or task.status != "pending"
    has_phone = bool(cust and cust.phone and cust.phone.strip())
    phone_hint = _phone_mask(cust.phone) if cust else ""
    # 审计日志：仅 pending 状态记 link_opened，避免每次刷新都记
    if task.status == "pending":
        _log_consent_audit(db, request, task.id, "link_opened",
                           note=f"verified={verified}")
    return templates.TemplateResponse(request, "consent_sign.html", {
        "task": task, "cust": cust, "pet": pet,
        "title": task.title or "协议签署",
        "verified": verified,
        "has_phone": has_phone,
        "phone_hint": phone_hint,
        "sms_mode": _consent_sms_enabled(),  # 是否走短信验证码模式
    })


@app.post("/consent/{token}/send-code")
async def consent_send_code(
    token: str, request: Request, db: Session = Depends(get_db),
):
    """向客户档案手机号发 6 位短信验证码（中期方案，需腾讯云模板审核通过）。"""
    task = db.query(ConsentTask).filter(ConsentTask.token == token).first()
    if not task:
        return {"ok": False, "error": "协议链接不存在或已失效"}
    if task.status != "pending":
        return {"ok": False, "error": "该协议已不可签字"}
    if not _consent_sms_enabled():
        return {"ok": False, "error": "短信服务未启用，请直接输入手机号验证"}
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    if not cust or not (cust.phone or "").strip():
        return {"ok": False, "error": "客户档案缺手机号，无法发送验证码"}
    import time as _t, secrets as _s
    now = _t.time()
    last = _CONSENT_LAST_SENT.get(token, 0)
    if now - last < _CONSENT_SEND_THROTTLE:
        wait = int(_CONSENT_SEND_THROTTLE - (now - last))
        return {"ok": False, "error": f"请 {wait} 秒后再获取"}
    code = "".join(_s.choice("0123456789") for _ in range(6))
    _CONSENT_CODES[token] = (code, now + _CONSENT_CODE_TTL)
    _CONSENT_LAST_SENT[token] = now
    # 发短信
    sent = False
    err = None
    try:
        from app.services.sms_tencent import send_sms_template
        sent, err = send_sms_template(
            cust.phone,
            settings.tencent_sms_tmpl_consent,
            [code, str(_CONSENT_CODE_TTL // 60)],  # 假设模板形如：您的验证码 {1}，{2} 分钟内有效
        )
    except Exception as _e:
        err = str(_e)
    _log_consent_audit(db, request, task.id, "code_sent",
                       phone=cust.phone,
                       note=f"sent={sent} err={err or ''}")
    if not sent:
        return {"ok": False, "error": f"短信发送失败：{err or '未知错误'}"}
    return {"ok": True, "phone_masked": _phone_mask(cust.phone)}


@app.post("/consent/{token}/verify")
async def consent_verify(
    token: str, request: Request, db: Session = Depends(get_db),
):
    """两种模式：
    - SMS 模式（短信验证码已启用）：客户输入收到的 6 位验证码
    - 手机号匹配模式（默认 / SMS 未启用）：客户输入手机号，与档案对比
    """
    task = db.query(ConsentTask).filter(ConsentTask.token == token).first()
    if not task:
        return {"ok": False, "error": "协议链接不存在或已失效"}
    if task.status != "pending":
        return {"ok": False, "error": "该协议已不可签字"}
    body = await request.json()
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    if not cust or not (cust.phone or "").strip():
        _log_consent_audit(db, request, task.id, "code_verify_fail",
                           note="no_phone_on_file")
        return {"ok": False, "error": "客户档案缺手机号，无法验证身份，请联系医院"}

    if _consent_sms_enabled():
        # 短信验证码模式
        code_input = (body.get("code") or "").strip()
        if not code_input or not code_input.isdigit() or len(code_input) != 6:
            return {"ok": False, "error": "请输入 6 位验证码"}
        import time as _t
        entry = _CONSENT_CODES.get(token)
        if not entry:
            _log_consent_audit(db, request, task.id, "code_verify_fail",
                               phone=cust.phone, note="no_code_issued")
            return {"ok": False, "error": "请先获取验证码"}
        code_correct, expires_at = entry
        if _t.time() > expires_at:
            _CONSENT_CODES.pop(token, None)
            _log_consent_audit(db, request, task.id, "code_verify_fail",
                               phone=cust.phone, note="code_expired")
            return {"ok": False, "error": "验证码已过期，请重新获取"}
        if code_input != code_correct:
            _log_consent_audit(db, request, task.id, "code_verify_fail",
                               phone=cust.phone, note="code_mismatch")
            return {"ok": False, "error": "验证码不正确"}
        # 通过 → 清掉一次性 code
        _CONSENT_CODES.pop(token, None)
        request.session[f"consent_verified_{token}"] = True
        _log_consent_audit(db, request, task.id, "code_verify_ok",
                           phone=cust.phone, note="mode=sms")
        return {"ok": True}
    else:
        # 手机号匹配模式
        phone_input = (body.get("phone") or "").strip()
        if not phone_input:
            return {"ok": False, "error": "请输入手机号"}
        if not _consent_phone_match(phone_input, cust.phone):
            _log_consent_audit(db, request, task.id, "code_verify_fail",
                               phone=phone_input, note="phone_mismatch")
            return {"ok": False, "error": "手机号与档案不符"}
        request.session[f"consent_verified_{token}"] = True
        _log_consent_audit(db, request, task.id, "code_verify_ok",
                           phone=cust.phone, note="mode=phone_match")
        return {"ok": True}


@app.post("/consent/{token}/sign")
async def consent_sign_submit(
    token: str, request: Request, db: Session = Depends(get_db),
):
    task = db.query(ConsentTask).filter(ConsentTask.token == token).first()
    if not task:
        return {"ok": False, "error": "协议链接不存在或已失效"}
    if task.status != "pending":
        return {"ok": False, "error": f"该协议已 {_CONSENT_STATUS_ZH.get(task.status, task.status)}，不可再次签字"}
    # 强制要求先通过手机号验证（防别人拿链接代签）
    if not _is_consent_verified(request, token):
        _log_consent_audit(db, request, task.id, "sign_fail", note="not_verified")
        return {"ok": False, "error": "请先完成身份验证"}
    body = await request.json()
    sig_data = (body.get("signature") or "").strip()
    if not sig_data.startswith("data:image/") or "," not in sig_data:
        _log_consent_audit(db, request, task.id, "sign_fail", note="invalid_data_url")
        return {"ok": False, "error": "签字数据无效"}
    payload = sig_data.split(",", 1)[1]
    if len(payload) < 800:
        _log_consent_audit(db, request, task.id, "sign_fail",
                           note=f"sig_too_small payload_len={len(payload)}")
        return {"ok": False, "error": "签字过于简单，请重新签字"}
    import base64
    try:
        raw = base64.b64decode(payload, validate=True)
    except Exception:
        _log_consent_audit(db, request, task.id, "sign_fail", note="b64_decode_failed")
        return {"ok": False, "error": "签字数据解码失败"}
    from pathlib import Path as _P
    sig_dir = _P("uploads/consent_signatures")
    sig_dir.mkdir(parents=True, exist_ok=True)
    fname = f"task_{task.id}_{secrets.token_hex(6)}.png"
    (sig_dir / fname).write_bytes(raw)

    # 计算文档/签字的 SHA256（证据链关键 — 后续如果文档或签字被改，哈希对不上）
    doc_hash = _sha256(task.snapshot_html or "")
    sig_hash = _sha256(raw)

    task.signature_path = f"consent_signatures/{fname}"
    task.signed_at = datetime.utcnow()
    task.signed_ip = (request.client.host if request.client else "")[:60]
    task.status = "signed"
    db.commit()
    # 记签字成功 + 文档/签字哈希
    _log_consent_audit(
        db, request, task.id, "sign_success",
        doc_sha256=doc_hash, sig_sha256=sig_hash,
        note=f"sig_bytes={len(raw)} sig_path={task.signature_path}",
    )
    # PDF 自动归档（失败不阻断签字成功 — 系统库缺也只是 PDF 不生成）
    try:
        from app.services.consent_pdf import generate_consent_pdf
        path, err = generate_consent_pdf(db, task.id)
        if err:
            logger.info("[consent] PDF 自动生成跳过 task=%s: %s", task.id, err)
    except Exception as _e:
        logger.warning("[consent] PDF 生成异常 task=%s: %s", task.id, _e)
    # 推送给发起人的企业微信
    try:
        from app.services import wecom_notify as _wn
        _wn.notify_consent_signed(db, task)
    except Exception as _e:
        logger.warning("[wecom] notify_consent_signed failed: %s", _e)
    return {"ok": True, "task_id": task.id}


@app.get("/admin/consent-tasks/{task_id}/printable", response_class=HTMLResponse)
async def admin_consent_task_printable(
    task_id: int, request: Request, db: Session = Depends(get_db),
):
    """打印友好版（A4 一页，含签字图）。浏览器→打印→另存为 PDF。
    不需要服务器装 weasyprint / pango。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    pet = db.get(Pet, task.pet_id) if task.pet_id else None
    clinic_name = "大风动物医院"
    if pet and pet.store:
        clinic_name = f"大风动物医院（{pet.store.replace('店', '分院')}）"
    return templates.TemplateResponse(request, "admin_consent_printable.html", {
        "task": task, "cust": cust, "pet": pet,
        "clinic_name": clinic_name,
    })


@app.post("/admin/consent-tasks/{task_id}/regenerate-pdf")
async def admin_consent_task_regen_pdf(
    task_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """手动重新生成 PDF（已签未归档时用）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    if task.status != "signed":
        return RedirectResponse(f"/admin/consent-tasks/{task_id}?msg=只有已签状态可生成 PDF", status_code=303)
    try:
        from app.services.consent_pdf import generate_consent_pdf
        path, err = generate_consent_pdf(db, task_id)
        if path:
            return RedirectResponse(f"/admin/consent-tasks/{task_id}?msg=PDF 已生成", status_code=303)
        from urllib.parse import quote
        return RedirectResponse(
            f"/admin/consent-tasks/{task_id}?msg=" + quote(f"PDF 失败：{err or '未知错误'}"),
            status_code=303,
        )
    except Exception as e:
        from urllib.parse import quote
        return RedirectResponse(
            f"/admin/consent-tasks/{task_id}?msg=" + quote(f"PDF 异常：{type(e).__name__}: {str(e)[:120]}"),
            status_code=303,
        )


@app.post("/admin/consent-tasks/{task_id}/delete")
async def admin_consent_task_delete(
    task_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    next_url: str = Form(""),
):
    """彻底删除任务（仅 pending / cancelled 状态可删 — 已签的有法律凭证不能动）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    if task.status not in ("pending", "cancelled"):
        return RedirectResponse(
            f"/admin/consent-tasks/{task_id}?msg=只能删除待签/已取消的任务（已签有法律凭证不可删，请用'取消'）",
            status_code=303,
        )
    cust_id = task.customer_id
    title = task.title
    db.delete(task)   # CASCADE 会顺带清掉关联的 ConsentDocument（如果有）
    db.commit()
    _audit(db, request, "consent_task_delete", application_id=None,
           detail={"task_id": task_id, "customer_id": cust_id, "title": title})
    db.commit()
    # 删除后任务已不存在，回跳列表（移动端回 /m/consents，桌面回客户档案）
    dest = _safe_next(next_url, f"/admin/customers/{cust_id}?tab=docs")
    sep = "&" if "?" in dest else "?"
    return RedirectResponse(f"{dest}{sep}msg=已删除协议任务", status_code=303)


@app.post("/admin/consent-tasks/{task_id}/cancel")
async def admin_consent_task_cancel(
    task_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    next_url: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    _base = _safe_next(next_url, f"/admin/consent-tasks/{task_id}")
    def _r(m):
        return RedirectResponse(f"{_base}{'&' if '?' in _base else '?'}msg={m}", status_code=303)
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    if task.status != "pending":
        return _r("只能取消待签状态")
    task.status = "cancelled"
    db.commit()
    return _r("已取消")


@app.post("/admin/consent-templates/{tid}/toggle")
async def admin_consent_template_toggle(
    tid: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    item = db.get(ConsentTemplate, tid)
    if not item:
        raise HTTPException(404)
    item.is_active = not item.is_active
    db.commit()
    return RedirectResponse(
        f"/admin/consent-templates?msg={'已上架' if item.is_active else '已下架'}：{item.name}",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# 优惠券 (Coupon) — 发放 / 核销 / 作废
# ---------------------------------------------------------------------------

_COUPON_KIND_ZH = {
    "cash":      "现金抵扣券",
    "discount":  "折扣券",
    "free_item": "兑换券",
}
_COUPON_STATUS_ZH = {
    "issued":    "未使用",
    "used":      "已核销",
    "expired":   "已过期",
    "cancelled": "已作废",
}


def _gen_coupon_code() -> str:
    """生成 12 位券码：日期 + 4 位随机。"""
    import secrets, string
    from datetime import date
    suffix = "".join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(6))
    return date.today().strftime("%y%m%d") + suffix


def _coupon_is_expired(c: Coupon) -> bool:
    if not c.expires_at:
        return False
    from datetime import date
    try:
        y, m, d = c.expires_at[:10].split("-")
        return date(int(y), int(m), int(d)) < date.today()
    except Exception:
        return False


def _coupon_compute_amount(c: Coupon, invoice_total: float) -> float:
    """按券类型 + 收费单总额 算出可抵扣多少。"""
    if c.min_amount and invoice_total < c.min_amount:
        return 0.0
    if c.kind == "cash":
        return float(min(c.face_value or 0, invoice_total))
    if c.kind == "discount":
        pct = float(c.discount_pct or 0)
        if pct <= 0 or pct >= 1:
            return 0.0
        # discount_pct=0.9 表示 9 折 → 抵扣 10%
        return round(invoice_total * (1 - pct), 2)
    if c.kind == "free_item":
        # 兑换券：用面额作参考价上限
        return float(min(c.face_value or 0, invoice_total))
    return 0.0


@app.get("/admin/coupons", response_class=HTMLResponse)
async def admin_coupons_list(
    request: Request,
    db: Session = Depends(get_db),
    status: str = Query(""),
    q: str = Query(""),
    store: str = Query(""),
):
    """优惠券总列表（发放管理）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        _wb_store = (store or "").strip()
    else:
        _wb_store = _admin_store
    qq = db.query(Coupon)
    qq = _apply_store_filter(qq, Coupon.store, _wb_store)
    if status:
        qq = qq.filter(Coupon.status == status)
    if q.strip():
        like = f"%{q.strip()}%"
        qq = qq.filter((Coupon.code.like(like)) | (Coupon.title.like(like)))
    rows = qq.order_by(Coupon.id.desc()).limit(500).all()
    # 顺手把过期未用的标 expired
    from datetime import date
    today = date.today().isoformat()
    flipped = 0
    for c in rows:
        if c.status == "issued" and c.expires_at and c.expires_at < today:
            c.status = "expired"; flipped += 1
    if flipped:
        db.commit()
    # 客户名映射
    cust_ids = [c.customer_id for c in rows if c.customer_id]
    cust_map = {x.id: x for x in db.query(Customer).filter(Customer.id.in_(cust_ids)).all()} if cust_ids else {}
    # 统计
    counts = {
        "issued":    db.query(Coupon).filter(Coupon.status == "issued").count(),
        "used":      db.query(Coupon).filter(Coupon.status == "used").count(),
        "expired":   db.query(Coupon).filter(Coupon.status == "expired").count(),
        "cancelled": db.query(Coupon).filter(Coupon.status == "cancelled").count(),
    }
    return templates.TemplateResponse(request, "uk/coupons.html", {
        "rows": rows,
        "cust_map": cust_map,
        "status": status,
        "q": q,
        "counts": counts,
        "kind_zh": _COUPON_KIND_ZH,
        "status_zh": _COUPON_STATUS_ZH,
        "csrf_token": _get_csrf_token(request),
        "wb_store": _wb_store,
        "is_superadmin": request.session.get("admin_role") == "superadmin",
    })


@app.post("/admin/coupons/issue")
async def admin_coupon_issue(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    title: str = Form(...),
    kind: str = Form("cash"),
    face_value: str = Form("0"),
    discount_pct: str = Form("0"),
    min_amount: str = Form("0"),
    expires_at: str = Form(""),
    customer_id: str = Form(""),   # 留空 = 通用券
    quantity: str = Form("1"),     # 批量发放数量
    code_prefix: str = Form(""),   # 自定义前缀（可选）
    notes: str = Form(""),
    store: str = Form(""),
):
    """发放优惠券（可指定客户 / 通用；可单张 / 批量）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    if kind not in _COUPON_KIND_ZH:
        kind = "cash"
    try:
        fv = max(0.0, float(face_value))
        pct = max(0.0, float(discount_pct))
        if pct > 1:
            pct = pct / 100.0  # 用户填 90 表示 9 折
        ma = max(0.0, float(min_amount))
        qty = max(1, min(500, int(quantity)))
    except (TypeError, ValueError):
        return RedirectResponse("/admin/coupons?msg=数值字段无效", status_code=303)
    cust_id_int = None
    if customer_id.strip().isdigit():
        cust_id_int = int(customer_id)
        if not db.get(Customer, cust_id_int):
            return RedirectResponse("/admin/coupons?msg=客户不存在", status_code=303)
    issued = 0
    for _ in range(qty):
        # 唯一码（碰撞重试 5 次）
        code = ""
        for _try in range(5):
            cand = (code_prefix.strip().upper()[:10] + _gen_coupon_code())[:40]
            if not db.query(Coupon.id).filter(Coupon.code == cand).first():
                code = cand
                break
        if not code:
            continue
        db.add(Coupon(
            code=code,
            customer_id=cust_id_int,
            title=title.strip()[:120],
            kind=kind,
            face_value=fv,
            discount_pct=pct,
            min_amount=ma,
            expires_at=(expires_at or "").strip()[:20],
            status="issued",
            issued_by=request.session.get("admin_username", "admin"),
            notes=(notes or "").strip(),
            store=_resolve_store_for_create(request, store),
        ))
        issued += 1
    db.commit()
    _audit(db, request, "coupon_issue", application_id=None,
           detail={"qty": issued, "kind": kind, "customer_id": cust_id_int})
    db.commit()
    return RedirectResponse(
        f"/admin/coupons?msg=已发放 {issued} 张{'（指定客户）' if cust_id_int else '（通用券）'}",
        status_code=303,
    )


@app.post("/admin/coupons/{cid}/cancel")
async def admin_coupon_cancel(
    cid: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    c = db.get(Coupon, cid)
    if not c:
        raise HTTPException(404)
    if c.status not in ("issued",):
        return RedirectResponse(f"/admin/coupons?msg=该券状态不允许作废", status_code=303)
    c.status = "cancelled"
    db.commit()
    return RedirectResponse(f"/admin/coupons?msg=已作废", status_code=303)


# ---------------------------------------------------------------------------
# 套餐 (Package) — 目录 + 售卖 + 核销
# ---------------------------------------------------------------------------

_PACKAGE_CATEGORY_ZH = {
    "beauty":  "美容",
    "bath":    "洗护",
    "medical": "医疗",
    "boarding":"寄养",
    "other":   "其他",
}


@app.get("/admin/packages", response_class=HTMLResponse)
async def admin_packages_list(request: Request, db: Session = Depends(get_db), store: str = ""):
    """套餐目录管理（创建/编辑/启停）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        _wb_store = (store or "").strip()
    else:
        _wb_store = _admin_store
    q = db.query(PackageProduct)
    q = _apply_store_filter(q, PackageProduct.store, _wb_store)
    items = q.order_by(PackageProduct.is_active.desc(), PackageProduct.id.desc()).all()
    # 统计：每个产品已售套餐数
    from sqlalchemy import func as _f
    sold_rows = (
        db.query(CustomerPackage.product_id, _f.count(CustomerPackage.id))
        .group_by(CustomerPackage.product_id)
        .all()
    )
    sold_map = {pid: cnt for pid, cnt in sold_rows if pid}
    return templates.TemplateResponse(request, "uk/packages.html", {
        "items": items,
        "sold_map": sold_map,
        "category_zh": _PACKAGE_CATEGORY_ZH,
        "csrf_token": _get_csrf_token(request),
        "wb_store": _wb_store,
        "is_superadmin": request.session.get("admin_role") == "superadmin",
    })


@app.post("/admin/packages/create", name="admin_packages_create")
async def admin_packages_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(...),
    category: str = Form("beauty"),
    total_uses: str = Form("10"),
    sell_price: str = Form("0"),
    unit_price: str = Form("0"),
    validity_days: str = Form("365"),
    notes: str = Form(""),
    store: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    try:
        n_total = max(1, int(total_uses))
        v_days  = max(0, int(validity_days))
        sp = max(0.0, float(sell_price))
        up = max(0.0, float(unit_price))
    except (TypeError, ValueError):
        return RedirectResponse("/admin/packages?msg=数值字段无效", status_code=303)
    if category not in _PACKAGE_CATEGORY_ZH:
        category = "other"
    p = PackageProduct(
        name=(name or "").strip()[:120],
        category=category,
        total_uses=n_total,
        sell_price=sp,
        unit_price=up,
        validity_days=v_days,
        notes=(notes or "").strip(),
        is_active=True,
        store=_resolve_store_for_create(request, store),
    )
    db.add(p); db.commit()
    return RedirectResponse(f"/admin/packages?msg=已创建套餐：{p.name}", status_code=303)


@app.post("/admin/packages/{pid}/edit")
async def admin_packages_edit(
    pid: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(...),
    category: str = Form("beauty"),
    total_uses: str = Form("10"),
    sell_price: str = Form("0"),
    unit_price: str = Form("0"),
    validity_days: str = Form("365"),
    notes: str = Form(""),
    store: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    p = db.get(PackageProduct, pid)
    if not p:
        raise HTTPException(404)
    try:
        p.total_uses    = max(1, int(total_uses))
        p.validity_days = max(0, int(validity_days))
        p.sell_price    = max(0.0, float(sell_price))
        p.unit_price    = max(0.0, float(unit_price))
    except (TypeError, ValueError):
        return RedirectResponse("/admin/packages?msg=数值字段无效", status_code=303)
    p.name     = (name or "").strip()[:120]
    p.category = category if category in _PACKAGE_CATEGORY_ZH else "other"
    p.notes    = (notes or "").strip()
    # superadmin 可改门店归属
    if request.session.get("admin_role") == "superadmin":
        p.store = (store or "").strip()
    db.commit()
    return RedirectResponse("/admin/packages?msg=已保存", status_code=303)


@app.post("/admin/packages/{pid}/toggle")
async def admin_packages_toggle(
    pid: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    p = db.get(PackageProduct, pid)
    if not p:
        raise HTTPException(404)
    p.is_active = not p.is_active
    db.commit()
    return RedirectResponse(
        f"/admin/packages?msg={'已上架' if p.is_active else '已下架'}：{p.name}",
        status_code=303,
    )


# ── 客户买套餐 ────────────────────────────────────────────────────
@app.post("/admin/customers/{customer_id}/packages/sell")
async def admin_customer_buy_package(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    product_id: int = Form(...),
    pay_method: str = Form("cash"),
    pet_id: int = Form(0),
    custom_price: str = Form(""),     # 留空 → 按目录价
    note: str = Form(""),
    # 补录老系统卡专用（可选）
    used_count: str = Form("0"),        # 已用次数（旧系统消耗过的）
    custom_purchase_date: str = Form(""),  # 真实购买日 YYYY-MM-DD（留空 = 今天）
    custom_total_uses: str = Form(""),  # 自定义总次（旧系统卡不同套餐时）
):
    """给客户售卖一份套餐 → 新建 CustomerPackage。
    若 pay_method == 'wallet' → 自动从钱包扣款；否则只记账（外部已收）。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    prod = db.get(PackageProduct, product_id)
    if not prod or not prod.is_active:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=packages&msg=套餐已下架或不存在", status_code=303)
    try:
        price = float(custom_price) if custom_price.strip() else float(prod.sell_price)
    except (TypeError, ValueError):
        price = float(prod.sell_price)
    if price < 0:
        price = 0.0

    from datetime import date, timedelta, datetime as _dt
    today = date.today()
    # 真实购买日（补录老卡用）
    purchase_d = today
    if (custom_purchase_date or "").strip():
        try:
            purchase_d = _dt.strptime(custom_purchase_date.strip(), "%Y-%m-%d").date()
        except ValueError:
            purchase_d = today
    expires = ""
    if prod.validity_days and prod.validity_days > 0:
        expires = (purchase_d + timedelta(days=prod.validity_days)).isoformat()
    # 已用次数（补录老卡用）
    try:
        used_n = max(0, int(used_count or 0))
    except (TypeError, ValueError):
        used_n = 0
    # 总次（旧系统不同套餐时允许覆盖）
    try:
        total_n = int((custom_total_uses or "").strip() or prod.total_uses or 0)
    except (TypeError, ValueError):
        total_n = prod.total_uses
    if total_n < 1:
        total_n = prod.total_uses
    if used_n > total_n:
        used_n = total_n
    # 状态：已用满 → exhausted
    initial_status = "exhausted" if used_n >= total_n else "active"

    cp = CustomerPackage(
        customer_id=customer_id,
        pet_id=pet_id or None,
        product_id=prod.id,
        name=prod.name,
        category=prod.category,
        total_uses=total_n,
        used_count=used_n,
        sell_price=price,
        unit_price=prod.unit_price,
        purchase_date=purchase_d.isoformat(),
        expires_at=expires,
        status=initial_status,
        store=_get_admin_store(request),
        operator=request.session.get("admin_username", "admin"),
        note=(note or "").strip(),
    )
    db.add(cp); db.flush()

    # 如果用钱包支付，立刻扣款
    # pay_method = 'external' / 'imported' / 'legacy' → 视为旧系统/外部已付，不动钱包不写发票
    if pay_method == "wallet":
        wallet = _get_or_create_wallet(db, customer_id)
        try:
            _wallet_apply_tx(
                db, wallet, tx_type="consume", amount=price,
                operator=request.session.get("admin_username", "admin"),
                store=_get_admin_store(request),
                note=f"购买套餐 {prod.name}",
            )
        except HTTPException as he:
            db.rollback()
            return RedirectResponse(
                f"/admin/customers/{customer_id}?tab=packages&msg={he.detail}", status_code=303
            )

    db.commit()
    _audit(db, request, "package_sell", application_id=None,
           detail={"customer_id": customer_id, "product": prod.name, "price": price, "pay": pay_method})
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{customer_id}?tab=packages&msg=已售卖：{prod.name}（¥{price:.2f}）",
        status_code=303,
    )


@app.post("/admin/customer-packages/{cp_id}/refund")
async def admin_customer_package_refund(
    cp_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    note: str = Form(""),
):
    """退掉未用完的套餐（按剩余次数比例退回钱包）。仅 superadmin。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    cp = db.get(CustomerPackage, cp_id)
    if not cp:
        raise HTTPException(404)
    if cp.status != "active":
        return RedirectResponse(
            f"/admin/customers/{cp.customer_id}?tab=packages&msg=该套餐已非激活状态", status_code=303,
        )
    remaining = max(0, cp.total_uses - cp.used_count)
    refund_amt = round(cp.sell_price * (remaining / cp.total_uses), 2) if cp.total_uses > 0 else 0.0
    if refund_amt > 0:
        wallet = _get_or_create_wallet(db, cp.customer_id)
        _wallet_apply_tx(
            db, wallet, tx_type="adjust", amount=refund_amt,
            operator=request.session.get("admin_username", "admin"),
            store=_get_admin_store(request),
            note=f"套餐退款 {cp.name}（剩 {remaining}/{cp.total_uses}）",
        )
    cp.status = "refunded"
    cp.note = ((cp.note or "") + f"\n[退款 ¥{refund_amt:.2f}：{note}]").strip()
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{cp.customer_id}?tab=packages&msg=已退款 ¥{refund_amt:.2f} 到钱包",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# 押金 (Deposit) — 手术 / 寄养 / 美容
# ---------------------------------------------------------------------------

_DEPOSIT_CATEGORY_ZH = {
    "surgery":     "手术押金",
    "boarding":    "寄养押金",
    "beauty":      "美容押金",
    "hospital":    "住院押金",
    "prepaid":     "预付款",
    "other":       "其他押金",
}
_DEPOSIT_STATUS_ZH = {
    "held":           "已收待结",
    "applied":        "已抵扣",
    "partial_refund": "部分退还",
    "refunded":       "已全额退款",
    "cancelled":      "已作废",
}


@app.get("/admin/deposits/{dep_id}", response_class=HTMLResponse)
@app.get("/admin/deposits/{dep_id}/detail", response_class=HTMLResponse)
async def admin_deposit_detail(dep_id: int, request: Request, db: Session = Depends(get_db)):
    """押金独立详情页：抬头 / 抵扣流水 / 退款流水 / 操作。"""
    require_admin(request)
    dep = db.get(Deposit, dep_id)
    if not dep:
        raise HTTPException(404, "押金记录不存在")
    # 限店权限
    admin_store = _get_admin_store(request)
    if admin_store and dep.store and dep.store != admin_store:
        raise HTTPException(403, "无权操作其他门店押金")
    cust = db.get(Customer, dep.customer_id) if dep.customer_id else None
    pet = db.get(Pet, dep.pet_id) if dep.pet_id else None
    visit = db.get(Visit, dep.visit_id) if dep.visit_id else None
    appointment = db.get(Appointment, dep.appointment_id) if dep.appointment_id else None
    applied_inv = db.get(Invoice, dep.applied_invoice_id) if dep.applied_invoice_id else None
    # 抵扣 / 退款流水从 AuditLog 解析
    audits = db.query(AuditLog).filter(
        AuditLog.action.in_(("deposit_apply", "deposit_refund", "deposit_cancel"))
    ).order_by(AuditLog.id.asc()).all()
    apply_rows = []
    refund_rows = []
    cancel_row = None
    for a in audits:
        try:
            d = json.loads(a.detail or "{}")
        except Exception:
            d = {}
        if d.get("deposit_id") != dep_id:
            continue
        if a.action == "deposit_apply":
            inv = db.get(Invoice, d.get("invoice_id")) if d.get("invoice_id") else None
            apply_rows.append({
                "amount": float(d.get("amount") or 0),
                "invoice": inv,
                "actor": a.actor, "at": a.created_at,
            })
        elif a.action == "deposit_refund":
            refund_rows.append({
                "amount": float(d.get("amount") or 0),
                "actor": a.actor, "at": a.created_at,
            })
        elif a.action == "deposit_cancel":
            cancel_row = {"actor": a.actor, "at": a.created_at}
    # 同客户其他未付发票（用于"抵扣到新单"下拉）
    other_invoices = []
    if cust:
        other_invoices = db.query(Invoice).filter(
            Invoice.customer_id == cust.id,
            Invoice.payment_status == "unpaid",
        ).order_by(Invoice.id.desc()).limit(20).all()
    remaining = float(dep.amount or 0) - float(dep.applied_amount or 0) - float(dep.refunded_amount or 0)
    if remaining < 0:
        remaining = 0.0
    return templates.TemplateResponse(request, "uk/deposit_detail.html", {
        "dep": dep, "cust": cust, "pet": pet, "visit": visit, "appointment": appointment,
        "applied_inv": applied_inv,
        "apply_rows": apply_rows, "refund_rows": refund_rows, "cancel_row": cancel_row,
        "remaining": round(remaining, 2),
        "other_invoices": other_invoices,
        "category_zh": _DEPOSIT_CATEGORY_ZH,
        "status_zh": _DEPOSIT_STATUS_ZH,
        "pay_zh": _INV_PAY_ZH,
        "msg": request.query_params.get("msg"),
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/follow-ups/{fu_id}", response_class=HTMLResponse)
@app.get("/admin/follow-ups/{fu_id}/detail", response_class=HTMLResponse)
async def admin_follow_up_detail(fu_id: int, request: Request, db: Session = Depends(get_db)):
    """回访独立详情页：抬头 / 时间线 / 客户反馈 / 状态切换 / 操作。"""
    require_admin(request)
    fu = db.get(FollowUp, fu_id)
    if not fu:
        raise HTTPException(404, "回访任务不存在")
    admin_store = _get_admin_store(request)
    if admin_store and fu.store and fu.store != admin_store:
        raise HTTPException(403, "无权操作其他门店回访")
    cust = db.get(Customer, fu.customer_id) if fu.customer_id else None
    pet = db.get(Pet, fu.pet_id) if fu.pet_id else None
    visit = db.get(Visit, fu.visit_id) if fu.visit_id else None
    # 反馈结构化数据
    response_items: list[dict] = []
    try:
        rd = json.loads(fu.response_data or "{}")
        if isinstance(rd, dict):
            for k, v in rd.items():
                response_items.append({"k": k, "v": v if not isinstance(v, (list, dict)) else json.dumps(v, ensure_ascii=False)})
    except Exception:
        pass
    # 推送 / 反馈日志（从 NotificationLog 反查 — 关键词匹配）
    from app.models import NotificationLog as _NL
    notif_rows = db.query(_NL).filter(
        _NL.payload.like(f"%followup#{fu_id}%")
    ).order_by(_NL.id.asc()).all() if cust else []
    # 时间线
    timeline = []
    if fu.created_at:
        timeline.append({"kind": "created", "at": fu.created_at, "text": "回访任务已生成"})
    if fu.sent_at:
        timeline.append({"kind": "sent", "at": fu.sent_at,
                         "text": f"已发送（{fu.channel or '渠道未指定'}）"})
    for n in notif_rows:
        timeline.append({"kind": "notif", "at": n.created_at,
                         "text": f"通知 · {n.channel} · {'成功' if n.success else '失败'}"})
    if fu.response_at:
        timeline.append({"kind": "responded", "at": fu.response_at,
                         "text": f"客户反馈 · {fu.response or ''} · {fu.response_note or ''}"})
    if fu.handled_at:
        timeline.append({"kind": "handled", "at": fu.handled_at,
                         "text": f"医院处理（{fu.handled_by or '—'}）· {fu.handle_note or ''}"})
    timeline.sort(key=lambda x: x["at"] or datetime.min)
    fu_status_zh = {
        "pending": "未到期",
        "due": "今日到期",
        "sent": "已发送",
        "responded": "客户已反馈",
        "phone_pending": "需电话",
        "closed": "已完成",
        "skipped": "已跳过",
    }
    return templates.TemplateResponse(request, "uk/follow_up_detail.html", {
        "fu": fu, "cust": cust, "pet": pet, "visit": visit,
        "response_items": response_items,
        "timeline": timeline,
        "fu_status_zh": fu_status_zh,
        "msg": request.query_params.get("msg"),
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/deposits/create")
async def admin_deposit_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(...),
    pet_id: int = Form(0),
    appointment_id: int = Form(0),
    visit_id: int = Form(0),
    category: str = Form("surgery"),
    amount: str = Form(...),
    pay_method: str = Form("cash"),
    note: str = Form(""),
):
    """收押金。amount > 0。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    try:
        amt = float(amount)
    except (TypeError, ValueError):
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=deposits&msg=金额无效", status_code=303)
    if amt <= 0:
        return RedirectResponse(f"/admin/customers/{customer_id}?tab=deposits&msg=押金需大于 0", status_code=303)
    if category not in _DEPOSIT_CATEGORY_ZH:
        category = "other"
    d = Deposit(
        customer_id=customer_id,
        pet_id=pet_id or None,
        appointment_id=appointment_id or None,
        visit_id=visit_id or None,
        category=category,
        amount=amt,
        pay_method=pay_method,
        status="held",
        store=_get_admin_store(request),
        operator=request.session.get("admin_username", "admin"),
        note=(note or "").strip()[:500],
    )
    db.add(d); db.commit()
    _audit(db, request, "deposit_create", application_id=None,
           detail={"customer_id": customer_id, "amount": amt, "category": category})
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{customer_id}?tab=deposits&msg=已收押金 ¥{amt:.2f}",
        status_code=303,
    )


@app.post("/admin/deposits/{dep_id}/apply")
async def admin_deposit_apply(
    dep_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    invoice_id: int = Form(...),
    apply_amount: str = Form(""),  # 留空 = 全用
):
    """把押金应用到一张收费单：
    - apply_amt = min(押金未用, 收费单未付)
    - 押金 status 转 applied（如果全部用完）或 partial_refund 占位（部分用、剩余等退）
    - 收费单 total 不变，但记账；剩余金额仍按 正常 pay 流程结算
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    d = db.get(Deposit, dep_id)
    if not d:
        raise HTTPException(404)
    if d.status not in ("held", "partial_refund"):
        return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=押金状态不允许使用", status_code=303)
    inv = db.get(Invoice, invoice_id)
    if not inv:
        raise HTTPException(404)
    if inv.customer_id != d.customer_id:
        return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=押金与客户不匹配", status_code=303)
    if inv.payment_status == "paid":
        return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=该单已收款", status_code=303)
    # 押金剩余
    remaining = d.amount - (d.applied_amount or 0.0) - (d.refunded_amount or 0.0)
    if remaining <= 0:
        return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=押金已无余额", status_code=303)
    try:
        want = float(apply_amount) if apply_amount.strip() else remaining
    except (TypeError, ValueError):
        want = remaining
    want = max(0.0, min(want, remaining, float(inv.total_amount or 0)))
    if want <= 0:
        return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=应用金额需大于 0", status_code=303)

    d.applied_invoice_id = inv.id
    d.applied_amount = (d.applied_amount or 0.0) + want
    # 折算后续状态
    new_remaining = d.amount - d.applied_amount - (d.refunded_amount or 0.0)
    if new_remaining <= 1e-6:
        d.status = "applied"
    else:
        d.status = "partial_refund"  # 占位待退
    # 写 Payment 流水：让发票的 paid_sum / 状态都正确反映押金抵扣
    operator = request.session.get("admin_username", "admin")
    store = _get_admin_store(request)
    db.add(Payment(
        invoice_id=inv.id, customer_id=inv.customer_id,
        method="deposit", amount=round(want, 2),
        ref_id=d.id, ref_no="", status="success",
        store=store, operator=operator,
        note=f"押金抵扣 #{d.id}（{_DEPOSIT_CATEGORY_ZH.get(d.category, d.category)}）",
    ))
    db.flush()
    inv.notes = ((inv.notes or "") + f"\n[抵扣押金 #{d.id} ¥{want:.2f}]").strip()
    # 重算 invoice 状态（paid / partial / unpaid）
    _invoice_recompute_status(db, inv)
    db.commit()
    _audit(db, request, "deposit_apply", application_id=None,
           detail={"deposit_id": dep_id, "invoice_id": invoice_id, "amount": want})
    db.commit()
    return RedirectResponse(f"/admin/invoices/{invoice_id}?msg=已抵扣押金 ¥{want:.2f}", status_code=303)


@app.post("/admin/deposits/{dep_id}/refund")
async def admin_deposit_refund(
    dep_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    refund_amount: str = Form(""),
    note: str = Form(""),
):
    """退还押金剩余部分。默认退完所有未用余额。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    d = db.get(Deposit, dep_id)
    if not d:
        raise HTTPException(404)
    if d.status in ("refunded", "cancelled"):
        return RedirectResponse(
            f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=押金已结清",
            status_code=303,
        )
    remaining = d.amount - (d.applied_amount or 0.0) - (d.refunded_amount or 0.0)
    if remaining <= 0:
        return RedirectResponse(
            f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=押金无可退余额",
            status_code=303,
        )
    try:
        want = float(refund_amount) if refund_amount.strip() else remaining
    except (TypeError, ValueError):
        want = remaining
    want = max(0.0, min(want, remaining))
    if want <= 0:
        return RedirectResponse(
            f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=退款金额无效",
            status_code=303,
        )
    d.refunded_amount = (d.refunded_amount or 0.0) + want
    d.refunded_at = datetime.utcnow()
    new_remaining = d.amount - (d.applied_amount or 0.0) - d.refunded_amount
    if new_remaining <= 1e-6:
        d.status = "refunded" if not d.applied_amount else "applied"
    else:
        d.status = "partial_refund"
    d.note = ((d.note or "") + f"\n[退 ¥{want:.2f}：{note}]").strip()
    db.commit()
    _audit(db, request, "deposit_refund", application_id=None,
           detail={"deposit_id": dep_id, "amount": want})
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=已退押金 ¥{want:.2f}",
        status_code=303,
    )


@app.post("/admin/deposits/{dep_id}/cancel")
async def admin_deposit_cancel(
    dep_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """作废押金（误收时用）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    d = db.get(Deposit, dep_id)
    if not d:
        raise HTTPException(404)
    if d.applied_amount and d.applied_amount > 0:
        return RedirectResponse(
            f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=已有抵扣记录，不能作废",
            status_code=303,
        )
    d.status = "cancelled"
    db.commit()
    return RedirectResponse(
        f"/admin/customers/{d.customer_id or 0}?tab=deposits&msg=已作废",
        status_code=303,
    )


# ---------------------------------------------------------------------------
# Phase 2 — 就诊病历 (Visits)
# ---------------------------------------------------------------------------

_VISIT_TYPE_ZH = {
    "outpatient": "门诊",
    "followup": "复诊",
    "postop": "术后复查",
    "vaccine": "疫苗接种",
    "surgery_consult": "手术咨询",
    "other": "其他",
}


@app.get("/admin/visits", response_class=HTMLResponse)
async def page_admin_visits(
    request: Request,
    db: Session = Depends(get_db),
    q: str = Query(""),
    visit_type: str = Query(""),
    vet: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    pet_id: int = Query(0),
    customer_id: int = Query(0),
    page: int = Query(1),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    query = db.query(Visit)
    if q:
        query = query.join(Customer, Visit.customer_id == Customer.id, isouter=True)\
                     .join(Pet, Visit.pet_id == Pet.id, isouter=True)\
                     .filter(or_(
                         Customer.name.ilike(f"%{q}%"),
                         Customer.phone.ilike(f"%{q}%"),
                         Pet.name.ilike(f"%{q}%"),
                         Visit.diagnosis.ilike(f"%{q}%"),
                     ))
    if visit_type:
        query = query.filter(Visit.visit_type == visit_type)
    if vet:
        query = query.filter(Visit.vet_name.ilike(f"%{vet}%"))
    if date_from:
        query = query.filter(Visit.visit_date >= date_from)
    if date_to:
        query = query.filter(Visit.visit_date <= date_to)
    if pet_id:
        query = query.filter(Visit.pet_id == pet_id)
    if customer_id:
        query = query.filter(Visit.customer_id == customer_id)
    total = query.count()
    page_size = 30
    visits = query.order_by(Visit.visit_date.desc(), Visit.id.desc())\
                  .offset((page - 1) * page_size).limit(page_size).all()
    # 预加载 customer/pet 名字
    cust_map = {}
    pet_map = {}
    for v in visits:
        if v.customer_id and v.customer_id not in cust_map:
            c = db.get(Customer, v.customer_id)
            if c:
                cust_map[v.customer_id] = c
        if v.pet_id and v.pet_id not in pet_map:
            p = db.get(Pet, v.pet_id)
            if p:
                pet_map[v.pet_id] = p
    return templates.TemplateResponse(request, "uk/visits.html", {
        "visits": visits,
        "cust_map": cust_map,
        "pet_map": pet_map,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "total": total,
        "page": page,
        "page_size": page_size,
        "filters": {"q": q, "visit_type": visit_type, "vet": vet,
                    "date_from": date_from, "date_to": date_to},
    })


# ---------------------------------------------------------------------------
# 回访 (FollowUp) — 触发规则 + 同步辅助
# ---------------------------------------------------------------------------

# visit_type → 默认回访间隔天数；0/缺省 = 不主动回访
_FOLLOWUP_RULES: dict[str, int] = {
    "surgery":         3,
    "postop":          2,
    "outpatient":      7,
    "beauty":          14,
    "followup":        0,   # 本身就是复诊，不再产生新回访
    "vaccine":         0,
    "surgery_consult": 0,
    "other":           7,
}


def _gen_followup_token() -> str:
    import secrets
    return secrets.token_urlsafe(12)[:16]


def _visit_store_short(db: Session, v: Visit) -> str:
    """从 Visit → Pet.store 推出门店短名（用于回访的门店隔离）。"""
    if not v.pet_id:
        return ""
    pet = db.get(Pet, v.pet_id)
    return (pet.store or "") if pet else ""


def _resolve_vet_username(db: Session, vet_name: str) -> str:
    """把 Visit.vet_name（医生真名）映射到 AdminUser.username。
    用户名本身就是医生真名，所以只需 username 完全匹配 → 找不到则返回原样。
    用于 FollowUp.assigned_to，让"只看我的"能正确过滤。
    """
    name = (vet_name or "").strip()
    if not name:
        return ""
    u = db.query(AdminUser).filter(AdminUser.username == name, AdminUser.is_active == True).first()
    if u:
        return u.username
    return name  # 找不到也存原文，至少能显示出来


def _compute_followup_planned_date(v: Visit) -> str:
    """旧版兜底：visit_type 死规则。新系统用 _match_followup_templates。"""
    if v.follow_up_at and v.follow_up_at.strip():
        return v.follow_up_at.strip()
    days = _FOLLOWUP_RULES.get((v.visit_type or "outpatient").strip(), 0)
    if not days or days <= 0:
        return ""
    return _add_days_to_date(v.visit_date, days)


def _add_days_to_date(date_str: str, days: int) -> str:
    """YYYY-MM-DD + days → YYYY-MM-DD。"""
    base = (date_str or "").strip()[:10]
    if not base or days <= 0:
        return ""
    try:
        from datetime import date, timedelta
        y, m, d = base.split("-")
        dt = date(int(y), int(m), int(d)) + timedelta(days=int(days))
        return dt.isoformat()
    except Exception:
        return ""


def _match_followup_templates(db: Session, diagnosis: str, visit_type: str = "") -> list:
    """按 diagnosis 关键词匹配 FollowUpTemplate 列表（去重 + 按 priority 排序）。

    匹配规则：
      - 遍历 is_active=True 模板，按 priority desc
      - 任一 keyword 出现在 diagnosis 中 → 命中
      - 不区分大小写、忽略空白
      - 没命中任何模板 + visit_type 是门诊类 → 用「一般门诊（默认）」兜底
    """
    diag = (diagnosis or "").lower().strip()
    templates = db.query(FollowUpTemplate).filter(FollowUpTemplate.is_active == True)\
        .order_by(FollowUpTemplate.priority.desc(), FollowUpTemplate.id).all()
    matched: list = []
    seen_ids: set = set()
    for tpl in templates:
        if tpl.id in seen_ids:
            continue
        kws = [k.strip().lower() for k in (tpl.keywords or "").split(",") if k.strip()]
        if not kws:
            continue  # 空关键词模板不参与匹配（如默认门诊）
        for kw in kws:
            if kw and kw in diag:
                matched.append(tpl)
                seen_ids.add(tpl.id)
                break
    if not matched and visit_type in ("outpatient", "other", "followup", ""):
        default = db.query(FollowUpTemplate).filter(FollowUpTemplate.name == "一般门诊（默认）").first()
        if default:
            matched = [default]
    return matched


def _sync_followup_for_visit(db: Session, v: Visit) -> None:
    """根据诊断匹配模板，自动衍生/更新多轮 FollowUp。

    语义：
      - 命中模板列表 × 每模板的 rounds → 每个 (template_id, round_no) 一条 FollowUp
      - 未发送的 (status in pending/due) 可被重建：planned_date/assignee 覆盖
      - 已发送/已反馈/已完成的（sent/responded/closed/phone_pending）一律保留
      - 诊断变化导致模板不再命中时：未发送的旧轮次会被删除，已发送的保留作为历史
    """
    if not v or not v.id:
        return
    import json as _json

    templates = _match_followup_templates(db, v.diagnosis or "", v.visit_type or "")
    existing = db.query(FollowUp).filter(FollowUp.visit_id == v.id).all()
    existing_by_key = {(fu.template_id, fu.round_no): fu for fu in existing}

    base_date = (v.visit_date or "").strip()[:10]
    assignee = _resolve_vet_username(db, v.vet_name or "")[:80]
    store = _visit_store_short(db, v)

    # 病历级别关闭回访（主人带回家自治）→ 取消所有未发送的，已发送的保留作历史
    if getattr(v, "followup_disabled", False):
        for fu in existing:
            if fu.status in ("pending", "due"):
                db.delete(fu)
        return

    if not templates or not base_date:
        # 没匹配到 / 无就诊日期 → 删除所有未发送的
        for fu in existing:
            if fu.status in ("pending", "due"):
                db.delete(fu)
        return

    desired_keys: set = set()
    for tpl in templates:
        try:
            rounds = _json.loads(tpl.rounds_json or "[]")
        except Exception:
            rounds = []
        for round_idx, rnd in enumerate(rounds, start=1):
            day_offset = int(rnd.get("day_offset", 0) or 0)
            if day_offset <= 0:
                continue
            planned = _add_days_to_date(base_date, day_offset)
            if not planned:
                continue
            key = (tpl.id, round_idx)
            desired_keys.add(key)
            fu = existing_by_key.get(key)
            round_name = (rnd.get("round_name") or f"第 {round_idx} 轮")[:80]
            if fu is None:
                fu = FollowUp(
                    visit_id=v.id,
                    customer_id=v.customer_id,
                    pet_id=v.pet_id,
                    template_id=tpl.id,
                    template_name=tpl.name[:120],
                    round_no=round_idx,
                    round_name=round_name,
                    store=store,
                    assigned_to=assignee,
                    planned_date=planned,
                    status="pending",
                    feedback_token=_gen_followup_token(),
                )
                db.add(fu)
            else:
                fu.customer_id = v.customer_id
                fu.pet_id = v.pet_id
                fu.store = store
                fu.template_name = tpl.name[:120]
                fu.round_name = round_name
                if fu.status in ("pending", "due"):
                    fu.planned_date = planned
                    fu.assigned_to = assignee or fu.assigned_to
                if not fu.feedback_token:
                    fu.feedback_token = _gen_followup_token()

    # 删除不在 desired 里且未发送的
    for key, fu in existing_by_key.items():
        if key not in desired_keys and fu.status in ("pending", "due"):
            db.delete(fu)


@app.get("/admin/visits/create", response_class=HTMLResponse)
async def page_admin_visit_create(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = Query(0),
    pet_id: int = Query(0),
    appointment_id: int = Query(0),
    search_q: str = Query(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    appt = db.get(Appointment, appointment_id) if appointment_id else None
    pets = db.query(Pet).filter(Pet.customer_id == customer_id).all() if customer_id else []
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v[0] for v in vets]
    today = datetime.utcnow().strftime("%Y-%m-%d")
    # 客户搜索结果
    search_results = []
    if search_q and not customer_id:
        search_results = db.query(Customer).filter(
            or_(
                Customer.name.ilike(f"%{search_q}%"),
                Customer.phone.ilike(f"%{search_q}%"),
            )
        ).limit(10).all()
    return templates.TemplateResponse(request, "uk/visit_create.html", {  # B 补 - UK 重写
        "cust": cust,
        "pet": pet,
        "pets": pets,
        "appt": appt,
        "vet_names": vet_names,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "today": today,
        "csrf_token": _get_csrf_token(request),
        "mode": "create",
        "search_q": search_q,
        "search_results": search_results,
    })


@app.post("/admin/visits/create")
async def admin_visit_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
    appointment_id: int = Form(0),
    visit_date: str = Form(""),
    visit_type: str = Form("outpatient"),
    chief_complaint: str = Form(""),
    physical_exam: str = Form(""),
    diagnosis: str = Form(""),
    treatment_plan: str = Form(""),
    notes: str = Form(""),
    vet_name: str = Form(""),
    next_url: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    v = Visit(
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        appointment_id=appointment_id or None,
        visit_date=visit_date.strip()[:20],
        visit_type=visit_type.strip()[:40] or "outpatient",
        chief_complaint=chief_complaint.strip(),
        physical_exam=physical_exam.strip(),
        diagnosis=diagnosis.strip(),
        treatment_plan=treatment_plan.strip(),
        notes=notes.strip(),
        vet_name=vet_name.strip()[:80],
        created_by=request.session.get("admin_username", "admin"),
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    # 创建回访任务（按 visit_type 规则）
    _sync_followup_for_visit(db, v)
    db.commit()
    # 如果是从预约完成时创建，更新预约状态
    if appointment_id:
        appt = db.get(Appointment, appointment_id)
        if appt and appt.status == AppointmentStatus.confirmed.value:
            appt.status = AppointmentStatus.completed.value
            db.commit()
    # 移动端 next_url 可带 {id} 占位
    if next_url:
        nu = next_url.replace("{id}", str(v.id))
        return RedirectResponse(_safe_next(nu, f"/admin/visits/{v.id}?msg=就诊记录已创建"), status_code=303)
    # 桌面端：创建后直接进新病历的 SOAP 页（不再跳回客户档案落到第一只宠物，省两步）
    return RedirectResponse(f"/admin/visits/{v.id}?msg=就诊记录已创建", status_code=303)


@app.get("/admin/visits/{visit_id}/print", response_class=HTMLResponse)
async def admin_visit_print(visit_id: int, request: Request, db: Session = Depends(get_db)):
    """病历单打印（一次就诊完整记录：SOAP + 处方 + 检查）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    visit = db.get(Visit, visit_id)
    if not visit:
        raise HTTPException(404, "就诊记录不存在")
    cust = db.get(Customer, visit.customer_id) if visit.customer_id else None
    pet  = db.get(Pet,      visit.pet_id)      if visit.pet_id      else None
    prescriptions = db.query(Prescription).filter(Prescription.visit_id == visit_id).order_by(Prescription.id.asc()).all()
    exam_orders   = db.query(ExamOrder).filter(ExamOrder.visit_id == visit_id).order_by(ExamOrder.id.asc()).all()
    # 解析 exam_orders.items_json 一次，避免模板里调用 json
    # 同时给每个 report 计算 page_count（PDF 渲染嵌入用），上限 5 页避免打印爆炸
    from app.services.pdf_render import get_pdf_page_count
    for eo in exam_orders:
        try:
            eo._items_parsed = json.loads(eo.items_json or "[]")
        except Exception:
            eo._items_parsed = []
        for rpt in eo.reports:
            if rpt.file_type == "pdf":
                cnt = get_pdf_page_count(rpt.file_path)
                rpt.page_count = min(cnt, 5) if cnt else 0
            else:
                rpt.page_count = 1  # 图片只有 1 张

    # 体重 + 年龄（同 prescription print）
    pet_weight = 0.0
    if pet:
        last_w = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id).order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if last_w:
            pet_weight = float(last_w.weight_kg or 0)
    pet_age = ""
    if pet and pet.birthday_estimate:
        try:
            from datetime import date as _date
            parts = pet.birthday_estimate.split("-")
            by = int(parts[0]); bm = int(parts[1]) if len(parts) > 1 else 1
            today = _date.today()
            years = today.year - by - (1 if (today.month, 1) < (bm, 1) else 0)
            pet_age = (f"{years} 岁" if years > 0 else f"{max(0, (today.year - by) * 12 + (today.month - bm))} 个月")
        except Exception:
            pet_age = pet.birthday_estimate or ""

    clinic_name_zh = "大风动物医院"
    clinic_name_en = "DaFo Animal Hospital"
    if pet and pet.store:
        clinic_name_zh = f"大风动物医院（{pet.store.replace('店', '分院')}）"
        clinic_name_en = f"DaFo Animal Hospital · {pet.store.replace('店', '')}"

    return templates.TemplateResponse(request, "admin_visit_print.html", {
        "visit": visit, "cust": cust, "pet": pet,
        "prescriptions": prescriptions,
        "exam_orders": exam_orders,
        "pet_weight": pet_weight, "pet_age": pet_age,
        "clinic_name_zh": clinic_name_zh, "clinic_name_en": clinic_name_en,
        "visit_type_zh": _VISIT_TYPE_ZH,
    })


@app.get("/admin/visits/{visit_id}/discharge-print", response_class=HTMLResponse)
async def admin_visit_discharge_print(visit_id: int, request: Request, db: Session = Depends(get_db)):
    """医嘱单独立打印（突出医嘱内容 + 处方 + 复诊建议，无完整 SOAP）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    visit = db.get(Visit, visit_id)
    if not visit:
        raise HTTPException(404, "就诊记录不存在")
    cust = db.get(Customer, visit.customer_id) if visit.customer_id else None
    pet  = db.get(Pet,      visit.pet_id)      if visit.pet_id      else None
    prescriptions = db.query(Prescription).filter(Prescription.visit_id == visit_id).order_by(Prescription.id.asc()).all()

    pet_weight = 0.0
    if pet:
        last_w = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id).order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if last_w:
            pet_weight = float(last_w.weight_kg or 0)
    pet_age = ""
    if pet and pet.birthday_estimate:
        try:
            from datetime import date as _date
            parts = pet.birthday_estimate.split("-")
            by = int(parts[0]); bm = int(parts[1]) if len(parts) > 1 else 1
            today = _date.today()
            years = today.year - by - (1 if (today.month, 1) < (bm, 1) else 0)
            pet_age = (f"{years} 岁" if years > 0 else f"{max(0, (today.year - by) * 12 + (today.month - bm))} 个月")
        except Exception:
            pet_age = pet.birthday_estimate or ""

    clinic_name_zh = "大风动物医院"
    clinic_name_en = "DaFo Animal Hospital"
    if pet and pet.store:
        clinic_name_zh = f"大风动物医院（{pet.store.replace('店', '分院')}）"
        clinic_name_en = f"DaFo Animal Hospital · {pet.store.replace('店', '')}"

    return templates.TemplateResponse(request, "admin_discharge_print.html", {
        "visit": visit, "cust": cust, "pet": pet,
        "prescriptions": prescriptions,
        "pet_weight": pet_weight, "pet_age": pet_age,
        "clinic_name_zh": clinic_name_zh, "clinic_name_en": clinic_name_en,
        "visit_type_zh": _VISIT_TYPE_ZH,
    })


@app.get("/admin/visits/{visit_id}", response_class=HTMLResponse)
async def page_admin_visit_detail(
    visit_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    cust = db.get(Customer, v.customer_id) if v.customer_id else None
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == v.customer_id).all() if v.customer_id else []
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v2[0] for v2 in vets]
    prescriptions = db.query(Prescription).filter(Prescription.visit_id == visit_id).order_by(Prescription.id.desc()).all()
    anesth_orders = db.query(AnesthesiaOrder).filter(AnesthesiaOrder.visit_id == visit_id).order_by(AnesthesiaOrder.id.desc()).all()
    sales_orders = db.query(SalesOrder).filter(SalesOrder.visit_id == visit_id).order_by(SalesOrder.id.desc()).all()
    invoices = db.query(Invoice).filter(Invoice.visit_id == visit_id).order_by(Invoice.id.desc()).all()
    exam_orders = db.query(ExamOrder).filter(ExamOrder.visit_id == visit_id).order_by(ExamOrder.id.desc()).all()
    # 本 visit 的所有回访轮次（按计划日 + round_no 排序）
    followups = db.query(FollowUp).filter(FollowUp.visit_id == visit_id)\
        .order_by(FollowUp.planned_date, FollowUp.round_no).all()
    for fu in followups:
        try:
            fu._answers = json.loads(fu.response_data or "") if fu.response_data else None
        except Exception:
            fu._answers = None
    # 解析检查项目，让列表能展示开了哪些项
    for eo in exam_orders:
        try:
            eo._items_parsed = json.loads(eo.items_json or "[]")
        except Exception:
            eo._items_parsed = []
    _PRESC_STATUS_ZH = {"draft": "草稿", "issued": "已开具", "dispensed": "已发药"}
    _SO_STATUS_ZH = {"pending": "待付款", "paid": "已收款", "cancelled": "已取消"}
    # B5.1+B5.2 UK 重写：默认走只读 uk/visit.html；/edit-form 或 ?mode=edit 走 uk/visit_edit.html
    _mode = request.query_params.get("mode") or ("edit" if "/edit-form" in str(request.url.path) else "view")
    _template = "uk/visit.html" if _mode == "view" else "uk/visit_edit.html"
    return templates.TemplateResponse(request, _template, {
        "visit": v,
        "cust": cust,
        "pet": pet,
        "pets": pets,
        "vet_names": vet_names,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "prescriptions": prescriptions,
        "anesth_orders": anesth_orders,
        "sales_orders": sales_orders,
        "invoices": invoices,
        "exam_orders": exam_orders,
        "followups": followups,
        "presc_status_zh": _PRESC_STATUS_ZH,
        "so_status_zh": _SO_STATUS_ZH,
        "inv_status_zh": _INV_STATUS_ZH,
        "fu_status_zh": {
            "pending": "待回访", "due": "今日到期", "sent": "已发送等反馈",
            "responded": "客户已反馈", "phone_pending": "待联系",
            "closed": "已完成", "skipped": "已忽略",
        },
        "csrf_token": _get_csrf_token(request),
        "mode": "edit",
        "msg": request.query_params.get("msg"),
        "is_superadmin": _is_superadmin(request),
    })


@app.get("/admin/visits/{visit_id}/edit-form", response_class=HTMLResponse)
async def page_admin_visit_edit_form(visit_id: int, request: Request, db: Session = Depends(get_db)):
    """B5.1 病历编辑（暂时复用老模板，B5.2 会 UK 重写）。"""
    return await page_admin_visit_detail(visit_id, request, db)


@app.post("/admin/visits/{visit_id}/edit")
async def admin_visit_edit(
    visit_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    pet_id: int = Form(0),
    visit_date: str = Form(""),
    visit_type: str = Form("outpatient"),
    chief_complaint: str = Form(""),
    physical_exam: str = Form(""),
    diagnosis: str = Form(""),
    treatment_plan: str = Form(""),
    notes: str = Form(""),
    vet_name: str = Form(""),
    follow_up_note: str = Form(""),
    follow_up_at: str = Form(""),
    return_to: str = Form(""),  # "customer" 时保存后跳回客户档案
    next_url: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    # 病历是法定档案：基础信息（日期/类型/医生/宠物）非超管不允许改
    _new_pet_id = pet_id or v.pet_id
    _new_date   = visit_date.strip()[:20]
    _new_type   = visit_type.strip()[:40] or "outpatient"
    _new_vet    = vet_name.strip()[:80]
    _meta_changed = (
        _new_pet_id != (v.pet_id or 0) or
        _new_date != (v.visit_date or "") or
        _new_type != (v.visit_type or "") or
        _new_vet  != (v.vet_name or "")
    )
    if _meta_changed and not _is_superadmin(request):
        raise HTTPException(status_code=403, detail="病历基础信息（日期/类型/医生/宠物）仅超级管理员可修改")
    if (v.status or "open") == "closed":
        raise HTTPException(status_code=403, detail="病历已结束，不可修改（合规要求）。如需追加请新建病历。")
    v.pet_id = _new_pet_id
    v.visit_date = _new_date
    v.visit_type = _new_type
    v.chief_complaint = chief_complaint.strip()
    v.physical_exam = physical_exam.strip()
    v.diagnosis = diagnosis.strip()
    v.treatment_plan = treatment_plan.strip()
    v.notes = notes.strip()
    v.vet_name = _new_vet
    v.follow_up_note = follow_up_note.strip()
    v.follow_up_at = follow_up_at.strip()[:20]
    db.commit()
    # 同步回访任务（visit_type / follow_up_at / vet_name 可能都变了）
    _sync_followup_for_visit(db, v)
    db.commit()
    # 若来自客户档案，保存后回去
    if return_to == "customer" and v.customer_id:
        return RedirectResponse(f"/admin/customers/{v.customer_id}?pet_id={v.pet_id or 0}&tab=visits&msg=就诊已保存", status_code=303)
    return RedirectResponse(
        _safe_next(next_url, f"/admin/visits/{visit_id}?msg=已保存"),
        status_code=303,
    )


@app.post("/admin/visits/{visit_id}/close")
async def admin_visit_close(visit_id: int, request: Request,
                            csrf_token: str = Form(""),
                            db: Session = Depends(get_db)):
    """结束病历。结束后病历及关联处方/检查不可改；按合规要求不可重开。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    # 限店员工
    admin_store = _get_admin_store(request)
    if admin_store and v.pet_id:
        pet = db.get(Pet, v.pet_id)
        if pet and pet.store and pet.store != admin_store:
            raise HTTPException(403, "无权操作其他门店的就诊记录")
    if (v.status or "open") == "closed":
        return RedirectResponse(f"/admin/visits/{visit_id}?msg=该病历已是结束状态", status_code=303)
    v.status = "closed"
    v.closed_at = datetime.utcnow()
    v.closed_by = request.session.get("admin_username", "") or ""
    db.commit()
    _audit(db, request, "visit_close",
           detail={"visit_id": v.id, "pet_id": v.pet_id, "customer_id": v.customer_id})
    db.commit()
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=病历已结束", status_code=303)


@app.post("/admin/visits/{visit_id}/followup-toggle")
async def admin_visit_followup_toggle(visit_id: int, request: Request,
                                       csrf_token: str = Form(""),
                                       next_url: str = Form(""),
                                       db: Session = Depends(get_db)):
    """切换病历级别「自动回访」开关。关闭 → 已 pending 全取消、未来不衍生；重新开启 → 重新衍生。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    admin_store = _get_admin_store(request)
    if admin_store and v.pet_id:
        pet = db.get(Pet, v.pet_id)
        if pet and pet.store and pet.store != admin_store:
            raise HTTPException(403, "无权操作其他门店的就诊记录")
    v.followup_disabled = not bool(getattr(v, "followup_disabled", False))
    db.commit()
    # sync：关闭 → 删 pending；开启 → 重新衍生
    _sync_followup_for_visit(db, v)
    db.commit()
    msg = "已关闭后续回访" if v.followup_disabled else "已重新开启回访"
    fb = f"/admin/visits/{visit_id}?msg={msg}"
    return RedirectResponse(_safe_next(next_url, fb), status_code=303)


@app.post("/api/visits/{visit_id}/autosave")
async def api_visit_autosave(visit_id: int, request: Request, db: Session = Depends(get_db)):
    """SOAP 7 步工作流的实时自动保存。仅接受 JSON。只更新文本字段，不动 pet_id/date/type 等。"""
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    v = db.get(Visit, visit_id)
    if not v:
        return {"ok": False, "error": "记录不存在"}
    if (v.status or "open") == "closed":
        return {"ok": False, "error": "病历已结束，不可修改"}
    # 限店员工：只能改本店宠物的诊疗记录
    admin_store = _get_admin_store(request)
    if admin_store and v.pet_id:
        pet = db.get(Pet, v.pet_id)
        if pet and pet.store and pet.store != admin_store:
            return {"ok": False, "error": "无权操作其他门店的就诊记录"}
    # 只允许这几个字段，避免 JS 注入修改 customer/pet
    allowed = {"chief_complaint", "physical_exam", "diagnosis", "treatment_plan", "follow_up_note", "follow_up_at", "notes"}
    changed = []
    for k, val in body.items():
        if k in allowed and isinstance(val, str):
            cur = getattr(v, k, "") or ""
            if cur != val[:2000]:
                setattr(v, k, val[:2000])
                changed.append(k)
    if changed:
        v.updated_at = datetime.utcnow()
        db.commit()
        # 诊断 / 复诊日期 任一变化都要重新匹配模板 + 衍生多轮回访
        if any(k in changed for k in ("follow_up_at", "diagnosis")):
            _sync_followup_for_visit(db, v)
            db.commit()
    # 本地时间显示
    from datetime import timezone, timedelta
    cn_tz = timezone(timedelta(hours=8))
    return {"ok": True, "changed": changed, "saved_at": datetime.now(cn_tz).strftime("%H:%M:%S")}


@app.post("/admin/visits/{visit_id}/delete")
async def admin_visit_delete(
    visit_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    # 病历法定档案：只有超管能删
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    customer_id = v.customer_id
    db.delete(v)
    db.commit()
    if customer_id:
        return RedirectResponse(f"/admin/customers/{customer_id}?msg=就诊记录已删除", status_code=303)
    return RedirectResponse("/admin/visits?msg=就诊记录已删除", status_code=303)


# ---------------------------------------------------------------------------
# 回访管理 — 列表 / 操作
# ---------------------------------------------------------------------------

_FOLLOWUP_STATUS_ZH = {
    "pending":       "待回访",
    "due":           "今日到期",
    "sent":          "已发送",
    "responded":     "客户已反馈",
    "phone_pending": "待联系",
    "closed":        "已完成",
    "skipped":       "已忽略",
}

_FOLLOWUP_RESPONSE_ZH = {
    "recovered":    "已好转",
    "needs_visit":  "需复诊",
    "no_reply":     "无回应",
}

_FOLLOWUP_CHANNEL_ZH = {
    "miniapp": "小程序",
    "sms":     "短信",
    "phone":   "电话",
}


def _followup_filtered_query(db: Session, request: Request):
    """门店隔离 + 我的过滤的基础查询。"""
    q = db.query(FollowUp)
    admin_store = _get_admin_store(request)
    if admin_store:
        q = q.filter((FollowUp.store == admin_store) | (FollowUp.store == ""))
    return q


@app.get("/admin/follow-ups", response_class=HTMLResponse)
async def page_admin_follow_ups(
    request: Request,
    db: Session = Depends(get_db),
    tab: str = Query("today"),     # today / overdue / sent / done
    mine: int = Query(0),          # 1=只看我的
    page: int = Query(1),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    from datetime import date
    today = date.today().isoformat()

    base = _followup_filtered_query(db, request)
    username = request.session.get("admin_username") or ""
    if mine and username:
        base = base.filter(FollowUp.assigned_to == username)

    # 4 个 tab 的过滤条件
    if tab == "today":
        # 今日 = 计划日期 ≤ 今天 (含逾期) 且仍可操作（与工作台「待回访任务」口径一致）
        q = base.filter(FollowUp.planned_date != "",
                        FollowUp.planned_date <= today,
                        FollowUp.status.in_(["pending", "due", "phone_pending"]))
        order = (FollowUp.planned_date.asc(), FollowUp.id.desc())
    elif tab == "overdue":
        q = base.filter(FollowUp.planned_date < today,
                        FollowUp.status.in_(["pending", "due", "phone_pending"]))
        order = (FollowUp.planned_date.asc(), FollowUp.id.desc())
    elif tab == "sent":
        # 已发送 + 已反馈待处理（needs_visit 的客户反馈需要医生跟进）
        q = base.filter(FollowUp.status.in_(["sent", "responded"]))
        # 把 responded 排前面，让医生先看到需复诊的
        order = (FollowUp.response.desc(), FollowUp.sent_at.desc(), FollowUp.id.desc())
    else:  # done
        tab = "done"
        from datetime import timedelta
        cutoff = (date.today() - timedelta(days=30)).isoformat()
        q = base.filter(
            FollowUp.status.in_(["closed", "skipped"]),
            FollowUp.planned_date >= cutoff,
        )
        order = (FollowUp.updated_at.desc(), FollowUp.id.desc())

    page_size = 30
    total = q.count()
    rows = q.order_by(*order).offset((page - 1) * page_size).limit(page_size).all()
    total_pages = max(1, (total + page_size - 1) // page_size)

    # 4 个 tab 各自总数（用于 tab 上的数字徽章）
    def _count(filter_):
        qq = _followup_filtered_query(db, request)
        if mine and username:
            qq = qq.filter(FollowUp.assigned_to == username)
        return filter_(qq).count()
    counts = {
        "today":   _count(lambda x: x.filter(FollowUp.planned_date != "", FollowUp.planned_date <= today, FollowUp.status.in_(["pending", "due", "phone_pending"]))),
        "overdue": _count(lambda x: x.filter(FollowUp.planned_date < today, FollowUp.status.in_(["pending", "due", "phone_pending"]))),
        "sent":    _count(lambda x: x.filter(FollowUp.status.in_(["sent", "responded"]))),
        "done":    _count(lambda x: x.filter(FollowUp.status.in_(["closed", "skipped"]))),
        "needs_visit": _count(lambda x: x.filter(FollowUp.status == "responded", FollowUp.response == "needs_visit")),
    }

    # 预取 visit / pet / customer 信息（避免模板里 N+1）
    visit_ids = [r.visit_id for r in rows]
    visits = {v.id: v for v in db.query(Visit).filter(Visit.id.in_(visit_ids)).all()} if visit_ids else {}
    pet_ids = list({r.pet_id for r in rows if r.pet_id})
    pets = {p.id: p for p in db.query(Pet).filter(Pet.id.in_(pet_ids)).all()} if pet_ids else {}
    cust_ids = list({r.customer_id for r in rows if r.customer_id})
    custs = {c.id: c for c in db.query(Customer).filter(Customer.id.in_(cust_ids)).all()} if cust_ids else {}

    return templates.TemplateResponse(request, "uk/follow_ups.html", {
        "title": "回访管理",
        "rows": rows,
        "visits": visits,
        "pets": pets,
        "customers": custs,
        "counts": counts,
        "tab": tab,
        "mine": mine,
        "page": page,
        "total_pages": total_pages,
        "total": total,
        "status_zh":   _FOLLOWUP_STATUS_ZH,
        "response_zh": _FOLLOWUP_RESPONSE_ZH,
        "channel_zh":  _FOLLOWUP_CHANNEL_ZH,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "today_str": today,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/follow-ups/{fu_id}/handle")
async def admin_followup_handle(
    fu_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    action: str = Form(...),       # contacted / refer_visit / skip / reopen
    note: str = Form(""),
    tab_redirect: str = Form("today"),
    next_url: str = Form(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    fu = db.get(FollowUp, fu_id)
    if not fu:
        raise HTTPException(404, "回访任务不存在")
    # 门店隔离
    admin_store = _get_admin_store(request)
    if admin_store and fu.store and fu.store != admin_store:
        raise HTTPException(403, "无权操作其他门店")
    username = request.session.get("admin_username") or "admin"
    now = datetime.utcnow()
    if action == "contacted":
        fu.status = "closed"
        fu.response = fu.response or "recovered"
        fu.handled_by = username
        fu.handled_at = now
        fu.handle_note = note.strip()[:500]
    elif action == "refer_visit":
        fu.status = "responded"
        fu.response = "needs_visit"
        fu.response_at = now
        fu.handled_by = username
        fu.handled_at = now
        fu.handle_note = note.strip()[:500]
    elif action == "skip":
        fu.status = "skipped"
        fu.handled_by = username
        fu.handled_at = now
        fu.handle_note = note.strip()[:500]
    elif action == "reopen":
        fu.status = "pending"
        fu.response = ""
        fu.response_at = None
        fu.handle_note = ""
        fu.handled_at = None
    else:
        raise HTTPException(400, f"未知操作 {action}")
    fu.updated_at = now
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/follow-ups?tab={tab_redirect}&msg=已更新"),
        status_code=303,
    )


# ═════════════════════════════════════════════════════════════════
# 回访模板管理（C）
# ═════════════════════════════════════════════════════════════════
_QUESTION_TYPES = {
    "scale1to5": "1-5 评分",
    "select":    "单选",
    "multi":     "多选",
    "number":    "数字（如体重）",
    "text":      "文本",
    "upload":    "上传照片",
}


@app.get("/admin/follow-up-templates", response_class=HTMLResponse)
async def page_admin_futpl_list(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    from app.data.vet_seed import SYSTEMS as _SYS
    tpls = db.query(FollowUpTemplate).order_by(
        FollowUpTemplate.is_active.desc(),
        FollowUpTemplate.priority.desc(),
        FollowUpTemplate.system,
        FollowUpTemplate.name,
    ).all()
    # 解析 rounds_json 用于列表显示
    import json as _json
    for t in tpls:
        try:
            t._rounds = _json.loads(t.rounds_json or "[]")
        except Exception:
            t._rounds = []
    return templates.TemplateResponse(request, "uk/follow_up_templates.html", {
        "tpls": tpls, "systems": _SYS,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.get("/admin/follow-up-templates/new", response_class=HTMLResponse)
async def page_admin_futpl_new(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    from app.data.vet_seed import SYSTEMS as _SYS
    return templates.TemplateResponse(request, "uk/follow_up_template_form.html", {
        "tpl": None, "rounds": [], "systems": _SYS,
        "question_types": _QUESTION_TYPES,
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/follow-up-templates/{tpl_id}/edit", response_class=HTMLResponse)
async def page_admin_futpl_edit(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    tpl = db.get(FollowUpTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404)
    from app.data.vet_seed import SYSTEMS as _SYS
    import json as _json
    try:
        rounds = _json.loads(tpl.rounds_json or "[]")
    except Exception:
        rounds = []
    return templates.TemplateResponse(request, "uk/follow_up_template_form.html", {
        "tpl": tpl, "rounds": rounds, "systems": _SYS,
        "question_types": _QUESTION_TYPES,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/follow-up-templates/save")
async def admin_futpl_save(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    tpl_id = int(form.get("tpl_id", 0) or 0)
    name = str(form.get("name", "")).strip()[:120]
    system = str(form.get("system", "")).strip()[:40]
    keywords = str(form.get("keywords", "")).strip()
    try:
        priority = int(form.get("priority", 50) or 50)
    except Exception:
        priority = 50
    rounds_json = str(form.get("rounds_json", "[]")).strip() or "[]"
    is_active = str(form.get("is_active", "0")) in ("1", "true", "on")
    notes = str(form.get("notes", "")).strip()

    if not name:
        raise HTTPException(400, "模板名为必填")

    # 校验 rounds_json 是合法 JSON
    import json as _json
    try:
        rounds = _json.loads(rounds_json)
        assert isinstance(rounds, list)
        for rnd in rounds:
            assert isinstance(rnd, dict)
            assert "day_offset" in rnd
    except Exception as e:
        raise HTTPException(400, f"轮次配置 JSON 不合法：{e}")

    if tpl_id:
        tpl = db.get(FollowUpTemplate, tpl_id)
        if not tpl:
            raise HTTPException(404)
        # 防止重名
        dup = db.query(FollowUpTemplate).filter(
            FollowUpTemplate.name == name, FollowUpTemplate.id != tpl_id
        ).first()
        if dup:
            raise HTTPException(400, "已有同名模板")
        tpl.name = name
        tpl.system = system
        tpl.keywords = keywords
        tpl.priority = priority
        tpl.rounds_json = rounds_json
        tpl.is_active = is_active
        tpl.notes = notes
        tpl.updated_at = datetime.utcnow()
    else:
        if db.query(FollowUpTemplate).filter(FollowUpTemplate.name == name).first():
            raise HTTPException(400, "已有同名模板")
        tpl = FollowUpTemplate(
            name=name, system=system, keywords=keywords,
            priority=priority, rounds_json=rounds_json,
            is_active=is_active, is_builtin=False, notes=notes,
        )
        db.add(tpl)
    db.commit()
    return RedirectResponse(f"/admin/follow-up-templates?msg=已保存：{name}", status_code=303)


@app.post("/admin/follow-up-templates/{tpl_id}/toggle")
async def admin_futpl_toggle(tpl_id: int, request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form("")):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    tpl = db.get(FollowUpTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404)
    tpl.is_active = not tpl.is_active
    db.commit()
    return RedirectResponse(f"/admin/follow-up-templates?msg={'已启用' if tpl.is_active else '已禁用'}：{tpl.name}", status_code=303)


@app.post("/admin/follow-up-templates/{tpl_id}/delete")
async def admin_futpl_delete(tpl_id: int, request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form("")):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    tpl = db.get(FollowUpTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404)
    if tpl.is_builtin:
        raise HTTPException(400, "内置模板不可删除，请改为禁用")
    name = tpl.name
    db.delete(tpl)
    db.commit()
    return RedirectResponse(f"/admin/follow-up-templates?msg=已删除：{name}", status_code=303)


# ─── 客户反馈短链（无登录，token 校验） ───────────────────────
def _load_followup_questions(db: Session, fu: FollowUp) -> list:
    """根据 FollowUp 的 template_id + round_no 拿到本轮要问的问题列表。

    模板被删/找不到时返回空列表，前端会用兜底的「2 选 1」表单。
    """
    if not fu.template_id:
        return []
    tpl = db.get(FollowUpTemplate, fu.template_id)
    if not tpl:
        return []
    import json as _json
    try:
        rounds = _json.loads(tpl.rounds_json or "[]")
    except Exception:
        return []
    idx = max(0, (fu.round_no or 1) - 1)
    if idx >= len(rounds):
        return []
    return rounds[idx].get("questions", []) or []


@app.get("/follow-up/{token}", response_class=HTMLResponse)
async def page_followup_feedback(token: str, request: Request, db: Session = Depends(get_db)):
    fu = db.query(FollowUp).filter(FollowUp.feedback_token == token).first()
    if not fu:
        raise HTTPException(404, "反馈链接已失效")
    pet = db.get(Pet, fu.pet_id) if fu.pet_id else None
    cust = db.get(Customer, fu.customer_id) if fu.customer_id else None
    visit = db.get(Visit, fu.visit_id) if fu.visit_id else None
    questions = _load_followup_questions(db, fu)
    return templates.TemplateResponse(request, "follow_up_feedback.html", {
        "fu": fu, "pet": pet, "cust": cust, "visit": visit,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "questions": questions,
        "submitted": False,
        "csrf_token": "",
    })


@app.post("/follow-up/{token}", response_class=HTMLResponse)
async def submit_followup_feedback(token: str, request: Request, db: Session = Depends(get_db)):
    fu = db.query(FollowUp).filter(FollowUp.feedback_token == token).first()
    if not fu:
        raise HTTPException(404, "反馈链接已失效")
    form = await request.form()
    questions = _load_followup_questions(db, fu)

    # ── 收集结构化答案 ──
    import json as _json
    answers: dict = {}
    photo_urls: list = []
    for q in questions:
        key = q.get("key")
        qtype = q.get("type")
        if not key:
            continue
        if qtype == "upload":
            files = form.getlist(f"q_{key}_files")
            saved = await _save_followup_photos(token, files, q.get("max", 3))
            if saved:
                answers[key] = saved
                photo_urls.extend(saved)
        else:
            raw = form.get(f"q_{key}", "")
            if raw is None:
                continue
            val = str(raw).strip()
            if not val:
                continue
            if qtype == "multi":
                answers[key] = [s for s in val.split("|") if s]
            elif qtype == "number":
                try:
                    answers[key] = float(val)
                except Exception:
                    answers[key] = val
            elif qtype == "scale1to5":
                try:
                    answers[key] = int(val)
                except Exception:
                    answers[key] = val
            else:
                answers[key] = val[:1000]

    # ── 兜底：旧模板/无模板时按 q_status 走 #
    if not questions:
        st = str(form.get("q_status", "")).strip()
        note = str(form.get("q_note", "")).strip()
        if st == "recovered":
            fu.status = "closed"; fu.response = "recovered"
        elif st == "needs_visit":
            fu.status = "responded"; fu.response = "needs_visit"
        fu.response_note = note[:500]
        fu.response_at = datetime.utcnow()
        fu.updated_at = datetime.utcnow()
        db.commit()
    else:
        # ── 推断高层 status：看是否有 needs_visit 类信号 ──
        nv = answers.get("needs_visit") or answers.get("status") or ""
        if isinstance(nv, str) and ("复诊" in nv or "紧急" in nv or "立即" in nv):
            fu.response = "needs_visit"
            fu.status = "responded"
        else:
            fu.response = "recovered"
            fu.status = "closed"
        # response_note: 把 'note' 答案显式同步出来，列表更直观
        note_val = answers.get("note") or ""
        if note_val:
            fu.response_note = str(note_val)[:500]
        fu.response_data = _json.dumps(answers, ensure_ascii=False)
        fu.response_at = datetime.utcnow()
        fu.updated_at = datetime.utcnow()
        db.commit()

        # ── 客户标紧急 / 需复诊 → 立即推送主治医师（企微，best-effort） ──
        if fu.response == "needs_visit":
            try:
                _push_urgent_feedback_to_vet(db, fu, answers)
            except Exception as e:
                logger.warning(f"urgent push failed for fu#{fu.id}: {e}")

    pet = db.get(Pet, fu.pet_id) if fu.pet_id else None
    cust = db.get(Customer, fu.customer_id) if fu.customer_id else None
    visit = db.get(Visit, fu.visit_id) if fu.visit_id else None
    return templates.TemplateResponse(request, "follow_up_feedback.html", {
        "fu": fu, "pet": pet, "cust": cust, "visit": visit,
        "visit_type_zh": _VISIT_TYPE_ZH,
        "questions": [],
        "submitted": True,
        "csrf_token": "",
    })


async def _save_followup_photos(token: str, files: list, max_count: int = 3) -> list[str]:
    """保存客户反馈附带的照片，返回 URL 列表。

    存储位置：uploads/followup/{token}/<idx>_<safename>
    """
    from pathlib import Path as _P
    import re as _re
    saved_urls: list[str] = []
    base = _P("uploads") / "followup" / token
    base.mkdir(parents=True, exist_ok=True)
    n = 0
    for f in files:
        if n >= max_count:
            break
        try:
            content = await f.read()
        except Exception:
            continue
        if not content or len(content) > 8 * 1024 * 1024:  # 8MB
            continue
        safename = _re.sub(r"[^a-zA-Z0-9._-]", "_", getattr(f, "filename", "img.jpg"))[:64] or "img.jpg"
        path = base / f"{n}_{safename}"
        try:
            with open(path, "wb") as out:
                out.write(content)
            saved_urls.append(f"/uploads/followup/{token}/{n}_{safename}")
            n += 1
        except Exception:
            continue
    return saved_urls


def _push_urgent_feedback_to_vet(db: Session, fu: FollowUp, answers: dict) -> None:
    """客户反馈需复诊 / 紧急 → 推送主治医师企微消息。
    需 AdminUser.wecom_userid 已绑定。
    """
    try:
        from app.services.wecom_client import send_app_message as _send
    except Exception:
        return
    assignee = (fu.assigned_to or "").strip()
    if not assignee:
        return
    user = db.query(AdminUser).filter(AdminUser.username == assignee).first()
    if not user or not user.wecom_userid:
        return
    pet = db.get(Pet, fu.pet_id) if fu.pet_id else None
    cust = db.get(Customer, fu.customer_id) if fu.customer_id else None
    pet_name = pet.name if pet else "客户宠物"
    cust_name = cust.name if cust else "客户"
    phone = (cust.phone if cust else "") or ""
    parts = [f"⚠️ 客户反馈紧急 / 需复诊：{pet_name}（{cust_name}）"]
    if phone:
        parts.append(f"电话：{phone}")
    if fu.template_name:
        parts.append(f"回访模板：{fu.template_name} · {fu.round_name or ''}")
    # 关键答案摘要
    for k in ("vomit", "stool", "cough", "breath", "wound", "needs_visit", "note"):
        v = answers.get(k)
        if v:
            parts.append(f"· {k}: {v}")
    parts.append(f"详情：https://dafopet.com/admin/follow-ups?tab=overdue")
    try:
        _send(user.wecom_userid, "\n".join(parts))
    except Exception:
        pass


@app.get("/api/follow-ups/badge")
async def api_followup_badge(request: Request, db: Session = Depends(get_db)):
    """工作台/导航栏轮询用：返回今日待回访 + 逾期 + 需电话兜底 的总数。"""
    if not request.session.get("admin"):
        return {"count": 0}
    from datetime import date
    today = date.today().isoformat()
    base = _followup_filtered_query(db, request)
    username = request.session.get("admin_username") or ""
    mine = base.filter(FollowUp.assigned_to == username) if username else base
    n_today = mine.filter(
        FollowUp.planned_date == today,
        FollowUp.status.in_(["pending", "due"]),
    ).count()
    n_overdue = mine.filter(
        FollowUp.planned_date < today,
        FollowUp.status.in_(["pending", "due", "phone_pending"]),
    ).count()
    # 客户反馈"需复诊"的也算紧急（医生需要主动联系排期）
    n_needs = mine.filter(
        FollowUp.status == "responded",
        FollowUp.response == "needs_visit",
    ).count()
    return {
        "count":   n_today + n_overdue + n_needs,
        "today":   n_today,
        "overdue": n_overdue,
        "needs_visit": n_needs,
    }


# ---------------------------------------------------------------------------
# Phase 3 — 处方单 (Prescriptions)
# ---------------------------------------------------------------------------

_PRESC_STATUS_ZH = {"draft": "草稿", "issued": "已开具", "dispensed": "已发药"}
_DRUG_TYPE_ZH = {
    "oral": "口服", "topical": "外用",
    "subcutaneous": "皮下", "intramuscular": "肌肉", "intravenous": "静脉",
    "injection": "注射（其他）",
    "eye_drop": "滴眼", "ear_drop": "滴耳", "nasal": "鼻喷",
    "nebulize": "雾化", "enema": "灌肠", "rectal": "直肠给药",
    "service": "处置 / 服务", "other": "其他",
}


def _customer_is_internal(db: Session, customer_id: int) -> bool:
    """检查客户是否为员工内购档案。"""
    if not customer_id:
        return False
    c = db.get(Customer, customer_id)
    return bool(c and c.is_internal)


def _apply_internal_pricing(db: Session, items: list[dict], customer_id: int) -> bool:
    """如果客户是员工内购档案，把 items 列表里每行的 unit_price 替换为 InventoryItem.cost_price，
    并重算 subtotal。返回 True 表示已应用内购价。
    items 字典需要至少包含 item_id / quantity_num / unit_price / subtotal 键。"""
    if not _customer_is_internal(db, customer_id):
        return False
    for it in items:
        iid = it.get("item_id")
        if not iid:
            continue
        inv = db.get(InventoryItem, int(iid))
        if not inv:
            continue
        cost = float(inv.cost_price or 0)
        # 数量字段在 prescription_items=quantity_num，sales_items=quantity，exam_items=qty
        qty = float(it.get("quantity_num") or it.get("quantity") or it.get("qty") or 1)
        it["unit_price"] = cost
        it["subtotal"] = round(qty * cost, 2)
    return True


def _apply_single_use_pack_billing(db: Session, items: list[dict]) -> None:
    """整支/整瓶计费：扫描 items，若品目 single_use_pack=True，则把数量向上取整到副单位整数倍，并重算 subtotal。
    数量字段兼容：prescription_items=quantity_num，sales_items=quantity（数字版 quantity_num/qty），exam_items=qty
    """
    for it in items:
        iid = it.get("item_id")
        if not iid:
            continue
        inv = db.get(InventoryItem, int(iid))
        if not inv or not getattr(inv, "single_use_pack", False):
            continue
        # 取数字 qty 字段
        qty_keys = ("quantity_num", "quantity", "qty")
        qty_key = None
        for k in qty_keys:
            v = it.get(k)
            if isinstance(v, (int, float)) and v > 0:
                qty_key = k; break
        if qty_key is None:
            continue
        old_qty = float(it[qty_key] or 0)
        new_qty = _billable_qty(inv, old_qty)
        if abs(new_qty - old_qty) < 1e-6:
            continue
        it[qty_key] = new_qty
        # 重算 subtotal
        up = float(it.get("unit_price") or 0)
        it["subtotal"] = round(new_qty * up, 2)
        # 销售单的 quantity 字符串字段也同步（如有）
        if "quantity" in it and not isinstance(it.get("quantity"), (int, float)):
            it["quantity"] = f"{new_qty:g}"


def _billable_qty(item: "InventoryItem | None", qty: float) -> float:
    """整支/整瓶计费：开 0.1ml = 开 1 整支 → 向上取整到副单位整数倍。
    - item.single_use_pack=False 或 item 为空 → 原样返回
    - ratio = max(1.0, unit2_ratio)；qty 向上取整到 ratio 的整数倍
    - 例：unit2_ratio=1.0（1支=1ml），qty=0.1 → 1.0；qty=1.5 → 2.0
    """
    if not item or not getattr(item, "single_use_pack", False):
        return qty
    try:
        ratio = float(item.unit2_ratio or 1.0)
    except (TypeError, ValueError):
        ratio = 1.0
    if ratio <= 0:
        ratio = 1.0
    import math as _math
    units = _math.ceil((qty - 1e-9) / ratio) if qty > 0 else 0
    return round(units * ratio, 4)


def _parse_presc_items(form_data) -> list[dict]:
    items = []
    i = 0
    while True:
        name = form_data.get(f"drug_name_{i}", "").strip()
        if not name and i > 20:
            break
        if name:
            try:
                qty_num = float(form_data.get(f"quantity_num_{i}", 1) or 1)
                unit_price = float(form_data.get(f"unit_price_{i}", 0) or 0)
            except ValueError:
                qty_num, unit_price = 1.0, 0.0
            raw_item_id = form_data.get(f"item_id_{i}", "").strip()
            item_id = int(raw_item_id) if raw_item_id and raw_item_id.isdigit() else None
            subtotal = round(qty_num * unit_price, 2)
            try:
                dose_amount = float(form_data.get(f"dose_amount_{i}", 0) or 0)
                times_per_day = float(form_data.get(f"times_per_day_{i}", 0) or 0)
            except ValueError:
                dose_amount, times_per_day = 0.0, 0.0
            # 把 dose_amount + dose_unit 合成回 dosage（向后兼容显示）
            dose_unit = form_data.get(f"dose_unit_{i}", "").strip()
            dosage_legacy = form_data.get(f"dosage_{i}", "").strip()
            if not dosage_legacy and dose_amount:
                dosage_legacy = (f"{dose_amount:g}{dose_unit}").strip()
            # frequency 文本：若有 times_per_day，自动生成"每日 N 次"
            freq_legacy = form_data.get(f"frequency_{i}", "").strip()
            if not freq_legacy and times_per_day:
                freq_legacy = f"每日{times_per_day:g}次"
            items.append({
                "item_id": item_id,
                "drug_name": name,
                "drug_type": form_data.get(f"drug_type_{i}", "oral").strip(),
                "dosage": dosage_legacy,
                "frequency": freq_legacy,
                "duration_days": form_data.get(f"duration_days_{i}", "").strip(),
                "quantity_num": qty_num,
                "quantity": form_data.get(f"quantity_{i}", "").strip(),
                "unit_price": unit_price,
                "subtotal": subtotal,
                "instructions": form_data.get(f"instructions_{i}", "").strip(),
                "dose_amount": dose_amount,
                "dose_unit": dose_unit,
                "times_per_day": times_per_day,
                "item_unit": form_data.get(f"item_unit_{i}", "").strip(),
                "print_note": form_data.get(f"print_note_{i}", "").strip(),
                "schedule_times": form_data.get(f"schedule_times_{i}", "").strip(),
            })
        i += 1
    return items


def _deduct_inventory(db: Session, item_id: int, qty: float, ref_type: str, ref_id: int, operator: str, note: str = "") -> None:
    """出库：减少库存，写流水。"""
    inv = db.get(InventoryItem, item_id)
    if not inv or inv.is_service:
        return
    before = inv.stock_qty
    inv.stock_qty = round(before - qty, 4)
    db.add(InventoryTransaction(
        item_id=item_id, tx_type="out", qty=qty,
        qty_before=before, qty_after=inv.stock_qty,
        unit_price=inv.sell_price,
        ref_type=ref_type, ref_id=ref_id,
        operator=operator, note=note,
    ))


def _restore_inventory(db: Session, item_id: int, qty: float, ref_type: str, ref_id: int, operator: str, note: str = "") -> None:
    """退回库存（删单时）：增加库存，写流水。"""
    inv = db.get(InventoryItem, item_id)
    if not inv or inv.is_service:
        return
    before = inv.stock_qty
    inv.stock_qty = round(before + qty, 4)
    db.add(InventoryTransaction(
        item_id=item_id, tx_type="return", qty=qty,
        qty_before=before, qty_after=inv.stock_qty,
        unit_price=inv.sell_price,
        ref_type=ref_type, ref_id=ref_id,
        operator=operator, note=note,
    ))


# ════════════════════════════════════════════════════════════════════
# 单据锁定：医疗 / 财务档案的不可篡改保护
#
# 锁定规则：
#   - 处方单：dispensed (已发药) 或 关联 Visit 有已付 Invoice → 锁
#   - 麻醉单：issued (已开具) 或 关联 Visit 有已付 Invoice → 锁
#   - 销售单：paid 或 关联 Visit 有已付 Invoice → 锁
#   - 检查单：有 ExamReport 或 关联 Visit 有已付 Invoice → 锁（仅项目，报告仍可上传）
#   - 疫苗单：invoice_id 对应 Invoice 已付 → 锁
#   - 驱虫单：invoice_id 对应 Invoice 已付 → 锁
#   - 狂犬疫苗：cert_no 已填 → 锁
#   - 任何 status='voided' 都视为锁
#
# 锁定后允许的操作：
#   1. 看 / 打印
#   2. 复制为新单（以本单为模板新建一张）
#   3. 作废（status→voided + 库存回退 + 写审计日志）
# ════════════════════════════════════════════════════════════════════

def _invoice_paid_for_visit(db: Session, visit_id: int) -> bool:
    """Visit 是否有任何 payment_status='paid' 的 Invoice。"""
    if not visit_id:
        return False
    return db.query(Invoice).filter(
        Invoice.visit_id == visit_id,
        Invoice.payment_status == "paid",
    ).first() is not None


def _is_prescription_locked(db: Session, p: "Prescription") -> tuple[bool, str]:
    if not p:
        return False, ""
    if getattr(p, "status", "") == "voided":
        return True, "已作废"
    if getattr(p, "status", "") == "dispensed":
        return True, "已发药"
    if p.visit_id and _invoice_paid_for_visit(db, p.visit_id):
        return True, "关联收费单已付款"
    return False, ""


def _is_sales_order_locked(db: Session, so: "SalesOrder") -> tuple[bool, str]:
    if not so:
        return False, ""
    if getattr(so, "status", "") == "voided":
        return True, "已作废"
    if getattr(so, "status", "") == "paid":
        return True, "已付款"
    if so.visit_id and _invoice_paid_for_visit(db, so.visit_id):
        return True, "关联收费单已付款"
    return False, ""


def _is_anesthesia_locked(db: Session, a: "AnesthesiaOrder") -> tuple[bool, str]:
    """对齐处方单：voided / 关联收费已付 才锁。issued 本身不锁，可删。"""
    if not a:
        return False, ""
    if getattr(a, "status", "") == "voided":
        return True, "已作废"
    if a.visit_id and _invoice_paid_for_visit(db, a.visit_id):
        return True, "关联收费单已付款"
    return False, ""


def _is_exam_order_locked(db: Session, eo: "ExamOrder") -> tuple[bool, str]:
    """检查单特殊：报告已上传或付款 → 锁项目；但报告仍可继续上传。"""
    if not eo:
        return False, ""
    if getattr(eo, "status", "") == "voided":
        return True, "已作废"
    has_report = db.query(ExamReport).filter(ExamReport.exam_order_id == eo.id).first() is not None
    if has_report:
        return True, "检查报告已上传"
    if eo.visit_id and _invoice_paid_for_visit(db, eo.visit_id):
        return True, "关联收费单已付款"
    return False, ""


def _is_vaccination_locked(db: Session, v: "Vaccination") -> tuple[bool, str]:
    if not v:
        return False, ""
    if getattr(v, "status", "") == "voided":
        return True, "已作废"
    if v.invoice_id:
        inv = db.get(Invoice, v.invoice_id)
        if inv and inv.payment_status == "paid":
            return True, "关联收费单已付款"
    return False, ""


def _is_deworming_locked(db: Session, d: "DewormingRecord") -> tuple[bool, str]:
    if not d:
        return False, ""
    if getattr(d, "status", "") == "voided":
        return True, "已作废"
    if getattr(d, "invoice_id", None):
        inv = db.get(Invoice, d.invoice_id)
        if inv and inv.payment_status == "paid":
            return True, "关联收费单已付款"
    return False, ""


def _is_grooming_locked(db: Session, g: "GroomingOrder") -> tuple[bool, str]:
    if not g:
        return False, ""
    if getattr(g, "status", "") == "voided":
        return True, "已作废"
    if getattr(g, "invoice_id", None):
        inv = db.get(Invoice, g.invoice_id)
        if inv and inv.payment_status == "paid":
            return True, "关联收费单已付款"
    return False, ""


def _is_rabies_locked(db: Session, r: "RabiesVaccineRecord") -> tuple[bool, str]:
    if not r:
        return False, ""
    if getattr(r, "cert_no", ""):
        return True, "免疫证号已上传"
    return False, ""


def _doc_paid_amount(db: Session, doc_type: str, doc_id: int) -> float:
    """通过 InvoiceItem.ref_type+ref_id 反查该单据在已付 Invoice 中的小计金额。

    返回 0 表示客户未付（拒绝退款入口），返回正值 = 可退到客户钱包的金额。
    """
    if not doc_id or not doc_type:
        return 0.0
    rows = (
        db.query(InvoiceItem)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .filter(
            InvoiceItem.ref_type == doc_type,
            InvoiceItem.ref_id == doc_id,
            Invoice.payment_status == "paid",
        )
        .all()
    )
    return round(sum((r.subtotal or 0) for r in rows), 2)


def _refund_to_wallet(
    db: Session, customer_id: int, amount: float, operator: str,
    note: str = "", invoice_id: int | None = None, store: str = "",
) -> "WalletTransaction | None":
    """退款入钱包：自动建钱包（如缺）+ 写 WalletTransaction(type=refund)。

    返回写入的 WalletTransaction，失败返回 None。
    """
    if not customer_id or amount <= 0:
        return None
    try:
        wallet = db.query(Wallet).filter(Wallet.customer_id == customer_id).first()
        if not wallet:
            wallet = Wallet(customer_id=customer_id, balance=0.0,
                            balance_principal=0.0, balance_bonus=0.0,
                            lifetime_recharge=0.0, lifetime_consume=0.0)
            db.add(wallet)
            db.flush()
        amount = round(float(amount), 2)
        wallet.balance = round((wallet.balance or 0) + amount, 2)
        # 退款全部进本金桶（赠送部分不退）
        wallet.balance_principal = round((wallet.balance_principal or 0) + amount, 2)
        wallet.updated_at = datetime.utcnow()
        tx = WalletTransaction(
            wallet_id=wallet.id,
            customer_id=customer_id,
            type="refund",
            amount=amount,
            balance_after=wallet.balance,
            invoice_id=invoice_id,
            store=store,
            note=note[:500],
            operator=operator,
        )
        db.add(tx)
        return tx
    except Exception as e:
        logger.warning(f"refund_to_wallet failed: {e}")
        return None


def _audit_doc_action(db: Session, doc_type: str, doc_id: int, action: str,
                       operator: str, reason: str = "", extra: str = "") -> None:
    """单据锁定/作废审计日志，写入 AuditLog 表。

    action: void / unlock / copy_from
    """
    try:
        db.add(AuditLog(
            application_id=None,
            actor=operator or "system",
            action=f"{doc_type}.{action}",
            detail=(f"id={doc_id} reason={reason} {extra}").strip()[:500],
        ))
    except Exception:
        pass


@app.get("/admin/prescriptions/create", response_class=HTMLResponse)
async def page_admin_presc_create(
    request: Request,
    db: Session = Depends(get_db),
    visit_id: int = Query(0),
    customer_id: int = Query(0),
    pet_id: int = Query(0),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    visit = db.get(Visit, visit_id) if visit_id else None
    if visit:
        customer_id = customer_id or visit.customer_id or 0
        pet_id = pet_id or visit.pet_id or 0
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == customer_id).all() if customer_id else []
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v[0] for v in vets]
    today = datetime.utcnow().strftime("%Y-%m-%d")
    history = []
    if pet_id:
        history = db.query(Prescription).filter(Prescription.pet_id == pet_id)\
            .order_by(Prescription.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/prescription.html", {  # B8.6 UK 重写
        "presc": None, "visit": visit, "cust": cust, "pet": pet, "pets": pets,
        "vet_names": vet_names, "drug_type_zh": _DRUG_TYPE_ZH,
        "presc_status_zh": _PRESC_STATUS_ZH,
        "presc_history": history,
        "today": today, "csrf_token": _get_csrf_token(request), "mode": "create",
    })


@app.post("/admin/prescriptions/create")
async def admin_presc_create(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    visit_id = int(form.get("visit_id", 0) or 0)
    customer_id = int(form.get("customer_id", 0) or 0)
    pet_id = int(form.get("pet_id", 0) or 0)
    # 病历已结束 → 不能再开处方
    if visit_id:
        _v = db.get(Visit, visit_id)
        if _v and (_v.status or "open") == "closed":
            raise HTTPException(403, "该病历已结束，不可新增处方；如需追加请新建病历")
    parsed_items = _parse_presc_items(form)
    # 整支/整瓶计费：开 0.1ml = 开 1 整支（向上取整到副单位整数倍）
    _apply_single_use_pack_billing(db, parsed_items)
    # 员工内购档案：单价改填进价
    _apply_internal_pricing(db, parsed_items, customer_id)
    total = round(sum(it["subtotal"] for it in parsed_items), 2)
    operator = request.session.get("admin_username", "admin")
    presc = Prescription(
        visit_id=visit_id or None,
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        prescribed_date=str(form.get("prescribed_date", "")).strip()[:20],
        vet_name=str(form.get("vet_name", "")).strip()[:80],
        status=str(form.get("status", "issued")).strip(),
        total_amount=total,
        notes=str(form.get("notes", "")).strip(),
        created_by=operator,
    )
    db.add(presc)
    db.flush()
    for it in parsed_items:
        db.add(PrescriptionItem(prescription_id=presc.id, **it))
        if it["item_id"] and it["quantity_num"] > 0:
            _deduct_inventory(db, it["item_id"], it["quantity_num"],
                              "prescription", presc.id, operator, f"处方#{presc.id}")
    db.commit()
    # 自动同步到收费单（草稿）
    if visit_id and presc.status != "draft":
        _sync_visit_invoice(db, visit_id, operator)
        db.commit()
    # 关联住院 → 自动生成发药任务
    try:
        if _generate_med_logs_for_prescription(db, presc) > 0:
            db.commit()
    except Exception:
        logger.exception("[med-logs] generate after presc create failed")
    # 移动端 next_url 支持，{id} 占位 = 新建出的处方 id
    next_url_raw = str(form.get("next_url") or "")
    if next_url_raw:
        nu = next_url_raw.replace("{id}", str(presc.id))
        fallback = f"/admin/visits/{visit_id}?msg=处方单已开具" if visit_id else f"/admin/prescriptions/{presc.id}?msg=处方单已创建"
        return RedirectResponse(_safe_next(nu, fallback), status_code=303)
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=处方单已开具" if visit_id else f"/admin/prescriptions/{presc.id}?msg=处方单已创建", status_code=303)


@app.get("/admin/prescriptions/{presc_id}", response_class=HTMLResponse)
async def page_admin_presc_detail(presc_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    presc = db.get(Prescription, presc_id)
    if not presc:
        raise HTTPException(404, "处方单不存在")
    visit = db.get(Visit, presc.visit_id) if presc.visit_id else None
    cust = db.get(Customer, presc.customer_id) if presc.customer_id else None
    pet = db.get(Pet, presc.pet_id) if presc.pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == presc.customer_id).all() if presc.customer_id else []
    vets = db.query(Staff.name).filter(Staff.status.in_(["active", "probation"]), Staff.position.ilike("%医%")).all()
    vet_names = [v[0] for v in vets]
    locked, lock_reason = _is_prescription_locked(db, presc)
    paid_amount = _doc_paid_amount(db, "prescription", presc_id) if locked else 0.0
    history = []
    if presc.pet_id:
        history = db.query(Prescription).filter(
            Prescription.pet_id == presc.pet_id,
            Prescription.id != presc_id,
        ).order_by(Prescription.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/prescription.html", {  # B8.6 UK 重写
        "presc": presc, "visit": visit, "cust": cust, "pet": pet, "pets": pets,
        "vet_names": vet_names, "drug_type_zh": _DRUG_TYPE_ZH,
        "presc_status_zh": _PRESC_STATUS_ZH,
        "presc_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "csrf_token": _get_csrf_token(request), "mode": "edit",
        "msg": request.query_params.get("msg"),
    })


@app.get("/admin/prescriptions/{presc_id}/print", response_class=HTMLResponse)
async def admin_presc_print(presc_id: int, request: Request, db: Session = Depends(get_db)):
    """处方笺打印页（国标 A5 横版）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    presc = db.get(Prescription, presc_id)
    if not presc:
        raise HTTPException(404, "处方单不存在")
    visit = db.get(Visit, presc.visit_id) if presc.visit_id else None
    cust = db.get(Customer, presc.customer_id) if presc.customer_id else None
    pet = db.get(Pet, presc.pet_id) if presc.pet_id else None
    # 最近一次体重（用于处方笺顶部"体重"字段）
    pet_weight = 0.0
    if pet:
        last_w = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id).order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if last_w:
            pet_weight = float(last_w.weight_kg or 0)
    # 年龄字符串（用 birthday_estimate 推断）
    pet_age = ""
    if pet and pet.birthday_estimate:
        try:
            from datetime import date as _date
            parts = pet.birthday_estimate.split("-")
            by = int(parts[0]); bm = int(parts[1]) if len(parts) > 1 else 1
            today = _date.today()
            years = today.year - by - (1 if (today.month, 1) < (bm, 1) else 0)
            if years <= 0:
                months = (today.year - by) * 12 + (today.month - bm)
                pet_age = f"{max(0, months)} 个月"
            else:
                pet_age = f"{years} 岁"
        except Exception:
            pet_age = pet.birthday_estimate or ""
    # 门店全名（处方笺标题里 "（横岗分院）" 等）
    clinic_name = "大风动物医院"
    if pet and pet.store:
        clinic_name = f"大风动物医院（{pet.store.replace('店', '分院')}）"
    return templates.TemplateResponse(request, "admin_prescription_print.html", {
        "presc": presc, "visit": visit, "cust": cust, "pet": pet,
        "pet_weight": pet_weight, "pet_age": pet_age,
        "clinic_name": clinic_name,
        "dispenser": "",  # 发药人留空，发药完成后可以打印第二联
    })


@app.post("/admin/prescriptions/{presc_id}/edit")
async def admin_presc_edit(presc_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    presc = db.get(Prescription, presc_id)
    if not presc:
        raise HTTPException(404)
    locked, reason = _is_prescription_locked(db, presc)
    if locked:
        raise HTTPException(400, f"处方单已锁定（{reason}），不可修改。如需调整请「复制为新单」或「作废」后重开。")
    # 病历已结束 → 不能改处方
    if presc.visit_id:
        _v = db.get(Visit, presc.visit_id)
        if _v and (_v.status or "open") == "closed":
            raise HTTPException(403, "所属病历已结束，处方不可修改")
    operator = request.session.get("admin_username", "admin")
    # 先把旧明细的库存退回 + 清掉本处方的 pending 发药任务
    # （旧 PrescriptionItem 删除后，孤儿 MedicationAdminLog 会指向不存在的 item_id）
    db.query(MedicationAdminLog).filter(
        MedicationAdminLog.prescription_id == presc_id,
        MedicationAdminLog.status == "pending",
    ).delete(synchronize_session=False)
    for old in presc.items:
        if old.item_id and old.quantity_num > 0:
            _restore_inventory(db, old.item_id, old.quantity_num,
                               "prescription", presc_id, operator, f"编辑处方#{presc_id}退回")
        db.delete(old)
    db.flush()
    parsed_items = _parse_presc_items(form)
    total = round(sum(it["subtotal"] for it in parsed_items), 2)
    presc.prescribed_date = str(form.get("prescribed_date", "")).strip()[:20]
    presc.vet_name = str(form.get("vet_name", "")).strip()[:80]
    presc.pet_id = int(form.get("pet_id", 0) or 0) or presc.pet_id
    presc.status = str(form.get("status", "issued")).strip()
    presc.total_amount = total
    presc.notes = str(form.get("notes", "")).strip()
    for it in parsed_items:
        db.add(PrescriptionItem(prescription_id=presc_id, **it))
        if it["item_id"] and it["quantity_num"] > 0:
            _deduct_inventory(db, it["item_id"], it["quantity_num"],
                              "prescription", presc_id, operator, f"处方#{presc_id}")
    db.commit()
    # 同步收费单
    if presc.visit_id and presc.status != "draft":
        _sync_visit_invoice(db, presc.visit_id, operator)
        db.commit()
    # 关联住院 → 重生发药任务（保留已 done/skipped 的）
    try:
        if _generate_med_logs_for_prescription(db, presc) > 0:
            db.commit()
    except Exception:
        logger.exception("[med-logs] generate after presc edit failed")
    return RedirectResponse(f"/admin/prescriptions/{presc_id}?msg=已保存", status_code=303)


@app.post("/admin/prescriptions/{presc_id}/delete")
async def admin_presc_delete(presc_id: int, request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form("")):
    """未锁定的处方单（draft / issued 且未付款）可删；锁定的请走 /void。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    presc = db.get(Prescription, presc_id)
    if not presc:
        raise HTTPException(404)
    locked, reason = _is_prescription_locked(db, presc)
    if locked:
        raise HTTPException(400, f"处方单已锁定（{reason}），不可删除。请使用「作废」。")
    if presc.visit_id:
        _v = db.get(Visit, presc.visit_id)
        if _v and (_v.status or "open") == "closed":
            raise HTTPException(403, "所属病历已结束，处方不可删除。如确需作废请使用「作废」。")
    operator = request.session.get("admin_username", "admin")
    visit_id = presc.visit_id
    # 清掉关联的住院发药任务（SQLite FK CASCADE 默认关，必须显式删）
    db.query(MedicationAdminLog).filter(
        MedicationAdminLog.prescription_id == presc_id
    ).delete(synchronize_session=False)
    for it in presc.items:
        if it.item_id and it.quantity_num > 0:
            _restore_inventory(db, it.item_id, it.quantity_num,
                               "prescription", presc_id, operator, f"删除处方#{presc_id}退回")
    db.delete(presc)
    db.commit()
    if visit_id:
        _sync_visit_invoice(db, visit_id, operator)
        db.commit()
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=处方单已删除" if visit_id else "/admin/visits", status_code=303)


@app.post("/admin/prescriptions/{presc_id}/void")
async def admin_presc_void(presc_id: int, request: Request, db: Session = Depends(get_db),
                            csrf_token: str = Form(""), void_reason: str = Form(""),
                            refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0)):
    """锁定的处方作废：库存回退 + 写审计 + status=voided + 可选退款入钱包。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    presc = db.get(Prescription, presc_id)
    if not presc:
        raise HTTPException(404)
    if presc.status == "voided":
        return RedirectResponse(f"/admin/prescriptions/{presc_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    visit_id = presc.visit_id
    # 库存回退
    for it in presc.items:
        if it.item_id and it.quantity_num > 0:
            _restore_inventory(db, it.item_id, it.quantity_num,
                               "prescription_void", presc_id, operator, f"作废处方#{presc_id}回退")
    presc.status = "voided"
    presc.voided_by = operator
    presc.voided_at = datetime.utcnow()
    presc.void_reason = (void_reason or "")[:200]
    # 作废 → 删未执行的发药任务（保留 done/skipped 作为历史，不可篡改）
    db.query(MedicationAdminLog).filter(
        MedicationAdminLog.prescription_id == presc_id,
        MedicationAdminLog.status == "pending",
    ).delete(synchronize_session=False)
    # 退款入钱包
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and presc.customer_id and refund_amount > 0:
        tx = _refund_to_wallet(
            db, presc.customer_id, float(refund_amount), operator,
            note=f"作废处方#{presc_id} 退款 · {void_reason}"[:500],
        )
        if tx:
            refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
            _audit_doc_action(db, "prescription", presc_id, "refund_to_wallet",
                              operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "prescription", presc_id, "void", operator, void_reason)
    db.commit()
    if visit_id:
        try:
            _sync_visit_invoice(db, visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=处方单已作废{refund_msg}" if visit_id else f"/admin/prescriptions/{presc_id}?msg=已作废{refund_msg}", status_code=303)


@app.post("/admin/prescriptions/{presc_id}/copy-as-new")
async def admin_presc_copy_as_new(presc_id: int, request: Request, db: Session = Depends(get_db),
                                    csrf_token: str = Form("")):
    """以本单为模板新建一张处方单（同 visit/customer/pet/vet/医生 + 全部明细）。
    新单 status=draft，未发药未付款，扣库存在保存时统一处理。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    src = db.get(Prescription, presc_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_presc = Prescription(
        visit_id=src.visit_id,
        customer_id=src.customer_id,
        pet_id=src.pet_id,
        prescribed_date=datetime.utcnow().strftime("%Y-%m-%d"),
        vet_name=src.vet_name,
        status="issued",
        total_amount=src.total_amount,
        notes=src.notes,
        created_by=operator,
    )
    db.add(new_presc)
    db.flush()
    for old in src.items:
        new_it = PrescriptionItem(
            prescription_id=new_presc.id,
            item_id=old.item_id,
            drug_name=old.drug_name,
            drug_type=old.drug_type,
            dosage=old.dosage,
            frequency=old.frequency,
            duration_days=old.duration_days,
            quantity_num=old.quantity_num,
            quantity=old.quantity,
            unit_price=old.unit_price,
            subtotal=old.subtotal,
            instructions=old.instructions,
            dose_amount=old.dose_amount,
            dose_unit=old.dose_unit,
            times_per_day=old.times_per_day,
            item_unit=old.item_unit,
            print_note=old.print_note,
        )
        db.add(new_it)
        if old.item_id and old.quantity_num > 0:
            _deduct_inventory(db, old.item_id, old.quantity_num,
                              "prescription", new_presc.id, operator,
                              f"处方#{new_presc.id}（复制自 #{presc_id}）")
    _audit_doc_action(db, "prescription", new_presc.id, "copy_from", operator,
                      extra=f"src={presc_id}")
    db.commit()
    if new_presc.visit_id:
        _sync_visit_invoice(db, new_presc.visit_id, operator)
        db.commit()
    return RedirectResponse(f"/admin/prescriptions/{new_presc.id}?msg=已复制为新单 · 可继续编辑", status_code=303)


# ═════════════════════════════════════════════════════════════════════
# 麻醉单（独立于处方单，国标要求）+ 麻醉/管控药台账
# ═════════════════════════════════════════════════════════════════════
_ANESTH_ROUTES = ["IV", "IM", "SC", "吸入", "硬膜外", "局部浸润", "口服"]
_ASA_GRADES = ["I", "II", "III", "IV", "V", "E"]


def _ledger_balance_for(db: Session, item_id: int) -> float:
    """取该药品台账当前余额（最后一条 balance_after）。"""
    if not item_id:
        return 0.0
    last = db.query(NarcoticsLedger).filter(NarcoticsLedger.item_id == item_id)\
        .order_by(NarcoticsLedger.id.desc()).first()
    if last:
        return float(last.balance_after or 0)
    # 没台账记录 → 用库存当前数量作起点
    inv = db.get(InventoryItem, item_id)
    return float(inv.stock_qty or 0) if inv else 0.0


def _write_narcotics_ledger(db: Session, *, item_id: int, item_name: str,
                            direction: str, source: str, qty: float, unit: str,
                            operator: str, cosigner: str = "",
                            visit_id: int | None = None,
                            anesth_order_id: int | None = None,
                            store: str = "", notes: str = "",
                            event_date: str = "") -> "NarcoticsLedger":
    """写一条台账，自动算 balance_after。direction: in=入/out=出/loss=损耗"""
    prev = _ledger_balance_for(db, item_id) if item_id else 0.0
    delta = qty if direction == "in" else -qty
    new_balance = round(prev + delta, 4)
    row = NarcoticsLedger(
        event_date=event_date or datetime.utcnow().strftime("%Y-%m-%d"),
        item_id=item_id or None,
        item_name=(item_name or "")[:120],
        direction=direction, source=source,
        qty=float(qty), unit=unit or "",
        balance_after=new_balance,
        operator=operator, cosigner=cosigner,
        visit_id=visit_id, anesth_order_id=anesth_order_id,
        store=store, notes=notes,
    )
    db.add(row)
    return row


def _parse_anesth_items(form) -> list[dict]:
    """解析麻醉单明细。字段：drug_name[]/item_id[]/route[]/concentration[]/dose_amount[]/dose_unit[]/total_qty[]/total_unit[]/unit_price[]/is_service[]/note[]"""
    names = form.getlist("drug_name[]")
    items = []
    for i, name in enumerate(names):
        name = str(name or "").strip()
        if not name:
            continue
        def _g(k, d=""):
            v = form.getlist(f"{k}[]")
            return v[i] if i < len(v) else d
        def _f(k):
            try:
                return float(_g(k) or 0)
            except Exception:
                return 0.0
        item_id = 0
        try:
            item_id = int(_g("item_id") or 0)
        except Exception:
            item_id = 0
        is_service_raw = _g("is_service") or ""
        unit_price = _f("unit_price")
        total_qty = _f("total_qty")
        subtotal = round(unit_price * (total_qty if total_qty > 0 else 1), 2)
        items.append({
            "item_id": item_id or None,
            "drug_name": name[:120],
            "route": (_g("route") or "IV")[:20],
            "concentration": (_g("concentration") or "")[:40],
            "dose_amount": _f("dose_amount"),
            "dose_unit": (_g("dose_unit") or "mg")[:20],
            "total_qty": total_qty,
            "total_unit": (_g("total_unit") or "")[:20],
            "unit_price": unit_price,
            "subtotal": subtotal,
            "is_service": is_service_raw in ("1", "true", "on"),
            "note": (_g("note") or "")[:200],
        })
    return items


def _anesth_form_context(request, db, *, order=None, visit=None, cust=None,
                         pet=None, pets=None, mode="create"):
    # 麻醉医师/复核人候选
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v[0] for v in vets]
    nurses = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
    ).all()
    cosigner_names = [n[0] for n in nurses if n[0] not in vet_names] + vet_names
    # 候选药：麻醉/管控药 + 服务类（吸入麻醉等）
    store = _get_op_store(request)
    cand_q = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    cand_q = cand_q.filter(
        (InventoryItem.is_controlled == True) |
        (InventoryItem.subcategory == "controlled") |
        (InventoryItem.name.ilike("%麻%"))
    )
    if store:
        cand_q = cand_q.filter(InventoryItem.store == store)
    candidates = cand_q.order_by(InventoryItem.name).limit(200).all()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    return {
        "order": order, "visit": visit, "cust": cust, "pet": pet, "pets": pets or [],
        "vet_names": vet_names, "cosigner_names": cosigner_names,
        "candidates": candidates,
        "routes": _ANESTH_ROUTES, "asa_grades": _ASA_GRADES,
        "today": today, "mode": mode,
        "csrf_token": _get_csrf_token(request),
    }


@app.get("/admin/visits/{visit_id}/anesthesia/new", response_class=HTMLResponse)
async def page_admin_anesth_new(visit_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    visit = db.get(Visit, visit_id)
    if not visit:
        raise HTTPException(404, "病例不存在")
    cust = db.get(Customer, visit.customer_id) if visit.customer_id else None
    pet = db.get(Pet, visit.pet_id) if visit.pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == visit.customer_id).all() if visit.customer_id else []
    ctx = _anesth_form_context(request, db, visit=visit, cust=cust, pet=pet, pets=pets, mode="create")
    if pet:
        ctx["anesth_history"] = db.query(AnesthesiaOrder).filter(
            AnesthesiaOrder.pet_id == pet.id
        ).order_by(AnesthesiaOrder.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/anesthesia.html", ctx)  # B8.5 UK 重写


@app.post("/admin/anesthesia/create")
async def admin_anesth_create(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    visit_id = int(form.get("visit_id", 0) or 0)
    customer_id = int(form.get("customer_id", 0) or 0)
    pet_id = int(form.get("pet_id", 0) or 0)
    vet_name = str(form.get("vet_name", "")).strip()[:80]
    cosigner = str(form.get("cosigner", "")).strip()[:80]
    if not vet_name:
        raise HTTPException(400, "麻醉医师为必填")
    if not cosigner:
        raise HTTPException(400, "国标要求双人复核：请选择第二签字人")
    if cosigner == vet_name:
        raise HTTPException(400, "第二签字人不能与麻醉医师相同")
    items = _parse_anesth_items(form)
    if not items:
        raise HTTPException(400, "请至少填写一项麻醉药品")
    total = round(sum(it["subtotal"] for it in items), 2)
    operator = request.session.get("admin_username", "admin")
    store = ""
    if visit_id:
        v = db.get(Visit, visit_id)
        if v and v.pet_id:
            p = db.get(Pet, v.pet_id)
            if p and p.store:
                store = p.store
    if not store:
        store = _get_op_store(request) or ""
    order = AnesthesiaOrder(
        visit_id=visit_id or None,
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        anesth_date=str(form.get("anesth_date", "")).strip()[:20] or datetime.utcnow().strftime("%Y-%m-%d"),
        asa_grade=str(form.get("asa_grade", "")).strip()[:10],
        vet_name=vet_name, cosigner=cosigner,
        start_time=str(form.get("start_time", "")).strip()[:10],
        end_time=str(form.get("end_time", "")).strip()[:10],
        recovery=str(form.get("recovery", "")).strip()[:40],
        status="issued",
        total_amount=total,
        store=store,
        notes=str(form.get("notes", "")).strip(),
        created_by=operator,
    )
    db.add(order)
    db.flush()
    # 写明细 + 出库（有库存的管控药）+ 台账留痕（所有项）
    for it in items:
        db.add(AnesthesiaOrderItem(order_id=order.id, **it))
        inv = db.get(InventoryItem, it["item_id"]) if it["item_id"] else None
        # 1) 实物扣库存：非服务类 + 有库存品目
        if inv and not inv.is_service and not it["is_service"] and it["total_qty"] > 0:
            _deduct_inventory(db, inv.id, it["total_qty"], "anesthesia",
                              order.id, operator, f"麻醉单#{order.id}")
        # 2) 写台账（包括服务类，国标要求所有麻醉/管控药都有痕迹）
        if inv and (inv.is_controlled or (inv.subcategory == "controlled")) or (not inv and it["is_service"]):
            _write_narcotics_ledger(
                db, item_id=inv.id if inv else 0, item_name=it["drug_name"],
                direction="out", source="anesth_order",
                qty=it["total_qty"] if it["total_qty"] > 0 else 1,
                unit=it["total_unit"] or it["dose_unit"],
                operator=vet_name, cosigner=cosigner,
                visit_id=visit_id or None, anesth_order_id=order.id,
                store=store, event_date=order.anesth_date,
                notes=f"{it['route']} {it['dose_amount']}{it['dose_unit']} · 麻醉单#{order.id}",
            )
    db.commit()
    # 同步收费单
    if visit_id:
        try:
            _sync_visit_invoice(db, visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=麻醉单已开具" if visit_id else f"/admin/anesthesia/{order.id}", status_code=303)


@app.get("/admin/anesthesia/{order_id}", response_class=HTMLResponse)
async def page_admin_anesth_detail(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    order = db.get(AnesthesiaOrder, order_id)
    if not order:
        raise HTTPException(404, "麻醉单不存在")
    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust = db.get(Customer, order.customer_id) if order.customer_id else None
    pet = db.get(Pet, order.pet_id) if order.pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == order.customer_id).all() if order.customer_id else []
    ctx = _anesth_form_context(request, db, order=order, visit=visit, cust=cust, pet=pet, pets=pets, mode="edit")
    locked, reason = _is_anesthesia_locked(db, order)
    ctx["locked"] = locked
    ctx["lock_reason"] = reason
    ctx["paid_amount"] = _doc_paid_amount(db, "anesthesia", order_id) if locked else 0.0
    if order.pet_id:
        ctx["anesth_history"] = db.query(AnesthesiaOrder).filter(
            AnesthesiaOrder.pet_id == order.pet_id,
            AnesthesiaOrder.id != order_id,
        ).order_by(AnesthesiaOrder.id.desc()).limit(10).all()
    ctx["msg"] = request.query_params.get("msg")
    return templates.TemplateResponse(request, "uk/anesthesia.html", ctx)  # B8.5 UK 重写


@app.post("/admin/anesthesia/{order_id}/copy-as-new")
async def admin_anesth_copy_as_new(order_id: int, request: Request, db: Session = Depends(get_db),
                                     csrf_token: str = Form("")):
    """以本麻醉单为模板新建一张（同 visit/pet/医师 + 全部明细 + 扣库存）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    src = db.get(AnesthesiaOrder, order_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_order = AnesthesiaOrder(
        visit_id=src.visit_id,
        customer_id=src.customer_id,
        pet_id=src.pet_id,
        anesth_date=datetime.utcnow().strftime("%Y-%m-%d"),
        asa_grade=src.asa_grade,
        vet_name=src.vet_name,
        cosigner=src.cosigner,
        start_time="",
        end_time="",
        recovery="",
        status="issued",
        total_amount=src.total_amount,
        store=src.store,
        notes=src.notes,
        created_by=operator,
    )
    db.add(new_order)
    db.flush()
    for old in src.items:
        new_it = AnesthesiaOrderItem(
            order_id=new_order.id,
            item_id=old.item_id,
            drug_name=old.drug_name,
            route=old.route,
            concentration=old.concentration,
            dose_amount=old.dose_amount,
            dose_unit=old.dose_unit,
            total_qty=old.total_qty,
            total_unit=old.total_unit,
            unit_price=old.unit_price,
            subtotal=old.subtotal,
            is_service=old.is_service,
            note=old.note,
        )
        db.add(new_it)
        inv = db.get(InventoryItem, old.item_id) if old.item_id else None
        if inv and not inv.is_service and not old.is_service and old.total_qty > 0:
            _deduct_inventory(db, inv.id, old.total_qty, "anesthesia",
                              new_order.id, operator, f"麻醉单#{new_order.id}（复制自 #{order_id}）")
        # 写台账
        if inv and (inv.is_controlled or inv.subcategory == "controlled"):
            _write_narcotics_ledger(
                db, item_id=inv.id, item_name=old.drug_name,
                direction="out", source="anesth_order",
                qty=old.total_qty if old.total_qty > 0 else 1,
                unit=old.total_unit or old.dose_unit,
                operator=src.vet_name, cosigner=src.cosigner,
                visit_id=src.visit_id, anesth_order_id=new_order.id,
                store=src.store, event_date=new_order.anesth_date,
                notes=f"复制自麻醉单#{order_id}",
            )
    _audit_doc_action(db, "anesthesia", new_order.id, "copy_from", operator, extra=f"src={order_id}")
    db.commit()
    if new_order.visit_id:
        try:
            _sync_visit_invoice(db, new_order.visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/anesthesia/{new_order.id}?msg=已复制为新单", status_code=303)


@app.get("/admin/anesthesia/{order_id}/print", response_class=HTMLResponse)
async def admin_anesth_print(order_id: int, request: Request, db: Session = Depends(get_db)):
    """麻醉记录单打印（国标三联，A5 横版，与处方笺同规格）"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    order = db.get(AnesthesiaOrder, order_id)
    if not order:
        raise HTTPException(404, "麻醉单不存在")
    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust = db.get(Customer, order.customer_id) if order.customer_id else None
    pet = db.get(Pet, order.pet_id) if order.pet_id else None
    pet_weight = 0.0
    if pet:
        last_w = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id)\
            .order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if last_w:
            pet_weight = float(last_w.weight_kg or 0)
    pet_age = ""
    if pet and pet.birthday_estimate:
        try:
            from datetime import date as _date
            parts = pet.birthday_estimate.split("-")
            by = int(parts[0]); bm = int(parts[1]) if len(parts) > 1 else 1
            today = _date.today()
            years = today.year - by - (1 if (today.month, 1) < (bm, 1) else 0)
            if years <= 0:
                months = (today.year - by) * 12 + (today.month - bm)
                pet_age = f"{max(0, months)} 个月"
            else:
                pet_age = f"{years} 岁"
        except Exception:
            pet_age = pet.birthday_estimate or ""
    clinic_name = "大风动物医院"
    store_for_title = order.store or (pet.store if pet else "")
    if store_for_title:
        clinic_name = f"大风动物医院（{store_for_title.replace('店', '分院')}）"
    return templates.TemplateResponse(request, "admin_anesthesia_print.html", {
        "order": order, "visit": visit, "cust": cust, "pet": pet,
        "pet_weight": pet_weight, "pet_age": pet_age,
        "clinic_name": clinic_name,
    })


@app.post("/admin/anesthesia/{order_id}/delete")
async def admin_anesth_delete(order_id: int, request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form("")):
    """未锁定的麻醉单 (issued 但 visit 未结、收费未付) 可真删除：
       - 回库
       - 物理删除该单关联的管控药台账条目（视为从未发生）
       - 删 items + order
       - 重同步发票
    已锁定（关联收费已付）→ 必须走 /void。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    order = db.get(AnesthesiaOrder, order_id)
    if not order:
        raise HTTPException(404)
    locked, reason = _is_anesthesia_locked(db, order)
    if locked:
        raise HTTPException(400, f"麻醉单已锁定（{reason}），不可删除。请使用「作废」。")
    if order.visit_id:
        v = db.get(Visit, order.visit_id)
        if v and (v.status or "open") == "closed":
            raise HTTPException(403, "所属病历已结束，麻醉单不可删除。如确需作废请使用「作废」。")
    operator = request.session.get("admin_username", "admin")
    visit_id = order.visit_id
    # 1) 回库（非服务类）
    for it in order.items:
        inv = db.get(InventoryItem, it.item_id) if it.item_id else None
        if inv and not inv.is_service and not it.is_service and (it.total_qty or 0) > 0:
            _restore_inventory(db, inv.id, it.total_qty, "anesthesia",
                               order_id, operator, f"删除麻醉单#{order_id}回库")
    # 2) 物理删该单关联的台账条目（真删 = 视为从未发生）
    db.query(NarcoticsLedger).filter(
        NarcoticsLedger.anesth_order_id == order_id
    ).delete(synchronize_session=False)
    # 3) 删 order（cascade 删 items）
    db.delete(order)
    db.commit()
    # 4) 重同步病例发票
    if visit_id:
        _sync_visit_invoice(db, visit_id, operator)
        db.commit()
    return RedirectResponse(
        f"/admin/visits/{visit_id}?msg=麻醉单已删除" if visit_id else "/admin/visits",
        status_code=303,
    )


@app.post("/admin/anesthesia/{order_id}/void")
async def admin_anesth_void(order_id: int, request: Request, db: Session = Depends(get_db),
                            csrf_token: str = Form(""), void_reason: str = Form(""),
                            refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    order = db.get(AnesthesiaOrder, order_id)
    if not order:
        raise HTTPException(404)
    if order.status == "voided":
        return RedirectResponse(f"/admin/anesthesia/{order_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    visit_id = order.visit_id
    # 回库 + 反向台账
    for it in order.items:
        inv = db.get(InventoryItem, it.item_id) if it.item_id else None
        if inv and not inv.is_service and not it.is_service and it.total_qty > 0:
            _restore_inventory(db, inv.id, it.total_qty, "anesthesia",
                               order_id, operator, f"作废麻醉单#{order_id}回库")
        if inv and (inv.is_controlled or inv.subcategory == "controlled"):
            _write_narcotics_ledger(
                db, item_id=inv.id, item_name=it.drug_name,
                direction="in", source="anesth_order",
                qty=it.total_qty if it.total_qty > 0 else 1,
                unit=it.total_unit or it.dose_unit,
                operator=operator, cosigner=order.cosigner,
                visit_id=order.visit_id, anesth_order_id=order.id,
                store=order.store, notes=f"作废麻醉单#{order.id}回退",
            )
    order.status = "voided"
    order.voided_by = operator
    order.voided_at = datetime.utcnow()
    order.void_reason = (void_reason or "")[:200]
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and order.customer_id and refund_amount > 0:
        tx = _refund_to_wallet(
            db, order.customer_id, float(refund_amount), operator,
            note=f"作废麻醉单#{order_id} 退款 · {void_reason}"[:500], store=order.store,
        )
        if tx:
            refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
            _audit_doc_action(db, "anesthesia", order_id, "refund_to_wallet",
                              operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "anesthesia", order_id, "void", operator, void_reason)
    db.commit()
    if visit_id:
        try:
            _sync_visit_invoice(db, visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=麻醉单已作废{refund_msg}" if visit_id else f"/admin/anesthesia/{order_id}?msg=已作废{refund_msg}", status_code=303)


# ════════════════════════════════════════════════════════════════
# 麻醉监护表（手术中逐时段生命体征 · 手机录入 + PDF 导出）
# 与麻醉单（AnesthesiaOrder）刻意分开：那是药/计费/管控药台账，这是逐时段监护。
# ════════════════════════════════════════════════════════════════
_ANMON_DEPTH_ZH = {"light": "偏浅", "adequate": "适宜", "deep": "偏深"}


def _anmon_now() -> datetime:
    """北京时间（监护时间戳人面向，统一 utc+8 存+显，自洽）。"""
    return datetime.utcnow() + timedelta(hours=8)


def _anmon_guard(request: Request, sheet: "AnesthesiaMonitorSheet") -> None:
    """门店隔离：staff 只能看本店的监护表。"""
    store_short = _get_admin_store(request)
    if store_short and sheet.store:
        store_full = _STORE_SHORT_TO_FULL.get(store_short, "")
        if sheet.store not in (store_short, store_full):
            raise HTTPException(403, "无权查看其他门店的麻醉监护表")


def _anmon_entry_flag(species: str, e: "AnesthesiaMonitorEntry") -> dict:
    """麻醉监护常见报警标记（参考，universal）：'bad' 红 / 'warn' 琥珀。"""
    f = {}
    if e.spo2:
        f["spo2"] = "bad" if e.spo2 < 90 else ("warn" if e.spo2 < 95 else "")
    if e.temperature_c:
        t = e.temperature_c
        f["temp"] = "bad" if (t < 36.0 or t > 40.0) else ("warn" if (t < 37.0 or t > 39.5) else "")
    if e.hr:
        f["hr"] = "bad" if e.hr < 50 else ("warn" if e.hr < 60 else "")
    if e.rr:
        f["rr"] = "bad" if e.rr < 6 else ("warn" if e.rr < 8 else "")
    return f


@app.post("/admin/visits/{visit_id}/anesthesia-monitor/create")
async def admin_anmon_create(visit_id: int, request: Request, db: Session = Depends(get_db),
                             csrf_token: str = Form(""), next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "病历不存在")
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    # 一台手术一张监护表：已有未结束的直接复用，避免重复建
    existing = db.query(AnesthesiaMonitorSheet).filter(
        AnesthesiaMonitorSheet.visit_id == visit_id,
        AnesthesiaMonitorSheet.status == "open",
    ).order_by(AnesthesiaMonitorSheet.id.desc()).first()
    if existing:
        return RedirectResponse(f"/m/anesthesia-monitor/{existing.id}", status_code=303)
    store = (pet.store if pet else "") or _get_op_store(request)
    weight = 0.0
    if pet:
        lw = db.query(WeightRecord).filter(WeightRecord.pet_id == pet.id)\
            .order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc()).first()
        if lw:
            weight = float(lw.weight_kg or 0)
    now = _anmon_now()
    sheet = AnesthesiaMonitorSheet(
        visit_id=visit_id, customer_id=v.customer_id, pet_id=v.pet_id,
        monitor_date=now.strftime("%Y-%m-%d"),
        start_time=now.strftime("%H:%M"),
        anesthetist=request.session.get("admin_username", "") or "",
        weight_kg=weight, store=store,
        created_by=request.session.get("admin_username", "") or "",
    )
    db.add(sheet)
    db.commit()
    return RedirectResponse(f"/m/anesthesia-monitor/{sheet.id}", status_code=303)


@app.get("/m/anesthesia-monitor/{sheet_id}", response_class=HTMLResponse)
async def m_anmon_detail(sheet_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/anesthesia-monitor/{sheet_id}", status_code=303)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    pet = db.get(Pet, sheet.pet_id) if sheet.pet_id else None
    cust = db.get(Customer, sheet.customer_id) if sheet.customer_id else None
    species = pet.species if pet else ""
    entries = list(sheet.entries)
    flags = {e.id: _anmon_entry_flag(species, e) for e in entries}

    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "sheet": sheet, "pet": pet, "cust": cust,
        "entries": entries, "flags": flags,
        "depth_zh": _ANMON_DEPTH_ZH,
        "now_hhmm": _anmon_now().strftime("%H:%M"),
        "next_url": f"/m/anesthesia-monitor/{sheet_id}",
    })
    return templates.TemplateResponse(request, "m_uk/anesthesia_monitor.html", ctx)


@app.post("/admin/anesthesia-monitor/{sheet_id}/entry")
async def admin_anmon_entry(sheet_id: int, request: Request, db: Session = Depends(get_db),
                            csrf_token: str = Form(""), next_url: str = Form(""),
                            time_hhmm: str = Form(""),
                            hr: int = Form(0), rr: int = Form(0), spo2: int = Form(0),
                            etco2: int = Form(0), temperature_c: float = Form(0.0),
                            bp_sys: int = Form(0), bp_dia: int = Form(0), bp_map: int = Form(0),
                            agent_pct: float = Form(0.0), o2_flow: float = Form(0.0),
                            depth: str = Form(""), event: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    fb = f"/m/anesthesia-monitor/{sheet_id}"
    if sheet.status == "closed":
        return RedirectResponse(_safe_next(next_url, fb + "?err=监护已结束，不可再记录"), status_code=303)
    if not any([hr, rr, spo2, etco2, temperature_c, bp_sys, bp_dia, bp_map, agent_pct, o2_flow, depth, (event or "").strip()]):
        return RedirectResponse(_safe_next(next_url, fb + "?err=至少填一项再记录#log"), status_code=303)
    # 时间：默认现在（北京时间）；可手动覆盖 HH:MM（补录）
    rec = _anmon_now()
    th = (time_hhmm or "").strip()
    if th:
        try:
            hh, mm = th.split(":")
            rec = rec.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)
        except Exception:
            pass
    if depth not in _ANMON_DEPTH_ZH:
        depth = ""
    db.add(AnesthesiaMonitorEntry(
        sheet_id=sheet_id, recorded_at=rec,
        recorded_by=request.session.get("admin_username", "") or "",
        hr=max(0, int(hr or 0)), rr=max(0, int(rr or 0)), spo2=max(0, int(spo2 or 0)),
        etco2=max(0, int(etco2 or 0)), temperature_c=max(0.0, float(temperature_c or 0)),
        bp_sys=max(0, int(bp_sys or 0)), bp_dia=max(0, int(bp_dia or 0)), bp_map=max(0, int(bp_map or 0)),
        agent_pct=max(0.0, float(agent_pct or 0)), o2_flow=max(0.0, float(o2_flow or 0)),
        depth=depth, event=(event or "").strip()[:200],
    ))
    sheet.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(_safe_next(next_url, fb + "?msg=已记录#log"), status_code=303)


@app.post("/admin/anesthesia-monitor/{sheet_id}/entry/{entry_id}/delete")
async def admin_anmon_entry_delete(sheet_id: int, entry_id: int, request: Request,
                                   db: Session = Depends(get_db),
                                   csrf_token: str = Form(""), next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    e = db.get(AnesthesiaMonitorEntry, entry_id)
    if e and e.sheet_id == sheet_id:
        db.delete(e)
        db.commit()
    return RedirectResponse(_safe_next(next_url, f"/m/anesthesia-monitor/{sheet_id}?msg=已删除#log"), status_code=303)


@app.post("/admin/anesthesia-monitor/{sheet_id}/update-header")
async def admin_anmon_update_header(sheet_id: int, request: Request, db: Session = Depends(get_db),
                                    csrf_token: str = Form(""), next_url: str = Form(""),
                                    procedure: str = Form(""), anesthetist: str = Form(""),
                                    surgeon: str = Form(""), asa_grade: str = Form(""),
                                    agent: str = Form(""), weight_kg: float = Form(0.0),
                                    monitor_date: str = Form(""), start_time: str = Form(""),
                                    end_time: str = Form(""), notes: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    sheet.procedure = (procedure or "").strip()[:200]
    sheet.anesthetist = (anesthetist or "").strip()[:80]
    sheet.surgeon = (surgeon or "").strip()[:80]
    sheet.asa_grade = (asa_grade or "").strip()[:10]
    sheet.agent = (agent or "").strip()[:80]
    sheet.weight_kg = max(0.0, float(weight_kg or 0))
    sheet.monitor_date = (monitor_date or "").strip()[:20]
    sheet.start_time = (start_time or "").strip()[:10]
    sheet.end_time = (end_time or "").strip()[:10]
    sheet.notes = (notes or "").strip()
    sheet.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(_safe_next(next_url, f"/m/anesthesia-monitor/{sheet_id}?msg=已保存表头"), status_code=303)


@app.post("/admin/anesthesia-monitor/{sheet_id}/close")
async def admin_anmon_close(sheet_id: int, request: Request, db: Session = Depends(get_db),
                            csrf_token: str = Form(""), next_url: str = Form(""), reopen: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    if reopen:
        sheet.status = "open"
        sheet.closed_at = None
        msg = "已重新开启"
    else:
        sheet.status = "closed"
        sheet.closed_at = datetime.utcnow()
        if not sheet.end_time:
            sheet.end_time = _anmon_now().strftime("%H:%M")
        msg = "监护已结束"
    sheet.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(_safe_next(next_url, f"/m/anesthesia-monitor/{sheet_id}?msg={msg}"), status_code=303)


@app.get("/admin/anesthesia-monitor/{sheet_id}/pdf")
async def admin_anmon_pdf(sheet_id: int, request: Request, db: Session = Depends(get_db),
                          download: int = Query(0)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/anesthesia-monitor/{sheet_id}", status_code=303)
    sheet = db.get(AnesthesiaMonitorSheet, sheet_id)
    if not sheet:
        raise HTTPException(404)
    _anmon_guard(request, sheet)
    from app.services.anesthesia_monitor_pdf import generate_monitor_pdf
    rel, err = generate_monitor_pdf(db, sheet_id)
    if not rel:
        return RedirectResponse(f"/m/anesthesia-monitor/{sheet_id}?err=" + quote(f"PDF 生成失败：{err}"), status_code=303)
    from fastapi.responses import FileResponse
    abs_path = Path("uploads") / rel
    fname = f"麻醉监护_{(sheet.pet.name if sheet.pet else sheet.id)}_{sheet.monitor_date or ''}.pdf"
    safe_name = quote(fname)
    disp = "attachment" if download else "inline"
    return FileResponse(str(abs_path), media_type="application/pdf",
                        headers={"Content-Disposition": f"{disp}; filename*=UTF-8''{safe_name}"})


# ─── 麻醉/管控药台账 ─────────────────────────────────────────────
@app.get("/admin/inventory/narcotics-ledger", response_class=HTMLResponse)
async def page_admin_narcotics_ledger(
    request: Request, db: Session = Depends(get_db),
    item_id: int = Query(0), source: str = Query(""),
    date_from: str = Query(""), date_to: str = Query(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    store = _get_admin_store(request) or ""
    q = db.query(NarcoticsLedger)
    if store:
        q = q.filter(NarcoticsLedger.store == store)
    if item_id:
        q = q.filter(NarcoticsLedger.item_id == item_id)
    if source:
        q = q.filter(NarcoticsLedger.source == source)
    if date_from:
        q = q.filter(NarcoticsLedger.event_date >= date_from)
    if date_to:
        q = q.filter(NarcoticsLedger.event_date <= date_to)
    rows = q.order_by(NarcoticsLedger.id.desc()).limit(500).all()
    # 候选药品下拉
    items_q = db.query(InventoryItem).filter(
        (InventoryItem.is_controlled == True) |
        (InventoryItem.subcategory == "controlled")
    )
    if store:
        items_q = items_q.filter((InventoryItem.store == store) | (InventoryItem.store == "") | (InventoryItem.store.is_(None)))
    items_list = items_q.order_by(InventoryItem.name).all()
    return templates.TemplateResponse(request, "uk/narcotics_ledger.html", {
        "rows": rows, "items_list": items_list,
        "item_id": item_id, "source": source,
        "date_from": date_from, "date_to": date_to,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.get("/admin/inventory/narcotics-ledger/manual", response_class=HTMLResponse)
async def page_admin_narcotics_manual(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    store = _get_op_store(request)
    items_q = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    items_q = items_q.filter(
        (InventoryItem.is_controlled == True) |
        (InventoryItem.subcategory == "controlled") |
        (InventoryItem.name.ilike("%麻%"))
    )
    if store:
        items_q = items_q.filter(InventoryItem.store == store)
    items_list = items_q.order_by(InventoryItem.name).all()
    staff_list = db.query(Staff.name).filter(Staff.status.in_(["active", "probation"])).all()
    staff_names = [s[0] for s in staff_list]
    return templates.TemplateResponse(request, "uk/narcotics_manual.html", {
        "items_list": items_list, "staff_names": staff_names,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/inventory/narcotics-ledger/manual")
async def admin_narcotics_manual_submit(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    item_id = int(form.get("item_id", 0) or 0)
    direction = str(form.get("direction", "")).strip()  # in / out / loss
    source = str(form.get("source", "manual_consume")).strip()  # manual_refill / manual_consume / stocktake / loss
    qty = float(form.get("qty", 0) or 0)
    unit = str(form.get("unit", "")).strip()[:20]
    operator = str(form.get("operator", "")).strip() or request.session.get("admin_username", "admin")
    cosigner = str(form.get("cosigner", "")).strip()[:80]
    notes = str(form.get("notes", "")).strip()
    event_date = str(form.get("event_date", "")).strip() or datetime.utcnow().strftime("%Y-%m-%d")
    if direction not in ("in", "out", "loss"):
        raise HTTPException(400, "方向无效")
    if qty <= 0:
        raise HTTPException(400, "数量必须大于 0")
    if not cosigner:
        raise HTTPException(400, "国标要求复核人签字")
    if cosigner == operator:
        raise HTTPException(400, "复核人不能与经办人相同")
    inv = db.get(InventoryItem, item_id) if item_id else None
    item_name = (inv.name if inv else str(form.get("item_name", "")).strip())[:120]
    if not item_name:
        raise HTTPException(400, "请选择或填写药品名称")
    store = (inv.store if inv else "") or _get_op_store(request) or ""
    # 同步实物库存（仅对有库存的非服务类品目）
    if inv and not inv.is_service:
        if direction == "in":
            _restore_inventory(db, inv.id, qty, "narcotics_manual", 0, operator, notes or f"手动入账 {source}")
        else:
            _deduct_inventory(db, inv.id, qty, "narcotics_manual", 0, operator, notes or f"手动出账 {source}")
    _write_narcotics_ledger(
        db, item_id=inv.id if inv else 0, item_name=item_name,
        direction=direction, source=source,
        qty=qty, unit=unit or (inv.unit if inv else ""),
        operator=operator, cosigner=cosigner,
        store=store, notes=notes, event_date=event_date,
    )
    db.commit()
    return RedirectResponse("/admin/inventory/narcotics-ledger?msg=已记录", status_code=303)


@app.get("/admin/inventory/narcotics-ledger/export")
async def admin_narcotics_export(
    request: Request, db: Session = Depends(get_db),
    item_id: int = Query(0), source: str = Query(""),
    date_from: str = Query(""), date_to: str = Query(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装")
    store = _get_admin_store(request) or ""
    q = db.query(NarcoticsLedger)
    if store:
        q = q.filter(NarcoticsLedger.store == store)
    if item_id:
        q = q.filter(NarcoticsLedger.item_id == item_id)
    if source:
        q = q.filter(NarcoticsLedger.source == source)
    if date_from:
        q = q.filter(NarcoticsLedger.event_date >= date_from)
    if date_to:
        q = q.filter(NarcoticsLedger.event_date <= date_to)
    rows = q.order_by(NarcoticsLedger.event_date, NarcoticsLedger.id).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "麻醉管控药台账"
    headers = ["日期", "药品", "方向", "来源", "数量", "单位", "余额", "经办人", "复核人", "门店", "关联病例", "关联麻醉单", "备注"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    _DIR = {"in": "入", "out": "出", "loss": "损耗"}
    _SRC = {"anesth_order": "麻醉单", "manual_refill": "手动补充",
            "manual_consume": "手动消耗", "stocktake": "盘点", "loss": "损耗"}
    for r in rows:
        ws.append([
            r.event_date, r.item_name, _DIR.get(r.direction, r.direction),
            _SRC.get(r.source, r.source), r.qty, r.unit, r.balance_after,
            r.operator, r.cosigner, r.store,
            r.visit_id or "", r.anesth_order_id or "", r.notes or "",
        ])
    for col in "ABCDEFGHIJKLM":
        ws.column_dimensions[col].width = 14
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"narcotics_ledger_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ---------------------------------------------------------------------------
# 发药工作台 (Dispensing Workbench)
# ---------------------------------------------------------------------------

@app.get("/admin/dispensing", response_class=HTMLResponse)
async def admin_dispensing(
    request: Request, db: Session = Depends(get_db),
    q: str = Query(""),
    status: str = Query("issued"),
    vet: str = Query(""),
    date_from: str = Query(""),
    date_to: str = Query(""),
    quick: str = Query(""),        # today / week / month
):
    require_admin(request)
    from datetime import date as _date, timedelta as _timedelta

    # 快捷时间范围
    today = _date.today()
    if quick == "today":
        date_from = date_to = today.isoformat()
    elif quick == "week":
        date_from = (today - _timedelta(days=today.weekday())).isoformat()
        date_to = today.isoformat()
    elif quick == "month":
        date_from = today.replace(day=1).isoformat()
        date_to = today.isoformat()

    qry = db.query(Prescription)
    if status and status != "all":
        qry = qry.filter(Prescription.status == status)
    if vet:
        qry = qry.filter(Prescription.vet_name == vet)
    if date_from:
        qry = qry.filter(Prescription.prescribed_date >= date_from)
    if date_to:
        qry = qry.filter(Prescription.prescribed_date <= date_to)
    if q:
        # 搜索患者姓名 / 宠物名 / 药品名（通过 subquery）
        drug_ids = db.query(PrescriptionItem.prescription_id).filter(
            PrescriptionItem.drug_name.ilike(f"%{q}%")
        ).subquery()
        pet_ids = db.query(Pet.id).filter(Pet.name.ilike(f"%{q}%")).subquery()
        cust_ids = db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).subquery()
        qry = qry.filter(or_(
            Prescription.id.in_(drug_ids),
            Prescription.pet_id.in_(pet_ids),
            Prescription.customer_id.in_(cust_ids),
        ))

    # 待发药队列：最早开单排前；其余倒序
    if status == "issued":
        prescs = qry.order_by(Prescription.prescribed_date.asc(), Prescription.id.asc()).limit(200).all()
    else:
        prescs = qry.order_by(Prescription.id.desc()).limit(200).all()

    # 预载关联数据
    cust_map = {}
    pet_map = {}
    for p in prescs:
        if p.customer_id and p.customer_id not in cust_map:
            c = db.get(Customer, p.customer_id)
            if c:
                cust_map[p.customer_id] = c
        if p.pet_id and p.pet_id not in pet_map:
            pt = db.get(Pet, p.pet_id)
            if pt:
                pet_map[p.pet_id] = pt

    # 统计数字
    pending_count = db.query(Prescription).filter(Prescription.status == "issued").count()
    today_str = today.isoformat()
    today_dispensed = db.query(Prescription).filter(
        Prescription.status == "dispensed",
        Prescription.prescribed_date == today_str,
    ).count()
    month_start = today.replace(day=1).isoformat()
    month_total = db.query(Prescription).filter(
        Prescription.prescribed_date >= month_start,
    ).count()

    # 在职医生列表
    vets = [r[0] for r in db.query(Prescription.vet_name).filter(
        Prescription.vet_name != ""
    ).distinct().order_by(Prescription.vet_name).all()]

    return templates.TemplateResponse(request, "uk/dispensing.html", {
        "request": request,
        "prescs": prescs, "cust_map": cust_map, "pet_map": pet_map,
        "q": q, "status": status, "vet": vet,
        "date_from": date_from, "date_to": date_to, "quick": quick,
        "pending_count": pending_count,
        "today_dispensed": today_dispensed,
        "month_total": month_total,
        "vets": vets,
        "presc_status_zh": _PRESC_STATUS_ZH,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": "发药工作台",
    })


@app.post("/admin/dispensing/{presc_id}/dispense")
async def admin_dispensing_dispense(
    presc_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    presc = db.get(Prescription, presc_id)
    if not presc or presc.status != "issued":
        raise HTTPException(400, "处方单状态不符，无法发药")
    presc.status = "dispensed"
    db.commit()
    return RedirectResponse("/admin/dispensing?msg=已确认发药", status_code=303)


@app.post("/admin/dispensing/{presc_id}/undispense")
async def admin_dispensing_undispense(
    presc_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    if request.session.get("admin_role") != "superadmin":
        raise HTTPException(403, "仅超级管理员可撤回发药")
    presc = db.get(Prescription, presc_id)
    if not presc or presc.status != "dispensed":
        raise HTTPException(400, "仅已发药的处方单可撤回")
    presc.status = "issued"
    db.commit()
    return RedirectResponse(f"/admin/dispensing?msg=已撤回至待发药&status=all", status_code=303)


@app.post("/admin/dispensing/bulk-dispense")
async def admin_dispensing_bulk(
    request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    form = await request.form()
    ids = [int(v) for k, v in form.multi_items() if k == "presc_ids" and v.isdigit()]
    if not ids:
        return RedirectResponse("/admin/dispensing?err=未选择处方单", status_code=303)
    updated = 0
    for pid in ids:
        p = db.get(Prescription, pid)
        if p and p.status == "issued":
            p.status = "dispensed"
            updated += 1
    db.commit()
    return RedirectResponse(f"/admin/dispensing?msg=已批量确认发药+{updated}+张", status_code=303)


# ---------------------------------------------------------------------------
# Phase 3 — 销售单 (Sales Orders)
# ---------------------------------------------------------------------------

_SO_STATUS_ZH = {"pending": "待付款", "paid": "已收款", "cancelled": "已取消"}
_SO_ITEM_TYPE_ZH = {"product": "商品", "service": "服务", "medication": "药品", "vaccine": "疫苗"}
_PAYMENT_METHOD_OPTIONS = ["现金", "微信", "支付宝", "银行卡", "挂账"]


def _parse_so_items(form_data) -> list[dict]:
    items = []
    i = 0
    while True:
        name = form_data.get(f"item_name_{i}", "").strip()
        if not name and i > 20:
            break
        if name:
            try:
                unit_price = float(form_data.get(f"unit_price_{i}", 0) or 0)
                quantity = float(form_data.get(f"quantity_{i}", 1) or 1)
            except ValueError:
                unit_price, quantity = 0.0, 1.0
            subtotal = round(unit_price * quantity, 2)
            raw_item_id = form_data.get(f"item_id_{i}", "").strip()
            item_id = int(raw_item_id) if raw_item_id and raw_item_id.isdigit() else None
            items.append({
                "item_id": item_id,
                "item_name": name,
                "item_type": form_data.get(f"item_type_{i}", "product").strip(),
                "unit_price": unit_price,
                "quantity": quantity,
                "subtotal": subtotal,
                "notes": form_data.get(f"item_notes_{i}", "").strip(),
            })
        i += 1
    return items


@app.get("/admin/sales-orders/create", response_class=HTMLResponse)
async def page_admin_so_create(
    request: Request, db: Session = Depends(get_db),
    customer_id: int = Query(0), visit_id: int = Query(0), pet_id: int = Query(0),
    search_q: str = Query(""),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    visit = db.get(Visit, visit_id) if visit_id else None
    if visit:
        customer_id = customer_id or visit.customer_id or 0
        pet_id = pet_id or visit.pet_id or 0
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == customer_id).all() if customer_id else []
    search_results = []
    if search_q and not customer_id:
        search_results = db.query(Customer).filter(
            or_(Customer.name.ilike(f"%{search_q}%"), Customer.phone.ilike(f"%{search_q}%"))
        ).limit(10).all()
    today = datetime.utcnow().strftime("%Y-%m-%d")
    history = []
    if customer_id:
        history = db.query(SalesOrder).filter(SalesOrder.customer_id == customer_id)\
            .order_by(SalesOrder.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/sales_order_form.html", {
        "order": None, "visit": visit, "cust": cust, "pet": pet, "pets": pets,
        "so_status_zh": _SO_STATUS_ZH, "item_type_zh": _SO_ITEM_TYPE_ZH,
        "payment_methods": _PAYMENT_METHOD_OPTIONS,
        "so_history": history,
        "today": today, "csrf_token": _get_csrf_token(request), "mode": "create",
        "search_q": search_q, "search_results": search_results,
    })


@app.post("/admin/sales-orders/create")
async def admin_so_create(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    customer_id = int(form.get("customer_id", 0) or 0)
    visit_id = int(form.get("visit_id", 0) or 0)
    pet_id = int(form.get("pet_id", 0) or 0)
    items = _parse_so_items(form)
    # 整支/整瓶计费：开 0.1ml = 开 1 整支
    _apply_single_use_pack_billing(db, items)
    # 员工内购档案：单价改填进价
    _apply_internal_pricing(db, items, customer_id)
    total = round(sum(it["subtotal"] for it in items), 2)
    operator = request.session.get("admin_username", "admin")
    order = SalesOrder(
        customer_id=customer_id or None,
        visit_id=visit_id or None,
        pet_id=pet_id or None,
        order_date=str(form.get("order_date", "")).strip()[:20],
        status=str(form.get("status", "pending")).strip(),
        total_amount=total,
        payment_method=str(form.get("payment_method", "")).strip()[:40],
        notes=str(form.get("notes", "")).strip(),
        created_by=operator,
    )
    db.add(order)
    db.flush()
    for it in items:
        db.add(SalesOrderItem(order_id=order.id, **it))
        if it["item_id"] and it["quantity"] > 0:
            _deduct_inventory(db, it["item_id"], it["quantity"],
                              "sales_order", order.id, operator, f"销售单#{order.id}")
    db.commit()
    # 同步收费单
    inv = None
    if visit_id:
        _sync_visit_invoice(db, visit_id, operator)
    else:
        inv = _sync_sales_order_invoice(db, order.id, operator)
    db.commit()
    # next_url：支持 {order_id} / {invoice_id} 占位（手机端跳收款页用）
    nu_raw = str(form.get("next_url", "")).strip()
    if nu_raw:
        nu = nu_raw.replace("{order_id}", str(order.id))
        nu = nu.replace("{invoice_id}", str(inv.id) if inv else "0")
        fb = f"/admin/visits/{visit_id}" if visit_id else f"/admin/sales-orders/{order.id}"
        return RedirectResponse(_safe_next(nu, fb), status_code=303)
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=销售单已创建" if visit_id else f"/admin/sales-orders/{order.id}?msg=销售单已创建·收银台已生成待收款单", status_code=303)


@app.get("/admin/sales-orders", response_class=HTMLResponse)
async def page_admin_so_list(
    request: Request, db: Session = Depends(get_db),
    q: str = Query(""), status: str = Query(""),
    date_from: str = Query(""), date_to: str = Query(""),
    page: int = Query(1),
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    query = db.query(SalesOrder)
    if q:
        query = query.join(Customer, SalesOrder.customer_id == Customer.id, isouter=True).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        )
    if status:
        query = query.filter(SalesOrder.status == status)
    if date_from:
        query = query.filter(SalesOrder.order_date >= date_from)
    if date_to:
        query = query.filter(SalesOrder.order_date <= date_to)
    total = query.count()
    page_size = 30
    orders = query.order_by(SalesOrder.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    cust_map = {}
    for o in orders:
        if o.customer_id and o.customer_id not in cust_map:
            c = db.get(Customer, o.customer_id)
            if c:
                cust_map[o.customer_id] = c
    return templates.TemplateResponse(request, "uk/sales_orders.html", {
        "orders": orders, "cust_map": cust_map,
        "so_status_zh": _SO_STATUS_ZH,
        "total": total, "page": page, "page_size": page_size,
        "filters": {"q": q, "status": status, "date_from": date_from, "date_to": date_to},
    })


@app.get("/admin/sales-orders/{order_id}", response_class=HTMLResponse)
async def page_admin_so_detail(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404, "销售单不存在")
    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust = db.get(Customer, order.customer_id) if order.customer_id else None
    pet = db.get(Pet, order.pet_id) if order.pet_id else None
    pets = db.query(Pet).filter(Pet.customer_id == order.customer_id).all() if order.customer_id else []
    locked, lock_reason = _is_sales_order_locked(db, order)
    paid_amount = _doc_paid_amount(db, "sales_order", order_id) if locked else 0.0
    history = []
    if order.customer_id:
        history = db.query(SalesOrder).filter(
            SalesOrder.customer_id == order.customer_id,
            SalesOrder.id != order_id,
        ).order_by(SalesOrder.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/sales_order_form.html", {
        "order": order, "visit": visit, "cust": cust, "pet": pet, "pets": pets,
        "so_status_zh": _SO_STATUS_ZH, "item_type_zh": _SO_ITEM_TYPE_ZH,
        "payment_methods": _PAYMENT_METHOD_OPTIONS,
        "so_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "csrf_token": _get_csrf_token(request), "mode": "edit",
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/sales-orders/{order_id}/edit")
async def admin_so_edit(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404)
    locked, reason = _is_sales_order_locked(db, order)
    if locked:
        raise HTTPException(400, f"销售单已锁定（{reason}），不可修改。请「复制为新单」或「作废」后重开。")
    operator = request.session.get("admin_username", "admin")
    # 先退回旧明细库存
    for old in order.items:
        if old.item_id and old.quantity > 0:
            _restore_inventory(db, old.item_id, old.quantity,
                               "sales_order", order_id, operator, f"编辑销售单#{order_id}退回")
        db.delete(old)
    db.flush()
    items = _parse_so_items(form)
    total = round(sum(it["subtotal"] for it in items), 2)
    order.order_date = str(form.get("order_date", "")).strip()[:20]
    order.status = str(form.get("status", "pending")).strip()
    order.payment_method = str(form.get("payment_method", "")).strip()[:40]
    order.total_amount = total
    order.notes = str(form.get("notes", "")).strip()
    order.pet_id = int(form.get("pet_id", 0) or 0) or order.pet_id
    for it in items:
        db.add(SalesOrderItem(order_id=order_id, **it))
        if it["item_id"] and it["quantity"] > 0:
            _deduct_inventory(db, it["item_id"], it["quantity"],
                              "sales_order", order_id, operator, f"销售单#{order_id}")
    db.commit()
    if order.visit_id:
        _sync_visit_invoice(db, order.visit_id, operator)
    else:
        _sync_sales_order_invoice(db, order.id, operator)
    db.commit()
    return RedirectResponse(f"/admin/sales-orders/{order_id}?msg=已保存", status_code=303)


@app.post("/admin/sales-orders/{order_id}/pay")
async def admin_so_pay(order_id: int, request: Request, db: Session = Depends(get_db),
                        csrf_token: str = Form(""), payment_method: str = Form("")):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404)
    order.status = "paid"
    if payment_method:
        order.payment_method = payment_method.strip()[:40]
    db.commit()
    return RedirectResponse(f"/admin/sales-orders/{order_id}?msg=已标记收款", status_code=303)


@app.post("/admin/sales-orders/{order_id}/delete")
async def admin_so_delete(order_id: int, request: Request, db: Session = Depends(get_db),
                           csrf_token: str = Form("")):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404)
    locked, reason = _is_sales_order_locked(db, order)
    if locked:
        raise HTTPException(400, f"销售单已锁定（{reason}），不可删除。请使用「作废」。")
    operator = request.session.get("admin_username", "admin")
    visit_id = order.visit_id
    for it in order.items:
        if it.item_id and it.quantity > 0:
            _restore_inventory(db, it.item_id, it.quantity,
                               "sales_order", order_id, operator, f"删除销售单#{order_id}退回")
    # 先清理独立发票（在删 order 之前，避免 FK 风险）
    if not visit_id:
        _delete_so_invoice(db, order_id)
    db.delete(order)
    db.commit()
    if visit_id:
        _sync_visit_invoice(db, visit_id, operator)
        db.commit()
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=销售单已删除" if visit_id else "/admin/sales-orders", status_code=303)


@app.post("/admin/sales-orders/{order_id}/void")
async def admin_so_void(order_id: int, request: Request, db: Session = Depends(get_db),
                         csrf_token: str = Form(""), void_reason: str = Form(""),
                         refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404)
    if order.status == "voided":
        return RedirectResponse(f"/admin/sales-orders/{order_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    visit_id = order.visit_id
    for it in order.items:
        if it.item_id and it.quantity > 0:
            _restore_inventory(db, it.item_id, it.quantity,
                               "sales_order_void", order_id, operator, f"作废销售单#{order_id}回退")
    order.status = "voided"
    order.voided_by = operator
    order.voided_at = datetime.utcnow()
    order.void_reason = (void_reason or "")[:200]
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and order.customer_id and refund_amount > 0:
        tx = _refund_to_wallet(
            db, order.customer_id, float(refund_amount), operator,
            note=f"作废销售单#{order_id} 退款 · {void_reason}"[:500],
        )
        if tx:
            refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
            _audit_doc_action(db, "sales_order", order_id, "refund_to_wallet",
                              operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "sales_order", order_id, "void", operator, void_reason)
    db.commit()
    if visit_id:
        try:
            _sync_visit_invoice(db, visit_id, operator)
            db.commit()
        except Exception:
            pass
    else:
        # 独立销售单作废 → 清掉对应的未付发票（已付的保留作历史档案）
        try:
            _delete_so_invoice(db, order_id)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=销售单已作废{refund_msg}" if visit_id else f"/admin/sales-orders/{order_id}?msg=已作废{refund_msg}", status_code=303)


@app.post("/admin/sales-orders/{order_id}/copy-as-new")
async def admin_so_copy_as_new(order_id: int, request: Request, db: Session = Depends(get_db),
                                csrf_token: str = Form("")):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    src = db.get(SalesOrder, order_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_order = SalesOrder(
        customer_id=src.customer_id,
        visit_id=src.visit_id,
        pet_id=src.pet_id,
        order_date=datetime.utcnow().strftime("%Y-%m-%d"),
        status="pending",
        total_amount=src.total_amount,
        payment_method="",
        notes=src.notes,
        created_by=operator,
    )
    db.add(new_order)
    db.flush()
    for old in src.items:
        new_it = SalesOrderItem(
            order_id=new_order.id,
            item_id=old.item_id,
            item_name=old.item_name,
            item_type=old.item_type,
            unit_price=old.unit_price,
            quantity=old.quantity,
            subtotal=old.subtotal,
            notes=old.notes,
        )
        db.add(new_it)
        if old.item_id and old.quantity > 0:
            _deduct_inventory(db, old.item_id, old.quantity,
                              "sales_order", new_order.id, operator,
                              f"销售单#{new_order.id}（复制自 #{order_id}）")
    _audit_doc_action(db, "sales_order", new_order.id, "copy_from", operator, extra=f"src={order_id}")
    db.commit()
    if new_order.visit_id:
        try:
            _sync_visit_invoice(db, new_order.visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/sales-orders/{new_order.id}?msg=已复制为新单", status_code=303)


# ─────────────────────────────────────────────
#  库存管理 Inventory
# ─────────────────────────────────────────────

INVENTORY_CATEGORIES = {
    "reception": {"label": "接待", "subs": {
        "registration": "挂号",
        "consultation": "咨询",
    }},
    "medication": {"label": "药品", "subs": {
        "controlled": "麻药/精神类",
        "general":    "普通药品",
    }},
    "vaccine": {"label": "疫苗", "subs": {
        "rabies":     "狂犬疫苗",
        "combo":      "联苗",
        "other":      "其他疫苗",
    }},
    "antiparasitic": {"label": "驱虫", "subs": {
        "internal":   "体内驱虫",
        "external":   "体外驱虫",
        "both":       "体内外同驱",
    }},
    "consumable": {"label": "耗材", "subs": {
        "general": "普通耗材",
    }},
    "product": {"label": "商品", "subs": {
        "general": "普通商品",
    }},
    "grooming": {"label": "美容", "subs": {
        "washcare": "洗护",
        "styling":  "造型",
        "addon":    "附加服务",
    }},
    "lab": {"label": "化验", "subs": {
        "routine_lab":  "常规化验",
        "external_lab": "院外送检",
    }},
    "imaging": {"label": "影像", "subs": {
        "dr":        "DR",
        "ct":        "CT",
        "mri":       "核磁共振",
        "ultrasound":"B超",
    }},
    "microscopy": {"label": "显微镜", "subs": {
        "optical":   "常规光学显微镜",
        "electron":  "电子显微镜",
    }},
    "surgery": {"label": "手术", "subs": {
        "general":    "普外科手术",
        "ophthalmic": "眼科手术",
        "orthopedic": "骨科手术",
        "dental":     "口腔手术",
        "neuro":      "神经外科",
    }},
    "treatment": {"label": "处置", "subs": {
        "routine":    "一般处置",
        "emergency":  "紧急处置",
        "anesthesia": "麻醉处置",
    }},
    "nursing": {"label": "护理", "subs": {
        "general":   "普通护理",
        "isolation": "隔离护理",
    }},
}


@app.get("/api/diseases/search")
async def api_diseases_search(
    request: Request,
    q: str = Query(""),
    db: Session = Depends(get_db),
):
    """诊断 autocomplete：在疾病字典 + 别名里模糊搜索。

    返回：[{name, system, system_zh, aliases, severity}]
    排序：use_count desc（常用的在前） + name 字母序
    """
    if not request.session.get("admin"):
        return []
    from sqlalchemy import or_ as _or_
    from app.data.vet_seed import SYSTEMS as _SYS
    qs = (q or "").strip()
    query = db.query(Disease)
    if qs:
        like = f"%{qs}%"
        query = query.filter(_or_(Disease.name.ilike(like), Disease.aliases.ilike(like)))
    rows = query.order_by(Disease.use_count.desc(), Disease.name).limit(20).all()
    return [
        {
            "name": d.name,
            "system": d.system,
            "system_zh": _SYS.get(d.system, d.system),
            "severity": d.severity,
            "aliases": d.aliases,
        }
        for d in rows
    ]


@app.get("/api/inventory/search")
async def api_inventory_search(
    request: Request,
    q: str = Query(""),
    category: str = Query(""),
    db: Session = Depends(get_db),
):
    """JSON autocomplete for inventory items — used by prescription/sales order forms."""
    query = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    # 多门店：所有角色按 session.admin_store 过滤（含超管），库存物理隔离
    query = _apply_store_filter(query, InventoryItem.store, _get_op_store(request))
    if category:
        query = query.filter(InventoryItem.category == category)
    if q:
        query = query.filter(InventoryItem.name.ilike(f"%{q}%"))
    items = query.order_by(InventoryItem.name).limit(30).all()
    # 按当前员工的门店取「有效售价」（默认价 + 该店覆盖）
    from app.services.pricing import effective_sell_price as _eff
    _cur_store = _get_op_store(request)
    return [
        {
            "id": it.id,
            "name": it.name,
            "category": it.category,
            "unit": it.unit,
            "unit2": it.unit2 or "",
            "unit2_ratio": float(it.unit2_ratio or 1.0),
            "sell_price": _eff(it, _cur_store),  # 关键：替换为有效价
            "default_price": it.sell_price,       # 留默认价做参考
            "cost_price": float(it.cost_price or 0),
            "stock_qty": it.stock_qty,
            "is_service": it.is_service,
            "is_controlled": it.is_controlled,
            "store": it.store or "",
        }
        for it in items
    ]


# ── 库存：批量编辑 ───────────────────────────────────────
def _redirect_back_with_msg(request: Request, fallback: str, msg: str = "", err: str = "") -> RedirectResponse:
    """带筛选参数回到 referer 页（编辑后保留 q/category/subcategory 等）。

    referer 不可信时回退到 fallback。msg 或 err 自动 url 编码追加。
    """
    from urllib.parse import urlparse, urlencode, parse_qsl, quote as _q
    referer = (request.headers.get("referer") or "").strip()
    target = fallback
    if referer:
        try:
            u = urlparse(referer)
            # 只接受同源同路径前缀的 referer，安全
            if u.path and (u.path == fallback.split("?")[0]
                           or u.path.startswith(fallback.split("?")[0] + "?")):
                params = dict(parse_qsl(u.query, keep_blank_values=False))
                # 清掉旧的 msg/err
                params.pop("msg", None); params.pop("err", None)
                if msg: params["msg"] = msg
                if err: params["err"] = err
                target = u.path
                if params:
                    target += "?" + urlencode(params, doseq=False)
                return RedirectResponse(target, status_code=303)
        except Exception:
            pass
    # 回退路径
    suffix = ""
    if msg: suffix = f"?msg={_q(msg, safe='')}"
    elif err: suffix = f"?err={_q(err, safe='')}"
    return RedirectResponse(fallback + suffix, status_code=303)


@app.post("/admin/inventory/bulk-edit")
async def admin_inventory_bulk_edit(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    item_ids: list[int] = Form(...),
    category: str = Form(""),
    subcategory: str = Form(""),
    supplier: str = Form(""),
    store: str = Form("__keep__"),
    is_service: str = Form("__keep__"),
    is_controlled: str = Form("__keep__"),
    requires_report: str = Form("__keep__"),
    single_use_pack: str = Form("__keep__"),
    unit: str = Form(""),
    unit2: str = Form(""),
    unit2_ratio: str = Form(""),
    notes: str = Form(""),
    supplier_clear: str = Form(""),
    notes_clear: str = Form(""),
):
    """对一组品目批量改 大类 / 小类 / 供应商 / 归属门店 / 服务类目 / 麻醉管控。
    每个字段留空（store/is_service/is_controlled 为 "__keep__"）= 不修改；至少要改 1 个字段。
    门店字段仅超级管理员可改；staff 提交会被忽略。
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    _require_csrf(request, csrf_token)
    if not item_ids:
        return RedirectResponse("/admin/inventory?msg=未选择品目", status_code=303)
    # 校验 category（小类必须属于该大类）
    if category and category not in INVENTORY_CATEGORIES:
        return RedirectResponse("/admin/inventory?msg=大类无效", status_code=303)
    if subcategory:
        if not category:
            return RedirectResponse("/admin/inventory?msg=改小类时必须同时选大类", status_code=303)
        if subcategory not in INVENTORY_CATEGORIES[category].get("subs", {}):
            return RedirectResponse("/admin/inventory?msg=小类不属于所选大类", status_code=303)
    # 门店：__keep__ = 不改；空字符串 = 改成通用；东环店/横岗店 = 改成指定店
    change_store = (store != "__keep__") and (request.session.get("admin_role") == "superadmin")
    if change_store and store not in ("", "东环店", "横岗店"):
        return RedirectResponse("/admin/inventory?msg=门店值无效", status_code=303)
    change_is_service = is_service in ("0", "1")
    change_is_controlled = is_controlled in ("0", "1")
    change_requires_report = requires_report in ("0", "1")
    change_single_use_pack = single_use_pack in ("0", "1")
    unit = (unit or "").strip()[:20]
    unit2 = (unit2 or "").strip()[:20]
    ratio_val = None
    if (unit2_ratio or "").strip():
        try:
            ratio_val = float(unit2_ratio.strip())
            if ratio_val < 1:
                ratio_val = 1.0
        except (ValueError, TypeError):
            ratio_val = None
    do_supplier_clear = supplier_clear == "1"
    do_notes_clear = notes_clear == "1"
    if not (category or supplier or do_supplier_clear or notes or do_notes_clear
            or change_store or change_is_service or change_is_controlled
            or change_requires_report or change_single_use_pack
            or unit or unit2 or ratio_val is not None):
        return RedirectResponse("/admin/inventory?msg=请至少选一个要修改的字段", status_code=303)

    rows = db.query(InventoryItem).filter(InventoryItem.id.in_(item_ids)).all()
    updated = 0
    for it in rows:
        if category:
            it.category = category
            # 改大类 → 小类同时清空或换；若用户明确选了小类用它，否则清空（防留旧大类的小类）
            it.subcategory = subcategory or ""
        elif subcategory:
            it.subcategory = subcategory
        if do_supplier_clear:
            it.supplier = ""
        elif supplier:
            it.supplier = supplier[:200]
        if do_notes_clear:
            it.notes = ""
        elif notes:
            it.notes = notes[:1000]
        if change_store:
            it.store = store
        if change_is_service:
            it.is_service = (is_service == "1")
        if change_is_controlled:
            it.is_controlled = (is_controlled == "1")
        if change_requires_report:
            it.requires_report = (requires_report == "1")
        if change_single_use_pack:
            it.single_use_pack = (single_use_pack == "1")
        if unit:
            it.unit = unit
        if unit2:
            it.unit2 = unit2
        if ratio_val is not None:
            it.unit2_ratio = ratio_val
        it.updated_at = datetime.utcnow()
        updated += 1
    db.commit()
    _audit(db, request, "inventory_bulk_edit", application_id=None,
           detail={"count": updated, "category": category, "subcategory": subcategory, "supplier": supplier})
    db.commit()
    parts = []
    if category: parts.append(f"大类={INVENTORY_CATEGORIES[category]['label']}")
    if subcategory: parts.append(f"小类={INVENTORY_CATEGORIES[category]['subs'].get(subcategory, subcategory)}")
    if supplier: parts.append(f"供应商={supplier}")
    if do_supplier_clear: parts.append("供应商=清空")
    if do_notes_clear: parts.append("备注=清空")
    if unit: parts.append(f"主单位={unit}")
    if unit2: parts.append(f"副单位={unit2}")
    msg = f"已批量更新 {updated} 个品目"
    if parts:
        msg += f"（{' · '.join(parts)}）"
    return _redirect_back_with_msg(request, "/admin/inventory", msg=msg)


# ── 库存品目搜索 API（拍照入库 / 各种映射场景共用） ───────
@app.get("/api/inventory/search")
async def api_inventory_search(
    request: Request,
    db: Session = Depends(get_db),
    q: str = Query(""),
    limit: int = Query(15),
):
    """按名称模糊搜索品目。返回 JSON 列表。"""
    if not request.session.get("admin"):
        return {"items": []}
    q = (q or "").strip()
    qq = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    # 日常开单/选品按当前操作门店过滤，含超管
    qq = _apply_store_filter(qq, InventoryItem.store, _get_op_store(request))
    if q:
        like = f"%{q}%"
        qq = qq.filter(InventoryItem.name.like(like))
    rows = qq.order_by(InventoryItem.name).limit(min(max(limit, 1), 50)).all()
    return {"items": [
        {
            "id": r.id, "name": r.name, "unit": r.unit,
            "unit2": r.unit2 or "",
            "unit2_ratio": float(r.unit2_ratio or 1.0),
            "stock_qty": float(r.stock_qty or 0),
            "cost_price": float(r.cost_price or 0),
            "sell_price": float(r.sell_price or 0),
        }
        for r in rows
    ]}


# ── 进货单照片识别入库 ─────────────────────────────────────
@app.get("/admin/inventory/import-xls", response_class=HTMLResponse)
async def admin_inventory_import_xls_page(request: Request):
    require_admin(request)
    require_superadmin(request)
    return templates.TemplateResponse(request, "uk/inventory_import_xls.html", {
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
        "err": request.query_params.get("err"),
    })


# 导入 records 暂存目录（替代 session，避免 cookie 容量超限）
_INV_IMPORT_CACHE_DIR = Path("data/inv_import_cache")
_INV_IMPORT_CACHE_DIR.mkdir(parents=True, exist_ok=True)


def _inv_import_cache_path(token: str) -> Path:
    # token 是 secrets.token_urlsafe，只允许 url-safe 字符，安全
    safe = "".join(c for c in token if c.isalnum() or c in "-_")[:80]
    return _INV_IMPORT_CACHE_DIR / f"{safe}.json"


@app.post("/admin/inventory/import-xls/preview")
async def admin_inventory_import_xls_preview(
    request: Request,
    db: Session = Depends(get_db),
    file: UploadFile = File(...),
    store: str = Form(""),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    import tempfile, os, json as _json
    from app.services.inventory_import import parse_xls_to_records, preview_import

    suffix = ".xlsx" if file.filename.lower().endswith(".xlsx") else ".xls"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        records, warnings = parse_xls_to_records(tmp_path)
        if not records:
            return RedirectResponse(f"/admin/inventory/import-xls?err=未解析出任何记录", status_code=303)
        preview = preview_import(db, records, store=store or "")
        # 大数据集（583+ 条）不能塞 session cookie（4KB 上限），落盘 cache
        token = secrets.token_urlsafe(20)
        cache_path = _inv_import_cache_path(token)
        cache_path.write_text(
            _json.dumps({"records": records, "store": store or ""}, ensure_ascii=False),
            encoding="utf-8",
        )
        # session 里只存 token（短字符串），并清理 1 小时前的旧缓存
        request.session["_inv_import_token"] = token
        try:
            now = datetime.utcnow().timestamp()
            for f in _INV_IMPORT_CACHE_DIR.glob("*.json"):
                if now - f.stat().st_mtime > 3600:
                    f.unlink(missing_ok=True)
        except Exception:
            pass
        return templates.TemplateResponse(request, "uk/inventory_import_xls.html", {
            "csrf_token": _get_csrf_token(request),
            "preview": preview,
            "warnings": warnings[:20],
            "filename": file.filename,
            "store": store,
        })
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


@app.post("/admin/inventory/import-xls/commit")
async def admin_inventory_import_xls_commit(
    request: Request,
    db: Session = Depends(get_db),
    strategy: str = Form("skip"),
    inherit_meta: str = Form(""),
    csrf_token: str = Form(""),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    import json as _json
    token = request.session.get("_inv_import_token", "")
    cache_path = _inv_import_cache_path(token) if token else None
    if not cache_path or not cache_path.exists():
        return RedirectResponse("/admin/inventory/import-xls?err=会话已过期或缓存丢失，请重新上传", status_code=303)
    try:
        cached = _json.loads(cache_path.read_text(encoding="utf-8"))
        records = cached.get("records", [])
        store = cached.get("store", "")
    except Exception as e:
        return RedirectResponse(f"/admin/inventory/import-xls?err=读取缓存失败：{e}", status_code=303)
    from app.services.inventory_import import commit_import
    _inherit = inherit_meta.lower() in ("1", "true", "on", "yes")
    stat = commit_import(db, records, store=store, strategy=strategy, inherit_meta=_inherit)
    # 清理缓存 + session 凭证
    try:
        cache_path.unlink(missing_ok=True)
    except Exception:
        pass
    request.session.pop("_inv_import_token", None)
    parts = [f"新建 {stat['created']}", f"更新 {stat['updated']}", f"跳过 {stat['skipped']}"]
    if stat.get("cross_inherited"):
        parts.append(f"继承别店属性 {stat['cross_inherited']}")
    msg = "导入完成：" + " / ".join(parts)
    return RedirectResponse(f"/admin/inventory?msg={msg}", status_code=303)


@app.post("/admin/inventory/batch-action")
async def admin_inventory_batch_action(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    action: str = Form(...),  # deactivate / activate / delete
):
    """库存品目批量操作：下架 / 启用 / 真删除（带引用保护）。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    form = await request.form()
    raw_ids = form.getlist("item_ids") if hasattr(form, "getlist") else []
    ids = []
    for x in raw_ids:
        try:
            ids.append(int(x))
        except (TypeError, ValueError):
            continue
    if not ids:
        return RedirectResponse("/admin/inventory?err=未选中任何品目", status_code=303)

    affected = 0
    blocked = 0
    if action in ("deactivate", "activate"):
        new_val = (action == "activate")
        for iid in ids:
            it = db.get(InventoryItem, iid)
            if it:
                it.is_active = new_val
                affected += 1
        db.commit()
        verb = "启用" if new_val else "下架"
        return _redirect_back_with_msg(request, "/admin/inventory", msg=f"已{verb} {affected} 条")
    if action == "delete":
        # 业务引用检查（影响病历准确性的引用 → 阻止删除，建议下架）
        #   PrescriptionItem / SalesOrderItem / Vaccination / Deworming / PackageRedemption
        # 内部引用（库存自身衍生数据 → 删除时手动级联清掉，SQLite 默认不启用 FK CASCADE）
        #   InventoryTransaction / InventoryBatch
        from app.models import (
            PrescriptionItem, InventoryTransaction, InventoryBatch,
            Vaccination,
        )
        try:
            from app.models import SalesOrderItem
        except ImportError:
            SalesOrderItem = None
        try:
            from app.models import Deworming
        except ImportError:
            Deworming = None
        try:
            from app.models import PackageRedemption
        except ImportError:
            PackageRedemption = None

        for iid in ids:
            it = db.get(InventoryItem, iid)
            if not it:
                continue
            # 1. 检查业务引用（处方/销售/疫苗/驱虫/套餐核销）— 任一存在则阻止
            ref_count = 0
            try:
                ref_count += db.query(PrescriptionItem).filter(PrescriptionItem.item_id == iid).count()
            except Exception:
                pass
            if SalesOrderItem is not None:
                try:
                    ref_count += db.query(SalesOrderItem).filter(SalesOrderItem.item_id == iid).count()
                except Exception:
                    pass
            try:
                ref_count += db.query(Vaccination).filter(Vaccination.inventory_item_id == iid).count()
            except Exception:
                pass
            if Deworming is not None:
                try:
                    ref_count += db.query(Deworming).filter(Deworming.inventory_item_id == iid).count()
                except Exception:
                    pass
            if PackageRedemption is not None:
                try:
                    ref_count += db.query(PackageRedemption).filter(PackageRedemption.item_id == iid).count()
                except Exception:
                    pass
            if ref_count > 0:
                blocked += 1
                continue
            # 2. 手动清理内部引用（SQLite 不自动 CASCADE）
            try:
                db.query(InventoryTransaction).filter(InventoryTransaction.item_id == iid).delete(synchronize_session=False)
            except Exception:
                pass
            try:
                db.query(InventoryBatch).filter(InventoryBatch.item_id == iid).delete(synchronize_session=False)
            except Exception:
                pass
            # 3. 删除品目本身
            db.delete(it)
            affected += 1
        db.commit()
        msg = f"已删除 {affected} 条"
        if blocked:
            msg += f"，{blocked} 条因有业务记录（处方/销售/疫苗等）引用被保留（请改用下架）"
        return _redirect_back_with_msg(request, "/admin/inventory", msg=msg)
    return RedirectResponse("/admin/inventory?err=未知操作", status_code=303)


@app.get("/admin/inventory/import-photo", response_class=HTMLResponse)
async def admin_inventory_import_photo_page(request: Request, db: Session = Depends(get_db)):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    return templates.TemplateResponse(request, "uk/inventory_import_photo.html", {
        "csrf_token": _get_csrf_token(request),
        "categories": INVENTORY_CATEGORIES,
    })


@app.post("/admin/inventory/import-photo/recognize")
async def admin_inventory_import_photo_recognize(
    request: Request,
    db: Session = Depends(get_db),
    files: list[UploadFile] = File(...),
):
    """接收 1~5 张进货单图片，调多模态 OCR，返回识别结果 JSON。"""
    if not request.session.get("admin"):
        return {"ok": False, "error": "未登录"}
    if not files:
        return {"ok": False, "error": "未上传图片"}
    if len(files) > 5:
        return {"ok": False, "error": "最多 5 张图片"}

    import tempfile, os as _os
    from app.services.purchase_ocr import recognize_purchase_photo, match_item_by_name

    saved_paths: list[Path] = []
    tmp_dir = Path(tempfile.mkdtemp(prefix="purchase_ocr_"))
    try:
        for f in files:
            suf = (Path(f.filename or "").suffix or ".jpg").lower()
            if suf not in (".jpg", ".jpeg", ".png", ".webp"):
                continue
            p = tmp_dir / f"{secrets.token_hex(8)}{suf}"
            p.write_bytes(await f.read())
            saved_paths.append(p)
        if not saved_paths:
            return {"ok": False, "error": "没有有效图片（仅支持 jpg/png/webp）"}
        result = await recognize_purchase_photo(saved_paths)
    finally:
        # 清理临时图（保留 30 秒供调试看错误时也无所谓，直接清）
        for p in saved_paths:
            try: p.unlink()
            except Exception: pass
        try: tmp_dir.rmdir()
        except Exception: pass

    if not result["ok"]:
        return {"ok": False, "error": result.get("error", "识别失败"), "raw": result.get("raw", "")}

    # 对每行做品目匹配（按当前用户门店过滤，避免跨店误匹配）
    _match_q = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    _match_q = _apply_store_filter(_match_q, InventoryItem.store, _get_op_store(request))
    all_items = _match_q.all()
    for it in result["items"]:
        matched_id, conf = match_item_by_name(it["name"], all_items)
        it["matched_id"] = matched_id
        it["match_confidence"] = round(conf, 2)
        if matched_id:
            m = next((x for x in all_items if x.id == matched_id), None)
            if m:
                it["matched_name"] = m.name
                it["matched_unit"] = m.unit
                it["matched_unit2"] = m.unit2 or ""
                it["matched_unit2_ratio"] = float(m.unit2_ratio or 1.0)
                it["matched_stock"] = float(m.stock_qty or 0)
                it["matched_sell_price"] = float(m.sell_price or 0)
                it["matched_cost_price"] = float(m.cost_price or 0)
                # 高置信匹配时，把已有品目的 spec/unit 反填进 OCR 行，省得用户手填
                if conf >= 0.85:
                    if not it.get("spec") and m.notes:
                        # InventoryItem 没有专门 spec 字段，规格通常在 notes 里
                        it["spec"] = (m.notes or "").strip()[:40]
                    if not it.get("main_unit"):
                        it["main_unit"] = m.unit or ""
                    if not it.get("pack_unit"):
                        it["pack_unit"] = m.unit2 or m.unit or ""
                    if not it.get("pack_size") and (m.unit2_ratio or 0) > 1:
                        it["pack_size"] = float(m.unit2_ratio)
    return {"ok": True, "items": result["items"]}


@app.post("/admin/inventory/import-photo/commit")
async def admin_inventory_import_photo_commit(
    request: Request,
    db: Session = Depends(get_db),
):
    """提交确认后的入库表单。表单字段（多行）：
       row_count = N
       row{i}_action = create | reuse | skip
       row{i}_item_id = N (reuse 时必填)
       row{i}_name / spec / qty / unit / unit_price / batch_no / expiry_date
    """
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    try:
        n = int(form.get("row_count") or 0)
    except (TypeError, ValueError):
        n = 0
    if n <= 0:
        return RedirectResponse("/admin/inventory/import-photo?msg=没有行可入库", status_code=303)

    operator = request.session.get("admin_username", "admin")
    created = 0
    stocked_in = 0
    skipped = 0
    merged = 0
    # 同批次内的 create 去重：normalize(name+spec) → 已创建 item
    from app.services.purchase_ocr import dedup_key as _dedup_key
    inbatch_created: dict[str, InventoryItem] = {}

    for i in range(n):
        action = (form.get(f"row{i}_action") or "skip").strip()
        if action == "skip":
            skipped += 1
            continue
        name = (form.get(f"row{i}_name") or "").strip()
        if not name:
            skipped += 1
            continue
        try:
            qty = float(form.get(f"row{i}_qty") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        if qty <= 0:
            skipped += 1
            continue
        try:
            unit_price = float(form.get(f"row{i}_unit_price") or 0)
        except (TypeError, ValueError):
            unit_price = 0.0
        unit = (form.get(f"row{i}_unit") or "").strip() or "个"
        batch_no = (form.get(f"row{i}_batch_no") or "").strip()
        expiry = (form.get(f"row{i}_expiry_date") or "").strip()[:10]
        spec = (form.get(f"row{i}_spec") or "").strip()
        # 大类 / 小类（用户在表单选；空 fallback 到 medication / general 兼容旧行为）
        row_category = (form.get(f"row{i}_category") or "").strip() or "medication"
        row_subcategory = (form.get(f"row{i}_subcategory") or "").strip()
        # 包装信息（OCR 新增字段）— 创建时记到 unit2/unit2_ratio；reuse 时用于单位换算
        try:
            pack_size = float(form.get(f"row{i}_pack_size") or 0)
        except (TypeError, ValueError):
            pack_size = 0.0
        main_unit_in = (form.get(f"row{i}_main_unit") or "").strip()
        # 兼容：副单位字段名同时支持 row{i}_pack_unit（旧字段）与 row{i}_unit（OCR 购买单位 = 副单位）
        pack_unit_in = (form.get(f"row{i}_pack_unit") or form.get(f"row{i}_unit") or "").strip()
        # 默认售价（人工填，可空 → 不动）
        try:
            sell_price_in = float(form.get(f"row{i}_sell_price") or 0)
        except (TypeError, ValueError):
            sell_price_in = 0.0

        # 取得或新建 item
        if action == "reuse":
            try:
                item_id = int(form.get(f"row{i}_item_id") or 0)
            except (TypeError, ValueError):
                item_id = 0
            item = db.get(InventoryItem, item_id) if item_id else None
            if not item:
                skipped += 1
                continue
            # 累加时把进货单上的标准名追加到 aliases，下次同样名能命中
            from app.services.purchase_ocr import add_alias_to_item as _add_alias
            _add_alias(item, name)
        else:  # create
            # 同批次内去重：相同 normalize(name+spec) 已经在本次创建过 → 直接复用
            dkey = _dedup_key(name, spec)
            existing_in_batch = inbatch_created.get(dkey)
            if existing_in_batch:
                item = existing_in_batch
                merged += 1
            else:
                # 还要再防一道：可能本次跑里有 reuse 操作刚把它建过；
                # 或同名 item 已在 DB 但客户端没匹配上（保险查一遍）
                from app.services.purchase_ocr import _normalize, _strip_brand_prefix
                tnorm = _normalize(_strip_brand_prefix(name))
                if tnorm:
                    _dup_q = db.query(InventoryItem).filter(
                        InventoryItem.is_active == True,
                        InventoryItem.name == name,
                    )
                    _dup_q = _apply_store_filter(_dup_q, InventoryItem.store, _get_op_store(request))
                    db_dup = _dup_q.first()
                    if db_dup:
                        item = db_dup
                        merged += 1
                        inbatch_created[dkey] = item
                    else:
                        item = None
                else:
                    item = None
                if item is None:
                    # 优先使用 OCR 识别的 main_unit；若无则退回 unit
                    _main_u = main_unit_in or unit
                    _pack_u = pack_unit_in or unit
                    _ratio = pack_size if pack_size > 1 else 1.0
                    # 若包装单位与主单位不同 → 把包装写进 unit2
                    if _pack_u and _pack_u != _main_u and _ratio > 1:
                        _unit2 = _pack_u
                        _unit2_ratio = _ratio
                    else:
                        _unit2 = ""
                        _unit2_ratio = 1.0
                    # 进价：OCR 是 ¥/pack_unit，统一换算成 ¥/main_unit 入库
                    _cost_per_main = unit_price / _ratio if _ratio > 1 else unit_price
                    item = InventoryItem(
                        name=name[:200],
                        category=row_category[:60],
                        subcategory=row_subcategory[:60],
                        unit=_main_u[:20],
                        unit2=_unit2[:20],
                        unit2_ratio=_unit2_ratio,
                        sell_price=sell_price_in if sell_price_in > 0 else 0.0,
                        cost_price=_cost_per_main,
                        stock_qty=0.0,
                        low_stock_min=0.0,
                        notes=spec,
                        created_by=operator,
                        is_active=True,
                        store=_resolve_store_for_create(request),
                    )
                    db.add(item)
                    db.flush()
                    inbatch_created[dkey] = item
                    created += 1

        # 单位换算：OCR 行的 unit 可能是包装单位（盒），item.unit 是主单位（片）
        # 若 unit == item.unit2（且有 ratio）→ 实际入库数量 = qty * ratio
        effective_qty = qty
        effective_unit_price = unit_price
        if item.unit2 and unit and unit == item.unit2 and (item.unit2_ratio or 1) > 1:
            effective_qty = qty * float(item.unit2_ratio)
            # 进价同步换算成 ¥/主单位
            effective_unit_price = unit_price / float(item.unit2_ratio) if unit_price > 0 else 0.0
        # 累加库存 + 写流水
        qty_before = float(item.stock_qty or 0)
        item.stock_qty = qty_before + effective_qty
        if effective_unit_price > 0:
            item.cost_price = effective_unit_price  # 用最新进价更新（按主单位）
        # 售价：用户在表单显式填了且与现有不同 → 更新（按主单位）
        if sell_price_in > 0 and abs(float(item.sell_price or 0) - sell_price_in) > 0.001:
            item.sell_price = sell_price_in
        # 流水按"换算后"的主单位记
        _note_pack = ""
        if effective_qty != qty:
            _note_pack = f" · 原 {qty}{unit} × {item.unit2_ratio:g}/{unit} = {effective_qty:g}{item.unit}"
        db.add(InventoryTransaction(
            item_id=item.id,
            tx_type="in",
            qty=effective_qty,
            qty_before=qty_before,
            qty_after=item.stock_qty,
            unit_price=effective_unit_price,
            ref_type="manual",
            operator=operator,
            note=f"进货单照片识别入库" + (f" · 批号{batch_no}" if batch_no else "") + _note_pack,
        ))
        # 批次
        if batch_no or expiry:
            from datetime import date as _date
            db.add(InventoryBatch(
                item_id=item.id,
                batch_no=batch_no[:80],
                quantity=effective_qty,
                expiry_date=expiry,
                received_date=_date.today().isoformat(),
                notes=spec[:500] if spec else "",
            ))
        stocked_in += 1

    db.commit()
    _audit(db, request, "inventory_import_photo", application_id=None,
           detail={"new": created, "stocked": stocked_in, "skipped": skipped, "merged": merged})
    db.commit()
    parts = [f"新增 {created} 个品目", f"{stocked_in} 笔入库"]
    if merged: parts.append(f"自动合并 {merged} 行同名")
    if skipped: parts.append(f"跳过 {skipped} 行")
    msg = f"入库完成：{'，'.join(parts)}"
    nu = str(form.get("next_url", "")).strip()
    if nu:
        sep = "&" if "?" in nu else "?"
        return RedirectResponse(_safe_next(nu + sep + "msg=" + msg, f"/admin/inventory?msg={msg}"), status_code=303)
    return RedirectResponse(f"/admin/inventory?msg={msg}", status_code=303)


@app.get("/admin/inventory", response_class=HTMLResponse)
async def admin_inventory_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    category: str = "",
    subcategory: str = "",
    low_stock: str = "",
    zero_stock: str = "",
    controlled: str = "",
    service_only: str = "",
    expiry_alert: str = "",
    store: str = "",
    page: int = 1,
):
    require_admin(request)
    from datetime import date as _date, timedelta as _timedelta
    page_size = 50
    query = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    # 多门店过滤：staff 自动看本店+通用；superadmin 可通过 ?store= 切
    _admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        _wb_store = (store or "").strip()
    else:
        _wb_store = _admin_store
    query = _apply_store_filter(query, InventoryItem.store, _wb_store)
    if q:
        query = query.filter(
            or_(InventoryItem.name.ilike(f"%{q}%"),
                InventoryItem.supplier.ilike(f"%{q}%"))
        )
    if category:
        query = query.filter(InventoryItem.category == category)
    if subcategory:
        query = query.filter(InventoryItem.subcategory == subcategory)
    if low_stock == "1":
        query = query.filter(
            InventoryItem.is_service == False,
            InventoryItem.stock_qty <= InventoryItem.low_stock_min,
            InventoryItem.low_stock_min > 0,
        )
    if zero_stock == "1":
        query = query.filter(InventoryItem.is_service == False, InventoryItem.stock_qty <= 0)
    if controlled == "1":
        query = query.filter(InventoryItem.is_controlled == True)
    if service_only == "1":
        query = query.filter(InventoryItem.is_service == True)
    if expiry_alert == "1":
        alert_date = (_date.today() + _timedelta(days=90)).isoformat()
        expiry_ids = (db.query(InventoryBatch.item_id)
                      .filter(InventoryBatch.is_depleted == False,
                              InventoryBatch.expiry_date != "",
                              InventoryBatch.expiry_date <= alert_date)
                      .distinct().subquery())
        query = query.filter(InventoryItem.id.in_(expiry_ids))
    total = query.count()
    items = query.order_by(InventoryItem.category, InventoryItem.name).offset((page - 1) * page_size).limit(page_size).all()
    total_pages = max(1, (total + page_size - 1) // page_size)
    low_count = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _wb_store
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= InventoryItem.low_stock_min,
        InventoryItem.low_stock_min > 0,
    ).count()
    zero_count = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _wb_store
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= 0,
    ).count()
    _alert_date = (_date.today() + _timedelta(days=90)).isoformat()
    expiry_count = (db.query(InventoryBatch.item_id)
                    .filter(InventoryBatch.is_depleted == False,
                            InventoryBatch.expiry_date != "",
                            InventoryBatch.expiry_date <= _alert_date)
                    .distinct().count())
    return templates.TemplateResponse(request, "uk/inventory.html", {  # B4 UK 重写；旧模板暂留
        "request": request, "items": items, "total": total,
        "page": page, "total_pages": total_pages,
        "q": q, "category": category, "subcategory": subcategory, "low_stock": low_stock,
        "zero_stock": zero_stock, "controlled": controlled, "service_only": service_only,
        "expiry_alert": expiry_alert,
        "categories": INVENTORY_CATEGORIES, "low_count": low_count, "zero_count": zero_count,
        "expiry_count": expiry_count,
        "csrf_token": _get_csrf_token(request),
        "title": "库存管理",
        "wb_store": _wb_store,
        "is_superadmin": request.session.get("admin_role") == "superadmin",
    })


@app.get("/admin/inventory/create", response_class=HTMLResponse)
async def admin_inventory_create_form(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    return templates.TemplateResponse(request, "uk/inventory_form.html", {
        "request": request, "item": None,
        "categories": INVENTORY_CATEGORIES,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": "新增品目",
        "default_store": _get_admin_store(request),
        "is_superadmin": request.session.get("admin_role") == "superadmin",
    })


@app.post("/admin/inventory/create")
async def admin_inventory_create(
    request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""), category: str = Form("medication"),
    subcategory: str = Form(""), is_service: str = Form("0"),
    is_controlled: str = Form("0"),
    report_exempt: str = Form("0"),
    single_use_pack: str = Form("0"),
    unit: str = Form("个"), unit2: str = Form(""), unit2_ratio: float = Form(1.0),
    sell_price: float = Form(0.0), cost_price: float = Form(0.0),
    stock_qty: float = Form(0.0), low_stock_min: float = Form(0.0),
    supplier: str = Form(""), notes: str = Form(""),
    store: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    if not name.strip():
        raise HTTPException(400, "品名不能为空")
    operator = request.session.get("admin_username", "")
    item = InventoryItem(
        name=name.strip(), category=category, subcategory=subcategory,
        is_service=(is_service == "1"), is_controlled=(is_controlled == "1"),
        requires_report=(report_exempt != "1"),  # 反向：勾免报告 → requires_report=False
        single_use_pack=(single_use_pack == "1"),
        unit=unit, unit2=unit2, unit2_ratio=unit2_ratio,
        sell_price=sell_price, cost_price=cost_price,
        stock_qty=stock_qty, low_stock_min=low_stock_min,
        supplier=supplier, notes=notes, created_by=operator,
        store=_resolve_store_for_create(request, store),
    )
    db.add(item)
    db.flush()
    # 若初始库存 > 0，记录入库流水
    if stock_qty > 0 and not (is_service == "1"):
        db.add(InventoryTransaction(
            item_id=item.id, tx_type="in", qty=stock_qty,
            qty_before=0, qty_after=stock_qty,
            unit_price=cost_price, ref_type="manual",
            operator=operator, note="初始库存录入",
        ))
    db.commit()
    return RedirectResponse(f"/admin/inventory?msg=已创建：{item.name}", status_code=303)


@app.get("/admin/inventory/{item_id}/edit", response_class=HTMLResponse)
async def admin_inventory_edit_form(item_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404)
    from app.services.pricing import parse_overrides as _po
    item.store_overrides_parsed = _po(item)
    # 别名：JSON → 一行一个，给 textarea 用
    aliases_text = ""
    try:
        arr = json.loads(item.aliases) if (item.aliases or "").strip() else []
        if isinstance(arr, list):
            aliases_text = "\n".join(str(a) for a in arr if str(a).strip())
    except Exception:
        aliases_text = ""
    return templates.TemplateResponse(request, "uk/inventory_form.html", {
        "request": request, "item": item,
        "categories": INVENTORY_CATEGORIES,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": f"编辑品目：{item.name}",
        "default_store": item.store or _get_admin_store(request),
        "is_superadmin": request.session.get("admin_role") == "superadmin",
        "aliases_text": aliases_text,
    })


@app.get("/admin/inventory/{item_id}", response_class=HTMLResponse)
async def admin_inventory_detail(item_id: int, request: Request, db: Session = Depends(get_db),
                                  page: int = 1):
    require_admin(request)
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404)
    page_size = 30
    txs = (db.query(InventoryTransaction)
           .filter(InventoryTransaction.item_id == item_id)
           .order_by(InventoryTransaction.created_at.desc())
           .offset((page - 1) * page_size).limit(page_size).all())
    tx_total = db.query(InventoryTransaction).filter(InventoryTransaction.item_id == item_id).count()
    batches = (db.query(InventoryBatch)
               .filter(InventoryBatch.item_id == item_id)
               .order_by(InventoryBatch.expiry_date)
               .all())
    from datetime import date as _date, timedelta as _timedelta
    today_str = _date.today().isoformat()
    alert_date_str = (_date.today() + _timedelta(days=90)).isoformat()
    return templates.TemplateResponse(request, "uk/inventory_detail.html", {
        "request": request, "item": item, "txs": txs,
        "tx_total": tx_total, "page": page,
        "tx_pages": max(1, (tx_total + page_size - 1) // page_size),
        "batches": batches, "today_str": today_str, "alert_date_str": alert_date_str,
        "categories": INVENTORY_CATEGORIES,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": f"品目详情：{item.name}",
    })


@app.post("/admin/inventory/{item_id}/edit")
async def admin_inventory_edit(
    item_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    name: str = Form(""), category: str = Form("medication"),
    subcategory: str = Form(""), is_service: str = Form("0"),
    is_controlled: str = Form("0"),
    report_exempt: str = Form("0"),
    single_use_pack: str = Form("0"),
    unit: str = Form("个"), unit2: str = Form(""), unit2_ratio: float = Form(1.0),
    sell_price: float = Form(0.0), cost_price: float = Form(0.0),
    low_stock_min: float = Form(0.0),
    supplier: str = Form(""), notes: str = Form(""),
    store: str = Form(""),
    # 门店覆盖价（方案 H）
    override_sell_donghuan: str = Form(""), override_cost_donghuan: str = Form(""),
    override_sell_henggang: str = Form(""), override_cost_henggang: str = Form(""),
    # 进货识别别名（一行一个）
    aliases_text: str = Form(""),
    # 从列表第几页跳来的（保存后回该页，避免每次回到第 1 页）
    from_page: str = Form(""),
    # 列表筛选条件 querystring（保存后回原筛选状态，
    # 例如选了大类=medication & 小类=controlled & page=3）
    from_qs: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404)
    item.name = name.strip() or item.name
    # 别名：去重 + 去空，按行解析
    if aliases_text is not None:
        lines = [ln.strip() for ln in aliases_text.splitlines() if ln.strip()]
        seen, deduped = set(), []
        for ln in lines:
            k = ln.lower().replace(" ", "")
            if k in seen:
                continue
            seen.add(k)
            deduped.append(ln[:200])
        item.aliases = json.dumps(deduped[:8], ensure_ascii=False) if deduped else ""
    item.category = category; item.subcategory = subcategory
    item.is_service = (is_service == "1"); item.is_controlled = (is_controlled == "1")
    item.requires_report = (report_exempt != "1")
    item.single_use_pack = (single_use_pack == "1")
    item.unit = unit; item.unit2 = unit2; item.unit2_ratio = unit2_ratio
    item.sell_price = sell_price; item.cost_price = cost_price
    item.low_stock_min = low_stock_min
    item.supplier = supplier; item.notes = notes
    # 门店价格覆盖已废弃（每个品目独立归属一家店）。任何编辑都强制清空 store_overrides，
    # 避免老数据继续影响 eff_price 计算。
    item.store_overrides = ""
    # 门店：staff 无权改；superadmin 可改，但必须是 东环店 / 横岗店 之一
    if request.session.get("admin_role") == "superadmin":
        new_store = (store or "").strip()
        if new_store not in ("东环店", "横岗店"):
            raise HTTPException(400, "归属门店必须是 东环店 或 横岗店，不允许空（每个品目必须归属一家店）")
        item.store = new_store
    item.updated_at = datetime.utcnow()
    db.commit()
    # 保存后优先回列表的原筛选状态 + 原页码（避免编辑后弹回第 1 页 / 弹回"全部"）
    # from_qs 是 URL-encoded querystring，含 q / category / subcategory / page 等
    _msg = f"已保存：{item.name[:30]}"
    if from_qs:
        # 安全：from_qs 是表单填的，可能被篡改；只接受 & 字符分隔的 k=v 串
        # 不允许出现 / 开头（防开放重定向）
        if "/" not in from_qs and "\\" not in from_qs:
            from urllib.parse import quote
            return RedirectResponse(
                f"/admin/inventory?{from_qs}&msg={quote(_msg)}",
                status_code=303,
            )
    if (from_page or "").strip().isdigit():
        return RedirectResponse(
            f"/admin/inventory?page={int(from_page)}&msg={_msg}",
            status_code=303,
        )
    return RedirectResponse(f"/admin/inventory/{item_id}?msg=已保存", status_code=303)


# ─── 小类规范化工具：扫脏数据 + 自动匹配 + 审批后批量改 ───
def _build_subcat_reverse_index() -> dict:
    """{中文 label → [(cat_key, sub_key, label), ...]}，含 alias 多对一。"""
    idx: dict[str, list] = {}
    # 直接 label 反查
    for cat_key, cat in INVENTORY_CATEGORIES.items():
        for sub_key, label in cat.get("subs", {}).items():
            idx.setdefault(label, []).append((cat_key, sub_key, label))
    # 常见 alias 手工映射（你描述的脏数据里出现的）
    alias = {
        "兽药+保健":   ("medication", "general", "普通药品"),
        "一般处置":     ("treatment",  "routine", "一般处置"),
        "院外实验室":   ("lab",        "external_lab", "院外送检"),
        "院外送检":     ("lab",        "external_lab", "院外送检"),
        "骨科手术":     ("surgery",    "orthopedic", "骨科手术"),
        "软组织手术":   ("surgery",    "general",    "普外科手术"),
        "眼科手术":     ("surgery",    "ophthalmic", "眼科手术"),
        "口腔手术":     ("surgery",    "dental",     "口腔手术"),
        "神经外科":     ("surgery",    "neuro",      "神经外科"),
        "麻醉处置":     ("treatment",  "anesthesia", "麻醉处置"),
        "紧急处置":     ("treatment",  "emergency",  "紧急处置"),
        "挂号":         ("reception",  "registration", "挂号"),
        "咨询":         ("reception",  "consultation", "咨询"),
        "接诊":         ("reception",  "registration", "挂号"),
        "门诊挂号":     ("reception",  "registration", "挂号"),
        "麻药/精神类":  ("medication", "controlled", "麻药/精神类"),
        "麻药":         ("medication", "controlled", "麻药/精神类"),
        "精神类":       ("medication", "controlled", "麻药/精神类"),
        "管控":         ("medication", "controlled", "麻药/精神类"),
        "普通药品":     ("medication", "general",    "普通药品"),
        "狂犬":         ("vaccine",    "rabies",     "狂犬疫苗"),
        "狂犬疫苗":     ("vaccine",    "rabies",     "狂犬疫苗"),
        "联苗":         ("vaccine",    "combo",      "联苗"),
        "其他疫苗":     ("vaccine",    "other",      "其他疫苗"),
        "体内驱虫":     ("antiparasitic", "internal", "体内驱虫"),
        "体外驱虫":     ("antiparasitic", "external", "体外驱虫"),
        "体内外同驱":   ("antiparasitic", "both",     "体内外同驱"),
        "普通耗材":     ("consumable", "general",    "普通耗材"),
        "耗材":         ("consumable", "general",    "普通耗材"),
        "普通商品":     ("product",    "general",    "普通商品"),
        "商品":         ("product",    "general",    "普通商品"),
        "洗护":         ("grooming",   "washcare",   "洗护"),
        "造型":         ("grooming",   "styling",    "造型"),
        "附加服务":     ("grooming",   "addon",      "附加服务"),
        "常规化验":     ("lab",        "routine_lab","常规化验"),
        "DR":           ("imaging",    "dr",         "DR"),
        "CT":           ("imaging",    "ct",         "CT"),
        "核磁共振":     ("imaging",    "mri",        "核磁共振"),
        "B超":          ("imaging",    "ultrasound", "B超"),
        "常规光学显微镜": ("microscopy","optical",   "常规光学显微镜"),
        "电子显微镜":   ("microscopy", "electron",   "电子显微镜"),
        "普外科手术":   ("surgery",    "general",    "普外科手术"),
        "普通护理":     ("nursing",    "general",    "普通护理"),
        "隔离护理":     ("nursing",    "isolation",  "隔离护理"),
    }
    for k, v in alias.items():
        idx.setdefault(k, []).append(v)
    return idx


def _match_subcat(item, idx: dict) -> dict:
    """返回 {status, new_category, new_subcategory, new_label, hint}。
    status: 'clean' 已规范 / 'match' 自动匹配到 / 'ambiguous' 多候选 / 'empty' 原值空 / 'unknown' 找不到
    """
    cur_cat = item.category or ""
    cur_sub = (item.subcategory or "").strip()
    # 已规范
    if cur_cat in INVENTORY_CATEGORIES and cur_sub in INVENTORY_CATEGORIES[cur_cat].get("subs", {}):
        return {"status": "clean", "new_category": cur_cat, "new_subcategory": cur_sub,
                "new_label": INVENTORY_CATEGORIES[cur_cat]["subs"][cur_sub], "hint": ""}
    if not cur_sub:
        return {"status": "empty", "new_category": cur_cat, "new_subcategory": "",
                "new_label": "", "hint": "原 subcategory 为空"}
    # 反查
    cands = idx.get(cur_sub) or []
    if not cands:
        # fuzzy
        import difflib
        all_labels = list(idx.keys())
        close = difflib.get_close_matches(cur_sub, all_labels, n=1, cutoff=0.7)
        if close:
            cands = idx.get(close[0]) or []
            hint_prefix = f"模糊匹配「{close[0]}」"
        else:
            return {"status": "unknown", "new_category": cur_cat, "new_subcategory": cur_sub,
                    "new_label": cur_sub, "hint": "找不到对应规范小类，建议人工处理"}
    else:
        hint_prefix = ""
    # 多候选：优先保留当前 category
    same_cat = [c for c in cands if c[0] == cur_cat]
    picked = same_cat[0] if same_cat else cands[0]
    status = "match" if len(cands) == 1 else "ambiguous"
    hint = (hint_prefix + f"候选 {len(cands)} 个").strip()
    if picked[0] != cur_cat:
        hint = (hint + f" · 大类将由 {cur_cat} → {picked[0]}").strip(" ·")
    return {"status": status, "new_category": picked[0], "new_subcategory": picked[1],
            "new_label": picked[2], "hint": hint}


@app.get("/admin/inventory-cleanup", response_class=HTMLResponse)
async def admin_inv_cleanup_preview(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    require_superadmin(request)
    idx = _build_subcat_reverse_index()
    items = db.query(InventoryItem).filter(InventoryItem.is_active == True).order_by(InventoryItem.id).all()
    rows = []
    counts = {"clean": 0, "match": 0, "ambiguous": 0, "unknown": 0, "empty": 0}
    for it in items:
        r = _match_subcat(it, idx)
        counts[r["status"]] += 1
        if r["status"] == "clean":
            continue
        cur_cat_label = INVENTORY_CATEGORIES.get(it.category, {}).get("label", it.category or "—")
        new_cat_label = INVENTORY_CATEGORIES.get(r["new_category"], {}).get("label", r["new_category"] or "—")
        rows.append({
            "id": it.id, "name": it.name,
            "cur_cat": it.category or "", "cur_cat_label": cur_cat_label,
            "cur_sub": it.subcategory or "",
            "new_cat": r["new_category"], "new_cat_label": new_cat_label,
            "new_sub": r["new_subcategory"], "new_label": r["new_label"],
            "status": r["status"], "hint": r["hint"],
            "cat_change": (it.category or "") != r["new_category"],
        })
    return templates.TemplateResponse(request, "uk/inventory_cleanup.html", {
        "rows": rows, "counts": counts,
        "csrf_token": _get_csrf_token(request),
        "title": "小类规范化工具",
    })


@app.post("/admin/inventory-cleanup")
async def admin_inv_cleanup_apply(
    request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    apply_ids: list[int] = Form(default=[]),
):
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    if not apply_ids:
        return RedirectResponse("/admin/inventory-cleanup?msg=未选择任何条目", status_code=303)
    idx = _build_subcat_reverse_index()
    updated = 0
    cat_changed = 0
    for iid in apply_ids:
        it = db.get(InventoryItem, iid)
        if not it:
            continue
        r = _match_subcat(it, idx)
        if r["status"] in ("clean", "unknown"):
            continue
        if r["new_category"] != (it.category or ""):
            cat_changed += 1
            it.category = r["new_category"]
        it.subcategory = r["new_subcategory"]
        it.updated_at = datetime.utcnow()
        updated += 1
    db.commit()
    return RedirectResponse(
        f"/admin/inventory-cleanup?msg=已规范化 {updated} 个品目（其中 {cat_changed} 个连大类一起改了）",
        status_code=303,
    )


# ─── 列表 vs 编辑表单一致性审计（只读） ───
@app.get("/admin/inventory-audit", response_class=HTMLResponse)
async def admin_inventory_audit(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    require_superadmin(request)
    items = db.query(InventoryItem).filter(InventoryItem.is_active == True)\
        .order_by(InventoryItem.id).all()
    issues = []  # 仅记录不一致
    n_total = 0
    n_ok = 0
    for it in items:
        n_total += 1
        problems = []

        # === 大类一致性 ===
        # 列表渲染：categories[it.category].label if found else it.category（fallback 原字符串）
        # 表单渲染：select 里如果 it.category 是 categories 的 key → 显示该 cat.label
        #          否则没有 selected option → 浏览器默认选第一个 = "药品"
        cat_in_dict = it.category in INVENTORY_CATEGORIES
        list_cat = INVENTORY_CATEGORIES[it.category]["label"] if cat_in_dict else (it.category or "—")
        if cat_in_dict:
            form_cat = INVENTORY_CATEGORIES[it.category]["label"]
        else:
            # 无匹配 → 表单 select 落到第一个 option = "药品"（medication）
            form_cat = INVENTORY_CATEGORIES["medication"]["label"] + " ⚠默认"
            problems.append(f"大类不一致：列表「{list_cat}」 vs 表单「{form_cat}」")

        # === 小类一致性 ===
        cur_sub = (it.subcategory or "").strip()
        # 列表 fallback：categories[it.category].subs[cur_sub] if found else cur_sub or "—"
        if cat_in_dict and cur_sub in INVENTORY_CATEGORIES[it.category].get("subs", {}):
            list_sub = INVENTORY_CATEGORIES[it.category]["subs"][cur_sub]
            form_sub = list_sub  # ✓ 完全匹配
        elif not cur_sub:
            list_sub = "—"
            form_sub = "—（不分小类）"
        else:
            # 不在该 cat 的标准 subs 里：列表显示 raw，表单显示 ⚠ 原值 也是 raw
            list_sub = cur_sub
            # 还要看是不是在其他大类的 subs 里
            in_other = any(cur_sub in c.get("subs", {}) for ck, c in INVENTORY_CATEGORIES.items() if ck != it.category)
            if in_other:
                form_sub = cur_sub + " ⚠ 原值（跨大类）"
                problems.append(f"小类「{cur_sub}」属于其他大类，建议规范化")
            else:
                form_sub = cur_sub + " ⚠ 原值（脏数据）"
                problems.append(f"小类「{cur_sub}」不在任何标准字典")

        if problems:
            issues.append({
                "id": it.id, "name": it.name,
                "list_cat": list_cat, "form_cat": form_cat,
                "list_sub": list_sub, "form_sub": form_sub,
                "unit": it.unit or "—",
                "supplier": it.supplier or "—",
                "stock_qty": it.stock_qty if not it.is_service else None,
                "is_service": it.is_service,
                "store": it.store or "（通用）",
                "problems": problems,
            })
        else:
            n_ok += 1

    return templates.TemplateResponse(request, "uk/inventory_audit.html", {
        "issues": issues, "n_total": n_total, "n_ok": n_ok,
        "csrf_token": _get_csrf_token(request),
        "title": "列表 vs 编辑表单一致性审计",
    })


# ─── UK Minimal 风格 Demo（只读，仅 superadmin）───
@app.get("/admin/uk-demo/inventory", response_class=HTMLResponse)
async def uk_demo_inventory(request: Request, db: Session = Depends(get_db),
                              q: str = "", category: str = ""):
    require_admin(request); require_superadmin(request)
    query = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    if q:
        query = query.filter(or_(InventoryItem.name.ilike(f"%{q}%"),
                                  InventoryItem.supplier.ilike(f"%{q}%")))
    if category:
        query = query.filter(InventoryItem.category == category)
    items = query.order_by(InventoryItem.name).limit(50).all()
    total = db.query(InventoryItem).filter(InventoryItem.is_active == True).count()
    zero_n = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= 0,
    ).count()
    low_n = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_service == False,
        InventoryItem.low_stock_min > 0,
        InventoryItem.stock_qty <= InventoryItem.low_stock_min,
        InventoryItem.stock_qty > 0,
    ).count()
    controlled_n = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_controlled == True,
    ).count()
    return templates.TemplateResponse(request, "uk_demo/inventory.html", {
        "items": items, "total": total,
        "zero_n": zero_n, "low_n": low_n, "controlled_n": controlled_n,
        "q": q, "category": category,
        "categories": INVENTORY_CATEGORIES,
        "active": "inventory",
    })


@app.get("/admin/uk-demo/customer/{cust_id}", response_class=HTMLResponse)
async def uk_demo_customer(cust_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request); require_superadmin(request)
    cust = db.get(Customer, cust_id)
    if not cust:
        raise HTTPException(404)
    pets = db.query(Pet).filter(Pet.customer_id == cust_id).order_by(Pet.id).all()
    wallet = db.query(Wallet).filter(Wallet.customer_id == cust_id).first()
    packages = db.query(CustomerPackage).filter(
        CustomerPackage.customer_id == cust_id,
        CustomerPackage.used_count < CustomerPackage.total_uses,
    ).order_by(CustomerPackage.id.desc()).all()
    visits = db.query(Visit).filter(Visit.customer_id == cust_id)\
        .order_by(Visit.id.desc()).limit(8).all()
    pet_map = {p.id: p for p in pets}
    # 未付款金额估算（简单求和 unpaid invoices）
    from app.models import Invoice
    from sqlalchemy import func as _sqfn
    unpaid_total = db.query(_sqfn.sum(Invoice.total_amount)).filter(
        Invoice.customer_id == cust_id,
        Invoice.payment_status == "unpaid",
    ).scalar() or 0
    return templates.TemplateResponse(request, "uk_demo/customer.html", {
        "cust": cust, "pets": pets, "wallet": wallet,
        "packages": packages, "visits": visits,
        "pet_map": pet_map,
        "unpaid_total": float(unpaid_total),
        "active": "customer",
    })


@app.post("/admin/inventory/{item_id}/stock-in")
async def admin_inventory_stock_in(
    item_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    qty: float = Form(...), unit_price: float = Form(0.0),
    batch_no: str = Form(""), expiry_date: str = Form(""),
    note: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    item = db.get(InventoryItem, item_id)
    if not item or item.is_service:
        raise HTTPException(404)
    if qty <= 0:
        raise HTTPException(400, "入库数量必须大于 0")
    from datetime import date as _date
    before = item.stock_qty
    item.stock_qty += qty
    item.updated_at = datetime.utcnow()
    tx_note = note or ("手动入库" + (f"（批次：{batch_no}）" if batch_no else ""))
    db.add(InventoryTransaction(
        item_id=item_id, tx_type="in", qty=qty,
        qty_before=before, qty_after=item.stock_qty,
        unit_price=unit_price or item.cost_price,
        ref_type="manual", operator=request.session.get("admin_username", ""),
        note=tx_note,
    ))
    if expiry_date:
        db.add(InventoryBatch(
            item_id=item_id,
            batch_no=batch_no,
            quantity=qty,
            expiry_date=expiry_date,
            received_date=_date.today().isoformat(),
            notes=note,
        ))
    db.commit()
    return RedirectResponse(f"/admin/inventory/{item_id}?msg=入库成功+{qty}{item.unit}", status_code=303)


@app.post("/admin/inventory/{item_id}/adjust")
async def admin_inventory_adjust(
    item_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    new_qty: float = Form(...), note: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    item = db.get(InventoryItem, item_id)
    if not item or item.is_service:
        raise HTTPException(404)
    before = item.stock_qty
    diff = new_qty - before
    item.stock_qty = new_qty
    item.updated_at = datetime.utcnow()
    db.add(InventoryTransaction(
        item_id=item_id, tx_type="adjust", qty=abs(diff),
        qty_before=before, qty_after=new_qty,
        unit_price=0, ref_type="manual",
        operator=request.session.get("admin_username", ""),
        note=note or f"盘点调整（{'+' if diff >= 0 else ''}{diff:.1f}）",
    ))
    db.commit()
    return RedirectResponse(f"/admin/inventory/{item_id}?msg=库存已调整", status_code=303)


@app.post("/admin/inventory/{item_id}/batch/{batch_id}/adjust")
async def admin_batch_adjust(
    item_id: int, batch_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    new_qty: float = Form(...),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    batch = (db.query(InventoryBatch)
             .filter(InventoryBatch.id == batch_id, InventoryBatch.item_id == item_id)
             .first())
    if not batch:
        raise HTTPException(404)
    item = db.get(InventoryItem, item_id)
    diff = new_qty - batch.quantity
    before = item.stock_qty
    batch.quantity = new_qty
    if new_qty <= 0:
        batch.is_depleted = True
    item.stock_qty = max(0.0, item.stock_qty + diff)
    item.updated_at = datetime.utcnow()
    db.add(InventoryTransaction(
        item_id=item_id, tx_type="adjust", qty=abs(diff),
        qty_before=before, qty_after=item.stock_qty,
        unit_price=0, ref_type="batch",
        operator=request.session.get("admin_username", ""),
        note=f"批次{batch.batch_no or batch_id}数量修正（{'+' if diff >= 0 else ''}{diff:.1f}）",
    ))
    db.commit()
    return RedirectResponse(f"/admin/inventory/{item_id}?msg=批次已更新", status_code=303)


@app.post("/admin/inventory/{item_id}/batch/{batch_id}/deplete")
async def admin_batch_deplete(
    item_id: int, batch_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    batch = (db.query(InventoryBatch)
             .filter(InventoryBatch.id == batch_id, InventoryBatch.item_id == item_id)
             .first())
    if not batch:
        raise HTTPException(404)
    item = db.get(InventoryItem, item_id)
    before = item.stock_qty
    remaining = batch.quantity
    batch.quantity = 0
    batch.is_depleted = True
    if remaining > 0:
        item.stock_qty = max(0.0, item.stock_qty - remaining)
        item.updated_at = datetime.utcnow()
        db.add(InventoryTransaction(
            item_id=item_id, tx_type="adjust", qty=remaining,
            qty_before=before, qty_after=item.stock_qty,
            unit_price=0, ref_type="batch",
            operator=request.session.get("admin_username", ""),
            note=f"批次{batch.batch_no or batch_id}标记耗尽",
        ))
    db.commit()
    return RedirectResponse(f"/admin/inventory/{item_id}?msg=批次已标记耗尽", status_code=303)


# ──────────────────────────────────────────────────────────────
# 循环盘点
# ──────────────────────────────────────────────────────────────

@app.get("/admin/stocktake", response_class=HTMLResponse)
async def admin_stocktake_list(request: Request, db: Session = Depends(get_db),
                                store: str = Query("")):
    require_admin(request)
    from datetime import date as _date, timedelta as _timedelta
    # 决定本次浏览的门店：
    #   - staff：永远是本店，忽略 ?store=
    #   - superadmin：?store= 优先；没传则 session.admin_store；都没有 → 强制选一个
    is_super = request.session.get("admin_role") == "superadmin"
    admin_store = request.session.get("admin_store", "") or ""
    if is_super:
        wb_store = (store or "").strip()
        if wb_store not in ("东环店", "横岗店"):
            wb_store = admin_store if admin_store in ("东环店", "横岗店") else "横岗店"
    else:
        wb_store = admin_store if admin_store in ("东环店", "横岗店") else ""
        if not wb_store:
            raise HTTPException(400, "员工账号缺少有效门店归属，请联系管理员")
    sessions = (db.query(StocktakeSession)
                .order_by(StocktakeSession.created_at.desc())
                .limit(30).all())
    # 统计每个大类最近盘点时间（仅本店）
    cycle_stats = []
    today = _date.today()
    for cat_key, cat_info in INVENTORY_CATEGORIES.items():
        base_q = db.query(InventoryItem).filter(
            InventoryItem.is_active == True,
            InventoryItem.is_service == False,
            InventoryItem.category == cat_key,
            InventoryItem.store == wb_store,
        )
        item_cnt = base_q.count()
        if item_cnt == 0:
            continue
        last_row = (db.query(InventoryItem.last_counted_at)
                    .filter(InventoryItem.is_active == True,
                            InventoryItem.is_service == False,
                            InventoryItem.category == cat_key,
                            InventoryItem.store == wb_store,
                            InventoryItem.last_counted_at.isnot(None))
                    .order_by(InventoryItem.last_counted_at.desc())
                    .first())
        last_dt = last_row[0].date() if last_row and last_row[0] else None
        days_ago = (today - last_dt).days if last_dt else None
        cycle_stats.append({
            "key": cat_key,
            "label": cat_info["label"],
            "item_count": item_cnt,
            "last_counted": last_dt,
            "days_ago": days_ago,
        })
    open_sessions = [s for s in sessions if s.status == "open"]
    return templates.TemplateResponse(request, "uk/stocktake.html", {
        "request": request, "sessions": sessions,
        "open_sessions": open_sessions,
        "cycle_stats": cycle_stats,
        "categories": INVENTORY_CATEGORIES,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": "循环盘点",
        "wb_store": wb_store,
        "is_super": is_super,
    })


@app.post("/admin/stocktake/create")
async def admin_stocktake_create(
    request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    category_filter: str = Form(""),
    name: str = Form(""),
    store: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    operator = request.session.get("admin_username", "")
    # 决定本次盘点的归属门店
    is_super = request.session.get("admin_role") == "superadmin"
    admin_store = request.session.get("admin_store", "") or ""
    if is_super:
        wb_store = (store or "").strip()
        if wb_store not in ("东环店", "横岗店"):
            wb_store = admin_store if admin_store in ("东环店", "横岗店") else ""
        if wb_store not in ("东环店", "横岗店"):
            return RedirectResponse("/admin/stocktake?err=请先选择门店再发起盘点", status_code=303)
    else:
        wb_store = admin_store if admin_store in ("东环店", "横岗店") else ""
        if not wb_store:
            return RedirectResponse("/admin/stocktake?err=员工账号缺少有效门店归属，请联系管理员", status_code=303)
    query = db.query(InventoryItem).filter(
        InventoryItem.is_active == True,
        InventoryItem.is_service == False,
        InventoryItem.store == wb_store,
    )
    if category_filter:
        query = query.filter(InventoryItem.category == category_filter)
    items = query.order_by(InventoryItem.category, InventoryItem.name).all()
    if not items:
        return RedirectResponse(f"/admin/stocktake?store={wb_store}&err={wb_store}该类别暂无品目", status_code=303)
    cat_label = INVENTORY_CATEGORIES.get(category_filter, {}).get("label", "全部") if category_filter else "全部"
    from datetime import date as _date
    session_name = name.strip() or f"{_date.today()} {wb_store} {cat_label}盘点"
    sess = StocktakeSession(
        name=session_name,
        category_filter=category_filter,
        operator=operator,
        item_count=len(items),
    )
    db.add(sess)
    db.flush()
    for it in items:
        db.add(StocktakeItem(
            session_id=sess.id,
            item_id=it.id,
            item_name=it.name,
            category=it.category,
            unit=it.unit,
            system_qty=it.stock_qty,
        ))
    db.commit()
    return RedirectResponse(f"/admin/stocktake/{sess.id}", status_code=303)


@app.get("/admin/stocktake/{session_id}", response_class=HTMLResponse)
async def admin_stocktake_session(
    session_id: int, request: Request, db: Session = Depends(get_db),
    q: str = "",
):
    require_admin(request)
    sess = db.get(StocktakeSession, session_id)
    if not sess:
        raise HTTPException(404)
    items_q = db.query(StocktakeItem).filter(StocktakeItem.session_id == session_id)
    if q:
        items_q = items_q.filter(StocktakeItem.item_name.ilike(f"%{q}%"))
    sit_items = items_q.order_by(StocktakeItem.category, StocktakeItem.item_name).all()
    counted = sum(1 for x in sit_items if x.actual_qty is not None)
    return templates.TemplateResponse(request, "uk/stocktake_session.html", {
        "request": request, "sess": sess, "sit_items": sit_items,
        "q": q, "counted": counted,
        "categories": INVENTORY_CATEGORIES,
        "csrf_token": request.session.get("csrf_token", ""),
        "title": f"盘点：{sess.name}",
    })


@app.post("/admin/stocktake/{session_id}/delete")
async def admin_stocktake_delete(
    session_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """删除盘点会话（仅 status=open 且超管可操作，已完成的盘点不允许删，保留审计）。"""
    require_admin(request)
    require_superadmin(request)
    _require_csrf(request, csrf_token)
    sess = db.get(StocktakeSession, session_id)
    if not sess:
        raise HTTPException(404)
    if sess.status != "open":
        return RedirectResponse(f"/admin/stocktake?err=该盘点已 {sess.status}，不可删除（已完成的盘点保留审计）", status_code=303)
    name = sess.name
    db.delete(sess)  # 级联删除 stocktake_items
    db.commit()
    from urllib.parse import quote as _q
    return RedirectResponse(f"/admin/stocktake?msg={_q('已删除盘点会话：' + name)}", status_code=303)


@app.get("/admin/stocktake/{session_id}/export")
async def admin_stocktake_export(
    session_id: int, request: Request, db: Session = Depends(get_db),
):
    """导出盘点表为 Excel：打印出来给员工现场实盘填写，汇总后再回到系统录入。"""
    require_admin(request)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl 未安装")
    sess = db.get(StocktakeSession, session_id)
    if not sess:
        raise HTTPException(404)
    rows = (
        db.query(StocktakeItem)
        .filter(StocktakeItem.session_id == session_id)
        .order_by(StocktakeItem.category, StocktakeItem.item_name)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "现场盘点单"
    # 标题
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"现场盘点单 · {sess.name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    # 元信息行
    ws.merge_cells("A2:H2")
    meta_cell = ws["A2"]
    cat_label = INVENTORY_CATEGORIES.get(sess.category_filter, {}).get("label", "全部") if sess.category_filter else "全部"
    meta_cell.value = f"盘点日期：____ 年 __ 月 __ 日    门店：____________    类别：{cat_label}    盘点人：____________    复核人：____________"
    meta_cell.alignment = Alignment(horizontal="left", vertical="center")
    meta_cell.font = Font(size=10)
    # 表头
    headers = ["序号", "品名", "大类", "系统库存", "单位", "实盘数量", "差异", "备注"]
    ws.append([])  # 空行
    ws.append(headers)
    head_row = ws.max_row
    head_fill = PatternFill("solid", fgColor="EFEFEF")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[head_row]:
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = head_fill
        c.border = border
    # 数据行（实盘数量 / 差异 / 备注 留空给手写）
    _CAT = {k: v.get("label", k) for k, v in INVENTORY_CATEGORIES.items()}
    for idx, r in enumerate(rows, start=1):
        ws.append([
            idx,
            r.item_name,
            _CAT.get(r.category, r.category),
            r.system_qty,
            r.unit,
            "",  # 实盘数量 — 手写
            "",  # 差异 — 手算
            "",  # 备注 — 手写
        ])
        # 给每行加边框 + 设行高便于手写
        cur_row = ws.max_row
        ws.row_dimensions[cur_row].height = 22
        for col_idx, c in enumerate(ws[cur_row], start=1):
            c.border = border
            c.alignment = Alignment(horizontal="center" if col_idx not in (2, 8) else "left", vertical="center")
            if col_idx == 4:
                c.number_format = "0.##"
    # 列宽（A4 打印友好）
    widths = {"A": 6, "B": 32, "C": 10, "D": 10, "E": 8, "F": 12, "G": 10, "H": 24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    # 打印设置：A4 横向 + 适配宽度
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"{head_row}:{head_row}"  # 每页重复表头
    # 底部签字栏
    ws.append([])
    last = ws.max_row + 1
    ws.cell(row=last, column=1).value = f"共 {len(rows)} 项 · 已盘 ____ 项 · 差异 ____ 项"
    ws.merge_cells(start_row=last, start_column=1, end_row=last, end_column=4)
    ws.cell(row=last, column=5).value = "盘点员签字："
    ws.cell(row=last, column=7).value = "复核员签字："
    ws.row_dimensions[last].height = 36

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = "".join(c for c in (sess.name or f"session_{session_id}") if c.isalnum() or c in "._- ")[:40].strip() or f"session_{session_id}"
    fname = f"盘点单_{safe_name}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    from urllib.parse import quote as _q
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_q(fname)}"},
    )


@app.post("/admin/stocktake/{session_id}/save")
async def admin_stocktake_save(
    session_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sess = db.get(StocktakeSession, session_id)
    if not sess or sess.status != "open":
        raise HTTPException(400, "盘点已完成或不存在")
    form = await request.form()
    sit_map = {si.id: si for si in db.query(StocktakeItem).filter(StocktakeItem.session_id == session_id).all()}
    for key, val in form.items():
        if key.startswith("actual_"):
            try:
                si_id = int(key.split("_", 1)[1])
                si = sit_map.get(si_id)
                if si and val.strip() != "":
                    si.actual_qty = float(val)
                    si.variance = si.actual_qty - si.system_qty
                elif si and val.strip() == "":
                    si.actual_qty = None
                    si.variance = 0.0
            except (ValueError, IndexError):
                pass
        elif key.startswith("notes_"):
            try:
                si_id = int(key.split("_", 1)[1])
                si = sit_map.get(si_id)
                if si:
                    si.notes = val
            except (ValueError, IndexError):
                pass
    db.commit()
    return RedirectResponse(f"/admin/stocktake/{session_id}?msg=已暂存", status_code=303)


@app.post("/admin/stocktake/{session_id}/submit")
async def admin_stocktake_submit(
    session_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    sess = db.get(StocktakeSession, session_id)
    if not sess or sess.status != "open":
        raise HTTPException(400, "盘点已完成或不存在")
    form = await request.form()
    sit_map = {si.id: si for si in db.query(StocktakeItem).filter(StocktakeItem.session_id == session_id).all()}
    # 先把表单值写入
    for key, val in form.items():
        if key.startswith("actual_"):
            try:
                si_id = int(key.split("_", 1)[1])
                si = sit_map.get(si_id)
                if si and val.strip() != "":
                    si.actual_qty = float(val)
                    si.variance = si.actual_qty - si.system_qty
                elif si and val.strip() == "":
                    si.actual_qty = None
                    si.variance = 0.0
            except (ValueError, IndexError):
                pass
        elif key.startswith("notes_"):
            try:
                si_id = int(key.split("_", 1)[1])
                si = sit_map.get(si_id)
                if si:
                    si.notes = val
            except (ValueError, IndexError):
                pass
    operator = request.session.get("admin_username", "")
    now = datetime.utcnow()
    variance_count = 0
    for si in sit_map.values():
        if si.actual_qty is None:
            continue
        inv_item = db.get(InventoryItem, si.item_id) if si.item_id else None
        if not inv_item:
            continue
        # 以实盘数量直接覆盖当前系统库存（基准是提交时的当前值，不是建单快照）
        # 这样盘点期间正常发生的出入库不会被重复计算
        before = inv_item.stock_qty
        after = si.actual_qty
        diff = after - before          # 与"当前"系统值的差，而非与快照的差
        si.variance = diff             # 更新为真实差异
        if abs(diff) > 0.001:
            variance_count += 1
            inv_item.stock_qty = after
            inv_item.updated_at = now
            db.add(InventoryTransaction(
                item_id=inv_item.id, tx_type="adjust", qty=abs(diff),
                qty_before=before, qty_after=after,
                unit_price=0, ref_type="stocktake", ref_id=session_id,
                operator=operator,
                note=f"循环盘点#{session_id}（{'+' if diff >= 0 else ''}{diff:.1f}）",
            ))
            si.is_adjusted = True
        inv_item.last_counted_at = now
    sess.variance_count = variance_count
    sess.status = "completed"
    sess.completed_at = now
    db.commit()
    return RedirectResponse(f"/admin/stocktake/{session_id}?msg=盘点已提交", status_code=303)


@app.post("/admin/inventory/{item_id}/deactivate")
async def admin_inventory_deactivate(
    item_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404)
    item.is_active = False
    db.commit()
    return RedirectResponse("/admin/inventory?msg=已下架", status_code=303)


# ─────────────────────────────────────────────────────────────────────────────
#  狂犬疫苗免疫登记  Rabies Vaccine Registration
# ─────────────────────────────────────────────────────────────────────────────

_INVALID_NAMES = {"先生", "女士", "小姐", "太太", "夫人", "mr", "mrs", "ms", "主人", "不详"}
_GENERIC_SUFFIXES = ("小姐", "先生", "女士", "太太", "夫人")
_SIG_DIR = Path("data/signatures")
_SIG_DIR.mkdir(parents=True, exist_ok=True)

_RABIES_STATUS_ZH = {
    "owner_pending": "待医护填写",
    "staff_pending": "待完成签字",
    "completed": "已完成",
}


def _is_invalid_name(name: str) -> bool:
    """判断姓名是否为「不规范 / 占位」格式。

    捕获：
      - 完全的占位词：先生、女士、小姐、太太、夫人、mr、不详 …
      - 单姓 + 通用后缀：「高小姐」「刘先生」「李女士」等导入老系统的脏数据
    """
    if not name:
        return True
    n = name.strip().lower().replace(" ", "").replace(".", "")
    if n in _INVALID_NAMES:
        return True
    raw = name.strip()
    # 「X小姐」「X先生」… 长度 ≤ 3 且尾部是通用后缀 → 老系统脏数据
    if len(raw) <= 3:
        for suf in _GENERIC_SUFFIXES:
            if raw.endswith(suf) and len(raw) > len(suf):
                return True
    return False


def _save_signature(data_url: str, prefix: str) -> str:
    """将 base64 data URL 保存为 PNG 文件，返回相对路径。"""
    import base64
    if not data_url or not data_url.startswith("data:image/"):
        return ""
    try:
        header, b64 = data_url.split(",", 1)
        img_bytes = base64.b64decode(b64)
        fname = f"{prefix}_{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}.png"
        fpath = _SIG_DIR / fname
        fpath.write_bytes(img_bytes)
        return str(fpath)
    except Exception:
        return ""


# ── 公开 API：手机号查询客户 ──────────────────────────────────────────────────

@app.get("/api/wechat/my-profile")
async def api_wechat_my_profile(openid: str = Query(""), db: Session = Depends(get_db)):
    """小程序预约页用：按 openid 查是否已绑定档案。已绑→返回客户+宠物列表"""
    openid = (openid or "").strip()
    if not openid:
        return {"bound": False}
    cust = db.query(Customer).filter(Customer.wechat_openid == openid).first()
    if not cust:
        return {"bound": False}
    pets = db.query(Pet).filter(Pet.customer_id == cust.id).all()
    return {
        "bound": True,
        "customer_id": cust.id,
        "name": cust.name or "",
        "phone": cust.phone or "",
        "pets": [
            {
                "id": p.id,
                "name": p.name,
                "breed": p.breed,
                "gender": p.gender,
                "species": p.species,
            }
            for p in pets
        ],
    }


@app.get("/api/customer/lookup")
async def api_customer_lookup(phone: str = Query(""), db: Session = Depends(get_db)):
    if not phone or len(phone) < 6:
        return {"found": False}
    p = phone.strip()
    # 主手机号匹配（可能有多条重客户档案）
    matched = db.query(Customer).filter(Customer.phone == p).all()
    # 主号匹配不到才查备用号 CSV
    if not matched:
        candidates = db.query(Customer).filter(Customer.phones_extra.like(f"%{p}%")).all()
        for c in candidates:
            extras = [x.strip() for x in (c.phones_extra or "").split(",") if x.strip()]
            if p in extras:
                matched.append(c)
    if not matched:
        return {"found": False}
    # 主客户：选名下宠物最多的那条作为代表（用于填客户名/地址）
    cust_with_pets: list[tuple] = []
    all_pets_seen: set[tuple] = set()  # (name_lower, species) 去重
    aggregated_pets: list[dict] = []
    for c in matched:
        pets = db.query(Pet).filter(Pet.customer_id == c.id).all()
        cust_with_pets.append((c, len(pets), pets))
        for pet in pets:
            key = ((pet.name or "").strip().lower(), pet.species or "")
            if key in all_pets_seen:
                continue
            all_pets_seen.add(key)
            aggregated_pets.append({
                "id": pet.id,
                "customer_id": c.id,    # 关键：每只宠物携带自己的 customer_id
                "name": pet.name,
                "breed": pet.breed,
                "gender": pet.gender,
                "birthday_estimate": pet.birthday_estimate,
                "color_pattern": pet.color_pattern,
                "species": pet.species,
            })
    cust_with_pets.sort(key=lambda x: -x[1])
    primary = cust_with_pets[0][0]
    invalid_name = _is_invalid_name(primary.name)
    return {
        "found": True,
        "customer_id": primary.id,
        "name": primary.name or "",
        "name_invalid": invalid_name,
        "address": primary.address,
        "duplicate_count": len(matched) if len(matched) > 1 else 0,
        "pets": aggregated_pets,
    }


# ── 公开表单：主人填写 ────────────────────────────────────────────────────────

@app.get("/rabies", response_class=HTMLResponse)
async def page_rabies_form(request: Request):
    return templates.TemplateResponse(request, "rabies_form.html", {
        "msg": request.query_params.get("msg"),
    })


@app.post("/rabies")
async def submit_rabies_form(request: Request, db: Session = Depends(get_db)):
    form = await request.form()

    owner_name    = str(form.get("owner_name", "")).strip()
    owner_phone   = str(form.get("owner_phone", "")).strip()
    owner_address = str(form.get("owner_address", "")).strip()
    animal_name   = str(form.get("animal_name", "")).strip()
    animal_breed  = str(form.get("animal_breed", "")).strip()
    animal_dob    = str(form.get("animal_dob", "")).strip()
    animal_gender = str(form.get("animal_gender", "")).strip()
    animal_color  = str(form.get("animal_color", "")).strip()
    owner_sig_data = str(form.get("owner_signature", "")).strip()
    customer_id_raw = str(form.get("customer_id", "")).strip()
    pet_id_raw      = str(form.get("pet_id", "")).strip()

    # 校验姓名
    if _is_invalid_name(owner_name) or not owner_name:
        return RedirectResponse("/rabies?msg=请填写真实姓名（不可填写先生/女士）", status_code=303)
    if not owner_phone:
        return RedirectResponse("/rabies?msg=请填写手机号", status_code=303)
    if not owner_sig_data or len(owner_sig_data) < 100:
        return RedirectResponse("/rabies?msg=请完成签名", status_code=303)

    # 保存签名
    sig_path = _save_signature(owner_sig_data, f"owner_{owner_phone}")

    # 提前读出门店（Pet 病历号生成、Customer 归属都会用到）
    clinic_store = str(form.get("clinic_store", "横岗店")).strip() or "横岗店"

    # 查找或创建客户
    customer_id = int(customer_id_raw) if customer_id_raw.isdigit() else None
    if not customer_id:
        cust = db.query(Customer).filter(Customer.phone == owner_phone).first()
        if cust:
            customer_id = cust.id
            # 如果档案名字是无效的，更新为真实姓名
            if _is_invalid_name(cust.name):
                cust.name = owner_name
            if owner_address and not cust.address:
                cust.address = owner_address
        else:
            cust = Customer(name=owner_name, phone=owner_phone, address=owner_address,
                            source="rabies")
            db.add(cust)
            db.flush()
            customer_id = cust.id
    else:
        cust = db.get(Customer, customer_id)
        if cust and _is_invalid_name(cust.name):
            cust.name = owner_name

    # 查找或创建宠物
    # 修复：若传入 pet_id 但其名字与本次提交的 animal_name 不一致，
    # 说明是同一主人的另一只宠物，应按 (customer_id, animal_name) 找已有，再否则新建，
    # 避免把不同动物的狂犬记录全挂在第一只宠物身上
    pet_id = int(pet_id_raw) if pet_id_raw.isdigit() else None
    pet = None
    if pet_id:
        pet = db.get(Pet, pet_id)
        if pet and animal_name and pet.name and pet.name.strip() != animal_name.strip():
            # 名字不一致 → 视为不同动物
            pet = None
            pet_id = None
    if not pet_id and animal_name:
        # 先按 (customer_id, name) 找已有，避免重复
        existing = (
            db.query(Pet)
            .filter(Pet.customer_id == customer_id, Pet.name == animal_name)
            .first()
        )
        if existing:
            pet = existing
            pet_id = existing.id
        else:
            pet = Pet(
                customer_id=customer_id,
                name=animal_name,
                breed=animal_breed,
                gender=animal_gender,
                birthday_estimate=animal_dob,
                color_pattern=animal_color,
                species="dog",
                store=clinic_store,
                medical_record_no=_gen_medical_record_no(db, clinic_store),
            )
            db.add(pet)
            db.flush()
            pet_id = pet.id
    if pet:
        if animal_color and not pet.color_pattern:
            pet.color_pattern = animal_color
        if animal_dob and not pet.birthday_estimate:
            pet.birthday_estimate = animal_dob
        # 老宠物没病历号 → 借这次登记补一个
        if not pet.medical_record_no:
            pet.store = pet.store or clinic_store
            pet.medical_record_no = _gen_medical_record_no(db, pet.store)

    record = RabiesVaccineRecord(
        customer_id=customer_id,
        pet_id=pet_id,
        owner_name=owner_name,
        owner_address=owner_address,
        owner_phone=owner_phone,
        animal_name=animal_name,
        animal_breed=animal_breed,
        animal_dob=animal_dob,
        animal_gender=animal_gender,
        animal_color=animal_color,
        owner_signature_path=sig_path,
        owner_signed_at=datetime.utcnow(),
        status="staff_pending",
        clinic_store=clinic_store,
    )
    db.add(record)
    db.commit()
    try:
        from app.services import wecom_notify as _wn
        _wn.notify_rabies_submitted(db, record)
    except Exception as _e:
        logger.warning("[wecom] notify_rabies_submitted failed: %s", _e)
    return RedirectResponse(f"/rabies/done?id={record.id}", status_code=303)


@app.get("/rabies/done", response_class=HTMLResponse)
async def page_rabies_done(request: Request, id: int = Query(0), db: Session = Depends(get_db)):
    rec = db.get(RabiesVaccineRecord, id) if id else None
    return templates.TemplateResponse(request, "rabies_done.html", {"rec": rec})


@app.post("/api/rabies/submit")
async def api_rabies_submit(request: Request, db: Session = Depends(get_db)):
    body = await request.json()
    owner_name    = str(body.get("owner_name", "")).strip()
    owner_phone   = str(body.get("owner_phone", "")).strip()
    owner_address = str(body.get("owner_address", "")).strip()
    animal_name   = str(body.get("animal_name", "")).strip()
    animal_breed  = str(body.get("animal_breed", "")).strip()
    animal_dob    = str(body.get("animal_dob", "")).strip()
    animal_gender = str(body.get("animal_gender", "")).strip()
    animal_color  = str(body.get("animal_color", "")).strip()
    owner_sig_data  = str(body.get("owner_signature", "")).strip()
    customer_id_raw = body.get("customer_id")
    pet_id_raw      = body.get("pet_id")

    if _is_invalid_name(owner_name) or not owner_name:
        raise HTTPException(400, detail="请填写真实姓名（不可填写先生/女士）")
    if not owner_phone:
        raise HTTPException(400, detail="请填写手机号")
    if not owner_address:
        raise HTTPException(400, detail="请填写联系地址")
    if not animal_name:
        raise HTTPException(400, detail="请填写动物名称")
    if not animal_breed:
        raise HTTPException(400, detail="请填写动物品种")
    if not animal_dob:
        raise HTTPException(400, detail="请选择动物出生年月")
    if not animal_color:
        raise HTTPException(400, detail="请填写动物毛色")
    if not owner_sig_data or len(owner_sig_data) < 100:
        raise HTTPException(400, detail="请完成签名")

    sig_path = _save_signature(owner_sig_data, f"owner_{owner_phone}")

    # 提前读出门店（Pet 病历号生成、Customer 归属都会用到）
    clinic_store = str(body.get("clinic_store", "横岗店")).strip() or "横岗店"

    customer_id = int(customer_id_raw) if isinstance(customer_id_raw, int) or (isinstance(customer_id_raw, str) and customer_id_raw.isdigit()) else None
    if not customer_id:
        cust = db.query(Customer).filter(Customer.phone == owner_phone).first()
        if cust:
            customer_id = cust.id
            if _is_invalid_name(cust.name):
                cust.name = owner_name
            if owner_address and not cust.address:
                cust.address = owner_address
        else:
            cust = Customer(name=owner_name, phone=owner_phone, address=owner_address,
                            source="rabies")
            db.add(cust)
            db.flush()
            customer_id = cust.id
    else:
        cust = db.get(Customer, customer_id)
        if cust and _is_invalid_name(cust.name):
            cust.name = owner_name

    # 修复：若传入 pet_id 但其名字与本次提交的 animal_name 不一致，
    # 视为不同动物（参见 /rabies 表单同样的修复）
    pet_id = int(pet_id_raw) if isinstance(pet_id_raw, int) or (isinstance(pet_id_raw, str) and str(pet_id_raw).isdigit()) else None
    pet = None
    if pet_id:
        pet = db.get(Pet, pet_id)
        if pet and animal_name and pet.name and pet.name.strip() != animal_name.strip():
            pet = None
            pet_id = None
    if not pet_id and animal_name:
        existing = (
            db.query(Pet)
            .filter(Pet.customer_id == customer_id, Pet.name == animal_name)
            .first()
        )
        if existing:
            pet = existing
            pet_id = existing.id
        else:
            pet = Pet(
                customer_id=customer_id,
                name=animal_name,
                breed=animal_breed,
                gender=animal_gender,
                birthday_estimate=animal_dob,
                color_pattern=animal_color,
                species="dog",
                store=clinic_store,
                medical_record_no=_gen_medical_record_no(db, clinic_store),
            )
            db.add(pet)
            db.flush()
            pet_id = pet.id
    if pet:
        if animal_color and not pet.color_pattern:
            pet.color_pattern = animal_color
        if animal_dob and not pet.birthday_estimate:
            pet.birthday_estimate = animal_dob
        # 老宠物没病历号 → 借这次登记补一个
        if not pet.medical_record_no:
            pet.store = pet.store or clinic_store
            pet.medical_record_no = _gen_medical_record_no(db, pet.store)

    record = RabiesVaccineRecord(
        customer_id=customer_id,
        pet_id=pet_id,
        owner_name=owner_name,
        owner_address=owner_address,
        owner_phone=owner_phone,
        animal_name=animal_name,
        animal_breed=animal_breed,
        animal_dob=animal_dob,
        animal_gender=animal_gender,
        animal_color=animal_color,
        owner_signature_path=sig_path,
        owner_signed_at=datetime.utcnow(),
        status="staff_pending",
        clinic_store=clinic_store,
    )
    db.add(record)
    db.commit()
    try:
        from app.services import wecom_notify as _wn
        _wn.notify_rabies_submitted(db, record)
    except Exception as _e:
        logger.warning("[wecom] notify_rabies_submitted failed: %s", _e)
    return {"id": record.id, "status": record.status}


# ── 后台：收费单 ─────────────────────────────────────────────────────────────

_INV_STATUS_ZH = {"unpaid": "待收款", "paid": "已收款", "cancelled": "已取消"}
_INV_PAY_ZH    = {
    "cash": "现金", "wechat": "微信", "alipay": "支付宝",
    "shouqianba": "收钱吧", "meituan": "美团", "third_party": "第三方",
    "wallet": "钱包", "package": "套餐", "deposit": "押金", "coupon": "优惠券",
    "mixed": "混合支付",
    "card": "刷卡", "groupbuy": "团购", "prepaid": "预付款",
    "free": "赠送 / 零结算",
}


def _resolve_invoice_store(db: Session, *, visit_id=None, pet_id=None, customer_id=None, fallback: str = "") -> str:
    """推断发票应归属的门店短名。
    优先级：pet.store → visit.pet.store → fallback (_get_op_store(request))"""
    if pet_id:
        p = db.get(Pet, pet_id)
        if p and p.store:
            return p.store
    if visit_id:
        v = db.get(Visit, visit_id)
        if v and v.pet_id:
            p = db.get(Pet, v.pet_id)
            if p and p.store:
                return p.store
    return fallback or ""


def _gen_invoice_no(db: Session) -> str:
    """生成收费单号：YYYYMMDD-N（当天第几张）"""
    from datetime import date
    today_str = date.today().strftime("%Y%m%d")
    count = db.query(func.count(Invoice.id)).filter(
        Invoice.invoice_no.like(f"{today_str}-%")
    ).scalar() or 0
    return f"{today_str}-{count + 1}"


def _calc_hosp_days(admitted_at, discharged_at) -> int:
    """住院天数：过夜算 1 天。当天进当天出 = 0 天（笼费 0）。
       进 6/2 22:00 → 出 6/3 06:00 = 1 天（过 1 夜）
       进 6/2 06:00 → 出 6/4 22:00 = 2 天（过 2 夜）
    """
    if not admitted_at or not discharged_at:
        return 0
    return max(0, (discharged_at.date() - admitted_at.date()).days)


def _sync_sales_order_invoice(db: Session, order_id: int, admin_name: str = "") -> "Invoice | None":
    """独立销售单（无关联病例）自动生成 / 更新一张「未关联病例」的发票，让收银台能收到款。
    - 已 paid / cancelled / refunded → 不动
    - 已存在 unpaid → 替换明细 + 重算总额
    - 无 → 新建 unpaid 发票
    - 销售单本身被删 → 调用方传 0 表示清空对应发票（外部判断后调用 _delete_so_invoice）
    """
    order = db.get(SalesOrder, order_id)
    if not order:
        return None
    # 已关联病例的销售单不走这里，走 _sync_visit_invoice
    if order.visit_id:
        return None
    if order.status == "cancelled":
        _delete_so_invoice(db, order_id)
        return None

    # 找已有 invoice：通过 InvoiceItem.ref_type=sales_order + ref_id=order_id 反查
    existing_item = (
        db.query(InvoiceItem)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .filter(
            InvoiceItem.ref_type == "sales_order",
            InvoiceItem.ref_id == order_id,
            Invoice.visit_id.is_(None),
        )
        .first()
    )
    inv = db.get(Invoice, existing_item.invoice_id) if existing_item else None

    # 已结清 → 不能改
    if inv and inv.payment_status in ("paid", "refunded", "cancelled"):
        return inv

    # 算明细
    line_items = []
    subtotal_sum = 0.0
    for it in (order.items or []):
        if (it.subtotal or 0) <= 0:
            continue
        line_items.append({
            "ref_type": "sales_order", "ref_id": order_id,
            "description": f"[销售#{order_id}] {it.item_name}",
            "quantity": float(it.quantity or 0),
            "unit_price": float(it.unit_price or 0),
            "subtotal": float(it.subtotal or 0),
        })
        subtotal_sum += float(it.subtotal or 0)

    if not line_items:
        # 销售单变空 → 删旧发票
        _delete_so_invoice(db, order_id)
        return None

    if inv is None:
        from datetime import date as _date
        inv = Invoice(
            customer_id=order.customer_id,
            visit_id=None,
            pet_id=order.pet_id,
            invoice_date=order.order_date or _date.today().isoformat(),
            subtotal=round(subtotal_sum, 2),
            discount_amount=0.0,
            total_amount=round(subtotal_sum, 2),
            payment_status="unpaid",
            notes=f"销售单 #{order_id}",
            store=_resolve_invoice_store(db, pet_id=order.pet_id, customer_id=order.customer_id),
            created_by=admin_name or "system",
        )
        db.add(inv)
        db.flush()
        # 生成 invoice_no
        try:
            inv.invoice_no = _gen_invoice_no(db)
        except Exception:
            pass
    else:
        # 清旧明细
        for old in list(inv.items):
            db.delete(old)
        db.flush()
        inv.subtotal = round(subtotal_sum, 2)
        inv.total_amount = round(subtotal_sum - (inv.discount_amount or 0), 2)
        inv.updated_at = datetime.utcnow()

    for li in line_items:
        db.add(InvoiceItem(invoice_id=inv.id, **li))
    db.flush()
    return inv


def _delete_so_invoice(db: Session, order_id: int) -> None:
    """销售单删除 / 作废时同步删未付的独立发票（已付的保留作为档案）"""
    item = (
        db.query(InvoiceItem)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .filter(
            InvoiceItem.ref_type == "sales_order",
            InvoiceItem.ref_id == order_id,
            Invoice.visit_id.is_(None),
        )
        .first()
    )
    if not item:
        return
    inv = db.get(Invoice, item.invoice_id)
    if not inv:
        return
    if inv.payment_status == "unpaid":
        db.delete(inv)
        db.flush()


def _sync_visit_invoice(db: Session, visit_id: int, admin_name: str = "") -> "Invoice | None":
    """把就诊产生的处方 / 检查单 / 销售单自动同步到一张「待收款」收费单。

    规则：
    - 已 paid / cancelled / refunded 的发票不动（已结清不能改）
    - 已存在 unpaid 的发票 → 替换其所有明细
    - 无 → 创建新 unpaid 发票
    - 若没有任何明细 → 不创建（但已存在的 unpaid 发票若变空，会被清空明细但保留单据，方便后续追加）
    """
    if not visit_id:
        return None
    visit = db.get(Visit, visit_id)
    if not visit:
        return None
    from datetime import date as _date

    # ── 关键：先收集已结清 (paid / refunded / partial) 发票里覆盖的 (ref_type, ref_id) ──
    # 这些条目已经走过收款链路，不能再次进新发票，否则同一笔处方/检查会被重复计费。
    settled_refs: set[tuple[str, int]] = set()
    settled_invs = db.query(Invoice).filter(
        Invoice.visit_id == visit_id,
        Invoice.payment_status.in_(("paid", "partial", "refunded")),
    ).all()
    for s_inv in settled_invs:
        for s_it in (s_inv.items or []):
            if s_it.ref_type and s_it.ref_id:
                settled_refs.add((s_it.ref_type, int(s_it.ref_id)))

    line_items: list[dict] = []
    subtotal_sum = 0.0

    # ── 1) 处方 ──
    prescs = db.query(Prescription).filter(
        Prescription.visit_id == visit_id,
        Prescription.status != "draft",
    ).all()
    for p in prescs:
        if ("prescription", p.id) in settled_refs:
            continue  # 整张处方已在已结清发票里，跳过
        for it in (p.items or []):
            if not it.drug_name or (it.subtotal or 0) <= 0:
                continue
            line_items.append({
                "ref_type": "prescription", "ref_id": p.id,
                "description": f"[处方#{p.id}] {it.drug_name}",
                "quantity": float(it.quantity_num or 0),
                "unit_price": float(it.unit_price or 0),
                "subtotal": float(it.subtotal or 0),
            })
            subtotal_sum += float(it.subtotal or 0)

    # ── 2) 检查单 ──
    exams = db.query(ExamOrder).filter(ExamOrder.visit_id == visit_id).all()
    for eo in exams:
        if ("exam_order", eo.id) in settled_refs:
            continue  # 已结清，跳过
        try:
            eitems = json.loads(eo.items_json or "[]")
        except Exception:
            eitems = []
        for it in eitems:
            name = (it.get("name") or "").strip()
            if not name:
                continue
            qty = float(it.get("qty") or 1)
            price = float(it.get("unit_price") or 0)
            sub = round(qty * price, 2)
            if sub <= 0:
                continue
            line_items.append({
                "ref_type": "exam_order", "ref_id": eo.id,
                "description": f"[检查#{eo.id}] {name}",
                "quantity": qty, "unit_price": price, "subtotal": sub,
            })
            subtotal_sum += sub

    # ── 3) 销售单 ──
    sos = db.query(SalesOrder).filter(
        SalesOrder.visit_id == visit_id,
        SalesOrder.status != "cancelled",
    ).all()
    for so in sos:
        if ("sales_order", so.id) in settled_refs:
            continue
        for it in (so.items or []):
            if (it.subtotal or 0) <= 0:
                continue
            line_items.append({
                "ref_type": "sales_order", "ref_id": so.id,
                "description": f"[销售#{so.id}] {it.item_name}",
                "quantity": float(it.quantity or 0),
                "unit_price": float(it.unit_price or 0),
                "subtotal": float(it.subtotal or 0),
            })
            subtotal_sum += float(it.subtotal or 0)

    # ── 4) 住院笼费（仅 discharged 计） ──
    hosps = db.query(Hospitalization).filter(
        Hospitalization.visit_id == visit_id,
        Hospitalization.status == "discharged",
    ).all()
    for h in hosps:
        if ("hospitalization", h.id) in settled_refs:
            continue
        days = _calc_hosp_days(h.admitted_at, h.discharged_at)
        if days <= 0:
            continue
        rate = float(h.daily_rate_override or 0)
        if rate <= 0 and h.cage_id:
            _c = db.get(Cage, h.cage_id)
            rate = float(_c.daily_rate if _c else 0)
        sub = round(days * rate, 2)
        if sub <= 0:
            continue
        cage_name = ""
        if h.cage_id:
            _c = db.get(Cage, h.cage_id)
            cage_name = _c.code if _c else ""
        line_items.append({
            "ref_type": "hospitalization", "ref_id": h.id,
            "description": f"[住院#{h.id}] 笼费 · {cage_name or '—'} × {days} 天",
            "quantity": float(days),
            "unit_price": rate,
            "subtotal": sub,
        })
        subtotal_sum += sub

    subtotal_sum = round(subtotal_sum, 2)

    # 查找现有未支付发票
    inv = db.query(Invoice).filter(
        Invoice.visit_id == visit_id,
        Invoice.payment_status == "unpaid",
    ).first()

    if not line_items:
        # 没有明细：若已有 unpaid 发票就清空它，否则不创建
        if inv:
            for old in list(inv.items):
                db.delete(old)
            inv.subtotal = 0.0
            inv.total_amount = 0.0
        return inv

    _resolved_store = _resolve_invoice_store(db, visit_id=visit_id, pet_id=visit.pet_id, customer_id=visit.customer_id)
    if inv is None:
        inv = Invoice(
            invoice_no=_gen_invoice_no(db),
            visit_id=visit_id,
            customer_id=visit.customer_id,
            pet_id=visit.pet_id,
            invoice_date=_date.today().isoformat(),
            payment_status="unpaid",
            subtotal=subtotal_sum,
            total_amount=subtotal_sum,
            store=_resolved_store,
            created_by=admin_name or "auto",
        )
        db.add(inv)
        db.flush()
    else:
        # 替换明细：先删旧明细 flush 后再插入新明细，避免冲突
        for old in list(inv.items):
            db.delete(old)
        db.flush()
        inv.subtotal = subtotal_sum
        inv.total_amount = round(subtotal_sum - (inv.discount_amount or 0), 2)
        inv.updated_at = datetime.utcnow()
        if not (inv.store or "") and _resolved_store:
            inv.store = _resolved_store
        # 顺便同步客户/宠物（visit 可能换过宠物或主人）
        if visit.customer_id:
            inv.customer_id = visit.customer_id
        if visit.pet_id:
            inv.pet_id = visit.pet_id

    for li in line_items:
        db.add(InvoiceItem(invoice_id=inv.id, **li))

    return inv


# ---------------------------------------------------------------------------
# 收款统计分析 — 日结 / 月结 / 各种维度分析
# ---------------------------------------------------------------------------

_REVENUE_PAY_ZH = {
    "cash":        "现金",
    "wechat":      "微信",
    "alipay":      "支付宝",
    "shouqianba":  "收钱吧",
    "meituan":     "美团",
    "third_party": "第三方",
    "card":        "刷卡",
    "groupbuy":    "团购",
    "prepaid":     "预付款",
    "wallet":      "钱包",
    "package":     "套餐核销",
    "deposit":     "押金抵扣",
    "coupon":      "优惠券",
    "free":        "赠送 / 零结算",
    "mixed":       "混合支付",
    "other":       "其他",
    "":            "未指定",
}


def _revenue_date_range(preset: str, date_from: str, date_to: str) -> tuple[str, str, str]:
    """根据预设/自定义返回 (from, to, label)。"""
    from datetime import date, timedelta
    today = date.today()
    if preset == "today":
        return today.isoformat(), today.isoformat(), "今日"
    if preset == "yesterday":
        d = today - timedelta(days=1)
        return d.isoformat(), d.isoformat(), "昨日"
    if preset == "week":
        start = today - timedelta(days=today.weekday())
        return start.isoformat(), today.isoformat(), "本周"
    if preset == "month":
        start = today.replace(day=1)
        return start.isoformat(), today.isoformat(), "本月"
    if preset == "last_month":
        first_this = today.replace(day=1)
        last_prev = first_this - timedelta(days=1)
        first_prev = last_prev.replace(day=1)
        return first_prev.isoformat(), last_prev.isoformat(), "上月"
    if preset == "year":
        return today.replace(month=1, day=1).isoformat(), today.isoformat(), "本年"
    # custom
    df = (date_from or "").strip() or today.isoformat()
    dt = (date_to   or "").strip() or today.isoformat()
    return df, dt, f"{df} ~ {dt}"


@app.get("/admin/reports/revenue", response_class=HTMLResponse)
async def admin_reports_revenue(
    request: Request,
    db: Session = Depends(get_db),
    preset: str = Query("month"),
    date_from: str = Query(""),
    date_to: str = Query(""),
    store: str = Query(""),  # 仅 superadmin 可选；员工自动锁本店
):
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    admin_store_short = _get_admin_store(request)
    if admin_store_short:
        store = admin_store_short  # 员工强制本店

    df, dt, label = _revenue_date_range(preset, date_from, date_to)

    # 已收款的收费单（paid_at 在区间内）
    # 员工内购档案不计入业绩报表（is_internal=True 的 customer 全部排除）
    _internal_ids_sub = db.query(Customer.id).filter(Customer.is_internal == True).subquery()
    base_q = db.query(Invoice).filter(
        Invoice.payment_status == "paid",
        Invoice.invoice_date >= df,
        Invoice.invoice_date <= dt,
        ~Invoice.customer_id.in_(_internal_ids_sub),
    )
    rows = base_q.order_by(Invoice.paid_at.desc()).all()

    # 如果有门店筛选：通过 visit.pet.store 关联（invoice 没有 store 字段，但能通过 pet）
    if store:
        from app.models import Pet as _Pet
        pet_store_map = {}
        pet_ids = list({r.pet_id for r in rows if r.pet_id})
        if pet_ids:
            for p in db.query(_Pet).filter(_Pet.id.in_(pet_ids)).all():
                pet_store_map[p.id] = p.store or ""
        rows = [r for r in rows if pet_store_map.get(r.pet_id, "") == store]

    # 汇总
    total_amount = sum(float(r.total_amount or 0) for r in rows)
    total_count = len(rows)
    avg = (total_amount / total_count) if total_count else 0.0

    # 按支付方式：从 Payment 表聚合（混合支付时单张单可拆多笔）
    invoice_ids = [r.id for r in rows]
    pay_rows = []
    if invoice_ids:
        pay_rows = db.query(Payment).filter(
            Payment.invoice_id.in_(invoice_ids),
            Payment.status == "success",
        ).all()
    by_method: dict[str, dict] = {}
    for p in pay_rows:
        m = (p.method or "").strip() or "other"
        if m not in by_method:
            by_method[m] = {"amount": 0.0, "count": 0}
        by_method[m]["amount"] += float(p.amount or 0)
        by_method[m]["count"] += 1
    # 兜底：如果完全没有 Payment 行（老数据），fall back 用 invoice.payment_method
    if not by_method:
        for r in rows:
            m = (r.payment_method or "").strip() or "other"
            if m not in by_method:
                by_method[m] = {"amount": 0.0, "count": 0}
            by_method[m]["amount"] += float(r.total_amount or 0)
            by_method[m]["count"] += 1
    by_method_list = sorted(
        [{"method": m, "label": _REVENUE_PAY_ZH.get(m, m), **v} for m, v in by_method.items()],
        key=lambda x: -x["amount"],
    )

    # 拆分：财务现金类（实收现金/电子钱）vs 业务非现金类（钱包/套餐/押金/券）
    _CASH_METHODS = {"cash", "wechat", "alipay", "shouqianba", "meituan", "third_party"}
    _NONCASH_METHODS = {"wallet", "package", "deposit", "coupon"}
    finance_methods_list = [m for m in by_method_list if m["method"] in _CASH_METHODS]
    business_noncash_list = [m for m in by_method_list if m["method"] in _NONCASH_METHODS]
    finance_cash_total = sum(m["amount"] for m in finance_methods_list)
    business_noncash_total = sum(m["amount"] for m in business_noncash_list)

    # 按收款员（不论现金还是非现金，看人均收款情况）
    by_operator: dict[str, dict] = {}
    for p in pay_rows:
        op = (p.operator or "").strip() or "未指定"
        if op not in by_operator:
            by_operator[op] = {"amount": 0.0, "count": 0}
        by_operator[op]["amount"] += float(p.amount or 0)
        by_operator[op]["count"] += 1
    by_operator_list = sorted(
        [{"operator": o, **v} for o, v in by_operator.items()],
        key=lambda x: -x["amount"],
    )

    # 按门店 × 支付方式 二维拆分（superadmin 一眼看两店对比）
    # 用 Payment.store 字段；fallback 用 pet 推断 store
    store_x_method: dict[str, dict[str, float]] = {}
    store_x_method_meta: dict[str, dict] = {}  # 每家店的合计 + 笔数
    for p in pay_rows:
        s = (p.store or "").strip() or "未指定"
        m = (p.method or "").strip() or "other"
        if s not in store_x_method:
            store_x_method[s] = {}
            store_x_method_meta[s] = {"cash_total": 0.0, "noncash_total": 0.0, "count": 0}
        store_x_method[s][m] = store_x_method[s].get(m, 0.0) + float(p.amount or 0)
        store_x_method_meta[s]["count"] += 1
        if m in _CASH_METHODS:
            store_x_method_meta[s]["cash_total"] += float(p.amount or 0)
        elif m in _NONCASH_METHODS:
            store_x_method_meta[s]["noncash_total"] += float(p.amount or 0)
    # 排序：按总金额倒序
    by_store_method_list = []
    for s in sorted(store_x_method.keys(), key=lambda x: -(store_x_method_meta[x]["cash_total"] + store_x_method_meta[x]["noncash_total"])):
        row = {
            "store": s,
            "count": store_x_method_meta[s]["count"],
            "cash_total": store_x_method_meta[s]["cash_total"],
            "noncash_total": store_x_method_meta[s]["noncash_total"],
            "total": store_x_method_meta[s]["cash_total"] + store_x_method_meta[s]["noncash_total"],
            "methods": store_x_method[s],
        }
        by_store_method_list.append(row)

    # 钱包消费"按比例扣的本金 vs 赠送"—— 用 WalletTransaction 上记的精确值
    wallet_consume_total = sum(m["amount"] for m in by_method_list if m["method"] == "wallet")
    wallet_consume_principal_est = 0.0
    wallet_consume_bonus_est = 0.0
    try:
        _wt_consume_q = db.query(
            func.coalesce(func.sum(WalletTransaction.consumed_principal), 0),
            func.coalesce(func.sum(WalletTransaction.consumed_bonus), 0),
        ).filter(
            WalletTransaction.type == "consume",
            WalletTransaction.created_at >= df + " 00:00:00",
            WalletTransaction.created_at <= dt + " 23:59:59",
        )
        if store:
            _wt_consume_q = _wt_consume_q.filter(WalletTransaction.store == store)
        wp, wb = _wt_consume_q.one()
        wallet_consume_principal_est = float(wp or 0)
        wallet_consume_bonus_est = float(wb or 0)
    except Exception:
        pass

    # 按门店（仅 superadmin 看；用 pet.store 推断）
    by_store_list: list = []
    if not admin_store_short:
        from app.models import Pet as _Pet
        pet_ids2 = list({r.pet_id for r in rows if r.pet_id})
        psmap = {}
        if pet_ids2:
            for p in db.query(_Pet).filter(_Pet.id.in_(pet_ids2)).all():
                psmap[p.id] = p.store or "未指定"
        by_store: dict[str, dict] = {}
        for r in rows:
            s = psmap.get(r.pet_id, "未指定") or "未指定"
            if s not in by_store:
                by_store[s] = {"amount": 0.0, "count": 0}
            by_store[s]["amount"] += float(r.total_amount or 0)
            by_store[s]["count"] += 1
        by_store_list = sorted(
            [{"store": s, **v} for s, v in by_store.items()],
            key=lambda x: -x["amount"],
        )

    # 日趋势（区间内每日）
    from datetime import date as _date, timedelta as _td
    def _parse(s):
        try:
            y, m, d = s.split("-")
            return _date(int(y), int(m), int(d))
        except Exception:
            return None
    d0 = _parse(df) or _date.today()
    d1 = _parse(dt) or _date.today()
    if (d1 - d0).days < 0:
        d0, d1 = d1, d0
    daily_amount: dict[str, float] = {}
    cur = d0
    while cur <= d1:
        daily_amount[cur.isoformat()] = 0.0
        cur += _td(days=1)
    for r in rows:
        k = (r.invoice_date or "")[:10]
        if k in daily_amount:
            daily_amount[k] += float(r.total_amount or 0)
    daily_series = [{"date": k, "amount": round(v, 2)} for k, v in sorted(daily_amount.items())]

    # 按收费来源类型（处方/检查/手术/其他）—— 通过 invoice.notes / visit 关联推断
    # 简化：用 invoice_no 前缀 / notes 关键词 / visit_id 关联推断
    by_category: dict[str, float] = {"处方": 0.0, "检查单": 0.0, "销售单": 0.0, "其他": 0.0}
    for r in rows:
        n = (r.notes or "") + (r.invoice_no or "")
        if "处方" in n or "Rx" in n.upper() or "PRESC" in n.upper():
            by_category["处方"] += float(r.total_amount or 0)
        elif "检查" in n or "EXAM" in n.upper():
            by_category["检查单"] += float(r.total_amount or 0)
        elif "销售" in n or "SO" in n.upper():
            by_category["销售单"] += float(r.total_amount or 0)
        else:
            by_category["其他"] += float(r.total_amount or 0)
    by_category_list = [{"label": k, "amount": round(v, 2)} for k, v in by_category.items() if v > 0]

    # 钱包充值（区间内）
    wallet_recharge_q = db.query(WalletTransaction).filter(
        WalletTransaction.type == "recharge",
        WalletTransaction.created_at >= df + " 00:00:00",
        WalletTransaction.created_at <= dt + " 23:59:59",
    )
    if store:
        wallet_recharge_q = wallet_recharge_q.filter(WalletTransaction.store == store)
    wallet_recharges = wallet_recharge_q.all()
    wallet_recharge_total = sum(float(t.amount or 0) for t in wallet_recharges)

    # 套餐售卖（区间内）
    pkg_sold_q = db.query(CustomerPackage).filter(
        CustomerPackage.purchase_date >= df,
        CustomerPackage.purchase_date <= dt,
    )
    if store:
        pkg_sold_q = pkg_sold_q.filter(CustomerPackage.store == store)
    pkg_sold = pkg_sold_q.all()
    pkg_sold_total = sum(float(p.sell_price or 0) for p in pkg_sold)

    # 押金净流入（收 - 退）
    dep_q = db.query(Deposit).filter(
        Deposit.created_at >= df + " 00:00:00",
        Deposit.created_at <= dt + " 23:59:59",
    )
    if store:
        dep_q = dep_q.filter(Deposit.store == store)
    deps = dep_q.all()
    deposit_in = sum(float(d.amount or 0) for d in deps)
    deposit_refund = sum(float(d.refunded_amount or 0) for d in deps if d.refunded_at and df <= d.refunded_at.strftime("%Y-%m-%d") <= dt)

    return templates.TemplateResponse(request, "uk/reports_revenue.html", {
        "label": label,
        "df": df, "dt": dt,
        "preset": preset,
        "store": store,
        "is_superadmin": (request.session.get("admin_role") == "superadmin"),
        "store_options": _STORE_OPTIONS if not admin_store_short else [admin_store_short],
        "admin_store_short": admin_store_short,
        # 汇总
        "total_amount": total_amount,
        "total_count": total_count,
        "avg": avg,
        # 分类
        "by_method_list": by_method_list,
        "by_store_list": by_store_list,
        "by_category_list": by_category_list,
        "daily_series": daily_series,
        # 财务 vs 业务 拆分
        "finance_methods_list": finance_methods_list,        # 现金类（含微信/支付宝/收钱吧/美团/第三方/现金）
        "finance_cash_total": finance_cash_total,             # 现金类总额
        "business_noncash_list": business_noncash_list,       # 钱包/套餐/押金/券
        "business_noncash_total": business_noncash_total,
        # 钱包消费拆解
        "wallet_consume_total": wallet_consume_total,
        "wallet_consume_principal_est": wallet_consume_principal_est,
        "wallet_consume_bonus_est": wallet_consume_bonus_est,
        # 按收款员
        "by_operator_list": by_operator_list,
        # 按门店 × 支付方式
        "by_store_method_list": by_store_method_list,
        # 其他财务流入
        "wallet_recharge_total": wallet_recharge_total,
        "wallet_recharges_count": len(wallet_recharges),
        "pkg_sold_total": pkg_sold_total,
        "pkg_sold_count": len(pkg_sold),
        "deposit_in": deposit_in,
        "deposit_refund": deposit_refund,
        # 财务收入合计 / 业务收入合计
        "finance_total": finance_cash_total + wallet_recharge_total,
        "business_total": total_amount,  # = 已收款收费单总额 = 现金类+钱包+套餐+押金+券
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/reports/revenue/export")
async def admin_reports_revenue_export(
    request: Request,
    db: Session = Depends(get_db),
    preset: str = Query("month"),
    date_from: str = Query(""),
    date_to: str = Query(""),
    store: str = Query(""),
):
    """导出 Excel：收费单明细 + 汇总。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    admin_store_short = _get_admin_store(request)
    if admin_store_short:
        store = admin_store_short
    df, dt, label = _revenue_date_range(preset, date_from, date_to)
    # 员工内购档案不计入业绩报表
    _internal_ids_sub = db.query(Customer.id).filter(Customer.is_internal == True).subquery()
    rows = db.query(Invoice).filter(
        Invoice.payment_status == "paid",
        Invoice.invoice_date >= df,
        Invoice.invoice_date <= dt,
        ~Invoice.customer_id.in_(_internal_ids_sub),
    ).order_by(Invoice.paid_at.asc()).all()

    if store:
        from app.models import Pet as _Pet
        ps = {p.id: (p.store or "") for p in db.query(_Pet).filter(_Pet.id.in_({r.pet_id for r in rows if r.pet_id})).all()}
        rows = [r for r in rows if ps.get(r.pet_id, "") == store]

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "收款明细"
    headers = ["收款时间", "单号", "客户", "宠物ID", "金额", "支付方式", "备注"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for r in rows:
        cust = db.get(Customer, r.customer_id) if r.customer_id else None
        ws.append([
            r.paid_at.strftime("%Y-%m-%d %H:%M") if r.paid_at else "",
            r.invoice_no or "",
            (cust.name if cust else ""),
            r.pet_id or "",
            float(r.total_amount or 0),
            _REVENUE_PAY_ZH.get(r.payment_method or "", r.payment_method or ""),
            (r.notes or "")[:200],
        ])
    # 汇总
    ws2 = wb.create_sheet("按支付方式")
    ws2.append(["支付方式", "笔数", "金额"])
    for c in range(1, 4):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    by_m: dict[str, dict] = {}
    for r in rows:
        m = r.payment_method or "未指定"
        d = by_m.setdefault(m, {"count": 0, "amount": 0.0})
        d["count"] += 1
        d["amount"] += float(r.total_amount or 0)
    for m, d in sorted(by_m.items(), key=lambda x: -x[1]["amount"]):
        ws2.append([_REVENUE_PAY_ZH.get(m, m), d["count"], round(d["amount"], 2)])

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    from urllib.parse import quote
    fname = quote(f"收款明细_{label.replace(' ', '_').replace('~','-')}.xlsx")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@app.get("/admin/invoices", response_class=HTMLResponse)
async def admin_invoices_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    status: str = "",
    store: str = "",
):
    require_admin(request)
    # 门店过滤：staff 自动锁本店；superadmin 通过 ?store= 切换（"" = 全部）
    admin_store = _get_admin_store(request)
    if admin_store:
        wb_store = admin_store
    else:
        wb_store = (store or "").strip()
    is_super = (request.session.get("admin_role") == "superadmin")

    query = db.query(Invoice).order_by(Invoice.id.desc())
    if wb_store:
        query = query.filter(Invoice.store == wb_store)
    if status:
        # 「待收款」tab 同时显示 unpaid + partial（部分收过的也需要继续收）
        if status == "unpaid":
            query = query.filter(Invoice.payment_status.in_(("unpaid", "partial")))
        else:
            query = query.filter(Invoice.payment_status == status)
    if q:
        from sqlalchemy import or_
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(Invoice.customer_id.in_(cids))
    invoices = query.limit(200).all()
    cust_map = {}
    for inv in invoices:
        if inv.customer_id and inv.customer_id not in cust_map:
            cust_map[inv.customer_id] = db.get(Customer, inv.customer_id)
    # 统计数据（同样按 wb_store 过滤；员工内购档案排除）
    from datetime import date as _date
    today_str = _date.today().isoformat()
    _internal_ids_sub = db.query(Customer.id).filter(Customer.is_internal == True).subquery()
    def _stat_q():
        q2 = db.query(Invoice).filter(~Invoice.customer_id.in_(_internal_ids_sub))
        if wb_store:
            q2 = q2.filter(Invoice.store == wb_store)
        return q2
    today_paid_sum = _stat_q().filter(
        Invoice.payment_status == "paid",
        Invoice.invoice_date == today_str,
    ).all()
    unpaid_all = _stat_q().filter(Invoice.payment_status.in_(("unpaid", "partial"))).all()
    inv_stats = {
        "today_paid_total": round(sum((i.total_amount or 0) for i in today_paid_sum), 2),
        "today_paid_count": len(today_paid_sum),
        "unpaid_count": len(unpaid_all),
        "unpaid_total": round(sum((i.total_amount or 0) for i in unpaid_all), 2),
    }
    # 同客户多张未付发票聚合（用于顶部合并结算入口）
    multi_pay_groups = []
    if status in ("", "unpaid"):
        _unpaid_q = db.query(Invoice).filter(
            Invoice.payment_status.in_(("unpaid", "partial")),
            Invoice.customer_id.is_not(None),
        )
        if wb_store:
            _unpaid_q = _unpaid_q.filter(Invoice.store == wb_store)
        _by_cust: dict[int, list] = {}
        for u in _unpaid_q.all():
            _by_cust.setdefault(u.customer_id, []).append(u)
        for cid, lst in _by_cust.items():
            if len(lst) >= 2:
                c = db.get(Customer, cid)
                multi_pay_groups.append({
                    "customer": c,
                    "count": len(lst),
                    "total": round(sum(float(i.total_amount or 0) for i in lst), 2),
                })
        multi_pay_groups.sort(key=lambda g: -g["total"])
    return templates.TemplateResponse(request, "uk/invoices.html", {
        "invoices": invoices,
        "cust_map": cust_map,
        "inv_status_zh": _INV_STATUS_ZH,
        "inv_pay_zh": _INV_PAY_ZH,
        "inv_stats": inv_stats,
        "multi_pay_groups": multi_pay_groups,
        "q": q,
        "status": status,
        "wb_store": wb_store,
        "is_super": is_super,
        "stores": ["东环店", "横岗店"],
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/cashier/multi-pay", response_class=HTMLResponse)
async def admin_cashier_multi_pay_page(
    customer_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """同客户多张未付发票合并结算页。"""
    require_admin(request)
    cust = db.get(Customer, customer_id)
    if not cust:
        raise HTTPException(404, "客户不存在")
    admin_store = _get_admin_store(request)
    q = db.query(Invoice).filter(
        Invoice.customer_id == customer_id,
        Invoice.payment_status == "unpaid",
    )
    if admin_store:
        q = q.filter(Invoice.store == admin_store)
    invoices = q.order_by(Invoice.id.asc()).all()
    # 计算每张 outstanding（=total - 已付）
    rows = []
    grand_total = 0.0
    for inv in invoices:
        paid = _invoice_paid_sum(db, inv.id)
        outstanding = max(0.0, float(inv.total_amount or 0) - paid)
        if outstanding <= 0:
            continue
        rows.append({"inv": inv, "outstanding": outstanding, "pet": inv.pet})
        grand_total += outstanding
    return templates.TemplateResponse(request, "uk/cashier_multi_pay.html", {
        "cust": cust,
        "rows": rows,
        "grand_total": round(grand_total, 2),
        "wb_store": admin_store or "",
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/cashier/multi-pay")
async def admin_cashier_multi_pay_submit(
    request: Request,
    db: Session = Depends(get_db),
):
    """提交合并结算：对选中的每张发票按 outstanding 创建 Payment + 自动结清。"""
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    customer_id = int(form.get("customer_id") or 0)
    if not customer_id:
        raise HTTPException(400, "customer_id 必填")
    method = str(form.get("method") or "cash").strip()
    ref_no = (form.get("ref_no") or "").strip()[:120]
    operator = request.session.get("admin_username", "admin")
    store = _get_admin_store(request)
    # 仅支持简单支付方式（钱包/套餐/押金/优惠券需精确调度，请单笔结算）
    simple_methods = {"cash", "wechat", "alipay", "shouqianba", "meituan", "third_party"}
    if method not in simple_methods:
        return RedirectResponse(
            f"/admin/cashier/multi-pay?customer_id={customer_id}&msg=合并结算暂只支持现金/微信/支付宝/收钱吧/美团/第三方",
            status_code=303,
        )
    inv_ids = [int(x) for x in form.getlist("invoice_ids") if str(x).isdigit()]
    if not inv_ids:
        return RedirectResponse(
            f"/admin/cashier/multi-pay?customer_id={customer_id}&msg=请勾选至少一张待收单",
            status_code=303,
        )
    # 折扣 / 减免（可选）— 按选中单 outstanding 比例分摊
    disc_mode = (form.get("discount_mode") or "none").strip()  # none / pct / amount
    try:
        disc_value = float((form.get("discount_value") or "0").strip() or 0)
    except ValueError:
        disc_value = 0.0
    if disc_mode == "pct":
        if disc_value > 1.0:
            disc_value = disc_value / 100.0  # 80 → 0.8
        if disc_value <= 0 or disc_value > 1.0:
            disc_mode = "none"; disc_value = 0.0
    elif disc_mode == "amount":
        if disc_value <= 0:
            disc_mode = "none"; disc_value = 0.0
    else:
        disc_mode = "none"; disc_value = 0.0

    # 第一遍：算出每张 outstanding，过滤掉已付清
    inv_rows = []
    for iid in inv_ids:
        inv = db.get(Invoice, iid)
        if not inv or inv.customer_id != customer_id or inv.payment_status == "paid":
            continue
        outstanding = max(0.0, float(inv.total_amount or 0) - _invoice_paid_sum(db, inv.id))
        if outstanding <= 0:
            continue
        inv_rows.append({"inv": inv, "out": outstanding})
    if not inv_rows:
        return RedirectResponse(
            f"/admin/cashier/multi-pay?customer_id={customer_id}&msg=勾选的单都无未收金额",
            status_code=303,
        )
    grand_out = sum(r["out"] for r in inv_rows)
    if disc_mode == "pct":
        total_discount = round(grand_out * (1.0 - disc_value), 2)
    elif disc_mode == "amount":
        total_discount = round(min(disc_value, grand_out), 2)
    else:
        total_discount = 0.0

    # 折扣只在第一笔混合支付时一次性应用（避免每轮重复加减免流水）
    # 通过 form.discount_applied 标记：已经应用过的不再写减免
    discount_already_applied = (form.get("discount_applied") == "1")
    if discount_already_applied:
        total_discount = 0.0

    # 应收（折扣分摊后） = out - share_per_inv
    inv_actuals: list[float] = []
    discount_share_used = 0.0
    for idx, r in enumerate(inv_rows):
        out = r["out"]
        if total_discount > 0:
            if idx < len(inv_rows) - 1:
                share = round(total_discount * (out / grand_out), 2)
            else:
                share = round(total_discount - discount_share_used, 2)
            share = max(0.0, min(share, out))
        else:
            share = 0.0
        inv_actuals.append(round(out - share, 2))
        if share > 0:
            # ★ 折扣直接折进发票本体：discount_amount += share, total_amount -= share
            # 这样三联（应付/已收/未收）反映真实情况，不再写 Payment(method=free) 误导
            inv_obj = r["inv"]
            inv_obj.discount_amount = round(float(inv_obj.discount_amount or 0) + share, 2)
            inv_obj.total_amount = round(float(inv_obj.total_amount or 0) - share, 2)
            inv_obj.notes = (inv_obj.notes or "") + f"\n[合并结算减免 {datetime.now().strftime('%Y-%m-%d %H:%M')}] 分摊 ¥{share:.2f}（共 {len(inv_rows)} 单）"
            db.flush()
            discount_share_used += share
    actuals_total = round(sum(inv_actuals), 2)

    # 本轮要消化的金额（混合支付时可指定本笔金额，默认 = 全部应收）
    try:
        round_amount = float((form.get("round_amount") or "").strip() or actuals_total)
    except ValueError:
        round_amount = actuals_total
    round_amount = max(0.0, min(round_amount, actuals_total))
    if round_amount <= 0:
        return RedirectResponse(
            f"/admin/cashier/multi-pay?customer_id={customer_id}&msg=本笔金额为 0",
            status_code=303,
        )

    # 按顺序消化：每张单先满足，剩余转下一张
    paid_count = 0
    paid_total = 0.0
    note_tag = f"合并结算（{len(inv_rows)} 单 · {method}）"
    remaining = round_amount
    for idx, r in enumerate(inv_rows):
        if remaining <= 0:
            break
        inv = r["inv"]
        actual_due = inv_actuals[idx]
        # 现在的 outstanding（考虑减免已写流水后）
        cur_out = max(0.0, actual_due)
        if cur_out <= 0:
            continue
        take = min(remaining, cur_out)
        # store 优先用发票本身的（保证收款统计「按门店」准确）
        _pay_store = store or (inv.store or "")
        db.add(Payment(
            invoice_id=inv.id, customer_id=customer_id,
            method=method, amount=round(take, 2),
            ref_no=ref_no, status="success",
            store=_pay_store, operator=operator,
            note=note_tag,
        ))
        db.flush()
        _invoice_recompute_status(db, inv)
        paid_count += 1
        paid_total += take
        remaining = round(remaining - take, 2)
    db.commit()

    # 剩余未付（含其他单的）
    still_outstanding = round(actuals_total - round_amount, 2)
    from urllib.parse import quote as _q
    if still_outstanding > 0.005:
        # 还有未付 → 回合并结算页让用户走下一种支付方式
        msg_text = f"本笔 {method} 收 ¥{paid_total:.2f}（{paid_count} 单部分/已收）· 剩余 ¥{still_outstanding:.2f} 请继续选支付方式"
        return RedirectResponse(
            f"/admin/cashier/multi-pay?customer_id={customer_id}&msg={_q(msg_text, safe='')}",
            status_code=303,
        )
    # 全部付清 → 回已收款 tab
    if total_discount > 0:
        msg = f"已合并结算 {len(inv_rows)} 单 · 应收 ¥{grand_out:.2f} − 减免 ¥{total_discount:.2f} = 实收 ¥{actuals_total:.2f}"
    else:
        msg = f"已合并结算 {len(inv_rows)} 单 · ¥{actuals_total:.2f}"
    return RedirectResponse(
        f"/admin/invoices?status=paid&msg={_q(msg, safe='')}",
        status_code=303,
    )


@app.get("/admin/invoices/create", response_class=HTMLResponse)
async def admin_invoice_create_page(
    request: Request,
    db: Session = Depends(get_db),
    visit_id: int = 0,
    customer_id: int = 0,
):
    require_admin(request)
    from datetime import date
    visit, cust, pet = None, None, None
    prefill_items = []

    if visit_id:
        visit = db.get(Visit, visit_id)
        if visit:
            if visit.customer_id:
                cust = db.get(Customer, visit.customer_id)
            if visit.pet_id:
                pet = db.get(Pet, visit.pet_id)
            # 预填：处方单
            prescs = db.query(Prescription).filter(
                Prescription.visit_id == visit_id,
                Prescription.status != "draft",
            ).all()
            for p in prescs:
                for item in p.items:
                    prefill_items.append({
                        "ref_type": "prescription",
                        "ref_id": p.id,
                        "description": f"[处方#{p.id}] {item.drug_name}",
                        "quantity": item.quantity_num,
                        "unit_price": item.unit_price,
                        "subtotal": item.subtotal,
                    })
            # 预填：销售单
            sos = db.query(SalesOrder).filter(
                SalesOrder.visit_id == visit_id,
                SalesOrder.status != "cancelled",
            ).all()
            for so in sos:
                for item in so.items:
                    prefill_items.append({
                        "ref_type": "sales_order",
                        "ref_id": so.id,
                        "description": f"[销售单#{so.id}] {item.item_name}",
                        "quantity": item.quantity,
                        "unit_price": item.unit_price,
                        "subtotal": item.subtotal,
                    })
    elif customer_id:
        cust = db.get(Customer, customer_id)

    return templates.TemplateResponse(request, "uk/invoice_detail.html", {
        "mode": "create",
        "visit": visit,
        "cust": cust,
        "pet": pet,
        "prefill_items": prefill_items,
        "today": date.today().isoformat(),
        "inv_status_zh": _INV_STATUS_ZH,
        "inv_pay_zh": _INV_PAY_ZH,
        "csrf_token": _get_csrf_token(request),
        "msg": None,
    })


@app.post("/admin/invoices/create")
async def admin_invoice_create(
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    from datetime import date

    visit_id    = int(form.get("visit_id") or 0) or None
    customer_id = int(form.get("customer_id") or 0) or None
    pet_id      = int(form.get("pet_id") or 0) or None
    invoice_date = str(form.get("invoice_date") or date.today().isoformat())
    discount     = float(form.get("discount_amount") or 0)
    notes        = str(form.get("notes") or "")
    admin_name   = request.session.get("admin_username", "")

    # 明细行
    descs      = form.getlist("desc[]")
    qtys       = form.getlist("qty[]")
    prices     = form.getlist("price[]")
    ref_types  = form.getlist("ref_type[]")
    ref_ids    = form.getlist("ref_id[]")

    subtotal = 0.0
    line_items = []
    for i, desc in enumerate(descs):
        desc = desc.strip()
        if not desc:
            continue
        qty   = float(qtys[i]) if i < len(qtys) else 1.0
        price = float(prices[i]) if i < len(prices) else 0.0
        sub   = round(qty * price, 2)
        subtotal += sub
        line_items.append(InvoiceItem(
            ref_type    = ref_types[i] if i < len(ref_types) else "manual",
            ref_id      = int(ref_ids[i]) if i < len(ref_ids) and ref_ids[i] else None,
            description = desc,
            quantity    = qty,
            unit_price  = price,
            subtotal    = sub,
        ))

    total = round(subtotal - discount, 2)
    inv = Invoice(
        invoice_no      = _gen_invoice_no(db),
        customer_id     = customer_id,
        visit_id        = visit_id,
        pet_id          = pet_id,
        invoice_date    = invoice_date,
        subtotal        = round(subtotal, 2),
        discount_amount = discount,
        total_amount    = total,
        payment_status  = "unpaid",
        notes           = notes,
        store           = _resolve_invoice_store(db, visit_id=visit_id, pet_id=pet_id, customer_id=customer_id, fallback=_get_op_store(request)),
        created_by      = admin_name,
    )
    db.add(inv)
    db.flush()
    for li in line_items:
        li.invoice_id = inv.id
        db.add(li)
    db.commit()
    return RedirectResponse(f"/admin/invoices/{inv.id}?msg=收费单已创建", status_code=303)


@app.get("/admin/invoices/{inv_id}", response_class=HTMLResponse)
async def admin_invoice_detail(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin(request)
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404, "收费单不存在")
    cust  = db.get(Customer, inv.customer_id) if inv.customer_id else None
    pet   = db.get(Pet,      inv.pet_id)      if inv.pet_id      else None
    visit = db.get(Visit,    inv.visit_id)    if inv.visit_id    else None
    # 该客户可用的钱包余额 + 有效套餐（用于结算时选择）
    wallet_balance = 0.0
    active_packages = []
    if inv.customer_id:
        w = db.query(Wallet).filter(Wallet.customer_id == inv.customer_id).first()
        wallet_balance = float(w.balance) if w else 0.0
        active_packages = (
            db.query(CustomerPackage)
            .filter(
                CustomerPackage.customer_id == inv.customer_id,
                CustomerPackage.status == "active",
            )
            .order_by(CustomerPackage.id.desc())
            .all()
        )
    # 该收费单已用的钱包/套餐流水
    paid_wallet_txs = (
        db.query(WalletTransaction)
        .filter(WalletTransaction.invoice_id == inv_id)
        .order_by(WalletTransaction.id.desc())
        .all()
    )
    paid_redeems = (
        db.query(PackageRedemption)
        .filter(PackageRedemption.invoice_id == inv_id)
        .order_by(PackageRedemption.id.desc())
        .all()
    )
    # 客户有未结清的押金（held / partial_refund 且还有剩余）→ 可抵扣
    available_deposits = []
    if inv.customer_id:
        for d in db.query(Deposit).filter(
            Deposit.customer_id == inv.customer_id,
            Deposit.status.in_(["held", "partial_refund"]),
        ).order_by(Deposit.id.desc()).all():
            remaining = d.amount - (d.applied_amount or 0) - (d.refunded_amount or 0)
            if remaining > 0:
                d._remaining = remaining
                available_deposits.append(d)

    # 客户可用的优惠券
    available_coupons = []
    if inv.customer_id:
        for c in db.query(Coupon).filter(
            ((Coupon.customer_id == inv.customer_id) | (Coupon.customer_id.is_(None))),
            Coupon.status == "issued",
        ).order_by(Coupon.id.desc()).all():
            if _coupon_is_expired(c):
                continue
            usable = _coupon_compute_amount(c, float(inv.total_amount or 0))
            if usable > 0:
                c._usable_amount = usable
                available_coupons.append(c)

    # 已加的 Payment 流水（含已撤销）
    payments = (
        db.query(Payment)
        .filter(Payment.invoice_id == inv_id)
        .order_by(Payment.id.desc())
        .all()
    )
    paid_sum = sum(float(p.amount or 0) for p in payments if p.status == "success")
    outstanding = max(0.0, float(inv.total_amount or 0) - paid_sum)
    return templates.TemplateResponse(request, "uk/invoice.html", {  # B8 UK 重写（view mode）
        "mode": "view",
        "inv": inv,
        "cust": cust,
        "pet": pet,
        "visit": visit,
        "inv_status_zh": _INV_STATUS_ZH,
        "inv_pay_zh": _INV_PAY_ZH,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
        "wallet_balance": wallet_balance,
        "active_packages": active_packages,
        "paid_wallet_txs": paid_wallet_txs,
        "paid_redeems": paid_redeems,
        "available_deposits": available_deposits,
        "deposit_category_zh": _DEPOSIT_CATEGORY_ZH,
        "available_coupons": available_coupons,
        "coupon_kind_zh": _COUPON_KIND_ZH,
        "payments": payments,
        "paid_sum": paid_sum,
        "outstanding": outstanding,
        "method_zh": _REVENUE_PAY_ZH,
        # 同客户其他待收单（用于本页直接勾选合并结算）
        "other_unpaid": _other_unpaid_for_invoice(db, inv) if inv.customer_id else [],
    })


def _other_unpaid_for_invoice(db: Session, current_inv: Invoice) -> list:
    """同客户其他未结清的发票（unpaid + partial），按 id 升序。返回 [{inv, outstanding, items_preview}]"""
    rows = db.query(Invoice).filter(
        Invoice.customer_id == current_inv.customer_id,
        Invoice.id != current_inv.id,
        Invoice.payment_status.in_(("unpaid", "partial")),
    ).order_by(Invoice.id.asc()).all()
    out = []
    for r in rows:
        paid = _invoice_paid_sum(db, r.id)
        outstanding = max(0.0, float(r.total_amount or 0) - paid)
        if outstanding <= 0:
            continue
        # 取明细描述前 60 字符做预览
        items_preview = " · ".join(
            (i.description or "")[:24] for i in (r.items or [])[:3]
        )
        if len(r.items or []) > 3:
            items_preview += f" 等 {len(r.items)} 项"
        out.append({"inv": r, "outstanding": round(outstanding, 2),
                    "items_preview": items_preview, "pet": r.pet})
    return out


def _invoice_paid_sum(db: Session, inv_id: int) -> float:
    """已收款金额合计（仅 success 状态）。"""
    rows = db.query(Payment).filter(
        Payment.invoice_id == inv_id,
        Payment.status == "success",
    ).all()
    return sum(float(r.amount or 0) for r in rows)


def _invoice_recompute_status(db: Session, inv: Invoice) -> None:
    """根据 Payments 合计自动调整 invoice 状态。"""
    paid = _invoice_paid_sum(db, inv.id)
    total = float(inv.total_amount or 0)
    if paid >= total - 1e-6 and total > 0:
        inv.payment_status = "paid"
        if not inv.paid_at:
            inv.paid_at = datetime.utcnow()
        # payment_method = 笔数最多的那种
        from collections import Counter
        methods = [p.method for p in db.query(Payment).filter(Payment.invoice_id == inv.id, Payment.status == "success").all()]
        if methods:
            inv.payment_method = Counter(methods).most_common(1)[0][0]
    else:
        inv.payment_status = "unpaid"
        inv.paid_at = None


@app.post("/admin/invoices/{inv_id}/add-payment")
async def admin_invoice_add_payment(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """添加一笔收款（混合支付：可重复调用直到 sum >= total）。
    支持「同客户多张待收单一起付」— 前端勾选其他单 id 通过 invoice_ids[] 传入：
      - 金额按勾选顺序 (current invoice 第一，其余按 id 升序) 分摊
      - 现金/微信/支付宝/钱包等 simple 方式可跨多单
      - 套餐/押金/优惠券 只支持单张 (current invoice)，因为绑特定明细行
    """
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    if inv.payment_status == "paid":
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已收款，请勿重复", status_code=303)

    method = str(form.get("method") or "cash").strip()
    operator = request.session.get("admin_username", "admin")
    # store 优先取发票本身的 store（更准确，超管未挂店时 _get_admin_store=""），
    # 这样收款统计「按门店」就不会出现「未指定」
    store = _get_admin_store(request) or (inv.store or "")

    # ── 合并结算：解析勾选的其他待收单 id ──
    extra_ids = [int(x) for x in form.getlist("extra_invoice_ids") if str(x).isdigit()]
    # 跨单收款方式白名单（钱包/套餐/押金/优惠券因为带特殊 ref，多单暂不支持）
    cross_methods = {"cash", "wechat", "alipay", "shouqianba", "meituan", "third_party"}
    multi_target = bool(extra_ids) and method in cross_methods
    if extra_ids and not multi_target:
        # 选了多张但方式不支持 → 提示后只对当前单生效
        from urllib.parse import quote as _q
        _warn = _q("钱包/套餐/押金/优惠券暂只支持单张发票结算；已只结当前单。", safe="")
        # 继续走 single-invoice 流程
    target_invs = [inv]
    if multi_target:
        for eid in extra_ids:
            e = db.get(Invoice, eid)
            if not e or e.customer_id != inv.customer_id:
                continue
            if e.payment_status == "paid":
                continue
            target_invs.append(e)

    outstanding_current = max(0.0, float(inv.total_amount or 0) - _invoice_paid_sum(db, inv.id))
    outstanding_total = outstanding_current
    if multi_target:
        for t in target_invs[1:]:
            outstanding_total += max(0.0, float(t.total_amount or 0) - _invoice_paid_sum(db, t.id))
    if outstanding_total <= 0:
        _invoice_recompute_status(db, inv); db.commit()
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已无欠款", status_code=303)
    try:
        want = float(form.get("amount") or outstanding_total)
    except (TypeError, ValueError):
        want = outstanding_total
    want = max(0.0, min(want, outstanding_total))
    if want <= 0:
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=金额需大于 0", status_code=303)
    if not multi_target and want > outstanding_current:
        want = outstanding_current

    ref_id: int | None = None
    ref_no = (form.get("ref_no") or "").strip()[:120]
    note   = (form.get("note") or "").strip()[:500]
    valid_methods = {
        "cash","wechat","alipay","shouqianba","meituan","third_party",
        "wallet","package","deposit","coupon",
    }
    if method not in valid_methods:
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=未知支付方式 {method}", status_code=303)

    # ── 钱包扣款 ──
    if method == "wallet":
        if not inv.customer_id:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=无客户绑定，无法用钱包", status_code=303)
        wallet = _get_or_create_wallet(db, inv.customer_id)
        try:
            tx = _wallet_apply_tx(
                db, wallet, tx_type="consume", amount=want,
                invoice_id=inv.id, operator=operator, store=store,
                note=f"收费单 {inv.invoice_no or inv.id}",
            )
            ref_id = tx.id
        except HTTPException as he:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg={he.detail}", status_code=303)

    # ── 套餐核销 ──（套餐按"次"扣，金额 want 仅用于记账）
    elif method == "package":
        cp_id_raw = (form.get("customer_package_id") or "").strip()
        if not cp_id_raw.isdigit():
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=请选择要核销的套餐", status_code=303)
        cp = db.get(CustomerPackage, int(cp_id_raw))
        if not cp or cp.customer_id != inv.customer_id:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=套餐与客户不匹配", status_code=303)
        if cp.status != "active" or cp.used_count >= cp.total_uses:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=套餐已失效或用完", status_code=303)
        # 解析勾选的明细项 ids → 拼明细描述写到流水里
        covered_ids = [int(x) for x in form.getlist("pkg_covered_item_ids") if str(x).isdigit()]
        covered_desc = ""
        if covered_ids:
            covered_items = db.query(InvoiceItem).filter(
                InvoiceItem.invoice_id == inv.id,
                InvoiceItem.id.in_(covered_ids),
            ).all()
            if covered_items:
                covered_desc = " · ".join(
                    f"{i.description.split('] ')[-1] if '] ' in i.description else i.description}（¥{i.subtotal:.2f}）"
                    for i in covered_items
                )
        cp.used_count += 1
        if cp.used_count >= cp.total_uses:
            cp.status = "exhausted"
        _redemption_note = f"收费单 {inv.invoice_no or inv.id}"
        if covered_desc:
            _redemption_note += f" · 抵扣：{covered_desc}"
        db.add(PackageRedemption(
            customer_package_id=cp.id, customer_id=cp.customer_id,
            pet_id=inv.pet_id or cp.pet_id, visit_id=inv.visit_id, invoice_id=inv.id,
            used_count=1, remaining_after=cp.total_uses - cp.used_count,
            store=store, operator=operator,
            note=_redemption_note,
        ))
        ref_id = cp.id
        # 加到 Payment.note 让收款流水也看到
        if covered_desc:
            note = (note + " · " if note else "") + f"套餐抵扣：{covered_desc}"

    # ── 押金抵扣 ──
    elif method == "deposit":
        d_id_raw = (form.get("deposit_id") or "").strip()
        if not d_id_raw.isdigit():
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=请选择要使用的押金", status_code=303)
        d = db.get(Deposit, int(d_id_raw))
        if not d or d.customer_id != inv.customer_id:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=押金与客户不匹配", status_code=303)
        remaining = d.amount - (d.applied_amount or 0) - (d.refunded_amount or 0)
        if remaining <= 0 or d.status in ("refunded", "cancelled"):
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=押金已无余额", status_code=303)
        want = min(want, remaining)
        d.applied_amount = (d.applied_amount or 0) + want
        d.applied_invoice_id = inv.id
        d_remaining = d.amount - d.applied_amount - (d.refunded_amount or 0)
        d.status = "applied" if d_remaining <= 1e-6 else "partial_refund"
        ref_id = d.id

    # ── 优惠券核销 ──
    elif method == "coupon":
        c_id_raw = (form.get("coupon_id") or "").strip()
        if not c_id_raw.isdigit():
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=请选择要使用的优惠券", status_code=303)
        c = db.get(Coupon, int(c_id_raw))
        if not c:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=优惠券不存在", status_code=303)
        if c.status != "issued":
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=优惠券状态不允许使用", status_code=303)
        if _coupon_is_expired(c):
            c.status = "expired"; db.commit()
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=优惠券已过期", status_code=303)
        if c.customer_id and c.customer_id != inv.customer_id:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=该券仅指定客户可用", status_code=303)
        # 计算可抵扣
        usable = _coupon_compute_amount(c, float(inv.total_amount or 0))
        if usable <= 0:
            return RedirectResponse(f"/admin/invoices/{inv_id}?msg=未达使用门槛或券无效", status_code=303)
        want = min(want, usable, outstanding)
        c.status = "used"
        c.used_invoice_id = inv.id
        c.used_amount = want
        c.used_at = datetime.utcnow()
        ref_id = c.id

    # cash / wechat / alipay / shouqianba / meituan / third_party — 无 side effect
    # 单张：直接给当前发票写一笔
    # 多张（multi_target=True）：按 target_invs 顺序分摊 want，每张独立 Payment
    if multi_target:
        remaining = want
        for t in target_invs:
            if remaining <= 0:
                break
            t_out = max(0.0, float(t.total_amount or 0) - _invoice_paid_sum(db, t.id))
            if t_out <= 0:
                continue
            pay_amt = round(min(remaining, t_out), 2)
            if pay_amt <= 0:
                continue
            _t_note = note
            if t.id != inv.id and not _t_note:
                _t_note = f"合并结算 (源单 #{inv.id})"
            db.add(Payment(
                invoice_id=t.id, customer_id=t.customer_id,
                method=method, amount=pay_amt,
                ref_id=None, ref_no=ref_no, status="success",
                store=store, operator=operator, note=_t_note,
            ))
            db.flush()
            _invoice_recompute_status(db, t)
            remaining = round(remaining - pay_amt, 2)
    else:
        db.add(Payment(
            invoice_id=inv.id,
            customer_id=inv.customer_id,
            method=method,
            amount=want,
            ref_id=ref_id,
            ref_no=ref_no,
            status="success",
            store=store,
            operator=operator,
            note=note,
        ))
        db.flush()
        _invoice_recompute_status(db, inv)
    # 充值单付清 → 把钱真正打进钱包（幂等）
    try:
        _maybe_credit_wallet_from_invoice(db, inv, request)
    except Exception as _e:
        logger.warning("[wallet_recharge_credit] failed: %s", _e)
    db.commit()
    nu = (form.get("next_url") or "").strip()
    if nu:
        target = _safe_next(nu.replace("{id}", str(inv_id)), f"/admin/invoices/{inv_id}")
        sep = "&" if "?" in target else "?"
        return RedirectResponse(f"{target}{sep}msg=已加 ¥{want:.2f}", status_code=303)
    return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已加 ¥{want:.2f}", status_code=303)


@app.post("/admin/invoices/{inv_id}/payments/{pay_id}/void")
async def admin_invoice_payment_void(
    inv_id: int,
    pay_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    next_url: str = Form(""),
):
    """撤销一笔收款（错收时用，会回滚钱包/套餐/押金/优惠券副作用）。仅 superadmin。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    p = db.get(Payment, pay_id)
    if not p or p.invoice_id != inv_id:
        raise HTTPException(404)
    if p.status != "success":
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=该笔已撤销", status_code=303)
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    # 回滚副作用
    if p.method == "wallet" and p.ref_id:
        wallet = _get_or_create_wallet(db, p.customer_id)
        # 把这笔 consume 退回（用 adjust + 加 amount）
        try:
            _wallet_apply_tx(
                db, wallet, tx_type="adjust", amount=float(p.amount),
                operator=request.session.get("admin_username", "admin"),
                store=p.store, note=f"撤销收费单 {inv.invoice_no or inv.id} 的钱包扣款",
            )
        except HTTPException:
            pass
    elif p.method == "package" and p.ref_id:
        cp = db.get(CustomerPackage, p.ref_id)
        if cp and cp.used_count > 0:
            cp.used_count -= 1
            if cp.status == "exhausted" and cp.used_count < cp.total_uses:
                cp.status = "active"
        # 标 redeem 行
        for r in db.query(PackageRedemption).filter(
            PackageRedemption.invoice_id == inv_id,
            PackageRedemption.customer_package_id == p.ref_id,
        ).all():
            db.delete(r)
    elif p.method == "deposit" and p.ref_id:
        d = db.get(Deposit, p.ref_id)
        if d:
            d.applied_amount = max(0.0, (d.applied_amount or 0) - float(p.amount))
            d.status = "held" if d.applied_amount <= 0 else "partial_refund"
            if d.applied_amount <= 0:
                d.applied_invoice_id = None
    elif p.method == "coupon" and p.ref_id:
        c = db.get(Coupon, p.ref_id)
        if c and c.status == "used" and c.used_invoice_id == inv_id:
            c.status = "issued"
            c.used_invoice_id = None
            c.used_amount = 0.0
            c.used_at = None
    p.status = "cancelled"
    db.flush()
    _invoice_recompute_status(db, inv)
    db.commit()
    if next_url:
        target = _safe_next(next_url.replace("{id}", str(inv_id)), f"/admin/invoices/{inv_id}")
        sep = "&" if "?" in target else "?"
        return RedirectResponse(f"{target}{sep}msg=已撤销 ¥{p.amount:.2f}", status_code=303)
    return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已撤销 ¥{p.amount:.2f}", status_code=303)


@app.post("/admin/invoices/{inv_id}/pay")
async def admin_invoice_pay_legacy(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """老接口：单笔全额支付。内部转发到 add-payment 用 outstanding 金额。"""
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    if inv.payment_status == "paid":
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已收款，请勿重复", status_code=303)
    # 把表单的 payment_method 映射成新接口的 method
    method = str(form.get("payment_method") or "cash").strip()
    # 兼容旧 prepaid → wallet
    if method == "prepaid":
        method = "wallet"
    # 重写表单参数
    forwarded = {
        "csrf_token": str(form.get("csrf_token", "")),
        "method": method,
        "amount": "",  # 留空 → 用 outstanding
        "customer_package_id": str(form.get("customer_package_id", "")),
        "deposit_id": str(form.get("deposit_id", "")),
        "coupon_id": str(form.get("coupon_id", "")),
        "ref_no": str(form.get("ref_no", "")),
        "note": str(form.get("note", "")),
    }
    # 用 starlette 的 _Form 模拟（直接调函数会更省事）
    from starlette.datastructures import FormData
    request._form = FormData(forwarded)
    # 委托给新接口
    return await admin_invoice_add_payment(inv_id=inv_id, request=request, db=db)


@app.post("/admin/invoices/{inv_id}/refund")
async def admin_invoice_refund(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """整单退款：把所有 success 状态的 Payment 逐一 void，
    自动回滚钱包/套餐/押金/优惠券副作用，invoice.payment_status 标 refunded。
    """
    require_admin(request)
    _require_csrf(request, csrf_token)
    require_superadmin(request)
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    if inv.payment_status != "paid":
        return RedirectResponse(f"/admin/invoices/{inv_id}?msg=只有已支付的单可整单退", status_code=303)
    operator = request.session.get("admin_username", "admin")
    pays = db.query(Payment).filter(
        Payment.invoice_id == inv_id, Payment.status == "success"
    ).all()
    voided = 0
    for p in pays:
        # 重用 void 逻辑（复制实现，避免循环 await）
        if p.method == "wallet" and p.ref_id:
            wallet = _get_or_create_wallet(db, p.customer_id)
            try:
                _wallet_apply_tx(
                    db, wallet, tx_type="adjust", amount=float(p.amount),
                    operator=operator, store=p.store,
                    note=f"整单退款 {inv.invoice_no or inv.id} 钱包返还",
                )
            except HTTPException:
                pass
        elif p.method == "package" and p.ref_id:
            cp = db.get(CustomerPackage, p.ref_id)
            if cp and cp.used_count > 0:
                cp.used_count -= 1
                if cp.status == "exhausted" and cp.used_count < cp.total_uses:
                    cp.status = "active"
            for r in db.query(PackageRedemption).filter(
                PackageRedemption.invoice_id == inv_id,
                PackageRedemption.customer_package_id == p.ref_id,
            ).all():
                db.delete(r)
        elif p.method == "deposit" and p.ref_id:
            d = db.get(Deposit, p.ref_id)
            if d:
                d.applied_amount = max(0.0, (d.applied_amount or 0) - float(p.amount))
                d.status = "held" if d.applied_amount <= 0 else "partial_refund"
                if d.applied_amount <= 0:
                    d.applied_invoice_id = None
        elif p.method == "coupon" and p.ref_id:
            c = db.get(Coupon, p.ref_id)
            if c and c.status == "used" and c.used_invoice_id == inv_id:
                c.status = "issued"; c.used_invoice_id = None; c.used_amount = 0.0; c.used_at = None
        p.status = "cancelled"
        voided += 1
    inv.payment_status = "refunded"
    inv.paid_at = None
    inv.notes = ((inv.notes or "") + f"\n[整单退款 by {operator} · 撤销 {voided} 笔]").strip()
    db.commit()
    _audit(db, request, "invoice_refund", application_id=None,
           detail={"invoice_id": inv_id, "voided_payments": voided})
    db.commit()
    return RedirectResponse(f"/admin/invoices/{inv_id}?msg=已整单退款（撤销 {voided} 笔收款）", status_code=303)


@app.post("/admin/invoices/{inv_id}/cancel")
async def admin_invoice_cancel(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    inv = db.get(Invoice, inv_id)
    if inv and inv.payment_status == "unpaid":
        inv.payment_status = "cancelled"
        db.commit()
    return RedirectResponse(f"/admin/invoices/{inv_id}?msg=收费单已取消", status_code=303)


@app.post("/admin/invoices/{inv_id}/apply-discount")
async def admin_invoice_apply_discount(
    inv_id: int, request: Request, db: Session = Depends(get_db),
):
    """整单折扣：填折率（如 0.7 = 7折）或折扣金额，写 Invoice.discount_amount 并重算 total。
    规则：
    - 已付清（paid）拒绝改 — 必须先撤回收款
    - 部分已收时仍可改，但折后总额必须 ≥ 已收 = sum(Payments)
    - mode=pct：折率 0.1-1.0；mode=amount：直接填减免金额
    """
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404, "收费单不存在")
    if inv.payment_status == "paid":
        return RedirectResponse(
            f"/admin/invoices/{inv_id}?msg=已付清的单子不能改折扣，请先撤回收款",
            status_code=303,
        )
    if inv.payment_status == "cancelled":
        return RedirectResponse(
            f"/admin/invoices/{inv_id}?msg=已取消的单子不能改折扣",
            status_code=303,
        )
    mode = (form.get("mode") or "pct").strip()
    reason = (form.get("reason") or "").strip()[:200]
    subtotal = float(inv.subtotal or 0)
    if subtotal <= 0:
        return RedirectResponse(
            f"/admin/invoices/{inv_id}?msg=小计为 0，无可折扣",
            status_code=303,
        )

    # 先算已收金额（套餐 / 现金 / 微信 等所有 success 流水）
    paid_sum = sum(float(p.amount or 0) for p in db.query(Payment).filter(
        Payment.invoice_id == inv_id, Payment.status == "success"
    ).all())
    # 折扣基数 = 未付部分 (subtotal - 已付)
    # 这样套餐已经核销过的次数不会被二次打折
    discount_base = max(0.0, subtotal - paid_sum)

    if mode == "pct":
        # 用户填 70 或 0.7 都识别为 7 折
        raw = (form.get("discount_pct") or "").strip()
        try:
            pct = float(raw)
        except Exception:
            raise HTTPException(400, "折率格式错误")
        if pct > 1.0:
            pct = pct / 100.0
        if pct <= 0 or pct > 1.0:
            raise HTTPException(400, "折率应在 1-100 之间（如填 70 = 7 折）")
        # 折扣只打在未付的部分上（套餐等已结算金额按原价算）
        discount_amt = round(discount_base * (1.0 - pct), 2)
    elif mode == "amount":
        try:
            discount_amt = float((form.get("discount_amount") or "").strip())
        except Exception:
            raise HTTPException(400, "折扣金额格式错误")
        if discount_amt < 0:
            raise HTTPException(400, "折扣金额不能为负")
        # 金额模式：仍允许整单扣（不限制只在未付部分），但不能超过 subtotal
        if discount_amt > subtotal:
            raise HTTPException(400, f"折扣金额应在 0-{subtotal:.2f} 之间")
    elif mode == "clear":
        discount_amt = 0.0
    else:
        raise HTTPException(400, "mode 参数错误")

    new_total = round(subtotal - discount_amt, 2)
    if new_total < paid_sum - 0.005:
        return RedirectResponse(
            f"/admin/invoices/{inv_id}?msg=折后总额 ¥{new_total:.2f} 小于已收 ¥{paid_sum:.2f}，请先撤回部分收款再改折扣",
            status_code=303,
        )

    inv.discount_amount = discount_amt
    inv.total_amount = new_total
    # 备注追加（不覆盖原 notes）
    if reason:
        suffix = f"\n[折扣 {datetime.now().strftime('%Y-%m-%d %H:%M')}] {reason}"
        inv.notes = (inv.notes or "") + suffix
    # 状态：
    # - 折后总额 = 0（整单赠送 / 100% 减免）→ 零结算，直接 paid，备注 method=free
    # - 已收 == 折后总额（> 0）→ paid
    # - 部分收款 → partial
    # - 否则 → unpaid
    if new_total <= 0.005:
        inv.payment_status = "paid"
        inv.paid_at = datetime.utcnow()
        if not inv.payment_method:
            inv.payment_method = "free"
    elif abs(paid_sum - new_total) < 0.005:
        inv.payment_status = "paid"
        inv.paid_at = inv.paid_at or datetime.utcnow()
    elif paid_sum > 0:
        inv.payment_status = "partial"
    else:
        inv.payment_status = "unpaid"
    inv.updated_at = datetime.utcnow()
    db.commit()
    if new_total <= 0.005 and discount_amt > 0:
        msg = f"已应用 100% 减免：¥{subtotal:.2f} 全额赠送 → 单据零结算"
    elif discount_amt > 0:
        msg = f"已应用折扣：原价 ¥{subtotal:.2f} − ¥{discount_amt:.2f} = ¥{new_total:.2f}"
    else:
        msg = "已清除折扣"
    return RedirectResponse(f"/admin/invoices/{inv_id}?msg={msg}", status_code=303)


@app.get("/admin/invoices/{inv_id}/print", response_class=HTMLResponse)
async def admin_invoice_print(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin(request)
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    cust  = db.get(Customer, inv.customer_id) if inv.customer_id else None
    pet   = db.get(Pet,      inv.pet_id)      if inv.pet_id      else None
    visit = db.get(Visit,    inv.visit_id)    if inv.visit_id    else None
    payments = (
        db.query(Payment)
        .filter(Payment.invoice_id == inv_id)
        .order_by(Payment.id.asc())
        .all()
    )
    # 门店全名 / 英文（根据宠物所属门店推断）
    clinic_name_zh = "大风动物医院"
    clinic_name_en = "DaFo Animal Hospital"
    if pet and pet.store:
        clinic_name_zh = f"大风动物医院（{pet.store.replace('店', '分院')}）"
        clinic_name_en = f"DaFo Animal Hospital · {pet.store.replace('店', '')}"
    return templates.TemplateResponse(request, "admin_invoice_print.html", {
        "inv": inv,
        "cust": cust,
        "pet": pet,
        "visit": visit,
        "payments": payments,
        "inv_status_zh": _INV_STATUS_ZH,
        "inv_pay_zh": _INV_PAY_ZH,
        "clinic_name_zh": clinic_name_zh,
        "clinic_name_en": clinic_name_en,
    })


@app.post("/admin/invoices/{inv_id}/delete")
async def admin_invoice_delete(
    inv_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    inv = db.get(Invoice, inv_id)
    if inv:
        db.delete(inv)
        db.commit()
    return RedirectResponse("/admin/invoices?msg=已删除", status_code=303)


# ── 后台：疫苗档案管理 ───────────────────────────────────────────────────────

_VACC_TYPE_ZH = {
    "rabies":    "狂犬疫苗",
    "combo":     "联苗",
    "combo_3":   "猫三联",     # 历史兼容
    "combo_6":   "猫六联",     # 历史兼容
    "canine_8":  "犬八联",     # 历史兼容
    "deworming": "驱虫",
    "other":     "其他",
}
# 新建/编辑表单只显示这两个；驱虫单独走 DewormingRecord，不出现在疫苗表
_VACC_TYPE_OPTIONS = {
    "rabies":    "狂犬疫苗",
    "combo":     "联苗",
}


def _attach_latest_batch(db: Session, items: list) -> None:
    """给一组 InventoryItem 附最新一批的 batch_no / expiry_date，
    供前端 JS 选品目时自动填 批次号 / 有效期。
    用属性 latest_batch_no / latest_expiry_date（不入库），不会影响 ORM。"""
    if not items:
        return
    ids = [it.id for it in items if getattr(it, "id", None)]
    if not ids:
        return
    rows = db.query(InventoryBatch).filter(
        InventoryBatch.item_id.in_(ids),
        InventoryBatch.is_depleted == False,  # noqa: E712
    ).order_by(InventoryBatch.expiry_date.asc(), InventoryBatch.id.desc()).all()
    by_item: dict[int, "InventoryBatch"] = {}
    for b in rows:
        by_item.setdefault(b.item_id, b)
    for it in items:
        b = by_item.get(it.id)
        it.latest_batch_no = (b.batch_no if b else "") or ""
        it.latest_expiry_date = (b.expiry_date if b else "") or ""
_DOSE_ZH = {1: "第1针", 2: "第2针", 3: "第3针", 99: "加强针"}


@app.get("/admin/vaccinations", response_class=HTMLResponse)
async def admin_vaccinations_list(
    request: Request, db: Session = Depends(get_db),
    q: str = Query(""), filter: str = Query("all"),
    vaccine_type: str = Query(""),
):
    require_admin(request)
    from datetime import date, timedelta
    from sqlalchemy import case as sa_case
    today = date.today().isoformat()
    soon  = (date.today() + timedelta(days=7)).isoformat()

    # 按到期日升序，NULL/空值排末尾
    due_order = sa_case(
        (Vaccination.next_due_date.is_(None), "9999-99-99"),
        (Vaccination.next_due_date == "",      "9999-99-99"),
        else_=Vaccination.next_due_date,
    )
    query = db.query(Vaccination).order_by(due_order.asc(), Vaccination.id.desc())

    if q:
        pet_ids  = [p.id for p in db.query(Pet.id).filter(Pet.name.ilike(f"%{q}%")).all()]
        cust_ids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(or_(
            Vaccination.pet_id.in_(pet_ids),
            Vaccination.customer_id.in_(cust_ids),
        ))
    if vaccine_type:
        query = query.filter(Vaccination.vaccine_type == vaccine_type)
    if filter == "soon":
        query = query.filter(
            Vaccination.next_due_date != "",
            Vaccination.next_due_date <= soon,
            Vaccination.next_due_date >= today,
        )
    elif filter == "overdue":
        query = query.filter(
            Vaccination.next_due_date != "",
            Vaccination.next_due_date < today,
        )

    records = query.limit(300).all()
    return templates.TemplateResponse(request, "uk/vaccinations.html", {
        "records": records, "q": q, "filter": filter,
        "vaccine_type": vaccine_type,
        "vacc_type_zh": _VACC_TYPE_ZH,
        "vacc_type_options": _VACC_TYPE_OPTIONS,
        "dose_zh": _DOSE_ZH,
        "today": today, "soon": soon,
        "title": "疫苗管理",
        "msg": request.query_params.get("msg"),
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/vaccinations/create", response_class=HTMLResponse)
async def admin_vaccination_create_page(
    request: Request, db: Session = Depends(get_db),
    pet_id: int = 0, customer_id: int = 0,
):
    require_admin(request)
    from datetime import date
    pet  = db.get(Pet, pet_id)  if pet_id  else None
    cust = db.get(Customer, customer_id) if customer_id else (
        db.get(Customer, pet.customer_id) if pet and pet.customer_id else None
    )
    vets = [v[0] for v in db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]), Staff.position.ilike("%医%")
    ).all()]
    vacc_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category.in_(["vaccine", "antiparasitic"]),
        InventoryItem.stock_qty > 0,
    ).order_by(InventoryItem.name).all()
    # 给每个品目附最近一批的 batch_no + expiry_date（用于 JS 自动填）
    _attach_latest_batch(db, vacc_items)
    # 该宠物的历史疫苗记录（倒序，最多 10 条）
    history = []
    if pet_id:
        history = db.query(Vaccination).filter(Vaccination.pet_id == pet_id)\
            .order_by(Vaccination.vaccinated_date.desc(), Vaccination.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/vaccination.html", {  # B 补 - UK 重写
        "mode": "create", "vacc": None,
        "pet": pet, "cust": cust,
        "vets": vets, "vacc_items": vacc_items,
        "vacc_type_zh": _VACC_TYPE_ZH,
        "vacc_type_options": _VACC_TYPE_OPTIONS,
        "dose_zh": _DOSE_ZH,
        "vacc_history": history,
        "today": date.today().isoformat(),
        "csrf_token": _get_csrf_token(request),
        "msg": None,
    })


@app.post("/admin/vaccinations/create")
async def admin_vaccination_create(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    from datetime import date

    pet_id      = int(form.get("pet_id") or 0) or None
    customer_id = int(form.get("customer_id") or 0) or None
    item_id     = int(form.get("inventory_item_id") or 0) or None
    is_free     = form.get("is_free") == "1"
    admin_name  = request.session.get("admin_username", "")

    vacc = Vaccination(
        pet_id            = pet_id,
        customer_id       = customer_id,
        vaccine_type      = str(form.get("vaccine_type") or "other"),
        vaccine_name      = str(form.get("vaccine_name") or "").strip()[:120],
        batch_no          = str(form.get("batch_no") or "").strip()[:80],
        dose_number       = int(form.get("dose_number") or 1),
        vaccinated_date   = str(form.get("vaccinated_date") or date.today().isoformat()),
        next_due_date     = str(form.get("next_due_date") or ""),
        inventory_item_id = item_id,
        is_free           = is_free,
        vet_name          = str(form.get("vet_name") or "").strip()[:80],
        notes             = str(form.get("notes") or "").strip(),
        created_by        = admin_name,
    )
    db.add(vacc)
    db.flush()

    # 库存出库
    if item_id:
        _deduct_inventory(db, item_id, 1.0, "vaccination", vacc.id, admin_name,
                          note=f"{vacc.vaccine_name or ''} 接种出库")

    # 需要收费 → 自动生成收费单
    # 优先用表单填的 charge_amount；为空才退回到 inventory item 的 sell_price
    try:
        charge_amount = float(form.get("charge_amount") or 0)
    except (ValueError, TypeError):
        charge_amount = 0.0
    if (not is_free) and charge_amount <= 0 and item_id:
        inv_item_for_price = db.get(InventoryItem, item_id)
        if inv_item_for_price:
            from app.services.pricing import effective_sell_price as _eff
            _price = _eff(inv_item_for_price, _get_admin_store(request))
            if _price > 0:
                charge_amount = _price

    if (not is_free) and charge_amount > 0:
        inv = Invoice(
            invoice_no      = _gen_invoice_no(db),
            customer_id     = customer_id,
            pet_id          = pet_id,
            invoice_date    = vacc.vaccinated_date,
            subtotal        = charge_amount,
            discount_amount = 0.0,
            total_amount    = charge_amount,
            payment_status  = "unpaid",
            notes           = f"疫苗接种 #{vacc.id}",
            store           = _resolve_invoice_store(db, pet_id=pet_id, customer_id=customer_id, fallback=_get_op_store(request)),
            created_by      = admin_name,
        )
        db.add(inv)
        db.flush()
        db.add(InvoiceItem(
            invoice_id  = inv.id,
            ref_type    = "vaccination",
            ref_id      = vacc.id,
            description = vacc.vaccine_name or vacc.vaccine_type,
            quantity    = 1.0,
            unit_price  = charge_amount,
            subtotal    = charge_amount,
        ))
        vacc.invoice_id = inv.id

    db.commit()
    redirect = f"/admin/customers/{db.get(Pet, pet_id).customer_id}" if pet_id and db.get(Pet, pet_id) else "/admin/vaccinations"
    msg_part = "疫苗记录已添加"
    if vacc.invoice_id:
        msg_part += f"，收费单 ¥{charge_amount:.2f} 已生成待收款"
    elif not is_free:
        msg_part += "（未生成收费单：金额=0 且未关联有售价的库存品目）"
    next_url_raw = str(form.get("next_url") or "")
    return RedirectResponse(
        _safe_next(next_url_raw, f"{redirect}?msg={msg_part}"),
        status_code=303,
    )


@app.get("/admin/vaccinations/{vacc_id}", response_class=HTMLResponse)
async def admin_vaccination_detail(vacc_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    vacc = db.get(Vaccination, vacc_id)
    if not vacc:
        raise HTTPException(404)
    pet  = db.get(Pet, vacc.pet_id) if vacc.pet_id else None
    cust = db.get(Customer, vacc.customer_id) if vacc.customer_id else None
    vets = [v[0] for v in db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]), Staff.position.ilike("%医%")
    ).all()]
    vacc_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category.in_(["vaccine", "antiparasitic"])
    ).order_by(InventoryItem.name).all()
    _attach_latest_batch(db, vacc_items)
    locked, lock_reason = _is_vaccination_locked(db, vacc)
    paid_amount = _doc_paid_amount(db, "vaccination", vacc_id) if locked else 0.0
    history = []
    if vacc.pet_id:
        history = db.query(Vaccination).filter(
            Vaccination.pet_id == vacc.pet_id,
            Vaccination.id != vacc_id,
        ).order_by(Vaccination.vaccinated_date.desc(), Vaccination.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/vaccination.html", {  # B 补 - UK 重写
        "mode": "edit", "vacc": vacc,
        "pet": pet, "cust": cust,
        "vets": vets, "vacc_items": vacc_items,
        "vacc_type_zh": _VACC_TYPE_ZH,
        "vacc_type_options": _VACC_TYPE_OPTIONS,
        "dose_zh": _DOSE_ZH,
        "vacc_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "today": vacc.vaccinated_date,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/vaccinations/{vacc_id}/edit")
async def admin_vaccination_edit(vacc_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    vacc = db.get(Vaccination, vacc_id)
    if not vacc:
        raise HTTPException(404)
    locked, reason = _is_vaccination_locked(db, vacc)
    if locked:
        raise HTTPException(400, f"疫苗单已锁定（{reason}），不可修改。请「复制为新单」或「作废」后重开。")
    vacc.vaccine_type    = str(form.get("vaccine_type") or "other")
    vacc.vaccine_name    = str(form.get("vaccine_name") or "").strip()[:120]
    vacc.batch_no        = str(form.get("batch_no") or "").strip()[:80]
    vacc.dose_number     = int(form.get("dose_number") or 1)
    vacc.vaccinated_date = str(form.get("vaccinated_date") or "")
    vacc.next_due_date   = str(form.get("next_due_date") or "")
    vacc.vet_name        = str(form.get("vet_name") or "").strip()[:80]
    vacc.notes           = str(form.get("notes") or "").strip()
    db.commit()
    return RedirectResponse(f"/admin/vaccinations/{vacc_id}?msg=已更新", status_code=303)


@app.post("/admin/vaccinations/{vacc_id}/delete")
async def admin_vaccination_delete(vacc_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    vacc = db.get(Vaccination, vacc_id)
    if not vacc:
        return RedirectResponse("/admin/vaccinations?msg=已删除", status_code=303)
    locked, reason = _is_vaccination_locked(db, vacc)
    if locked:
        raise HTTPException(400, f"疫苗单已锁定（{reason}），不可删除。请使用「作废」。")
    operator = request.session.get("admin_username", "")
    pet_cust_id = db.get(Pet, vacc.pet_id).customer_id if vacc.pet_id and db.get(Pet, vacc.pet_id) else None
    item_id = getattr(vacc, "inventory_item_id", None)
    if item_id:
        try:
            _restore_inventory(db, item_id, 1.0, "vaccination", vacc.id, operator, f"删除疫苗#{vacc.id}退回")
        except Exception:
            pass
    _audit(db, request, "vaccination_delete", detail={"vaccination_id": vacc_id})
    db.delete(vacc)
    db.commit()
    if pet_cust_id:
        return RedirectResponse(f"/admin/customers/{pet_cust_id}?msg=疫苗记录已删除", status_code=303)
    return RedirectResponse("/admin/vaccinations?msg=已删除", status_code=303)


@app.post("/admin/vaccinations/{vacc_id}/void")
async def admin_vaccination_void(vacc_id: int, request: Request, db: Session = Depends(get_db),
                                   csrf_token: str = Form(""), void_reason: str = Form(""),
                                   refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0)):
    require_admin(request)
    _require_csrf(request, csrf_token)
    vacc = db.get(Vaccination, vacc_id)
    if not vacc:
        raise HTTPException(404)
    if vacc.status == "voided":
        return RedirectResponse(f"/admin/vaccinations/{vacc_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    item_id = getattr(vacc, "inventory_item_id", None)
    if item_id:
        try:
            _restore_inventory(db, item_id, 1.0, "vaccination_void", vacc.id, operator, f"作废疫苗#{vacc.id}回退")
        except Exception:
            pass
    vacc.status = "voided"
    vacc.voided_by = operator
    vacc.voided_at = datetime.utcnow()
    vacc.void_reason = (void_reason or "")[:200]
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and vacc.customer_id and refund_amount > 0:
        tx = _refund_to_wallet(
            db, vacc.customer_id, float(refund_amount), operator,
            note=f"作废疫苗#{vacc_id} 退款 · {void_reason}"[:500],
        )
        if tx:
            refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
            _audit_doc_action(db, "vaccination", vacc_id, "refund_to_wallet",
                              operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "vaccination", vacc_id, "void", operator, void_reason)
    db.commit()
    return RedirectResponse(f"/admin/vaccinations/{vacc_id}?msg=已作废{refund_msg}", status_code=303)


@app.post("/admin/vaccinations/{vacc_id}/copy-as-new")
async def admin_vaccination_copy_as_new(vacc_id: int, request: Request, db: Session = Depends(get_db),
                                          csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    src = db.get(Vaccination, vacc_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_v = Vaccination(
        pet_id=src.pet_id, customer_id=src.customer_id,
        vaccine_type=src.vaccine_type, vaccine_name=src.vaccine_name,
        batch_no="", dose_number=src.dose_number + 1 if src.dose_number < 99 else 99,
        vaccinated_date=datetime.utcnow().strftime("%Y-%m-%d"),
        next_due_date="",
        inventory_item_id=src.inventory_item_id,
        is_free=src.is_free,
        rabies_record_id=None,
        invoice_id=None,
        vet_name=src.vet_name, notes=src.notes,
        status="active",
        created_by=operator,
    )
    db.add(new_v)
    if src.inventory_item_id:
        try:
            _deduct_inventory(db, src.inventory_item_id, 1.0, "vaccination",
                              0, operator, f"疫苗（复制自 #{vacc_id}）")
        except Exception:
            pass
    _audit_doc_action(db, "vaccination", 0, "copy_from", operator, extra=f"src={vacc_id}")
    db.commit()
    db.refresh(new_v)
    return RedirectResponse(f"/admin/vaccinations/{new_v.id}?msg=已复制为新单 · 请填写最新批次/日期", status_code=303)


@app.post("/admin/vaccinations/send-reminders")
async def admin_send_vaccine_reminders(request: Request, db: Session = Depends(get_db)):
    """手动批量发送 7 天内到期的疫苗提醒（幂等：同一条记录只发一次）。"""
    require_admin(request)
    result = _run_vaccine_reminders(db)
    msg = f"提醒发送完成：成功 {result['sent']} 条，跳过 {result['skipped']} 条（无 openid），失败 {result['errors']} 条"
    return RedirectResponse(f"/admin/vaccinations?msg={quote(msg, safe='')}", status_code=303)


# ── 后台：狂犬疫苗登记管理 ───────────────────────────────────────────────────

@app.get("/admin/rabies", response_class=HTMLResponse)
async def admin_rabies_list(
    request: Request, db: Session = Depends(get_db),
    q: str = Query(""), status: str = Query(""),
    date_from: str = Query(""), date_to: str = Query(""),
    page: int = Query(1),
):
    require_admin(request)
    query = db.query(RabiesVaccineRecord)
    if q:
        query = query.filter(or_(
            RabiesVaccineRecord.owner_name.ilike(f"%{q}%"),
            RabiesVaccineRecord.owner_phone.ilike(f"%{q}%"),
            RabiesVaccineRecord.cert_no.ilike(f"%{q}%"),
        ))
    if status:
        query = query.filter(RabiesVaccineRecord.status == status)
    if date_from:
        query = query.filter(RabiesVaccineRecord.created_at >= date_from)
    if date_to:
        query = query.filter(RabiesVaccineRecord.created_at <= date_to + " 23:59:59")
    total = query.count()
    page_size = 30
    records = query.order_by(RabiesVaccineRecord.id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    total_pages = max(1, (total + page_size - 1) // page_size)
    return templates.TemplateResponse(request, "uk/rabies_list.html", {
        "records": records, "total": total, "page": page,
        "total_pages": total_pages, "page_size": page_size,
        "q": q, "status": status, "date_from": date_from, "date_to": date_to,
        "status_zh": _RABIES_STATUS_ZH,
    })


@app.get("/admin/rabies/{rec_id}", response_class=HTMLResponse)
async def admin_rabies_detail(rec_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v[0] for v in vets]
    locked, lock_reason = _is_rabies_locked(db, rec)
    return templates.TemplateResponse(request, "uk/rabies_detail.html", {
        "rec": rec,
        "vet_names": vet_names,
        "status_zh": _RABIES_STATUS_ZH,
        "locked": locked, "lock_reason": lock_reason,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
        "err": request.query_params.get("err"),
        "owner_name_invalid": _is_invalid_name(rec.owner_name or ""),
    })


@app.post("/admin/rabies/{rec_id}/fill")
async def admin_rabies_fill(rec_id: int, request: Request, db: Session = Depends(get_db)):
    """医护填写疫苗信息 + 签名"""
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    # 医护信息允许在 cert_no 录入后继续修改（仅主人/宠物信息一并被锁）

    # 医护信息：即使 cert_no 已录入（lock = True）也允许修改
    # 用户痛点：常需补录批号 / 改医护名 / 修正日期；锁的应只是主人和宠物信息
    rec.vaccine_manufacturer = str(form.get("vaccine_manufacturer", "")).strip()[:120]
    rec.vaccine_batch_no     = str(form.get("vaccine_batch_no", "")).strip()[:80]
    rec.vaccine_date         = str(form.get("vaccine_date", "")).strip()[:20]
    rec.staff_name           = str(form.get("staff_name", "")).strip()[:80]

    staff_sig_data = str(form.get("staff_signature", "")).strip()
    if staff_sig_data and len(staff_sig_data) > 100:
        rec.staff_signature_path = _save_signature(staff_sig_data, f"staff_{rec_id}")
        rec.staff_signed_at = datetime.utcnow()

    # 状态：未录证号才推进到 completed；已录证号保持 completed 不变
    if rec.status != "completed":
        rec.status = "completed"
    rec.updated_at = datetime.utcnow()
    db.flush()

    # 自动同步到疫苗档案（如果尚未同步过）
    existing_vacc = db.query(Vaccination).filter(Vaccination.rabies_record_id == rec_id).first()
    if not existing_vacc:
        # 查找狂犬疫苗库存品目（优先匹配名称含"狂犬"的）
        rabies_item = _apply_store_filter(
            db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
        ).filter(
            InventoryItem.category == "vaccine",
            InventoryItem.name.ilike("%狂犬%"),
        ).first()
        vacc = Vaccination(
            pet_id            = rec.pet_id,
            customer_id       = rec.customer_id,
            vaccine_type      = "rabies",
            vaccine_name      = rec.vaccine_manufacturer or "狂犬疫苗",
            batch_no          = rec.vaccine_batch_no or "",
            dose_number       = 1,
            vaccinated_date   = rec.vaccine_date or datetime.utcnow().strftime("%Y-%m-%d"),
            next_due_date     = (datetime.strptime(rec.vaccine_date, "%Y-%m-%d") + timedelta(days=365)).strftime("%Y-%m-%d") if rec.vaccine_date else (datetime.utcnow() + timedelta(days=365)).strftime("%Y-%m-%d"),
            inventory_item_id = rabies_item.id if rabies_item else None,
            is_free           = True,
            rabies_record_id  = rec_id,
            vet_name          = rec.staff_name or "",
            created_by        = request.session.get("admin_username", ""),
        )
        db.add(vacc)
        db.flush()
        # 库存出库（有关联库存品目时）
        if rabies_item:
            _deduct_inventory(db, rabies_item.id, 1.0, "vaccination", vacc.id,
                              request.session.get("admin_username", ""), note=f"狂犬疫苗登记#{rec_id} 出库")

    db.commit()
    return RedirectResponse(f"/admin/rabies/{rec_id}?msg=已保存", status_code=303)


@app.post("/admin/rabies/{rec_id}/cert-no")
async def admin_rabies_cert_no(rec_id: int, request: Request, db: Session = Depends(get_db),
                                csrf_token: str = Form(""), cert_no: str = Form("")):
    """录入免疫证号（最后一步 · 录入即锁）"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.cert_no:
        raise HTTPException(400, "免疫证号已录入，记录已锁定。如需修改请「作废重开」。")
    rec.cert_no = cert_no.strip()[:60]
    rec.updated_at = datetime.utcnow()
    _audit_doc_action(db, "rabies", rec_id, "cert_locked",
                      request.session.get("admin_username", ""), extra=f"cert_no={rec.cert_no}")
    db.commit()
    return RedirectResponse(f"/admin/rabies/{rec_id}?msg=免疫证号已录入 · 记录已锁定", status_code=303)


@app.post("/admin/rabies/{rec_id}/delete")
async def admin_rabies_delete(rec_id: int, request: Request, db: Session = Depends(get_db),
                               csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "")
    # 级联：删除关联的疫苗记录 + 退回库存 + 删除签名图片文件
    linked_vaccs = db.query(Vaccination).filter(Vaccination.rabies_record_id == rec_id).all()
    for vacc in linked_vaccs:
        # 通过 vaccine_name 反查库存项（疫苗扣减是按 1.0 数量）
        # 直接走通用 restore：如果 vacc 没记 item_id 就跳过
        item_id = getattr(vacc, "item_id", None)
        if item_id:
            try:
                _restore_inventory(db, item_id, 1.0, "vaccination", vacc.id, operator,
                                   f"删除狂犬记录#{rec_id}退回")
            except Exception:
                pass
        db.delete(vacc)
    # 删除签名文件
    for sig_path in (rec.owner_signature_path, rec.staff_signature_path):
        if sig_path:
            try:
                Path(sig_path).unlink(missing_ok=True)
            except Exception:
                pass
    _audit(db, request, "rabies_delete", detail={"rabies_id": rec_id, "linked_vacc_count": len(linked_vaccs)})
    db.delete(rec)
    db.commit()
    return RedirectResponse("/admin/rabies?msg=记录及关联数据已删除", status_code=303)


@app.get("/admin/rabies/{rec_id}/signature/{who}")
async def admin_rabies_signature(rec_id: int, who: str, request: Request, db: Session = Depends(get_db)):
    """返回签名图片文件"""
    require_admin(request)
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    path = rec.owner_signature_path if who == "owner" else rec.staff_signature_path
    if not path or not Path(path).exists():
        raise HTTPException(404)
    return FileResponse(path, media_type="image/png")


@app.post("/admin/rabies/{rec_id}/edit-owner")
async def admin_rabies_edit_owner(rec_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    rec = db.get(RabiesVaccineRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    locked, reason = _is_rabies_locked(db, rec)
    if locked:
        raise HTTPException(400, f"狂犬登记已锁定（{reason}），不可修改。")
    form = await request.form()
    new_owner_name = str(form.get("owner_name", rec.owner_name)).strip()
    if new_owner_name and _is_invalid_name(new_owner_name):
        return RedirectResponse(
            f"/admin/rabies/{rec_id}?err=请填写真实姓名（不可仅填先生/女士/小姐）",
            status_code=303,
        )
    rec.owner_name    = new_owner_name or rec.owner_name
    rec.owner_phone   = str(form.get("owner_phone", rec.owner_phone)).strip() or rec.owner_phone
    rec.owner_address = str(form.get("owner_address", rec.owner_address or "")).strip()
    rec.animal_name   = str(form.get("animal_name", rec.animal_name or "")).strip()
    rec.animal_breed  = str(form.get("animal_breed", rec.animal_breed or "")).strip()
    rec.animal_dob    = str(form.get("animal_dob", rec.animal_dob or "")).strip()
    rec.animal_gender = str(form.get("animal_gender", rec.animal_gender or "")).strip()
    rec.animal_color  = str(form.get("animal_color", rec.animal_color or "")).strip()
    rec.updated_at = datetime.utcnow()

    # 同步客户档案：姓名若是规范的全名，覆盖客户主档（修复老数据脏名）
    synced_msg = ""
    if rec.customer_id and new_owner_name and not _is_invalid_name(new_owner_name):
        cust = db.get(Customer, rec.customer_id)
        if cust and cust.name != new_owner_name:
            old_name = cust.name
            cust.name = new_owner_name
            synced_msg = f" · 客户档案姓名已同步（{old_name} → {new_owner_name}）"
    # 同步宠物档案：动物名称变更时
    if rec.pet_id and rec.animal_name:
        pet = db.get(Pet, rec.pet_id)
        if pet and pet.name != rec.animal_name:
            pet.name = rec.animal_name
        if pet:
            if rec.animal_breed and pet.breed != rec.animal_breed:
                pet.breed = rec.animal_breed
            if rec.animal_color and pet.color_pattern != rec.animal_color:
                pet.color_pattern = rec.animal_color
            if rec.animal_dob and pet.birthday_estimate != rec.animal_dob:
                pet.birthday_estimate = rec.animal_dob

    db.commit()
    return RedirectResponse(f"/admin/rabies/{rec_id}?msg=信息已更新{synced_msg}", status_code=303)


# 注：上面已有完整版本的 admin_rabies_delete（带 CSRF + 级联清理），此处旧版本已移除


@app.get("/admin/rabies/export/excel")
async def admin_rabies_export(
    request: Request, db: Session = Depends(get_db),
    date_from: str = Query(""), date_to: str = Query(""),
    status: str = Query(""),
):
    """导出 Excel，含签名图片（Pillow 在启动时已自动安装）。"""
    require_admin(request)
    import io
    from fastapi.responses import Response as FastResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage

    query = db.query(RabiesVaccineRecord)
    if status:
        query = query.filter(RabiesVaccineRecord.status == status)
    if date_from:
        query = query.filter(RabiesVaccineRecord.created_at >= date_from)
    if date_to:
        query = query.filter(RabiesVaccineRecord.created_at <= date_to + " 23:59:59")
    records = query.order_by(RabiesVaccineRecord.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "狂犬疫苗免疫登记"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=True)
    hdr_font = Font(bold=True, size=10)
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")

    columns = [
        ("免疫证号",       14, lambda r: r.cert_no),
        ("动物主人姓名",   12, lambda r: r.owner_name),
        ("联系地址",       28, lambda r: r.owner_address),
        ("联系电话",       14, lambda r: r.owner_phone),
        ("动物名称",       10, lambda r: r.animal_name),
        ("品种",           10, lambda r: r.animal_breed),
        ("出生年月/年龄",   9, lambda r: r.animal_dob),
        ("性别",            6, lambda r: {"male": "公", "female": "母"}.get(r.animal_gender, r.animal_gender or "")),
        ("毛色",            6, lambda r: r.animal_color),
        ("疫苗厂家",       10, lambda r: r.vaccine_manufacturer),
        ("批号",            8, lambda r: r.vaccine_batch_no),
        ("免疫时间",       12, lambda r: r.vaccine_date),
        ("免疫人员",       10, lambda r: r.staff_name),
        ("医护签名",       15, None),
        ("主人签名",       15, None),
    ]
    SIG_STAFF_COL = 14  # 1-based
    SIG_OWNER_COL = 15
    ROW_H = 35

    for col_idx, (title, width, _) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[1].height = 35

    # 保持 BytesIO 引用防止 GC 在 wb.save() 前回收
    _img_bufs: list = []

    def _embed_sig(sig_path: str, col_1idx: int, row_idx: int) -> None:
        if not sig_path:
            return
        p = Path(sig_path)
        if not p.exists():
            return
        try:
            buf = io.BytesIO(p.read_bytes())
            _img_bufs.append(buf)
            xl_img = XLImage(buf)
            scale = min(180 / max(xl_img.width, 1), 40 / max(xl_img.height, 1), 1.0)
            xl_img.width  = int(xl_img.width  * scale)
            xl_img.height = int(xl_img.height * scale)
            xl_img.anchor = f"{get_column_letter(col_1idx)}{row_idx}"
            ws.add_image(xl_img)
        except Exception:
            pass

    for r_idx, rec in enumerate(records, 2):
        ws.row_dimensions[r_idx].height = ROW_H
        for c_idx, (_, _, getter) in enumerate(columns, 1):
            if getter is not None:
                try:
                    val = getter(rec)
                except Exception:
                    val = ""
            else:
                val = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = center; cell.border = border
        _embed_sig(rec.staff_signature_path, SIG_STAFF_COL, r_idx)
        _embed_sig(rec.owner_signature_path,  SIG_OWNER_COL,  r_idx)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    fname = f"狂犬疫苗登记_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return FastResponse(
        content=out.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ── 检查单 ────────────────────────────────────────────────────────────────────

def _exam_order_token(db: Session) -> tuple[str, "datetime"]:
    """生成唯一 upload_token + 24小时到期时间。"""
    while True:
        token = secrets.token_urlsafe(20)
        if not db.query(ExamOrder).filter(ExamOrder.upload_token == token).first():
            return token, datetime.utcnow() + timedelta(hours=24)


@app.get("/admin/exam-orders/create", response_class=HTMLResponse)
async def admin_exam_order_create_page(
    request: Request,
    db: Session = Depends(get_db),
    visit_id: int = Query(0),
):
    require_admin(request)
    visit = db.get(Visit, visit_id) if visit_id else None
    if not visit:
        return RedirectResponse("/admin/visits")
    cust = db.get(Customer, visit.customer_id) if visit.customer_id else None
    pet  = db.get(Pet, visit.pet_id) if visit.pet_id else None
    history = []
    if pet:
        history_q = db.query(ExamOrder).join(Visit, ExamOrder.visit_id == Visit.id)\
            .filter(Visit.pet_id == pet.id).order_by(ExamOrder.id.desc()).limit(10).all()
        for h in history_q:
            try:
                h._items_parsed = json.loads(h.items_json or "[]")
            except Exception:
                h._items_parsed = []
            h._has_report = db.query(ExamReport).filter(ExamReport.exam_order_id == h.id).first() is not None
        history = history_q
    return templates.TemplateResponse(request, "uk/exam_form.html", {  # B8.7 UK 重写
        "visit": visit, "cust": cust, "pet": pet,
        "exam_order": None, "mode": "create",
        "exam_history": history,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/exam-orders/create")
async def admin_exam_order_create(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    visit_id = int(form.get("visit_id") or 0)
    notes    = str(form.get("notes") or "").strip()
    # 病历已结束 → 不能开新检查
    if visit_id:
        _v = db.get(Visit, visit_id)
        if _v and (_v.status or "open") == "closed":
            raise HTTPException(403, "所属病历已结束，不可新增检查单；如需追加请新建病历")

    # 收集检查项目
    items: list[dict] = []
    idx = 0
    while idx < 30:
        name = str(form.get(f"item_name_{idx}") or "").strip()
        if name:
            try:
                qty        = float(form.get(f"item_qty_{idx}")   or 1)
                unit_price = float(form.get(f"item_price_{idx}") or 0)
            except (ValueError, TypeError):
                qty, unit_price = 1.0, 0.0
            subtotal = round(qty * unit_price, 2)
            items.append({
                "name":       name,
                "item_id":    int(form.get(f"item_id_{idx}") or 0) or None,
                "qty":        qty,
                "unit":       str(form.get(f"item_unit_{idx}")  or "").strip(),
                "unit_price": unit_price,
                "subtotal":   subtotal,
                "notes":      str(form.get(f"item_notes_{idx}") or "").strip(),
            })
        idx += 1

    # 整支/整瓶计费 + 员工内购档案：单价改填进价
    _apply_single_use_pack_billing(db, items)
    if visit_id:
        _v = db.get(Visit, visit_id)
        if _v and _v.customer_id:
            _apply_internal_pricing(db, items, _v.customer_id)

    token, exp = _exam_order_token(db)
    order = ExamOrder(
        visit_id=visit_id,
        items_json=json.dumps(items, ensure_ascii=False),
        notes=notes,
        upload_token=token,
        token_expires_at=exp,
        created_by=request.session.get("admin_username", ""),
    )
    db.add(order)
    db.commit()
    db.refresh(order)
    # 同步收费单
    if order.visit_id:
        _sync_visit_invoice(db, order.visit_id, request.session.get("admin_username", ""))
        db.commit()
    # M6 移动端 next_url（{id} 占位 = 新建检查单 id）
    next_url_raw = str(form.get("next_url") or "")
    if next_url_raw:
        nu = next_url_raw.replace("{id}", str(order.id))
        return RedirectResponse(_safe_next(nu, f"/admin/exam-orders/{order.id}"), status_code=303)
    return RedirectResponse(f"/admin/exam-orders/{order.id}", status_code=303)


# 报告类型自动识别 → 不同标题 / 颜色 / 网格
_REPORT_STYLES = {
    "ultrasound": {
        "title_zh": "B 超 检 查 报 告", "title_en": "Ultrasound Examination Report",
        "accent": "#0369a1", "frame_bg": "#000", "frame_border": "#1e3a8a",
        "grid_cols": "1fr 1fr", "media_label": "影 像 资 料",
        "impression_label": "超声所见与印象 · Findings & Impression",
        "keywords": ["B超", "超声", "ultrasound"],
    },
    "xray": {
        "title_zh": "X 光 检 查 报 告", "title_en": "Radiographic Examination Report",
        "accent": "#1f2937", "frame_bg": "#000", "frame_border": "#111",
        "grid_cols": "1fr 1fr", "media_label": "X 光 片",
        "impression_label": "影像所见与印象 · Findings & Impression",
        "keywords": ["X光", "DR", "放射", "x-ray", "xray"],
    },
    "microscope": {
        "title_zh": "显 微 镜 检 查 报 告", "title_en": "Microscopic Examination Report",
        "accent": "#7c3aed", "frame_bg": "#f9fafb", "frame_border": "#c4b5fd",
        "grid_cols": "1fr 1fr 1fr", "media_label": "镜 检 视 野",
        "impression_label": "镜检所见与诊断 · Findings & Diagnosis",
        "keywords": ["显微", "镜检", "涂片", "细胞学", "粪检"],
    },
    "lab": {
        "title_zh": "化 验 检 查 报 告", "title_en": "Laboratory Test Report",
        "accent": "#059669", "frame_bg": "#f9fafb", "frame_border": "#d1d5db",
        "grid_cols": "1fr 1fr", "media_label": "化 验 单 据",
        "impression_label": "化验结果与解读 · Results & Interpretation",
        "keywords": ["血常规", "生化", "尿检", "化验", "血液", "lab"],
    },
    "generic": {
        "title_zh": "检 查 报 告", "title_en": "Examination Report",
        "accent": "#374151", "frame_bg": "#f9fafb", "frame_border": "#d1d5db",
        "grid_cols": "1fr 1fr", "media_label": "检 查 资 料",
        "impression_label": "检查所见与结论",
        "keywords": [],
    },
}


def _detect_report_style(items: list) -> dict:
    """从检查项目名称推断报告类型样式。"""
    text = " ".join((it.get("name") or "") for it in items).lower()
    for key, style in _REPORT_STYLES.items():
        if key == "generic":
            continue
        for kw in style["keywords"]:
            if kw.lower() in text:
                return style
    return _REPORT_STYLES["generic"]


@app.get("/admin/exam-orders/{order_id}/print", response_class=HTMLResponse)
async def admin_exam_order_print(
    order_id: int, request: Request, db: Session = Depends(get_db),
):
    """检查报告打印（按项目类型自动选样式：B超/X光/显微镜/化验/通用）。"""
    if not request.session.get("admin"):
        return RedirectResponse("/admin/login")
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust = db.get(Customer, visit.customer_id) if visit and visit.customer_id else None
    pet = db.get(Pet, visit.pet_id) if visit and visit.pet_id else None
    items = json.loads(order.items_json or "[]")
    style = _detect_report_style(items)
    image_reports = [r for r in order.reports if (r.file_type or "image").lower() != "pdf"]
    pdf_reports   = [r for r in order.reports if (r.file_type or "").lower() == "pdf"]
    # clinic 名
    clinic_name_zh = "大风动物医院"
    if pet and pet.store:
        clinic_name_zh = f"大风动物医院（{pet.store.replace('店', '分院')}）"
    return templates.TemplateResponse(request, "admin_exam_print.html", {
        "order": order, "visit": visit, "cust": cust, "pet": pet,
        "items": items,
        "image_reports": image_reports,
        "pdf_reports": pdf_reports,
        "report_style": style,
        "clinic_name_zh": clinic_name_zh,
    })


@app.get("/admin/exam-reports/pending", response_class=HTMLResponse)
async def admin_exam_reports_pending(
    request: Request,
    db: Session = Depends(get_db),
    days: int = Query(30),
):
    """未出检查报告全院清单（工作台卡片『全部 →』跳过来）。
    粒度到「项」：一张检查单 6 项，已上传 X 个、缺 Y 个 → 列 Y 项。
    item_label="" 的通用报告视为覆盖全部项 → 不算缺。
    """
    require_admin(request)
    days = max(1, min(days, 365))
    cutoff = datetime.now() - timedelta(days=days)
    store_short = _get_admin_store(request)
    q = db.query(ExamOrder).filter(
        ExamOrder.status != "voided",
        ExamOrder.created_at >= cutoff,
    )
    if store_short:
        q = q.join(Visit, ExamOrder.visit_id == Visit.id, isouter=True)\
             .join(Pet, Visit.pet_id == Pet.id, isouter=True)\
             .filter(or_(Pet.store == store_short, Pet.store == None))
    rows = q.order_by(ExamOrder.created_at.asc()).all()  # 老的在前（紧迫）

    pending_rows = []
    for eo in rows:
        try:
            items = json.loads(eo.items_json or "[]")
        except Exception:
            items = []
        item_names = []
        for it in items:
            n = (it.get("name") or "").strip()
            if not n:
                continue
            iid = it.get("item_id")
            if iid:
                inv = db.get(InventoryItem, int(iid))
                if inv is not None and getattr(inv, "requires_report", True) is False:
                    continue
            item_names.append(n)
        if not item_names:
            continue
        reports = db.query(ExamReport).filter(ExamReport.exam_order_id == eo.id).all()
        # 显微镜报告也算
        micro = db.query(MicroscopyReport).filter(MicroscopyReport.exam_order_id == eo.id).all()
        reported_labels = {(r.item_label or "").strip() for r in reports}
        reported_labels |= {(m.item_label or "").strip() for m in micro}
        if "" in reported_labels and (reports or micro):
            continue  # 有通用报告
        missing = [n for n in item_names if n not in reported_labels]
        if not missing:
            continue
        visit = db.get(Visit, eo.visit_id) if eo.visit_id else None
        cust = db.get(Customer, visit.customer_id) if visit and visit.customer_id else None
        pet = db.get(Pet, visit.pet_id) if visit and visit.pet_id else None
        # 紧迫度：> 7 天红、> 3 天琥珀、其余正常
        age_days = (datetime.now() - eo.created_at).days if eo.created_at else 0
        urgency = "red" if age_days > 7 else ("amber" if age_days > 3 else "")
        pending_rows.append({
            "eo": eo, "visit": visit, "cust": cust, "pet": pet,
            "missing": missing, "done": len(item_names) - len(missing),
            "total": len(item_names), "age_days": age_days, "urgency": urgency,
        })

    return templates.TemplateResponse(request, "uk/exam_reports_pending.html", {
        "rows": pending_rows,
        "days": days,
        "csrf_token": _get_csrf_token(request),
        "store_label": store_short or "全部门店",
    })


@app.get("/admin/exam-orders/{order_id}", response_class=HTMLResponse)
async def admin_exam_order_detail(
    order_id: int, request: Request, db: Session = Depends(get_db),
    msg: str = Query(""),
):
    require_admin(request)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    visit = db.get(Visit, order.visit_id)
    cust  = db.get(Customer, visit.customer_id) if visit and visit.customer_id else None
    pet   = db.get(Pet, visit.pet_id) if visit and visit.pet_id else None
    items = json.loads(order.items_json or "[]")
    # Token 过期则自动刷新
    if not order.token_expires_at or order.token_expires_at < datetime.utcnow():
        order.upload_token, order.token_expires_at = _exam_order_token(db)
        db.commit()
    locked, lock_reason = _is_exam_order_locked(db, order)
    paid_amount = _doc_paid_amount(db, "exam_order", order_id) if locked else 0.0
    # 该宠物历史检查单（通过 Visit.pet_id 反查，最多 10 条）
    history = []
    if visit and visit.pet_id:
        history_q = db.query(ExamOrder).join(Visit, ExamOrder.visit_id == Visit.id)\
            .filter(Visit.pet_id == visit.pet_id, ExamOrder.id != order_id)\
            .order_by(ExamOrder.id.desc()).limit(10).all()
        for h in history_q:
            try:
                h._items_parsed = json.loads(h.items_json or "[]")
            except Exception:
                h._items_parsed = []
            h._has_report = db.query(ExamReport).filter(ExamReport.exam_order_id == h.id).first() is not None
        history = history_q
    # 判定本检查单是否包含显微镜类项目（→ 是否显示「生成显微镜报告」入口）
    micro_items = []
    for it in items:
        iid = it.get("item_id") if isinstance(it, dict) else None
        if iid:
            inv = db.get(InventoryItem, int(iid))
            if inv and (inv.category or "") == "microscopy":
                micro_items.append({"name": it.get("name") or inv.name, "item_id": iid})
        else:
            # 无 item_id 兜底：名称含关键词也算
            n = (it.get("name") if isinstance(it, dict) else "") or ""
            if any(k in n for k in ("镜检", "镜下", "刮片", "涂片", "粪检", "皮肤检查", "耳道分泌", "阴道脱落", "显微")):
                micro_items.append({"name": n, "item_id": None})
    has_microscopy = len(micro_items) > 0
    # 已生成的显微镜报告（按 exam_order_id 反查）
    micro_reports = db.query(MicroscopyReport).filter(
        MicroscopyReport.exam_order_id == order_id
    ).order_by(MicroscopyReport.id.desc()).all()
    # 已被某份报告认领的项目名 set（用于上传下拉灰显「已上传」+ 项目卡片状态）
    assigned_labels = set()
    for rpt in (order.reports or []):
        lbl = (rpt.item_label or "").strip()
        if lbl:
            assigned_labels.add(lbl)
    for mr in micro_reports:
        lbl = (mr.item_label or "").strip()
        if lbl:
            assigned_labels.add(lbl)
    return templates.TemplateResponse(request, "uk/exam_detail.html", {  # B8.7 UK 重写
        "order": order, "visit": visit, "cust": cust, "pet": pet,
        "items": items, "msg": msg,
        "exam_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "assigned_labels": assigned_labels,
        "has_microscopy": has_microscopy, "micro_items": micro_items,
        "micro_reports": micro_reports,
        "csrf_token": _get_csrf_token(request),
    })


_EXAM_REPORT_EXT_OK = {".pdf", ".jpg", ".jpeg", ".png", ".heic", ".webp"}
_EXAM_REPORT_MAX_BYTES = 20 * 1024 * 1024  # 20 MB


@app.post("/admin/exam-orders/{order_id}/upload")
async def admin_exam_order_upload(
    order_id: int, request: Request, db: Session = Depends(get_db),
    file: UploadFile = File(...),
    item_label: str = Form(""),
    csrf_token: str = Form(""),
    next_url: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)

    fname = file.filename or "report"
    ext = Path(fname).suffix.lower()
    if ext not in _EXAM_REPORT_EXT_OK:
        raise HTTPException(400, f"不支持的文件类型 {ext}，仅允许 PDF/JPG/PNG/HEIC/WEBP")
    data = await file.read()
    if len(data) > _EXAM_REPORT_MAX_BYTES:
        raise HTTPException(413, "文件超过 20MB 上限")
    ftype = "pdf" if ext == ".pdf" else "image"

    dest_dir = Path(settings.upload_dir) / "exam_reports" / str(order_id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f"rpt_{secrets.token_hex(8)}{ext}"
    dest.write_bytes(data)

    report = ExamReport(
        exam_order_id=order_id,
        file_path=str(dest),
        original_name=fname,
        file_type=ftype,
        item_label=(item_label or "").strip()[:120],
        uploaded_by=request.session.get("admin_username", "系统"),
    )
    db.add(report)
    order.status = "completed"
    order.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/exam-orders/{order_id}?msg=报告已上传"),
        status_code=303,
    )


@app.post("/admin/exam-orders/{order_id}/refresh-token")
async def admin_exam_order_refresh_token(
    order_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    order.upload_token, order.token_expires_at = _exam_order_token(db)
    db.commit()
    return RedirectResponse(f"/admin/exam-orders/{order_id}?msg=二维码已刷新", status_code=303)


@app.post("/admin/exam-orders/{order_id}/delete-report/{report_id}")
async def admin_exam_report_delete(
    order_id: int, report_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rpt = db.get(ExamReport, report_id)
    if rpt and rpt.exam_order_id == order_id:
        try:
            Path(rpt.file_path).unlink(missing_ok=True)
        except Exception:
            pass
        db.delete(rpt)
        db.commit()
    return RedirectResponse(f"/admin/exam-orders/{order_id}?msg=报告已删除", status_code=303)


@app.post("/admin/exam-orders/{order_id}/delete")
async def admin_exam_order_delete(
    order_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    return_to: str = Form(""),
):
    """删除整个检查单（连同所有报告）。锁定的请走 /void。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404, "检查单不存在")
    locked, reason = _is_exam_order_locked(db, order)
    if locked:
        raise HTTPException(400, f"检查单已锁定（{reason}），不可删除。请使用「作废」。")
    if order.visit_id:
        _v = db.get(Visit, order.visit_id)
        if _v and (_v.status or "open") == "closed":
            raise HTTPException(403, "所属病历已结束，检查单不可删除。如确需作废请使用「作废」。")
    visit_id = order.visit_id
    # 先删本地报告文件
    for rpt in list(order.reports or []):
        try:
            if rpt.file_path:
                Path(rpt.file_path).unlink(missing_ok=True)
        except Exception:
            pass
    db.delete(order)  # cascade 会删 reports
    db.commit()
    # 同步收费单
    if visit_id:
        _sync_visit_invoice(db, visit_id, request.session.get("admin_username", ""))
        db.commit()
    if return_to == "visit" and visit_id:
        return RedirectResponse(f"/admin/visits/{visit_id}?step=3&msg=检查单已删除", status_code=303)
    return RedirectResponse(f"/admin/visits/{visit_id}?msg=检查单已删除" if visit_id else "/admin/customers", status_code=303)


@app.post("/admin/exam-orders/{order_id}/void")
async def admin_exam_order_void(
    order_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""), void_reason: str = Form(""),
    refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    if order.status == "voided":
        return RedirectResponse(f"/admin/exam-orders/{order_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    order.status = "voided"
    order.voided_by = operator
    order.voided_at = datetime.utcnow()
    order.void_reason = (void_reason or "")[:200]
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and order.visit_id and refund_amount > 0:
        visit = db.get(Visit, order.visit_id)
        cust_id = visit.customer_id if visit else 0
        if cust_id:
            tx = _refund_to_wallet(
                db, cust_id, float(refund_amount), operator,
                note=f"作废检查单#{order_id} 退款 · {void_reason}"[:500],
            )
            if tx:
                refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
                _audit_doc_action(db, "exam_order", order_id, "refund_to_wallet",
                                  operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "exam_order", order_id, "void", operator, void_reason)
    db.commit()
    if order.visit_id:
        try:
            _sync_visit_invoice(db, order.visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/exam-orders/{order_id}?msg=已作废{refund_msg}", status_code=303)


@app.post("/admin/exam-orders/{order_id}/copy-as-new")
async def admin_exam_order_copy_as_new(
    order_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    src = db.get(ExamOrder, order_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    token, exp = _exam_order_token(db)
    new_order = ExamOrder(
        visit_id=src.visit_id,
        items_json=src.items_json,
        notes=src.notes,
        status="pending",
        upload_token=token,
        token_expires_at=exp,
        created_by=operator,
    )
    db.add(new_order)
    _audit_doc_action(db, "exam_order", new_order.id, "copy_from", operator, extra=f"src={order_id}")
    db.commit()
    db.refresh(new_order)
    if new_order.visit_id:
        try:
            _sync_visit_invoice(db, new_order.visit_id, operator)
            db.commit()
        except Exception:
            pass
    return RedirectResponse(f"/admin/exam-orders/{new_order.id}?msg=已复制为新单", status_code=303)


# ─── 显微镜检查报告（皮肤刮片 / 耳道分泌物 / 粪检 等手工出报告）─────────────
# 按 3 张 Word 模板拆字段：每张含 3 段 — 镜检可见 / 寄生虫定性 / 病理可见
_MICRO_GRADES = ["-", "+", "++", "+++"]   # 半定量
_MICRO_PARA_OPTS = ["阴性", "阳性"]
_MICRO_MAGS = ["10x", "40x", "100x（油镜）"]

_MICRO_TEMPLATES = {
    "skin": {
        "label": "皮肤检查",
        "microbe": [
            "球菌", "链球菌", "杆菌", "绿脓杆菌", "破伤风梭菌", "分歧杆菌",
            "酵母菌", "马拉色菌", "念珠菌", "真菌孢子",
            "嗜中性粒细胞", "嗜酸性粒细胞", "嗜碱性粒细胞",
            "巨噬细胞", "淋巴细胞", "红细胞", "上皮细胞", "角质细胞",
        ],
        "parasite": ["蠕形螨", "疥螨", "羌螨", "其他螨虫"],
        "pathology": [
            {"name": "伍德氏灯检查", "options": ["阴性", "阳性", "未检"]},
            {"name": "毛根",     "options": ["生长期>50%", "休止期>50%", "未观察"]},
            {"name": "毛干",     "options": ["无损伤", "朽木样", "断裂", "未观察"]},
        ],
    },
    "ear": {
        "label": "耳道检查",
        "microbe": [
            "脂滴", "球菌", "链球菌", "杆菌", "绿脓杆菌", "分歧杆菌",
            "酵母菌", "马拉色菌", "念珠菌", "真菌孢子",
            "嗜中性粒细胞", "嗜酸性粒细胞", "嗜碱性粒细胞",
            "巨噬细胞", "淋巴细胞", "红细胞", "上皮细胞", "角质细胞", "异常细胞",
        ],
        "parasite": ["蠕形螨", "疥螨", "羌螨", "其他螨虫"],
        "pathology": [
            {"name": "耳道皮肤",       "options": ["无损", "破溃", "化脓"]},
            {"name": "分泌物颜色",     "options": ["褐色", "黄绿色", "黑色", "无色透明"]},
            {"name": "耳道分泌物形态", "options": ["干燥", "粘稠", "油性", "水样"]},
        ],
    },
    "fecal": {
        "label": "粪便检查",
        "microbe": [
            "弯曲杆菌", "螺旋杆菌", "分歧杆菌", "球菌/杆菌比", "产气荚膜梭菌",
            "螺旋体", "酵母菌", "霍乱弧菌",
            "嗜中性粒细胞", "嗜酸性粒细胞", "巨噬细胞", "淋巴细胞", "成纤维细胞",
            "红细胞", "上皮细胞",
            "淀粉颗粒", "脂滴", "植物纤维", "肌纤维",
        ],
        "parasite": [
            "蛔虫卵", "钩虫卵", "绦虫卵", "鞭虫卵",
            "滴虫", "贾第鞭毛虫", "阿米巴虫", "球虫",
        ],
        "pathology": [
            {"name": "形状（布里斯托分级）", "options": ["1级", "2级", "3级", "4级", "5级", "6级", "7级"]},
            {"name": "颜色", "options": ["类红色", "类黄色", "类绿色", "类黑色", "类灰色"]},
            {"name": "气味", "options": ["微臭", "酸臭", "腥臭", "恶臭"]},
        ],
    },
    "general": {
        "label": "通用",
        "microbe": [
            "马拉色菌", "球菌", "杆菌", "真菌菌丝/孢子",
            "上皮细胞", "白细胞", "红细胞",
        ],
        "parasite": ["螨虫", "寄生虫卵"],
        "pathology": [],
    },
}

# 模板顺序（前端 chip 顺序 + 默认推断）
_MICRO_TEMPLATE_ORDER = ["skin", "ear", "fecal", "general"]


def _infer_micro_template(item_label: str) -> str:
    """根据归属检查项名字推断默认模板。"""
    s = (item_label or "")
    if any(k in s for k in ("皮肤", "刮片", "毛", "癣")):
        return "skin"
    if any(k in s for k in ("耳", "外耳", "中耳")):
        return "ear"
    if any(k in s for k in ("粪", "便", "肠")):
        return "fecal"
    return "general"
_MICRO_PHOTO_EXT_OK = {".jpg", ".jpeg", ".png", ".webp", ".heic"}
_MICRO_PHOTO_MAX = 10 * 1024 * 1024  # 10 MB / 张


def _micro_form_ctx(request: "Request", db: Session, order: "ExamOrder",
                    report: "MicroscopyReport | None" = None):
    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust = db.get(Customer, visit.customer_id) if visit and visit.customer_id else None
    pet = db.get(Pet, visit.pet_id) if visit and visit.pet_id else None
    items = json.loads(order.items_json or "[]")
    micro_items = []
    for it in items:
        iid = it.get("item_id") if isinstance(it, dict) else None
        is_micro = False
        if iid:
            inv = db.get(InventoryItem, int(iid))
            if inv and (inv.category or "") == "microscopy":
                is_micro = True
                micro_items.append({"name": it.get("name") or inv.name})
        if not is_micro:
            n = (it.get("name") if isinstance(it, dict) else "") or ""
            if any(k in n for k in ("镜检", "镜下", "刮片", "涂片", "粪检", "皮肤检查", "耳道分泌", "阴道脱落", "显微")):
                micro_items.append({"name": n})
    # 兽医候选
    vets = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vet_names = [v[0] for v in vets]
    # 已有 findings（编辑模式）：拆成 3 段字典 cur[cat][name] = grade
    cur = {"microbe": {}, "parasite": {}, "pathology": {}}
    if report:
        try:
            for f in json.loads(report.findings_json or "[]") or []:
                cat = f.get("cat", "microbe")
                if cat in cur:
                    cur[cat][f.get("name", "")] = f.get("grade", "")
        except Exception:
            pass
    cur_photos = []
    if report:
        try:
            cur_photos = json.loads(report.photos_json or "[]") or []
        except Exception:
            cur_photos = []
    # 默认模板
    if report and report.template_type in _MICRO_TEMPLATES:
        active_tpl = report.template_type
    else:
        # 没指定 → 按当前选中的检查项名字推断；若仍无 → general
        first_name = (micro_items[0]["name"] if micro_items else "")
        active_tpl = _infer_micro_template(first_name)
    # 模板列表（含 label）按固定顺序
    tpl_list = [{"key": k, "label": _MICRO_TEMPLATES[k]["label"]} for k in _MICRO_TEMPLATE_ORDER]
    return {
        "order": order, "visit": visit, "cust": cust, "pet": pet,
        "micro_items": micro_items, "vet_names": vet_names,
        "templates": _MICRO_TEMPLATES, "tpl_list": tpl_list,
        "active_tpl": active_tpl,
        "grades": _MICRO_GRADES, "para_opts": _MICRO_PARA_OPTS, "mags": _MICRO_MAGS,
        "report": report, "cur": cur, "cur_photos": cur_photos,
        "csrf_token": _get_csrf_token(request),
    }


@app.get("/admin/exam-orders/{order_id}/microscopy/new", response_class=HTMLResponse)
async def page_admin_microscopy_new(order_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    locked, _ = _is_exam_order_locked(db, order)
    # 报告允许在 locked 后继续补（与上传图片同口径）
    ctx = _micro_form_ctx(request, db, order, report=None)
    ctx["locked"] = locked
    return templates.TemplateResponse(request, "uk/microscopy_form.html", ctx)


@app.get("/admin/microscopy-reports/{report_id}/edit", response_class=HTMLResponse)
async def page_admin_microscopy_edit(report_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    report = db.get(MicroscopyReport, report_id)
    if not report:
        raise HTTPException(404)
    order = db.get(ExamOrder, report.exam_order_id)
    if not order:
        raise HTTPException(404)
    # 限店权限
    admin_store = _get_admin_store(request)
    if admin_store and report.store and report.store != admin_store:
        raise HTTPException(403, "无权操作其他门店")
    ctx = _micro_form_ctx(request, db, order, report=report)
    ctx["locked"] = False
    return templates.TemplateResponse(request, "uk/microscopy_form.html", ctx)


@app.post("/admin/microscopy/ai-draft")
async def admin_microscopy_ai_draft(request: Request, db: Session = Depends(get_db)):
    """根据医生已勾选的镜检数据，调 LLM 生成"镜下所见 / 结论 / 建议"三段文字稿。
    入参：application/json
      {
        template_type, item_label, sample_site, magnification, pet_species,
        findings: [{cat, name, grade}],
        extras:   [{name, grade}],
        narrative_user: str
      }
    出参：{ok, narrative, conclusion, advice, error?}
    """
    require_admin(request)
    try:
        payload = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "请求体不是 JSON"}, status_code=400)
    if not isinstance(payload, dict):
        return JSONResponse({"ok": False, "error": "请求体格式错误"}, status_code=400)
    from app.services.microscopy_ai import draft_microscopy_text
    result = await draft_microscopy_text(payload)
    return JSONResponse(result)


@app.post("/admin/exam-orders/{order_id}/microscopy/create")
async def admin_microscopy_create(order_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))

    visit = db.get(Visit, order.visit_id) if order.visit_id else None
    cust_id = visit.customer_id if visit else None
    pet_id = visit.pet_id if visit else None
    pet = db.get(Pet, pet_id) if pet_id else None

    item_label = str(form.get("item_label", "")).strip()[:120]
    template_type = str(form.get("template_type", "general")).strip()
    if template_type not in _MICRO_TEMPLATES:
        template_type = "general"
    vet_name = str(form.get("vet_name", "")).strip()[:80]
    magnification = str(form.get("magnification", "")).strip()[:20]
    sample_site = str(form.get("sample_site", "")).strip()[:120]
    narrative = str(form.get("narrative", "")).strip()
    conclusion = str(form.get("conclusion", "")).strip()
    advice = str(form.get("advice", "")).strip()

    if not item_label:
        raise HTTPException(400, "请选择归属检查项")
    if not vet_name:
        raise HTTPException(400, "兽医必填")

    # 结构化检出物：按 active 模板的字段列表 + 表单中提交的 grade_<cat>_<idx>
    tpl = _MICRO_TEMPLATES[template_type]
    findings: list[dict] = []
    # 1) 镜检可见（半定量）— 仅记非阴性（非"-"且非空）
    for i, name in enumerate(tpl["microbe"]):
        g = (form.get(f"grade_microbe_{i}", "") or "").strip()
        if g and g not in ("-", "阴性"):
            findings.append({"cat": "microbe", "name": name, "grade": g})
    # 2) 寄生虫定性 — 全记（阴性也保留，表明已检）
    for i, name in enumerate(tpl["parasite"]):
        g = (form.get(f"grade_parasite_{i}", "") or "").strip()
        if g:
            findings.append({"cat": "parasite", "name": name, "grade": g})
    # 3) 病理可见 — 全记（医生选了才有意义）
    for i, pa in enumerate(tpl["pathology"]):
        g = (form.get(f"grade_pathology_{i}", "") or "").strip()
        if g:
            findings.append({"cat": "pathology", "name": pa["name"], "grade": g})
    # 4) 额外自定义项（任意行，归入 microbe 段）
    ex_names = form.getlist("finding_extra_name") if hasattr(form, "getlist") else []
    ex_grades = form.getlist("finding_extra_grade") if hasattr(form, "getlist") else []
    for i, n in enumerate(ex_names):
        n = (n or "").strip()
        g = ((ex_grades[i] if i < len(ex_grades) else "") or "").strip()
        if n and g:
            findings.append({"cat": "microbe", "name": n[:60], "grade": g[:20]})

    store = _get_op_store(request) or (pet.store if pet else "") or ""

    report = MicroscopyReport(
        exam_order_id=order_id,
        customer_id=cust_id, pet_id=pet_id, visit_id=order.visit_id,
        item_label=item_label, template_type=template_type, vet_name=vet_name,
        magnification=magnification, sample_site=sample_site,
        findings_json=json.dumps(findings, ensure_ascii=False),
        narrative=narrative, conclusion=conclusion, advice=advice,
        photos_json="[]", store=store,
        operator=request.session.get("admin_username", "admin"),
    )
    db.add(report)
    db.flush()  # 拿到 id

    # 处理照片上传
    photo_files = form.getlist("photos") if hasattr(form, "getlist") else []
    saved_paths = []
    photo_dir = Path("uploads") / "microscopy_photos" / str(report.id)
    photo_dir.mkdir(parents=True, exist_ok=True)
    for f in photo_files:
        if not hasattr(f, "filename") or not f.filename:
            continue
        ext = Path(f.filename).suffix.lower()
        if ext not in _MICRO_PHOTO_EXT_OK:
            continue
        data = await f.read()
        if not data or len(data) > _MICRO_PHOTO_MAX:
            continue
        name = f"p_{secrets.token_hex(6)}{ext}"
        (photo_dir / name).write_bytes(data)
        saved_paths.append(f"microscopy_photos/{report.id}/{name}")
    report.photos_json = json.dumps(saved_paths, ensure_ascii=False)
    db.commit()

    # 渲染 PDF（失败不阻塞，给出提示信息）
    from app.services.microscopy_pdf import generate_microscopy_pdf
    _, err = generate_microscopy_pdf(db, report.id)
    if err:
        return RedirectResponse(
            f"/admin/exam-orders/{order_id}?msg=报告已保存，PDF 生成失败：{err}",
            status_code=303,
        )
    return RedirectResponse(
        f"/admin/exam-orders/{order_id}?msg=显微镜报告已生成并上传",
        status_code=303,
    )


@app.post("/admin/microscopy-reports/{report_id}/regen-pdf")
async def admin_microscopy_regen(report_id: int, request: Request, db: Session = Depends(get_db)):
    """重新生成 PDF（首次生成失败 / weasyprint 后装时补救）"""
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    report = db.get(MicroscopyReport, report_id)
    if not report:
        raise HTTPException(404)
    admin_store = _get_admin_store(request)
    if admin_store and report.store and report.store != admin_store:
        raise HTTPException(403, "无权操作其他门店")
    from app.services.microscopy_pdf import generate_microscopy_pdf
    _, err = generate_microscopy_pdf(db, report.id)
    if err:
        return RedirectResponse(
            f"/admin/exam-orders/{report.exam_order_id}?msg=PDF 重新生成失败：{err}",
            status_code=303,
        )
    return RedirectResponse(
        f"/admin/exam-orders/{report.exam_order_id}?msg=PDF 已重新生成",
        status_code=303,
    )


@app.post("/admin/microscopy-reports/{report_id}/delete")
async def admin_microscopy_delete(report_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    report = db.get(MicroscopyReport, report_id)
    if not report:
        raise HTTPException(404)
    admin_store = _get_admin_store(request)
    if admin_store and report.store and report.store != admin_store:
        raise HTTPException(403, "无权操作其他门店")
    order_id = report.exam_order_id
    # 删 PDF（ExamReport 一并删）
    if report.exam_report_id:
        er = db.get(ExamReport, report.exam_report_id)
        if er:
            try:
                if er.file_path:
                    Path(er.file_path).unlink(missing_ok=True)
            except Exception:
                pass
            db.delete(er)
    # 删原始照片目录
    try:
        import shutil
        ph_dir = Path("uploads") / "microscopy_photos" / str(report.id)
        if ph_dir.exists():
            shutil.rmtree(ph_dir, ignore_errors=True)
    except Exception:
        pass
    db.delete(report)
    db.commit()
    return RedirectResponse(f"/admin/exam-orders/{order_id}?msg=显微镜报告已删除", status_code=303)


@app.get("/admin/exam-orders/{order_id}/qr.png")
async def admin_exam_order_qr(
    order_id: int, request: Request, db: Session = Depends(get_db),
):
    require_admin(request)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    # 老订单可能缺 upload_token（早期建单遗漏）→ 自动补一个，并立即写库
    if not order.upload_token or not order.token_expires_at or order.token_expires_at < datetime.utcnow():
        try:
            order.upload_token, order.token_expires_at = _exam_order_token(db)
            db.commit()
        except Exception as _e:
            logger.warning("[exam-qr] regen token failed: %s", _e)
    # 优先用 PUBLIC_BASE_URL 配置（生产域名），否则 fallback 到 request.base_url
    base = (settings.public_base_url or str(request.base_url)).rstrip("/")
    url  = f"{base}/exam-upload/{order.upload_token or ''}"
    # 尝试 1: qrcode + PIL（PNG）
    try:
        import qrcode, io as _io
        img  = qrcode.make(url)
        buf  = _io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/png")
    except Exception as e1:
        logger.warning("[exam-qr] PNG 生成失败，尝试 SVG: %s", e1)
    # 尝试 2: qrcode SVG（不需要 PIL）
    try:
        import qrcode
        import qrcode.image.svg as _svg
        import io as _io
        factory = _svg.SvgImage
        img = qrcode.make(url, image_factory=factory)
        buf = _io.BytesIO()
        img.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="image/svg+xml")
    except Exception as e2:
        logger.error("[exam-qr] SVG 也失败: %s", e2)
    # 尝试 3: 第三方公网 API（兜底，无需任何依赖）
    try:
        from urllib.parse import quote as _q
        import httpx
        api_url = f"https://api.qrserver.com/v1/create-qr-code/?size=240x240&data={_q(url, safe='')}"
        with httpx.Client(timeout=5.0, follow_redirects=True) as cli:
            r = cli.get(api_url)
            if r.status_code == 200 and r.content:
                return Response(content=r.content, media_type=r.headers.get('content-type', 'image/png'))
    except Exception as e3:
        logger.error("[exam-qr] 第三方 API 也失败: %s", e3)
    raise HTTPException(500, "二维码生成失败 — 请检查服务器是否安装 qrcode 和 Pillow 库")


@app.get("/admin/exam-reports/{report_id}/page-image")
async def admin_exam_report_page_image(
    report_id: int,
    request: Request,
    db: Session = Depends(get_db),
    page: int = Query(0, ge=0, description="PDF 页码（0-based）；图片类型忽略此参数"),
):
    """检查报告渲染成 PNG 用于嵌入打印页。

    - 图片类型：直接返回原文件
    - PDF：用 PyMuPDF 渲染指定页为 PNG，缓存到 data/exam_pages_cache/
    """
    require_admin(request)
    rpt = db.get(ExamReport, report_id)
    if not rpt:
        raise HTTPException(404)
    p = Path(rpt.file_path)
    if not p.exists():
        raise HTTPException(404)
    if rpt.file_type != "pdf":
        return FileResponse(str(p), media_type="image/jpeg")
    from app.services.pdf_render import render_pdf_page
    out = render_pdf_page(str(p), page, rpt.id)
    if not out:
        raise HTTPException(500, "PDF 渲染失败（服务器可能未装 PyMuPDF）")
    return FileResponse(str(out), media_type="image/png")


@app.get("/admin/exam-reports/{report_id}/file")
async def admin_exam_report_file(
    report_id: int,
    request: Request,
    db: Session = Depends(get_db),
    download: int = Query(0, description="1=强制下载附件，0=浏览器内联预览"),
):
    require_admin(request)
    rpt = db.get(ExamReport, report_id)
    if not rpt:
        raise HTTPException(404)
    p = Path(rpt.file_path)
    if not p.exists():
        raise HTTPException(404)
    media = "application/pdf" if rpt.file_type == "pdf" else "image/jpeg"
    if download:
        # Content-Disposition: attachment + filename → 浏览器触发下载
        return FileResponse(str(p), media_type=media, filename=rpt.original_name)
    # 内联预览：Content-Disposition: inline → PDF 浏览器内打开，图片直接显示
    from urllib.parse import quote as _q
    safe_name = _q(rpt.original_name or "report")
    return FileResponse(
        str(p),
        media_type=media,
        headers={"Content-Disposition": f'inline; filename*=UTF-8\'\'{safe_name}'},
    )


# ── 手机端上传（无需登录，Token 校验）────────────────────────────────────────

@app.get("/exam-upload/{token}", response_class=HTMLResponse)
async def exam_upload_mobile_page(token: str, request: Request, db: Session = Depends(get_db)):
    order = db.query(ExamOrder).filter(ExamOrder.upload_token == token).first()
    if not order or (order.token_expires_at and order.token_expires_at < datetime.utcnow()):
        return HTMLResponse("<h2 style='font-family:sans-serif;padding:2rem;'>链接已失效，请联系医院前台刷新二维码。</h2>", status_code=410)
    visit = db.get(Visit, order.visit_id)
    cust  = db.get(Customer, visit.customer_id) if visit and visit.customer_id else None
    pet   = db.get(Pet, visit.pet_id) if visit and visit.pet_id else None
    items = json.loads(order.items_json or "[]")
    return templates.TemplateResponse(request, "exam_upload_mobile.html", {
        "order": order, "visit": visit, "cust": cust, "pet": pet,
        "items": items, "token": token,
        "msg": request.query_params.get("msg", ""),
    })


@app.post("/exam-upload/{token}/upload")
async def exam_upload_mobile_post(
    token: str, request: Request, db: Session = Depends(get_db),
    file: UploadFile = File(...),
    item_label: str = Form(""),
):
    order = db.query(ExamOrder).filter(ExamOrder.upload_token == token).first()
    if not order or (order.token_expires_at and order.token_expires_at < datetime.utcnow()):
        raise HTTPException(410, "链接已失效")

    fname = file.filename or "photo.jpg"
    ext   = Path(fname).suffix.lower() or ".jpg"
    if ext not in _EXAM_REPORT_EXT_OK:
        raise HTTPException(400, f"不支持的文件类型 {ext}")
    data = await file.read()
    if len(data) > _EXAM_REPORT_MAX_BYTES:
        raise HTTPException(413, "文件超过 20MB 上限")
    ftype = "pdf" if ext == ".pdf" else "image"

    dest_dir = Path(settings.upload_dir) / "exam_reports" / str(order.id)
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / f"mob_{secrets.token_hex(8)}{ext}"
    dest.write_bytes(data)

    db.add(ExamReport(
        exam_order_id=order.id,
        file_path=str(dest),
        original_name=fname,
        file_type=ftype,
        item_label=(item_label or "").strip()[:120],
        uploaded_by="手机上传",
    ))
    order.status = "completed"
    order.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/exam-upload/{token}?msg=上传成功", status_code=303)


# ── 待领养动物 ────────────────────────────────────────────────────────────────

_ADOPTION_DIR = Path("data/adoption")
_ADOPTION_STATUS_ZH = {"available": "待领养", "adopted": "已领养", "paused": "暂停"}


# 领养文件白名单
_ADOPTION_IMG_EXT = {".jpg", ".jpeg", ".png", ".webp", ".heic"}
_ADOPTION_VIDEO_EXT = {".mp4", ".mov", ".webm"}
_ADOPTION_DOC_EXT = {".pdf"}
_ADOPTION_MAX_IMG = 10 * 1024 * 1024   # 10 MB
_ADOPTION_MAX_VIDEO = 50 * 1024 * 1024 # 50 MB


def _save_adoption_file(upload: UploadFile, prefix: str) -> str:
    _ADOPTION_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(upload.filename or "").suffix.lower()
    # 白名单校验
    allowed = _ADOPTION_IMG_EXT | _ADOPTION_VIDEO_EXT | _ADOPTION_DOC_EXT
    if ext not in allowed:
        raise HTTPException(400, f"不支持的文件类型 {ext}")
    content = upload.file.read()
    # 大小限制
    max_size = _ADOPTION_MAX_VIDEO if ext in _ADOPTION_VIDEO_EXT else _ADOPTION_MAX_IMG
    if len(content) > max_size:
        raise HTTPException(413, f"文件超过 {max_size // (1024*1024)}MB 上限")
    fname = f"{prefix}_{int(datetime.utcnow().timestamp()*1000)}{ext}"
    dest = _ADOPTION_DIR / fname
    dest.write_bytes(content)
    return str(dest)


# 公开 API（小程序用）
@app.get("/api/adoption")
async def api_adoption_list(db: Session = Depends(get_db)):
    pets = db.query(AdoptionPet).filter(AdoptionPet.status == "available").order_by(AdoptionPet.sort_order, AdoptionPet.id.desc()).all()
    result = []
    for p in pets:
        result.append({
            "id": p.id,
            "name": p.name,
            "species": p.species,
            "breed": p.breed,
            "age_estimate": p.age_estimate,
            "gender": p.gender,
            "personality": p.personality,
            "health_note": p.health_note,
            "requirements": p.requirements,
            "has_image1": bool(p.image1_path),
            "has_image2": bool(p.image2_path),
            "has_video": bool(p.video_path),
            "status": p.status,
        })
    return result


@app.get("/api/adoption/{pet_id}/image/{n}")
async def api_adoption_image(pet_id: int, n: int, db: Session = Depends(get_db)):
    pet = db.get(AdoptionPet, pet_id)
    if not pet:
        raise HTTPException(404)
    path = pet.image1_path if n == 1 else pet.image2_path
    if not path or not Path(path).exists():
        raise HTTPException(404)
    from fastapi.responses import FileResponse
    return FileResponse(path)


@app.get("/api/adoption/{pet_id}/video")
async def api_adoption_video(pet_id: int, db: Session = Depends(get_db)):
    pet = db.get(AdoptionPet, pet_id)
    if not pet or not pet.video_path or not Path(pet.video_path).exists():
        raise HTTPException(404)
    from fastapi.responses import FileResponse
    return FileResponse(pet.video_path)


# 后台管理
@app.get("/admin/adoption", response_class=HTMLResponse)
async def admin_adoption_list(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    pets = db.query(AdoptionPet).order_by(AdoptionPet.sort_order, AdoptionPet.id.desc()).all()
    return templates.TemplateResponse(request, "uk/adoption_list.html", {
        "pets": pets, "status_zh": _ADOPTION_STATUS_ZH,
    })


@app.get("/admin/adoption/new", response_class=HTMLResponse)
async def admin_adoption_new_form(request: Request):
    require_admin(request)
    return templates.TemplateResponse(request, "uk/adoption_form.html", {
        "pet": None, "status_zh": _ADOPTION_STATUS_ZH,
    })


@app.post("/admin/adoption/new")
async def admin_adoption_create(request: Request, db: Session = Depends(get_db),
    name: str = Form(""), species: str = Form("cat"), breed: str = Form(""),
    age_estimate: str = Form(""), gender: str = Form("unknown"),
    personality: str = Form(""), health_note: str = Form(""), requirements: str = Form(""),
    sort_order: int = Form(0), csrf_token: str = Form(""),
    image1: UploadFile = File(None), image2: UploadFile = File(None),
    video: UploadFile = File(None),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    pet = AdoptionPet(
        name=name.strip(), species=species, breed=breed.strip(),
        age_estimate=age_estimate.strip(), gender=gender,
        personality=personality.strip(), health_note=health_note.strip(),
        requirements=requirements.strip(), sort_order=sort_order,
    )
    if image1 and image1.filename:
        pet.image1_path = _save_adoption_file(image1, f"img1_{name}")
    if image2 and image2.filename:
        pet.image2_path = _save_adoption_file(image2, f"img2_{name}")
    if video and video.filename:
        pet.video_path = _save_adoption_file(video, f"video_{name}")
    db.add(pet)
    db.commit()
    return RedirectResponse("/admin/adoption?msg=已添加", status_code=303)


@app.get("/admin/adoption/{pet_id}", response_class=HTMLResponse)
async def admin_adoption_detail(pet_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    pet = db.get(AdoptionPet, pet_id)
    if not pet:
        raise HTTPException(404)
    return templates.TemplateResponse(request, "uk/adoption_form.html", {
        "pet": pet, "status_zh": _ADOPTION_STATUS_ZH,
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/adoption/{pet_id}/edit")
async def admin_adoption_edit(pet_id: int, request: Request, db: Session = Depends(get_db),
    name: str = Form(""), species: str = Form("cat"), breed: str = Form(""),
    age_estimate: str = Form(""), gender: str = Form("unknown"),
    personality: str = Form(""), health_note: str = Form(""), requirements: str = Form(""),
    sort_order: int = Form(0), status: str = Form("available"), csrf_token: str = Form(""),
    image1: UploadFile = File(None), image2: UploadFile = File(None),
    video: UploadFile = File(None),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    pet = db.get(AdoptionPet, pet_id)
    if not pet:
        raise HTTPException(404)
    pet.name = name.strip(); pet.species = species; pet.breed = breed.strip()
    pet.age_estimate = age_estimate.strip(); pet.gender = gender
    pet.personality = personality.strip(); pet.health_note = health_note.strip()
    pet.requirements = requirements.strip(); pet.sort_order = sort_order
    pet.status = status
    if image1 and image1.filename:
        pet.image1_path = _save_adoption_file(image1, f"img1_{pet_id}")
    if image2 and image2.filename:
        pet.image2_path = _save_adoption_file(image2, f"img2_{pet_id}")
    if video and video.filename:
        pet.video_path = _save_adoption_file(video, f"video_{pet_id}")
    pet.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/admin/adoption/{pet_id}?msg=已保存", status_code=303)


@app.post("/admin/adoption/{pet_id}/adopt")
async def admin_adoption_mark_adopted(pet_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""), agreement: UploadFile = File(None),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    pet = db.get(AdoptionPet, pet_id)
    if not pet:
        raise HTTPException(404)
    pet.status = "adopted"
    if agreement and agreement.filename:
        pet.adoption_agreement_path = _save_adoption_file(agreement, f"agreement_{pet_id}")
    pet.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/admin/adoption/{pet_id}?msg=已标记为已领养", status_code=303)


@app.post("/admin/adoption/{pet_id}/delete")
async def admin_adoption_delete(pet_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    pet = db.get(AdoptionPet, pet_id)
    if not pet:
        raise HTTPException(404)
    db.delete(pet)
    db.commit()
    return RedirectResponse("/admin/adoption?msg=已删除", status_code=303)


@app.get("/admin/adoption/{pet_id}/agreement")
async def admin_adoption_agreement(pet_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    pet = db.get(AdoptionPet, pet_id)
    if not pet or not pet.adoption_agreement_path or not Path(pet.adoption_agreement_path).exists():
        raise HTTPException(404)
    from fastapi.responses import FileResponse
    return FileResponse(pet.adoption_agreement_path)


# ── 日历视图 ──────────────────────────────────────────────────────────────────

@app.get("/admin/calendar", response_class=HTMLResponse)
async def admin_calendar_page(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    return templates.TemplateResponse(request, "uk/admin_calendar.html", {  # B 补 - UK 重写
        "csrf_token": _get_csrf_token(request),
        "admin_store": _get_admin_store(request),
    })


@app.get("/api/calendar/events")
async def api_calendar_events(
    request: Request,
    start: str = Query(""),
    end:   str = Query(""),
    store: str = Query(""),
    db: Session = Depends(get_db),
):
    """返回指定日期范围内的预约 + 全天封锁日程（JSON）。已取消的预约不显示在日历上。"""
    require_admin(request)
    # 限店员工强制使用自己门店，忽略前端传入的 store
    admin_store = _get_admin_store(request)
    if admin_store:
        store = admin_store
    q = db.query(Appointment).filter(Appointment.status != AppointmentStatus.cancelled.value)
    if start:
        q = q.filter(Appointment.appointment_date >= start)
    if end:
        q = q.filter(Appointment.appointment_date <= end)
    if store:
        full_store = _STORE_SHORT_TO_FULL.get(store, store)
        q = q.filter(Appointment.store == full_store)
    appts = q.order_by(Appointment.appointment_date, Appointment.appointment_time).all()

    pet_map: dict = {}
    pet_ids = [a.pet_id for a in appts if a.pet_id]
    if pet_ids:
        pets_q = db.query(Pet).filter(Pet.id.in_(pet_ids)).all()
        pet_map = {p.id: p for p in pets_q}

    appt_list = []
    for a in appts:
        pet = pet_map.get(a.pet_id) if a.pet_id else None
        species = pet.species if pet else ("cat" if a.category == "tnr" else "")
        store_short = _STORE_FULL_TO_SHORT.get(a.store or "", a.store or "")
        appt_list.append({
            "id":             a.id,
            "type":           "appointment",
            "category":       a.category or "",
            "service_name":   a.service_name or "",
            "status":         a.status or "",
            "customer_name":  a.customer_name or "",
            "phone":          a.phone or "",
            "pet_name":       a.pet_name or "",
            "pet_gender":     a.pet_gender or "",
            "pet_species":    species,
            "store":          a.store or "",
            "store_short":    store_short,
            "date":           a.appointment_date or "",
            "time":           a.appointment_time or "",
            "duration":       a.duration_minutes or 30,
            "notes":          a.notes or "",
            "related_app_id": a.related_application_id,
            "created_at":     a.created_at.strftime("%m-%d %H:%M") if a.created_at else "",
            "customer_id":    a.customer_id,
            "pet_id":         a.pet_id,
        })

    bq = db.query(CalendarBlock)
    if start:
        bq = bq.filter(CalendarBlock.block_date >= start)
    if end:
        bq = bq.filter(CalendarBlock.block_date <= end)
    if store:
        bq = bq.filter((CalendarBlock.store == store) | (CalendarBlock.store == ""))
    blocks = bq.order_by(CalendarBlock.block_date).all()
    block_list = [{
        "id":    b.id,
        "type":  "block",
        "title": b.title,
        "date":  b.block_date,
        "store": b.store,
        "notes": b.notes,
    } for b in blocks]

    return {"appointments": appt_list, "blocks": block_list}


@app.post("/api/calendar/blocks/create")
async def api_calendar_block_create(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    # 限店员工只能创建本店封锁
    admin_store = _get_admin_store(request)
    block_store = str(body.get("store", "")).strip()[:40]
    if admin_store:
        block_store = admin_store
    block = CalendarBlock(
        title=str(body.get("title", "")).strip()[:200] or "全天封锁",
        block_date=str(body.get("date", "")).strip()[:20],
        store=block_store,
        notes=str(body.get("notes", "")).strip()[:500],
        created_by=request.session.get("admin_username", ""),
    )
    db.add(block)
    db.commit()
    db.refresh(block)
    return {"ok": True, "id": block.id}


@app.post("/api/calendar/blocks/{block_id}/delete")
async def api_calendar_block_delete(block_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    block = db.get(CalendarBlock, block_id)
    if not block:
        return {"ok": False, "error": "不存在"}
    # 限店员工只能删本店封锁
    admin_store = _get_admin_store(request)
    if admin_store and block.store and block.store != admin_store:
        return {"ok": False, "error": "无权删除其他门店的封锁"}
    db.delete(block)
    db.commit()
    return {"ok": True}


@app.post("/api/calendar/appt/{appt_id}/status")
async def api_calendar_appt_status(appt_id: int, request: Request, db: Session = Depends(get_db)):
    """日历弹窗 AJAX 状态更新，复用现有状态转换逻辑，返回 JSON。"""
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    status = str(body.get("status", "")).strip()
    cancel_reason = str(body.get("cancel_reason", "")).strip()[:300]
    if status not in _ALLOWED_APPOINTMENT_STATUSES:
        return {"ok": False, "error": "无效状态"}
    row = db.get(Appointment, appt_id)
    if not row:
        return {"ok": False, "error": "预约不存在"}
    # 限店员工只能改本店预约
    admin_store = _get_admin_store(request)
    if admin_store:
        full_store = _STORE_SHORT_TO_FULL.get(admin_store, admin_store)
        if row.store and row.store != full_store:
            return {"ok": False, "error": "无权操作其他门店的预约"}
    old_status = row.status
    row.status = status
    row.updated_at = datetime.utcnow()
    if cancel_reason:
        existing_notes = (row.notes or "").strip()
        row.notes = (existing_notes + f"\n[取消原因] {cancel_reason}").strip()
    _audit(db, request, "appointment_status_update",
           application_id=row.related_application_id,
           detail={"appointment_id": row.id, "old_status": old_status, "status": status})
    if row.related_application_id:
        app_row = db.get(Application, row.related_application_id)
        if app_row:
            if status == AppointmentStatus.confirmed.value and app_row.status in (
                ApplicationStatus.approved.value, ApplicationStatus.pre_approved.value,
            ):
                app_row.status = ApplicationStatus.scheduled.value
                app_row.appointment_at = row.appointment_date
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.arrived.value and app_row.status in (
                ApplicationStatus.scheduled.value, ApplicationStatus.approved.value,
            ):
                app_row.status = ApplicationStatus.arrived_verified.value
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.cancelled.value and app_row.status == ApplicationStatus.scheduled.value:
                app_row.status = ApplicationStatus.approved.value
                app_row.updated_at = datetime.utcnow()
            elif status == AppointmentStatus.no_show.value:
                app_row.status = ApplicationStatus.no_show.value
                app_row.updated_at = datetime.utcnow()
    db.commit()
    openid = (row.wechat_openid or "").strip()
    if openid and status in (AppointmentStatus.confirmed.value, AppointmentStatus.cancelled.value):
        status_label = "已确认，请按约定时间到院" if status == AppointmentStatus.confirmed.value else "已取消"
        push_appointment_status(
            db, appointment_id=row.id, openid=openid,
            status_text=status_label, service_name=row.service_name or "",
            store=row.store or "", appointment_date=row.appointment_date or "",
            appointment_time=row.appointment_time or "",
            phone=row.phone or "", customer_name=row.customer_name or "",
            note=cancel_reason or status_label,
        )
    return {"ok": True}


@app.post("/api/calendar/appt/{appt_id}/reschedule")
async def api_calendar_appt_reschedule(appt_id: int, request: Request, db: Session = Depends(get_db)):
    """日历弹窗 AJAX 改约。"""
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    new_date = str(body.get("date", "")).strip()[:20]
    new_time = str(body.get("time", "")).strip()[:10]
    if not new_date or not new_time:
        return {"ok": False, "error": "请填写新日期和时间"}
    row = db.get(Appointment, appt_id)
    if not row:
        return {"ok": False, "error": "预约不存在"}
    # 限店员工只能改本店预约
    admin_store = _get_admin_store(request)
    if admin_store:
        full_store = _STORE_SHORT_TO_FULL.get(admin_store, admin_store)
        if row.store and row.store != full_store:
            return {"ok": False, "error": "无权操作其他门店的预约"}
    # 已完成 / 已取消 / 爽约 状态不允许改约
    if row.status in (AppointmentStatus.completed.value, AppointmentStatus.cancelled.value, AppointmentStatus.no_show.value):
        return {"ok": False, "error": f"当前状态（{_APPOINTMENT_STATUS_LABELS.get(row.status, row.status)}）不允许改约"}
    old_date, old_time = row.appointment_date, row.appointment_time
    row.appointment_date = new_date
    row.appointment_time = new_time
    row.updated_at = datetime.utcnow()
    _audit(db, request, "appointment_reschedule",
           application_id=row.related_application_id,
           detail={"appointment_id": row.id, "old_date": old_date, "old_time": old_time,
                   "new_date": new_date, "new_time": new_time})
    db.commit()
    return {"ok": True}



# ── 驱虫记录 / 体重记录 / 医疗文书 ────────────────────────────────

_DEWORM_DIR = Path("data/dewormings")
_MEDDOC_DIR = Path("data/medical_docs")


_DEW_TYPE_ZH = {"external": "体外驱虫", "internal": "体内驱虫", "combo": "体内外同驱"}


def _split_dew_notes(notes: str) -> tuple[str, str]:
    """从 notes 里抽出『批号：xxx』，返回 (清干净的 notes, batch_no)"""
    if not notes:
        return "", ""
    import re
    m = re.search(r"批号[:：]\s*([^\n]+)", notes)
    if m:
        batch = m.group(1).strip()
        clean = re.sub(r"批号[:：][^\n]+\n?", "", notes).strip()
        return clean, batch
    return notes, ""


@app.get("/admin/dewormings/create", response_class=HTMLResponse)
async def page_admin_deworming_create(
    request: Request, db: Session = Depends(get_db),
    customer_id: int = Query(0), pet_id: int = Query(0),
):
    require_admin(request)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    vets_q = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vets = [v[0] for v in vets_q]
    dew_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category == "antiparasitic",
        InventoryItem.is_active == True,
    ).order_by(InventoryItem.name).all()
    from datetime import date as _d, timedelta as _td
    today = _d.today().isoformat()
    next_due = (_d.today() + _td(days=30)).isoformat()
    history = []
    if pet_id:
        history = db.query(DewormingRecord).filter(DewormingRecord.pet_id == pet_id)\
            .order_by(DewormingRecord.deworm_date.desc(), DewormingRecord.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/deworming.html", {  # B 补 - UK 重写
        "mode": "create", "rec": None,
        "cust": cust, "pet": pet, "vets": vets, "dew_items": dew_items,
        "today": today, "next_due_default": next_due,
        "rec_batch_no": "", "rec_clean_notes": "", "rec_charge_amount": 0.0,
        "deworm_history": history,
        "locked": False, "lock_reason": "", "paid_amount": 0.0,
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/dewormings/{rec_id}", response_class=HTMLResponse)
async def page_admin_deworming_detail(rec_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    rec = db.get(DewormingRecord, rec_id)
    if not rec:
        raise HTTPException(404, "驱虫记录不存在")
    cust = db.get(Customer, rec.customer_id) if rec.customer_id else None
    pet = db.get(Pet, rec.pet_id) if rec.pet_id else None
    vets_q = db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
        Staff.position.ilike("%医%")
    ).all()
    vets = [v[0] for v in vets_q]
    dew_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category == "antiparasitic",
        InventoryItem.is_active == True,
    ).order_by(InventoryItem.name).all()
    locked, lock_reason = _is_deworming_locked(db, rec)
    paid_amount = _doc_paid_amount(db, "deworming", rec_id) if locked else 0.0
    clean_notes, batch_no = _split_dew_notes(rec.notes or "")
    charge = 0.0
    if rec.invoice_id:
        inv = db.get(Invoice, rec.invoice_id)
        if inv:
            charge = float(inv.total_amount or 0)
    history = []
    if rec.pet_id:
        history = db.query(DewormingRecord).filter(
            DewormingRecord.pet_id == rec.pet_id,
            DewormingRecord.id != rec_id,
        ).order_by(DewormingRecord.deworm_date.desc(), DewormingRecord.id.desc()).limit(10).all()
    return templates.TemplateResponse(request, "uk/deworming.html", {  # B 补 - UK 重写
        "mode": "edit", "rec": rec,
        "cust": cust, "pet": pet, "vets": vets, "dew_items": dew_items,
        "today": rec.deworm_date, "next_due_default": rec.next_due_date,
        "rec_batch_no": batch_no, "rec_clean_notes": clean_notes, "rec_charge_amount": charge,
        "deworm_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/dewormings/{rec_id}/edit")
async def admin_deworming_edit(
    rec_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    deworm_date: str = Form(""),
    deworm_type: str = Form("external"),
    product_name: str = Form(""),
    weight_kg: float = Form(0.0),
    dose: str = Form(""),
    next_due_date: str = Form(""),
    notes: str = Form(""),
    vet_name: str = Form(""),
    batch_no: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(DewormingRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    locked, reason = _is_deworming_locked(db, rec)
    if locked:
        raise HTTPException(400, f"驱虫记录已锁定（{reason}），不可修改。请「复制为新单」或「作废」。")
    notes_full = notes.strip()
    if batch_no.strip():
        notes_full = (notes_full + "\n批号：" + batch_no.strip()).strip()
    rec.deworm_date = deworm_date.strip()[:20]
    rec.deworm_type = (deworm_type or "external")[:40]
    rec.product_name = product_name.strip()[:120]
    rec.weight_kg = weight_kg or 0.0
    rec.dose = dose.strip()[:80]
    rec.next_due_date = next_due_date.strip()[:20]
    rec.vet_name = vet_name.strip()[:80]
    rec.notes = notes_full
    rec.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/admin/dewormings/{rec_id}?msg=已保存", status_code=303)


@app.post("/admin/dewormings/create")
async def admin_deworming_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
    deworm_date: str = Form(""),
    deworm_type: str = Form("external"),
    product_name: str = Form(""),
    weight_kg: float = Form(0.0),
    dose: str = Form(""),
    next_due_date: str = Form(""),
    notes: str = Form(""),
    vet_name: str = Form(""),
    inventory_item_id: int = Form(0),
    batch_no: str = Form(""),
    charge_amount: float = Form(0.0),
    next_url: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    # batch_no 临时存到 notes 末尾（DewormingRecord 模型没该字段，避免迁移）
    notes_full = notes.strip()
    if batch_no.strip():
        notes_full = (notes_full + "\n批号：" + batch_no.strip()).strip()
    rec = DewormingRecord(
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        deworm_date=deworm_date.strip()[:20],
        deworm_type=deworm_type.strip()[:40] or "external",
        product_name=product_name.strip()[:120],
        weight_kg=weight_kg or 0.0,
        dose=dose.strip()[:80],
        next_due_date=next_due_date.strip()[:20],
        vet_name=vet_name.strip()[:80],
        notes=notes_full,
        created_by=request.session.get("admin_username", ""),
    )
    db.add(rec)
    db.flush()

    admin_name = request.session.get("admin_username", "")
    msg = "驱虫已录入"

    # 库存出库 1 单位
    if inventory_item_id:
        try:
            _deduct_inventory(db, inventory_item_id, 1.0, "deworming", rec.id, admin_name,
                              note=f"{rec.product_name or '驱虫'} 使用出库")
        except Exception as _e:
            logger.warning("[deworming] inventory deduct failed: %s", _e)

    # 自动生成收费单（不收费=金额≤0）
    if charge_amount and charge_amount > 0:
        inv = Invoice(
            invoice_no      = _gen_invoice_no(db),
            customer_id     = customer_id or None,
            pet_id          = pet_id or None,
            invoice_date    = rec.deworm_date or datetime.now().strftime("%Y-%m-%d"),
            subtotal        = charge_amount,
            discount_amount = 0.0,
            total_amount    = charge_amount,
            payment_status  = "unpaid",
            notes           = f"驱虫 #{rec.id}",
            store           = _resolve_invoice_store(db, pet_id=pet_id, customer_id=customer_id, fallback=_get_op_store(request)),
            created_by      = admin_name,
        )
        db.add(inv)
        db.flush()
        db.add(InvoiceItem(
            invoice_id  = inv.id,
            ref_type    = "deworming",
            ref_id      = rec.id,
            description = rec.product_name or "驱虫",
            quantity    = 1.0,
            unit_price  = charge_amount,
            subtotal    = charge_amount,
        ))
        # 反向挂到驱虫记录上（支持锁定/退款路径）
        rec.invoice_id = inv.id
        msg += f"，收费单 ¥{charge_amount:.2f} 已生成待收款"

    db.commit()
    # 跳到驱虫详情页（与疫苗一致），用户可继续编辑/锁定/作废
    return RedirectResponse(
        _safe_next(next_url, f"/admin/dewormings/{rec.id}?msg={msg}"),
        status_code=303,
    )


@app.post("/admin/dewormings/{rec_id}/delete")
async def admin_deworming_delete(
    rec_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(DewormingRecord, rec_id)
    if rec:
        locked, reason = _is_deworming_locked(db, rec)
        if locked:
            raise HTTPException(400, f"驱虫记录已锁定（{reason}），不可删除。请使用「作废」。")
        db.delete(rec)
        db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=vaccines&msg=驱虫记录已删除", status_code=303)


@app.post("/admin/dewormings/{rec_id}/void")
async def admin_deworming_void(rec_id: int, request: Request, db: Session = Depends(get_db),
                                 csrf_token: str = Form(""), void_reason: str = Form(""),
                                 customer_id: int = Form(0), pet_id: int = Form(0)):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(DewormingRecord, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.status == "voided":
        return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=vaccines&msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    rec.status = "voided"
    rec.voided_by = operator
    rec.voided_at = datetime.utcnow()
    rec.void_reason = (void_reason or "")[:200]
    _audit_doc_action(db, "deworming", rec_id, "void", operator, void_reason)
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=vaccines&msg=驱虫记录已作废", status_code=303)


@app.post("/admin/dewormings/{rec_id}/copy-as-new")
async def admin_deworming_copy_as_new(rec_id: int, request: Request, db: Session = Depends(get_db),
                                        csrf_token: str = Form(""),
                                        customer_id: int = Form(0), pet_id: int = Form(0)):
    require_admin(request)
    _require_csrf(request, csrf_token)
    src = db.get(DewormingRecord, rec_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_rec = DewormingRecord(
        customer_id=src.customer_id, pet_id=src.pet_id,
        deworm_date=datetime.utcnow().strftime("%Y-%m-%d"),
        deworm_type=src.deworm_type,
        product_name=src.product_name,
        weight_kg=src.weight_kg,
        dose=src.dose,
        next_due_date="",
        vet_name=src.vet_name,
        notes=src.notes,
        status="active",
        created_by=operator,
    )
    db.add(new_rec)
    _audit_doc_action(db, "deworming", 0, "copy_from", operator, extra=f"src={rec_id}")
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=vaccines&msg=已复制为新驱虫记录", status_code=303)


# ═════════════════════════════════════════════════════════════════════
# 美容单 (GroomingOrder) — 独立单据，与驱虫/疫苗一致体验
# ═════════════════════════════════════════════════════════════════════
_GROOMING_PET_SIZES = {"small": "小型", "medium": "中型", "large": "大型", "xlarge": "巨型"}
_GROOMING_COAT_LENGTHS = {"short": "短毛", "medium": "中毛", "long": "长毛"}


def _parse_grooming_services(form) -> list:
    """解析美容服务清单。字段：service_name[]/item_id[]/qty[]/price[]/note[]"""
    names = form.getlist("service_name[]")
    items = []
    for i, name in enumerate(names):
        name = str(name or "").strip()
        if not name:
            continue
        def _g(k, d=""):
            v = form.getlist(f"{k}[]")
            return v[i] if i < len(v) else d
        try:
            qty = float(_g("qty") or 1)
        except Exception:
            qty = 1.0
        try:
            price = float(_g("price") or 0)
        except Exception:
            price = 0.0
        try:
            item_id = int(_g("item_id") or 0)
        except Exception:
            item_id = 0
        items.append({
            "name": name[:120],
            "item_id": item_id or None,
            "qty": qty,
            "price": price,
            "subtotal": round(qty * price, 2),
            "notes": (_g("note") or "")[:200],
        })
    return items


def _query_grooming_items(db: Session, request: Request):
    """美容服务项库存（category=grooming）。"""
    q = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.category == "grooming",
        InventoryItem.is_active == True,
    ).order_by(InventoryItem.subcategory, InventoryItem.name)
    return q.all()


@app.get("/admin/grooming-orders/create", response_class=HTMLResponse)
async def page_admin_grooming_create(
    request: Request, db: Session = Depends(get_db),
    customer_id: int = Query(0), pet_id: int = Query(0), appointment_id: int = Query(0),
):
    require_admin(request)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    appt = db.get(Appointment, appointment_id) if appointment_id else None
    if appt and not cust:
        cust = db.get(Customer, appt.customer_id) if appt.customer_id else None
        pet = db.get(Pet, appt.pet_id) if appt.pet_id else pet
    groomers = [s[0] for s in db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
    ).all()]
    # 助理候选 = 所有在职员工（医生 / 美容师 / 助理 都可担任助理）
    assistants = groomers[:]
    today = datetime.utcnow().strftime("%Y-%m-%d")
    history = []
    if pet:
        history = db.query(GroomingOrder).filter(GroomingOrder.pet_id == pet.id)\
            .order_by(GroomingOrder.id.desc()).limit(10).all()
    groom_items = _query_grooming_items(db, request)
    return templates.TemplateResponse(request, "uk/grooming.html", {  # B 补 - UK 重写
        "mode": "create", "rec": None,
        "cust": cust, "pet": pet, "appt": appt,
        "groomers": groomers, "assistants": assistants, "groom_items": groom_items,
        "pet_sizes": _GROOMING_PET_SIZES, "coat_lengths": _GROOMING_COAT_LENGTHS,
        "today": today,
        "grooming_history": history,
        "locked": False, "lock_reason": "", "paid_amount": 0.0,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/grooming-orders/create")
async def admin_grooming_create(request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    customer_id = int(form.get("customer_id", 0) or 0)
    pet_id = int(form.get("pet_id", 0) or 0)
    appointment_id = int(form.get("appointment_id", 0) or 0)
    services = _parse_grooming_services(form)
    total = round(sum(s["subtotal"] for s in services), 2)
    charge_amount = float(form.get("charge_amount", 0) or 0)
    operator = request.session.get("admin_username", "admin")

    pet = db.get(Pet, pet_id) if pet_id else None
    store = (pet.store if pet else "") or _get_admin_store(request) or ""

    rec = GroomingOrder(
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        appointment_id=appointment_id or None,
        groom_date=str(form.get("groom_date", "")).strip()[:20],
        groomer_name=str(form.get("groomer_name", "")).strip()[:80],
        assistant_name=str(form.get("assistant_name", "")).strip()[:80],
        services_json=json.dumps(services, ensure_ascii=False),
        total_amount=total,
        skin_condition=str(form.get("skin_condition", "")).strip()[:200],
        behavior_note=str(form.get("behavior_note", "")).strip()[:200],
        store=store,
        notes=str(form.get("notes", "")).strip(),
        created_by=operator,
    )
    db.add(rec)
    db.flush()

    # 关联库存品目自动扣库存（仅非服务类）
    for s in services:
        iid = s.get("item_id")
        qty = float(s.get("qty") or 0)
        if iid and qty > 0:
            inv = db.get(InventoryItem, iid)
            if inv and not inv.is_service:
                _deduct_inventory(db, iid, qty, "grooming", rec.id, operator,
                                  note=f"美容#{rec.id} {s.get('name','')}")

    msg = "美容单已开具"
    # 自动生成收费单（金额 > 0 时）
    actual_charge = charge_amount if charge_amount > 0 else total
    if actual_charge > 0:
        inv = Invoice(
            invoice_no=_gen_invoice_no(db),
            customer_id=customer_id or None,
            pet_id=pet_id or None,
            invoice_date=rec.groom_date or datetime.now().strftime("%Y-%m-%d"),
            subtotal=actual_charge, discount_amount=0.0, total_amount=actual_charge,
            payment_status="unpaid",
            notes=f"美容 #{rec.id}",
            store=_resolve_invoice_store(db, pet_id=pet_id, customer_id=customer_id, fallback=_get_op_store(request)),
            created_by=operator,
        )
        db.add(inv)
        db.flush()
        # 按每项服务拆明细，让客户看清楚 — 而不是合并成「美容服务（3 项）」
        # 如果有 charge_amount 覆盖总价 (custom_charge != sum(subtotals))，差额作为「折扣 / 加价」单独成行
        items_total = round(sum(float(s.get("subtotal") or 0) for s in services), 2)
        for s in services:
            sub = round(float(s.get("subtotal") or 0), 2)
            if sub <= 0:
                continue
            qty = float(s.get("qty") or 1)
            price = float(s.get("price") or 0)
            name = (s.get("name") or "美容服务").strip()
            note = (s.get("notes") or "").strip()
            db.add(InvoiceItem(
                invoice_id=inv.id, ref_type="grooming", ref_id=rec.id,
                description=f"[美容#{rec.id}] {name}" + (f" · {note}" if note else ""),
                quantity=qty, unit_price=price, subtotal=sub,
            ))
        # 价格覆盖差额（手动改总价时）
        if abs(actual_charge - items_total) > 0.005:
            diff = round(actual_charge - items_total, 2)
            db.add(InvoiceItem(
                invoice_id=inv.id, ref_type="grooming", ref_id=rec.id,
                description=("[美容#" + str(rec.id) + "] 整单折扣 / 减免") if diff < 0 else ("[美容#" + str(rec.id) + "] 整单加价"),
                quantity=1.0, unit_price=diff, subtotal=diff,
            ))
        rec.invoice_id = inv.id
        msg += f"，收费单 ¥{actual_charge:.2f} 已生成待收款"

    db.commit()
    return RedirectResponse(f"/admin/grooming-orders/{rec.id}?msg={msg}", status_code=303)


@app.get("/admin/grooming-orders/{rec_id}", response_class=HTMLResponse)
async def page_admin_grooming_detail(rec_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404, "美容单不存在")
    cust = db.get(Customer, rec.customer_id) if rec.customer_id else None
    pet = db.get(Pet, rec.pet_id) if rec.pet_id else None
    appt = db.get(Appointment, rec.appointment_id) if rec.appointment_id else None
    groomers = [s[0] for s in db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
    ).all()]
    assistants = groomers[:]
    locked, lock_reason = _is_grooming_locked(db, rec)
    paid_amount = _doc_paid_amount(db, "grooming", rec_id) if locked else 0.0
    try:
        services = json.loads(rec.services_json or "[]")
    except Exception:
        services = []
    history = []
    if rec.pet_id:
        history = db.query(GroomingOrder).filter(
            GroomingOrder.pet_id == rec.pet_id, GroomingOrder.id != rec_id,
        ).order_by(GroomingOrder.id.desc()).limit(10).all()
    groom_items = _query_grooming_items(db, request)
    return templates.TemplateResponse(request, "uk/grooming.html", {  # B 补 - UK 重写
        "mode": "edit", "rec": rec, "rec_services": services,
        "cust": cust, "pet": pet, "appt": appt,
        "groomers": groomers, "assistants": assistants, "groom_items": groom_items,
        "pet_sizes": _GROOMING_PET_SIZES, "coat_lengths": _GROOMING_COAT_LENGTHS,
        "today": rec.groom_date,
        "grooming_history": history,
        "locked": locked, "lock_reason": lock_reason, "paid_amount": paid_amount,
        "csrf_token": _get_csrf_token(request),
        "msg": request.query_params.get("msg"),
    })


@app.post("/admin/grooming-orders/{rec_id}/edit")
async def admin_grooming_edit(rec_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    locked, reason = _is_grooming_locked(db, rec)
    if locked:
        raise HTTPException(400, f"美容单已锁定（{reason}），不可修改。请「复制为新单」或「作废」。")
    services = _parse_grooming_services(form)
    rec.groom_date = str(form.get("groom_date", "")).strip()[:20]
    rec.groomer_name = str(form.get("groomer_name", "")).strip()[:80]
    rec.assistant_name = str(form.get("assistant_name", "")).strip()[:80]
    rec.services_json = json.dumps(services, ensure_ascii=False)
    rec.total_amount = round(sum(s["subtotal"] for s in services), 2)
    rec.skin_condition = str(form.get("skin_condition", "")).strip()[:200]
    rec.behavior_note = str(form.get("behavior_note", "")).strip()[:200]
    rec.notes = str(form.get("notes", "")).strip()
    rec.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=已保存", status_code=303)


@app.post("/admin/grooming-orders/{rec_id}/delete")
async def admin_grooming_delete(rec_id: int, request: Request, db: Session = Depends(get_db),
                                 csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        return RedirectResponse("/admin/customers", status_code=303)
    locked, reason = _is_grooming_locked(db, rec)
    if locked:
        raise HTTPException(400, f"美容单已锁定（{reason}），不可删除。请使用「作废」。")
    cust_id = rec.customer_id
    pet_id = rec.pet_id
    db.delete(rec)
    db.commit()
    if cust_id:
        return RedirectResponse(f"/admin/customers/{cust_id}?pet_id={pet_id}&msg=美容单已删除", status_code=303)
    return RedirectResponse("/admin/customers?msg=美容单已删除", status_code=303)


@app.post("/admin/grooming-orders/{rec_id}/void")
async def admin_grooming_void(rec_id: int, request: Request, db: Session = Depends(get_db),
                                csrf_token: str = Form(""), void_reason: str = Form(""),
                                refund_to_wallet: str = Form(""), refund_amount: float = Form(0.0)):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.status == "voided":
        return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=该单已作废", status_code=303)
    operator = request.session.get("admin_username", "admin")
    # 回库（仅非服务类品目）
    try:
        services = json.loads(rec.services_json or "[]")
        for s in services:
            iid = s.get("item_id")
            qty = float(s.get("qty") or 0)
            if iid and qty > 0:
                inv = db.get(InventoryItem, iid)
                if inv and not inv.is_service:
                    _restore_inventory(db, iid, qty, "grooming_void", rec_id, operator,
                                       note=f"作废美容#{rec_id} 回库")
    except Exception:
        pass
    rec.status = "voided"
    rec.voided_by = operator
    rec.voided_at = datetime.utcnow()
    rec.void_reason = (void_reason or "")[:200]
    refund_msg = ""
    if refund_to_wallet in ("1", "true", "on") and rec.customer_id and refund_amount > 0:
        tx = _refund_to_wallet(
            db, rec.customer_id, float(refund_amount), operator,
            note=f"作废美容单#{rec_id} 退款 · {void_reason}"[:500],
        )
        if tx:
            refund_msg = f" · ¥{refund_amount:.2f} 已退入客户钱包"
            _audit_doc_action(db, "grooming", rec_id, "refund_to_wallet",
                              operator, extra=f"amount={refund_amount}")
    _audit_doc_action(db, "grooming", rec_id, "void", operator, void_reason)
    db.commit()
    return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=已作废{refund_msg}", status_code=303)


@app.post("/admin/grooming-orders/{rec_id}/copy-as-new")
async def admin_grooming_copy_as_new(rec_id: int, request: Request, db: Session = Depends(get_db),
                                       csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    src = db.get(GroomingOrder, rec_id)
    if not src:
        raise HTTPException(404)
    operator = request.session.get("admin_username", "admin")
    new_rec = GroomingOrder(
        customer_id=src.customer_id, pet_id=src.pet_id,
        appointment_id=None,
        groom_date=datetime.utcnow().strftime("%Y-%m-%d"),
        start_time="", end_time="",
        groomer_name=src.groomer_name,
        services_json=src.services_json,
        total_amount=src.total_amount,
        pet_size=src.pet_size, coat_length=src.coat_length,
        skin_condition="", behavior_note="",
        store=src.store,
        notes=src.notes,
        status="active",
        created_by=operator,
    )
    db.add(new_rec)
    _audit_doc_action(db, "grooming", 0, "copy_from", operator, extra=f"src={rec_id}")
    db.commit()
    db.refresh(new_rec)
    return RedirectResponse(f"/admin/grooming-orders/{new_rec.id}?msg=已复制为新美容单 · 请填本次时间/状态", status_code=303)


# ── 美容前后照片：上传 / 删除（CSV 存路径） ──
def _grooming_photos_list(csv: str) -> list[str]:
    return [p.strip() for p in (csv or "").split(",") if p.strip()]


def _grooming_photos_join(items: list[str]) -> str:
    return ",".join(items)


@app.post("/admin/grooming-orders/{rec_id}/upload-photos")
async def admin_grooming_upload_photos(
    rec_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    kind: str = Form(""),               # before / after
    photos: list[UploadFile] | None = File(None),
    next_url: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.status == "voided":
        return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=已作废单无法上传", status_code=303)
    if kind not in ("before", "after"):
        raise HTTPException(400, "kind 必须是 before 或 after")
    if not photos:
        return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=未选择文件", status_code=303)
    base = Path(settings.upload_dir) / "grooming" / str(rec_id)
    base.mkdir(parents=True, exist_ok=True)
    added = 0
    for uf in photos:
        if not uf.filename:
            continue
        ext = _image_ext(uf.filename)
        dest = base / f"{kind}_{secrets.token_hex(6)}{ext}"
        dest.write_bytes(await uf.read())
        try:
            dest = _compress_image(dest)
        except Exception:
            pass
        try:
            rel = Path(dest).resolve().relative_to(Path(settings.upload_dir).resolve())
            url = "/uploads/" + str(rel).replace("\\", "/")
        except Exception:
            url = "/uploads/grooming/" + str(rec_id) + "/" + dest.name
        added += 1
        if kind == "before":
            items = _grooming_photos_list(rec.before_photos)
            items.append(url)
            rec.before_photos = _grooming_photos_join(items)
        else:
            items = _grooming_photos_list(rec.after_photos)
            items.append(url)
            rec.after_photos = _grooming_photos_join(items)
    rec.updated_at = datetime.utcnow()
    db.commit()
    label = "美容前" if kind == "before" else "美容后"
    return RedirectResponse(
        _safe_next(next_url, f"/admin/grooming-orders/{rec_id}?msg=已上传{added}张{label}照片"),
        status_code=303,
    )


@app.post("/admin/grooming-orders/{rec_id}/delete-photo")
async def admin_grooming_delete_photo(
    rec_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    kind: str = Form(""),
    photo_url: str = Form(""),
    next_url: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.status == "voided":
        return RedirectResponse(f"/admin/grooming-orders/{rec_id}?msg=已作废单不可改", status_code=303)
    if kind not in ("before", "after"):
        raise HTTPException(400, "kind 错误")
    field_val = rec.before_photos if kind == "before" else rec.after_photos
    items = [p for p in _grooming_photos_list(field_val) if p != photo_url]
    if kind == "before":
        rec.before_photos = _grooming_photos_join(items)
    else:
        rec.after_photos = _grooming_photos_join(items)
    # 尝试物理删除文件
    try:
        if photo_url.startswith("/uploads/"):
            rel = photo_url[len("/uploads/"):]
            p = (Path(settings.upload_dir) / rel).resolve()
            root = Path(settings.upload_dir).resolve()
            if str(p).startswith(str(root)) and p.is_file():
                p.unlink()
    except Exception:
        pass
    rec.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/grooming-orders/{rec_id}?msg=已删除"),
        status_code=303,
    )


@app.post("/admin/weight-records/create")
async def admin_weight_create(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
    record_date: str = Form(""),
    weight_kg: float = Form(0.0),
    notes: str = Form(""),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    # 防呆：体重必须 > 0 且 ≤ 200kg（兜底防误输入负数或异常大值）
    if not pet_id or weight_kg <= 0 or weight_kg > 200:
        return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=weight&msg=体重需在 0~200kg 之间", status_code=303)
    rec = WeightRecord(
        pet_id=pet_id,
        record_date=record_date.strip()[:20],
        weight_kg=weight_kg,
        notes=notes.strip(),
        created_by=request.session.get("admin_username", ""),
    )
    db.add(rec)
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=weight&msg=体重已记录", status_code=303)


@app.post("/admin/weight-records/{rec_id}/delete")
async def admin_weight_delete(
    rec_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    rec = db.get(WeightRecord, rec_id)
    if rec:
        db.delete(rec)
        db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=weight&msg=体重记录已删除", status_code=303)


@app.post("/admin/medical-docs/upload")
async def admin_medical_doc_upload(
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
    visit_id: int = Form(0),
    doc_type: str = Form("consent"),
    title: str = Form(""),
    notes: str = Form(""),
    file: UploadFile = File(...),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _MEDDOC_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename or "").suffix.lower() or ".bin"
    fname = f"doc_{pet_id}_{int(datetime.utcnow().timestamp()*1000)}{ext}"
    dest = _MEDDOC_DIR / fname
    content = await file.read()
    dest.write_bytes(content)
    doc = MedicalDocument(
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        visit_id=visit_id or None,
        doc_type=doc_type.strip()[:40] or "consent",
        title=title.strip()[:200] or (file.filename or "未命名"),
        file_path=str(dest),
        original_name=file.filename or "",
        file_type="pdf" if ext == ".pdf" else "image",
        file_size=len(content),
        notes=notes.strip(),
        uploaded_by=request.session.get("admin_username", ""),
    )
    db.add(doc)
    db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=docs&msg=文书已上传", status_code=303)


@app.get("/admin/medical-docs/{doc_id}/file")
async def admin_medical_doc_file(doc_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    doc = db.get(MedicalDocument, doc_id)
    if not doc or not doc.file_path or not Path(doc.file_path).exists():
        raise HTTPException(404)
    from fastapi.responses import FileResponse
    return FileResponse(doc.file_path, filename=doc.original_name)


@app.post("/admin/medical-docs/{doc_id}/delete")
async def admin_medical_doc_delete(
    doc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    customer_id: int = Form(0),
    pet_id: int = Form(0),
):
    require_admin(request)
    _require_csrf(request, csrf_token)
    doc = db.get(MedicalDocument, doc_id)
    if doc:
        try:
            if doc.file_path and Path(doc.file_path).exists():
                Path(doc.file_path).unlink()
        except Exception:
            pass
        db.delete(doc)
        db.commit()
    return RedirectResponse(f"/admin/customers/{customer_id}?pet_id={pet_id}&tab=docs&msg=文书已删除", status_code=303)



# ── 处方模板（套餐） ─────────────────────────────────────────────

@app.get("/api/prescription-templates")
async def api_presc_templates_list(request: Request, db: Session = Depends(get_db)):
    """列出所有处方模板。"""
    require_admin(request)
    rows = db.query(PrescriptionTemplate).order_by(
        PrescriptionTemplate.use_count.desc(), PrescriptionTemplate.id.desc()
    ).limit(200).all()
    return [{
        "id": r.id, "name": r.name, "category": r.category,
        "use_count": r.use_count,
        "item_count": len(json.loads(r.items_json or "[]")),
    } for r in rows]


@app.get("/api/prescription-templates/{tpl_id}")
async def api_presc_template_get(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    tpl = db.get(PrescriptionTemplate, tpl_id)
    if not tpl:
        return {"ok": False, "error": "模板不存在"}
    # 使用计数 +1
    tpl.use_count = (tpl.use_count or 0) + 1
    db.commit()
    return {
        "ok": True,
        "id": tpl.id, "name": tpl.name, "notes": tpl.notes,
        "items": json.loads(tpl.items_json or "[]"),
    }


@app.post("/api/prescription-templates/create")
async def api_presc_template_create(request: Request, db: Session = Depends(get_db)):
    """从当前处方表单保存为模板。"""
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    name = (body.get("name") or "").strip()[:120]
    if not name:
        return {"ok": False, "error": "请填写模板名称"}
    items = body.get("items", [])
    if not isinstance(items, list) or not items:
        return {"ok": False, "error": "模板至少包含 1 个药品"}
    tpl = PrescriptionTemplate(
        name=name,
        category=(body.get("category") or "").strip()[:40],
        items_json=json.dumps(items, ensure_ascii=False),
        notes=(body.get("notes") or "").strip()[:500],
        created_by=request.session.get("admin_username", ""),
    )
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return {"ok": True, "id": tpl.id}


@app.post("/api/prescription-templates/{tpl_id}/delete")
async def api_presc_template_delete(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    tpl = db.get(PrescriptionTemplate, tpl_id)
    if tpl:
        db.delete(tpl)
        db.commit()
    return {"ok": True}


# ── 检查单模板（套餐） ─────────────────────────────────────────────

@app.get("/api/exam-templates")
async def api_exam_templates_list(request: Request, db: Session = Depends(get_db)):
    """列出所有检查模板（按使用次数倒序）。"""
    require_admin(request)
    rows = db.query(ExamTemplate).order_by(
        ExamTemplate.use_count.desc(), ExamTemplate.id.desc()
    ).limit(200).all()
    out = []
    for r in rows:
        try:
            items = json.loads(r.items_json or "[]")
        except Exception:
            items = []
        out.append({
            "id": r.id, "name": r.name, "category": r.category,
            "use_count": r.use_count, "item_count": len(items),
            # 让前端一次拿到摘要预览
            "item_names": [str(it.get("name") or "")[:14] for it in items[:4]],
        })
    return out


@app.get("/api/exam-templates/{tpl_id}")
async def api_exam_template_get(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    """获取模板详情 + 使用计数 +1。"""
    require_admin(request)
    tpl = db.get(ExamTemplate, tpl_id)
    if not tpl:
        return {"ok": False, "error": "模板不存在"}
    tpl.use_count = (tpl.use_count or 0) + 1
    db.commit()
    # 重算单价：若 item_id 仍存在 → 用最新库存价覆盖（套餐价格随时间漂移）
    raw_items = json.loads(tpl.items_json or "[]")
    op_store = _get_op_store(request)
    refreshed = []
    for it in raw_items:
        iid = it.get("item_id") or 0
        inv = db.get(InventoryItem, int(iid)) if iid else None
        if inv:
            from app.services.pricing import effective_sell_price as _eff
            it["unit_price"] = float(_eff(inv, op_store))
            it["unit"] = inv.unit or it.get("unit") or ""
        refreshed.append(it)
    return {
        "ok": True,
        "id": tpl.id, "name": tpl.name, "notes": tpl.notes,
        "items": refreshed,
    }


@app.post("/api/exam-templates/create")
async def api_exam_template_create(request: Request, db: Session = Depends(get_db)):
    """从当前检查单表单保存为模板。"""
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    name = (body.get("name") or "").strip()[:120]
    if not name:
        return {"ok": False, "error": "请填写模板名称"}
    items = body.get("items", [])
    if not isinstance(items, list) or not items:
        return {"ok": False, "error": "模板至少包含 1 个检查项"}
    tpl = ExamTemplate(
        name=name,
        category=(body.get("category") or "").strip()[:40],
        items_json=json.dumps(items, ensure_ascii=False),
        notes=(body.get("notes") or "").strip()[:500],
        created_by=request.session.get("admin_username", ""),
    )
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return {"ok": True, "id": tpl.id, "name": tpl.name}


@app.post("/api/exam-templates/{tpl_id}/delete")
async def api_exam_template_delete(tpl_id: int, request: Request, db: Session = Depends(get_db)):
    require_admin(request)
    body = await request.json()
    _require_csrf(request, body.get("csrf_token", ""))
    tpl = db.get(ExamTemplate, tpl_id)
    if tpl:
        db.delete(tpl)
        db.commit()
    return {"ok": True}


@app.get("/api/prescriptions/recent")
async def api_prescription_recent(
    pet_id: int = Query(0),
    customer_id: int = Query(0),
    exclude_visit_id: int = Query(0),
    request: Request = None,
    db: Session = Depends(get_db),
):
    """返回该宠物（或客户）最近一张处方的明细，供「复制上次处方」用。"""
    require_admin(request)
    q = db.query(Prescription)
    if pet_id:
        q = q.filter(Prescription.pet_id == pet_id)
    elif customer_id:
        q = q.filter(Prescription.customer_id == customer_id)
    else:
        return {"ok": False, "error": "缺少 pet_id 或 customer_id"}
    if exclude_visit_id:
        q = q.filter(Prescription.visit_id != exclude_visit_id)
    p = q.order_by(Prescription.id.desc()).first()
    if not p:
        return {"ok": False, "error": "无历史处方"}
    items = [{
        "drug_name": it.drug_name, "item_id": it.item_id,
        "drug_type": it.drug_type, "dosage": it.dosage,
        "frequency": it.frequency, "duration_days": it.duration_days,
        "quantity_num": it.quantity_num, "quantity": it.quantity,
        "unit_price": it.unit_price, "subtotal": it.subtotal,
        "instructions": it.instructions,
        # 新细化字段
        "dose_amount": it.dose_amount or 0,
        "dose_unit": it.dose_unit or "",
        "times_per_day": it.times_per_day or 0,
        "item_unit": it.item_unit or "",
        "print_note": it.print_note or "",
    } for it in p.items]
    return {
        "ok": True, "id": p.id, "prescribed_date": p.prescribed_date,
        "vet_name": p.vet_name, "notes": p.notes, "items": items,
    }


# ════════════════════════════════════════════════════════════════════════
# 住院管理（D1）：笼位 + 入院 + 出院（自动结账）
# ════════════════════════════════════════════════════════════════════════

_CAGE_KIND_ZH = {"general": "普通笼", "iso": "隔离笼", "icu": "ICU", "other": "其他"}
_HOSP_STATUS_ZH = {"admitted": "住院中", "discharged": "已出院", "cancelled": "已取消"}


def _gen_hosp_token(db: Session, field: str) -> str:
    """生成唯一短 token（staff/owner 各一个）。"""
    while True:
        tk = secrets.token_urlsafe(12)
        col = getattr(Hospitalization, field)
        exists = db.query(Hospitalization.id).filter(col == tk).first()
        if not exists:
            return tk


# ─── 笼位管理 ───
@app.get("/admin/cages", response_class=HTMLResponse)
async def admin_cages_list(request: Request, db: Session = Depends(get_db),
                            store: str = ""):
    require_admin(request)
    admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        wb_store = (store or "").strip()
    else:
        wb_store = admin_store
    q = db.query(Cage).filter(Cage.is_active == True)
    if wb_store:
        q = q.filter(Cage.store == wb_store)
    cages = q.order_by(Cage.store, Cage.sort_order, Cage.code).all()
    # 占用情况：每个 cage 当前是否有 admitted 的住院
    occupied_ids = {h.cage_id for h in db.query(Hospitalization)
                    .filter(Hospitalization.status == "admitted",
                            Hospitalization.cage_id != None).all()}
    return templates.TemplateResponse(request, "uk/cages.html", {
        "request": request, "cages": cages, "kind_zh": _CAGE_KIND_ZH,
        "occupied_ids": occupied_ids,
        "wb_store": wb_store, "csrf_token": _get_csrf_token(request),
        "is_superadmin": request.session.get("admin_role") == "superadmin",
        "title": "笼位管理",
    })


@app.post("/admin/cages/create")
async def admin_cages_create(request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form(""),
                              code: str = Form(""), kind: str = Form("general"),
                              daily_rate: float = Form(0.0),
                              store: str = Form(""), sort_order: int = Form(0),
                              notes: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    code = code.strip()[:40]
    if not code:
        return RedirectResponse("/admin/cages?msg=笼号不能为空", status_code=303)
    if kind not in _CAGE_KIND_ZH:
        kind = "general"
    # 超管可指定门店；员工自动归本店
    if request.session.get("admin_role") == "superadmin":
        cage_store = (store or "").strip()
    else:
        cage_store = _get_admin_store(request)
    # 同店内笼号唯一
    dup = db.query(Cage).filter(Cage.store == cage_store, Cage.code == code,
                                 Cage.is_active == True).first()
    if dup:
        return RedirectResponse(f"/admin/cages?msg=笼号「{code}」已存在", status_code=303)
    c = Cage(
        store=cage_store, code=code, kind=kind,
        daily_rate=max(0.0, float(daily_rate or 0)),
        sort_order=int(sort_order or 0),
        notes=notes.strip()[:500],
        created_by=request.session.get("admin_username", ""),
    )
    db.add(c)
    db.commit()
    return RedirectResponse(f"/admin/cages?msg=已添加笼位 {code}", status_code=303)


@app.post("/admin/cages/{cage_id}/edit")
async def admin_cages_edit(cage_id: int, request: Request, db: Session = Depends(get_db),
                            csrf_token: str = Form(""),
                            code: str = Form(""), kind: str = Form("general"),
                            daily_rate: float = Form(0.0),
                            sort_order: int = Form(0), notes: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    c = db.get(Cage, cage_id)
    if not c:
        raise HTTPException(404, "笼位不存在")
    c.code = (code or "").strip()[:40] or c.code
    if kind in _CAGE_KIND_ZH:
        c.kind = kind
    c.daily_rate = max(0.0, float(daily_rate or 0))
    c.sort_order = int(sort_order or 0)
    c.notes = (notes or "").strip()[:500]
    db.commit()
    return RedirectResponse(f"/admin/cages?msg=已保存", status_code=303)


@app.post("/admin/cages/{cage_id}/delete")
async def admin_cages_delete(cage_id: int, request: Request, db: Session = Depends(get_db),
                              csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    c = db.get(Cage, cage_id)
    if not c:
        raise HTTPException(404)
    # 占用中不让删
    occ = db.query(Hospitalization).filter(Hospitalization.cage_id == cage_id,
                                            Hospitalization.status == "admitted").first()
    if occ:
        return RedirectResponse(f"/admin/cages?msg=笼位 {c.code} 当前有动物住院，不能删除", status_code=303)
    c.is_active = False
    db.commit()
    return RedirectResponse(f"/admin/cages?msg=已删除笼位 {c.code}", status_code=303)


# ─── 住院档案 ───
@app.get("/admin/inpatient/new", response_class=HTMLResponse)
async def admin_inpatient_new_page(request: Request, db: Session = Depends(get_db),
                                     visit_id: int = 0):
    require_admin(request)
    v = db.get(Visit, visit_id) if visit_id else None
    if not v:
        raise HTTPException(404, "请从某次就诊记录发起入院")
    if (v.status or "open") == "closed":
        raise HTTPException(400, "病历已结束，不能再开住院")
    cust = db.get(Customer, v.customer_id) if v.customer_id else None
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    admin_store = _get_admin_store(request)
    store_short = admin_store or (pet.store if pet else "") or ""
    cages_q = db.query(Cage).filter(Cage.is_active == True)
    if store_short:
        cages_q = cages_q.filter(Cage.store == store_short)
    cages = cages_q.order_by(Cage.sort_order, Cage.code).all()
    occupied_ids = {h.cage_id for h in db.query(Hospitalization)
                    .filter(Hospitalization.status == "admitted",
                            Hospitalization.cage_id != None).all()}
    return templates.TemplateResponse(request, "uk/inpatient_new.html", {
        "request": request, "visit": v, "cust": cust, "pet": pet,
        "cages": cages, "occupied_ids": occupied_ids,
        "kind_zh": _CAGE_KIND_ZH, "store_short": store_short,
        "csrf_token": _get_csrf_token(request),
        "title": "新建住院",
    })


@app.post("/admin/inpatient/admit")
async def admin_inpatient_admit(request: Request, db: Session = Depends(get_db),
                                  csrf_token: str = Form(""),
                                  visit_id: int = Form(...),
                                  cage_id: int = Form(0),
                                  reason: str = Form(""),
                                  expected_discharge_date: str = Form(""),
                                  daily_rate_override: float = Form(0.0)):
    require_admin(request)
    _require_csrf(request, csrf_token)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404, "就诊记录不存在")
    # 已有"住院中"档案 → 跳到那张
    existing = db.query(Hospitalization).filter(
        Hospitalization.pet_id == v.pet_id,
        Hospitalization.status == "admitted",
    ).first()
    if existing:
        return RedirectResponse(f"/admin/inpatient/{existing.id}?msg=该宠物已有住院中档案",
                                 status_code=303)
    # 校验笼位空闲
    if cage_id:
        cage = db.get(Cage, cage_id)
        if not cage or not cage.is_active:
            raise HTTPException(400, "笼位无效")
        occ = db.query(Hospitalization).filter(
            Hospitalization.cage_id == cage_id,
            Hospitalization.status == "admitted",
        ).first()
        if occ:
            raise HTTPException(400, f"笼位 {cage.code} 已被占用")
    # 门店：取员工绑定店；超管取 visit 关联 pet.store
    admin_store = _get_admin_store(request)
    if admin_store:
        store_short = admin_store
    else:
        pet = db.get(Pet, v.pet_id) if v.pet_id else None
        store_short = (pet.store if pet else "") or ""
    h = Hospitalization(
        pet_id=v.pet_id, customer_id=v.customer_id, visit_id=visit_id,
        cage_id=cage_id or None,
        store=store_short,
        reason=(reason or v.chief_complaint or "")[:2000],
        expected_discharge_date=(expected_discharge_date or "").strip()[:20],
        daily_rate_override=max(0.0, float(daily_rate_override or 0)),
        status="admitted",
        staff_token=_gen_hosp_token(db, "staff_token"),
        owner_token=_gen_hosp_token(db, "owner_token"),
        created_by=request.session.get("admin_username", ""),
    )
    db.add(h)
    db.flush()
    _audit(db, request, "hospitalization_admit", detail={
        "id": h.id, "pet_id": h.pet_id, "visit_id": visit_id,
        "cage_id": cage_id, "store": store_short,
    })
    db.commit()
    return RedirectResponse(f"/admin/inpatient/{h.id}?msg=已入院",
                             status_code=303)


@app.post("/admin/inpatient/{hosp_id}/discharge")
async def admin_inpatient_discharge(hosp_id: int, request: Request,
                                      db: Session = Depends(get_db),
                                      csrf_token: str = Form(""),
                                      discharge_summary: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404)
    if h.status != "admitted":
        return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=当前状态不可出院",
                                 status_code=303)
    h.status = "discharged"
    h.discharged_at = datetime.utcnow()
    h.discharge_summary = (discharge_summary or "").strip()[:5000]
    h.closed_by = request.session.get("admin_username", "")
    db.flush()
    # 出院后同步收费单：加笼费明细
    operator = request.session.get("admin_username", "admin")
    if h.visit_id:
        inv = _sync_visit_invoice(db, h.visit_id, operator)
        if inv:
            h.invoice_id = inv.id
    _audit(db, request, "hospitalization_discharge",
           detail={"id": h.id, "days": _calc_hosp_days(h.admitted_at, h.discharged_at)})
    db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已出院 · 账单已同步",
                             status_code=303)


@app.post("/admin/inpatient/{hosp_id}/transfer")
async def admin_inpatient_transfer(hosp_id: int, request: Request,
                                     db: Session = Depends(get_db),
                                     csrf_token: str = Form(""),
                                     cage_id: int = Form(0)):
    """换笼。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    h = db.get(Hospitalization, hosp_id)
    if not h or h.status != "admitted":
        raise HTTPException(400, "状态不允许换笼")
    if cage_id:
        cage = db.get(Cage, cage_id)
        if not cage or not cage.is_active:
            raise HTTPException(400, "笼位无效")
        occ = db.query(Hospitalization).filter(
            Hospitalization.cage_id == cage_id,
            Hospitalization.status == "admitted",
            Hospitalization.id != hosp_id,
        ).first()
        if occ:
            raise HTTPException(400, f"笼位 {cage.code} 已被占用")
    h.cage_id = cage_id or None
    _audit(db, request, "hospitalization_transfer",
           detail={"id": h.id, "cage_id": cage_id})
    db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已换笼", status_code=303)


@app.post("/admin/inpatient/{hosp_id}/cancel")
async def admin_inpatient_cancel(hosp_id: int, request: Request,
                                   db: Session = Depends(get_db),
                                   csrf_token: str = Form("")):
    """误开作废（仅 admitted 可取消）。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    h = db.get(Hospitalization, hosp_id)
    if not h or h.status != "admitted":
        raise HTTPException(400)
    h.status = "cancelled"
    h.closed_by = request.session.get("admin_username", "")
    _audit(db, request, "hospitalization_cancel", detail={"id": h.id})
    db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已取消", status_code=303)


@app.get("/admin/inpatient", response_class=HTMLResponse)
async def admin_inpatient_board(request: Request, db: Session = Depends(get_db),
                                   status: str = "admitted", store: str = "",
                                   view: str = "cards"):
    """住院看板：卡片视图（默认）/ 笼位图视图。"""
    require_admin(request)
    admin_store = _get_admin_store(request)
    if request.session.get("admin_role") == "superadmin":
        wb_store = (store or "").strip()
    else:
        wb_store = admin_store
    q = db.query(Hospitalization)
    if status in ("admitted", "discharged", "cancelled"):
        q = q.filter(Hospitalization.status == status)
    if wb_store:
        q = q.filter(Hospitalization.store == wb_store)
    rows = q.order_by(Hospitalization.admitted_at.desc()).limit(200).all()

    # 给每张卡片附加：处方数、活跃药品数
    presc_map: dict[int, dict] = {}
    visit_ids = [h.visit_id for h in rows if h.visit_id]
    if visit_ids:
        from sqlalchemy import func
        rows_p = (db.query(Prescription.visit_id,
                            func.count(Prescription.id).label("cnt"))
                  .filter(Prescription.visit_id.in_(visit_ids),
                          Prescription.status.in_(["issued", "dispensed", "draft"]))
                  .group_by(Prescription.visit_id).all())
        for vid, cnt in rows_p:
            presc_map[vid] = {"count": cnt}

    # 笼位图：所有笼（按当前 wb_store 过滤），叠加占用映射
    cages_q = db.query(Cage).filter(Cage.is_active == True)
    if wb_store:
        cages_q = cages_q.filter(Cage.store == wb_store)
    all_cages = cages_q.order_by(Cage.store, Cage.sort_order, Cage.code).all()
    occ_map: dict[int, Hospitalization] = {}
    if view == "cages" or status == "admitted":
        # 用全量 admitted 的占用映射（不受 status 筛选影响）
        adm_q = db.query(Hospitalization).filter(Hospitalization.status == "admitted")
        if wb_store:
            adm_q = adm_q.filter(Hospitalization.store == wb_store)
        for h in adm_q.all():
            if h.cage_id:
                occ_map[h.cage_id] = h

    return templates.TemplateResponse(request, "uk/inpatient.html", {  # B6 UK 重写
        "request": request, "rows": rows, "status": status, "view": view,
        "status_zh": _HOSP_STATUS_ZH, "kind_zh": _CAGE_KIND_ZH,
        "wb_store": wb_store, "csrf_token": _get_csrf_token(request),
        "calc_days": _calc_hosp_days,
        "now": datetime.utcnow(),
        "presc_map": presc_map,
        "all_cages": all_cages, "occ_map": occ_map,
        "title": "住院管理",
    })


@app.get("/admin/inpatient/{hosp_id}", response_class=HTMLResponse)
async def admin_inpatient_detail(hosp_id: int, request: Request,
                                   db: Session = Depends(get_db)):
    require_admin(request)
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404)
    cust = db.get(Customer, h.customer_id) if h.customer_id else None
    pet = db.get(Pet, h.pet_id) if h.pet_id else None
    cage = db.get(Cage, h.cage_id) if h.cage_id else None
    visit = db.get(Visit, h.visit_id) if h.visit_id else None
    # 可换笼位（仅未占用 + 同店）
    avail_q = db.query(Cage).filter(Cage.is_active == True)
    if h.store:
        avail_q = avail_q.filter(Cage.store == h.store)
    avail_cages = avail_q.order_by(Cage.sort_order, Cage.code).all()
    occupied_ids = {x.cage_id for x in db.query(Hospitalization).filter(
        Hospitalization.status == "admitted",
        Hospitalization.id != hosp_id,
        Hospitalization.cage_id != None,
    ).all()}
    # 处方 + 用药任务（按时间排序）
    prescs = db.query(Prescription).filter(
        Prescription.visit_id == h.visit_id if h.visit_id else False,
    ).order_by(Prescription.id.desc()).all()
    # 今日发药任务（含漏药）+ 明日预览
    from datetime import date as _date2, timedelta as _td
    today_start = datetime.combine(_date2.today(), datetime.min.time())
    today_end   = today_start + _td(days=1)
    today_logs = db.query(MedicationAdminLog).filter(
        MedicationAdminLog.hospitalization_id == h.id,
        MedicationAdminLog.scheduled_at >= today_start,
        MedicationAdminLog.scheduled_at < today_end,
    ).order_by(MedicationAdminLog.scheduled_at, MedicationAdminLog.id).all()
    # 历史漏药（昨日及之前仍 pending）
    overdue_logs = db.query(MedicationAdminLog).filter(
        MedicationAdminLog.hospitalization_id == h.id,
        MedicationAdminLog.scheduled_at < today_start,
        MedicationAdminLog.status == "pending",
    ).order_by(MedicationAdminLog.scheduled_at).limit(20).all()
    # D4：生命体征 / I/O / 进食 最近记录 + 24h 汇总
    vitals = db.query(VitalSignsLog).filter(
        VitalSignsLog.hospitalization_id == h.id,
    ).order_by(VitalSignsLog.recorded_at.desc()).limit(20).all()
    io_logs = db.query(IOLog).filter(
        IOLog.hospitalization_id == h.id,
    ).order_by(IOLog.recorded_at.desc()).limit(20).all()
    feed_logs = db.query(FeedingLog).filter(
        FeedingLog.hospitalization_id == h.id,
    ).order_by(FeedingLog.recorded_at.desc()).limit(20).all()
    # 24h I/O 净平衡
    from datetime import timedelta as _td2
    cutoff = datetime.utcnow() - _td2(hours=24)
    io_24h = db.query(IOLog).filter(
        IOLog.hospitalization_id == h.id,
        IOLog.recorded_at >= cutoff,
    ).all()
    io_in_24h = sum((x.amount_ml or 0) for x in io_24h if x.direction == "in")
    io_out_24h = sum((x.amount_ml or 0) for x in io_24h if x.direction == "out")
    # 体征异常标记
    vital_flag_map = {v.id: _vital_flags(pet.species if pet else "", v) for v in vitals}
    # 交班记录（最新 10 条）
    handovers = db.query(HandoverNote).filter(
        HandoverNote.hospitalization_id == h.id,
    ).order_by(HandoverNote.recorded_at.desc()).limit(10).all()
    latest_handover = handovers[0] if handovers else None
    return templates.TemplateResponse(request, "uk/inpatient_detail.html", {  # B7 UK 重写
        "request": request, "h": h, "cust": cust, "pet": pet, "cage": cage,
        "visit": visit, "avail_cages": avail_cages, "occupied_ids": occupied_ids,
        "status_zh": _HOSP_STATUS_ZH, "kind_zh": _CAGE_KIND_ZH,
        "calc_days": _calc_hosp_days,
        "now": datetime.utcnow(),
        "prescs": prescs,
        "today_logs": today_logs, "overdue_logs": overdue_logs,
        "vitals": vitals, "io_logs": io_logs, "feed_logs": feed_logs,
        "io_in_24h": io_in_24h, "io_out_24h": io_out_24h,
        "vital_flag_map": vital_flag_map,
        "mm_color_zh": _MM_COLOR_ZH, "io_cat_zh": _IO_CATEGORY_ZH,
        "appetite_zh": _APPETITE_ZH,
        "handovers": handovers, "latest_handover": latest_handover,
        "shift_zh": _SHIFT_ZH, "current_shift": _guess_current_shift(),
        "csrf_token": _get_csrf_token(request),
        "title": f"住院 #{h.id}",
    })


# ─── 发药日志生成 + 操作 ───
def _parse_schedule_times(s: str) -> list[tuple[int, int]]:
    """解析 "08:00,14:00,20:00" → [(8,0),(14,0),(20,0)]。容错：跳过非法项。"""
    out = []
    for chunk in (s or "").replace("，", ",").replace("、", ",").split(","):
        c = chunk.strip()
        if not c:
            continue
        # 支持 "8", "8:30", "08:00"
        if ":" in c:
            try:
                hh, mm = c.split(":", 1)
                h = int(hh); m = int(mm)
            except Exception:
                continue
        else:
            try:
                h = int(c); m = 0
            except Exception:
                continue
        if 0 <= h < 24 and 0 <= m < 60:
            out.append((h, m))
    return out


# 住院处方医生没填时刻表时，按给药频次推默认发药时刻（覆盖绝大多数病例）：
#   SID/qd/q24h（一天一次）→ 以开方时间（北京整点）为默认
#   BID/q12h（一天两次）   → 10,20      （上午10点 / 晚8点）
#   TID/q8h（一天三次）    → 10,15,20   （上午10点 / 下午3点 / 晚8点）
#   QID/q6h（一天四次）    → 10,13,17,21（上午10点 / 下午1点 / 下午5点 / 晚9点）
#   q48h / prn / 一天四次以上 / 未知   → 空串（需人工排时间）
def _default_schedule_for_freq(freq: str, opened_hour: int) -> str:
    f = (freq or "").strip().lower()
    if f in ("qd", "q24h", "sid"):
        h = opened_hour if 0 <= opened_hour <= 23 else 10
        return str(h)
    if f in ("bid", "q12h"):
        return "10,20"
    if f in ("tid", "q8h"):
        return "10,15,20"
    if f in ("qid", "q6h"):
        return "10,13,17,21"
    return ""


def _generate_med_logs_for_prescription(db: Session, presc: "Prescription") -> int:
    """为处方批量生成用药日志（仅当处方关联到 admitted 住院时）。

    规则：
    - presc.status 必须 != 'draft' 且 != 'voided'
    - 找到该 visit_id 对应的 admitted Hospitalization
    - 每个 PrescriptionItem：
      * schedule_times 空 → 按给药频次推默认（见 _default_schedule_for_freq）；
        频次也推不出（prn/q48h/未知）→ 跳过
      * 否则按 (duration_days × schedule_times) 生成日志
      * 起始日：prescribed_date 或今天
      * 先删本 item 的现有 pending 日志（重生），保留 done/skipped
    返回新生成日志数。
    """
    if not presc or presc.status in ("draft", "voided"):
        return 0
    if not presc.visit_id:
        return 0
    hosp = db.query(Hospitalization).filter(
        Hospitalization.visit_id == presc.visit_id,
        Hospitalization.status == "admitted",
    ).first()
    if not hosp:
        return 0
    from datetime import date as _date2, timedelta as _td
    # 起始日
    start_date = None
    if presc.prescribed_date:
        try:
            start_date = datetime.strptime(presc.prescribed_date, "%Y-%m-%d").date()
        except Exception:
            start_date = None
    if not start_date:
        start_date = _date2.today()

    # 开方时间（北京整点），SID 默认用它
    try:
        opened_hour = (presc.created_at + _td(hours=8)).hour
    except Exception:
        opened_hour = 10

    created = 0
    for it in (presc.items or []):
        times = _parse_schedule_times(it.schedule_times or "")
        if not times:
            # 住院处方但医生没填时刻表 → 按给药频次套默认（SID 用开方时间 / BID 10,20 /
            # TID 10,15,20 / QID 10,13,17,21；q48h·prn·>4次/天 → 空，需人工排）。
            # 注：本函数只在存在 admitted 住院时才会执行到这里，故默认仅作用于住院处方。
            times = _parse_schedule_times(_default_schedule_for_freq(it.frequency, opened_hour))
        if not times:
            continue
        # 解析天数：duration_days 字段可能是 "7" 或 "症状缓解为止" 等
        try:
            n_days = int(str(it.duration_days or "").strip())
        except Exception:
            n_days = 7  # 默认 7 天，医生可以手动撤销/延长
        n_days = max(1, min(n_days, 14))  # 上限 14 天
        # 删本 item 的 pending（保留 done/skipped/refused）
        db.query(MedicationAdminLog).filter(
            MedicationAdminLog.prescription_item_id == it.id,
            MedicationAdminLog.status == "pending",
        ).delete(synchronize_session=False)
        for day_n in range(n_days):
            d = start_date + _td(days=day_n)
            for dose_idx, (h, m) in enumerate(times, 1):
                sched_at = datetime.combine(d, datetime.min.time()).replace(hour=h, minute=m)
                db.add(MedicationAdminLog(
                    hospitalization_id=hosp.id,
                    prescription_id=presc.id,
                    prescription_item_id=it.id,
                    scheduled_at=sched_at,
                    day_index=day_n + 1,
                    dose_index=dose_idx,
                    status="pending",
                ))
                created += 1
    db.flush()
    return created


def _safe_next(next_url: str, fallback: str) -> str:
    """防开放重定向：必须以 / 开头且非 //。"""
    nu = (next_url or "").strip()
    if nu and nu.startswith("/") and not nu.startswith("//"):
        return nu
    return fallback


@app.post("/admin/medication-log/{log_id}/check")
async def admin_medication_log_check(log_id: int, request: Request,
                                       db: Session = Depends(get_db),
                                       csrf_token: str = Form(""),
                                       dose_actual: str = Form(""),
                                       notes: str = Form(""),
                                       next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    log = db.get(MedicationAdminLog, log_id)
    if not log:
        raise HTTPException(404)
    hosp = db.get(Hospitalization, log.hospitalization_id)
    if not hosp or hosp.status != "admitted":
        raise HTTPException(400, "住院已结束，不可补打卡")
    log.status = "done"
    log.administered_at = datetime.utcnow()
    log.administered_by = request.session.get("admin_username", "")
    log.dose_actual = (dose_actual or "").strip()[:80]
    log.notes = (notes or "").strip()[:300]
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{log.hospitalization_id}#meds"),
        status_code=303,
    )


@app.post("/admin/medication-log/{log_id}/skip")
async def admin_medication_log_skip(log_id: int, request: Request,
                                      db: Session = Depends(get_db),
                                      csrf_token: str = Form(""),
                                      notes: str = Form(""),
                                      next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    log = db.get(MedicationAdminLog, log_id)
    if not log:
        raise HTTPException(404)
    log.status = "skipped"
    log.administered_at = datetime.utcnow()
    log.administered_by = request.session.get("admin_username", "")
    log.notes = (notes or "").strip()[:300]
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{log.hospitalization_id}#meds"),
        status_code=303,
    )


@app.post("/admin/medication-log/{log_id}/uncheck")
async def admin_medication_log_uncheck(log_id: int, request: Request,
                                         db: Session = Depends(get_db),
                                         csrf_token: str = Form(""),
                                         next_url: str = Form("")):
    """撤销打卡（误操作时）。"""
    require_admin(request)
    _require_csrf(request, csrf_token)
    log = db.get(MedicationAdminLog, log_id)
    if not log:
        raise HTTPException(404)
    log.status = "pending"
    log.administered_at = None
    log.administered_by = ""
    log.dose_actual = ""
    log.notes = ""
    # 撤销后清掉漏药推送标记，超过 grace 仍不打勾时会再次提醒
    log.reminder_sent_at = None
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{log.hospitalization_id}#meds"),
        status_code=303,
    )


# ════════════════════════════════════════════════════════════════════════
# 住院 D4：生命体征 / I/O / 进食记录
# ════════════════════════════════════════════════════════════════════════

_MM_COLOR_ZH = {
    "pink": "粉红 ✓", "pale": "苍白 ⚠", "cyanotic": "发绀 ⚠",
    "jaundice": "黄染 ⚠", "brick_red": "砖红 ⚠",
}
_IO_CATEGORY_ZH = {
    # in
    "iv_fluid": "静脉输液", "oral": "口服", "injection": "注射", "in_other": "其他(入)",
    # out
    "urine": "尿", "stool": "便", "vomit": "呕吐", "drainage": "引流", "out_other": "其他(出)",
}
_APPETITE_ZH = {
    0: "拒食 0", 1: "强饲 1", 2: "少量 2", 3: "正常 3", 4: "旺盛 4",
}
_SHIFT_ZH = {"morning": "早班", "afternoon": "中班", "night": "夜班"}


def _guess_current_shift() -> str:
    """按当前小时推断班次（北京时间）：早班 7-15，中班 15-22，夜班 22-7。"""
    from datetime import timezone, timedelta as _td3
    cn = timezone(_td3(hours=8))
    h = datetime.now(cn).hour
    if 7 <= h < 15:
        return "morning"
    if 15 <= h < 22:
        return "afternoon"
    return "night"


def _check_hosp_writable(db: Session, hosp_id: int):
    """统一校验住院档案可写。"""
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404, "住院档案不存在")
    if h.status != "admitted":
        raise HTTPException(400, "已出院 / 取消的住院档案不可记录")
    return h


@app.post("/admin/inpatient/{hosp_id}/vitals")
async def admin_inpatient_vitals_create(hosp_id: int, request: Request,
                                          db: Session = Depends(get_db),
                                          csrf_token: str = Form(""),
                                          temperature_c: float = Form(0.0),
                                          hr: int = Form(0),
                                          rr: int = Form(0),
                                          mm_color: str = Form(""),
                                          crt_sec: float = Form(0.0),
                                          weight_kg: float = Form(0.0),
                                          notes: str = Form(""),
                                          next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    if not any([temperature_c, hr, rr, mm_color, crt_sec, weight_kg]):
        return RedirectResponse(
            _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=至少填一项体征#vitals"),
            status_code=303,
        )
    if mm_color not in _MM_COLOR_ZH and mm_color != "":
        mm_color = ""
    log = VitalSignsLog(
        hospitalization_id=hosp_id,
        recorded_by=request.session.get("admin_username", ""),
        temperature_c=max(0.0, float(temperature_c or 0)),
        hr=max(0, int(hr or 0)),
        rr=max(0, int(rr or 0)),
        mm_color=mm_color,
        crt_sec=max(0.0, float(crt_sec or 0)),
        weight_kg=max(0.0, float(weight_kg or 0)),
        notes=(notes or "").strip()[:300],
    )
    db.add(log)
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=体征已记录#vitals"),
        status_code=303,
    )


@app.post("/admin/inpatient/{hosp_id}/vitals/{log_id}/delete")
async def admin_inpatient_vitals_delete(hosp_id: int, log_id: int, request: Request,
                                          db: Session = Depends(get_db),
                                          csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    log = db.get(VitalSignsLog, log_id)
    if log and log.hospitalization_id == hosp_id:
        db.delete(log)
        db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已删除#vitals",
                             status_code=303)


@app.post("/admin/inpatient/{hosp_id}/io")
async def admin_inpatient_io_create(hosp_id: int, request: Request,
                                      db: Session = Depends(get_db),
                                      csrf_token: str = Form(""),
                                      direction: str = Form("in"),
                                      category: str = Form("other"),
                                      amount_ml: float = Form(0.0),
                                      notes: str = Form(""),
                                      next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    if direction not in ("in", "out"):
        direction = "in"
    if amount_ml < 0:
        amount_ml = 0
    log = IOLog(
        hospitalization_id=hosp_id,
        recorded_by=request.session.get("admin_username", ""),
        direction=direction,
        category=category[:20],
        amount_ml=float(amount_ml),
        notes=(notes or "").strip()[:300],
    )
    db.add(log)
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=I/O 已记录#io"),
        status_code=303,
    )


@app.post("/admin/inpatient/{hosp_id}/io/{log_id}/delete")
async def admin_inpatient_io_delete(hosp_id: int, log_id: int, request: Request,
                                      db: Session = Depends(get_db),
                                      csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    log = db.get(IOLog, log_id)
    if log and log.hospitalization_id == hosp_id:
        db.delete(log)
        db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已删除#io",
                             status_code=303)


@app.post("/admin/inpatient/{hosp_id}/feeding")
async def admin_inpatient_feeding_create(hosp_id: int, request: Request,
                                           db: Session = Depends(get_db),
                                           csrf_token: str = Form(""),
                                           food_type: str = Form(""),
                                           offered_g: float = Form(0.0),
                                           eaten_g: float = Form(0.0),
                                           appetite_score: int = Form(3),
                                           notes: str = Form(""),
                                           next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    if appetite_score < 0 or appetite_score > 4:
        appetite_score = 3
    log = FeedingLog(
        hospitalization_id=hosp_id,
        recorded_by=request.session.get("admin_username", ""),
        food_type=(food_type or "").strip()[:120],
        offered_g=max(0.0, float(offered_g or 0)),
        eaten_g=max(0.0, float(eaten_g or 0)),
        appetite_score=int(appetite_score),
        notes=(notes or "").strip()[:300],
    )
    db.add(log)
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=进食已记录#feeding"),
        status_code=303,
    )


@app.post("/admin/inpatient/{hosp_id}/feeding/{log_id}/delete")
async def admin_inpatient_feeding_delete(hosp_id: int, log_id: int, request: Request,
                                           db: Session = Depends(get_db),
                                           csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    log = db.get(FeedingLog, log_id)
    if log and log.hospitalization_id == hosp_id:
        db.delete(log)
        db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已删除#feeding",
                             status_code=303)


# ─── D6：笼牌二维码 + 业主扫码登陆页 + 员工扫码跳转 ───
def _hosp_external_url(token: str, kind: str) -> str:
    """生成扫码外链。基址优先 settings.public_base_url，否则相对路径。"""
    base = (settings.public_base_url or "").rstrip("/")
    path = f"/inpatient/{kind}/{token}"
    return f"{base}{path}" if base else path


@app.get("/inpatient/staff/{token}")
async def inpatient_staff_scan(token: str, request: Request, db: Session = Depends(get_db)):
    """员工扫码 → 找到对应住院档案 → 已登录直跳详情，未登录引到登录页。"""
    h = db.query(Hospitalization).filter(Hospitalization.staff_token == token).first()
    if not h:
        raise HTTPException(404, "二维码无效或档案已删除")
    if not request.session.get("admin"):
        # 带上目的地，登录后回来
        return RedirectResponse(f"/admin/login?next=/admin/inpatient/{h.id}",
                                 status_code=303)
    return RedirectResponse(f"/admin/inpatient/{h.id}", status_code=303)


@app.get("/inpatient/owner/{token}", response_class=HTMLResponse)
async def inpatient_owner_scan(token: str, request: Request, db: Session = Depends(get_db)):
    """业主扫码 → 只读 H5（D7 会渲染真正内容；D6 先放最小占位 + 友好提示）。"""
    h = db.query(Hospitalization).filter(Hospitalization.owner_token == token).first()
    if not h:
        # 不暴露存在性 — 统一给一个"页面不存在"
        return templates.TemplateResponse(request, "inpatient_owner.html", {
            "request": request, "h": None, "title": "宠物住院信息",
        })
    pet = db.get(Pet, h.pet_id) if h.pet_id else None
    cust = db.get(Customer, h.customer_id) if h.customer_id else None
    # 业主可看的事件时间轴（最近 3 天）：用药打勾 / 喂食 / 体温
    from datetime import timedelta as _td3
    cutoff = datetime.utcnow() - _td3(days=3)
    events: list[dict] = []
    # 已完成用药
    for ml in db.query(MedicationAdminLog).filter(
        MedicationAdminLog.hospitalization_id == h.id,
        MedicationAdminLog.status == "done",
        MedicationAdminLog.administered_at >= cutoff,
    ).order_by(MedicationAdminLog.administered_at.desc()).limit(30).all():
        drug = ml.prescription_item.drug_name if ml.prescription_item else "药物"
        events.append({
            "kind": "med", "icon": "💊", "color": "#3b82f6",
            "title": drug,
            "subtitle": "已喂药",
            "at": ml.administered_at,
        })
    # 喂食
    for f in db.query(FeedingLog).filter(
        FeedingLog.hospitalization_id == h.id,
        FeedingLog.recorded_at >= cutoff,
    ).order_by(FeedingLog.recorded_at.desc()).limit(30).all():
        appetite_emoji = ["😢 拒食", "😟 强饲", "🙂 少量", "😊 正常", "😋 旺盛"][min(max(f.appetite_score, 0), 4)]
        eaten_text = ""
        if f.offered_g and f.eaten_g:
            eaten_text = f" · 吃了 {f.eaten_g:g}/{f.offered_g:g}g"
        elif f.eaten_g:
            eaten_text = f" · 吃了 {f.eaten_g:g}g"
        events.append({
            "kind": "feed", "icon": "🍽", "color": "#f97316",
            "title": (f.food_type or "进食"),
            "subtitle": f"{appetite_emoji}{eaten_text}",
            "at": f.recorded_at,
        })
    # 体温（只展示体温，不展示 HR/RR 等专业指标）
    for v in db.query(VitalSignsLog).filter(
        VitalSignsLog.hospitalization_id == h.id,
        VitalSignsLog.recorded_at >= cutoff,
        VitalSignsLog.temperature_c > 0,
    ).order_by(VitalSignsLog.recorded_at.desc()).limit(20).all():
        # 体温标签
        sp = (pet.species if pet else "").lower()
        if sp == "cat":
            normal = 38.0 <= v.temperature_c <= 39.5
        else:
            normal = 37.5 <= v.temperature_c <= 39.0
        events.append({
            "kind": "vital", "icon": "🌡", "color": "#ef4444" if not normal else "#10b981",
            "title": f"体温 {v.temperature_c:.1f}℃",
            "subtitle": "正常" if normal else "偏离正常范围（医生关注中）",
            "at": v.recorded_at,
        })
    events.sort(key=lambda e: e["at"] or datetime.min, reverse=True)
    events = events[:30]
    # 入院多少天
    days = _calc_hosp_days(h.admitted_at, h.discharged_at or datetime.utcnow())
    return templates.TemplateResponse(request, "inpatient_owner.html", {
        "request": request, "h": h, "pet": pet, "cust": cust,
        "events": events, "days": days,
        "title": "宠物住院信息",
    })


@app.get("/admin/inpatient/{hosp_id}/qr/{kind}.png")
async def admin_inpatient_qr_png(hosp_id: int, kind: str, request: Request,
                                   db: Session = Depends(get_db)):
    """生成二维码 PNG。kind ∈ staff/owner。"""
    require_admin(request)
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404)
    if kind not in ("staff", "owner"):
        raise HTTPException(400)
    token = h.staff_token if kind == "staff" else h.owner_token
    if not token:
        raise HTTPException(404, "token 缺失")
    try:
        import qrcode
        from io import BytesIO
    except ImportError:
        raise HTTPException(500, "服务端缺少 qrcode 库")
    url = _hosp_external_url(token, kind)
    qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_M,
                        box_size=10, border=2)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = BytesIO()
    img.save(buf, "PNG")
    buf.seek(0)
    return Response(content=buf.read(), media_type="image/png",
                    headers={"Cache-Control": "private, max-age=3600"})


@app.get("/admin/inpatient/{hosp_id}/cage-tag", response_class=HTMLResponse)
async def admin_inpatient_cage_tag(hosp_id: int, request: Request,
                                     db: Session = Depends(get_db)):
    """笼牌打印页：含员工 / 业主双二维码，工作人员打印后贴笼。"""
    require_admin(request)
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404)
    pet = db.get(Pet, h.pet_id) if h.pet_id else None
    cust = db.get(Customer, h.customer_id) if h.customer_id else None
    cage = db.get(Cage, h.cage_id) if h.cage_id else None
    return templates.TemplateResponse(request, "admin_cage_tag.html", {
        "request": request, "h": h, "pet": pet, "cust": cust, "cage": cage,
        "kind_zh": _CAGE_KIND_ZH,
        "calc_days": _calc_hosp_days, "now": datetime.utcnow(),
        "staff_url": _hosp_external_url(h.staff_token, "staff"),
        "owner_url": _hosp_external_url(h.owner_token, "owner"),
        "title": "笼牌",
    })


@app.post("/admin/inpatient/{hosp_id}/handover")
async def admin_inpatient_handover_create(hosp_id: int, request: Request,
                                            db: Session = Depends(get_db),
                                            csrf_token: str = Form(""),
                                            shift: str = Form(""),
                                            content: str = Form(""),
                                            next_url: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    if shift not in _SHIFT_ZH:
        shift = _guess_current_shift()
    content = (content or "").strip()
    if not content:
        return RedirectResponse(
            _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=内容不能为空#handover"),
            status_code=303,
        )
    log = HandoverNote(
        hospitalization_id=hosp_id,
        shift=shift,
        content=content[:2000],
        recorded_by=request.session.get("admin_username", ""),
    )
    db.add(log)
    db.commit()
    return RedirectResponse(
        _safe_next(next_url, f"/admin/inpatient/{hosp_id}?msg=交班已留言#handover"),
        status_code=303,
    )


@app.post("/admin/inpatient/{hosp_id}/handover/{log_id}/delete")
async def admin_inpatient_handover_delete(hosp_id: int, log_id: int, request: Request,
                                            db: Session = Depends(get_db),
                                            csrf_token: str = Form("")):
    require_admin(request)
    _require_csrf(request, csrf_token)
    _check_hosp_writable(db, hosp_id)
    log = db.get(HandoverNote, log_id)
    if log and log.hospitalization_id == hosp_id:
        db.delete(log)
        db.commit()
    return RedirectResponse(f"/admin/inpatient/{hosp_id}?msg=已删除#handover",
                             status_code=303)


def _vital_flags(species: str, log: "VitalSignsLog") -> dict:
    """判断哪些体征异常。species: cat/dog/其他。"""
    sp = (species or "").lower()
    flags = {"T": "", "HR": "", "RR": "", "MM": "", "CRT": ""}
    if log.temperature_c:
        if sp == "cat":
            if log.temperature_c < 38.0 or log.temperature_c > 39.5:
                flags["T"] = "high" if log.temperature_c > 39.5 else "low"
        else:  # dog 或未知，用犬阈值
            if log.temperature_c < 37.5 or log.temperature_c > 39.0:
                flags["T"] = "high" if log.temperature_c > 39.0 else "low"
    if log.hr:
        if sp == "cat":
            if log.hr < 120 or log.hr > 220:
                flags["HR"] = "high" if log.hr > 220 else "low"
        else:
            if log.hr < 60 or log.hr > 160:
                flags["HR"] = "high" if log.hr > 160 else "low"
    if log.rr:
        if sp == "cat":
            if log.rr < 16 or log.rr > 40:
                flags["RR"] = "high" if log.rr > 40 else "low"
        else:
            if log.rr < 10 or log.rr > 30:
                flags["RR"] = "high" if log.rr > 30 else "low"
    if log.mm_color and log.mm_color != "pink":
        flags["MM"] = "abnormal"
    if log.crt_sec and log.crt_sec >= 2:
        flags["CRT"] = "high"
    return flags


# ═════════════════════════════════════════════════════════════════════════════
# M1 · 手机 PWA 路由骨架
# /m            根入口，按 mobile_role 派发
# /m/doctor     医生首页（M3 起填内容）
# /m/nurse      助理首页（M2 起填内容）
# /m/groomer    美容师首页（M2.5 起填内容）
# /m/me         「我」页（账号/门店/扫码/切换视图）
# /m/desktop    强制切桌面（写 cookie + 跳 /admin）
# /m/switch     切换 mobile_role 视图（仅 session，不改库）
# ═════════════════════════════════════════════════════════════════════════════
def _current_mobile_role(request: Request, db: Session) -> str:
    """优先 session（允许临时切视图），否则查 DB。"""
    # session 里如果存了 override，优先用
    sv = request.session.get("mobile_role_override") or request.session.get("mobile_role")
    if not sv:
        uname = request.session.get("admin_username") or ""
        u = db.query(AdminUser).filter(AdminUser.username == uname).first() if uname else None
        sv = (u.mobile_role if u else "auto") or "auto"
        request.session["mobile_role"] = sv
    return _resolve_mobile_role(request.session.get("admin_role", "staff"), sv)


@app.get("/m", response_class=HTMLResponse)
async def m_root(request: Request, db: Session = Depends(get_db)):
    """UK-Minimal 统一首页（取代原三角色分离）。"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m", status_code=303)
    ctx = _m_ctx(request, db, active_tab="today")
    ctx["badges"] = _m_badges(request, db)

    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    today_str = datetime.utcnow().strftime("%Y-%m-%d")

    # 今日预约（前 8 条）
    appt_q = db.query(Appointment).filter(
        Appointment.appointment_date == today_str,
        Appointment.status.notin_(["cancelled", "rejected", "no_show"]),
    )
    if store_short:
        appt_q = appt_q.filter(or_(Appointment.store == store_full,
                                    Appointment.store == store_short))
    today_appts = appt_q.order_by(Appointment.appointment_time.asc()).limit(8).all()
    today_appts_total = appt_q.count()

    # 待审 TNR
    pending_tnr = db.query(Application).filter(
        Application.status == ApplicationStatus.pending_manual.value,
    )
    if store_short:
        pending_tnr = pending_tnr.filter(or_(
            Application.clinic_store == store_full,
            Application.clinic_store == store_short,
        ))
    pending_tnr_n = pending_tnr.count()

    # 我最近 5 个病历
    uname = request.session.get("admin_username") or ""
    my_visits = db.query(Visit).filter(
        or_(Visit.vet_name == uname, Visit.created_by == uname),
    ).order_by(Visit.id.desc()).limit(5).all()
    enriched_visits = []
    for v in my_visits:
        pet = db.get(Pet, v.pet_id) if v.pet_id else None
        cust = db.get(Customer, v.customer_id) if v.customer_id else None
        enriched_visits.append({"v": v, "pet": pet, "cust": cust})

    ctx.update({
        "today_appts": today_appts,
        "today_appts_total": today_appts_total,
        "pending_tnr_n": pending_tnr_n,
        "my_visits": enriched_visits,
        "today_str": today_str,
    })
    return templates.TemplateResponse(request, "m_uk/home.html", ctx)


@app.get("/m/desktop")
async def m_force_desktop(request: Request):
    """强制切桌面视图（写 cookie）。"""
    resp = RedirectResponse("/admin", status_code=303)
    resp.set_cookie("force_desktop", "1", max_age=60 * 60 * 24 * 30, httponly=False, samesite="lax")
    return resp


@app.get("/m/auto")
async def m_clear_force_desktop(request: Request):
    """清除 force_desktop cookie，恢复 UA 自动检测，跳回 /m。"""
    resp = RedirectResponse("/m", status_code=303)
    resp.delete_cookie("force_desktop")
    return resp


@app.post("/m/switch")
async def m_switch_role(
    request: Request,
    target: str = Form(""),
    csrf_token: str = Form(""),
):
    """临时切手机端视图（仅 session，不写库）。"""
    if not _admin_ok(request):
        raise HTTPException(401)
    _require_csrf(request, csrf_token)
    t = (target or "").strip().lower()
    if t not in ("doctor", "nurse", "groomer", "auto"):
        t = "auto"
    request.session["mobile_role_override"] = t
    return RedirectResponse("/m", status_code=303)


@app.post("/m/switch-store")
async def m_switch_store(
    request: Request,
    store: str = Form(""),
    csrf_token: str = Form(""),
    next_url: str = Form("/m/me"),
):
    """超管手机端切换"当前挂靠门店"（仅 session，不写库）。

    - 仅 superadmin 可用；店员的 session.admin_store 锁死本店，不允许切。
    - 写 session['admin_store']，由 _get_op_store 读取 → 影响手机列表/首页待办/开单品目过滤。
    - 空字符串 = 全部门店（超管看全店）。
    """
    if not _admin_ok(request):
        raise HTTPException(401)
    _require_csrf(request, csrf_token)
    if request.session.get("admin_role") != "superadmin":
        # 店员无权切店，原样回去
        return RedirectResponse(_safe_next(next_url, "/m/me"), status_code=303)
    s = (store or "").strip()
    if s not in ("东环店", "横岗店", ""):
        s = ""
    request.session["admin_store"] = s
    return RedirectResponse(_safe_next(next_url, "/m/me"), status_code=303)


def _m_ctx(request: Request, db: Session, *, active_tab: str) -> dict:
    """所有 /m/* 模板共用上下文。"""
    role = _current_mobile_role(request, db)
    uname = request.session.get("admin_username") or ""
    store = _get_admin_store(request)  # superadmin = "" = 全店
    return {
        "request": request,
        "csrf_token": _get_csrf_token(request),
        "mobile_role": role,
        "active_tab": active_tab,
        "admin_username": uname,
        "admin_role": request.session.get("admin_role", "staff"),
        "admin_store": store,
        # 简单顶部条用：拿原始 store 显示
        "admin_store_label": request.session.get("admin_store") or "全部门店",
    }


def _m_badges(request: Request, db: Session) -> dict:
    """首页待办徽章计数（漏药 / 待配药 / 回访 / 协议待签）。

    门店隔离：staff 只数本店；superadmin 数全部。
    住院相关的 store 比对用全名，所以要把短名转全名。
    """
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    now = datetime.utcnow()

    # 1. 漏药：scheduled_at <= now+5min，status=pending，关联 hosp 仍在 admitted
    med_q = db.query(MedicationAdminLog).join(
        Hospitalization, Hospitalization.id == MedicationAdminLog.hospitalization_id
    ).filter(
        MedicationAdminLog.status == "pending",
        MedicationAdminLog.scheduled_at <= now,
        Hospitalization.status == "admitted",
    )
    if store_short:
        med_q = med_q.filter(or_(Hospitalization.store == store_short,
                                  Hospitalization.store == store_full))
    overdue_meds = med_q.count()

    # 2. 待配药：status=issued 且 dispensed_at 空 且 未作废
    presc_q = db.query(Prescription).filter(
        Prescription.status == "issued",
        Prescription.dispensed_at == None,  # noqa: E711
        Prescription.voided_at == None,     # noqa: E711
    )
    # 处方表没存 store；通过 pet.store 关联过滤
    if store_short:
        presc_q = presc_q.outerjoin(Pet, Pet.id == Prescription.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )
    pending_dispense = presc_q.count()

    # 3. 回访：due / sent / phone_pending
    fu_q = db.query(FollowUp).filter(
        FollowUp.status.in_(["due", "sent", "phone_pending"])
    )
    if store_short:
        fu_q = fu_q.filter(or_(FollowUp.store == store_short, FollowUp.store == ""))
    due_followups = fu_q.count()

    # 4. 协议待签
    cs_q = db.query(ConsentTask).filter(ConsentTask.status == "pending")
    if store_short:
        cs_q = cs_q.filter(or_(ConsentTask.store == store_short, ConsentTask.store == ""))
    pending_consents = cs_q.count()

    # 5. 进行中的盘点：StocktakeSession 没有 store 字段，按 StocktakeItem 关联的 InventoryItem.store 判断
    st_q = db.query(StocktakeSession).filter(StocktakeSession.status == "open")
    if store_short:
        # 通过 session.name 包含门店名作为近似判断（创建时 session_name 含 store）
        st_q = st_q.filter(StocktakeSession.name.like(f"%{store_short}%"))
    open_stocktakes = st_q.count()

    # 6. 待录入报告：30 天内已开、未上传报告、且至少一个项目需要报告的检查单
    from datetime import timedelta as _td
    exam_cutoff = datetime.utcnow() - _td(days=30)
    exam_q = db.query(ExamOrder).filter(
        ExamOrder.status != "voided",
        ExamOrder.created_at >= exam_cutoff,
        ~db.query(ExamReport.id).filter(ExamReport.exam_order_id == ExamOrder.id).exists(),
    )
    if store_short:
        exam_q = exam_q.outerjoin(Visit, Visit.id == ExamOrder.visit_id).outerjoin(
            Pet, Pet.id == Visit.pet_id
        ).filter(or_(Pet.store == store_short, Pet.store == "", Pet.store == None))  # noqa: E711
    exam_report_pending = 0
    for eo in exam_q.all():
        try:
            idata = json.loads(eo.items_json or "[]")
        except Exception:
            idata = []
        # 至少一个项目需要报告（跳过保定费/操作费等纯收费项）
        for it in idata:
            iid = it.get("item_id") if isinstance(it, dict) else None
            inv = db.get(InventoryItem, iid) if iid else None
            if inv is not None and not inv.requires_report:
                continue
            exam_report_pending += 1
            break

    # 7. 库存预警（低/零库存）+ 库存过期（90 天内到期批次）
    from datetime import date as _date, timedelta as _timedelta
    inv_base = _apply_store_filter(
        db.query(InventoryItem).filter(InventoryItem.is_active == True),
        InventoryItem.store, store_short,
    )
    low_stock = inv_base.filter(
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= InventoryItem.low_stock_min,
        InventoryItem.low_stock_min > 0,
    ).count()
    _alert_date = (_date.today() + _timedelta(days=90)).isoformat()
    exp_q = db.query(InventoryBatch.item_id).join(
        InventoryItem, InventoryItem.id == InventoryBatch.item_id
    ).filter(
        InventoryBatch.expiry_date != "",
        InventoryBatch.expiry_date <= _alert_date,
    )
    exp_q = _apply_store_filter(exp_q, InventoryItem.store, store_short)
    inventory_expiry = exp_q.distinct().count()

    return {
        "overdue_meds": overdue_meds,
        "pending_dispense": pending_dispense,
        "due_followups": due_followups,
        "pending_consents": pending_consents,
        "open_stocktakes": open_stocktakes,
        "exam_report_pending": exam_report_pending,
        "low_stock": low_stock,
        "inventory_expiry": inventory_expiry,
    }


@app.get("/m/doctor", response_class=HTMLResponse)
async def m_doctor_home_legacy(request: Request):
    """旧角色首页 → 统一首页（UK 重构）"""
    return RedirectResponse("/m", status_code=303)


async def m_doctor_home_OBSOLETE(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/doctor", status_code=303)
    ctx = _m_ctx(request, db, active_tab="today")
    ctx["mobile_role"] = "doctor"
    ctx["badges"] = _m_badges(request, db)

    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    today_str = datetime.utcnow().strftime("%Y-%m-%d")

    # 今日预约（所有类别，按时间排）
    appt_q = db.query(Appointment).filter(
        Appointment.appointment_date == today_str,
        Appointment.status.notin_(["cancelled", "rejected", "no_show"]),
    )
    if store_short:
        appt_q = appt_q.filter(or_(Appointment.store == store_full,
                                    Appointment.store == store_short))
    today_appts = appt_q.order_by(Appointment.appointment_time.asc()).all()

    # 待审 TNR
    pending_tnr = db.query(Application).filter(
        Application.status == ApplicationStatus.pending_manual.value,
    )
    if store_short:
        pending_tnr = pending_tnr.filter(or_(
            Application.clinic_store == store_full,
            Application.clinic_store == store_short,
        ))
    pending_tnr_n = pending_tnr.count()

    # 最近 5 个我开的病历
    uname = request.session.get("admin_username") or ""
    my_visits = db.query(Visit).filter(
        or_(Visit.vet_name == uname, Visit.created_by == uname),
    ).order_by(Visit.id.desc()).limit(5).all()
    enriched_visits = []
    for v in my_visits:
        pet = db.get(Pet, v.pet_id) if v.pet_id else None
        cust = db.get(Customer, v.customer_id) if v.customer_id else None
        enriched_visits.append({"v": v, "pet": pet, "cust": cust})

    ctx.update({
        "today_appts": today_appts,
        "pending_tnr_n": pending_tnr_n,
        "my_visits": enriched_visits,
    })
    return templates.TemplateResponse(request, "m/doctor_home.html", ctx)


@app.get("/m/nurse", response_class=HTMLResponse)
async def m_nurse_home_legacy(request: Request):
    """旧角色首页 → 统一首页"""
    return RedirectResponse("/m", status_code=303)


# M1 占位的 /m/groomer 由 M2.5 的 m_groomer_home_v2 覆盖


@app.get("/m/medical", response_class=HTMLResponse)
async def m_medical_hub(request: Request, db: Session = Depends(get_db)):
    """医疗 tab：病历 / 处方 / 检查 / 住院 / 配药 / 疫苗 / 驱虫 / 回访"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/medical", status_code=303)
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx["badges"] = _m_badges(request, db)
    return templates.TemplateResponse(request, "m_uk/medical_hub.html", ctx)


@app.get("/m/finance", response_class=HTMLResponse)
async def m_finance_hub(request: Request, db: Session = Depends(get_db)):
    """财务 tab：收费单 / 钱包 / 套餐 / 押金 / 优惠券 / 收款报表"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/finance", status_code=303)
    ctx = _m_ctx(request, db, active_tab="finance")
    return templates.TemplateResponse(request, "m_uk/finance_hub.html", ctx)


@app.get("/m/soon", response_class=HTMLResponse)
async def m_soon(request: Request, title: str = "敬请期待", db: Session = Depends(get_db)):
    """占位页：还没开发的功能跳这里。"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m", status_code=303)
    ctx = _m_ctx(request, db, active_tab="")
    ctx["soon_title"] = title
    return templates.TemplateResponse(request, "m/soon.html", ctx)


@app.get("/m/me", response_class=HTMLResponse)
async def m_me(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/me", status_code=303)
    ctx = _m_ctx(request, db, active_tab="me")
    # 当前 mobile_role 的"原始值"，和 override 区分开
    uname = request.session.get("admin_username") or ""
    u = db.query(AdminUser).filter(AdminUser.username == uname).first() if uname else None
    ctx["mobile_role_raw"] = (u.mobile_role if u else "auto") or "auto"
    ctx["override_active"] = bool(request.session.get("mobile_role_override"))
    return templates.TemplateResponse(request, "m_uk/me.html", ctx)


# ═════════════════════════════════════════════════════════════════════════════
# M2 · 助理版核心页：住院 / 回访 / 待配药
# ═════════════════════════════════════════════════════════════════════════════
def _m_store_filters_hosp(query, store_short: str, store_full: str):
    if store_short:
        return query.filter(or_(Hospitalization.store == store_short,
                                 Hospitalization.store == store_full))
    return query


@app.get("/m/inpatient", response_class=HTMLResponse)
async def m_inpatient_list(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/inpatient", status_code=303)
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    q = db.query(Hospitalization).filter(Hospitalization.status == "admitted")
    q = _m_store_filters_hosp(q, store_short, store_full)
    hosps = q.order_by(Hospitalization.admitted_at.desc()).all()

    now = datetime.utcnow()
    cards = []
    for h in hosps:
        pet = db.get(Pet, h.pet_id) if h.pet_id else None
        cust = db.get(Customer, h.customer_id) if h.customer_id else None
        cage = db.get(Cage, h.cage_id) if h.cage_id else None
        # overdue 数
        overdue_n = db.query(MedicationAdminLog).filter(
            MedicationAdminLog.hospitalization_id == h.id,
            MedicationAdminLog.status == "pending",
            MedicationAdminLog.scheduled_at <= now,
        ).count()
        days = _calc_hosp_days(h.admitted_at, h.discharged_at or now)
        cards.append({
            "h": h, "pet": pet, "cust": cust, "cage": cage,
            "overdue_n": overdue_n, "days": days,
        })
    ctx = _m_ctx(request, db, active_tab="inpatient")
    ctx["cards"] = cards
    ctx["total_n"] = len(cards)
    return templates.TemplateResponse(request, "m_uk/inpatient_list.html", ctx)


@app.get("/m/inpatient/{hosp_id}", response_class=HTMLResponse)
async def m_inpatient_detail(hosp_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/inpatient/{hosp_id}", status_code=303)
    h = db.get(Hospitalization, hosp_id)
    if not h:
        raise HTTPException(404)
    # 门店校验
    store_short = _get_admin_store(request)
    if store_short:
        store_full = _STORE_SHORT_TO_FULL.get(store_short, "")
        if h.store not in (store_short, store_full):
            raise HTTPException(403, "无权查看其他门店住院")

    pet = db.get(Pet, h.pet_id) if h.pet_id else None
    cust = db.get(Customer, h.customer_id) if h.customer_id else None
    cage = db.get(Cage, h.cage_id) if h.cage_id else None

    now = datetime.utcnow()
    # 今日发药任务
    today_start = datetime(now.year, now.month, now.day)
    today_end = today_start + timedelta(days=1)
    today_logs = db.query(MedicationAdminLog).filter(
        MedicationAdminLog.hospitalization_id == hosp_id,
        MedicationAdminLog.scheduled_at >= today_start,
        MedicationAdminLog.scheduled_at < today_end,
    ).order_by(MedicationAdminLog.scheduled_at).all()
    # 漏药（昨天/更早的 pending）
    overdue_logs = db.query(MedicationAdminLog).filter(
        MedicationAdminLog.hospitalization_id == hosp_id,
        MedicationAdminLog.status == "pending",
        MedicationAdminLog.scheduled_at < today_start,
    ).order_by(MedicationAdminLog.scheduled_at).all()

    # 最新交班
    latest_handover = db.query(HandoverNote).filter(
        HandoverNote.hospitalization_id == hosp_id
    ).order_by(HandoverNote.recorded_at.desc()).first()

    # 最近 3 条体征
    recent_vitals = db.query(VitalSignsLog).filter(
        VitalSignsLog.hospitalization_id == hosp_id
    ).order_by(VitalSignsLog.recorded_at.desc()).limit(3).all()

    days = _calc_hosp_days(h.admitted_at, h.discharged_at or now)
    ctx = _m_ctx(request, db, active_tab="inpatient")
    ctx.update({
        "h": h, "pet": pet, "cust": cust, "cage": cage,
        "today_logs": [l for l in today_logs if l.prescription_item],
        "overdue_logs": [l for l in overdue_logs if l.prescription_item],
        "latest_handover": latest_handover,
        "recent_vitals": recent_vitals,
        "days": days,
        "now": now,
        "shift_zh": _SHIFT_ZH,
        "current_shift": _guess_current_shift(),
        "next_url": f"/m/inpatient/{hosp_id}",
    })
    return templates.TemplateResponse(request, "m_uk/inpatient_detail.html", ctx)


# ─── 回访 ───
@app.get("/m/follow-ups", response_class=HTMLResponse)
async def m_follow_ups(request: Request, tab: str = "today", db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/follow-ups", status_code=303)
    store_short = _get_op_store(request)
    base = db.query(FollowUp)
    if store_short:
        base = base.filter(or_(FollowUp.store == store_short, FollowUp.store == ""))
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    if tab == "today":
        rows = base.filter(
            FollowUp.status.in_(["due", "sent", "phone_pending"]),
            FollowUp.planned_date <= today_str,
        ).order_by(FollowUp.planned_date.asc()).all()
    elif tab == "responded":
        rows = base.filter(
            FollowUp.status == "responded",
        ).order_by(FollowUp.response_at.desc().nullslast()).limit(50).all()
    elif tab == "closed":
        rows = base.filter(
            FollowUp.status == "closed",
        ).order_by(FollowUp.handled_at.desc().nullslast()).limit(50).all()
    else:
        rows = base.order_by(FollowUp.id.desc()).limit(50).all()
    # 富化每行
    enriched = []
    for r in rows:
        pet = db.get(Pet, r.pet_id) if r.pet_id else None
        cust = db.get(Customer, r.customer_id) if r.customer_id else None
        enriched.append({"fu": r, "pet": pet, "cust": cust})
    ctx = _m_ctx(request, db, active_tab="follow_ups")
    ctx.update({"rows": enriched, "tab": tab})
    return templates.TemplateResponse(request, "m_uk/follow_ups.html", ctx)


# ─── 待配药 ───
@app.get("/m/dispensing", response_class=HTMLResponse)
async def m_dispensing_list(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/dispensing", status_code=303)
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    q = db.query(Prescription).filter(
        Prescription.status == "issued",
        Prescription.dispensed_at == None,   # noqa: E711
        Prescription.voided_at == None,      # noqa: E711
    )
    if store_short:
        q = q.outerjoin(Pet, Pet.id == Prescription.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )
    rows = q.order_by(Prescription.created_at.desc()).limit(100).all()
    enriched = []
    for p in rows:
        pet = db.get(Pet, p.pet_id) if p.pet_id else None
        cust = db.get(Customer, p.customer_id) if p.customer_id else None
        item_cnt = len(p.items)
        enriched.append({"p": p, "pet": pet, "cust": cust, "item_cnt": item_cnt})
    ctx = _m_ctx(request, db, active_tab="dispensing")
    ctx["rows"] = enriched
    return templates.TemplateResponse(request, "m_uk/dispensing_list.html", ctx)


@app.get("/m/dispensing/{presc_id}", response_class=HTMLResponse)
async def m_dispensing_detail(presc_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/dispensing/{presc_id}", status_code=303)
    p = db.get(Prescription, presc_id)
    if not p:
        raise HTTPException(404)
    pet = db.get(Pet, p.pet_id) if p.pet_id else None
    cust = db.get(Customer, p.customer_id) if p.customer_id else None
    # 检查管控药
    has_controlled = False
    for it in p.items:
        if it.item_id:
            inv = db.get(InventoryItem, it.item_id)
            if inv and getattr(inv, "is_controlled", False):
                has_controlled = True
                break
    ctx = _m_ctx(request, db, active_tab="dispensing")
    ctx.update({
        "p": p, "pet": pet, "cust": cust,
        "has_controlled": has_controlled,
    })
    return templates.TemplateResponse(request, "m_uk/dispensing_detail.html", ctx)


@app.post("/m/dispensing/{presc_id}/mark")
async def m_dispensing_mark(
    presc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    if not _admin_ok(request):
        raise HTTPException(401)
    _require_csrf(request, csrf_token)
    p = db.get(Prescription, presc_id)
    if not p:
        raise HTTPException(404)
    if p.dispensed_at is not None:
        return RedirectResponse(f"/m/dispensing/{presc_id}?msg=该处方已配齐", status_code=303)
    if p.status == "voided":
        return RedirectResponse(f"/m/dispensing/{presc_id}?err=该处方已作废", status_code=303)
    p.dispensed_at = datetime.utcnow()
    p.dispensed_by = request.session.get("admin_username", "")
    if p.status == "issued":
        p.status = "dispensed"
    db.commit()
    return RedirectResponse("/m/dispensing?msg=已配齐 ✓", status_code=303)


# ═════════════════════════════════════════════════════════════════════════════
# M7 · 移动端循环盘点：列表 / 会话详情 / 保存
# 用法：超管在桌面发起盘点单 → 医生/助理拿手机进入对应会话 → 逐项填实盘 → 保存
# 提交（status=open → completed）仍由桌面操作，避免误提交
# ═════════════════════════════════════════════════════════════════════════════

@app.get("/m/stocktake", response_class=HTMLResponse)
async def m_stocktake_list(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/stocktake", status_code=303)
    store_short = _get_op_store(request)
    q = db.query(StocktakeSession).filter(StocktakeSession.status == "open")
    if store_short:
        q = q.filter(StocktakeSession.name.like(f"%{store_short}%"))
    sessions = q.order_by(StocktakeSession.created_at.desc()).all()
    # 每个会话算下进度
    enriched = []
    for s in sessions:
        total = db.query(StocktakeItem).filter(StocktakeItem.session_id == s.id).count()
        counted = db.query(StocktakeItem).filter(
            StocktakeItem.session_id == s.id,
            StocktakeItem.actual_qty.isnot(None),
        ).count()
        enriched.append({"sess": s, "total": total, "counted": counted})
    ctx = _m_ctx(request, db, active_tab="")
    ctx["rows"] = enriched
    return templates.TemplateResponse(request, "m/stocktake_list.html", ctx)


@app.get("/m/stocktake/{session_id}", response_class=HTMLResponse)
async def m_stocktake_detail(session_id: int, request: Request, db: Session = Depends(get_db),
                              q: str = Query(""), filter: str = Query("all")):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/stocktake/{session_id}", status_code=303)
    sess = db.get(StocktakeSession, session_id)
    if not sess:
        raise HTTPException(404)
    items_q = db.query(StocktakeItem).filter(StocktakeItem.session_id == session_id)
    if q:
        items_q = items_q.filter(StocktakeItem.item_name.ilike(f"%{q}%"))
    if filter == "uncounted":
        items_q = items_q.filter(StocktakeItem.actual_qty.is_(None))
    elif filter == "variance":
        items_q = items_q.filter(StocktakeItem.actual_qty.isnot(None),
                                  StocktakeItem.variance != 0)
    sit_items = items_q.order_by(StocktakeItem.category, StocktakeItem.item_name).all()
    # 全量统计（不受筛选影响）
    total = db.query(StocktakeItem).filter(StocktakeItem.session_id == session_id).count()
    counted = db.query(StocktakeItem).filter(
        StocktakeItem.session_id == session_id,
        StocktakeItem.actual_qty.isnot(None),
    ).count()
    ctx = _m_ctx(request, db, active_tab="")
    ctx.update({
        "sess": sess, "sit_items": sit_items,
        "q": q, "filter": filter,
        "total": total, "counted": counted,
        "categories": INVENTORY_CATEGORIES,
    })
    return templates.TemplateResponse(request, "m/stocktake_detail.html", ctx)


@app.post("/m/stocktake/{session_id}/save-one")
async def m_stocktake_save_one(
    session_id: int, request: Request, db: Session = Depends(get_db),
    csrf_token: str = Form(""),
    item_id: int = Form(...),
    actual_qty: str = Form(""),
    notes: str = Form(""),
    next_url: str = Form(""),
):
    """单条保存（移动端逐项录入，省得整页提交）。"""
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/stocktake/{session_id}", status_code=303)
    _require_csrf(request, csrf_token)
    sess = db.get(StocktakeSession, session_id)
    if not sess or sess.status != "open":
        return RedirectResponse(f"/m/stocktake/{session_id}?err=盘点已完成或不存在", status_code=303)
    si = db.query(StocktakeItem).filter(
        StocktakeItem.id == item_id,
        StocktakeItem.session_id == session_id,
    ).first()
    if not si:
        return RedirectResponse(f"/m/stocktake/{session_id}?err=该项不存在", status_code=303)
    v = (actual_qty or "").strip()
    if v == "":
        si.actual_qty = None
        si.variance = 0.0
    else:
        try:
            si.actual_qty = float(v)
            si.variance = si.actual_qty - si.system_qty
        except ValueError:
            return RedirectResponse(f"/m/stocktake/{session_id}?err=数量格式错误", status_code=303)
    si.notes = (notes or "").strip()[:500]
    db.commit()
    target = _safe_next(next_url, f"/m/stocktake/{session_id}?msg=已保存")
    return RedirectResponse(target, status_code=303)


# ═════════════════════════════════════════════════════════════════════════════
# M2.5 · 美容师页：今日预约 / 美容单 / 新建 / 拍照
# ═════════════════════════════════════════════════════════════════════════════
_M_BEAUTY_CATS = ("beauty", "grooming", "washcare")


def _m_today_beauty_appts(db: Session, store_short: str, store_full: str):
    """今日的美容预约（按时间排序）。"""
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    q = db.query(Appointment).filter(
        Appointment.category.in_(_M_BEAUTY_CATS),
        Appointment.appointment_date == today_str,
        Appointment.status.notin_(["cancelled", "rejected", "no_show"]),
    )
    if store_short:
        q = q.filter(or_(Appointment.store == store_full,
                          Appointment.store == store_short))
    return q.order_by(Appointment.appointment_time.asc()).all()


@app.get("/m/groomer", response_class=HTMLResponse)
async def m_groomer_home_legacy(request: Request):
    """旧角色首页 → 统一首页"""
    return RedirectResponse("/m", status_code=303)


async def m_groomer_home_OBSOLETE(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/groomer", status_code=303)
    ctx = _m_ctx(request, db, active_tab="today")
    ctx["mobile_role"] = "groomer"
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    appts = _m_today_beauty_appts(db, store_short, store_full)
    # 最近 3 个美容单
    q = db.query(GroomingOrder).filter(GroomingOrder.status != "voided")
    if store_short:
        q = q.filter(or_(GroomingOrder.store == store_short,
                          GroomingOrder.store == store_full,
                          GroomingOrder.store == ""))
    recent = q.order_by(GroomingOrder.id.desc()).limit(5).all()
    # 未拍照单（today + 没 after_photos）
    no_photo = 0
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    for r in recent:
        if r.groom_date == today_str and not (r.after_photos or "").strip():
            no_photo += 1
    ctx.update({
        "appts": appts, "recent": recent, "no_photo": no_photo,
    })
    return templates.TemplateResponse(request, "m/groomer_home.html", ctx)


@app.get("/m/grooming", response_class=HTMLResponse)
async def m_grooming_list(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/grooming", status_code=303)
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    q = db.query(GroomingOrder)
    if store_short:
        q = q.filter(or_(GroomingOrder.store == store_short,
                          GroomingOrder.store == store_full,
                          GroomingOrder.store == ""))
    rows = q.order_by(GroomingOrder.id.desc()).limit(80).all()
    enriched = []
    for r in rows:
        pet = db.get(Pet, r.pet_id) if r.pet_id else None
        cust = db.get(Customer, r.customer_id) if r.customer_id else None
        before_n = len([p for p in (r.before_photos or "").split(",") if p.strip()])
        after_n = len([p for p in (r.after_photos or "").split(",") if p.strip()])
        enriched.append({
            "r": r, "pet": pet, "cust": cust,
            "before_n": before_n, "after_n": after_n,
        })
    ctx = _m_ctx(request, db, active_tab="grooming")
    ctx["mobile_role"] = "groomer"
    ctx["rows"] = enriched
    return templates.TemplateResponse(request, "m_uk/grooming_list.html", ctx)


@app.get("/m/grooming/new", response_class=HTMLResponse)
async def m_grooming_new(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
    appointment_id: int = 0,
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/grooming/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    appt = db.get(Appointment, appointment_id) if appointment_id else None
    if appt and not cust:
        cust = db.get(Customer, appt.customer_id) if appt.customer_id else None
        pet = db.get(Pet, appt.pet_id) if appt.pet_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    groomers = [s[0] for s in db.query(Staff.name).filter(
        Staff.status.in_(["active", "probation"]),
    ).all()]
    groom_items = _query_grooming_items(db, request)
    ctx = _m_ctx(request, db, active_tab="grooming")
    ctx["mobile_role"] = "groomer"
    ctx.update({
        "cust": cust, "pet": pet, "appt": appt, "pets": pets,
        "groomers": groomers, "groom_items": groom_items,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "now_time": datetime.utcnow().strftime("%H:%M"),
        "default_groomer": request.session.get("admin_username", ""),
    })
    return templates.TemplateResponse(request, "m_uk/grooming_new.html", ctx)


@app.post("/m/grooming/create")
async def m_grooming_create(request: Request, db: Session = Depends(get_db)):
    """手机精简新建：不出收费单，留到桌面端收费。"""
    if not _admin_ok(request):
        raise HTTPException(401)
    form = await request.form()
    _require_csrf(request, str(form.get("csrf_token", "")))
    customer_id = int(form.get("customer_id", 0) or 0)
    pet_id = int(form.get("pet_id", 0) or 0)
    appointment_id = int(form.get("appointment_id", 0) or 0)
    if not pet_id or not customer_id:
        return RedirectResponse("/m/grooming/new?err=请选择客户和宠物", status_code=303)
    # 服务：item_id[] CSV 形式（chip 多选）
    service_ids = form.getlist("service_id[]")
    services = []
    for sid in service_ids:
        try:
            iid = int(sid)
        except Exception:
            continue
        inv = db.get(InventoryItem, iid)
        if not inv:
            continue
        price = float(inv.sell_price or 0)
        services.append({
            "name": inv.name,
            "item_id": iid,
            "qty": 1.0,
            "price": price,
            "subtotal": round(price, 2),
            "notes": "",
        })
    total = round(sum(s["subtotal"] for s in services), 2)
    operator = request.session.get("admin_username", "admin")
    pet = db.get(Pet, pet_id)
    store = (pet.store if pet else "") or _get_admin_store(request) or ""
    rec = GroomingOrder(
        customer_id=customer_id or None,
        pet_id=pet_id or None,
        appointment_id=appointment_id or None,
        groom_date=str(form.get("groom_date", "")).strip()[:20] or datetime.utcnow().strftime("%Y-%m-%d"),
        groomer_name=str(form.get("groomer_name", "")).strip()[:80] or operator,
        assistant_name=str(form.get("assistant_name", "")).strip()[:80],
        services_json=json.dumps(services, ensure_ascii=False),
        total_amount=total,
        skin_condition=str(form.get("skin_condition", "")).strip()[:200],
        behavior_note=str(form.get("behavior_note", "")).strip()[:200],
        store=store,
        notes=str(form.get("notes", "")).strip(),
        created_by=operator,
    )
    db.add(rec)
    db.flush()
    # 自动扣库存（仅非服务）
    for s in services:
        iid = s.get("item_id")
        qty = float(s.get("qty") or 0)
        if iid and qty > 0:
            inv = db.get(InventoryItem, iid)
            if inv and not inv.is_service:
                _deduct_inventory(db, iid, qty, "grooming", rec.id, operator,
                                  note=f"美容#{rec.id} {s.get('name','')}")
    db.commit()
    return RedirectResponse(f"/m/grooming/{rec.id}?msg=美容单已建", status_code=303)


@app.get("/m/grooming/{rec_id}", response_class=HTMLResponse)
async def m_grooming_detail(rec_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/grooming/{rec_id}", status_code=303)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    pet = db.get(Pet, rec.pet_id) if rec.pet_id else None
    cust = db.get(Customer, rec.customer_id) if rec.customer_id else None
    try:
        services = json.loads(rec.services_json or "[]")
    except Exception:
        services = []
    before_list = [p for p in (rec.before_photos or "").split(",") if p.strip()]
    after_list = [p for p in (rec.after_photos or "").split(",") if p.strip()]
    locked, lock_reason = _is_grooming_locked(db, rec)
    ctx = _m_ctx(request, db, active_tab="grooming")
    ctx["mobile_role"] = "groomer"
    ctx.update({
        "rec": rec, "pet": pet, "cust": cust,
        "services": services,
        "before_list": before_list, "after_list": after_list,
        "locked": locked, "lock_reason": lock_reason,
        "next_url": f"/m/grooming/{rec_id}",
    })
    return templates.TemplateResponse(request, "m_uk/grooming_detail.html", ctx)


@app.post("/m/grooming/{rec_id}/done")
async def m_grooming_mark_done(rec_id: int, request: Request,
                                 db: Session = Depends(get_db),
                                 csrf_token: str = Form("")):
    """手机端「完成美容」按钮：写 end_time。"""
    if not _admin_ok(request):
        raise HTTPException(401)
    _require_csrf(request, csrf_token)
    rec = db.get(GroomingOrder, rec_id)
    if not rec:
        raise HTTPException(404)
    if rec.status == "voided":
        return RedirectResponse(f"/m/grooming/{rec_id}?err=已作废", status_code=303)
    rec.end_time = datetime.utcnow().strftime("%H:%M")
    rec.updated_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/m/grooming/{rec_id}?msg=已标记完成", status_code=303)


# 简易客户搜索：JSON，给 grooming_new.html 用
@app.get("/m/api/search-customer")
async def m_api_search_customer(
    request: Request,
    q: str = "",
    db: Session = Depends(get_db),
):
    if not _admin_ok(request):
        return {"results": []}
    q = (q or "").strip()
    if len(q) < 2:
        return {"results": []}
    # 宠物名也命中
    _pet_owner_ids = db.query(Pet.customer_id).filter(Pet.name.ilike(f"%{q}%"))
    custs = db.query(Customer).filter(
        or_(
            Customer.name.ilike(f"%{q}%"),
            Customer.phone.ilike(f"%{q}%"),
            Customer.phones_extra.ilike(f"%{q}%"),
            Customer.id.in_(_pet_owner_ids),
        )
    ).order_by(Customer.id.desc()).limit(10).all()
    results = []
    for c in custs:
        pets = db.query(Pet).filter(Pet.customer_id == c.id).order_by(Pet.id).all()
        results.append({
            "id": c.id,
            "name": c.name or "",
            "phone": c.phone or "",
            "pets": [{
                "id": p.id, "name": p.name or "",
                "species": p.species or "", "breed": p.breed or "",
                "gender": p.gender or "unknown",
            } for p in pets],
        })
    return {"results": results}


# ═════════════════════════════════════════════════════════════════════════════
# M3 · 医生只读层：客户搜索 / 客户档案 / 病历详情
# ═════════════════════════════════════════════════════════════════════════════
@app.get("/m/customers", response_class=HTMLResponse)
async def m_customers_search(request: Request, q: str = "", db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/customers", status_code=303)
    q = (q or "").strip()
    results = []
    if len(q) >= 2:
        # 宠物名命中的客户 id（子查询）
        _pet_owner_ids = db.query(Pet.customer_id).filter(Pet.name.ilike(f"%{q}%"))
        custs = db.query(Customer).filter(
            or_(
                Customer.name.ilike(f"%{q}%"),
                Customer.phone.ilike(f"%{q}%"),
                Customer.phones_extra.ilike(f"%{q}%"),   # 备用号
                Customer.id.in_(_pet_owner_ids),          # 按宠物名搜
            )
        ).order_by(Customer.id.desc()).limit(30).all()
        for c in custs:
            pets = db.query(Pet).filter(Pet.customer_id == c.id).order_by(Pet.id).all()
            results.append({"c": c, "pets": pets})
    ctx = _m_ctx(request, db, active_tab="customers")
    ctx.update({"q": q, "results": results})
    return templates.TemplateResponse(request, "m_uk/customers.html", ctx)


@app.get("/m/customers/new", response_class=HTMLResponse)
async def m_customer_new_form(request: Request, db: Session = Depends(get_db)):
    """手机端新建客户表单"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/customers/new", status_code=303)
    ctx = _m_ctx(request, db, active_tab="customers")
    return templates.TemplateResponse(request, "m_uk/customer_new.html", ctx)


@app.get("/m/customer/{cust_id}/pets/new", response_class=HTMLResponse)
async def m_pet_new_form(cust_id: int, request: Request, db: Session = Depends(get_db)):
    """手机端给已有客户加宠物"""
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/customer/{cust_id}/pets/new", status_code=303)
    cust = db.get(Customer, cust_id)
    if not cust:
        raise HTTPException(404)
    ctx = _m_ctx(request, db, active_tab="customers")
    ctx["cust"] = cust
    # 默认门店：自己绑的店；超管空就让选
    ctx["default_store"] = request.session.get("admin_store") or ""
    ctx["is_superadmin"] = (request.session.get("admin_role") == "superadmin")
    return templates.TemplateResponse(request, "m_uk/pet_new.html", ctx)


@app.get("/m/customer/{cust_id}", response_class=HTMLResponse)
async def m_customer_profile(cust_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/customer/{cust_id}", status_code=303)
    cust = db.get(Customer, cust_id)
    if not cust:
        raise HTTPException(404)
    pets = db.query(Pet).filter(Pet.customer_id == cust_id).order_by(Pet.id).all()
    wallet = db.query(Wallet).filter(Wallet.customer_id == cust_id).first()
    packages = db.query(CustomerPackage).filter(
        CustomerPackage.customer_id == cust_id,
        CustomerPackage.used_count < CustomerPackage.total_uses,
    ).order_by(CustomerPackage.id.desc()).all()
    # 最近 10 个病历
    visits = db.query(Visit).filter(Visit.customer_id == cust_id)\
        .order_by(Visit.id.desc()).limit(10).all()
    # 最近的住院（admitted 才显示）
    admitted = db.query(Hospitalization).filter(
        Hospitalization.customer_id == cust_id,
        Hospitalization.status == "admitted",
    ).order_by(Hospitalization.id.desc()).all()
    enriched_visits = []
    for v in visits:
        p = db.get(Pet, v.pet_id) if v.pet_id else None
        enriched_visits.append({"v": v, "pet": p})
    ctx = _m_ctx(request, db, active_tab="customers")
    ctx.update({
        "cust": cust, "pets": pets,
        "wallet": wallet, "packages": packages,
        "visits": enriched_visits,
        "admitted": admitted,
    })
    return templates.TemplateResponse(request, "m_uk/customer_profile.html", ctx)


_VISIT_TYPE_ZH = {
    "outpatient": "门诊",
    "followup":   "复诊",
    "postop":     "术后",
    "vaccine":    "疫苗",
    "surgery_consult": "术前面诊",
    "other":      "其他",
}


@app.get("/m/pet/{pet_id}", response_class=HTMLResponse)
async def m_pet_profile(pet_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/pet/{pet_id}", status_code=303)
    pet = db.get(Pet, pet_id)
    if not pet:
        raise HTTPException(404)
    cust = db.get(Customer, pet.customer_id) if pet.customer_id else None
    # 最近 10 病历（该宠物）
    visits = db.query(Visit).filter(Visit.pet_id == pet_id)\
        .order_by(Visit.id.desc()).limit(10).all()
    # 疫苗 + 驱虫
    vaccinations = db.query(Vaccination).filter(
        Vaccination.pet_id == pet_id,
        Vaccination.status == "active",
    ).order_by(Vaccination.vaccinated_date.desc()).limit(10).all()
    dewormings = db.query(DewormingRecord).filter(
        DewormingRecord.pet_id == pet_id,
        DewormingRecord.status == "active",
    ).order_by(DewormingRecord.deworm_date.desc()).limit(10).all()
    # 体重曲线（最近 6 条）
    weights = db.query(WeightRecord).filter(WeightRecord.pet_id == pet_id)\
        .order_by(WeightRecord.record_date.desc(), WeightRecord.id.desc())\
        .limit(6).all()
    # 当前住院（如果有）
    admitted = db.query(Hospitalization).filter(
        Hospitalization.pet_id == pet_id,
        Hospitalization.status == "admitted",
    ).first()
    ctx = _m_ctx(request, db, active_tab="customers")
    ctx.update({
        "pet": pet, "cust": cust,
        "visits": visits, "vaccinations": vaccinations,
        "dewormings": dewormings, "weights": weights,
        "admitted": admitted,
    })
    return templates.TemplateResponse(request, "m_uk/pet_profile.html", ctx)


@app.get("/m/visits", response_class=HTMLResponse)
async def m_visits_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    scope: str = "today",   # today / mine / all
):
    """病历列表：今日 / 我的 / 全部，可搜索宠物名/客户名/病历号"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/visits", status_code=303)
    store_short = _get_op_store(request)
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    uname = request.session.get("admin_username") or ""

    base = db.query(Visit)
    if scope == "today":
        base = base.filter(Visit.visit_date == today_str)
    elif scope == "mine":
        base = base.filter(or_(Visit.vet_name == uname, Visit.created_by == uname))
    # all：不加额外条件，默认按 id 倒序取最近

    # 门店隔离：通过 pet.store
    if store_short:
        base = base.outerjoin(Pet, Pet.id == Visit.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )

    # 搜索
    q = (q or "").strip()
    if q:
        if q.isdigit():
            base = base.filter(Visit.id == int(q))
        else:
            base = base.outerjoin(Customer, Customer.id == Visit.customer_id).filter(
                or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
            )

    visits = base.order_by(Visit.id.desc()).limit(50).all()
    rows = []
    for v in visits:
        p = db.get(Pet, v.pet_id) if v.pet_id else None
        c = db.get(Customer, v.customer_id) if v.customer_id else None
        rows.append({"v": v, "pet": p, "cust": c})

    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "scope": scope, "q": q, "today_str": today_str})
    return templates.TemplateResponse(request, "m_uk/visits.html", ctx)


@app.get("/m/visit/new", response_class=HTMLResponse)
async def m_visit_new(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
    appointment_id: int = 0,
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/visit/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    appt = db.get(Appointment, appointment_id) if appointment_id else None
    if appt and not cust:
        cust = db.get(Customer, appt.customer_id) if appt.customer_id else None
        pet = db.get(Pet, appt.pet_id) if appt.pet_id else None
    if pet and not cust:
        # 从宠物档案页只带了 pet_id 进来时，反查所属客户，免得再搜一遍
        cust = db.get(Customer, pet.customer_id) if pet.customer_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    uname = request.session.get("admin_username") or ""
    default_vet = uname or ""
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "mode": "new", "v": None,
        "cust": cust, "pet": pet, "pets": pets, "appt": appt,
        "default_vet": default_vet,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "visit_types": _VISIT_TYPE_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/visit_edit.html", ctx)


@app.get("/m/visit/{visit_id}/edit", response_class=HTMLResponse)
async def m_visit_edit_form(visit_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/visit/{visit_id}/edit", status_code=303)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404)
    if (v.status or "open") == "closed":
        return RedirectResponse(f"/m/visit/{visit_id}?err=病历已结束，不可编辑", status_code=303)
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    cust = db.get(Customer, v.customer_id) if v.customer_id else None
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "mode": "edit", "v": v,
        "cust": cust, "pet": pet, "pets": [],
        "default_vet": v.vet_name or "",
        "today": v.visit_date,
        "visit_types": _VISIT_TYPE_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/visit_edit.html", ctx)


@app.get("/m/vaccination/new", response_class=HTMLResponse)
async def m_vaccination_new(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/vaccination/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    if pet and not cust:
        cust = db.get(Customer, pet.customer_id) if pet.customer_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    # 疫苗品目（category=vaccine）
    vacc_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(InventoryItem.category == "vaccine", InventoryItem.is_active == True).all()
    uname = request.session.get("admin_username") or ""
    default_vet = uname or ""
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "cust": cust, "pet": pet, "pets": pets,
        "vacc_items": vacc_items,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "default_vet": default_vet,
    })
    return templates.TemplateResponse(request, "m_uk/vaccination_new.html", ctx)


@app.get("/m/deworming/new", response_class=HTMLResponse)
async def m_deworming_new(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/deworming/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    if pet and not cust:
        cust = db.get(Customer, pet.customer_id) if pet.customer_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    deworm_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(InventoryItem.category == "antiparasitic", InventoryItem.is_active == True).all()
    uname = request.session.get("admin_username") or ""
    u = db.query(AdminUser).filter(AdminUser.username == uname).first() if uname else None
    default_vet = (u.display_name if u and u.display_name else uname) or ""
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "cust": cust, "pet": pet, "pets": pets,
        "deworm_items": deworm_items,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "default_vet": default_vet,
    })
    return templates.TemplateResponse(request, "m_uk/deworming_new.html", ctx)


@app.get("/m/sales/new", response_class=HTMLResponse)
async def m_sales_new(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/sales/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    if pet and not cust:
        cust = db.get(Customer, pet.customer_id) if pet.customer_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({
        "cust": cust, "pet": pet, "pets": pets,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
    })
    return templates.TemplateResponse(request, "m_uk/sales_new.html", ctx)


@app.get("/m/visit/{visit_id}/exam", response_class=HTMLResponse)
async def m_exam_new(visit_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/visit/{visit_id}/exam", status_code=303)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404)
    if (v.status or "open") == "closed":
        return RedirectResponse(f"/m/visit/{visit_id}?err=病历已结束，不可开检查", status_code=303)
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    cust = db.get(Customer, v.customer_id) if v.customer_id else None
    # 检查品目：category in (lab/imaging/microscopy)
    exam_items = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, _get_op_store(request)
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.category.in_(["lab", "imaging", "microscopy"]),
    ).order_by(InventoryItem.category, InventoryItem.subcategory, InventoryItem.name).all()
    # 按 category 分组
    grouped = {"lab": [], "imaging": [], "microscopy": []}
    for it in exam_items:
        grouped.setdefault(it.category, []).append(it)
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "v": v, "pet": pet, "cust": cust,
        "grouped": grouped,
    })
    return templates.TemplateResponse(request, "m_uk/exam_new.html", ctx)


@app.get("/m/invoices", response_class=HTMLResponse)
async def m_invoices_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    status: str = "unpaid",       # unpaid / paid / all
    customer_id: int = 0,
):
    """收费单列表：未结清 / 已收款 / 全部 + 按客户筛选 + 搜索"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/invoices", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(Invoice).order_by(Invoice.id.desc())
    if store_short:
        query = query.filter(Invoice.store == store_short)
    if status == "unpaid":
        query = query.filter(Invoice.payment_status.in_(("unpaid", "partial")))
    elif status == "paid":
        query = query.filter(Invoice.payment_status == "paid")
    if customer_id:
        query = query.filter(Invoice.customer_id == customer_id)
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        if q.isdigit():
            query = query.filter(or_(
                Invoice.id == int(q),
                Invoice.customer_id.in_(cids),
            ))
        else:
            query = query.filter(Invoice.customer_id.in_(cids))
    invoices = query.limit(80).all()
    rows = []
    for inv in invoices:
        c = db.get(Customer, inv.customer_id) if inv.customer_id else None
        p = db.get(Pet, inv.pet_id) if inv.pet_id else None
        paid = _invoice_paid_sum(db, inv.id)
        outstanding = max(0.0, float(inv.total_amount or 0) - paid)
        rows.append({"inv": inv, "cust": c, "pet": p,
                     "paid": paid, "outstanding": outstanding})
    # KPI: 今日已收 + 未付总欠款（员工内购档案排除）
    from datetime import date as _date
    today_str = _date.today().isoformat()
    _internal_ids_sub = db.query(Customer.id).filter(Customer.is_internal == True).subquery()
    base_q = db.query(Invoice).filter(~Invoice.customer_id.in_(_internal_ids_sub))
    if store_short:
        base_q = base_q.filter(Invoice.store == store_short)
    today_paid = base_q.filter(
        Invoice.payment_status == "paid",
        Invoice.invoice_date == today_str,
    ).all()
    unpaid_all = base_q.filter(Invoice.payment_status.in_(("unpaid", "partial"))).all()
    kpi = {
        "today_paid_total": round(sum(float(i.total_amount or 0) for i in today_paid), 2),
        "today_paid_count": len(today_paid),
        "unpaid_count": len(unpaid_all),
        "unpaid_total": round(sum(
            max(0.0, float(i.total_amount or 0) - _invoice_paid_sum(db, i.id))
            for i in unpaid_all
        ), 2),
    }
    customer = db.get(Customer, customer_id) if customer_id else None
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({
        "rows": rows, "q": q, "status": status,
        "customer": customer, "customer_id": customer_id,
        "kpi": kpi, "today_str": today_str,
    })
    return templates.TemplateResponse(request, "m_uk/invoices.html", ctx)


@app.get("/m/invoices/{inv_id}", response_class=HTMLResponse)
async def m_invoice_detail(inv_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/invoices/{inv_id}", status_code=303)
    inv = db.get(Invoice, inv_id)
    if not inv:
        raise HTTPException(404)
    cust  = db.get(Customer, inv.customer_id) if inv.customer_id else None
    pet   = db.get(Pet, inv.pet_id) if inv.pet_id else None
    visit = db.get(Visit, inv.visit_id) if inv.visit_id else None
    # 钱包
    wallet_balance = 0.0
    if inv.customer_id:
        w = db.query(Wallet).filter(Wallet.customer_id == inv.customer_id).first()
        wallet_balance = float(w.balance) if w else 0.0
    # 有效套餐
    active_packages = []
    if inv.customer_id:
        active_packages = db.query(CustomerPackage).filter(
            CustomerPackage.customer_id == inv.customer_id,
            CustomerPackage.status == "active",
        ).order_by(CustomerPackage.id.desc()).all()
    # 押金
    available_deposits = []
    if inv.customer_id:
        for d in db.query(Deposit).filter(
            Deposit.customer_id == inv.customer_id,
            Deposit.status.in_(["held", "partial_refund"]),
        ).order_by(Deposit.id.desc()).all():
            remaining = d.amount - (d.applied_amount or 0) - (d.refunded_amount or 0)
            if remaining > 0:
                d._remaining = remaining
                available_deposits.append(d)
    # 优惠券
    available_coupons = []
    if inv.customer_id:
        for c in db.query(Coupon).filter(
            ((Coupon.customer_id == inv.customer_id) | (Coupon.customer_id.is_(None))),
            Coupon.status == "issued",
        ).order_by(Coupon.id.desc()).all():
            if _coupon_is_expired(c):
                continue
            usable = _coupon_compute_amount(c, float(inv.total_amount or 0))
            if usable > 0:
                c._usable_amount = usable
                available_coupons.append(c)
    # 已加的 Payment 流水
    payments = db.query(Payment).filter(
        Payment.invoice_id == inv_id
    ).order_by(Payment.id.desc()).all()
    paid_sum = sum(float(p.amount or 0) for p in payments if p.status == "success")
    outstanding = max(0.0, float(inv.total_amount or 0) - paid_sum)
    is_super = (request.session.get("admin_role") == "superadmin")
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({
        "inv": inv, "cust": cust, "pet": pet, "visit": visit,
        "wallet_balance": wallet_balance,
        "active_packages": active_packages,
        "available_deposits": available_deposits,
        "available_coupons": available_coupons,
        "payments": payments,
        "paid_sum": paid_sum,
        "outstanding": outstanding,
        "method_zh": _REVENUE_PAY_ZH,
        "other_unpaid": _other_unpaid_for_invoice(db, inv) if inv.customer_id else [],
        "is_super": is_super,
        "deposit_category_zh": _DEPOSIT_CATEGORY_ZH,
        "coupon_kind_zh": _COUPON_KIND_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/invoice_detail.html", ctx)


@app.get("/m/customer/{cust_id}/wallet/recharge", response_class=HTMLResponse)
async def m_wallet_recharge_form(cust_id: int, request: Request, db: Session = Depends(get_db)):
    """手机端钱包充值表单"""
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/customer/{cust_id}/wallet/recharge", status_code=303)
    cust = db.get(Customer, cust_id)
    if not cust:
        raise HTTPException(404)
    wallet = db.query(Wallet).filter(Wallet.customer_id == cust_id).first()
    ctx = _m_ctx(request, db, active_tab="customers")
    ctx.update({
        "cust": cust,
        "wallet_balance": float(wallet.balance) if wallet else 0.0,
    })
    return templates.TemplateResponse(request, "m_uk/wallet_recharge.html", ctx)


@app.get("/m/reports/revenue", response_class=HTMLResponse)
async def m_reports_revenue(
    request: Request,
    db: Session = Depends(get_db),
    preset: str = "today",  # today / 7d / 30d / month
):
    """手机端收款报表（精简版：当日 / 近 7 天 / 近 30 天 / 本月）"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/reports/revenue", status_code=303)
    from datetime import date as _date, timedelta as _td
    store_short = _get_op_store(request)
    today = _date.today()
    if preset == "today":
        d_from = today.isoformat(); d_to = today.isoformat(); label = "今日"
    elif preset == "7d":
        d_from = (today - _td(days=6)).isoformat(); d_to = today.isoformat(); label = "近 7 天"
    elif preset == "30d":
        d_from = (today - _td(days=29)).isoformat(); d_to = today.isoformat(); label = "近 30 天"
    else:  # month
        d_from = today.replace(day=1).isoformat(); d_to = today.isoformat(); label = "本月"
    # 按 Payment 表聚合（员工内购档案排除）
    _internal_ids_sub = db.query(Customer.id).filter(Customer.is_internal == True).subquery()
    q = db.query(Payment).filter(
        Payment.status == "success",
        Payment.created_at >= datetime.strptime(d_from, "%Y-%m-%d"),
        Payment.created_at < datetime.strptime(d_to, "%Y-%m-%d") + timedelta(days=1),
        ~Payment.customer_id.in_(_internal_ids_sub),
    )
    if store_short:
        q = q.filter(Payment.store == store_short)
    rows = q.all()
    total = round(sum(float(p.amount or 0) for p in rows), 2)
    count = len(rows)
    by_method = {}
    for p in rows:
        m = p.method or "other"
        by_method.setdefault(m, {"count": 0, "total": 0.0})
        by_method[m]["count"] += 1
        by_method[m]["total"] = round(by_method[m]["total"] + float(p.amount or 0), 2)
    # 按金额排序
    method_list = sorted(by_method.items(), key=lambda x: -x[1]["total"])
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({
        "preset": preset, "label": label, "d_from": d_from, "d_to": d_to,
        "total": total, "count": count,
        "method_list": method_list,
        "method_zh": _REVENUE_PAY_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/revenue_report.html", ctx)


@app.get("/m/prescriptions", response_class=HTMLResponse)
async def m_prescriptions_list(request: Request, db: Session = Depends(get_db),
                                q: str = "", status: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/prescriptions", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(Prescription).order_by(Prescription.id.desc())
    if store_short:
        query = query.outerjoin(Pet, Pet.id == Prescription.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )
    if status:
        query = query.filter(Prescription.status == status)
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        if q.isdigit():
            query = query.filter(or_(Prescription.id == int(q), Prescription.customer_id.in_(cids)))
        else:
            query = query.filter(Prescription.customer_id.in_(cids))
    items = query.limit(80).all()
    rows = []
    for p in items:
        pet = db.get(Pet, p.pet_id) if p.pet_id else None
        c = db.get(Customer, p.customer_id) if p.customer_id else None
        rows.append({"p": p, "pet": pet, "cust": c})
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "q": q, "status": status})
    return templates.TemplateResponse(request, "m_uk/prescriptions.html", ctx)


@app.get("/m/exam-orders", response_class=HTMLResponse)
async def m_exam_orders_list(request: Request, db: Session = Depends(get_db),
                              q: str = "", status: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/exam-orders", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(ExamOrder).order_by(ExamOrder.id.desc())
    if store_short:
        query = query.outerjoin(Visit, Visit.id == ExamOrder.visit_id).outerjoin(
            Pet, Pet.id == Visit.pet_id
        ).filter(or_(Pet.store == store_short, Pet.store == "", Pet.store == None))  # noqa: E711
    if status:
        query = query.filter(ExamOrder.status == status)
    if q:
        if q.isdigit():
            query = query.filter(ExamOrder.id == int(q))
    items = query.limit(80).all()
    rows = []
    for eo in items:
        v = db.get(Visit, eo.visit_id) if eo.visit_id else None
        pet = db.get(Pet, v.pet_id) if v and v.pet_id else None
        c = db.get(Customer, v.customer_id) if v and v.customer_id else None
        try:
            items_data = json.loads(eo.items_json or "[]")
        except Exception:
            items_data = []
        try:
            reports_cnt = len(list(eo.reports or []))
        except Exception:
            reports_cnt = 0
        rows.append({"eo": eo, "v": v, "pet": pet, "cust": c,
                     "items_n": len(items_data), "reports_n": reports_cnt})
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "q": q, "status": status})
    return templates.TemplateResponse(request, "m_uk/exam_orders.html", ctx)


@app.get("/m/vaccinations", response_class=HTMLResponse)
async def m_vaccinations_list(request: Request, db: Session = Depends(get_db), q: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/vaccinations", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(Vaccination).filter(Vaccination.status == "active").order_by(Vaccination.id.desc())
    if store_short:
        query = query.outerjoin(Pet, Pet.id == Vaccination.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(Vaccination.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for v in items:
        p = db.get(Pet, v.pet_id) if v.pet_id else None
        c = db.get(Customer, v.customer_id) if v.customer_id else None
        rows.append({"v": v, "pet": p, "cust": c})
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "q": q,
                "title": "疫苗", "subtitle": "接种记录 · 到期提醒",
                "back_url": "/m/medical"})
    return templates.TemplateResponse(request, "m_uk/vaccinations.html", ctx)


@app.get("/m/dewormings", response_class=HTMLResponse)
async def m_dewormings_list(request: Request, db: Session = Depends(get_db), q: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/dewormings", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(DewormingRecord).filter(DewormingRecord.status == "active").order_by(DewormingRecord.id.desc())
    if store_short:
        query = query.outerjoin(Pet, Pet.id == DewormingRecord.pet_id).filter(
            or_(Pet.store == store_short, Pet.store == "", Pet.store == None)  # noqa: E711
        )
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(DewormingRecord.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for d in items:
        p = db.get(Pet, d.pet_id) if d.pet_id else None
        c = db.get(Customer, d.customer_id) if d.customer_id else None
        rows.append({"d": d, "pet": p, "cust": c})
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "q": q})
    return templates.TemplateResponse(request, "m_uk/dewormings.html", ctx)


@app.get("/m/sales", response_class=HTMLResponse)
async def m_sales_list(request: Request, db: Session = Depends(get_db), q: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/sales", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(SalesOrder).order_by(SalesOrder.id.desc())
    if store_short:
        query = query.filter(or_(SalesOrder.store == store_short, SalesOrder.store == ""))
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(SalesOrder.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for s in items:
        c = db.get(Customer, s.customer_id) if s.customer_id else None
        rows.append({"s": s, "cust": c})
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({"rows": rows, "q": q})
    return templates.TemplateResponse(request, "m_uk/sales.html", ctx)


@app.get("/m/sales/{order_id}", response_class=HTMLResponse)
async def m_sales_detail(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/sales/{order_id}", status_code=303)
    order = db.get(SalesOrder, order_id)
    if not order:
        raise HTTPException(404)
    cust = db.get(Customer, order.customer_id) if order.customer_id else None
    pet = db.get(Pet, order.pet_id) if order.pet_id else None
    # 关联 invoice（独立销售单自动同步）
    inv = None
    if not order.visit_id:
        inv_item = (
            db.query(InvoiceItem)
            .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
            .filter(InvoiceItem.ref_type == "sales_order", InvoiceItem.ref_id == order_id,
                    Invoice.visit_id.is_(None))
            .first()
        )
        if inv_item:
            inv = db.get(Invoice, inv_item.invoice_id)
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({
        "order": order, "cust": cust, "pet": pet, "inv": inv,
        "so_status_zh": _SO_STATUS_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/sale_detail.html", ctx)


@app.get("/m/packages", response_class=HTMLResponse)
async def m_packages_list(request: Request, db: Session = Depends(get_db), q: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/packages", status_code=303)
    query = db.query(CustomerPackage).order_by(CustomerPackage.id.desc())
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(CustomerPackage.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for cp in items:
        c = db.get(Customer, cp.customer_id) if cp.customer_id else None
        rows.append({"cp": cp, "cust": c})
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({"rows": rows, "q": q})
    return templates.TemplateResponse(request, "m_uk/packages.html", ctx)


@app.get("/m/deposits", response_class=HTMLResponse)
async def m_deposits_list(request: Request, db: Session = Depends(get_db), q: str = "",
                          status: str = "active"):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/deposits", status_code=303)
    query = db.query(Deposit).order_by(Deposit.id.desc())
    if status == "active":
        query = query.filter(Deposit.status.in_(["held", "partial_refund"]))
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(Deposit.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for d in items:
        c = db.get(Customer, d.customer_id) if d.customer_id else None
        remaining = max(0.0, float(d.amount or 0) - float(d.applied_amount or 0) - float(d.refunded_amount or 0))
        rows.append({"d": d, "cust": c, "remaining": remaining})
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({"rows": rows, "q": q, "status": status,
                "deposit_category_zh": _DEPOSIT_CATEGORY_ZH})
    return templates.TemplateResponse(request, "m_uk/deposits.html", ctx)


@app.get("/m/coupons", response_class=HTMLResponse)
async def m_coupons_list(request: Request, db: Session = Depends(get_db), q: str = "",
                         status: str = "issued"):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/coupons", status_code=303)
    query = db.query(Coupon).order_by(Coupon.id.desc())
    if status:
        query = query.filter(Coupon.status == status)
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(or_(Coupon.customer_id.in_(cids), Coupon.customer_id == None))  # noqa: E711
    items = query.limit(50).all()
    rows = []
    for c in items:
        cust = db.get(Customer, c.customer_id) if c.customer_id else None
        rows.append({"c": c, "cust": cust})
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({"rows": rows, "q": q, "status": status,
                "coupon_kind_zh": _COUPON_KIND_ZH})
    return templates.TemplateResponse(request, "m_uk/coupons.html", ctx)


@app.get("/m/wallets", response_class=HTMLResponse)
async def m_wallets_list(request: Request, db: Session = Depends(get_db), q: str = ""):
    """钱包列表 = 有余额的客户钱包；空时引导去客户搜索充值"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/wallets", status_code=303)
    query = db.query(Wallet).filter(Wallet.balance > 0).order_by(Wallet.balance.desc())
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(Wallet.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for w in items:
        c = db.get(Customer, w.customer_id) if w.customer_id else None
        rows.append({"w": w, "cust": c})
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx.update({"rows": rows, "q": q})
    return templates.TemplateResponse(request, "m_uk/wallets.html", ctx)


@app.get("/m/consents", response_class=HTMLResponse)
async def m_consents_list(request: Request, db: Session = Depends(get_db), q: str = "",
                          status: str = ""):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/consents", status_code=303)
    store_short = _get_op_store(request)
    query = db.query(ConsentTask).order_by(ConsentTask.id.desc())
    if store_short:
        query = query.filter(or_(ConsentTask.store == store_short, ConsentTask.store == ""))
    if status:
        query = query.filter(ConsentTask.status == status)
    if q:
        cids = [c.id for c in db.query(Customer.id).filter(
            or_(Customer.name.ilike(f"%{q}%"), Customer.phone.ilike(f"%{q}%"))
        ).all()]
        query = query.filter(ConsentTask.customer_id.in_(cids))
    items = query.limit(50).all()
    rows = []
    for t in items:
        c = db.get(Customer, t.customer_id) if t.customer_id else None
        p = db.get(Pet, t.pet_id) if t.pet_id else None
        rows.append({"t": t, "cust": c, "pet": p})
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"rows": rows, "q": q, "status": status})
    return templates.TemplateResponse(request, "m_uk/consents.html", ctx)


@app.get("/m/consent-task/{task_id}", response_class=HTMLResponse)
async def m_consent_task_detail(task_id: int, request: Request, db: Session = Depends(get_db)):
    """手机端协议任务详情：已签可看正文+签名+打印存档；待签可复制链接/发短信。"""
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/consent-task/{task_id}", status_code=303)
    task = db.get(ConsentTask, task_id)
    if not task:
        raise HTTPException(404)
    cust = db.get(Customer, task.customer_id) if task.customer_id else None
    pet = db.get(Pet, task.pet_id) if task.pet_id else None
    doc = db.query(ConsentDocument).filter(ConsentDocument.task_id == task_id).first()
    sign_url = f"{request.url.scheme}://{request.url.netloc}/consent/{task.token}"
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "task": task, "cust": cust, "pet": pet, "doc": doc,
        "sign_url": sign_url,
        "status_zh": _CONSENT_STATUS_ZH,
        "category_zh": _CONSENT_CATEGORY_ZH,
    })
    return templates.TemplateResponse(request, "m_uk/consent_detail.html", ctx)


@app.get("/m/inventory", response_class=HTMLResponse)
async def m_inventory_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    filter: str = "all",   # all / low / zero / controlled / expiry
    category: str = "",
):
    """库存列表：搜索 + 筛选 + 低/零库存预警"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/inventory", status_code=303)
    from datetime import date as _date, timedelta as _timedelta
    store_short = _get_op_store(request)
    query = db.query(InventoryItem).filter(InventoryItem.is_active == True)
    query = _apply_store_filter(query, InventoryItem.store, store_short)
    if q:
        query = query.filter(or_(
            InventoryItem.name.ilike(f"%{q}%"),
            InventoryItem.supplier.ilike(f"%{q}%"),
        ))
    if category:
        query = query.filter(InventoryItem.category == category)
    if filter == "low":
        query = query.filter(
            InventoryItem.is_service == False,
            InventoryItem.stock_qty <= InventoryItem.low_stock_min,
            InventoryItem.low_stock_min > 0,
        )
    elif filter == "zero":
        query = query.filter(InventoryItem.is_service == False, InventoryItem.stock_qty <= 0)
    elif filter == "controlled":
        query = query.filter(InventoryItem.is_controlled == True)
    elif filter == "expiry":
        alert_date = (_date.today() + _timedelta(days=90)).isoformat()
        expiry_ids = (db.query(InventoryBatch.item_id)
                      .filter(InventoryBatch.is_depleted == False,
                              InventoryBatch.expiry_date != "",
                              InventoryBatch.expiry_date <= alert_date)
                      .distinct().subquery())
        query = query.filter(InventoryItem.id.in_(expiry_ids))
    items = query.order_by(InventoryItem.category, InventoryItem.name).limit(80).all()
    # KPI
    base_q = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, store_short
    ).filter(InventoryItem.is_active == True)
    low_count = base_q.filter(
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= InventoryItem.low_stock_min,
        InventoryItem.low_stock_min > 0,
    ).count()
    zero_count = base_q.filter(
        InventoryItem.is_service == False,
        InventoryItem.stock_qty <= 0,
    ).count()
    controlled_count = base_q.filter(InventoryItem.is_controlled == True).count()
    alert_date = (_date.today() + _timedelta(days=90)).isoformat()
    expiry_count = (db.query(InventoryBatch.item_id)
                    .filter(InventoryBatch.is_depleted == False,
                            InventoryBatch.expiry_date != "",
                            InventoryBatch.expiry_date <= alert_date)
                    .distinct().count())
    ctx = _m_ctx(request, db, active_tab="me")
    ctx.update({
        "items": items, "q": q, "filter": filter, "category": category,
        "low_count": low_count, "zero_count": zero_count,
        "controlled_count": controlled_count, "expiry_count": expiry_count,
        "categories": INVENTORY_CATEGORIES,
    })
    return templates.TemplateResponse(request, "m_uk/inventory.html", ctx)


@app.get("/m/inventory/import-photo", response_class=HTMLResponse)
async def m_inventory_import_photo(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/inventory/import-photo", status_code=303)
    ctx = _m_ctx(request, db, active_tab="finance")
    ctx["categories"] = INVENTORY_CATEGORIES
    return templates.TemplateResponse(request, "m_uk/inventory_import_photo.html", ctx)


@app.get("/m/inventory/{item_id}", response_class=HTMLResponse)
async def m_inventory_detail(item_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/inventory/{item_id}", status_code=303)
    item = db.get(InventoryItem, item_id)
    if not item:
        raise HTTPException(404)
    txs = (db.query(InventoryTransaction)
           .filter(InventoryTransaction.item_id == item_id)
           .order_by(InventoryTransaction.created_at.desc()).limit(20).all())
    batches = (db.query(InventoryBatch)
               .filter(InventoryBatch.item_id == item_id)
               .order_by(InventoryBatch.expiry_date).all())
    from datetime import date as _date, timedelta as _timedelta
    today_str = _date.today().isoformat()
    alert_date_str = (_date.today() + _timedelta(days=90)).isoformat()
    ctx = _m_ctx(request, db, active_tab="me")
    ctx.update({
        "item": item, "txs": txs, "batches": batches,
        "today_str": today_str, "alert_date_str": alert_date_str,
        "categories": INVENTORY_CATEGORIES,
    })
    return templates.TemplateResponse(request, "m_uk/inventory_detail.html", ctx)


@app.get("/m/appointment/{appt_id}", response_class=HTMLResponse)
async def m_appointment_detail(appt_id: int, request: Request, db: Session = Depends(get_db)):
    """手机端预约详情：信息 + 状态动作（确认/到店/完成/取消/爽约）"""
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/appointment/{appt_id}", status_code=303)
    a = db.get(Appointment, appt_id)
    if not a:
        raise HTTPException(404)
    # 关联客户
    cust = db.get(Customer, a.customer_id) if a.customer_id else None
    pet  = db.get(Pet, a.pet_id) if a.pet_id else None
    # 关联 TNR 申请（如果有）
    related_app = None
    if a.related_application_id:
        related_app = db.get(Application, a.related_application_id)
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({"a": a, "cust": cust, "pet": pet, "related_app": related_app})
    return templates.TemplateResponse(request, "m_uk/appointment_detail.html", ctx)


@app.get("/m/calendar", response_class=HTMLResponse)
async def m_calendar(
    request: Request,
    db: Session = Depends(get_db),
    ym: str = "",   # YYYY-MM；空则当月
):
    """手机端月历总览：每天一格 + 该天预约数 + 状态色点"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/calendar", status_code=303)
    from datetime import date as _date, timedelta as _td
    from calendar import monthrange

    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""

    today = _date.today()
    today_iso = today.isoformat()
    # 解析 ym
    try:
        if ym:
            yy, mm = map(int, ym.split("-"))
        else:
            yy, mm = today.year, today.month
        cur_first = _date(yy, mm, 1)
    except Exception:
        cur_first = _date(today.year, today.month, 1)
        yy, mm = cur_first.year, cur_first.month

    # 上 / 下个月
    if mm == 1:
        prev_ym = f"{yy-1:04d}-12"
    else:
        prev_ym = f"{yy:04d}-{mm-1:02d}"
    if mm == 12:
        next_ym = f"{yy+1:04d}-01"
    else:
        next_ym = f"{yy:04d}-{mm+1:02d}"

    # 查这个月所有预约
    _, days_in_month = monthrange(yy, mm)
    month_start = cur_first.isoformat()
    month_end = _date(yy, mm, days_in_month).isoformat()

    q = db.query(Appointment).filter(
        Appointment.appointment_date >= month_start,
        Appointment.appointment_date <= month_end,
        Appointment.status.notin_(["cancelled", "rejected"]),
    )
    if store_short:
        q = q.filter(or_(Appointment.store == store_short, Appointment.store == store_full))
    appts = q.all()

    # 按日期分桶 + 状态计数
    day_buckets: dict[str, dict] = {}
    for a in appts:
        d = a.appointment_date or ""
        if not d:
            continue
        b = day_buckets.setdefault(d, {"total": 0, "pending": 0, "confirmed": 0, "arrived": 0, "completed": 0, "no_show": 0})
        b["total"] += 1
        st = a.status if a.status in b else "confirmed"
        b[st] = b.get(st, 0) + 1

    # 生成 6 周 × 7 天的网格（多出来的前后补灰）
    # 周日开头 = 0
    first_weekday = cur_first.weekday()  # 周一=0
    # 转成 周日=0
    first_weekday_sun = (first_weekday + 1) % 7
    grid_start = cur_first - _td(days=first_weekday_sun)
    cells = []
    for i in range(42):
        d = grid_start + _td(days=i)
        iso = d.isoformat()
        in_month = (d.month == mm and d.year == yy)
        b = day_buckets.get(iso, None)
        cells.append({
            "iso": iso,
            "day": d.day,
            "in_month": in_month,
            "is_today": iso == today_iso,
            "bucket": b,
        })
        if i >= 35 and d.month != mm:
            # 第 6 周如果全在下个月，可剪掉。这里保留以稳定 6 周高度。
            pass

    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "year": yy, "month": mm,
        "cells": cells,
        "prev_ym": prev_ym, "next_ym": next_ym,
        "today_iso": today_iso,
        "total_month": len(appts),
    })
    return templates.TemplateResponse(request, "m_uk/calendar.html", ctx)


@app.get("/m/appointments", response_class=HTMLResponse)
async def m_appointments_list(
    request: Request,
    db: Session = Depends(get_db),
    q: str = "",
    scope: str = "today",   # today / week / pending / all
    d: str = "",            # 指定日期
):
    """预约列表：今日/本周/待确认/全部 + 搜索"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/appointments", status_code=303)
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    today_str = d.strip() or datetime.utcnow().strftime("%Y-%m-%d")
    from datetime import datetime as _dt, timedelta as _td
    today_dt = _dt.strptime(today_str, "%Y-%m-%d")
    week_end_str = (today_dt + _td(days=6)).strftime("%Y-%m-%d")

    base = db.query(Appointment).order_by(
        Appointment.appointment_date.asc(),
        Appointment.appointment_time.asc(),
    )
    if store_short:
        base = base.filter(or_(Appointment.store == store_short, Appointment.store == store_full))

    if scope == "today":
        base = base.filter(Appointment.appointment_date == today_str)
    elif scope == "week":
        base = base.filter(Appointment.appointment_date >= today_str,
                            Appointment.appointment_date <= week_end_str)
    elif scope == "pending":
        base = base.filter(Appointment.status == AppointmentStatus.pending.value)

    # 默认排除取消 / 拒绝 / 爽约
    if scope != "all":
        base = base.filter(Appointment.status.notin_(["cancelled", "rejected", "no_show"]))

    if q:
        base = base.filter(or_(
            Appointment.customer_name.ilike(f"%{q}%"),
            Appointment.phone.ilike(f"%{q}%"),
            Appointment.pet_name.ilike(f"%{q}%"),
        ))

    appts = base.limit(80).all()

    # 按日期分组
    grouped: dict[str, list] = {}
    for a in appts:
        grouped.setdefault(a.appointment_date or "—", []).append(a)
    # 待确认计数
    pending_count = db.query(Appointment).filter(
        Appointment.status == AppointmentStatus.pending.value,
    )
    if store_short:
        pending_count = pending_count.filter(or_(Appointment.store == store_short, Appointment.store == store_full))
    pending_n = pending_count.count()

    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "appts": appts, "grouped": grouped,
        "scope": scope, "q": q, "d": today_str,
        "pending_n": pending_n,
    })
    return templates.TemplateResponse(request, "m_uk/appointments.html", ctx)


@app.get("/m/appointments/new", response_class=HTMLResponse)
async def m_appointment_new_form(
    request: Request,
    db: Session = Depends(get_db),
    customer_id: int = 0,
    pet_id: int = 0,
    category: str = "outpatient",
):
    """手机端新建预约表单"""
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/appointments/new", status_code=303)
    cust = db.get(Customer, customer_id) if customer_id else None
    pet = db.get(Pet, pet_id) if pet_id else None
    pets = []
    if cust:
        pets = db.query(Pet).filter(Pet.customer_id == cust.id).order_by(Pet.id).all()
    store_short = _get_admin_store(request) or (request.session.get("admin_store") or "")
    is_super = (request.session.get("admin_role") == "superadmin")
    # TNR 配额信息（如果当前选了某个店）
    tnr_info = None
    if store_short:
        cfg = _get_tnr_store_config(db, store_short)
        year_month = datetime.utcnow().strftime("%Y-%m")
        used = _get_tnr_monthly_confirmed_count(db, store_short, year_month)
        tnr_info = {
            "accepting": bool(cfg.tnr_accepting),
            "quota": cfg.tnr_monthly_quota,
            "used": used,
            "remaining": max(0, cfg.tnr_monthly_quota - used),
            "year_month": year_month,
        }
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "cust": cust, "pet": pet, "pets": pets,
        "default_store": store_short,
        "is_superadmin": is_super,
        "category_default": category,
        "tnr_info": tnr_info,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
    })
    return templates.TemplateResponse(request, "m_uk/appointment_new.html", ctx)


@app.get("/m/exam-order/{order_id}", response_class=HTMLResponse)
async def m_exam_detail(order_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/exam-order/{order_id}", status_code=303)
    order = db.get(ExamOrder, order_id)
    if not order:
        raise HTTPException(404)
    v = db.get(Visit, order.visit_id) if order.visit_id else None
    pet = db.get(Pet, v.pet_id) if v and v.pet_id else None
    cust = db.get(Customer, v.customer_id) if v and v.customer_id else None
    try:
        items = json.loads(order.items_json or "[]")
    except Exception:
        items = []
    # 是否含显微镜项目（用于显示"显微镜报告"入口）
    has_microscopy = False
    for it in items:
        iid = it.get("item_id") if isinstance(it, dict) else None
        if iid:
            inv = db.get(InventoryItem, int(iid))
            if inv and (inv.category or "") == "microscopy":
                has_microscopy = True
                break
        if not has_microscopy:
            n = (it.get("name") if isinstance(it, dict) else "") or ""
            if any(k in n for k in ("镜检", "镜下", "刮片", "涂片", "粪检", "皮肤检查", "耳道分泌", "阴道脱落", "显微")):
                has_microscopy = True
                break
    assigned_labels = set()
    for rpt in (order.reports or []):
        lbl = (rpt.item_label or "").strip()
        if lbl:
            assigned_labels.add(lbl)
    micro_reports_m = db.query(MicroscopyReport).filter(
        MicroscopyReport.exam_order_id == order_id
    ).all()
    for mr in micro_reports_m:
        lbl = (mr.item_label or "").strip()
        if lbl:
            assigned_labels.add(lbl)
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "order": order, "v": v, "pet": pet, "cust": cust,
        "items": items, "reports": order.reports,
        "has_microscopy": has_microscopy,
        "assigned_labels": assigned_labels,
        "next_url": f"/m/exam-order/{order_id}",
    })
    return templates.TemplateResponse(request, "m_uk/exam_detail.html", ctx)


# TNR 审核
@app.get("/m/tnr", response_class=HTMLResponse)
async def m_tnr_list(request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse("/admin/login?next=/m/tnr", status_code=303)
    store_short = _get_op_store(request)
    store_full = _STORE_SHORT_TO_FULL.get(store_short, "") if store_short else ""
    q = db.query(Application).filter(
        Application.status.in_([
            ApplicationStatus.pending_manual.value,
            ApplicationStatus.pre_approved.value,
        ])
    )
    if store_short:
        q = q.filter(or_(
            Application.clinic_store == store_full,
            Application.clinic_store == store_short,
        ))
    rows = q.order_by(Application.created_at.desc()).limit(50).all()
    ctx = _m_ctx(request, db, active_tab="tnr")
    ctx["rows"] = rows
    return templates.TemplateResponse(request, "m_uk/tnr_list.html", ctx)


@app.get("/m/tnr/{app_id}", response_class=HTMLResponse)
async def m_tnr_detail(app_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/tnr/{app_id}", status_code=303)
    row = db.get(Application, app_id)
    if not row:
        raise HTTPException(404)
    # 照片
    images = [m for m in row.media if m.kind == MediaKind.application_image.value]
    videos = [m for m in row.media if m.kind == MediaKind.application_video.value]
    ctx = _m_ctx(request, db, active_tab="tnr")
    ctx.update({"row": row, "images": images, "videos": videos})
    return templates.TemplateResponse(request, "m_uk/tnr_detail.html", ctx)


@app.get("/m/visit/{visit_id}/prescribe", response_class=HTMLResponse)
async def m_prescribe_new(visit_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/visit/{visit_id}/prescribe", status_code=303)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404)
    if (v.status or "open") == "closed":
        return RedirectResponse(f"/m/visit/{visit_id}?err=病历已结束，不可开处方", status_code=303)
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    cust = db.get(Customer, v.customer_id) if v.customer_id else None
    # 模板：按使用次数热排
    templates_list = db.query(PrescriptionTemplate).order_by(
        PrescriptionTemplate.use_count.desc(), PrescriptionTemplate.id.desc()
    ).limit(20).all()
    uname = request.session.get("admin_username") or ""
    u = db.query(AdminUser).filter(AdminUser.username == uname).first() if uname else None
    default_vet = (u.display_name if u and u.display_name else uname) or v.vet_name or ""
    # 该宠物是否住院（决定是否要 schedule_times）
    hosp_admitted = False
    if v.pet_id:
        hosp_admitted = db.query(Hospitalization).filter(
            Hospitalization.pet_id == v.pet_id,
            Hospitalization.status == "admitted",
        ).first() is not None
    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "v": v, "pet": pet, "cust": cust,
        "templates_list": templates_list,
        "default_vet": default_vet,
        "today": datetime.utcnow().strftime("%Y-%m-%d"),
        "hosp_admitted": hosp_admitted,
    })
    return templates.TemplateResponse(request, "m_uk/prescription_new.html", ctx)


@app.get("/m/api/search-drug")
async def m_api_search_drug(
    request: Request,
    q: str = "",
    customer_id: int = 0,
    db: Session = Depends(get_db),
):
    if not _admin_ok(request):
        return {"results": []}
    q = (q or "").strip()
    if not q:
        return {"results": []}
    # 员工内购客户 → 单价按进价填
    is_internal = False
    if customer_id:
        _c = db.get(Customer, customer_id)
        is_internal = bool(_c and _c.is_internal)
    # 开处方场景：按当前操作门店过滤（含超管）
    store_short = _get_op_store(request)
    query = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, store_short
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.category.in_(["medication", "consumable"]),
        InventoryItem.name.ilike(f"%{q}%"),
    ).order_by(InventoryItem.is_controlled.desc(), InventoryItem.name).limit(15)
    results = []
    for it in query.all():
        # 库存等级：> low_stock_min = green, > 0 = yellow, 0 = red
        stock_level = "green"
        if it.is_service:
            stock_level = "green"
        elif it.stock_qty <= 0:
            stock_level = "red"
        elif it.stock_qty <= (it.low_stock_min or 0):
            stock_level = "yellow"
        # 内购客户：sell_price 透明替换为 cost_price，前端 JS 无需改动
        eff_price = float(it.cost_price or 0) if is_internal else float(it.sell_price or 0)
        results.append({
            "id": it.id, "name": it.name,
            "unit": it.unit or "", "unit2": it.unit2 or "",
            "sell_price": eff_price,
            "cost_price": float(it.cost_price or 0),
            "is_internal_pricing": is_internal,
            "stock_qty": float(it.stock_qty or 0),
            "stock_level": stock_level,
            "is_controlled": bool(it.is_controlled),
            "is_service": bool(it.is_service),
            "single_use_pack": bool(getattr(it, "single_use_pack", False)),
            "unit2": it.unit2 or "",
            "unit2_ratio": float(it.unit2_ratio or 1.0),
        })
    return {"results": results, "is_internal_pricing": is_internal}


@app.get("/m/api/search-product")
async def m_api_search_product(
    request: Request,
    q: str = "",
    customer_id: int = 0,
    db: Session = Depends(get_db),
):
    """销售单专用：全品类（药品/耗材/食品/用品/服务都允许），按操作门店过滤。"""
    if not _admin_ok(request):
        return {"results": []}
    q = (q or "").strip()
    if not q:
        return {"results": []}
    is_internal = False
    if customer_id:
        _c = db.get(Customer, customer_id)
        is_internal = bool(_c and _c.is_internal)
    store_short = _get_op_store(request)
    query = _apply_store_filter(
        db.query(InventoryItem), InventoryItem.store, store_short
    ).filter(
        InventoryItem.is_active == True,
        InventoryItem.name.ilike(f"%{q}%"),
    ).order_by(InventoryItem.name).limit(15)
    results = []
    for it in query.all():
        if it.is_service:
            stock_level = "green"
        elif (it.stock_qty or 0) <= 0:
            stock_level = "red"
        elif (it.stock_qty or 0) <= (it.low_stock_min or 0):
            stock_level = "yellow"
        else:
            stock_level = "green"
        eff_price = float(it.cost_price or 0) if is_internal else float(it.sell_price or 0)
        results.append({
            "id": it.id, "name": it.name,
            "unit": it.unit or "",
            "sell_price": eff_price,
            "stock_qty": float(it.stock_qty or 0),
            "stock_level": stock_level,
            "is_service": bool(it.is_service),
        })
    return {"results": results, "is_internal_pricing": is_internal}


@app.get("/m/api/recent-drugs")
async def m_api_recent_drugs(
    request: Request,
    db: Session = Depends(get_db),
):
    """当前用户最近 60 天开过的不同药品，按频次排序。"""
    if not _admin_ok(request):
        return {"results": []}
    uname = request.session.get("admin_username") or ""
    u = db.query(AdminUser).filter(AdminUser.username == uname).first() if uname else None
    vet_names = list({n for n in [uname, (u.display_name if u else None)] if n})
    if not vet_names:
        return {"results": []}
    from sqlalchemy import func as _f
    since = datetime.utcnow() - timedelta(days=60)
    rows = (
        db.query(PrescriptionItem.item_id, PrescriptionItem.drug_name,
                 _f.count(PrescriptionItem.id).label("n"))
        .join(Prescription, Prescription.id == PrescriptionItem.prescription_id)
        .filter(
            Prescription.vet_name.in_(vet_names),
            Prescription.created_at >= since,
            Prescription.status != "voided",
            PrescriptionItem.item_id.isnot(None),
        )
        .group_by(PrescriptionItem.item_id, PrescriptionItem.drug_name)
        .order_by(_f.count(PrescriptionItem.id).desc())
        .limit(12)
        .all()
    )
    op_store = _get_op_store(request)
    results = []
    for r in rows:
        inv = db.get(InventoryItem, r.item_id) if r.item_id else None
        if not inv:
            continue
        # 跨店历史不展示：当前操作门店不一致 → 跳过
        if op_store and (inv.store or "") != op_store:
            continue
        stock_level = "green"
        if inv.is_service:
            stock_level = "green"
        elif inv.stock_qty <= 0:
            stock_level = "red"
        elif inv.stock_qty <= (inv.low_stock_min or 0):
            stock_level = "yellow"
        results.append({
            "id": inv.id, "name": inv.name,
            "unit": inv.unit or "",
            "sell_price": float(inv.sell_price or 0),
            "stock_qty": float(inv.stock_qty or 0),
            "stock_level": stock_level,
            "is_controlled": bool(inv.is_controlled),
            "n": int(r.n),
        })
    return {"results": results}


@app.get("/m/api/presc-template/{tpl_id}")
async def m_api_presc_template(
    tpl_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    if not _admin_ok(request):
        raise HTTPException(401)
    tpl = db.get(PrescriptionTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404)
    try:
        items = json.loads(tpl.items_json or "[]")
    except Exception:
        items = []
    # 富化：补 item info
    out = []
    for it in items:
        iid = it.get("item_id")
        inv = db.get(InventoryItem, iid) if iid else None
        out.append({
            "item_id": iid,
            "drug_name": it.get("drug_name", ""),
            "drug_type": it.get("drug_type", "oral"),
            "dose_amount": it.get("dose_amount", 0),
            "dose_unit": it.get("dose_unit", ""),
            "times_per_day": it.get("times_per_day", 0),
            "frequency": it.get("frequency", ""),
            "duration_days": it.get("duration_days", ""),
            "quantity_num": it.get("quantity_num", 1),
            "instructions": it.get("instructions", ""),
            "schedule_times": it.get("schedule_times", ""),
            "unit_price": float(inv.sell_price) if inv else float(it.get("unit_price", 0)),
            "item_unit": (inv.unit if inv else "") or it.get("item_unit", ""),
        })
    return {"name": tpl.name, "notes": tpl.notes or "", "items": out}


@app.get("/m/visit/{visit_id}", response_class=HTMLResponse)
async def m_visit_detail(visit_id: int, request: Request, db: Session = Depends(get_db)):
    if not _admin_ok(request):
        return RedirectResponse(f"/admin/login?next=/m/visit/{visit_id}", status_code=303)
    v = db.get(Visit, visit_id)
    if not v:
        raise HTTPException(404)
    pet = db.get(Pet, v.pet_id) if v.pet_id else None
    cust = db.get(Customer, v.customer_id) if v.customer_id else None

    prescriptions = db.query(Prescription).filter(Prescription.visit_id == visit_id)\
        .order_by(Prescription.id.desc()).all()
    exam_orders = db.query(ExamOrder).filter(ExamOrder.visit_id == visit_id)\
        .order_by(ExamOrder.id.desc()).all()
    # 解析 exam items
    exam_rows = []
    for eo in exam_orders:
        try:
            items = json.loads(eo.items_json or "[]")
            if not isinstance(items, list):
                items = []
        except Exception:
            items = []
        try:
            reports = list(eo.reports or [])
        except Exception:
            reports = []
        exam_rows.append({"eo": eo, "items": items, "reports": reports})

    # 视觉对齐：vaccinations / dewormings 与 visit 时间窗口
    vaccinations = []
    dewormings = []
    if v.pet_id:
        # 同就诊日的疫苗/驱虫（粗略关联）
        vd = v.visit_date or ""
        if vd:
            vaccinations = db.query(Vaccination).filter(
                Vaccination.pet_id == v.pet_id,
                Vaccination.vaccinated_date == vd,
                Vaccination.status == "active",
            ).all()
            dewormings = db.query(DewormingRecord).filter(
                DewormingRecord.pet_id == v.pet_id,
                DewormingRecord.deworm_date == vd,
                DewormingRecord.status == "active",
            ).all()

    # 关联住院
    hospitalization = db.query(Hospitalization).filter(
        Hospitalization.visit_id == visit_id
    ).first()

    # 关联麻醉监护表
    anmon_sheets = db.query(AnesthesiaMonitorSheet).filter(
        AnesthesiaMonitorSheet.visit_id == visit_id
    ).order_by(AnesthesiaMonitorSheet.id.desc()).all()

    ctx = _m_ctx(request, db, active_tab="medical")
    ctx.update({
        "v": v, "pet": pet, "cust": cust,
        "prescriptions": prescriptions,
        "exam_rows": exam_rows,
        "vaccinations": vaccinations,
        "dewormings": dewormings,
        "hospitalization": hospitalization,
        "anmon_sheets": anmon_sheets,
    })
    return templates.TemplateResponse(request, "m_uk/visit_detail.html", ctx)


@app.post("/m/dispensing/{presc_id}/undo")
async def m_dispensing_undo(
    presc_id: int,
    request: Request,
    db: Session = Depends(get_db),
    csrf_token: str = Form(""),
):
    """撤销已配齐（误操作时）。"""
    if not _admin_ok(request):
        raise HTTPException(401)
    _require_csrf(request, csrf_token)
    p = db.get(Prescription, presc_id)
    if not p:
        raise HTTPException(404)
    p.dispensed_at = None
    p.dispensed_by = ""
    if p.status == "dispensed":
        p.status = "issued"
    db.commit()
    return RedirectResponse(f"/m/dispensing/{presc_id}?msg=已撤销配齐", status_code=303)
